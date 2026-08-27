"""Short Term Debt is repayment CLIPPED to the balance, not four quarters summed.

The workbook's Short Term Debt cell summed the next four quarters of Actual
Debt Repayment. On a revolver - a business that borrows and repays inside the
window - that counts money borrowed AFTER the balance-sheet date as if it had
been repaid out of the balance standing on it, so Short Term Debt, Current
Liabilities, Total Liabilities and Total Liabilities & Equity all come out
too high and the balance sheet does not balance.

Ironwood Precision Machining (1b9b4e45) is the case: a Q1 closing balance of
387,500 against repayments of 12,500 / 728,619 / 34,954 / 27,791 in Q2-Q5.
The bare sum is 803,864; the debt that actually exists is 387,500; the Q1
balance sheet was out by 416,364.

The engine has always done this correctly - finmo_model.py forward-simulates
and clips each quarter to what is left:

    _actual_clipped = min(_requested_repayment, max(0.0, _simulated_closing))
    _simulated_closing = max(0.0, _simulated_closing - _actual_clipped)

so this is the workbook learning what the engine already knew.

MEASURED across 400 drafts since 2026-07-13: 8 carry the class, errors from
$521 to $2,992,791 (median $1.4M). Only 3 of the 8 were ever visible, because
the Checks balance tie-out runs on Q1 and Q20 only - Cedar Ridge Bioenergy's
$2.99M error sits at Q4 and Meridian's $123,606 at Q16, and both delivered
without a word.
"""
from __future__ import annotations

import gzip
import json
import os
import shutil
import sys
import tempfile
import time
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
for path in (ROOT, os.path.join(ROOT, "python")):
  if path not in sys.path:
    sys.path.insert(0, path)

FIXTURE = os.path.join(HERE, "fixtures", "std_ironwood_revolver_export_row.json.gz")
PERIOD_START_COL = 3  # stub; Q1 = 4


def _fixture():
  with gzip.open(FIXTURE, "rt", encoding="utf-8") as fh:
    return json.load(fh)


def _quarter_rows(fx):
  return json.loads(fx["row"]["finmo_json"])["quarter_rows"]


def _build(fx, out_dir):
  from client_statements_output_excel.export_client_workbook import export_workbook_for_row
  row = dict(fx["row"])
  row["business_name"] = "STD revolver probe"
  return export_workbook_for_row(row, output_dir=out_dir, run_diagnostics=fx.get("run_diagnostics"))


def _excel_available() -> bool:
  try:
    import win32com.client as win32
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Quit()
    return True
  except Exception:
    return False


class TheFixtureCarriesTheRevolverClassTests(unittest.TestCase):
  """Offline: the stored payload really does repay more than it owes inside
  the window - if this stops being true, the fixture no longer guards."""

  def test_repayments_exceed_the_balance_inside_the_window(self):
    qr = _quarter_rows(_fixture())
    closing = float(qr[1].get("debt_closing_balance") or 0.0)
    window = [float(qr[q].get("debt_repayment") or 0.0) for q in range(2, 6)]
    self.assertGreater(sum(window), closing,
                       "fixture no longer exercises the revolver class - pick a draft that does")

  def test_the_engine_clips_and_the_bare_sum_does_not(self):
    qr = _quarter_rows(_fixture())
    closing = float(qr[1].get("debt_closing_balance") or 0.0)
    window = [float(qr[q].get("debt_repayment") or 0.0) for q in range(2, 6)]
    simulated, clipped = closing, 0.0
    for repayment in window:
      take = min(repayment, max(0.0, simulated))
      clipped += take
      simulated = max(0.0, simulated - take)
    self.assertAlmostEqual(clipped, float(qr[1].get("short_term_debt") or 0.0), places=2,
                           msg="the clipped walk must reproduce the engine's own STD")
    self.assertGreater(sum(window) - clipped, 1.0,
                       "the bare sum must overstate, or there is nothing to fix")


@unittest.skipUnless(_excel_available(), "Excel is required to evaluate the workbook's formulas")
class TheWorkbookMatchesTheEngineTests(unittest.TestCase):

  @classmethod
  def setUpClass(cls):
    import openpyxl
    import win32com.client as win32
    cls.fx = _fixture()
    cls.qr = _quarter_rows(cls.fx)
    cls.tmp = tempfile.mkdtemp(prefix="std_revolver_")
    path = str(_build(cls.fx, cls.tmp))
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Visible = False
    x.DisplayAlerts = False
    wb = x.Workbooks.Open(path)
    for _ in range(20):
      try:
        wb.Sheets(1).Name
        break
      except Exception:
        time.sleep(1.5)
    x.CalculateFullRebuild()
    wb.Save()
    wb.Close(False)
    x.Quit()
    cls.wb = openpyxl.load_workbook(path, data_only=True)
    cls.finmo = cls.wb["FINMO"]

  @classmethod
  def tearDownClass(cls):
    shutil.rmtree(cls.tmp, ignore_errors=True)

  def _row(self, label):
    for r in range(1, self.finmo.max_row + 1):
      if str(self.finmo.cell(r, 1).value or "").strip() == label:
        return r
    raise AssertionError(f"FINMO row {label!r} not found")

  def test_short_term_debt_equals_the_engine_every_quarter(self):
    row = self._row("Short Term Debt")
    misses = []
    for q in range(1, 21):
      got = float(self.finmo.cell(row, PERIOD_START_COL + q).value or 0.0)
      engine = float(self.qr[q].get("short_term_debt") or 0.0)
      if abs(got - engine) > 1.0:
        misses.append((f"Q{q}", engine, got))
    self.assertFalse(misses, f"workbook STD != engine: {misses[:5]}")

  def test_the_balance_sheet_balances_every_quarter(self):
    """Not just Q1 and Q20 - the Checks tie-out only covers those two, which
    is why the Cedar Ridge and Meridian errors were never seen."""
    assets, liabilities = self._row("Total Assets"), self._row("Total Liabilities & Equity")
    out = []
    for q in range(1, 21):
      a = float(self.finmo.cell(assets, PERIOD_START_COL + q).value or 0.0)
      le = float(self.finmo.cell(liabilities, PERIOD_START_COL + q).value or 0.0)
      if abs(a - le) > 1.0:
        out.append((f"Q{q}", round(a - le, 2)))
    self.assertFalse(out, f"balance sheet does not balance: {out[:5]}")

  def test_short_term_plus_long_term_equals_closing_debt(self):
    std, ltd = self._row("Short Term Debt"), self._row("Long Term Debt")
    misses = []
    for q in range(1, 21):
      total = (float(self.finmo.cell(std, PERIOD_START_COL + q).value or 0.0)
               + float(self.finmo.cell(ltd, PERIOD_START_COL + q).value or 0.0))
      engine = float(self.qr[q].get("debt_closing_balance") or 0.0)
      if abs(total - engine) > 1.0:
        misses.append((f"Q{q}", engine, total))
    self.assertFalse(misses, f"STD + LTD != closing debt: {misses[:5]}")


if __name__ == "__main__":
  unittest.main()
