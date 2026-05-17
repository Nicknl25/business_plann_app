"""Phase 9 P3.17 Phase 3c — workbook P&L lease-component flow tests.

Catches the regression class found in Phase 3c Phase 1: the
workbook FINMO P&L Interest and Depreciation cells must reference
BOTH the legacy debt-only / PPE-only Model Inputs cells AND the
lease component Model Inputs cells, so the rendered P&L line
equals the sum of components (matching the persisted FINMO
`interest` and `depreciation` totals, which include both).

Pre-iter, the workbook P&L formula referenced only the legacy
cells. The Debt Schedule reduced ROU correctly via its own
Excel formula, but the lease depreciation never flowed through
the P&L, breaking the workbook balance sheet by ~lease asset
depreciation per quarter.

Tests:
 1. Source-shape check on finmo_sheet.py — both legacy and
    lease Model Inputs references appear in the P&L Interest
    and Depreciation formula generators. Catches a future
    regression that drops a reference.
 2. Workbook-output check — if any recent ExpressLogix
    workbook exists from the test runs directory, load it and
    assert the FINMO Q1 Interest and Depreciation cell
    formulas reference both legacy and lease Model Inputs
    cells. Skipped if no workbook exists.
"""

from __future__ import annotations

import re
import sys
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
_PYTHON_DIR = _REPO_ROOT / "python"
if str(_PYTHON_DIR) not in sys.path:
  sys.path.insert(0, str(_PYTHON_DIR))


_FINMO_SHEET_SRC = (
  _REPO_ROOT / "client_statements_output_excel" / "finmo_sheet.py"
).read_text(encoding="utf-8")


class WorkbookPnlFormulaSourceTests(unittest.TestCase):
  """Static-source-shape regression check. The finmo_sheet.py
  P&L Interest and Depreciation `_set_formula(...)` lines must
  reference both legacy and lease Model Inputs cells.
  """

  def test_pnl_interest_formula_references_debt_and_lease(self) -> None:
    # The Interest line generator must reference both
    # 'is::Interest Expense' (debt-only legacy) AND
    # 'cash::Lease Interest Expense' (lease component).
    interest_line_re = re.compile(
      r'ctx\.finmo_row\("Income Statement",\s*"Interest"\).*?_set_formula',
      re.DOTALL,
    )
    # Find the Interest formula generator
    matches = re.findall(
      r'_set_formula\([^)]*ctx\.finmo_row\("Income Statement",\s*"Interest"\)[^"]*,\s*col,\s*(f"[^"]+")',
      _FINMO_SHEET_SRC,
    )
    self.assertEqual(len(matches), 1, f"Expected exactly one Interest formula generator; found {len(matches)}")
    formula = matches[0]
    self.assertIn("is::Interest Expense", formula, "P&L Interest formula must reference debt interest component")
    self.assertIn("cash::Lease Interest Expense", formula, "P&L Interest formula must reference lease interest component")

  def test_pnl_depreciation_formula_references_ppe_and_lease(self) -> None:
    matches = re.findall(
      r'_set_formula\([^)]*ctx\.finmo_row\("Income Statement",\s*"Depreciation"\)[^"]*,\s*col,\s*(f"[^"]+")',
      _FINMO_SHEET_SRC,
    )
    self.assertEqual(len(matches), 1, f"Expected exactly one Depreciation formula generator; found {len(matches)}")
    formula = matches[0]
    self.assertIn("is::Depreciation Expense", formula, "P&L Depreciation formula must reference PPE depreciation component")
    self.assertIn("cash::Lease Asset Depreciation", formula, "P&L Depreciation formula must reference lease asset depreciation component")


def _latest_express_workbook() -> Path | None:
  """Pick the most recently modified ExpressLogix workbook
  produced by a test run, if any exist."""
  base = Path("C:/dev/Cilient Plans")
  if not base.exists():
    return None
  matches = sorted(
    base.glob("ExpressLogix Shipping Services -- *.xlsx"),
    key=lambda p: p.stat().st_mtime,
    reverse=True,
  )
  return matches[0] if matches else None


_FINMO_SHEET_PATH = _REPO_ROOT / "client_statements_output_excel" / "finmo_sheet.py"


class WorkbookPnlFormulaOutputTests(unittest.TestCase):
  """End-to-end shape check on the most recent ExpressLogix
  workbook (if one exists AND was generated after the current
  finmo_sheet.py source). Verifies the Excel formula strings
  reference both the legacy and lease Model Inputs source cells.

  Skipped when no workbook exists or when the workbook is older
  than the source code — that state means a fresh E2E run is
  needed to regenerate the workbook before the output check is
  meaningful. The static source-shape check above covers the
  formula-generation correctness independently.
  """

  def setUp(self) -> None:
    self._workbook_path = _latest_express_workbook()
    if self._workbook_path is None:
      self.skipTest("No ExpressLogix workbook present on disk")
    workbook_mtime = self._workbook_path.stat().st_mtime
    source_mtime = _FINMO_SHEET_PATH.stat().st_mtime
    if workbook_mtime < source_mtime:
      self.skipTest(
        f"Latest workbook {self._workbook_path.name} is older than finmo_sheet.py "
        "source; rerun the lease-bearing E2E to regenerate before this check is meaningful."
      )

  def _load_pnl_formulas(self):
    import openpyxl
    wb = openpyxl.load_workbook(str(self._workbook_path), data_only=False)
    finmo = wb["FINMO"]
    mi = wb["Model Inputs"]
    interest_row = depreciation_row = None
    for r in range(1, 100):
      label = finmo.cell(row=r, column=1).value
      if label == "Interest":
        interest_row = r
      elif label == "Depreciation":
        depreciation_row = r
    self.assertIsNotNone(interest_row, "FINMO sheet must have Interest row")
    self.assertIsNotNone(depreciation_row, "FINMO sheet must have Depreciation row")
    # Build a Model Inputs label→row index so we can resolve cell refs
    mi_label_by_row = {}
    for r in range(1, 250):
      label = mi.cell(row=r, column=1).value
      if label is not None:
        mi_label_by_row[r] = str(label).strip()
    # Q1 is column D (index 4)
    interest_formula = finmo.cell(row=interest_row, column=4).value
    depreciation_formula = finmo.cell(row=depreciation_row, column=4).value
    return interest_formula, depreciation_formula, mi_label_by_row

  def _resolve_referenced_mi_labels(self, formula: str, mi_label_by_row: dict) -> set:
    """Extract any 'Model Inputs'!D<row> refs and resolve to labels."""
    refs = re.findall(r"'Model Inputs'!D(\d+)", formula or "")
    return {mi_label_by_row.get(int(r), f"<row {r}>") for r in refs}

  def test_workbook_pnl_interest_references_both_components(self) -> None:
    interest_formula, _, mi_label_by_row = self._load_pnl_formulas()
    labels = self._resolve_referenced_mi_labels(interest_formula, mi_label_by_row)
    self.assertIn(
      "Interest Expense", labels,
      f"Workbook P&L Interest at {self._workbook_path.name} must reference 'Interest Expense' Model Inputs row; got refs to {labels}",
    )
    self.assertIn(
      "Lease Interest Expense", labels,
      f"Workbook P&L Interest at {self._workbook_path.name} must reference 'Lease Interest Expense' Model Inputs row; got refs to {labels}",
    )

  def test_workbook_pnl_depreciation_references_both_components(self) -> None:
    _, depreciation_formula, mi_label_by_row = self._load_pnl_formulas()
    labels = self._resolve_referenced_mi_labels(depreciation_formula, mi_label_by_row)
    self.assertIn(
      "Depreciation Expense", labels,
      f"Workbook P&L Depreciation at {self._workbook_path.name} must reference 'Depreciation Expense' Model Inputs row; got refs to {labels}",
    )
    self.assertIn(
      "Lease Asset Depreciation", labels,
      f"Workbook P&L Depreciation at {self._workbook_path.name} must reference 'Lease Asset Depreciation' Model Inputs row; got refs to {labels}",
    )


if __name__ == "__main__":
  unittest.main()
