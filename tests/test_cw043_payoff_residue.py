"""A rounding crumb must never break the model.

CW-043 TURN A (mini's finding, 2026-08-24). The computed-principal Debt
Schedule (0b26ce8) lands principal as Excel's ``scheduled + extra``, which
reproduces the engine's repayment only to within an ulp. On Harrow Lane
Grooming (85f5825d) the Q7 cash sweep that pays the loan off landed at
11756.999999999998 instead of 11757, ``MAX(0, opening - principal)`` kept the
5.46e-12 of debt that was left, interest kept accruing on it for thirteen
more quarters, and FINMO's Interest Coverage / DSCR printed ~1e17 where the
previous build printed a dash. Checks!B2 said OK throughout - the tie-outs
are tolerance-based and a crumb is inside every tolerance.

Two layers are pinned here, on Harrow's REAL stored export row, built by the
production exporter and recalculated by Excel (the class lives in Excel's
arithmetic, so nothing short of Excel can measure it):

  1. The schedule lands on the engine's own grid: principal is ROUNDed to
     6 dp and a closing balance below half a cent is zero, so a payoff IS a
     payoff - closing exactly 0, no interest after it, on both blocks.
  2. Ratios never divide by a crumb: a denominator below half a cent prints
     the dash it would print for zero.

Skips (never silently passes) when Excel is not available.
"""
from __future__ import annotations

import gzip
import json
import os
import shutil
import sys
import tempfile
import time
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
for path in (ROOT, os.path.join(ROOT, "python")):
  if path not in sys.path:
    sys.path.insert(0, path)

FIXTURE = os.path.join(HERE, "fixtures", "cw043_harrow_export_row.json.gz")
PERIOD_START_COL = 3  # stub; Q1 = 4 ... Q20 = 23
PAYOFF_Q = 7


def _excel_available() -> bool:
  try:
    import win32com.client as win32  # noqa: F401
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Quit()
    return True
  except Exception:
    return False


def _recalc(path: str) -> None:
  import win32com.client as win32
  x = win32.gencache.EnsureDispatch("Excel.Application")
  x.Visible = False
  x.DisplayAlerts = False
  wb = x.Workbooks.Open(path)
  for _ in range(20):
    try:
      wb.Sheets(1).Name
      break
    except Exception:
      time.sleep(1.5)
  x.CalculateFullRebuild()
  wb.Save()
  wb.Close(False)
  x.Quit()


@unittest.skipUnless(_excel_available(), "Excel is required to evaluate the workbook's formulas")
class PayoffLandsExactlyTests(unittest.TestCase):

  @classmethod
  def setUpClass(cls):
    import openpyxl
    from client_statements_output_excel.export_client_workbook import export_workbook_for_row

    with gzip.open(FIXTURE, "rt", encoding="utf-8") as fh:
      fx = json.load(fh)
    row = dict(fx["row"])
    row["business_name"] = "CW043 residue probe"
    cls.tmp = tempfile.mkdtemp(prefix="cw043_residue_")
    path = export_workbook_for_row(row, output_dir=cls.tmp,
                                   run_diagnostics=fx.get("run_diagnostics"))
    _recalc(str(path))
    cls.wb = openpyxl.load_workbook(str(path), data_only=True)

  @classmethod
  def tearDownClass(cls):
    shutil.rmtree(cls.tmp, ignore_errors=True)

  # ---- helpers -----------------------------------------------------------
  def _debt_sheet(self):
    return self.wb["Debt Schedule"]

  def _header_cols(self, ws):
    return {str(ws.cell(6, c).value or "").strip(): c for c in range(1, 30)}

  def _finmo_row(self, label):
    ws = self.wb["FINMO"]
    for r in range(1, ws.max_row + 1):
      if str(ws.cell(r, 1).value or "").strip() == label:
        return ws, r
    raise AssertionError(f"FINMO row {label!r} not found")

  # ---- layer 1: the schedule ----------------------------------------------
  def test_the_payoff_quarter_closes_at_exactly_zero(self):
    ws = self._debt_sheet()
    r = 7 + PAYOFF_Q  # stub is row 7
    self.assertEqual(ws.cell(r, 1).value, f"Q{PAYOFF_Q}")
    closing = ws.cell(r, 11).value  # debt Closing
    self.assertEqual(closing, 0,
                     f"debt closing at the payoff quarter is {closing!r}, not 0 - "
                     "an ulp of principal survived MAX(0, ...)")

  def test_no_interest_accrues_after_the_payoff(self):
    ws = self._debt_sheet()
    for q in range(PAYOFF_Q + 1, 21):
      r = 7 + q
      self.assertEqual(ws.cell(r, 2).value, 0, f"Q{q} opening not 0")
      self.assertEqual(ws.cell(r, 9).value, 0, f"Q{q} interest {ws.cell(r, 9).value!r} on a paid-off loan")

  # ---- layer 2: the ratios ----------------------------------------------
  def test_interest_coverage_prints_a_dash_once_the_loan_is_gone(self):
    ws, r = self._finmo_row("Interest Coverage")
    bad = {}
    for q in range(PAYOFF_Q + 1, 21):
      v = ws.cell(r, PERIOD_START_COL + q).value
      if v != "-":
        bad[f"Q{q}"] = v
    self.assertFalse(bad, f"Interest Coverage on a paid-off loan should be '-': {bad}")

  def test_dscr_prints_a_dash_once_the_loan_is_gone(self):
    ws, r = self._finmo_row("Debt Service Coverage Ratio (DSCR)")
    bad = {}
    for q in range(PAYOFF_Q + 1, 21):
      v = ws.cell(r, PERIOD_START_COL + q).value
      if isinstance(v, (int, float)) and abs(v) > 1e6:
        bad[f"Q{q}"] = v
    self.assertFalse(bad, f"DSCR blew up on a crumb: {bad}")

  def test_no_crumbs_anywhere_on_the_schedule(self):
    """The class, not the instance: no cell of the Debt Schedule table holds
    a value that is nonzero but below a hundredth of a cent."""
    ws = self._debt_sheet()
    crumbs = []
    for r in range(7, 28):
      for c in range(2, 25):
        v = ws.cell(r, c).value
        if isinstance(v, float) and v != 0 and abs(v) < 1e-4:
          crumbs.append((ws.cell(6, c).value, ws.cell(r, 1).value, v))
    self.assertFalse(crumbs, f"rounding crumbs on the schedule: {crumbs[:8]}")


if __name__ == "__main__":
  unittest.main()
