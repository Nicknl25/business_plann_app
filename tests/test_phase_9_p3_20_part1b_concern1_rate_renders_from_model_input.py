"""Phase 9 P3.20 Part 1b Concern 1 — synthetic test.

Confirms the workbook builder faithfully renders whatever per-quarter
rate value is in the persisted model_input's `expenses::Interest Rate`
row. The Part 1 regenerated workbook from draft 3f7fd829 showed D12 =
0.1025 because that draft was persisted BEFORE bcf818d's bridge fix
(case (a): stale persisted data). A fresh model_input with 0.025625
in the Interest Rate row renders D12 = 0.025625.

There is no additional writer path that bypasses model_input.
"""
from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
_PYTHON_DIR = _REPO_ROOT / "python"
if str(_PYTHON_DIR) not in sys.path:
  sys.path.insert(0, str(_PYTHON_DIR))
if str(_REPO_ROOT) not in sys.path:
  sys.path.insert(0, str(_REPO_ROOT))


def _build_minimal_draft_row(
  *,
  interest_rate_per_quarter_q0: float,
  interest_rate_per_quarter_q1plus: float,
) -> dict:
  """Build a minimal `DraftWorkbookData`-compatible dict: a draft row
  with model_input + finmo that includes only what the Debt Schedule
  sheet builder reads. Other sheets may render with zeros; this test
  only inspects Debt Schedule cells D12 and the lease interest formula.
  """
  # 21-element values list: Q0 stub + Q1..Q20
  rate_values = [interest_rate_per_quarter_q0] + [
    interest_rate_per_quarter_q1plus for _ in range(20)
  ]
  expense_rows = [
    {
      "label": "Interest Rate",
      "lever_id": "expenses::Interest Rate",
      "values": rate_values,
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Depreciation",
      "lever_id": "expenses::Depreciation",
      "values": [0.0 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Taxes",
      "lever_id": "expenses::Taxes",
      "values": [0.21 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Cost of Goods Sold",
      "lever_id": "expenses::Cost of Goods Sold",
      "values": [0.7 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Marketing",
      "lever_id": "expenses::Marketing",
      "values": [0.05 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Payroll",
      "lever_id": "expenses::Payroll",
      "values": [10000.0 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Lease",
      "lever_id": "expenses::Lease",
      "values": [0.0 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "General & Administrative",
      "lever_id": "expenses::General & Administrative",
      "values": [0.03 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
    {
      "label": "Research & Development",
      "lever_id": "expenses::Research & Development",
      "values": [0.0 for _ in range(21)],
      "named_range": "model_input_expenses",
      "controller_write": True,
    },
  ]
  schedule_rows = [
    {
      "label": "Capital Expenditures",
      "lever_id": "schedules::Capital Expenditures",
      "values": [0.0 for _ in range(21)],
    },
    {
      "label": "Debt Issuance (New Borrowing)",
      "lever_id": "schedules::Debt Issuance (New Borrowing)",
      "values": [0.0 for _ in range(21)],
    },
    {
      "label": "Debt Repayment (Scheduled)",
      "lever_id": "schedules::Debt Repayment (Scheduled)",
      "values": [0.0 for _ in range(21)],
    },
    {
      "label": "Less: Principal Repayments",
      "lever_id": "schedules::Less: Principal Repayments",
      "values": [0.0 for _ in range(21)],
    },
    {
      "label": "Plus: Net Additions",
      "lever_id": "schedules::Plus: Net Additions",
      "values": [0.0 for _ in range(21)],
    },
  ]
  balance_sheet_rows = [
    {"label": "Accounts Receivable Days", "lever_id": "balance_sheet::Accounts Receivable Days", "values": [30.0 for _ in range(21)]},
    {"label": "Inventory Days", "lever_id": "balance_sheet::Inventory Days", "values": [30.0 for _ in range(21)]},
    {"label": "Accounts Payable Days", "lever_id": "balance_sheet::Accounts Payable Days", "values": [30.0 for _ in range(21)]},
    {"label": "Prepaid Expenses (% of Revenue)", "lever_id": "balance_sheet::Prepaid Expenses (% of Revenue)", "values": [0.0 for _ in range(21)]},
    {"label": "Deferred Revenue (% of Revenue)", "lever_id": "balance_sheet::Deferred Revenue (% of Revenue)", "values": [0.0 for _ in range(21)]},
    {"label": "Short Term Debt (% of LTD)", "lever_id": "balance_sheet::Short Term Debt (% of LTD)", "values": [0.0 for _ in range(21)]},
    {"label": "Owner's Capital", "lever_id": "balance_sheet::Owner's Capital", "values": [100000.0 for _ in range(21)]},
    {"label": "Distributions", "lever_id": "balance_sheet::Distributions", "values": [0.0 for _ in range(21)]},
    {"label": "Other Equity", "lever_id": "balance_sheet::Other Equity", "values": [0.0 for _ in range(21)]},
  ]
  revenue_rows = []  # not needed for this test
  model_input_json = {
    "business_name": "Synthetic Test Co",
    "draft_id": "synthetic_concern1_test",
    "start_date": "2026-01-01",
    "periods": [{"slot_index": i, "quarter": i, "year": 2026, "date": "", "days_in_quarter": 90, "is_stub": i == 0} for i in range(21)],
    "sections": {
      "revenue": revenue_rows,
      "expenses": expense_rows,
      "balance_sheet": balance_sheet_rows,
      "schedules": {
        "debt_opening_balance_seed": 500000.0,
        "lease_opening_balance_seed": 54000.0,
        "ppe_opening_balance_seed": 100000.0,
        "forecast_ppe_opening_balance_seed": 100000.0,
        "accumulated_depreciation_opening_seed": 0.0,
        "cash_opening_balance_seed": 50000.0,
        "accounts_receivable_opening_balance_seed": 0.0,
        "inventory_opening_balance_seed": 0.0,
        "accounts_payable_opening_balance_seed": 0.0,
        "short_term_debt_opening_balance_seed": 0.0,
        "rows": schedule_rows,
      },
    },
    "derived_driver_policies": {
      "debt_interest_rate_policy": {
        "policy_version": "sba_7a_business_loan_interest_rate_v1",
        "driver_source": "sba_loan_7a_raw",
        "lever_id": "expenses::Interest Rate",
        "annual_rate_decimal": 0.1025,
        "quarterly_rate_decimal": 0.025625,
        "source_detail": {"source": "sba_loan_7a_raw", "annual_rate_decimal": 0.1025},
      },
    },
  }
  # Minimal finmo_json: 21 quarter_rows
  finmo_quarter_rows = []
  for q in range(21):
    finmo_quarter_rows.append({
      "quarter_index": q, "quarter": q, "year": 2026, "date": "2026-03-31",
      "days_in_quarter": 90,
      "revenue": 0.0, "cost_of_goods_sold": 0.0, "gross_profit": 0.0,
      "marketing": 0.0, "research_and_development": 0.0, "lease_rent": 0.0,
      "payroll": 10000.0, "general_and_administrative": 0.0,
      "ebitda": 0.0, "interest": 0.0, "depreciation": 0.0, "taxes": 0.0,
      "net_income": 0.0, "cash": 50000.0, "accounts_receivable": 0.0, "inventory": 0.0,
      "current_assets": 50000.0, "ppe": 100000.0, "accumulated_depreciation": 0.0,
      "total_assets": 150000.0, "accounts_payable": 0.0, "prepaid_expenses": 0.0,
      "short_term_debt": 0.0, "deferred_revenue": 0.0, "current_liabilities": 0.0,
      "long_term_debt": 500000.0, "total_liabilities": 500000.0,
      "owners_capital": 100000.0, "distributions": 0.0, "retained_earnings": -450000.0,
      "other_equity": 0.0, "total_equity": -350000.0, "total_liabilities_and_equity": 150000.0,
      "ending_cash": 50000.0,
      "debt_opening_balance": 500000.0, "debt_closing_balance": 500000.0,
      "debt_interest_rate": 0.025625, "debt_interest_expense": 0.0,
      "lease_opening_balance_total": 54000.0, "lease_principal_repayments": 0.0,
      "lease_net_additions": 0.0, "lease_closing_balance_total": 54000.0,
      "lease_interest_expense": 0.0, "lease_asset_depreciation_expense": 0.0,
      "ppe_depreciation_expense": 0.0, "debt_interest_expense_only": 0.0,
      "right_of_use_asset": 54000.0, "right_of_use_asset_opening": 54000.0,
      "capital_lease_obligation": 54000.0,
      "capital_expenditures": 0.0, "investing_cash_flow": 0.0,
      "debt_issuance": 0.0, "debt_repayment": 0.0, "debt_receive_repay": 0.0,
      "equity": 0.0, "owner_distributions": 0.0, "financing_cash_flow": 0.0,
      "net_cash_flow": 0.0, "beginning_cash": 50000.0,
      "changes_in_current_assets": 0.0, "changes_in_current_liabilities": 0.0,
      "operating_cash_flow": 0.0,
      "debt_requested_issuance": 0.0, "debt_requested_repayment": 0.0,
      "debt_additions_repayments_net": 0.0, "accounting_equation_check": 0.0,
    })
  finmo_json = {"quarter_rows": finmo_quarter_rows}
  draft_row = {
    "draft_id": "synthetic_concern1_test",
    "client_id": "synthetic",
    "business_name": "Synthetic Test Co",
    "business_start_date": "2026-01-01",
  }
  return draft_row, model_input_json, finmo_json


class WorkbookRateCellRendersFromModelInputTests(unittest.TestCase):
  """When the Interest Rate row in model_input holds the per-quarter
  value 0.025625, the regenerated workbook's Debt Schedule!D12 holds
  0.025625. No other writer path overrides this.
  """

  def _build_and_inspect(
    self,
    *,
    rate_q0: float,
    rate_q1: float,
  ) -> dict:
    from client_statements_output_excel.data import DraftWorkbookData
    from client_statements_output_excel.workbook_builder import (
      build_client_financial_model_workbook,
    )
    import openpyxl

    draft_row, model_input, finmo = _build_minimal_draft_row(
      interest_rate_per_quarter_q0=rate_q0,
      interest_rate_per_quarter_q1plus=rate_q1,
    )
    data = DraftWorkbookData(
      draft_row=draft_row,
      model_input_json=model_input,
      finmo_json=finmo,
      payroll_headcount={"contract_version": "synthetic", "rows": []},
      debt_schedule={"contract_version": "synthetic", "rows": []},
      planning_run_json={"contract_version": "synthetic"},
    )
    wb = build_client_financial_model_workbook(data)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
      tmp_path = Path(tmp.name)
    try:
      wb.save(str(tmp_path))
      loaded = openpyxl.load_workbook(str(tmp_path), data_only=False)
      debt = loaded["Debt Schedule"]
      # Find Interest Rate, Interest Expense, Lease Interest, Lease Opening row indices
      labels = {}
      for r in range(1, 60):
        v = debt.cell(row=r, column=1).value
        if v:
          labels[v] = r
      return {
        "interest_rate_q0": debt.cell(row=labels["Interest Rate"], column=3).value,
        "interest_rate_q1": debt.cell(row=labels["Interest Rate"], column=4).value,
        "interest_expense_q1_formula": debt.cell(row=labels["Interest Expense"], column=4).value,
        "lease_interest_q1_formula": debt.cell(row=labels["Lease Interest Expense"], column=4).value,
        "lease_open_row": labels["Lease Opening Balance"],
        "interest_rate_row": labels["Interest Rate"],
      }
    finally:
      try:
        tmp_path.unlink()
      except Exception:
        pass

  def test_per_quarter_rate_renders_directly_into_d12(self) -> None:
    """Inject the per-quarter rate (annual / 4 = 0.025625) into the
    Interest Rate row; D12 must hold exactly that value."""
    result = self._build_and_inspect(rate_q0=0.0125, rate_q1=0.025625)
    self.assertEqual(
      result["interest_rate_q1"], 0.025625,
      f"D12 (Interest Rate Q1) must render the model_input value verbatim; got {result['interest_rate_q1']!r}",
    )
    self.assertEqual(
      result["interest_rate_q0"], 0.0125,
      f"C12 (Interest Rate Q0) must render the model_input stub value verbatim; got {result['interest_rate_q0']!r}",
    )

  def test_annual_rate_in_model_input_renders_as_annual_in_d12(self) -> None:
    """Symmetric test: if the model_input row STILL has the annual
    value (e.g., pre-bcf818d persisted draft), D12 renders the annual.
    This is what produced Part 1's observation: stale draft -> annual
    value in D12. The workbook builder is correct either way; the
    persisted source data is what determines the displayed value."""
    result = self._build_and_inspect(rate_q0=0.05, rate_q1=0.1025)
    self.assertEqual(
      result["interest_rate_q1"], 0.1025,
      "If model_input still holds annual rate, D12 holds annual rate (this was Part 1's stale-data case)",
    )

  def test_lease_interest_formula_references_d12_not_a_literal(self) -> None:
    """Confirms Part 1's cell-reference fix: lease interest formula
    references D12, NOT a Python-interpolated literal."""
    result = self._build_and_inspect(rate_q0=0.0125, rate_q1=0.025625)
    formula = str(result["lease_interest_q1_formula"])
    ir_row = result["interest_rate_row"]
    self.assertIn(f"D{ir_row}", formula, f"Lease interest formula must reference D{ir_row}; got {formula!r}")
    # No bare decimal literal in the formula (apart from cell refs)
    import re
    bare_rate = re.search(r"\*\s*0\.\d+", formula)
    self.assertIsNone(bare_rate, f"Lease interest formula must not contain *0.XX literal; got {formula!r}")


if __name__ == "__main__":
  unittest.main()
