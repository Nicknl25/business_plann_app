"""The payroll chain lands on the engine's grid, follows the person not the title,
and the hidden bridge reproduces the engine's rows.

Payroll Schedule (Nick, 2026-08-25, parts 2 and 3). The sheet is HORIZONTAL:
one block per role - a header row (Title/Person, Staffing Class), an OEWS
title row, a Wage source row, then nine period rows across the quarter
columns. Starting FTE(q) = ROUND(Ending FTE(q-1), 6); Annual Wage and
Benefits % = previous, or ROUND(previous +/- the engine's delta, 6) at a
bump; Hires literal. Below the blocks a HIDDEN BRIDGE reproduces the old
vertical block - one row per engine row, in the engine's order, every cell a
formula into the blocks - which is what the summary SUMIFS and the Checks
payroll tie-outs range over.

Pinned here, on real stored export rows:
  1. THE 2-DP GRID (Halbrook): the engine authors FTE and hires on a 2-dp
     grid; 6.06 + 0.35 in IEEE is 6.409999999999999; a bare chain compounds
     that crumb through Ending, Average, Wage Cost and the P&L. Every chained
     Starting FTE is ROUND(..., 6), and the exact emitted chain evaluated in
     doubles reproduces the engine bit for bit; the control test proves the
     fixture carries the class.
  2. ROLE IDENTITY (Understory: two Grow Technicians at DIFFERENT wages): a
     block per (class, title, person, ordinal); each person's chain reproduces
     THEIR wage series; the bridge row for a person points at that person's
     block. Under a title-only key the two merge and the wrong wage appears.
  3. THE BRIDGE: one hidden formula row per engine row in engine order, and
     with Excel its values equal the engine's (skips, never passes, without).
"""
from __future__ import annotations

import gzip
import json
import os
import re
import shutil
import sys
import tempfile
import time
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
for path in (ROOT, os.path.join(ROOT, "python")):
  if path not in sys.path:
    sys.path.insert(0, path)

HALBROOK = os.path.join(HERE, "fixtures", "cw043_halbrook_export_row.json.gz")
UNDERSTORY = os.path.join(HERE, "fixtures", "cw043_understory_export_row.json.gz")
PERIOD_START_COL = 3  # stub; Q1 = 4
#: The block's period rows, in order. "Wage source" leads them: it is a
#: per-quarter label (the owner-draw deferral changes it mid-horizon), written
#: only in the quarters where it CHANGES (2026-08-27).
PERIOD_ROWS = ["Wage source", "Starting FTE", "Hires", "Ending FTE", "Average FTE",
               "Annual Wage", "Benefits %", "Wage Cost", "Taxes & Benefits", "Total Payroll"]
NUMERIC_PERIOD_ROWS = [r for r in PERIOD_ROWS if r != "Wage source"]


def _fixture(path):
  with gzip.open(path, "rt", encoding="utf-8") as fh:
    return json.load(fh)


def _payroll_rows(fx):
  return [x for x in json.loads(fx["row"]["payroll_headcount"]).get("rows") or []
          if isinstance(x, dict)]


def _build_workbook(fx, out_dir):
  from client_statements_output_excel.export_client_workbook import export_workbook_for_row
  row = dict(fx["row"])
  row["business_name"] = "CW043 payroll probe"
  return export_workbook_for_row(row, output_dir=out_dir, run_diagnostics=fx.get("run_diagnostics"))


def _role_identity(rows):
  """(quarter, (class, title, person, ordinal-in-quarter)) per engine row - derived
  here from the stored rows, independently of the builder."""
  ids, seen, current_q = [], {}, None
  for item in rows:
    q = int(item.get("quarter_index") or 0)
    if q != current_q:
      current_q, seen = q, {}
    base = (str(item.get("staffing_class") or "").strip(),
            str(item.get("position_title") or "").strip(),
            str(item.get("person_name") or "").strip())
    ordinal = seen.get(base, 0)
    seen[base] = ordinal + 1
    ids.append((q, base + (ordinal,)))
  return ids


def _wage_source_expected(item):
  """The plain-English wage-source label for one engine row, derived here from
  the stored value so the test does not simply restate the builder."""
  raw = str(item.get("wage_source") or item.get("wage_source_code") or "").strip()
  if not raw:
    return ""
  base, _, rest = raw.partition("|")
  words = {
    "client_override": "Your stated wage",
    "intake_oews_key_person": "Your stated wage",
    "oews_title_catalog:oews_median": "Market median for this title (BLS)",
    "oews_median": "Market median for this title (BLS)",
    "oews_pct10": "Market 10th percentile for this title (BLS)",
    "oews_pct25": "Market 25th percentile for this title (BLS)",
    "oews_pct75": "Market 75th percentile for this title (BLS)",
    "oews_pct90": "Market 90th percentile for this title (BLS)",
  }.get(base.lower(), base)
  notes = {
    "floor_adapted": "raised to the wage floor",
    "part_time_hours_adapted": "adjusted for part-time hours",
    "owner_draw_deferred": "owner draw deferred",
  }
  extra = [notes.get(x.strip().lower(), x.strip()) for x in rest.split("|") if x.strip()]
  return words + (" (" + ", ".join(extra) + ")" if extra else "")


def _role_blocks(ws):
  """Visible role blocks: list of {label: row}, in sheet order.

  The geometry is VARIABLE: a staffed role carries an "OEWS title" row, a
  named person does not (no labelled empty rows, 2026-08-27), so the block is
  found by its period rows rather than a fixed offset."""
  blocks, r = [], 1
  while r <= ws.max_row:
    if (ws.cell(r, 1).value == PERIOD_ROWS[0]
        and ws.cell(r + 1, 1).value == PERIOD_ROWS[1]):
      has_oews = ws.cell(r - 1, 1).value == "OEWS title"
      block = {"header": r - 2 if has_oews else r - 1}
      if has_oews:
        block["oews"] = r - 1
      for n, label in enumerate(PERIOD_ROWS):
        assert ws.cell(r + n, 1).value == label, f"block at row {r} lost its {label} row"
        block[label] = r + n
      block["source"] = block["Wage source"]
      blocks.append(block)
      r += len(PERIOD_ROWS)
    else:
      r += 1
  return blocks


def _bridge_rows(ws):
  hdr = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == "Quarter"]
  assert hdr, "bridge header not found"
  first = hdr[-1] + 1
  last = first
  while isinstance(ws.cell(last + 1, 1).value, (int, float)):
    last += 1
  return first, last


def _evaluate_block(ws, block, rows_by_q):
  """Evaluate the block's emitted formulas in doubles; returns {label: {q: value}}."""
  row_label = {r_: l for l, r_ in block.items() if l in NUMERIC_PERIOD_ROWS}
  vals = {label: {} for label in NUMERIC_PERIOD_ROWS}
  for q in range(1, 21):
    if q not in rows_by_q:
      continue
    col = PERIOD_START_COL + q

    def resolve(v, label):
      if not isinstance(v, str):
        return float(v or 0.0)
      m = re.fullmatch(r"=ROUND\(([A-Z]+)(\d+)([+-][0-9.eE+-]+)?,6\)|=([A-Z]+)(\d+)", v)
      assert m, f"unexpected formula shape for {label} at column {col}: {v}"
      if m.group(1):
        src, delta = row_label[int(m.group(2))], m.group(3)
        base = vals[src][q - 1]
        return round(base + float(delta), 6) if delta else round(base, 6)
      return vals[row_label[int(m.group(5))]][q - 1]

    vals["Starting FTE"][q] = resolve(ws.cell(block["Starting FTE"], col).value, "Starting FTE")
    vals["Hires"][q] = float(ws.cell(block["Hires"], col).value or 0.0)
    vals["Ending FTE"][q] = vals["Starting FTE"][q] + vals["Hires"][q]
    vals["Annual Wage"][q] = resolve(ws.cell(block["Annual Wage"], col).value, "Annual Wage")
    vals["Benefits %"][q] = resolve(ws.cell(block["Benefits %"], col).value, "Benefits %")
  return vals


class _Built(unittest.TestCase):
  FIXTURE = HALBROOK

  @classmethod
  def setUpClass(cls):
    import openpyxl
    cls.fx = _fixture(cls.FIXTURE)
    cls.rows = _payroll_rows(cls.fx)
    cls.ids = _role_identity(cls.rows)
    cls.tmp = tempfile.mkdtemp(prefix="cw043_payroll_")
    cls.path = str(_build_workbook(cls.fx, cls.tmp))
    cls.ws = openpyxl.load_workbook(cls.path)["Payroll Schedule"]
    cls.blocks = _role_blocks(cls.ws)
    cls.by_identity = {}
    for (q, ident), item in zip(cls.ids, cls.rows):
      cls.by_identity.setdefault(ident, {})[q] = item

  @classmethod
  def tearDownClass(cls):
    shutil.rmtree(cls.tmp, ignore_errors=True)


class TheChainIsOnTheGridTests(_Built):
  """Halbrook - the 2-dp grid class."""

  def test_one_block_per_role_identity(self):
    self.assertEqual(len(self.blocks), len(self.by_identity))

  def test_every_chained_starting_fte_is_rounded_to_the_grid(self):
    chained = [self.ws.cell(b["Starting FTE"], c).value for b in self.blocks for c in range(4, 24)
               if isinstance(self.ws.cell(b["Starting FTE"], c).value, str)]
    self.assertTrue(chained, "no chained Starting FTE cells - the chain is not built")
    bad = [f for f in chained if not re.fullmatch(r"=ROUND\([A-Z]+\d+,6\)", f)]
    self.assertFalse(bad, f"Starting FTE chained without the 6-dp grid: {bad[:5]}")

  def test_the_emitted_chain_reproduces_the_engine_bit_for_bit(self):
    misses = []
    for block, (ident, by_q) in zip(self.blocks, self.by_identity.items()):
      vals = _evaluate_block(self.ws, block, by_q)
      for q, item in by_q.items():
        for label, key in (("Starting FTE", "starting_fte"), ("Annual Wage", "annual_wage"),
                           ("Benefits %", "payroll_taxes_benefits_percent")):
          if vals[label][q] != float(item.get(key) or 0.0):
            misses.append((ident[1][:20], q, key, float(item.get(key) or 0.0), vals[label][q]))
    self.assertFalse(misses, f"chain != engine: {misses[:6]}")

  def test_the_bare_chain_is_the_class_this_guard_exists_for(self):
    """CONTROL: prior Ending without the grid drifts off the engine's 2-dp
    values on this fixture - which is why the ROUND is not optional."""
    drift = 0
    for ident, by_q in self.by_identity.items():
      end = None
      for q in sorted(by_q):
        s = float(by_q[q].get("starting_fte") or 0.0)
        h = float(by_q[q].get("hires") or 0.0)
        contiguous = end is not None and (q - 1) in by_q
        if contiguous and end != s:
          drift += 1
        end = (end if contiguous else s) + h
    self.assertGreater(drift, 0, "fixture no longer exercises the 2-dp grid class - pick a draft that does")

  def test_the_bridge_has_one_hidden_formula_row_per_engine_row_in_engine_order(self):
    first, last = _bridge_rows(self.ws)
    self.assertEqual(last - first + 1, len(self.rows))
    for i, (q, ident) in enumerate(self.ids):
      r = first + i
      self.assertEqual(self.ws.cell(r, 1).value, q)
      self.assertTrue(self.ws.row_dimensions[r].hidden, f"bridge row {r} is visible")
      for c in range(2, 15):
        v = self.ws.cell(r, c).value
        self.assertTrue(isinstance(v, str) and v.startswith("="), f"bridge {r},{c} is not a formula: {v!r}")

  def test_the_bridge_wage_source_resolves_to_this_quarters_label(self):
    """Wage source is written only where it CHANGES (2026-08-27), so the
    bridge points at the cell that holds the label in force for its quarter -
    the most recent write at or before that quarter - and the TEXT it
    resolves to must equal the engine's label for that row."""
    from openpyxl.utils import column_index_from_string
    first, _ = _bridge_rows(self.ws)
    block_of = dict(zip(self.by_identity.keys(), self.blocks))
    wrong = []
    for i, (q, ident) in enumerate(self.ids):
      block = block_of[ident]
      src_row = block["Wage source"]
      # the label in force: the last written cell at or before this quarter
      effective_col, effective_text = None, None
      for c in range(PERIOD_START_COL + 1, PERIOD_START_COL + q + 1):
        v = self.ws.cell(src_row, c).value
        if v not in (None, ""):
          effective_col, effective_text = c, v
      v = str(self.ws.cell(first + i, 14).value)
      m = re.fullmatch(r'=IF\(([A-Z]+)(\d+)="","",\1\2\)', v)
      self.assertIsNotNone(m, f"bridge wage-source cell {first + i} is not a guarded reference: {v!r}")
      got = (int(m.group(2)), column_index_from_string(m.group(1)))
      engine = _wage_source_expected(self.rows[i])
      if got != (src_row, effective_col) or effective_text != engine:
        wrong.append((first + i, got, "expected", (src_row, effective_col), effective_text, engine))
    self.assertFalse(wrong, f"bridge wage source does not resolve to this quarter's label: {wrong[:4]}")

  def test_the_visible_numeric_surface_is_four_inputs_per_engine_row(self):
    import openpyxl
    amber = str(openpyxl.load_workbook(self.path)["Debt Schedule"].cell(8, 3).fill.fgColor.rgb)[-6:]
    got = 0
    for block in self.blocks:
      for label in ("Starting FTE", "Hires", "Annual Wage", "Benefits %"):
        got += sum(1 for c in range(4, 24) if str(self.ws.cell(block[label], c).fill.fgColor.rgb)[-6:] == amber)
    self.assertEqual(got, 4 * len(self.rows))


class TheChainFollowsThePersonNotTheTitleTests(_Built):
  """Understory - two Grow Technicians, different wages."""
  FIXTURE = UNDERSTORY

  def test_the_fixture_has_two_people_with_one_title_and_different_wages(self):
    q1 = [it for it in self.rows if int(it.get("quarter_index") or 0) == 1]
    grow = [it for it in q1 if it.get("position_title") == "Grow Technician"]
    self.assertEqual(len(grow), 2, "fixture no longer carries the duplicate-title class")
    self.assertNotEqual(grow[0].get("annual_wage"), grow[1].get("annual_wage"))

  def test_each_person_has_their_own_block(self):
    self.assertEqual(len(self.blocks), len(self.by_identity),
                     "two people with one title collapsed into one block")

  def test_each_grow_technician_keeps_their_own_wage_down_the_chain(self):
    misses = []
    for block, (ident, by_q) in zip(self.blocks, self.by_identity.items()):
      if ident[1] != "Grow Technician":
        continue
      vals = _evaluate_block(self.ws, block, by_q)
      for q, item in by_q.items():
        if vals["Annual Wage"][q] != float(item.get("annual_wage") or 0.0):
          misses.append((ident[2], q, float(item.get("annual_wage")), vals["Annual Wage"][q]))
    self.assertFalse(misses, f"a Grow Technician inherited the other's wage: {misses[:4]}")

  def test_every_bridge_cell_points_at_its_own_identitys_block(self):
    """F3 (mini, 2026-08-25): ALL thirteen bridge formulas of every engine
    row - not only Annual Wage - reference the block of that row's own
    identity, at the right row for the column's label and the right quarter
    column. A per-column regression would sum a same-title twin's FTE or
    payroll into the summary and FINMO unseen."""
    from openpyxl.utils import column_index_from_string
    first, last = _bridge_rows(self.ws)
    block_of = dict(zip(self.by_identity.keys(), self.blocks))
    expect = {2: ("header", 2), 3: ("header", 1),
              5: ("Starting FTE", None), 6: ("Hires", None), 7: ("Ending FTE", None),
              8: ("Average FTE", None), 9: ("Annual Wage", None), 10: ("Benefits %", None),
              11: ("Wage Cost", None), 12: ("Taxes & Benefits", None), 13: ("Total Payroll", None)}
    wrong = []
    for i, (q, ident) in enumerate(self.ids):
      r = first + i
      block = block_of[ident]
      for col, (label, fixed_col) in expect.items():
        v = str(self.ws.cell(r, col).value)
        m = re.search(r"([A-Z]+)(\d+)", v)
        self.assertIsNotNone(m, f"bridge {r},{col} has no reference: {v!r}")
        ref_col, ref_row = column_index_from_string(m.group(1)), int(m.group(2))
        want_row = block[label]
        want_col = fixed_col if fixed_col else PERIOD_START_COL + q
        if (ref_row, ref_col) != (want_row, want_col):
          wrong.append((r, col, ident[2] or ident[1], "refs", (ref_row, ref_col), "expected", (want_row, want_col)))
      # column 4 (OEWS title): a reference into THIS block's oews row, or an
      # empty string where the role has no OEWS title (a named person).
      v4 = str(self.ws.cell(r, 4).value)
      if "oews" in block:
        m4 = re.search(r"([A-Z]+)(\d+)", v4)
        self.assertIsNotNone(m4, f"bridge {r},4 has no reference: {v4!r}")
        if int(m4.group(2)) != block["oews"]:
          wrong.append((r, 4, ident[2] or ident[1], "oews refs", int(m4.group(2)), "expected", block["oews"]))
      elif v4 != '=""':
        wrong.append((r, 4, ident[2] or ident[1], "expected an empty OEWS cell, got", v4))
      # column 14 (wage source): into THIS block's Wage source row
      m14 = re.search(r"([A-Z]+)(\d+)", str(self.ws.cell(r, 14).value))
      self.assertIsNotNone(m14, f"bridge {r},14 has no reference")
      if int(m14.group(2)) != block["Wage source"]:
        wrong.append((r, 14, ident[2] or ident[1], "source refs", int(m14.group(2)), "expected", block["Wage source"]))
    self.assertFalse(wrong, f"bridge cells pointing outside their own block: {wrong[:4]}")

  def test_every_bridge_row_points_at_its_own_persons_block(self):
    first, last = _bridge_rows(self.ws)
    block_of = dict(zip(self.by_identity.keys(), self.blocks))
    wrong = []
    for i, (q, ident) in enumerate(self.ids):
      r = first + i
      v = self.ws.cell(r, 9).value  # Annual Wage -> '=<col><Annual Wage row of the block>'
      m = re.fullmatch(r"=([A-Z]+)(\d+)", str(v))
      self.assertIsNotNone(m, f"bridge wage cell {r} is not a plain reference: {v!r}")
      if int(m.group(2)) != block_of[ident]["Annual Wage"]:
        wrong.append((r, ident[2], "points at row", int(m.group(2)), "expected", block_of[ident]["Annual Wage"]))
    self.assertFalse(wrong, f"bridge rows pointing at another person's block: {wrong[:4]}")


def _excel_available() -> bool:
  try:
    import win32com.client as win32
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Quit()
    return True
  except Exception:
    return False


@unittest.skipUnless(_excel_available(), "Excel is required to evaluate the workbook's formulas")
class TheRecalculatedBridgeEqualsTheEngineTests(unittest.TestCase):

  @classmethod
  def setUpClass(cls):
    import openpyxl
    import win32com.client as win32
    cls.fx = _fixture(HALBROOK)
    cls.rows = _payroll_rows(cls.fx)
    cls.tmp = tempfile.mkdtemp(prefix="cw043_payroll_xl_")
    path = str(_build_workbook(cls.fx, cls.tmp))
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Visible = False
    x.DisplayAlerts = False
    wb = x.Workbooks.Open(path)
    for _ in range(20):
      try:
        wb.Sheets(1).Name
        break
      except Exception:
        time.sleep(1.5)
    x.CalculateFullRebuild()
    wb.Save()
    wb.Close(False)
    x.Quit()
    cls.ws = openpyxl.load_workbook(path, data_only=True)["Payroll Schedule"]
    cls.first, cls.last = _bridge_rows(openpyxl.load_workbook(path)["Payroll Schedule"])

  @classmethod
  def tearDownClass(cls):
    shutil.rmtree(cls.tmp, ignore_errors=True)

  def test_recalculated_bridge_inputs_equal_the_engine_bit_for_bit(self):
    misses = []
    for i, item in enumerate(self.rows):
      r = self.first + i
      for col, key in ((5, "starting_fte"), (6, "hires"), (9, "annual_wage"), (10, "payroll_taxes_benefits_percent")):
        got = float(self.ws.cell(r, col).value or 0.0)
        engine = float(item.get(key) or 0.0)
        if got != engine:
          misses.append((r, key, engine, got))
    self.assertFalse(misses, f"Excel bridge != engine: {misses[:6]}")


if __name__ == "__main__":
  unittest.main()
