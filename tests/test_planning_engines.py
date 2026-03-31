from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
PYTHON_DIR = ROOT / "python"
CLIENT_DIR = PYTHON_DIR / "client_intake_and_finmo"
for path in (str(PYTHON_DIR), str(CLIENT_DIR)):
  if path not in sys.path:
    sys.path.insert(0, path)


from api_handlers.intake_consult import (  # type: ignore  # noqa: E402
  _start_consistency_governance_if_needed,
  _build_consistency_finmo_attempts_payload,
  _build_consistency_modified_plan_payload,
  _serialize_debug_draft_row,
)
from api_handlers.shared_context import build_shared_context  # type: ignore  # noqa: E402
from consistency_flow import (  # type: ignore  # noqa: E402
  _select_best_finmo_attempts,
  _build_strategy_layer,
  apply_consistency_selected_path,
  build_consistency_governance_state,
)
from consistency_strategy_advisor import _normalize_strategy_selection_contract, advise_consistency_strategy_selection  # type: ignore  # noqa: E402
from client_intake_and_finmo.consistency_flow.controller import _gpt_blueprint_is_usable, _selection_lever_adjustment_plan  # type: ignore  # noqa: E402
from client_intake_and_finmo.consistency_flow.finmo_controller import (  # type: ignore  # noqa: E402
  _build_finmo_calibration_spec,
  _build_controller_input_seed_from_profile,
  build_controller_finmo_candidate,
)
from client_intake_and_finmo.finmo_bridge import build_python_finmo_json, build_python_model_input_json, build_consistency_forecast_view_from_finmo, normalize_model_input_forecast_anchor, _build_model_input_overlay, _execute_finmo_calibration_shell, _forecast_anchor_date_iso, _read_model_input_json, _write_model_input_json_to_workbook  # type: ignore  # noqa: E402
from client_intake_and_finmo.quarter_grid import (  # type: ignore  # noqa: E402
  available_planning_modes,
  classify_planning_mode,
  chunk_quarter_grid_rows,
  controls_from_quarter_grid,
  extract_quarter_grid_rows,
  planning_mode_text,
  resolve_planning_mode,
  targets_from_quarter_grid,
  validate_quarter_grid_response,
)
from intake_pipeline import _ensure_submission_finmo_path  # type: ignore  # noqa: E402


class PlanningEnginesTests(unittest.TestCase):
  def test_consistency_governance_prefers_authoritative_finmo_path(self) -> None:
    captured: dict = {}

    def _fake_build_consistency_governance_state(**kwargs):
      captured.update(kwargs)
      return {"status": "ok", "scenarios": []}

    with patch("consistency_flow.build_consistency_governance_state", side_effect=_fake_build_consistency_governance_state):
      governance_state, _runtime_payload = _start_consistency_governance_if_needed(
        authoritative_finmo_path="C:\\authoritative\\client.xlsx",
        intake_context={
          "finmo_path": "",
          "business_name": "CareFirst",
          "business_start_date": "2026-09-03",
          "address": "123 Main",
          "shared_context": {},
        },
        ops_json={},
        market_json={},
        people_json={},
        financials_json={},
        financials_year1_json={},
        marketing_model_json={},
      )

    self.assertEqual(governance_state, {"status": "ok", "scenarios": []})
    self.assertEqual(captured.get("finmo_path"), "C:\\authoritative\\client.xlsx")

  def test_selection_lever_adjustment_plan_preserves_gpt_bands(self) -> None:
    plan = _selection_lever_adjustment_plan(
      {
        "lever_adjustment_plan": [
          {
            "lever_id": "expenses::Payroll",
            "direction": "down",
            "quarter_start": 1,
            "quarter_end": 4,
            "min_value": 40000,
            "max_value": 60000,
            "rationale": "Trim early payroll",
          }
        ]
      }
    )

    self.assertEqual(len(plan), 1)
    self.assertEqual(plan[0]["min_value"], 40000)
    self.assertEqual(plan[0]["max_value"], 60000)

  def test_controller_input_seed_prefers_workbook_revenue_identity_and_slot_key(self) -> None:
    seed = _build_controller_input_seed_from_profile(
      profile={},
      direct_inputs={
        "product_driver_basis": [
          {
            "lob_name": "Primary",
            "product_name": "Product 1",
            "annual_capacity_units": 5200.0,
            "unit_price": 80.0,
            "utilization_rate": 0.75,
            "annual_revenue": 312000.0,
          }
        ],
        "model_input_json": {
          "lever_catalog": {
            "revenue::Primary line of business::In-home care hour::Capacity": {
              "section": "revenue",
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Capacity",
              "revenue_slot_key": "lob_1_product_1",
              "lob_slot_index": 0,
              "product_slot_index": 0,
            }
          }
        },
      },
      baseline_state={
        "financials_json": {
          "cogs_total_year1": 241427.43,
          "payroll_total_year1": 249990.0,
          "other_opex_absolute": 115280.94,
          "monthly_rent_expense": 1200.0,
        },
        "people_json": {"people": []},
        "marketing_model_json": {"marketing_percent_of_revenue": 0.15},
        "financials_year1_json": {"company_revenue_total_year1": 312000.0},
      },
    )

    revenue_product = (((seed[0] or {}).get("revenue_products") or [])[0]["products"] or [])[0]
    revenue_lob = (((seed[0] or {}).get("revenue_products") or [])[0] or {})
    self.assertEqual(revenue_lob["lob_name"], "Primary line of business")
    self.assertEqual(revenue_product["product_name"], "In-home care hour")
    self.assertEqual(revenue_product["revenue_slot_key"], "lob_1_product_1")
    self.assertEqual(revenue_product["capacity_units"], 1300.0)

  def test_build_model_input_overlay_uses_revenue_slot_key_when_names_drift(self) -> None:
    overlay = _build_model_input_overlay(
      baseline_model_input={
        "periods": [
          {"quarter": 0, "column_index": 6},
          {"quarter": 1, "column_index": 7},
          {"quarter": 2, "column_index": 8},
        ],
        "sections": {
          "revenue": [
            {
              "placeholder_lob": "LOB 1",
              "placeholder_product": "Product 1",
              "revenue_slot_key": "lob_1_product_1",
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Capacity",
              "values": [],
            },
            {
              "placeholder_lob": "LOB 1",
              "placeholder_product": "Product 1",
              "revenue_slot_key": "lob_1_product_1",
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Unit Price",
              "values": [],
            },
            {
              "placeholder_lob": "LOB 1",
              "placeholder_product": "Product 1",
              "revenue_slot_key": "lob_1_product_1",
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Utilization",
              "values": [],
            },
          ],
          "expenses": [],
          "balance_sheet": [],
          "schedules": {"rows": []},
        },
      },
      business_facts={"start_date": "2026-09-03"},
      ops_json={
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "In-home care hour",
                "unit_price": 80.0,
                "utilization_rate": 0.75,
                "operating_periods_per_year": 52,
                "units_per_period_capacity": 100,
              }
            ],
          }
        ]
      },
      people_json={"people": []},
      financials_json={},
      financials_year1_json={"company_revenue_total_year1": 312000.0},
      marketing_model_json={},
      controller_input_seed=[
        {
          "quarter_index": 1,
          "revenue_products": [
            {
              "lob_name": "Primary",
              "products": [
                {
                  "product_name": "Product 1",
                  "revenue_slot_key": "lob_1_product_1",
                  "capacity_units": 1300.0,
                  "price": 80.8,
                  "utilization": 0.675,
                }
              ],
            }
          ],
        },
        {
          "quarter_index": 2,
          "revenue_products": [
            {
              "lob_name": "Primary",
              "products": [
                {
                  "product_name": "Product 1",
                  "revenue_slot_key": "lob_1_product_1",
                  "capacity_units": 1400.0,
                  "price": 81.6,
                  "utilization": 0.69,
                }
              ],
            }
          ],
        },
      ],
      forecast_quarters=[],
    )

    revenue_rows = overlay["sections"]["revenue"]
    self.assertEqual(revenue_rows[0]["lob"], "Primary line of business")
    self.assertEqual(revenue_rows[0]["product"], "In-home care hour")
    self.assertEqual(revenue_rows[0]["values"], [1300.0, 1400.0])
    self.assertEqual(revenue_rows[1]["values"], [80.8, 81.6])
    self.assertEqual(revenue_rows[2]["values"], [0.675, 0.69])

  def test_write_model_input_json_to_workbook_matches_revenue_rows_by_slot_key(self) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Model Inputs"

    for label, row_idx in [("Days in QTR", 1), ("Date", 2), ("Quarter", 3), ("Year", 4), ("Year Fraction", 5)]:
      ws.cell(row=row_idx, column=5).value = label
    period_columns = [6, 7, 8]
    for col_idx, quarter, year in [(6, 0, 2026), (7, 1, 2026), (8, 2, 2027)]:
      ws.cell(row=2, column=col_idx).value = "2026-09-03"
      ws.cell(row=3, column=col_idx).value = quarter
      ws.cell(row=4, column=col_idx).value = year
      ws.cell(row=5, column=col_idx).value = 0.25
    revenue_rows = [
      ("Primary line of business", "In-home care hour", "Capacity", 7),
      ("Primary line of business", "In-home care hour", "Unit Price", 8),
      ("Primary line of business", "In-home care hour", "Utilization", 9),
    ]
    for lob, product, driver, row_idx in revenue_rows:
      ws.cell(row=row_idx, column=1).value = "Controller write"
      ws.cell(row=row_idx, column=3).value = lob
      ws.cell(row=row_idx, column=4).value = product
      ws.cell(row=row_idx, column=5).value = driver
    ws.cell(row=12, column=1).value = "Controller write"
    ws.cell(row=12, column=2).value = "Marketing"
    ws.cell(row=14, column=1).value = "Controller write"
    ws.cell(row=14, column=2).value = "Accounts Receivable Days"
    ws.cell(row=16, column=2).value = "Closing Balance"
    ws.cell(row=17, column=2).value = "Closing Balance (Total)"
    ws.cell(row=18, column=2).value = "Plus: Additions (repayments), net"

    wb.defined_names["model_input_periods"] = DefinedName("model_input_periods", attr_text="'Model Inputs'!$D$1:$H$5")
    wb.defined_names["model_input_startdate"] = DefinedName("model_input_startdate", attr_text="'Model Inputs'!$B$1")
    wb.defined_names["model_input_revenue"] = DefinedName("model_input_revenue", attr_text="'Model Inputs'!$A$7:$H$9")
    wb.defined_names["model_input_expenses"] = DefinedName("model_input_expenses", attr_text="'Model Inputs'!$A$12:$H$12")
    wb.defined_names["model_input_balancehseet"] = DefinedName("model_input_balancehseet", attr_text="'Model Inputs'!$A$14:$H$14")
    wb.defined_names["model_input_schedules"] = DefinedName("model_input_schedules", attr_text="'Model Inputs'!$A$16:$H$18")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
      workbook_path = handle.name
    try:
      wb.save(workbook_path)
      wb.close()

      _write_model_input_json_to_workbook(
        workbook_path,
        {
          "sections": {
            "revenue": [
              {
                "revenue_slot_key": "lob_1_product_1",
                "placeholder_lob": "LOB 1",
                "placeholder_product": "Product 1",
                "lob": "Primary line of business",
                "product": "In-home care hour",
                "driver": "Capacity",
                "values": [1300.0, 1400.0],
              },
              {
                "revenue_slot_key": "lob_1_product_1",
                "placeholder_lob": "LOB 1",
                "placeholder_product": "Product 1",
                "lob": "Primary line of business",
                "product": "In-home care hour",
                "driver": "Unit Price",
                "values": [80.0, 81.0],
              },
              {
                "revenue_slot_key": "lob_1_product_1",
                "placeholder_lob": "LOB 1",
                "placeholder_product": "Product 1",
                "lob": "Primary line of business",
                "product": "In-home care hour",
                "driver": "Utilization",
                "values": [0.75, 0.8],
              },
            ],
            "expenses": [],
            "balance_sheet": [],
            "schedules": {"rows": []},
          }
        },
      )
      model_input_json = _read_model_input_json(workbook_path)
      revenue_rows = model_input_json["sections"]["revenue"]
      self.assertEqual(revenue_rows[0]["values"][-2:], [1300.0, 1400.0])
      self.assertEqual(revenue_rows[1]["values"][-2:], [80.0, 81.0])
      self.assertEqual(revenue_rows[2]["values"][-2:], [0.75, 0.8])
    finally:
      Path(workbook_path).unlink(missing_ok=True)

  def test_build_python_model_input_json_uses_dynamic_intake_revenue_rows_only(self) -> None:
    payload = build_python_model_input_json(
      business_facts={"start_date": "2026-03-28", "business_name": "Sunrise Home Health Care"},
      ops_json={
        "business_type": "Home Health Agency",
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "visit of care",
                "unit_name": "visit of care",
                "units_per_week_capacity": 100,
                "operating_periods_per_year": 52,
                "unit_price": None,
                "utilization_rate": None,
              }
            ],
          }
        ],
      },
      people_json={},
      financials_json={},
      financials_year1_json={},
      marketing_model_json={},
      controller_input_seed=[],
      forecast_quarters=[],
    )

    revenue_rows = payload["sections"]["revenue"]
    self.assertEqual(len(revenue_rows), 3)
    self.assertEqual(
      [row["lever_id"] for row in revenue_rows],
      [
        "revenue::Primary line of business::visit of care::Capacity",
        "revenue::Primary line of business::visit of care::Unit Price",
        "revenue::Primary line of business::visit of care::Utilization",
      ],
    )
    self.assertEqual(payload["start_date"], _forecast_anchor_date_iso())
    self.assertEqual(payload["business_start_date"], "2026-03-28")
    self.assertEqual(len(payload["periods"]), 20)
    self.assertEqual(payload["periods"][0]["column_letter"], "H")
    self.assertEqual(payload["periods"][-1]["column_letter"], "AA")
    self.assertIn("expenses::Payroll", payload["lever_catalog"])
    self.assertIn("balance_sheet::Owner's Capital", payload["lever_catalog"])
    self.assertIn("schedules::Less: Principal Repayments", payload["lever_catalog"])

  def test_normalize_model_input_forecast_anchor_reanchors_period_years(self) -> None:
    payload = {
      "start_date": "2010-06-15",
      "business_start_date": "2010-06-15",
      "periods": [
        {"slot_index": 0, "column_index": 8, "column_letter": "H", "year": 2010.0, "quarter": 1.0, "date": "2010-06-15"},
        {"slot_index": 1, "column_index": 9, "column_letter": "I", "year": 2010.0, "quarter": 2.0, "date": "2010-09-15"},
      ],
      "sections": {"revenue": [], "expenses": [], "balance_sheet": [], "schedules": {"rows": []}},
    }

    normalized = normalize_model_input_forecast_anchor(payload, anchor_date_iso="2026-03-30")

    self.assertEqual(normalized["start_date"], "2026-03-30")
    self.assertEqual(normalized["business_start_date"], "2010-06-15")
    self.assertEqual(len(normalized["periods"]), 2)
    self.assertEqual(normalized["periods"][0]["date"], "2026-03-30")
    self.assertEqual(normalized["periods"][0]["year"], 2026.0)
    self.assertEqual(normalized["periods"][1]["year"], 2026.0)

  def test_build_python_model_input_json_derives_non_revenue_baselines_from_inputs(self) -> None:
    payload = build_python_model_input_json(
      business_facts={"start_date": "2026-03-28", "business_name": "CareFirst"},
      ops_json={
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "visit of care",
                "units_per_week_capacity": 100,
                "operating_periods_per_year": 52,
              }
            ],
          }
        ],
      },
      people_json={"people": [{"annual_wage": 52000.0}, {"annual_wage": 78000.0}]},
      financials_json={
        "monthly_rent_expense": 2000.0,
        "cogs_total_year1": 180000.0,
        "current_revenue": 400000.0,
        "other_operating_expense": 56000.0,
        "annual_interest_payment": 2400.0,
        "total_debt_outstanding": 120000.0,
        "accumulated_depreciation": 12000.0,
        "taxes_percent": 0.21,
        "initial_equity": 50000.0,
        "short_term_debt": 12000.0,
      },
      financials_year1_json={"company_revenue_total_year1": 400000.0},
      marketing_model_json={"marketing_percent_of_revenue": 0.08},
      controller_input_seed=[],
      forecast_quarters=[],
    )

    expense_rows = {row["label"]: row for row in payload["sections"]["expenses"]}
    balance_rows = {row["label"]: row for row in payload["sections"]["balance_sheet"]}
    self.assertEqual(expense_rows["Lease"]["values"][:3], [6000.0, 6000.0, 6000.0])
    self.assertEqual(expense_rows["Payroll"]["values"][:3], [32500.0, 32500.0, 32500.0])
    self.assertEqual(expense_rows["Marketing"]["values"][:3], [0.08, 0.08, 0.08])
    self.assertEqual(expense_rows["Cost of Goods Sold"]["values"][:3], [0.45, 0.45, 0.45])
    self.assertEqual(balance_rows["Owner's Capital"]["values"][:3], [50000.0, 50000.0, 50000.0])

  def test_build_python_finmo_json_matches_finmo_contract_shape(self) -> None:
    model_input_json = build_python_model_input_json(
      business_facts={"start_date": "2026-03-28", "business_name": "CareFirst"},
      ops_json={
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "visit of care",
                "units_per_week_capacity": 100,
                "operating_periods_per_year": 52,
                "unit_price": 100.0,
                "utilization_rate": 0.75,
              }
            ],
          }
        ],
      },
      people_json={"people": [{"annual_wage": 120000.0}]},
      financials_json={
        "monthly_rent_expense": 2000.0,
        "cogs_total_year1": 180000.0,
        "current_revenue": 400000.0,
        "other_operating_expense": 56000.0,
        "annual_interest_payment": 2400.0,
        "total_debt_outstanding": 120000.0,
        "accumulated_depreciation": 12000.0,
        "taxes_percent": 0.21,
        "initial_equity": 50000.0,
        "short_term_debt": 12000.0,
      },
      financials_year1_json={"company_revenue_total_year1": 400000.0},
      marketing_model_json={"marketing_percent_of_revenue": 0.08},
      controller_input_seed=[],
      forecast_quarters=[],
    )

    finmo_json = build_python_finmo_json(model_input_json=model_input_json, finmo_path="C:\\fake\\carefirst.xlsx")
    self.assertEqual(finmo_json["contract_version"], "finmo_output_v1")
    self.assertEqual(finmo_json["finmo_path"], "C:\\fake\\carefirst.xlsx")
    self.assertEqual(len(finmo_json["periods"]), 21)
    self.assertEqual(len(finmo_json["quarter_rows"]), 21)
    self.assertEqual(finmo_json["periods"][0]["quarter"], 0.0)
    self.assertEqual(finmo_json["quarter_rows"][0]["quarter"], 0.0)
    self.assertEqual(finmo_json["pl"][0]["label"], "Revenue")
    self.assertEqual(finmo_json["balance_sheet"][0]["label"], "Cash")
    self.assertEqual(finmo_json["cash_flow"][0]["label"], "Beginning Cash")
    self.assertIn("all_ok", finmo_json["accounting_check"])
    self.assertIn("numeric_values", finmo_json["accounting_check"])

    forecast_view = build_consistency_forecast_view_from_finmo(finmo_json)
    self.assertEqual(len(forecast_view["quarter_driver_path"]), 20)
    self.assertEqual(len(forecast_view["forecast_years"]), 5)
    self.assertEqual(forecast_view["quarter_driver_path"][0]["quarter_index"], 1)

  def test_build_python_model_input_json_infers_revenue_drivers_from_forecast_quarters(self) -> None:
    forecast_quarters = []
    for quarter_index in range(1, 21):
      revenue = 90000.0 if quarter_index <= 4 else 120000.0
      cogs = revenue * 0.55
      marketing = revenue * 0.05
      payroll = 40000.0 if quarter_index <= 4 else 57500.0
      g_and_a = revenue * 0.15 if quarter_index <= 4 else revenue * 0.12
      lease = 3600.0
      gross_profit = revenue - cogs
      ebitda = gross_profit - marketing - payroll - g_and_a - lease
      forecast_quarters.append(
        {
          "quarter_index": quarter_index,
          "revenue": revenue,
          "cogs": cogs,
          "marketing": marketing,
          "research_and_development": 0.0,
          "lease_rent": lease,
          "payroll": payroll,
          "g_and_a": g_and_a,
          "ebitda": ebitda,
          "interest": 0.0,
          "depreciation": 0.0,
          "taxes": 0.0,
          "net_income": ebitda,
          "ending_cash": ebitda,
        }
      )

    model_input_json = build_python_model_input_json(
      business_facts={"start_date": "2026-03-28", "business_name": "CareFirst"},
      ops_json={
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "visit of care",
                "units_per_week_capacity": 100,
                "operating_periods_per_year": 52,
                "unit_price": 80.0,
                "utilization_rate": 0.75,
              }
            ],
          }
        ],
      },
      people_json={"people": [{"annual_wage": 249990.0}]},
      financials_json={"monthly_rent_expense": 1200.0},
      financials_year1_json={},
      marketing_model_json={},
      controller_input_seed=[],
      forecast_quarters=forecast_quarters,
    )
    finmo_json = build_python_finmo_json(model_input_json=model_input_json)
    forecast_view = build_consistency_forecast_view_from_finmo(finmo_json)
    quarter_rows = [row for row in ((model_input_json.get("sections") or {}).get("revenue") or []) if isinstance(row, dict)]
    capacity_values = next(
      row["values"] for row in quarter_rows
      if row.get("lever_id") == "revenue::Primary line of business::visit of care::Capacity"
    )
    self.assertGreater(capacity_values[0], 0.0)
    self.assertAlmostEqual(forecast_view["quarter_driver_path"][0]["revenue"], 90000.0, places=3)
    self.assertAlmostEqual(forecast_view["quarter_driver_path"][7]["revenue"], 120000.0, places=3)
    self.assertTrue(finmo_json["accounting_check"]["all_ok"])

  def test_quarter_grid_extracts_live_rows_and_converts_to_solver_contract(self) -> None:
    model_input_json = build_python_model_input_json(
      business_facts={"start_date": "2026-03-28", "business_name": "CareFirst"},
      ops_json={
        "lob_models": [
          {
            "lob_name": "Primary line of business",
            "products": [
              {
                "product_name": "visit of care",
                "units_per_week_capacity": 100,
                "operating_periods_per_year": 52,
                "unit_price": 100.0,
                "utilization_rate": 0.75,
              }
            ],
          }
        ],
      },
      people_json={"people": [{"annual_wage": 120000.0}]},
      financials_json={
        "monthly_rent_expense": 2000.0,
        "cogs_total_year1": 180000.0,
        "current_revenue": 400000.0,
        "other_operating_expense": 56000.0,
        "initial_equity": 50000.0,
      },
      financials_year1_json={"company_revenue_total_year1": 400000.0},
      marketing_model_json={"marketing_percent_of_revenue": 0.08},
      controller_input_seed=[],
      forecast_quarters=[],
    )
    finmo_json = build_python_finmo_json(model_input_json=model_input_json)
    baseline_outputs = [item for item in (finmo_json.get("quarter_rows") or []) if isinstance(item, dict) and float(item.get("quarter") or 0.0) > 0.0]
    grid_rows = extract_quarter_grid_rows(model_input_json=model_input_json, baseline_outputs=baseline_outputs)
    self.assertTrue(any(item["row_id"] == "Revenue" for item in grid_rows))
    self.assertTrue(any(item["row_id"] == "EBITDA" for item in grid_rows))
    self.assertTrue(any(str(item["row_id"]).startswith("revenue::") for item in grid_rows))

    probe_json = {
      "summary": "ok",
      "rows": [
        {
          "row_id": "revenue::Primary line of business::visit of care::Unit Price",
          "row_type": "lever",
          "quarter_bands": [
            {"quarter_index": quarter_index, "min_value": 90.0, "max_value": 110.0}
            for quarter_index in range(1, 21)
          ],
        },
        {
          "row_id": "Revenue",
          "row_type": "output",
          "quarter_bands": [
            {"quarter_index": quarter_index, "min_value": 80000.0, "max_value": 120000.0}
            for quarter_index in range(1, 21)
          ],
        },
      ],
    }
    controls = controls_from_quarter_grid(probe_json)
    targets = targets_from_quarter_grid(probe_json)
    self.assertEqual(len(controls), 20)
    self.assertEqual(len(targets), 20)
    self.assertEqual(controls[0].lever_id, "revenue::Primary line of business::visit of care::Unit Price")
    self.assertEqual(targets[0].metric, "Revenue")

  def test_quarter_grid_validation_flags_missing_and_flat_rows(self) -> None:
    requested_rows = [
      {"row_id": "expenses::Payroll", "row_type": "lever"},
      {"row_id": "EBITDA", "row_type": "output"},
    ]
    response_json = {
      "summary": "bad",
      "rows": [
        {
          "row_id": "expenses::Payroll",
          "row_type": "lever",
          "quarter_bands": [
            {"quarter_index": quarter_index, "min_value": 40000.0, "max_value": 40000.0}
            for quarter_index in range(1, 21)
          ],
        }
      ],
    }
    validation = validate_quarter_grid_response(requested_rows=requested_rows, response_json=response_json)
    self.assertEqual(validation["missing_rows"], ["EBITDA"])
    self.assertEqual(validation["flat_rows"], ["expenses::Payroll"])
    self.assertEqual(len(chunk_quarter_grid_rows(requested_rows * 2, 2)), 2)

  def test_quarter_grid_prompt_modes_are_file_backed_and_resolved(self) -> None:
    modes = available_planning_modes()
    self.assertIn("turnaround", modes)
    self.assertIn("normalize", modes)
    self.assertIn("rebalance", modes)
    self.assertEqual(resolve_planning_mode("TURNAROUND"), "turnaround")
    self.assertEqual(resolve_planning_mode("unknown-mode"), "turnaround")
    self.assertIn("profitable as soon as it can become profitable", planning_mode_text("turnaround"))
    self.assertIn("timing_months_max", planning_mode_text("turnaround"))
    self.assertIn("months_until_hire", planning_mode_text("turnaround"))
    self.assertIn("over-optimistic or commercially overstated", planning_mode_text("normalize"))
    self.assertIn("timing_months_max", planning_mode_text("normalize"))
    self.assertIn("months_until_hire", planning_mode_text("rebalance"))

  def test_classify_planning_mode_keeps_internal_diagnosis_app_side(self) -> None:
    turnaround = classify_planning_mode(
      baseline_summary={"revenue": 100000.0, "ebitda": -30000.0},
      diagnosis={"severity_class": "severe", "primary_cause": "payroll-driven"},
    )
    normalize = classify_planning_mode(
      baseline_summary={"revenue": 100000.0, "ebitda": 45000.0},
      diagnosis={"severity_class": "moderate", "preferred_strategy_ids": ["reality_normalization_strategy"]},
    )
    rebalance = classify_planning_mode(
      baseline_summary={"revenue": 100000.0, "ebitda": 8000.0},
      diagnosis={"severity_class": "moderate", "primary_cause": "mixed"},
    )
    self.assertEqual(turnaround["planning_mode"], "turnaround")
    self.assertEqual(normalize["planning_mode"], "normalize")
    self.assertEqual(rebalance["planning_mode"], "rebalance")

  def test_strategy_selection_contract_normalizes_to_workbook_levers_and_finmo_lines(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha", "beta", "missing"],
      "allowed_model_input_levers": [
        "expenses::Marketing",
        "bad-token",
        "revenue::Home Health::Skilled Nursing::Unit Price",
      ],
      "forbidden_model_input_levers": ["expenses::Payroll", "legacy_lever"],
      "governed_period_groups": [
        {"quarter_start": 0, "quarter_end": 4},
        {
          "quarter_start": 5,
          "quarter_end": 20,
          "quarterly_expansion_levers": [
            "expenses::Marketing",
            "legacy_name",
          ],
        },
      ],
      "lever_adjustment_plan": [
        {
          "lever_id": "expenses::Marketing",
          "direction": "decrease",
          "quarter_start": 1,
          "quarter_end": 4,
        },
        {
          "lever_id": "legacy_lever",
          "direction": "increase",
          "quarter_start": 1,
          "quarter_end": 4,
        },
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "goal_band": {"min": 0, "max": 50}},
        {"line_item": "Legacy Output", "goal_band": {"min": 0, "max": 1}},
      ],
      "active_levers": ["expenses::Marketing", "legacy_lever"],
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": [
          "expenses::Marketing",
          "revenue::Home Health::Skilled Nursing::Unit Price",
        ],
      },
      {
        "strategy_id": "beta",
        "allowed_model_input_levers": ["expenses::Payroll"],
      },
    ]
    fixed_facts = {
      "finmo_json": {
        "pl": [{"label": "Revenue"}, {"label": "EBITDA"}],
        "balance_sheet": [{"label": "Cash"}],
        "cash_flow": [],
      }
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertEqual(normalized["selected_strategy_ids"], ["alpha", "beta"])
    self.assertEqual(
      normalized["allowed_model_input_levers"],
      [
        "expenses::Marketing",
        "revenue::Home Health::Skilled Nursing::Unit Price",
      ],
    )
    self.assertEqual(normalized["forbidden_model_input_levers"], ["expenses::Payroll"])
    self.assertEqual(
      normalized["governed_period_groups"][0]["quarter_start"],
      1,
    )
    self.assertEqual(
      normalized["governed_period_groups"][1]["quarterly_expansion_levers"],
      [],
    )
    self.assertEqual(
      [(item["quarter_start"], item["quarter_end"]) for item in normalized["governed_period_groups"]],
      [(1, 4), (5, 8), (9, 12), (13, 16), (17, 20)],
    )
    self.assertEqual(len(normalized["lever_adjustment_plan"]), 1)
    self.assertEqual(normalized["controlled_output_targets"][0]["line_item"], "EBITDA")
    self.assertEqual(normalized["active_levers"], ["expenses::Marketing"])
    self.assertTrue(
      any("lever_adjustment_plan_missing_coverage::expenses::Marketing" in item for item in normalized["coverage_issues"])
    )
    self.assertTrue(
      any("controlled_output_targets_missing_ebitda_full_coverage" in item for item in normalized["coverage_issues"])
    )

  def test_normalize_strategy_selection_contract_accepts_full_horizon_numeric_coverage(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "expenses::Marketing",
        "expenses::Payroll",
        "revenue::Home Health::Skilled Nursing::Unit Price",
      ],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 4, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 5, "quarter_end": 8, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 9, "quarter_end": 12, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 13, "quarter_end": 16, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 17, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {
          "lever_id": "expenses::Marketing",
          "direction": "down",
          "quarter_start": 1,
          "quarter_end": 4,
          "min_value": 0.08,
          "max_value": 0.12,
        },
        {
          "lever_id": "expenses::Marketing",
          "direction": "hold",
          "quarter_start": 5,
          "quarter_end": 20,
          "min_value": 0.08,
          "max_value": 0.12,
        },
        {
          "lever_id": "revenue::Home Health::Skilled Nursing::Unit Price",
          "direction": "up",
          "quarter_start": 1,
          "quarter_end": 8,
          "min_value": 85,
          "max_value": 95,
        },
        {
          "lever_id": "revenue::Home Health::Skilled Nursing::Unit Price",
          "direction": "hold",
          "quarter_start": 9,
          "quarter_end": 20,
          "min_value": 90,
          "max_value": 95,
        },
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 1, "quarter_end": 4, "min_value": -40000, "max_value": 0},
        {"line_item": "Revenue", "quarter_start": 1, "quarter_end": 4, "min_value": 80000, "max_value": 120000},
        {"line_item": "EBITDA", "quarter_start": 5, "quarter_end": 8, "min_value": 0, "max_value": 20000},
        {"line_item": "Revenue", "quarter_start": 5, "quarter_end": 8, "min_value": 90000, "max_value": 150000},
        {"line_item": "EBITDA", "quarter_start": 9, "quarter_end": 12, "min_value": 5000, "max_value": 25000},
        {"line_item": "Revenue", "quarter_start": 9, "quarter_end": 12, "min_value": 95000, "max_value": 160000},
        {"line_item": "EBITDA", "quarter_start": 13, "quarter_end": 16, "min_value": 10000, "max_value": 35000},
        {"line_item": "Revenue", "quarter_start": 13, "quarter_end": 16, "min_value": 110000, "max_value": 180000},
        {"line_item": "EBITDA", "quarter_start": 17, "quarter_end": 20, "min_value": 15000, "max_value": 50000},
        {"line_item": "Revenue", "quarter_start": 17, "quarter_end": 20, "min_value": 120000, "max_value": 250000},
      ],
      "active_levers": ["expenses::Marketing", "revenue::Home Health::Skilled Nursing::Unit Price"],
      "severity_class": "moderate",
      "minimum_package_strength": "moderate",
      "controller_directives": {
        "minimum_meaningful_levers": 2,
        "minimum_package_count": 3,
      },
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": [
          "expenses::Marketing",
          "expenses::Payroll",
          "revenue::Home Health::Skilled Nursing::Unit Price",
        ],
      },
    ]
    fixed_facts = {
      "finmo_json": {
        "pl": [{"label": "Revenue"}, {"label": "EBITDA"}],
        "balance_sheet": [{"label": "Cash"}],
        "cash_flow": [],
      }
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertEqual(normalized["coverage_issues"], [])
    self.assertTrue(_gpt_blueprint_is_usable(normalized))

  def test_normalize_strategy_selection_contract_derives_revenue_targets_from_revenue_levers(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
      ],
      "forbidden_model_input_levers": [],
      "lever_adjustment_plan": [
        {
          "lever_id": "revenue::Primary line of business::In-home care hour::Unit Price",
          "direction": "up",
          "quarter_start": 1,
          "quarter_end": 20,
          "min_value": 82,
          "max_value": 90,
        },
        {
          "lever_id": "revenue::Primary line of business::In-home care hour::Utilization",
          "direction": "up",
          "quarter_start": 1,
          "quarter_end": 20,
          "min_value": 0.78,
          "max_value": 0.9,
        },
      ],
      "controlled_output_targets": [],
      "target_posture": {
        "year1_ebitda_posture": "improving_negative",
        "year2_ebitda_posture": "breakeven",
        "year3_ebitda_posture": "profitable",
        "staffing_posture": "lean",
        "pricing_posture": "modest_increase",
        "demand_posture": "steady_growth",
        "cost_posture": "disciplined",
      },
      "severity_class": "moderate",
      "minimum_package_strength": "moderate",
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": selection["allowed_model_input_levers"],
      },
    ]
    fixed_facts = {
      "finmo_json": {
        "pl": [{"label": "Revenue"}, {"label": "EBITDA"}],
      },
      "model_input_json": {
        "start_date": "2026-09-03",
        "sections": {
          "revenue": [
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Capacity", "values": [1300.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Unit Price", "values": [80.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Utilization", "values": [0.75 for _ in range(20)]},
          ],
          "expenses": [
            {"label": "Cost of Goods Sold", "values": [0.78 for _ in range(20)]},
            {"label": "Marketing", "values": [0.14 for _ in range(20)]},
            {"label": "Research & Development", "values": [0.0 for _ in range(20)]},
            {"label": "Lease", "values": [3600.0 for _ in range(20)]},
            {"label": "Payroll", "values": [62497.5 for _ in range(20)]},
            {"label": "General & Administrative", "values": [0.277182 for _ in range(20)]},
            {"label": "Interest Rate", "values": [0.0 for _ in range(20)]},
            {"label": "Depreciation", "values": [0.0 for _ in range(20)]},
            {"label": "Taxes", "values": [0.0 for _ in range(20)]},
          ],
          "balance_sheet": [],
          "schedules": {"debt_opening_balance_seed": 0.0, "lease_opening_balance_seed": 0.0, "rows": []},
        },
      },
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    revenue_targets = [
      item for item in normalized["controlled_output_targets"]
      if str(item.get("line_item") or "").strip() == "Revenue"
    ]
    self.assertEqual(
      [(item["quarter_start"], item["quarter_end"]) for item in revenue_targets],
      [(1, 4), (5, 8), (9, 12), (13, 16), (17, 20)],
    )
    self.assertFalse(
      any("controlled_output_targets_missing_revenue_full_coverage" in item for item in normalized["coverage_issues"])
    )

  def test_normalize_strategy_selection_contract_flags_infeasible_revenue_targets(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
      ],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 4, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 5, "quarter_end": 8, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 9, "quarter_end": 12, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 13, "quarter_end": 16, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 17, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {
          "lever_id": "revenue::Primary line of business::In-home care hour::Unit Price",
          "direction": "up",
          "quarter_start": 1,
          "quarter_end": 20,
          "min_value": 90,
          "max_value": 100,
        },
        {
          "lever_id": "revenue::Primary line of business::In-home care hour::Utilization",
          "direction": "up",
          "quarter_start": 1,
          "quarter_end": 20,
          "min_value": 0.8,
          "max_value": 0.9,
        },
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 4, "quarter_end": 4, "min_value": -40000, "max_value": 0},
        {"line_item": "Revenue", "quarter_start": 4, "quarter_end": 4, "min_value": 200000, "max_value": 250000},
        {"line_item": "EBITDA", "quarter_start": 8, "quarter_end": 8, "min_value": -20000, "max_value": 10000},
        {"line_item": "Revenue", "quarter_start": 8, "quarter_end": 8, "min_value": 200000, "max_value": 250000},
        {"line_item": "EBITDA", "quarter_start": 12, "quarter_end": 12, "min_value": 0, "max_value": 20000},
        {"line_item": "Revenue", "quarter_start": 12, "quarter_end": 12, "min_value": 200000, "max_value": 250000},
        {"line_item": "EBITDA", "quarter_start": 16, "quarter_end": 16, "min_value": 10000, "max_value": 30000},
        {"line_item": "Revenue", "quarter_start": 16, "quarter_end": 16, "min_value": 200000, "max_value": 250000},
        {"line_item": "EBITDA", "quarter_start": 20, "quarter_end": 20, "min_value": 15000, "max_value": 50000},
        {"line_item": "Revenue", "quarter_start": 20, "quarter_end": 20, "min_value": 200000, "max_value": 250000},
      ],
      "active_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
      ],
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": [
          "revenue::Primary line of business::In-home care hour::Unit Price",
          "revenue::Primary line of business::In-home care hour::Utilization",
        ],
      },
    ]
    fixed_facts = {
      "finmo_json": {
        "pl": [{"label": "Revenue"}, {"label": "EBITDA"}],
      },
      "model_input_json": {
        "sections": {
          "revenue": [
            {
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Capacity",
              "values": [1300.0 for _ in range(20)],
            },
            {
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Unit Price",
              "values": [80.0 for _ in range(20)],
            },
            {
              "lob": "Primary line of business",
              "product": "In-home care hour",
              "driver": "Utilization",
              "values": [0.75 for _ in range(20)],
            },
          ]
        }
      },
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertTrue(
      any("controlled_output_targets_infeasible_revenue::Q4" in item for item in normalized["coverage_issues"])
    )
    self.assertFalse(_gpt_blueprint_is_usable(normalized))

  def test_normalize_strategy_selection_contract_flags_infeasible_ebitda_targets(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
        "expenses::Cost of Goods Sold",
        "expenses::Marketing",
        "expenses::Payroll",
        "expenses::General & Administrative",
      ],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 4, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 5, "quarter_end": 8, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 9, "quarter_end": 12, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 13, "quarter_end": 16, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 17, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {"lever_id": "revenue::Primary line of business::In-home care hour::Unit Price", "direction": "up", "quarter_start": 1, "quarter_end": 20, "min_value": 90, "max_value": 95},
        {"lever_id": "revenue::Primary line of business::In-home care hour::Utilization", "direction": "up", "quarter_start": 1, "quarter_end": 20, "min_value": 0.8, "max_value": 0.9},
        {"lever_id": "expenses::Cost of Goods Sold", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.55, "max_value": 0.65},
        {"lever_id": "expenses::Marketing", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.05, "max_value": 0.08},
        {"lever_id": "expenses::Payroll", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 55000, "max_value": 65000},
        {"lever_id": "expenses::General & Administrative", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.12, "max_value": 0.18},
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 4, "quarter_end": 4, "min_value": -60000, "max_value": -25000},
        {"line_item": "Revenue", "quarter_start": 4, "quarter_end": 4, "min_value": 90000, "max_value": 120000},
        {"line_item": "EBITDA", "quarter_start": 8, "quarter_end": 8, "min_value": -30000, "max_value": 5000},
        {"line_item": "Revenue", "quarter_start": 8, "quarter_end": 8, "min_value": 105000, "max_value": 135000},
        {"line_item": "EBITDA", "quarter_start": 12, "quarter_end": 12, "min_value": 0, "max_value": 22000},
        {"line_item": "Revenue", "quarter_start": 12, "quarter_end": 12, "min_value": 115000, "max_value": 145000},
        {"line_item": "EBITDA", "quarter_start": 16, "quarter_end": 16, "min_value": 10000, "max_value": 35000},
        {"line_item": "Revenue", "quarter_start": 16, "quarter_end": 16, "min_value": 120000, "max_value": 150000},
        {"line_item": "EBITDA", "quarter_start": 20, "quarter_end": 20, "min_value": 15000, "max_value": 40000},
        {"line_item": "Revenue", "quarter_start": 20, "quarter_end": 20, "min_value": 120000, "max_value": 130000},
      ],
      "active_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
        "expenses::Cost of Goods Sold",
        "expenses::Marketing",
        "expenses::Payroll",
        "expenses::General & Administrative",
      ],
      "severity_class": "severe",
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": selection["allowed_model_input_levers"],
      },
    ]
    fixed_facts = {
      "finmo_json": {
        "pl": [{"label": "Revenue"}, {"label": "EBITDA"}],
      },
      "model_input_json": {
        "start_date": "2026-09-03",
        "sections": {
          "revenue": [
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Capacity", "values": [1300.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Unit Price", "values": [80.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Utilization", "values": [0.75 for _ in range(20)]},
          ],
          "expenses": [
            {"label": "Cost of Goods Sold", "values": [0.78 for _ in range(20)]},
            {"label": "Marketing", "values": [0.14 for _ in range(20)]},
            {"label": "Research & Development", "values": [0.0 for _ in range(20)]},
            {"label": "Lease", "values": [3600.0 for _ in range(20)]},
            {"label": "Payroll", "values": [62497.5 for _ in range(20)]},
            {"label": "General & Administrative", "values": [0.277182 for _ in range(20)]},
            {"label": "Interest Rate", "values": [0.0 for _ in range(20)]},
            {"label": "Depreciation", "values": [0.0 for _ in range(20)]},
            {"label": "Taxes", "values": [0.0 for _ in range(20)]},
          ],
          "balance_sheet": [],
          "schedules": {"debt_opening_balance_seed": 0.0, "lease_opening_balance_seed": 0.0, "rows": []},
        },
      },
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertTrue(
      any("controlled_output_targets_infeasible_ebitda::Q20" in item for item in normalized["coverage_issues"])
    )
    self.assertFalse(_gpt_blueprint_is_usable(normalized))

  def test_normalize_strategy_selection_contract_rejects_negative_ebitda_targets_from_year_two_onward(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": ["expenses::Payroll"],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {"lever_id": "expenses::Payroll", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 45000, "max_value": 65000},
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 1, "quarter_end": 4, "min_value": -80000, "max_value": -20000},
        {"line_item": "EBITDA", "quarter_start": 5, "quarter_end": 20, "min_value": -1000, "max_value": 20000},
      ],
      "active_levers": ["expenses::Payroll"],
      "severity_class": "severe",
      "minimum_package_strength": "strong",
    }
    strategy_catalog = [{"strategy_id": "alpha", "allowed_model_input_levers": ["expenses::Payroll"]}]
    fixed_facts = {"finmo_json": {"pl": [{"label": "EBITDA"}]}}

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertTrue(
      any(
        "controlled_output_targets_below_minimum_viability_floor::Q5,Q6,Q7,Q8,Q9,Q10,Q11,Q12,Q13,Q14,Q15,Q16,Q17,Q18,Q19,Q20" in item
        for item in normalized["coverage_issues"]
      )
    )
    self.assertFalse(_gpt_blueprint_is_usable(normalized))

  def test_normalize_strategy_selection_contract_flags_need_for_capacity_when_year2_breakeven_is_impossible(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "revenue::Primary line of business::In-home care hour::Unit Price",
        "revenue::Primary line of business::In-home care hour::Utilization",
        "expenses::Payroll",
        "expenses::Marketing",
        "expenses::General & Administrative",
      ],
      "forbidden_model_input_levers": [],
      "lever_adjustment_plan": [
        {"lever_id": "revenue::Primary line of business::In-home care hour::Unit Price", "direction": "up", "quarter_start": 1, "quarter_end": 20, "min_value": 85, "max_value": 95},
        {"lever_id": "revenue::Primary line of business::In-home care hour::Utilization", "direction": "up", "quarter_start": 1, "quarter_end": 20, "min_value": 0.82, "max_value": 0.9},
        {"lever_id": "expenses::Payroll", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 45000, "max_value": 65000},
        {"lever_id": "expenses::Marketing", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.08, "max_value": 0.12},
        {"lever_id": "expenses::General & Administrative", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.18, "max_value": 0.24},
      ],
      "controlled_output_targets": [],
      "target_posture": {
        "year1_ebitda_posture": "improving_negative",
        "year2_ebitda_posture": "breakeven",
        "year3_ebitda_posture": "profitable",
        "staffing_posture": "lean",
        "pricing_posture": "modest_increase",
        "demand_posture": "steady_growth",
        "cost_posture": "disciplined",
      },
      "severity_class": "severe",
      "minimum_package_strength": "strong",
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": selection["allowed_model_input_levers"],
      },
    ]
    fixed_facts = {
      "finmo_json": {"pl": [{"label": "Revenue"}, {"label": "EBITDA"}]},
      "model_input_json": {
        "start_date": "2026-09-03",
        "sections": {
          "revenue": [
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Capacity", "values": [1300.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Unit Price", "values": [80.0 for _ in range(20)]},
            {"lob": "Primary line of business", "product": "In-home care hour", "driver": "Utilization", "values": [0.75 for _ in range(20)]},
          ],
          "expenses": [
            {"label": "Cost of Goods Sold", "values": [0.78 for _ in range(20)]},
            {"label": "Marketing", "values": [0.14 for _ in range(20)]},
            {"label": "Research & Development", "values": [0.0 for _ in range(20)]},
            {"label": "Lease", "values": [3600.0 for _ in range(20)]},
            {"label": "Payroll", "values": [62497.5 for _ in range(20)]},
            {"label": "General & Administrative", "values": [0.277182 for _ in range(20)]},
            {"label": "Interest Rate", "values": [0.0 for _ in range(20)]},
            {"label": "Depreciation", "values": [0.0 for _ in range(20)]},
            {"label": "Taxes", "values": [0.0 for _ in range(20)]},
          ],
          "balance_sheet": [],
          "schedules": {"debt_opening_balance_seed": 0.0, "lease_opening_balance_seed": 0.0, "rows": []},
        },
      },
    }

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertTrue(
      any(item == "viability_requires_capacity_lever_or_broader_scale_strategy" for item in normalized["coverage_issues"])
    )

  def test_normalize_strategy_selection_contract_applies_fixed_five_groups(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": ["expenses::Marketing"],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 4, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 5, "quarter_end": 12, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 13, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {"lever_id": "expenses::Marketing", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.05, "max_value": 0.1},
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 4, "quarter_end": 4, "min_value": -50000, "max_value": -10000},
        {"line_item": "EBITDA", "quarter_start": 12, "quarter_end": 12, "min_value": 0, "max_value": 10000},
        {"line_item": "EBITDA", "quarter_start": 20, "quarter_end": 20, "min_value": 10000, "max_value": 25000},
      ],
      "active_levers": ["expenses::Marketing"],
      "severity_class": "moderate",
    }
    strategy_catalog = [{"strategy_id": "alpha", "allowed_model_input_levers": ["expenses::Marketing"]}]
    fixed_facts = {"finmo_json": {"pl": [{"label": "EBITDA"}]}}

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertEqual(
      [(item["quarter_start"], item["quarter_end"]) for item in normalized["governed_period_groups"]],
      [(1, 4), (5, 8), (9, 12), (13, 16), (17, 20)],
    )

  def test_normalize_strategy_selection_contract_applies_app_owned_severe_posture(self) -> None:
    selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": [
        "expenses::Payroll",
        "expenses::Cost of Goods Sold",
        "expenses::General & Administrative",
        "expenses::Marketing",
      ],
      "forbidden_model_input_levers": [],
      "governed_period_groups": [
        {"quarter_start": 1, "quarter_end": 4, "input_granularity": "quarterly", "quarterly_expansion_levers": []},
        {"quarter_start": 5, "quarter_end": 8, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 9, "quarter_end": 12, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 13, "quarter_end": 16, "input_granularity": "grouped", "quarterly_expansion_levers": []},
        {"quarter_start": 17, "quarter_end": 20, "input_granularity": "grouped", "quarterly_expansion_levers": []},
      ],
      "lever_adjustment_plan": [
        {"lever_id": "expenses::Payroll", "direction": "down", "quarter_start": 1, "quarter_end": 4, "min_value": 52000, "max_value": 56000},
        {"lever_id": "expenses::Payroll", "direction": "down", "quarter_start": 5, "quarter_end": 8, "min_value": 52000, "max_value": 56000},
        {"lever_id": "expenses::Payroll", "direction": "hold", "quarter_start": 9, "quarter_end": 12, "min_value": 53000, "max_value": 57000},
        {"lever_id": "expenses::Payroll", "direction": "hold", "quarter_start": 13, "quarter_end": 16, "min_value": 54000, "max_value": 58000},
        {"lever_id": "expenses::Payroll", "direction": "hold", "quarter_start": 17, "quarter_end": 20, "min_value": 55000, "max_value": 59000},
        {"lever_id": "expenses::Cost of Goods Sold", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.6, "max_value": 0.72},
        {"lever_id": "expenses::General & Administrative", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.17, "max_value": 0.24},
        {"lever_id": "expenses::Marketing", "direction": "down", "quarter_start": 1, "quarter_end": 20, "min_value": 0.11, "max_value": 0.14},
      ],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "quarter_start": 1, "quarter_end": 1, "min_value": -80000, "max_value": -60000},
        {"line_item": "EBITDA", "quarter_start": 4, "quarter_end": 4, "min_value": -76000, "max_value": -55000},
        {"line_item": "EBITDA", "quarter_start": 8, "quarter_end": 8, "min_value": -70000, "max_value": -52000},
        {"line_item": "EBITDA", "quarter_start": 12, "quarter_end": 12, "min_value": -68000, "max_value": -50000},
        {"line_item": "EBITDA", "quarter_start": 16, "quarter_end": 16, "min_value": -65000, "max_value": -48000},
        {"line_item": "EBITDA", "quarter_start": 20, "quarter_end": 20, "min_value": -65000, "max_value": -48000},
      ],
      "active_levers": [
        "expenses::Payroll",
        "expenses::Cost of Goods Sold",
        "expenses::General & Administrative",
        "expenses::Marketing",
      ],
      "severity_class": "severe",
      "minimum_package_strength": "strong",
      "controller_directives": {
        "minimum_meaningful_levers": 3,
        "minimum_package_count": 1,
        "aggression_level": "moderate",
        "escalate_on_retry": False,
      },
    }
    strategy_catalog = [
      {
        "strategy_id": "alpha",
        "allowed_model_input_levers": selection["allowed_model_input_levers"],
      },
    ]
    fixed_facts = {"finmo_json": {"pl": [{"label": "EBITDA"}]}}

    normalized = _normalize_strategy_selection_contract(
      selection=selection,
      strategy_catalog=strategy_catalog,
      fixed_facts=fixed_facts,
    )

    self.assertEqual(normalized["controller_directives"]["aggression_level"], "high")
    self.assertEqual(normalized["controller_directives"]["escalate_on_retry"], True)
    self.assertEqual(normalized["controller_directives"]["minimum_package_count"], 2)
    self.assertFalse(
      any(item == "severe_blueprint_requires_aggression_level_high" for item in normalized["coverage_issues"])
    )
    self.assertFalse(
      any(item == "severe_blueprint_requires_escalate_on_retry_true" for item in normalized["coverage_issues"])
    )

  def test_advise_consistency_strategy_selection_returns_last_invalid_blueprint_details(self) -> None:
    class _FakeResponse:
      status_code = 200

      def json(self) -> dict:
        return {}

    invalid_selection = {
      "selected_strategy_ids": ["alpha"],
      "coverage_issues": [
        "severe_blueprint_requires_escalate_on_retry_true",
        "severe_blueprint_requires_aggression_level_high",
      ],
      "controller_directives": {
        "aggression_level": "moderate",
        "escalate_on_retry": False,
      },
    }

    with patch("consistency_strategy_advisor._strategy_layer_enabled", return_value=True), patch(
      "consistency_strategy_advisor._require_openai_key",
      return_value="test-key",
    ), patch(
      "consistency_strategy_advisor._post_openai",
      return_value=_FakeResponse(),
    ), patch(
      "consistency_strategy_advisor._parse_json_response",
      return_value={"selected_strategy_ids": ["alpha"]},
    ), patch(
      "consistency_strategy_advisor._normalize_strategy_selection_contract",
      return_value=invalid_selection,
    ):
      result = advise_consistency_strategy_selection(
        baseline_summary={},
        fixed_facts={},
        viability_mode=True,
        diagnosis={},
        strategy_catalog=[
          {
            "strategy_id": "alpha",
            "strategy_name": "Alpha",
            "archetype": "operations",
            "allowed_model_input_levers": ["expenses::Payroll"],
            "allowed_model_input_lever_details": [],
            "dominant_tradeoff": "",
          }
        ],
        retry_context=None,
      )

    self.assertEqual(result["error"], "strategy_advisor_no_selection")
    self.assertEqual(result["error_detail"], "strategy_blueprint_coverage_gaps")
    self.assertEqual(
      result["coverage_issues"],
      [
        "severe_blueprint_requires_escalate_on_retry_true",
        "severe_blueprint_requires_aggression_level_high",
      ],
    )
    self.assertEqual(result["last_invalid_selection"], invalid_selection)
    self.assertEqual(result["advisor_attempt_count"], 3)

  def test_build_strategy_layer_shortlists_catalog_from_toolset_selector(self) -> None:
    captured: dict = {}

    def _fake_gpt_strategy_selection(**kwargs):
      captured["strategy_catalog"] = kwargs.get("strategy_catalog")
      return {"error": "missing"}

    with patch("consistency_flow._gpt_strategy_required", return_value=True), patch(
      "consistency_flow._build_strategy_catalog",
      return_value=[
        {"strategy_id": "cost_structure_adjustment", "allowed_model_input_levers": ["expenses::Payroll"]},
        {"strategy_id": "demand_supported_growth", "allowed_model_input_levers": ["revenue::Primary::Product::Capacity"]},
        {"strategy_id": "staffing_ramp_adjustment", "allowed_model_input_levers": ["expenses::Payroll"]},
      ],
    ), patch(
      "consistency_flow._gpt_strategy_selection",
      side_effect=_fake_gpt_strategy_selection,
    ):
      layer = _build_strategy_layer(
        state_model={"fixed_facts": {}},
        direct_inputs={},
        baseline_summary={
          "revenue": 100.0,
          "ebitda": -80.0,
          "payroll": 60.0,
          "gross_profit": 20.0,
        },
        diagnostic_state=None,
        viability_mode=True,
      )

    self.assertEqual(layer["source"], "gpt_required_unavailable")
    passed_ids = [str(item.get("strategy_id") or "") for item in (captured.get("strategy_catalog") or []) if isinstance(item, dict)]
    self.assertEqual(passed_ids, ["demand_supported_growth", "staffing_ramp_adjustment"])

  def test_select_best_finmo_attempts_prefers_accepted_clear_attempt(self) -> None:
    selected = _select_best_finmo_attempts(
      [
        {
          "scenario_id": "blocked",
          "remaining_blocking_count": 1,
          "remaining_violation_count": 0,
          "presentation_issues": [],
          "forecast_years": [{"ebitda": 100.0}, {"ebitda": 150.0}, {"ebitda": 200.0}],
          "gpt_validation_result": {"validation_status": "accepted"},
        },
        {
          "scenario_id": "clear",
          "remaining_blocking_count": 0,
          "remaining_violation_count": 0,
          "presentation_issues": [],
          "forecast_years": [{"ebitda": 10.0}, {"ebitda": 25.0}, {"ebitda": 50.0}],
          "gpt_validation_result": {"validation_status": "accepted"},
        },
      ],
      require_clear=True,
    )

    self.assertEqual([item["scenario_id"] for item in selected], ["clear"])

  def test_select_best_finmo_attempts_returns_provisional_result_when_clear_not_required(self) -> None:
    selected = _select_best_finmo_attempts(
      [
        {
          "scenario_id": "accepted",
          "remaining_blocking_count": 1,
          "remaining_violation_count": 1,
          "presentation_issues": ["degrading_five_year_path"],
          "forecast_years": [{"ebitda": 20.0}, {"ebitda": 30.0}, {"ebitda": 40.0}],
          "gpt_validation_result": {"validation_status": "accepted"},
        },
        {
          "scenario_id": "rejected",
          "remaining_blocking_count": 0,
          "remaining_violation_count": 0,
          "presentation_issues": [],
          "forecast_years": [{"ebitda": 5.0}, {"ebitda": 10.0}, {"ebitda": 15.0}],
          "gpt_validation_result": {"validation_status": "rejected"},
        },
      ],
      require_clear=False,
    )

    self.assertEqual([item["scenario_id"] for item in selected], ["accepted"])

  def test_build_consistency_governance_state_uses_finmo_candidate_and_attempt_metadata(self) -> None:
    strategy_selection = {
      "selected_strategy_ids": ["alpha"],
      "allowed_model_input_levers": ["expenses::Marketing"],
      "lever_adjustment_plan": [
        {
          "lever_id": "expenses::Marketing",
          "direction": "decrease",
          "quarter_start": 1,
          "quarter_end": 4,
        }
      ],
      "governed_period_groups": [{"quarter_start": 1, "quarter_end": 20}],
      "controlled_output_targets": [
        {"line_item": "EBITDA", "goal_band": {"min": 0.0, "max": 10.0}}
      ],
    }
    candidate = {
      "scenario_id": "finmo-1",
      "strategy_id": "alpha",
      "strategy_name": "Alpha",
      "remaining_blocking_count": 0,
      "remaining_violation_count": 0,
      "remaining_violations": [],
      "remaining_blocking_violations": [],
      "presentation_issues": [],
      "forecast_years": [
        {"year_index": 1, "ebitda": 5.0},
        {"year_index": 2, "ebitda": 10.0},
        {"year_index": 3, "ebitda": 15.0},
      ],
      "forecast_quarters": [{"quarter_index": 1, "ebitda": 1.0}],
      "model_input_json": {"lever_catalog": {"expenses::Marketing": {}}},
      "finmo_json": {"pl": [{"label": "EBITDA", "period_values": [1.0, 2.0, 3.0]}]},
      "gpt_validation_request": {"validation_contract_version": "finmo_validation_request_v1"},
      "controller_calibration_request": {"allowed_model_input_levers": ["expenses::Marketing"]},
      "modified_state": {
        "ops_json": {"capacity_driver": "labor"},
        "people_json": {"people": []},
        "financials_json": {"payroll_total_year1": 100000},
        "financials_year1_json": {"company_revenue_total_year1": 250000},
        "marketing_model_json": {"expected_units_year1": 1000},
      },
    }

    with patch(
      "consistency_flow._build_consistency_state_model",
      return_value={"fixed_facts": {"business_type": "Home Health"}, "baseline_state": {}, "finmo_json": {}},
    ), patch(
      "consistency_flow._build_controller_inputs",
      return_value={"current_revenue": 250000.0},
    ), patch(
      "consistency_flow._build_strategy_catalog",
      return_value=[{"strategy_id": "alpha", "allowed_model_input_levers": ["expenses::Marketing"]}],
    ), patch(
      "consistency_flow._build_runtime_strategy",
      return_value={"strategy_id": "alpha", "strategy_name": "Alpha"},
    ), patch(
      "consistency_flow._gpt_blueprint_is_usable",
      return_value=True,
    ), patch(
      "consistency_flow._gpt_strategy_selection",
      return_value=strategy_selection,
    ), patch(
      "consistency_flow._controller_profiles",
      return_value=[{"profile_id": "alpha"}],
    ), patch(
      "consistency_flow._build_calibration_contract",
      return_value={"profile": {"strategy_id": "alpha"}, "diagnostics": {}, "direct_inputs": {}},
    ), patch(
      "consistency_flow.build_controller_finmo_candidate",
      return_value=candidate,
    ), patch(
      "consistency_flow._gpt_finmo_validation",
      return_value={"validation_status": "accepted", "issues": []},
    ):
      governance_state = build_consistency_governance_state(
        ops_json={"capacity_driver": "labor"},
        people_json={"people": []},
        financials_json={"payroll_total_year1": 100000},
        financials_year1_json={"company_revenue_total_year1": 250000},
        marketing_model_json={"expected_units_year1": 1000},
        model_input_json={"lever_catalog": {"expenses::Marketing": {}}},
        finmo_json={"pl": [{"label": "EBITDA", "period_values": [1.0, 2.0, 3.0]}]},
      )

    self.assertEqual(governance_state["status"], "awaiting_choice")
    self.assertEqual(governance_state["selection_status"], "finmo_validated")
    self.assertEqual(governance_state["governed_attempt_limit"], 3)
    self.assertEqual(governance_state["governed_attempt_count"], 1)
    self.assertEqual(governance_state["scenarios"][0]["scenario_id"], "finmo-1")

  def test_apply_consistency_selected_path_uses_modified_state_only(self) -> None:
    result = apply_consistency_selected_path(
      ops_json={},
      people_json={},
      financials_json={},
      financials_year1_json={},
      marketing_model_json={},
      governance_state={
        "scenarios": [
          {
            "scenario_id": "chosen",
            "modified_state": {
              "ops_json": {"capacity_driver": "labor"},
              "people_json": {"people": []},
              "financials_json": {"payroll_total_year1": 100000},
              "financials_year1_json": {"company_revenue_total_year1": 250000},
              "marketing_model_json": {"expected_units_year1": 1000},
            },
          }
        ]
      },
      selected_scenario_id="chosen",
      overrides={},
    )

    self.assertEqual(result["ops_json"]["capacity_driver"], "labor")
    self.assertEqual(result["financials_year1_json"]["company_revenue_total_year1"], 250000)

  def test_build_consistency_modified_plan_payload_is_finmo_authoritative(self) -> None:
    finmo_json = {
      "quarter_rows": [
        {"year": 2026, "quarter": 1, "date": "2026-01-01", "revenue": 100.0, "cogs": 40.0, "gross_profit": 60.0, "marketing": 5.0, "payroll": 20.0, "g_and_a": 10.0, "research_and_development": 0.0, "lease_rent": 0.0, "ebitda": 25.0, "interest": 2.0, "depreciation": 1.0, "taxes": 3.0, "net_income": 19.0, "cash": 50.0, "ending_cash": 50.0, "total_assets": 200.0, "total_liabilities_and_equity": 200.0},
        {"year": 2026, "quarter": 2, "date": "2026-04-01", "revenue": 120.0, "cogs": 48.0, "gross_profit": 72.0, "marketing": 6.0, "payroll": 22.0, "g_and_a": 11.0, "research_and_development": 0.0, "lease_rent": 0.0, "ebitda": 33.0, "interest": 2.0, "depreciation": 1.0, "taxes": 4.0, "net_income": 26.0, "cash": 60.0, "ending_cash": 60.0, "total_assets": 220.0, "total_liabilities_and_equity": 220.0},
      ],
      "accounting_check": {"status": "ok"},
    }

    payload = _build_consistency_modified_plan_payload(
      governance_state={"state_model": {"fixed_facts": {"business_type": "Home Health"}}},
      selected_scenario={
        "scenario_id": "finmo-1",
        "strategy_id": "alpha",
        "remaining_blocking_violations": [],
      },
      consistency_runtime_payload={},
      initial_ops_json={},
      initial_market_json={},
      initial_people_json={},
      initial_financials_json={"payroll_total_year1": 120000},
      initial_financials_year1_json={"company_revenue_total_year1": 250000},
      initial_marketing_model_json={},
      modified_ops_json={},
      modified_market_json={},
      modified_people_json={},
      modified_financials_json={"payroll_total_year1": 100000},
      modified_financials_year1_json={"company_revenue_total_year1": 300000},
      modified_marketing_model_json={},
      modified_forecast_quarters=[],
      finmo_json=finmo_json,
    )

    self.assertEqual(payload["forecast_meta"]["financial_authority"], "finmo")
    self.assertEqual(payload["forecast_meta"]["finmo_accounting_check"]["status"], "ok")
    self.assertEqual(payload["quarter_driver_path"][0]["quarter_index"], 1)
    self.assertTrue(payload["forecast_years"])

  def test_build_consistency_finmo_attempts_payload_tracks_attempts(self) -> None:
    payload = _build_consistency_finmo_attempts_payload(
      governance_state={
        "attempted_scenarios": [
          {
            "scenario_id": "a1",
            "strategy_id": "alpha",
            "strategy_name": "Alpha",
            "remaining_violations": ["gross_margin_too_low"],
            "remaining_blocking_violations": ["gross_margin_too_low"],
            "allowed_model_input_levers": ["expenses::Marketing"],
            "controller_calibration_request": {"allowed_model_input_levers": ["expenses::Marketing"]},
            "gpt_validation_request": {"validation_contract_version": "finmo_validation_request_v1"},
            "gpt_validation_result": {"validation_status": "rejected"},
            "model_input_json": {"lever_catalog": {"expenses::Marketing": {}}},
            "finmo_json": {"pl": [{"label": "EBITDA"}]},
          }
        ]
      },
      selected_scenario={"scenario_id": "a1"},
    )

    self.assertEqual(payload["attempt_count"], 1)
    self.assertEqual(payload["selected_scenario_id"], "a1")
    self.assertTrue(payload["attempts"][0]["accepted"])
    self.assertEqual(payload["attempts"][0]["finmo_status"], "persisted")

  def test_build_consistency_finmo_attempts_payload_includes_failed_attempts(self) -> None:
    payload = _build_consistency_finmo_attempts_payload(
      governance_state={
        "attempt_failures": [
          {
            "attempt_index": 1,
            "strategy_id": "alpha",
            "candidate_failure": {
              "failure_stage": "sync_exception",
              "error_type": "PermissionError",
              "error_message": "locked workbook",
            },
          }
        ]
      },
      selected_scenario=None,
    )

    self.assertEqual(payload["attempt_count"], 0)
    self.assertEqual(payload["failed_attempt_count"], 1)
    self.assertEqual(payload["failed_attempts"][0]["candidate_failure"]["error_type"], "PermissionError")

  def test_build_controller_finmo_candidate_returns_failure_payload_when_finmo_readback_fails(self) -> None:
    with patch(
      "client_intake_and_finmo.consistency_flow.finmo_controller._candidate_finmo_readback",
      return_value={
        "candidate_failure": {
          "failure_stage": "sync_exception",
          "error_type": "PermissionError",
          "error_message": "locked workbook",
        }
      },
    ):
      candidate = build_controller_finmo_candidate(
        profile={"strategy_id": "alpha", "strategy_name": "Alpha"},
        contract_bundle={"direct_inputs": {}, "diagnostics": {}},
        state_model={"baseline_state": {}},
        scenario_index=1,
      )

    self.assertEqual(candidate["strategy_id"], "alpha")
    self.assertEqual(candidate["candidate_failure"]["failure_stage"], "sync_exception")

  def test_finmo_calibration_spec_excludes_price_from_solver_inputs(self) -> None:
    calibration_spec = _build_finmo_calibration_spec(
      profile={
        "allowed_model_input_levers": [
          "revenue::Home Health::Skilled Nursing::Unit Price",
          "revenue::Home Health::Skilled Nursing::Utilization",
          "expenses::Payroll",
        ],
        "lever_adjustment_plan": [
          {
            "lever_id": "revenue::Home Health::Skilled Nursing::Unit Price",
            "direction": "up",
            "quarter_start": 1,
            "quarter_end": 20,
            "min_value": 90.0,
            "max_value": 95.0,
          },
          {
            "lever_id": "revenue::Home Health::Skilled Nursing::Utilization",
            "direction": "up",
            "quarter_start": 1,
            "quarter_end": 20,
            "min_value": 0.7,
            "max_value": 0.85,
          },
          {
            "lever_id": "expenses::Payroll",
            "direction": "down",
            "quarter_start": 1,
            "quarter_end": 20,
            "min_value": 40000.0,
            "max_value": 60000.0,
          },
        ],
        "governed_period_groups": [{"quarter_start": 1, "quarter_end": 20}],
        "controlled_output_targets": [
          {
            "line_item": "EBITDA",
            "quarter_start": 1,
            "quarter_end": 20,
            "min_value": 0.0,
            "max_value": 10000.0,
          }
        ],
      },
      direct_inputs={
        "model_input_json": {
          "lever_catalog": {
            "revenue::Home Health::Skilled Nursing::Unit Price": {
              "section": "revenue",
              "named_range": "model_input_revenue",
              "lob": "Home Health",
              "product": "Skilled Nursing",
              "driver": "Unit Price",
              "valid_quarter_indices": list(range(1, 21)),
              "value_kind": "direct_number",
              "input_semantics": "currency_per_unit",
            },
            "revenue::Home Health::Skilled Nursing::Utilization": {
              "section": "revenue",
              "named_range": "model_input_revenue",
              "lob": "Home Health",
              "product": "Skilled Nursing",
              "driver": "Utilization",
              "valid_quarter_indices": list(range(1, 21)),
              "value_kind": "ratio",
              "input_semantics": "utilization_ratio",
            },
            "expenses::Payroll": {
              "section": "expenses",
              "named_range": "model_input_expenses",
              "label": "Payroll",
              "valid_quarter_indices": list(range(1, 21)),
              "value_kind": "direct_number",
              "input_semantics": "quarter_currency",
            },
          }
        }
      },
      controller_input_seed=[],
    )

    solver_requests = calibration_spec["solver_requests"]
    self.assertTrue(solver_requests)
    changing_levers = {
      item["lever_id"]
      for request in solver_requests
      for item in request["changing_inputs"]
    }
    self.assertNotIn("revenue::Home Health::Skilled Nursing::Unit Price", changing_levers)
    self.assertIn("revenue::Home Health::Skilled Nursing::Utilization", changing_levers)
    self.assertIn("expenses::Payroll", changing_levers)
    self.assertEqual(solver_requests[0]["objective"]["objective_mode"], "maximize")

  def test_execute_finmo_calibration_shell_uses_objective_mode_not_midpoint_target(self) -> None:
    with patch("client_intake_and_finmo.finmo_bridge.subprocess.run") as subprocess_run:
      subprocess_run.return_value.stdout = json.dumps(
        {
          "goal_seek_results": [],
          "solver_results": [
            {
              "request_id": "solver_group_1_q1_q4",
              "success": True,
              "solver_result": "0",
              "solver_code": 0,
              "objective_mode": "maximize",
            }
          ],
        }
      )
      result = _execute_finmo_calibration_shell(
        finmo_path="C:\\temp\\client.xlsx",
        calibration_shell={
          "solver_requests": [
            {
              "request_id": "solver_group_1_q1_q4",
              "objective_cell": {"sheet": "Financial Model QTR", "cell": "G10"},
              "objective": {
                "goal_band": {"min": 0.0, "max": 100.0},
                "objective_mode": "maximize",
              },
              "changing_input_cells": [
                {"sheet": "Model Inputs", "cell": "G8"},
              ],
              "constraints": [],
            }
          ]
        },
      )

    self.assertTrue(result["solver_results"])
    called_script = subprocess_run.call_args.args[0][3]
    self.assertIn("$objectiveMode = ($request.objective_mode | Out-String).Trim().ToLower();", called_script)
    self.assertNotIn("$request.target_value", called_script)
    self.assertIn("SolverOK', $objectiveSpec, 1, $null, $changeSpec", called_script)

  def test_serialize_debug_draft_row_parses_current_finmo_fields(self) -> None:
    row = {
      "messages_json": json.dumps([{"role": "assistant", "content": "x"}]),
      "financials_json": json.dumps({"payroll_total_year1": 100000}),
      "model_input_json": json.dumps({"lever_catalog": {"expenses::Marketing": {}}}),
      "finmo_json": json.dumps({"pl": [{"label": "EBITDA"}]}),
      "consistency_finmo_attempts_json": json.dumps({"attempt_count": 1}),
      "consistency_gpt_governance_json": json.dumps({"status": "awaiting_choice"}),
      "consistency_controller_contract_json": json.dumps({"selection_status": "finmo_validated"}),
    }

    serialized = _serialize_debug_draft_row(row)

    self.assertEqual(serialized["model_input_json"]["lever_catalog"], {"expenses::Marketing": {}})
    self.assertEqual(serialized["finmo_json"]["pl"][0]["label"], "EBITDA")
    self.assertEqual(serialized["financials_json"]["payroll_total_year1"], 100000)

  def test_shared_context_returns_finmo_consistency_fields(self) -> None:
    consult_row = {
      "operating_model_json": json.dumps({"capacity_driver": "labor"}),
      "target_market_json": json.dumps({"customer_type": "b2c"}),
      "people_json": json.dumps({"people": []}),
      "financials_json": json.dumps({"payroll_total_year1": 100000}),
      "marketing_model_json": json.dumps({"expected_units_year1": 1000}),
      "financials_year1_json": json.dumps({"company_revenue_total_year1": 250000}),
      "model_input_json": json.dumps({"lever_catalog": {"expenses::Marketing": {}}}),
      "finmo_json": json.dumps({"pl": [{"label": "Revenue"}]}),
      "consistency_finmo_attempts_json": json.dumps({"attempt_count": 1}),
      "consistency_gpt_governance_json": json.dumps({"status": "awaiting_choice"}),
      "consistency_controller_contract_json": json.dumps({"selection_status": "finmo_validated"}),
    }

    with patch("intake_consult_draft.get_draft", return_value=consult_row):
      shared_context = build_shared_context(conn=object(), draft_id="draft-1")

    self.assertEqual(shared_context["model_input_json"]["lever_catalog"], {"expenses::Marketing": {}})
    self.assertEqual(shared_context["finmo_json"]["pl"][0]["label"], "Revenue")
    self.assertEqual(shared_context["consistency_finmo_attempts"]["attempt_count"], 1)

  def test_ensure_submission_finmo_path_reuses_existing_path(self) -> None:
    row = {"finmo_path": "C:/tmp/existing.xlsx", "business_name": "Firm", "created_at": "2026-03-29 10:00:00"}

    with patch("intake_pipeline.create_client_finmo_workbook") as create_mock, patch(
      "intake_pipeline.update_intake_submission_finmo_path"
    ) as update_mock:
      finmo_path = _ensure_submission_finmo_path(
        conn=object(),
        submission_id=1,
        submission_row=row,
        finmo_template_path="template.xlsx",
        client_finmo_dir="client-dir",
        client_id="client123",
      )

    self.assertEqual(finmo_path, "C:/tmp/existing.xlsx")
    create_mock.assert_not_called()
    update_mock.assert_not_called()


if __name__ == "__main__":
  unittest.main()
