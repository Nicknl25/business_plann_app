"""X1 — the GUARD TEST (2026-08-18).

This is the point of X1. The design system is only permanent if a later sheet
cannot quietly bypass it, so this walks the WHOLE built workbook — every
populated cell and every chart of every sheet — and fails on:
  * a font family other than Calibri (Nick's Q2 ruling),
  * a font size outside the type scale,
  * a font color or fill outside the palette,
  * a number format outside the design set,
  * a chart built outside `design.chart()` (the single door),
  * `varyColors` left on (Excel then legends every POINT — the rainbow bug),
  * a deleted axis (openpyxl hides axes unless told not to),
  * a series color outside the validated palette.

Because it walks the workbook rather than a list of sheets, a sheet added in
X2+/W3+ is covered the moment it exists — nobody has to remember to add it here.

The negative-control tests prove the guard actually bites: they inject each
class of bypass into a real built workbook and assert it is caught.
"""
from __future__ import annotations

import os
import sys
import unittest

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for path in (REPO, os.path.join(REPO, "python"), os.path.join(REPO, "tests")):
  if path not in sys.path:
    sys.path.insert(0, path)

from openpyxl.chart import LineChart, Reference  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402

from _p3_40_contract_2_fixtures import valid_workbook_payload_dict  # noqa: E402
from client_statements_output_excel import design  # noqa: E402
from client_statements_output_excel.cover_sheet import COVER_SHEET  # noqa: E402
from client_statements_output_excel.data import DraftWorkbookData  # noqa: E402
from client_statements_output_excel.workbook_builder import (  # noqa: E402
  build_client_financial_model_workbook,
)


def _workbook():
  payload = valid_workbook_payload_dict()
  data = DraftWorkbookData(
    draft_row={"business_name": "Fixture Co", "address_city": "Madison", "address_state": "WI"},
    model_input_json=payload.get("model_input_json") or {},
    finmo_json=payload.get("finmo_json") or {},
    payroll_headcount=payload.get("payroll_headcount") or {},
    debt_schedule=payload.get("debt_schedule") or {},
    planning_run_json=payload.get("planning_run_json") or {},
    run_diagnostics=payload.get("run_diagnostics"),
  )
  return build_client_financial_model_workbook(data)


class DesignSystemGuardTests(unittest.TestCase):
  @classmethod
  def setUpClass(cls):
    cls.wb = _workbook()

  def test_no_sheet_or_chart_bypasses_the_design_system(self):
    problems = design.audit_workbook(self.wb, max_report=40)
    self.assertEqual(problems, [], "\n".join(["design-system bypasses:"] + problems))

  def test_every_chart_came_through_the_single_door(self):
    charts = [(ws.title, c) for ws in self.wb.worksheets for c in (ws._charts or [])]
    self.assertTrue(charts, "the workbook should carry charts")
    for title, chart in charts:
      self.assertTrue(getattr(chart, design.MARKER_ATTR, False), f"{title}: chart bypassed design.chart()")
      self.assertFalse(chart.varyColors, f"{title}: varyColors on")

  def test_cover_is_first_and_carries_no_formulas(self):
    self.assertEqual(self.wb.sheetnames[0], COVER_SHEET)
    self.assertEqual(self.wb.active.title, COVER_SHEET)
    cover = self.wb[COVER_SHEET]
    formulas = [
      c.coordinate for row in cover.iter_rows() for c in row
      if isinstance(c.value, str) and c.value.startswith("=")
    ]
    self.assertEqual(formulas, [], "a formula-bearing cover would move the R32 grid")

  def test_sheet_order_follows_the_reading_order(self):
    names = self.wb.sheetnames
    expected = [n for n in design.SHEET_ORDER if n in names]
    self.assertEqual(names[: len(expected)], expected)
    self.assertLess(names.index("Dashboard"), names.index("FINMO"))

  def test_type_scale_is_pinned_to_calibri(self):
    title = self.wb["FINMO"]["A1"]
    self.assertEqual(title.font.name, design.FONT_FAMILY)
    self.assertIn(int(title.font.size), design.ALLOWED_FONT_SIZES)

  def test_no_text_cell_masquerades_as_a_formula(self):
    """One of the four confirmed corruption triggers (research §6.4)."""
    for ws in self.wb.worksheets:
      for row in ws.iter_rows():
        for cell in row:
          if isinstance(cell.value, str) and cell.value.startswith("="):
            self.assertNotRegex(cell.value, r"^=\s", f"{ws.title}!{cell.coordinate}")


class GuardNegativeControlTests(unittest.TestCase):
  """The guard is only worth having if it bites. Each of these injects one
  class of bypass into a real built workbook and asserts it is reported."""

  def setUp(self):
    self.wb = _workbook()
    self.assertEqual(design.audit_workbook(self.wb), [], "fixture must start clean")
    self.ws = self.wb["Dashboard"]

  def test_catches_a_raw_fill_color(self):
    self.ws["A1"].fill = PatternFill("solid", fgColor="FF00FF")
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("fill #FF00FF" in p for p in problems), problems)

  def test_catches_a_foreign_font_family(self):
    self.ws["A1"].font = Font(name="Comic Sans MS", size=11)
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("font family 'Comic Sans MS'" in p for p in problems), problems)

  def test_catches_a_font_size_outside_the_scale(self):
    self.ws["A1"].font = Font(name=design.FONT_FAMILY, size=7)
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("font size 7" in p for p in problems), problems)

  def test_catches_a_stray_number_format(self):
    self.ws["A1"].number_format = "0.000"
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("number format '0.000'" in p for p in problems), problems)

  def test_catches_a_chart_built_outside_the_helper(self):
    rogue = LineChart()
    rogue.add_data(Reference(self.ws, min_col=41, min_row=5, max_row=6))
    self.ws.add_chart(rogue, "AA100")
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("built outside design.chart()" in p for p in problems), problems)

  def test_catches_a_series_color_outside_the_palette(self):
    chart = [c for c in self.ws._charts if c.series][0]
    chart.series[0].graphicalProperties.line.solidFill = "FF00FF"
    problems = design.audit_workbook(self.wb)
    self.assertTrue(any("series color #FF00FF" in p for p in problems), problems)


if __name__ == "__main__":
  unittest.main()
