"""Phase 9 P3.20 Part 1b Concern 2 — Model Status fail-fast
behavior differentiation.

Differentiate environment failure (skip with warning) from
genuine status failure (always re-raise, in BOTH production and
test modes).

The fail-fast function `assert_workbook_model_status_ok` itself
already raises unconditionally on non-OK status — these tests
lock that contract in. The wiring change in intake_consult.py
(removing the test-mode gate) is exercised by the propagation
test below.
"""
from __future__ import annotations

import os
import sys
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
_PYTHON_DIR = _REPO_ROOT / "python"
if str(_PYTHON_DIR) not in sys.path:
  sys.path.insert(0, str(_PYTHON_DIR))


def _make_synthetic_workbook(model_status_value, tmp_path: Path) -> Path:
  import openpyxl
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Checks"
  ws.cell(row=2, column=1, value="Model Status")
  ws.cell(row=2, column=2, value=model_status_value)
  out = tmp_path / "synthetic_model_status_test.xlsx"
  wb.save(str(out))
  return out


class ModelStatusEnvFailureSkipTests(unittest.TestCase):
  """Env-failure path: when Excel COM recalc is unavailable, the
  fail-fast returns silently with a log warning (does NOT raise).
  This protects dev machines and production environments where
  Excel might be temporarily unavailable from spurious failures.
  """

  def setUp(self) -> None:
    self._tmp_dir = Path(_REPO_ROOT / "tmp" / "p3_20_concern2_env_failure_tests")
    self._tmp_dir.mkdir(parents=True, exist_ok=True)
    # Make sure test-mode does not influence behavior — the function
    # should behave the same regardless of CONVERGENCE_TEST_MODE.
    self._prev_test_mode = os.environ.get("CONVERGENCE_TEST_MODE")
    os.environ.pop("CONVERGENCE_TEST_MODE", None)

  def tearDown(self) -> None:
    if self._prev_test_mode is not None:
      os.environ["CONVERGENCE_TEST_MODE"] = self._prev_test_mode

  def _patch_recalc_to_env_failure(self, error_message: str):
    from client_intake_and_finmo.post_intake_runtime_validation import (  # type: ignore
      workbook_model_status as wms,
    )
    wms._recalc_workbook_via_excel_com = lambda _path: error_message
    return wms

  def test_pywin32_unavailable_skips_silently(self) -> None:
    """When the recalc helper returns a pywin32_unavailable error,
    the fail-fast returns without raising."""
    wms = self._patch_recalc_to_env_failure(
      "pywin32_unavailable: ImportError: No module named 'win32com'"
    )
    # Pre-populate the workbook with status FAIL to prove the env
    # failure short-circuits BEFORE the status check runs (no raise).
    wb_path = _make_synthetic_workbook("FAIL", self._tmp_dir)
    # Should not raise -- env failure path takes precedence.
    wms.assert_workbook_model_status_ok(str(wb_path))

  def test_excel_com_startup_failure_skips_silently(self) -> None:
    wms = self._patch_recalc_to_env_failure(
      "excel_com_failure: COMError: Excel.Application not available"
    )
    wb_path = _make_synthetic_workbook("FAIL", self._tmp_dir)
    wms.assert_workbook_model_status_ok(str(wb_path))


class ModelStatusAlwaysRaisesOnNonOkTests(unittest.TestCase):
  """When recalc succeeds and status is read, any non-'OK' value
  ALWAYS raises -- regardless of CONVERGENCE_TEST_MODE. This is
  the Concern 2 behavior change.
  """

  def setUp(self) -> None:
    self._tmp_dir = Path(_REPO_ROOT / "tmp" / "p3_20_concern2_always_raises_tests")
    self._tmp_dir.mkdir(parents=True, exist_ok=True)
    self._prev_test_mode = os.environ.get("CONVERGENCE_TEST_MODE")

  def tearDown(self) -> None:
    if self._prev_test_mode is None:
      os.environ.pop("CONVERGENCE_TEST_MODE", None)
    else:
      os.environ["CONVERGENCE_TEST_MODE"] = self._prev_test_mode

  def _patch_recalc_to_noop(self):
    from client_intake_and_finmo.post_intake_runtime_validation import (  # type: ignore
      workbook_model_status as wms,
    )
    wms._recalc_workbook_via_excel_com = lambda _path: None
    return wms

  def test_status_fail_in_production_mode_raises(self) -> None:
    """KEY CHECK: in production mode (no CONVERGENCE_TEST_MODE),
    a FAIL status still raises. Pre-Concern-2 the intake_consult
    wrapper suppressed this; the fail-fast function itself ALWAYS
    raised but the wrapper hid it. Test here verifies the function
    contract is correct independent of mode."""
    os.environ.pop("CONVERGENCE_TEST_MODE", None)
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = _make_synthetic_workbook("FAIL", self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed) as ctx:
      wms.assert_workbook_model_status_ok(str(wb_path))
    self.assertEqual(ctx.exception.operation, "workbook_model_status_fail")
    self.assertEqual(ctx.exception.actual, "FAIL")

  def test_status_fail_in_test_mode_raises(self) -> None:
    os.environ["CONVERGENCE_TEST_MODE"] = "true"
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = _make_synthetic_workbook("FAIL", self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed):
      wms.assert_workbook_model_status_ok(str(wb_path))

  def test_status_warn_in_production_mode_raises(self) -> None:
    """WARN is not OK; must raise in production."""
    os.environ.pop("CONVERGENCE_TEST_MODE", None)
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = _make_synthetic_workbook("WARN", self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed):
      wms.assert_workbook_model_status_ok(str(wb_path))

  def test_status_empty_in_production_mode_raises(self) -> None:
    """An empty Checks!B2 (None) must raise in production."""
    os.environ.pop("CONVERGENCE_TEST_MODE", None)
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = _make_synthetic_workbook(None, self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed):
      wms.assert_workbook_model_status_ok(str(wb_path))

  def test_status_ok_in_production_mode_passes(self) -> None:
    os.environ.pop("CONVERGENCE_TEST_MODE", None)
    wms = self._patch_recalc_to_noop()
    wb_path = _make_synthetic_workbook("OK", self._tmp_dir)
    wms.assert_workbook_model_status_ok(str(wb_path))


class IntakeConsultWrapperPropagatesNonOkTests(unittest.TestCase):
  """Confirms the Concern 2 wiring change: the intake_consult.py
  wrapper around assert_workbook_model_status_ok NO LONGER suppresses
  non-OK exceptions under production mode. The pre-Concern-2 code
  did a try/except that only re-raised under CONVERGENCE_TEST_MODE;
  the new code lets the exception propagate.

  Verified by checking the source code shape rather than triggering
  a full Flask request — the call site is a few lines deep in a
  multi-thousand-line API handler, so source-shape check is
  the focused regression guard.
  """

  def test_intake_consult_wrapper_no_longer_swallows_exceptions(self) -> None:
    src = (
      _REPO_ROOT / "python" / "api_handlers" / "intake_consult.py"
    ).read_text(encoding="utf-8")
    # The pre-Concern-2 pattern explicitly checked test-mode in the
    # except block. After Concern 2, that check is removed; the
    # exception from assert_workbook_model_status_ok propagates.
    self.assertNotIn(
      "if _cms_test_mode_enabled():",
      src,
      "intake_consult.py should no longer test CONVERGENCE_TEST_MODE around the workbook Model Status check",
    )
    # The current pattern: a separate try/except around the IMPORT
    # of the check module (env-failure-skip), followed by a bare
    # call to assert_workbook_model_status_ok (let it raise).
    self.assertIn(
      "workbook_model_status_check_module_unavailable",
      src,
      "intake_consult.py must keep the import-failure log-and-continue branch",
    )

  def test_intake_consult_wrapper_calls_check_before_email(self) -> None:
    src = (
      _REPO_ROOT / "python" / "api_handlers" / "intake_consult.py"
    ).read_text(encoding="utf-8")
    # The check should run AFTER export but BEFORE auto-email so a
    # broken workbook never reaches the customer.
    check_idx = src.find("assert_workbook_model_status_ok(client_workbook_path)")
    email_idx = src.find("# Auto-email the workbook")
    self.assertGreater(check_idx, 0, "Model Status check call not found")
    self.assertGreater(email_idx, 0, "Auto-email block not found")
    self.assertLess(check_idx, email_idx, "Model Status check must run BEFORE auto-email")


if __name__ == "__main__":
  unittest.main()
