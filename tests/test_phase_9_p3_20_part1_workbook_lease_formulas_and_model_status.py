"""Phase 9 P3.20 Part 1 — tests for:

1. Lease Interest formula now references the Interest Rate cell
   (not a Python literal)
2. Lease Asset Depreciation formula now references the Lease
   Opening Balance Q0 cell (not a Python literal)
3. Model Status fail-fast: returns silently on "OK", raises on
   any non-"OK" value, handles missing workbook gracefully.

These are workbook-builder-output tests (load the most recent
ExpressLogix workbook from disk and inspect formulas) plus
fail-fast unit tests using a synthetic minimal workbook for the
Model Status check.
"""

from __future__ import annotations

import os
import re

from openpyxl.utils import get_column_letter
import sys
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
_PYTHON_DIR = _REPO_ROOT / "python"
if str(_PYTHON_DIR) not in sys.path:
  sys.path.insert(0, str(_PYTHON_DIR))

_SCHEDULE_SHEETS_PATH = (
  _REPO_ROOT / "client_statements_output_excel" / "schedule_sheets.py"
)


class WorkbookLeaseFormulaSourceTests(unittest.TestCase):
  """Static source-shape checks: the formula generators must use
  cell references, not Python literal interpolation."""

  def setUp(self) -> None:
    self._src = _SCHEDULE_SHEETS_PATH.read_text(encoding="utf-8")

  def _row_formulas(self, rows, label):
    """The 21 quarter cells of one Debt Schedule row."""
    ws = self._debt_sheet()
    return [ws.cell(rows[label], 3 + i).value for i in range(21)]

  def _debt_rows(self):
    """Every Debt Schedule row KEY mapped to its row number.

    Read from the ctx registry, not from the text in column A. These tests
    used to key off the displayed label, so rewording the sheet broke four of
    them at once - which is the exact coupling the key/label separation was
    built to remove. The key is the contract; the label is presentation.
    """
    self._debt_sheet()
    return dict(type(self)._CTX.schedule_rows["Debt Schedule"])

  def _debt_sheet(self):
    """The built Debt Schedule, built once per process."""
    cached = getattr(type(self), "_SHEET", None)
    if cached is not None:
      return cached
    from replay_gate import _bootstrap
    _bootstrap.bind_root(None)
    from replay_gate.context import GateContext
    from client_statements_output_excel import data as wbdata, workbook_builder
    ctx = GateContext(_bootstrap.gate_connection(), _bootstrap.read_connection())
    wb, gap = ctx._build_workbook(
      workbook_builder.build_client_financial_model_workbook, wbdata.draft_data_from_row)
    assert not gap, gap
    type(self)._CTX = wb.build_context
    type(self)._SHEET = wb["Debt Schedule"]
    return type(self)._SHEET

  def test_lease_interest_formula_uses_cell_reference(self) -> None:
    """Lease interest reads a RATE CELL, never a Python literal.

    This used to regex the builder's source for one exact `ws.cell(...)`
    spelling, so it failed the moment the sheet was rearranged even though the
    property it names was untouched - the third time a source-text guard on
    this file has cost a rewrite. Asserted against the built sheet now.
    """
    rows = self._debt_rows()
    ws = self._debt_sheet()
    interest = rows["Lease Interest Expense"]
    opening = rows["Lease Opening Balance"]
    # The bridge row is a straight link into the schedule; the arithmetic
    # lives in the table it points at, so follow the one hop rather than
    # asserting against the link.
    from openpyxl.utils import column_index_from_string
    live = []
    for i in range(1, 21):
      link = ws.cell(interest, 3 + i).value
      if not isinstance(link, str) or not link.startswith("="):
        continue
      m = re.fullmatch(r"=([A-Z]{1,2})(\d+)", link.strip())
      self.assertIsNotNone(
        m, f"the lease interest bridge row is not a plain link: {link!r}")
      target = ws.cell(int(m.group(2)), column_index_from_string(m.group(1))).value
      if isinstance(target, str) and target.startswith("="):
        live.append(target)
    self.assertTrue(live, "no lease interest formulas were built")
    for formula in live:
      self.assertRegex(
        formula, r"[A-Z]{1,2}\d+\s*\*\s*[A-Z]{1,2}\d+",
        f"lease interest is not a balance times a rate CELL: {formula!r}")
      self.assertNotRegex(
        formula, r"\*\s*0?\.\d+",
        f"a literal rate is baked into the lease interest formula: {formula!r}")
      self.assertRegex(
        formula, r"^=[A-Z]{1,2}\d+\*[A-Z]{1,2}\d+$",
        f"lease interest is not exactly opening x rate: {formula!r}")

  def test_no_python_seed_literal_survives_in_the_generator(self):
    self.assertNotIn(
      "lease_seed_value = number(schedules.get(",
      self._src,
      "lease_seed_value local should be removed",
    )


def _latest_express_workbook() -> Path | None:
  base = Path("C:/dev/Cilient Plans")
  if not base.exists():
    return None
  matches = sorted(
    base.glob("ExpressLogix Shipping Services -- *.xlsx"),
    key=lambda p: p.stat().st_mtime,
    reverse=True,
  )
  return matches[0] if matches else None


class WorkbookLeaseFormulaOutputTests(unittest.TestCase):
  """If a recent ExpressLogix workbook exists AND was generated
  after the schedule_sheets.py source, load it and verify the
  Excel formulas use the cell-reference pattern."""

  def setUp(self) -> None:
    self._workbook_path = _latest_express_workbook()
    if self._workbook_path is None:
      self.skipTest("No ExpressLogix workbook present on disk")
    wb_mtime = self._workbook_path.stat().st_mtime
    src_mtime = _SCHEDULE_SHEETS_PATH.stat().st_mtime
    if wb_mtime < src_mtime:
      self.skipTest(
        f"Latest workbook {self._workbook_path.name} is older than "
        "schedule_sheets.py source; rerun an E2E to regenerate before this check is meaningful."
      )

  def test_lease_interest_formula_references_interest_rate_cell(self) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(str(self._workbook_path), data_only=False)
    debt = wb["Debt Schedule"]
    # Find rows for Lease Interest Expense and Interest Rate
    lease_int_row = ir_row = None
    for r in range(1, 60):
      label = debt.cell(row=r, column=1).value
      if label == "Lease Interest Expense":
        lease_int_row = r
      elif label == "Interest Rate":
        ir_row = r
    self.assertIsNotNone(lease_int_row, "Debt Schedule must have Lease Interest Expense row")
    self.assertIsNotNone(ir_row, "Debt Schedule must have Interest Rate row")
    # Q1 = column D. Formula should reference D<ir_row>.
    q1_formula = debt.cell(row=lease_int_row, column=4).value
    self.assertIsNotNone(q1_formula, "Lease Interest Q1 cell should not be None/blank")
    expected_ref = f"D{ir_row}"
    self.assertIn(
      expected_ref, str(q1_formula),
      f"Lease Interest Q1 formula must reference {expected_ref} (Interest Rate Q1). Got: {q1_formula!r}",
    )
    # Confirm there's NO bare numeric literal (other than the rate cell ref) — heuristic
    self.assertNotRegex(
      str(q1_formula), r"\*\s*0\.\d+",
      f"Lease Interest Q1 formula should not contain a bare *0.XX numeric literal. Got: {q1_formula!r}",
    )

  def test_lease_asset_depreciation_formula_references_lease_opening_q0_cell(self) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(str(self._workbook_path), data_only=False)
    debt = wb["Debt Schedule"]
    lease_dep_row = lease_open_row = None
    for r in range(1, 60):
      label = debt.cell(row=r, column=1).value
      if label == "Lease Asset Depreciation":
        lease_dep_row = r
      elif label == "Lease Opening Balance":
        lease_open_row = r
    self.assertIsNotNone(lease_dep_row, "Debt Schedule must have Lease Asset Depreciation row")
    self.assertIsNotNone(lease_open_row, "Debt Schedule must have Lease Opening Balance row")
    # Q0 = column C. Formula should reference C<lease_open_row>.
    q1_formula = debt.cell(row=lease_dep_row, column=4).value
    self.assertIsNotNone(q1_formula, "Lease Asset Depreciation Q1 cell should not be None")
    expected_ref = f"C{lease_open_row}"
    self.assertIn(
      expected_ref, str(q1_formula),
      f"Lease Asset Depreciation Q1 formula must reference {expected_ref} (Lease Opening Balance Q0). Got: {q1_formula!r}",
    )
    # Confirm there's no bare integer literal followed by /20 (the old pattern)
    self.assertNotRegex(
      str(q1_formula), r"\(\d{2,}/20\)",
      f"Lease Asset Depreciation Q1 formula should not contain a (NNNNN/20) literal. Got: {q1_formula!r}",
    )


class WorkbookModelStatusFailFastTests(unittest.TestCase):
  """Unit tests for assert_workbook_model_status_ok using a
  synthetic minimal workbook (no Excel recalc required since we
  pre-populate the Checks!B2 value)."""

  def _make_workbook(self, model_status_value, tmp_path: Path) -> Path:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checks"
    ws.cell(row=2, column=1, value="Model Status")
    ws.cell(row=2, column=2, value=model_status_value)
    out = tmp_path / "synthetic_workbook.xlsx"
    wb.save(str(out))
    return out

  def setUp(self) -> None:
    self._tmp_dir = Path(_REPO_ROOT / "tmp" / "p3_20_model_status_tests")
    self._tmp_dir.mkdir(parents=True, exist_ok=True)

  def _patch_recalc_to_noop(self):
    # Monkey-patch the recalc helper to return success without
    # actually opening Excel; the test workbooks are pre-populated.
    from client_intake_and_finmo.post_intake_runtime_validation import (  # type: ignore
      workbook_model_status as wms,
    )
    wms._recalc_workbook_via_excel_com = lambda _path: None
    return wms

  def test_status_ok_passes_silently(self) -> None:
    wms = self._patch_recalc_to_noop()
    wb_path = self._make_workbook("OK", self._tmp_dir)
    # Should not raise
    wms.assert_workbook_model_status_ok(str(wb_path))

  def test_status_fail_raises(self) -> None:
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = self._make_workbook("FAIL", self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed) as ctx:
      wms.assert_workbook_model_status_ok(str(wb_path))
    err = ctx.exception
    self.assertEqual(err.operation, "workbook_model_status_fail")
    self.assertEqual(err.actual, "FAIL")
    self.assertIn("app", err.details.get("guidance", "").lower())

  def test_status_warn_also_raises(self) -> None:
    """Anything other than the exact string 'OK' must fail."""
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = self._make_workbook("WARN", self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed):
      wms.assert_workbook_model_status_ok(str(wb_path))

  def test_status_none_raises(self) -> None:
    """An empty Checks!B2 (None) should still fail since it's not 'OK'."""
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    wms = self._patch_recalc_to_noop()
    wb_path = self._make_workbook(None, self._tmp_dir)
    with self.assertRaises(PostIntakePreconditionFailed):
      wms.assert_workbook_model_status_ok(str(wb_path))

  def test_missing_workbook_raises_precondition(self) -> None:
    from client_intake_and_finmo.fail_fast.common import PostIntakePreconditionFailed  # type: ignore
    from client_intake_and_finmo.post_intake_runtime_validation import (  # type: ignore
      workbook_model_status as wms,
    )
    bogus = self._tmp_dir / "definitely_not_present.xlsx"
    if bogus.exists():
      bogus.unlink()
    with self.assertRaises(PostIntakePreconditionFailed) as ctx:
      wms.assert_workbook_model_status_ok(str(bogus))
    self.assertEqual(ctx.exception.operation, "workbook_model_status_workbook_missing")


if __name__ == "__main__":
  unittest.main()
