import os
import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook

# ============================================================
# ENV
# ============================================================

load_dotenv()

FINMO_PATH = os.getenv("FINMO")

MYSQL_CONFIG = {
    "host": os.getenv("MYSQL_HOST"),
    "port": int(os.getenv("MYSQL_PORT", "3306")),
    "user": os.getenv("MYSQL_USER"),
    "password": os.getenv("MYSQL_PASSWORD"),
    "database": os.getenv("MYSQL_DB"),
}

# ============================================================
# DB HELPERS
# ============================================================

def get_conn():
    return mysql.connector.connect(**MYSQL_CONFIG)

# ============================================================
# LOOKUPS
# ============================================================

def get_naics_from_business_type(conn, business_type):
    """
    Resolve client-selected business_type -> 6-digit NAICS
    """
    bt_clean = business_type.strip()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT business_types, naics_6 FROM naics_master")
    rows = cur.fetchall()
    cur.close()

    for row in rows:
        bt_field = row.get("business_types") or ""
        tokens = [tok.strip() for tok in str(bt_field).split(",") if tok.strip()]
        for tok in tokens:
            if tok == bt_clean:
                return row["naics_6"]

    raise ValueError(f"No NAICS mapping found for business_type: {business_type}")


def get_growth_naics_from_index(conn, naics_6):
    """
    Pull the already-resolved growth NAICS from industry_growth_index
    """
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT naics_code, naics_level, trust
        FROM industry_growth_index
        WHERE %s LIKE CONCAT(naics_code, '%')
          AND usable = 1
        ORDER BY naics_level DESC
        LIMIT 1
    """, (naics_6,))
    row = cur.fetchone()
    cur.close()

    if not row:
        raise ValueError(f"No usable growth index entry for NAICS {naics_6}")

    return row


def get_soi_naics(conn, growth_naics, naics_level):
    """
    Confirm SOI exists at the chosen NAICS level.
    If missing (rare), fallback vertically.
    """
    cur = conn.cursor(dictionary=True)

    col_map = {
        6: "naics_6_digit",
        5: "naics_5_digit",
        4: "naics_4_digit",
        3: "naics_3_digit",
        2: "naics_2_digit",
    }

    for level in range(naics_level, 1, -1):
        col = col_map.get(level)
        if not col:
            continue
        cur.execute(f"""
            SELECT DISTINCT {col} AS naics_code
            FROM soi_corporate_tax_returns
            WHERE {col} = %s
            LIMIT 1
        """, (growth_naics[:level],))

        row = cur.fetchone()
        if row:
            cur.close()
            return row["naics_code"]

    cur.close()
    raise ValueError(f"No SOI data found for NAICS {growth_naics}")

# ============================================================
# EXCEL WRITER
# ============================================================

def write_named_cell(wb, name, value):
    """
    Write to a named range (single-cell)
    """
    defined_name = wb.defined_names[name]
    sheet_name, cell = list(defined_name.destinations)[0]
    ws = wb[sheet_name]
    ws[cell] = value

# ============================================================
# MAIN PIPELINE
# ============================================================

def populate_finmo(business_type):
    conn = get_conn()

    try:
        # --- Resolve NAICS ---
        naics_6 = get_naics_from_business_type(conn, business_type)

        growth_row = get_growth_naics_from_index(conn, naics_6)
        growth_naics = growth_row["naics_code"]
        growth_level = growth_row["naics_level"]

        soi_naics = get_soi_naics(conn, growth_naics, growth_level)

        # --- Load Excel ---
        wb = load_workbook(FINMO_PATH)

        write_named_cell(wb, "business_type", business_type)
        write_named_cell(wb, "business_type_naics", naics_6)
        write_named_cell(wb, "growth_rate_naics", growth_naics)
        write_named_cell(wb, "soi_corp_base", soi_naics)

        wb.save(FINMO_PATH)

        info = {
            "business_type": business_type,
            "business_type_naics": naics_6,
            "growth_rate_naics": growth_naics,
            "soi_corp_base": soi_naics,
        }
        print(f"Populated FINMO: {info}")
        return info

    finally:
        conn.close()

# ============================================================
# USAGE
# ============================================================

# Example:
# populate_finmo("Acting School")
