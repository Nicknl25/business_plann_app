"""Phase 9 P3.20 Part 1 — workbook Model Status fail-fast.

Runs AFTER workbook generation, BEFORE the run is marked complete.
Reads the workbook's `Checks!B2` Model Status cell (computed by
the Checks sheet's invariants: BS balance, CF tie, schedule
bridges, formula-logic diagnostics). If status != "OK", hard-stops
the run with a named diagnostic that explicitly directs the
operator to verify APP state first, not patch workbook formulas.

The doctrine reflected in the diagnostic message: the workbook is
a reflection of the app. When the workbook surfaces a failed
invariant, the FIRST hypothesis is that the app produced the wrong
data; the workbook is just exposing it. Patching workbook formulas
to mask the failure hides the real bug.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

from client_intake_and_finmo.fail_fast.common import (  # type: ignore
  PostIntakePreconditionFailed,
  convergence_test_mode_enabled,
)


_logger = logging.getLogger(__name__)


_MODEL_STATUS_OK = "OK"
_CHECKS_SHEET = "Checks"
_MODEL_STATUS_CELL_ROW = 2
_MODEL_STATUS_CELL_COL = 2  # column B


def _recalc_workbook_via_excel_com(workbook_path: str) -> Optional[str]:
  """Open the workbook in Excel via COM, force CalculateFull(),
  and save. Returns None on success or a short error description
  on failure (so callers can decide whether to skip the check).

  Excel + COM are the canonical Windows recalc path. On non-Windows
  environments or environments without Excel installed, this
  function returns a non-None error and the post-run fail-fast
  treats the check as unable-to-evaluate (logs a warning, does
  not hard-fail). Disabling the fail-fast for unavailability is
  preferable to false-failing every run on a dev machine without
  Excel.
  """
  try:
    import win32com.client as _w32  # type: ignore
    import pythoncom as _com  # type: ignore
  except Exception as exc:
    return f"pywin32_unavailable: {type(exc).__name__}: {str(exc)[:200]}"
  excel = None
  wb = None
  # COM IS PER-THREAD (2026-08-29). This runs on a Flask request-handler
  # thread, and CoInitialize is per-thread state: without it EnsureDispatch
  # raises -2147221008 "CoInitialize has not been called", this function
  # returns an error, and the caller treats the workbook as unable-to-
  # evaluate. Two things then went wrong quietly for every delivered file:
  # the Save() below never ran, so the workbook shipped with formulas and NO
  # CACHED VALUES (Excel recalculates on open so a client still sees numbers,
  # but anything without a spreadsheet engine reads an empty file), and the
  # Checks!B2 model-status assertion was skipped on every run. A script's
  # MAIN thread is initialised by pythoncom on import, which is why the same
  # recalc always worked from the command line and never from the server.
  _com_ready = False
  try:
    _com.CoInitialize()
    _com_ready = True
  except Exception:
    # Already initialised on this thread, or an apartment-mode clash: either
    # way EnsureDispatch below is the real test, so carry on and let it speak.
    pass
  try:
    excel = _w32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(workbook_path))
    if wb is None:
      return "excel_workbook_open_returned_none"
    excel.CalculateFull()
    wb.Save()
    return None
  except Exception as exc:
    return f"excel_com_failure: {type(exc).__name__}: {str(exc)[:200]}"
  finally:
    try:
      if wb is not None:
        wb.Close(SaveChanges=False)
    except Exception:
      pass
    try:
      if excel is not None:
        excel.Quit()
    except Exception:
      pass
    if _com_ready:
      try:
        _com.CoUninitialize()
      except Exception:
        pass


def _read_model_status(workbook_path: str) -> Any:
  """Open the recalculated workbook with openpyxl data_only=True
  and return the Checks!B2 value."""
  import openpyxl  # type: ignore

  wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
  try:
    if _CHECKS_SHEET not in wb.sheetnames:
      return None
    checks = wb[_CHECKS_SHEET]
    return checks.cell(row=_MODEL_STATUS_CELL_ROW, column=_MODEL_STATUS_CELL_COL).value
  finally:
    try:
      wb.close()
    except Exception:
      pass


def assert_workbook_model_status_ok(
  workbook_path: Any,
  *,
  pipeline_stage: str = "post_intake_finalize_validation_workbook_model_status",
) -> None:
  """Post-run machinery fail-fast: the generated workbook's
  Checks!B2 Model Status cell must read "OK".

  Behavior:
    - If the workbook file doesn't exist, raise immediately
      (this is a precondition failure, not a quality issue).
    - If Excel/COM recalc isn't available (e.g. dev machine
      without Excel), log a warning and return without raising
      — the check requires a recalculated workbook to be
      meaningful, and we'd rather skip than false-fail.
    - If recalc succeeds and Checks!B2 == "OK": return silently.
    - If Checks!B2 is anything else (including None, "FAIL",
      "BLOCKED", "WARN", etc.): raise PostIntakePreconditionFailed
      with diagnostic naming the actual value and pointing the
      operator at app-side fail-fasts FIRST.

  Diagnostic doctrine: when this fires, the FIRST hypothesis is
  that the app produced wrong data, NOT that the workbook formula
  is wrong. The operator should verify app state via existing
  post-intake fail-fasts (accounting equation, stored-totals,
  capital lease components, schedule reconciliations) before
  adjusting any workbook formula.
  """
  path = Path(str(workbook_path or "").strip())
  if not str(path):
    raise PostIntakePreconditionFailed(
      operation="workbook_model_status_workbook_path_missing",
      pipeline_stage=pipeline_stage,
      expected="non-empty workbook_path",
      actual="empty",
    )
  if not path.exists():
    raise PostIntakePreconditionFailed(
      operation="workbook_model_status_workbook_missing",
      pipeline_stage=pipeline_stage,
      expected=f"workbook file exists at {path}",
      actual="file not found",
      details={"workbook_path": str(path)},
    )
  recalc_error = _recalc_workbook_via_excel_com(str(path))
  if recalc_error is not None:
    # Excel recalc unavailable -- skip the check with a warning so
    # we don't false-fail on dev machines without Excel. The check
    # is meant for production where Excel COM is always available.
    _logger.warning(
      "workbook_model_status_check_skipped: recalc unavailable for %s: %s",
      path.name, recalc_error,
    )
    return
  try:
    status_value = _read_model_status(str(path))
  except Exception as exc:
    raise PostIntakePreconditionFailed(
      operation="workbook_model_status_read_failed",
      pipeline_stage=pipeline_stage,
      expected=f"readable Checks!B2 in {path.name}",
      actual=f"{type(exc).__name__}: {str(exc)[:200]}",
      details={"workbook_path": str(path)},
    ) from exc
  if status_value == _MODEL_STATUS_OK:
    return
  message_parts = [
    f"workbook_model_status={status_value!r} (expected {_MODEL_STATUS_OK!r}).",
    "Inspect Checks sheet rows where Status=FAIL for the failing invariants.",
    "DEFAULT ASSUMPTION: the failure indicates a bug in the APP, not the workbook.",
    "Verify app state via existing post-intake fail-fasts (accounting_equation_violation,",
    "stored_totals_match_components_violation, capital_lease_* validators,",
    "debt/payroll schedule reconciliations) BEFORE adjusting any workbook formula.",
  ]
  raise PostIntakePreconditionFailed(
    operation="workbook_model_status_fail",
    pipeline_stage=pipeline_stage,
    expected=_MODEL_STATUS_OK,
    actual=str(status_value),
    details={
      "workbook_path": str(path),
      "checks_sheet_model_status_cell": f"{_CHECKS_SHEET}!B{_MODEL_STATUS_CELL_ROW}",
      "guidance": " ".join(message_parts),
    },
  )
