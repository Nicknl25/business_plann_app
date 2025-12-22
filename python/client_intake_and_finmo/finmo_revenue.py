from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Optional

try:
  from dotenv import load_dotenv  # type: ignore
except Exception:  # pragma: no cover
  load_dotenv = None  # type: ignore

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

DEFAULT_STATE_PATH = (
  Path(__file__).resolve().parent / "temp" / "latest_intake.json"
)


def _maybe_load_dotenv() -> None:
  if load_dotenv is None:
    return
  try:
    load_dotenv()
  except Exception:
    pass


def _require_finmo_path(finmo_path: Optional[str] = None) -> str:
  if finmo_path is not None and str(finmo_path).strip():
    return str(finmo_path).strip()

  _maybe_load_dotenv()
  env_path = os.getenv("FINMO")
  if not env_path or not str(env_path).strip():
    raise RuntimeError("FINMO environment variable is not set.")
  return str(env_path).strip()


def _write_named_cell(wb, name: str, value: Any) -> None:
  if name not in wb.defined_names:
    raise KeyError(f"Named range not found in workbook: {name}")

  defined_name = wb.defined_names[name]
  destinations = list(defined_name.destinations)
  if not destinations:
    raise ValueError(f"Named range has no destinations: {name}")

  sheet_name, cell = destinations[0]
  ws = wb[sheet_name]
  ws[cell] = value


def _read_named_cell(wb, name: str) -> Any:
  if name not in wb.defined_names:
    raise KeyError(f"Named range not found in workbook: {name}")

  defined_name = wb.defined_names[name]
  destinations = list(defined_name.destinations)
  if not destinations:
    raise ValueError(f"Named range has no destinations: {name}")

  sheet_name, cell = destinations[0]
  ws = wb[sheet_name]
  return ws[cell].value


def _to_float(value: Any) -> Optional[float]:
  if value is None:
    return None
  if isinstance(value, (int, float)):
    return float(value)
  raw = str(value).strip()
  if not raw:
    return None
  raw = raw.replace(",", "")
  try:
    return float(raw)
  except Exception:
    return None


def record_latest_intake(
  *,
  current_revenue: float,
  state_path: Path = DEFAULT_STATE_PATH,
  extra: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
  state_path.parent.mkdir(parents=True, exist_ok=True)

  payload: Dict[str, Any] = {
    "current_revenue": float(current_revenue),
    "submitted_at_utc": datetime.now(timezone.utc).isoformat(),
  }
  if extra:
    payload.update(extra)

  state_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
  return payload


def read_latest_intake(
  *, state_path: Path = DEFAULT_STATE_PATH
) -> Dict[str, Any]:
  if not state_path.exists():
    raise FileNotFoundError(
      f"Latest intake state not found at {state_path}. Submit the intake form first."
    )
  data = json.loads(state_path.read_text(encoding="utf-8"))
  if not isinstance(data, dict):
    raise ValueError(f"Invalid latest intake state at {state_path}.")
  return data


def read_intake_submission_from_db(*, client_id: str) -> Dict[str, Any]:
  _maybe_load_dotenv()
  try:
    from intake_submission import get_mysql_connection  # type: ignore
  except Exception as exc:
    raise RuntimeError("Unable to import MySQL helper get_mysql_connection.") from exc

  if not client_id or not str(client_id).strip():
    raise ValueError("client_id is required to load an intake submission.")

  conn = get_mysql_connection()
  try:
    cur = conn.cursor(dictionary=True)
    try:
      cur.execute(
        """
        SELECT id, client_id, current_revenue, finmo_path, created_at
        FROM intake_submissions
        WHERE client_id = %s
        LIMIT 1
        """
        ,
        (str(client_id).strip(),),
      )
      row = cur.fetchone()
    finally:
      try:
        cur.close()
      except Exception:
        pass
  finally:
    try:
      conn.close()
    except Exception:
      pass

  if not row:
    raise RuntimeError(f"No intake submission found for client_id={client_id!r}.")
  if not isinstance(row, dict):
    raise RuntimeError("Unexpected DB row shape for intake submission.")
  return row


def calculate_soi_business_receipts_per_return(
  *,
  soi_corp_base: str,
) -> Dict[str, Any]:
  _maybe_load_dotenv()
  try:
    from intake_submission import get_mysql_connection  # type: ignore
  except Exception as exc:
    raise RuntimeError("Unable to import MySQL helper get_mysql_connection.") from exc

  code = str(soi_corp_base).strip()
  if not code:
    raise ValueError("soi_corp_base is required.")
  if not code.isdigit():
    raise ValueError(f"soi_corp_base must be numeric NAICS code, got {code!r}")

  col_map = {
    6: "naics_6_digit",
    5: "naics_5_digit",
    4: "naics_4_digit",
    3: "naics_3_digit",
    2: "naics_2_digit",
  }

  conn = get_mysql_connection()
  try:
    cur = conn.cursor(dictionary=True)
    try:
      rows = None
      matched_level = None
      for level in range(min(6, len(code)), 1, -1):
        col = col_map[level]
        cur.execute(
          f"SELECT * FROM soi_corporate_tax_returns WHERE {col} = %s",
          (code[:level],),
        )
        fetched = cur.fetchall()
        if fetched:
          rows = fetched
          matched_level = level
          break
    finally:
      try:
        cur.close()
      except Exception:
        pass
  finally:
    try:
      conn.close()
    except Exception:
      pass

  if not rows:
    raise RuntimeError(
      f"No soi_corporate_tax_returns row found for soi_corp_base={code!r}"
    )

  sum_all_returns = 0.0
  sum_total = 0.0
  skipped = 0
  for row in rows:
    if not isinstance(row, dict):
      skipped += 1
      continue
    returns_val = _to_float(row.get("all_returns"))
    total_val = _to_float(row.get("total"))
    if returns_val is None or total_val in (None, 0.0):
      skipped += 1
      continue
    sum_all_returns += returns_val
    sum_total += total_val

  if sum_total == 0.0:
    raise RuntimeError(
      f"SOI rows found for soi_corp_base={code!r}, but none had usable all_returns/total."
    )

  ratio = (sum_all_returns / sum_total) * 1000.0
  print(
    "SOI all_returns/total (ratio of sums) for soi_corp_base=%s "
    "(matched_level=%s, row_count=%s, skipped=%s): %s / %s = %s"
    % (
      code,
      matched_level,
      len(rows),
      skipped,
      sum_all_returns,
      sum_total,
      ratio,
    )
  )
  return {
    "soi_corp_base": code,
    "matched_level": matched_level,
    "row_count": len(rows),
    "sum_all_returns": sum_all_returns,
    "sum_total": sum_total,
    "skipped_rows": skipped,
    "ratio": ratio,
  }


def write_soi_revenue_total_all_firms_to_finmo(
  *,
  client_id: str,
  soi_corp_base: Optional[str] = None,
) -> Dict[str, Any]:
  submission = read_intake_submission_from_db(client_id=client_id)
  finmo_path = submission.get("finmo_path")
  if not finmo_path or not str(finmo_path).strip():
    raise ValueError("Intake submission is missing finmo_path.")

  wb = load_workbook(str(finmo_path).strip())
  soi_code = soi_corp_base
  if not soi_code:
    soi_code = str(_read_named_cell(wb, "soi_corp_base") or "").strip()
  if not soi_code:
    raise ValueError("Unable to determine soi_corp_base for this submission.")

  soi_calc = calculate_soi_business_receipts_per_return(soi_corp_base=soi_code)
  ratio = soi_calc["ratio"]

  if "model_selection_range" not in wb.defined_names:
    raise KeyError("Named range not found in workbook: model_selection_range")
  defined_name = wb.defined_names["model_selection_range"]
  destinations = list(defined_name.destinations)
  if not destinations:
    raise ValueError("Named range has no destinations: model_selection_range")

  sheet_name, rng = destinations[0]
  ws = wb[sheet_name]
  min_col, min_row, max_col, max_row = range_boundaries(rng)

  label = "SOI Revenue Total (all firms in industry)"
  found = False
  for row_idx in range(min_row, max_row + 1):
    for col_idx in range(min_col, max_col + 1):
      cell = ws.cell(row=row_idx, column=col_idx)
      val = cell.value
      if isinstance(val, str) and val.strip() == label:
        ws.cell(row=row_idx, column=col_idx + 1).value = float(ratio)
        found = True
        break
    if found:
      break

  if not found:
    raise RuntimeError(f"Label not found in model_selection_range: {label!r}")

  wb.save(str(finmo_path).strip())

  return {
    "finmo_path": str(finmo_path).strip(),
    "soi_corp_base": soi_code,
    "soi_receipts_per_return": float(ratio),
  }


def write_starting_revenue_intake_to_finmo(
  *,
  current_revenue: float,
  finmo_path: Optional[str] = None,
) -> Dict[str, Any]:
  finmo_path = _require_finmo_path(finmo_path)
  wb = load_workbook(finmo_path)
  _write_named_cell(wb, "starting_revenue_intake", float(current_revenue))
  wb.save(finmo_path)
  return {
    "finmo_path": finmo_path,
    "starting_revenue_intake": float(current_revenue),
  }


def sync_intake_revenue_to_finmo(
  *,
  client_id: str,
) -> Dict[str, Any]:
  submission = read_intake_submission_from_db(client_id=client_id)

  revenue = submission.get("current_revenue", None)
  if revenue is None:
    raise ValueError("Intake submission is missing current_revenue.")
  try:
    revenue_value = float(revenue)
  except Exception as exc:
    raise ValueError(
      f"Latest intake current_revenue is not numeric: {revenue}"
    ) from exc

  submission_finmo_path = submission.get("finmo_path")
  if not submission_finmo_path or not str(submission_finmo_path).strip():
    raise ValueError("Intake submission is missing finmo_path.")

  result = write_starting_revenue_intake_to_finmo(
    current_revenue=revenue_value,
    finmo_path=str(submission_finmo_path).strip(),
  )
  return {"intake_submission": submission, **result}


if __name__ == "__main__":  # pragma: no cover
  # Manual utility: write a specific submission's revenue into its FINMO workbook.
  arg_client_id = sys.argv[1] if len(sys.argv) > 1 else os.getenv("CLIENT_ID")
  if not arg_client_id:
    raise SystemExit("Usage: python finmo_revenue.py <client_id> (or set CLIENT_ID)")
  print(sync_intake_revenue_to_finmo(client_id=arg_client_id))
