"""Shared issue database for autonomous persona testing.

The Cowork testing app (persona driver) DETECTS issues and writes them here;
this side owns the schema, the recurrence/resolution accounting, and the
dashboard. Lives in the same MySQL store as the run_vitals_* tables so an
issue joins to its run's backend truth natively (same draft_id, one join).

Three tables
------------
- ``issues`` — the registry: ONE row per unique issue, keyed by
  ``signature``. This table is the deliberate UPDATE exception to the
  project's INSERT-only discipline: it is a REGISTRY (current state of a
  known issue), not a log. Every state change it undergoes is mirrored by an
  INSERT-only audit row in ``issue_resolution_events``, so history is never
  lost to an UPDATE.
- ``issue_occurrences`` — INSERT-only: one row per sighting of an issue on a
  specific run (draft_id + turn/section + what-happened/what-should-have).
- ``issue_resolution_events`` — INSERT-only audit of the resolution-sensing:
  exercised-clean ticks, recurrences, reopens, and the resolution verdicts.

Signature (the crux of resolution-sensing)
------------------------------------------
``signature`` is the stable identity of "this same issue" across runs.
Authored by the reporter (Cowork), following the documented convention
(docs/issue_database.md): ``<category>:<locus...>`` built ONLY from stable
coordinates — section, stage, field/question id, failure class, business
slug. NEVER from run-minted tokens (draft_id, timestamps) or verbatim GPT
text (varies run to run). The registry upserts on it; recurrence == a new
occurrence arriving under an existing signature.

Resolution-sensing honesty (resolution_class)
---------------------------------------------
- ``hard`` (default for hard_break / verdict / flow): deterministically
  resolvable. When a later run EXERCISES the issue's path (probe_json
  predicate against that run's vitals) and the failure does not recur, that
  is an exercised-clean tick; at ``hard_clean_threshold`` consecutive clean
  exercises the issue is resolved with resolution_confidence='confirmed'.
- ``soft`` (default for experience): NOT deterministically resolvable — GPT
  phrasing varies run to run, so absence is weak evidence. Soft issues can
  only ever reach resolution_basis='not_seen_n_runs' with
  resolution_confidence='observational', after ``soft_runs_threshold``
  exercised runs without a recurrence. The schema forbids 'confirmed' on a
  soft issue by construction (only the hard path writes it).

Artifacts, not intentions (the CW-031 law)
------------------------------------------
The probe clauses ``section``/``stage_like``/``business_like``/``min_turns``
answer only "did this run go DOWN the issue's path" — OPPORTUNITY. They cannot
see whether the path produced the right number, because they never read a
persisted value. For a long time that was the whole detector, so
'resolved confirmed' actually meant no more than "a run finished, visited the
same section, and the reporter did not re-file the signature". #138 was
resolved-confirmed by that rule on the very run whose workbook disproves it:
the app PROPOSED a per-line COGS split (prose), the reporter saw the proposal
and stayed quiet, and the four product rows were written null anyway.

So: an issue may only reach resolution_confidence='confirmed' when its probe
carries an ``artifact`` assertion and that assertion was READ on the run and
HELD. Everything else — opportunity clauses, reporter silence, metadata-only
probes — is capped at 'observational' and must clear the soft threshold. An
artifact assertion that FAILS is a recurrence, not a quiet run: the registry
reopens the issue on its own evidence without waiting to be told.

  "artifact": [{"kind": "ops_per_line_cogs", "min_lines": 2}]

An assertion whose precondition is absent on this run (e.g. a multi-line
assertion on a single-line business) returns not_applicable, which counts as
NOT EXERCISED — absence of opportunity is never evidence.

Any recurrence resets both counters; a recurrence after resolution flips the
issue to status='recurring' and increments reopened_count.

Write contract style: report_issue() FAILS LOUDLY on a bad category /
severity / status vocabulary — a typo in the write path must surface
immediately, not create a phantom taxonomy.
"""

from __future__ import annotations

import json
import re
from typing import Any, Dict, List, Optional


ISSUES_TABLE = "issues"
OCCURRENCES_TABLE = "issue_occurrences"
RESOLUTION_EVENTS_TABLE = "issue_resolution_events"

CATEGORIES = ("hard_break", "flow", "verdict", "experience")
SEVERITIES = ("blocker", "major", "minor", "note")
STATUSES = ("open", "resolved", "recurring")
RESOLUTION_CLASSES = ("hard", "soft")
RESOLUTION_BASES = ("artifact_verified", "retested_clean", "not_seen_n_runs",
                    "manual")
RESOLUTION_CONFIDENCES = ("confirmed", "observational")

# --- Probe vocabulary -------------------------------------------------------
# OPPORTUNITY clauses answer "did this run go down the issue's path?". They are
# necessary and NEVER sufficient: a run can walk the path and still produce the
# broken artifact. On their own they only ever earn 'observational'.
PROBE_OPPORTUNITY_KEYS = {
  "require_completed", "sections", "section", "call_label_like", "stage_like",
  "business_like", "min_turns", "manual_only", "auto_recur",
}
# METADATA carries no predicate at all. A probe made only of metadata is a
# probe with NO retest condition — it must never tick anything.
PROBE_METADATA_KEYS = {"note", "regression_pin", "title"}
# The ARTIFACT clause is the only evidence that can earn 'confirmed'.
PROBE_ARTIFACT_KEYS = {"artifact"}
PROBE_KEYS = PROBE_OPPORTUNITY_KEYS | PROBE_METADATA_KEYS | PROBE_ARTIFACT_KEYS

ARTIFACT_KINDS = ("ops_per_line_cogs", "ops_field_non_null", "workbook_cogs_rows",
                  "ops_cogs_shared_group")

# Category -> default resolution class. flow defaults to hard (loops /
# dead-ends are routing logic, deterministically re-testable); the reporter
# may override to soft for phrasing-dependent flow issues.
DEFAULT_RESOLUTION_CLASS = {
  "hard_break": "hard",
  "verdict": "hard",
  "flow": "hard",
  "experience": "soft",
}

DEFAULT_HARD_CLEAN_THRESHOLD = 1
DEFAULT_SOFT_RUNS_THRESHOLD = 5

_CREATE_ISSUES_SQL = f"""
CREATE TABLE IF NOT EXISTS {ISSUES_TABLE} (
  issue_id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
  signature VARCHAR(191) NOT NULL,
  category VARCHAR(24) NOT NULL,
  resolution_class VARCHAR(8) NOT NULL,
  severity VARCHAR(12) NOT NULL,
  title VARCHAR(255) NOT NULL DEFAULT '',
  status VARCHAR(16) NOT NULL DEFAULT 'open',
  first_seen_at DATETIME(6) NOT NULL,
  last_seen_at DATETIME(6) NOT NULL,
  occurrence_count INT NOT NULL DEFAULT 1,
  reopened_count INT NOT NULL DEFAULT 0,
  clean_exercise_count INT NOT NULL DEFAULT 0,
  runs_since_last_seen INT NOT NULL DEFAULT 0,
  resolved_detected_at DATETIME(6) NULL,
  resolution_basis VARCHAR(32) NULL,
  resolution_confidence VARCHAR(16) NULL,
  probe_json LONGTEXT NULL,
  created_at DATETIME(6) NOT NULL DEFAULT CURRENT_TIMESTAMP(6),
  updated_at DATETIME(6) NOT NULL DEFAULT CURRENT_TIMESTAMP(6)
    ON UPDATE CURRENT_TIMESTAMP(6),
  UNIQUE KEY uniq_signature (signature),
  KEY ix_issues_status (status),
  KEY ix_issues_category (category),
  KEY ix_issues_last_seen (last_seen_at)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""

_CREATE_OCCURRENCES_SQL = f"""
CREATE TABLE IF NOT EXISTS {OCCURRENCES_TABLE} (
  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
  issue_id BIGINT UNSIGNED NOT NULL,
  signature VARCHAR(191) NOT NULL,
  draft_id VARCHAR(64) NOT NULL DEFAULT '',
  planning_run_id VARCHAR(64) NOT NULL DEFAULT '',
  business_name VARCHAR(255) NULL,
  persona VARCHAR(64) NOT NULL DEFAULT '',
  turn_index INT NULL,
  section VARCHAR(32) NOT NULL DEFAULT '',
  stage VARCHAR(64) NOT NULL DEFAULT '',
  severity VARCHAR(12) NOT NULL DEFAULT '',
  observed LONGTEXT NULL,
  expected LONGTEXT NULL,
  evidence_json LONGTEXT NULL,
  source VARCHAR(16) NOT NULL DEFAULT 'cowork',
  created_at DATETIME(6) NOT NULL DEFAULT CURRENT_TIMESTAMP(6),
  KEY ix_occ_issue (issue_id, created_at),
  KEY ix_occ_signature (signature),
  KEY ix_occ_draft (draft_id),
  KEY ix_occ_created (created_at)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""

_CREATE_RESOLUTION_EVENTS_SQL = f"""
CREATE TABLE IF NOT EXISTS {RESOLUTION_EVENTS_TABLE} (
  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
  issue_id BIGINT UNSIGNED NOT NULL,
  signature VARCHAR(191) NOT NULL,
  draft_id VARCHAR(64) NOT NULL DEFAULT '',
  event_type VARCHAR(32) NOT NULL,
  detail_json LONGTEXT NULL,
  created_at DATETIME(6) NOT NULL DEFAULT CURRENT_TIMESTAMP(6),
  KEY ix_res_issue (issue_id, created_at),
  KEY ix_res_type (event_type),
  KEY ix_res_created (created_at)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""

_tables_ready = False


def ensure_tables(conn) -> None:
  global _tables_ready
  if _tables_ready:
    return
  cur = conn.cursor()
  try:
    cur.execute(_CREATE_ISSUES_SQL)
    cur.execute(_CREATE_OCCURRENCES_SQL)
    cur.execute(_CREATE_RESOLUTION_EVENTS_SQL)
    conn.commit()
    _tables_ready = True
  finally:
    cur.close()


def _require(value: str, allowed: tuple, field: str) -> str:
  v = str(value or "").strip()
  if v not in allowed:
    raise ValueError(f"{field} must be one of {allowed}, got {value!r}")
  return v


def _json_or_none(value: Any) -> Optional[str]:
  if value is None:
    return None
  if isinstance(value, str):
    return value
  return json.dumps(value, ensure_ascii=False, default=str)


def _normalize_probe(probe: Any) -> Any:
  """Validate a probe on the way in.

  A probe written as prose ("observe the banner across a price change") is a
  NOTE, not a predicate — 95 of the registry's first 129 probes were prose,
  each one silently discarded at read time and replaced by a derived
  section-only guess. Prose is kept as ``note`` so the intent survives, but it
  is never mistaken for a retest condition. A probe written as an object with
  an UNKNOWN key is a typo, and typos fail loudly here rather than quietly
  widening the probe to "any completed run".
  """
  if probe is None:
    return None
  if isinstance(probe, (str, bytes)):
    text = probe.decode() if isinstance(probe, bytes) else probe
    text = text.strip()
    if not text:
      return None
    try:
      parsed = json.loads(text)
    except Exception:
      return {"note": text}
    return _normalize_probe(parsed) if isinstance(parsed, (dict, list)) else {"note": text}
  if not isinstance(probe, dict):
    raise ValueError(f"probe must be an object or a note string, got {type(probe).__name__}")
  unknown = sorted(set(probe) - PROBE_KEYS)
  if unknown:
    raise ValueError(
      f"unknown probe key(s) {unknown}; allowed: {sorted(PROBE_KEYS)}. "
      "An unrecognized key is silently ignored by the checker, which widens "
      "the probe to 'any completed run' - author it correctly or put it in 'note'."
    )
  specs = probe.get("artifact")
  if specs is not None:
    if isinstance(specs, dict):
      specs = [specs]
    if not isinstance(specs, list) or not specs:
      raise ValueError("probe 'artifact' must be a non-empty object or list of objects")
    for spec in specs:
      if not isinstance(spec, dict):
        raise ValueError(f"artifact assertion must be an object, got {spec!r}")
      kind = str(spec.get("kind") or "").strip()
      if kind not in ARTIFACT_KINDS:
        raise ValueError(f"artifact kind must be one of {ARTIFACT_KINDS}, got {kind!r}")
  return probe


def _insert_resolution_event(
  cur, *, issue_id: int, signature: str, draft_id: str,
  event_type: str, detail: Any = None,
) -> None:
  cur.execute(
    f"""
    INSERT INTO {RESOLUTION_EVENTS_TABLE}
      (issue_id, signature, draft_id, event_type, detail_json)
    VALUES (%s, %s, %s, %s, %s)
    """,
    (issue_id, signature, draft_id, event_type, _json_or_none(detail)),
  )


def _fetch_issue(cur, signature: str) -> Optional[Dict[str, Any]]:
  cur.execute(
    f"SELECT * FROM {ISSUES_TABLE} WHERE signature = %s", (signature,)
  )
  row = cur.fetchone()
  if row is None:
    return None
  cols = [d[0] for d in cur.description]
  return dict(zip(cols, row))


# ----------------------------------------------------------------------------
# Write path (Cowork's contract; also used by the HTTP endpoint).
# ----------------------------------------------------------------------------

def report_issue(
  conn,
  *,
  signature: str,
  category: str,
  severity: str,
  observed: str,
  expected: str,
  draft_id: str = "",
  planning_run_id: str = "",
  business_name: str = "",
  persona: str = "",
  turn_index: Optional[int] = None,
  section: str = "",
  stage: str = "",
  title: str = "",
  resolution_class: str = "",
  probe: Any = None,
  evidence: Any = None,
  source: str = "cowork",
) -> Dict[str, Any]:
  """Report one sighting of an issue. Upserts the registry row keyed by
  ``signature`` and INSERTs the occurrence. Returns the resulting registry
  state plus flags (is_new, reopened).

  Vocabulary violations raise — the write contract fails loudly.
  """
  signature = str(signature or "").strip()
  if not signature or len(signature) > 191:
    raise ValueError("signature is required (non-empty, <=191 chars)")
  category = _require(category, CATEGORIES, "category")
  severity = _require(severity, SEVERITIES, "severity")
  rclass = str(resolution_class or "").strip() or DEFAULT_RESOLUTION_CLASS[category]
  rclass = _require(rclass, RESOLUTION_CLASSES, "resolution_class")
  observed = str(observed or "").strip()
  expected = str(expected or "").strip()
  if not observed or not expected:
    raise ValueError("observed and expected are both required (what the app "
                     "did vs what should have happened)")
  probe = _normalize_probe(probe)

  ensure_tables(conn)
  cur = conn.cursor()
  try:
    existing = _fetch_issue(cur, signature)
    is_new = existing is None
    reopened = False
    if is_new:
      cur.execute(
        f"""
        INSERT INTO {ISSUES_TABLE}
          (signature, category, resolution_class, severity, title, status,
           first_seen_at, last_seen_at, occurrence_count, probe_json)
        VALUES (%s, %s, %s, %s, %s, 'open', NOW(6), NOW(6), 1, %s)
        """,
        (signature, category, rclass, severity,
         str(title or "")[:255], _json_or_none(probe)),
      )
      issue_id = cur.lastrowid
    else:
      issue_id = int(existing["issue_id"])
      reopened = str(existing["status"]) == "resolved"
      next_status = "recurring" if reopened else str(existing["status"])
      # A sighting resets BOTH resolution counters and (on a resolved issue)
      # clears the resolution verdict — the audit rows keep the history.
      cur.execute(
        f"""
        UPDATE {ISSUES_TABLE}
        SET status = %s,
            last_seen_at = NOW(6),
            occurrence_count = occurrence_count + 1,
            reopened_count = reopened_count + %s,
            clean_exercise_count = 0,
            runs_since_last_seen = 0,
            severity = %s,
            resolved_detected_at = IF(%s, NULL, resolved_detected_at),
            resolution_basis = IF(%s, NULL, resolution_basis),
            resolution_confidence = IF(%s, NULL, resolution_confidence),
            probe_json = COALESCE(%s, probe_json)
        WHERE issue_id = %s
        """,
        (next_status, 1 if reopened else 0, severity,
         reopened, reopened, reopened, _json_or_none(probe), issue_id),
      )
      if reopened:
        _insert_resolution_event(
          cur, issue_id=issue_id, signature=signature, draft_id=draft_id,
          event_type="reopened",
          detail={"previous_basis": existing.get("resolution_basis"),
                  "previous_confidence": existing.get("resolution_confidence")},
        )

    cur.execute(
      f"""
      INSERT INTO {OCCURRENCES_TABLE}
        (issue_id, signature, draft_id, planning_run_id, business_name,
         persona, turn_index, section, stage, severity, observed, expected,
         evidence_json, source)
      VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
      """,
      (issue_id, signature, str(draft_id or ""), str(planning_run_id or ""),
       str(business_name or "")[:255] or None, str(persona or "")[:64],
       int(turn_index) if turn_index is not None else None,
       str(section or "")[:32], str(stage or "")[:64], severity,
       observed, expected, _json_or_none(evidence), str(source or "cowork")[:16]),
    )
    conn.commit()
  finally:
    cur.close()
  state = get_issue(conn, signature=signature)
  state["is_new"] = is_new
  state["reopened"] = reopened
  return state


def get_issue(conn, *, signature: str) -> Dict[str, Any]:
  ensure_tables(conn)
  cur = conn.cursor()
  try:
    row = _fetch_issue(cur, signature)
  finally:
    cur.close()
  if row is None:
    raise KeyError(f"no issue with signature {signature!r}")
  return row


def set_probe(conn, *, signature: str, probe: Any, note: str = "") -> Dict[str, Any]:
  """Attach/replace an issue's retest condition, audited.

  Probes are registry STATE, so upgrading one (typically: giving an issue a
  real ``artifact`` assertion in place of an opportunity-only guess) goes
  through here and leaves an INSERT-only audit row, rather than a hand-edited
  UPDATE. Validation is the same fail-loud path the write contract uses.
  Attaching a probe does NOT change status or counters; the next evaluation
  re-judges the issue on the new condition.
  """
  ensure_tables(conn)
  probe = _normalize_probe(probe)
  cur = conn.cursor()
  try:
    issue = _fetch_issue(cur, signature)
    if issue is None:
      raise KeyError(f"no issue with signature {signature!r}")
    cur.execute(
      f"UPDATE {ISSUES_TABLE} SET probe_json = %s WHERE issue_id = %s",
      (_json_or_none(probe), int(issue["issue_id"])),
    )
    _insert_resolution_event(
      cur, issue_id=int(issue["issue_id"]), signature=signature, draft_id="",
      event_type="probe_updated",
      detail={"previous": issue.get("probe_json"),
              "next": _json_or_none(probe), "note": note},
    )
    conn.commit()
  finally:
    cur.close()
  return get_issue(conn, signature=signature)


def resolve_manually(conn, *, signature: str, note: str = "") -> Dict[str, Any]:
  """Human override: mark resolved with basis 'manual'. Confidence follows
  the class honesty rule (hard -> confirmed, soft -> observational)."""
  ensure_tables(conn)
  cur = conn.cursor()
  try:
    issue = _fetch_issue(cur, signature)
    if issue is None:
      raise KeyError(f"no issue with signature {signature!r}")
    confidence = "confirmed" if issue["resolution_class"] == "hard" else "observational"
    cur.execute(
      f"""
      UPDATE {ISSUES_TABLE}
      SET status='resolved', resolved_detected_at=NOW(6),
          resolution_basis='manual', resolution_confidence=%s
      WHERE issue_id=%s
      """,
      (confidence, issue["issue_id"]),
    )
    _insert_resolution_event(
      cur, issue_id=int(issue["issue_id"]), signature=signature, draft_id="",
      event_type="manual_resolve", detail={"note": note},
    )
    conn.commit()
  finally:
    cur.close()
  return get_issue(conn, signature=signature)


# ----------------------------------------------------------------------------
# Resolution sensing (run after each completed persona run).
# ----------------------------------------------------------------------------

def _probe_from_issue(cur, issue: Dict[str, Any]) -> Dict[str, Any]:
  """The issue's probe predicate; when absent, derive a minimal one from its
  FIRST occurrence (section; business for verdict issues) so older issues
  still get honest exercised-detection instead of none."""
  raw = issue.get("probe_json")
  if raw:
    try:
      parsed = json.loads(raw) if isinstance(raw, (str, bytes)) else raw
      if isinstance(parsed, dict) and parsed:
        return parsed
    except Exception:
      pass
  cur.execute(
    f"""
    SELECT section, business_name FROM {OCCURRENCES_TABLE}
    WHERE issue_id = %s ORDER BY id ASC LIMIT 1
    """,
    (issue["issue_id"],),
  )
  row = cur.fetchone()
  derived: Dict[str, Any] = {}
  if row:
    section, business = row[0], row[1]
    if section:
      derived["section"] = str(section)
    if str(issue.get("category")) == "verdict" and business:
      derived["business_like"] = f"%{str(business)[:80]}%"
  return derived


def _run_exercised(cur, draft_id: str, probe: Dict[str, Any]) -> Dict[str, Any]:
  """Did this run exercise the issue's path? Evaluates the small documented
  predicate vocabulary against the run's vitals + planning rows. Every
  clause must pass. Empty probe -> not exercised (never guess)."""
  if not probe:
    return {"exercised": False, "reason": "no probe and none derivable"}
  # manual_only: the issue's retest condition cannot be expressed in the
  # vitals vocabulary (e.g. "anchor computed on stated revenue, not the
  # capacity model" — issue #10). Auto-sensing must never tick it: a
  # section-level coincidence confirming a real bug is exactly what the
  # schema exists to NOT trust. Resolution comes from resolve_manually or
  # a future probe upgrade.
  if probe.get("manual_only"):
    return {"exercised": False, "reason": "manual-retest-only (vitals cannot sense this condition)"}
  # require_completed is a GUARD, not a retest condition, and metadata is not a
  # predicate at all. A probe carrying only those says "any finished run
  # retests this", which is how nine regression_pin-only issues were resolving
  # on runs that never went near them. No condition => nothing to tick.
  conditions = (set(probe) & (PROBE_OPPORTUNITY_KEYS | PROBE_ARTIFACT_KEYS)) - {"require_completed"}
  if not conditions:
    return {"exercised": False,
            "reason": "probe states no retest condition (metadata/notes only)"}
  checks: List[str] = []

  if probe.get("require_completed", True):
    cur.execute(
      """
      SELECT run_status FROM planning_runs
      WHERE draft_id = %s ORDER BY started_at DESC LIMIT 1
      """,
      (draft_id,),
    )
    row = cur.fetchone()
    status = str(row[0]) if row else ""
    if status != "completed":
      return {"exercised": False,
              "reason": f"run not completed (status={status or 'none'})"}
    checks.append("run_completed")

  sections = probe.get("sections") or (
    [probe["section"]] if probe.get("section") else []
  )
  if sections:
    placeholders = ", ".join(["%s"] * len(sections))
    cur.execute(
      f"""
      SELECT COUNT(*) FROM run_vitals_turns
      WHERE draft_id = %s AND (section IN ({placeholders})
                               OR section_after IN ({placeholders}))
      """,
      (draft_id, *sections, *sections),
    )
    if int(cur.fetchone()[0] or 0) == 0:
      return {"exercised": False, "reason": f"sections {sections} not visited"}
    checks.append(f"sections={sections}")

  if probe.get("call_label_like"):
    cur.execute(
      """
      SELECT COUNT(*) FROM run_vitals_gpt_calls
      WHERE draft_id = %s AND call_label LIKE %s
      """,
      (draft_id, str(probe["call_label_like"])),
    )
    if int(cur.fetchone()[0] or 0) == 0:
      return {"exercised": False,
              "reason": f"no gpt call matching {probe['call_label_like']!r}"}
    checks.append("call_label")

  if probe.get("stage_like"):
    cur.execute(
      """
      SELECT COUNT(*) FROM planning_stage_events
      WHERE draft_id = %s AND stage LIKE %s
      """,
      (draft_id, str(probe["stage_like"])),
    )
    if int(cur.fetchone()[0] or 0) == 0:
      return {"exercised": False,
              "reason": f"stage {probe['stage_like']!r} not reached"}
    checks.append("stage")

  if probe.get("business_like"):
    cur.execute(
      "SELECT business_name FROM intake_consult_drafts WHERE draft_id = %s",
      (draft_id,),
    )
    row = cur.fetchone()
    name = str(row[0] or "") if row else ""
    like = str(probe["business_like"]).replace("%", "").lower()
    if like and like not in name.lower():
      return {"exercised": False,
              "reason": f"business {name!r} does not match {probe['business_like']!r}"}
    checks.append("business")

  if probe.get("min_turns"):
    cur.execute(
      "SELECT COUNT(*) FROM run_vitals_turns WHERE draft_id = %s",
      (draft_id,),
    )
    if int(cur.fetchone()[0] or 0) < int(probe["min_turns"]):
      return {"exercised": False, "reason": "fewer turns than min_turns"}
    checks.append("min_turns")

  return {"exercised": True, "reason": "; ".join(checks) or "probe empty-pass"}


def _auto_recurrence(cur, draft_id: str, probe: Dict[str, Any]) -> Optional[str]:
  """Hard signals the checker can detect WITHOUT a Cowork report, when the
  probe opts in: stalls, holds, matching GPT errors on this run."""
  auto = probe.get("auto_recur") or {}
  if not isinstance(auto, dict) or not auto:
    return None
  if auto.get("stall"):
    cur.execute(
      """
      SELECT COUNT(*) FROM run_vitals_events
      WHERE draft_id = %s AND event_type LIKE 'watch_end_stall%%'
      """,
      (draft_id,),
    )
    if int(cur.fetchone()[0] or 0) > 0:
      return "stall event on this run"
  if auto.get("hold"):
    cur.execute(
      """
      SELECT COUNT(*) FROM run_vitals_events
      WHERE draft_id = %s AND event_type = 'turn_hold'
      """,
      (draft_id,),
    )
    if int(cur.fetchone()[0] or 0) > 0:
      return "turn_hold event on this run"
  if auto.get("gpt_error_like"):
    cur.execute(
      """
      SELECT COUNT(*) FROM run_vitals_gpt_calls
      WHERE draft_id = %s AND error LIKE %s
      """,
      (draft_id, str(auto["gpt_error_like"])),
    )
    if int(cur.fetchone()[0] or 0) > 0:
      return f"gpt error matching {auto['gpt_error_like']!r}"
  return None


def _load_ops_model(cur, draft_id: str) -> Optional[Dict[str, Any]]:
  cur.execute(
    "SELECT operating_model_json FROM intake_consult_drafts WHERE draft_id = %s",
    (draft_id,),
  )
  row = cur.fetchone()
  if not row or not row[0]:
    return None
  try:
    parsed = json.loads(row[0]) if isinstance(row[0], (str, bytes)) else row[0]
  except Exception:
    return None
  return parsed if isinstance(parsed, dict) else None


def _ops_product_rows(ops: Dict[str, Any]) -> List[Dict[str, Any]]:
  rows: List[Dict[str, Any]] = []
  for lob in (ops or {}).get("lob_models") or []:
    if not isinstance(lob, dict):
      continue
    for product in lob.get("products") or []:
      if isinstance(product, dict):
        rows.append(product)
  return rows


def _assert_ops_per_line_cogs(cur, draft_id: str, spec: Dict[str, Any]) -> Dict[str, Any]:
  """THE artifact for the per-line COGS class: every product row on a
  multi-line business must carry a non-null cogs_percent_of_line_revenue.
  A proposal in chat is not this; only the written ops row is."""
  if "require_distinct_rates" in spec:
    raise ValueError(
      "require_distinct_rates is retired: distinct rates are now the DEFAULT "
      "for ops_per_line_cogs. Use allow_shared_rates=true to opt out, so an "
      "opt-out is always a stated decision rather than an absent flag."
    )
  min_lines = int(spec.get("min_lines") or 2)
  ops = _load_ops_model(cur, draft_id)
  if ops is None:
    return {"verdict": "not_applicable", "detail": "no operating_model_json on this draft"}
  products = _ops_product_rows(ops)
  if len(products) < min_lines:
    return {"verdict": "not_applicable",
            "detail": f"{len(products)} product row(s) < min_lines={min_lines}"}
  written = [p for p in products
             if p.get("cogs_percent_of_line_revenue") is not None]
  if len(written) < len(products):
    missing = [str(p.get("product_name") or p.get("name") or "?")
               for p in products if p.get("cogs_percent_of_line_revenue") is None]
    return {"verdict": "fail",
            "detail": (f"{len(written)}/{len(products)} product rows carry "
                       f"cogs_percent_of_line_revenue; null on {missing}")}
  detail = f"all {len(products)} product rows carry cogs_percent_of_line_revenue"
  # DISTINCT RATES ARE THE DEFAULT for this class. N rows all carrying the
  # same rate is a blend wearing per-line clothing: it satisfies the field
  # check while the model is exactly as wrong as before. The implemented
  # semantics are "at least TWO distinct rates", not "all pairwise distinct",
  # so a client-declared collapse (plants and hard goods share a structure,
  # install and design do not) still passes. The one shape this would
  # false-fail is a client declaring that ALL lines share one structure, which
  # no artifact can tell apart from the bug — that case must opt out
  # EXPLICITLY, from the recorded grouping, never by the check being absent.
  if not spec.get("allow_shared_rates", False):
    rates = {round(float(p["cogs_percent_of_line_revenue"]), 4) for p in products}
    if len(rates) < 2:
      return {"verdict": "fail",
              "detail": f"all {len(products)} rows share one rate {rates} - "
                        "a blend wearing per-line clothing"}
    detail += f"; {len(rates)} distinct rate(s)"
  return {"verdict": "pass", "detail": detail}


def _assert_ops_cogs_shared_group(cur, draft_id: str, spec: Dict[str, Any]) -> Dict[str, Any]:
  """A STORED cost-structure collapse must be coherent (#142).

  The client declaring "plants and hard goods share one cost structure" now
  writes cogs_cost_structure_group on the rows, so it is finally assertable:
  every member of a stored group carries the SAME rate, and a line outside it
  differs. Before tier 2 there was no field to read and this issue was
  correctly capped at observational.

  WHAT IT STILL CANNOT SEE, stated rather than papered over: whether the
  client asked for a collapse that was never stored. No artifact can - the
  absence of a group is indistinguishable from a client who never asked. That
  half is covered by Test Files/_live_cw031_cogs_door_turn.py, which says the
  sentence to the live router and reads the rows back.
  """
  ops = _load_ops_model(cur, draft_id)
  if ops is None:
    return {"verdict": "not_applicable", "detail": "no operating_model_json on this draft"}
  products = _ops_product_rows(ops)
  groups: Dict[str, List[Dict[str, Any]]] = {}
  for product in products:
    label = str(product.get("cogs_cost_structure_group") or "").strip()
    if label:
      groups.setdefault(label, []).append(product)
  if not groups:
    return {"verdict": "not_applicable",
            "detail": "no cost-structure collapse is stored on this draft"}
  details = []
  for label, members in groups.items():
    if len(members) < int(spec.get("min_group") or 2):
      return {"verdict": "fail",
              "detail": f"group {label!r} has {len(members)} member(s), expected >= 2"}
    rates = {round(float(m["cogs_percent_of_line_revenue"]), 4) for m in members
             if m.get("cogs_percent_of_line_revenue") is not None}
    if len(rates) > 1:
      return {"verdict": "fail",
              "detail": f"group {label!r} members carry {len(rates)} different rates {rates}"}
    outside = [p for p in products if p not in members
               and p.get("cogs_percent_of_line_revenue") is not None]
    if rates and outside and all(
      round(float(p["cogs_percent_of_line_revenue"]), 4) in rates for p in outside
    ):
      return {"verdict": "fail",
              "detail": (f"group {label!r} shares its rate with every line outside it - "
                         "a collapse that collapsed everything")}
    details.append(f"{label} ({len(members)} lines, rate {rates or 'unset'})")
  return {"verdict": "pass", "detail": "; ".join(details)}


def _assert_ops_field_non_null(cur, draft_id: str, spec: Dict[str, Any]) -> Dict[str, Any]:
  """Generic: a dotted path into operating_model_json must be non-null.
  ``products[]`` walks every product row and requires ALL of them."""
  path = str(spec.get("path") or "").strip()
  if not path:
    raise ValueError("artifact ops_field_non_null requires 'path'")
  ops = _load_ops_model(cur, draft_id)
  if ops is None:
    return {"verdict": "not_applicable", "detail": "no operating_model_json on this draft"}
  if path.startswith("products[]."):
    field = path.split(".", 1)[1]
    products = _ops_product_rows(ops)
    if not products:
      return {"verdict": "not_applicable", "detail": "no product rows"}
    missing = [str(p.get("product_name") or "?") for p in products
               if p.get(field) is None]
    if missing:
      return {"verdict": "fail", "detail": f"{field} null on {missing}"}
    return {"verdict": "pass", "detail": f"{field} non-null on all {len(products)} rows"}
  node: Any = ops
  for part in path.split("."):
    if not isinstance(node, dict) or part not in node:
      return {"verdict": "fail", "detail": f"path {path!r} absent at {part!r}"}
    node = node[part]
  if node is None:
    return {"verdict": "fail", "detail": f"path {path!r} is null"}
  return {"verdict": "pass", "detail": f"path {path!r} non-null"}


_SUM_FORMULA_RE = re.compile(
  r"^=\s*SUM\(\s*\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)\s*\)$",
  re.IGNORECASE,
)

# Reconciliation tolerance: the three routes to a quarter's COGS are computed
# from rates rounded at different places (the stub column stores 0.5405 where
# Q1 stores 0.540515), so they agree to a fraction of a percent, not to the
# cent. 0.5% with a $1 floor is far tighter than any real defect.
_RECONCILE_REL_TOL = 0.005
_RECONCILE_ABS_TOL = 1.0


def _numeric(value: Any) -> Optional[float]:
  if isinstance(value, bool) or value is None:
    return None
  if isinstance(value, (int, float)):
    return float(value)
  return None


_PERIOD_HEADER_RE = re.compile(r"^(stub|q\d+)$", re.IGNORECASE)


def _period_columns(ws) -> Dict[str, int]:
  """{'Q1': 4, ...} for the PERIOD columns only, by header label.

  The sheets carry annual roll-up columns (Y1..Y5) after the quarters, and
  those sum HORIZONTALLY across their own row (``=SUM(D11:G11)``). Treating
  one as a quarter turns a correct workbook into a failure — and multiplying
  a year's summed capacity by a year's summed unit price into nonsense — so
  every column-wise check below runs over these and nothing else.
  """
  for header_row in range(1, min(ws.max_row, 8) + 1):
    found = {}
    for col in range(1, ws.max_column + 1):
      value = ws.cell(row=header_row, column=col).value
      if value is not None and _PERIOD_HEADER_RE.match(str(value).strip()):
        found[str(value).strip()] = col
    if found:
      return found
  return {}


def _sheet_rows_by_label(ws) -> List[Any]:
  """[(row_index, column_A_text)] for every row carrying a label."""
  out = []
  for (cell,) in ws.iter_rows(min_col=1, max_col=1):
    if cell.value is not None and str(cell.value).strip():
      out.append((cell.row, str(cell.value).strip()))
  return out


def _assert_total_sums_over_lines(ws, per_line_rows, total_row, label):
  """Law bullet 2: the total row must be =SUM over EXACTLY the per-line rows.

  N per-line rows above a total that sums the wrong range is the same wrong
  number with better manners, so the range is checked span-for-span in every
  quarter column that carries the formula.
  """
  first, last = min(per_line_rows), max(per_line_rows)
  if last - first + 1 != len(per_line_rows):
    return (f"the {len(per_line_rows)} per-line {label!r} rows are not contiguous "
            f"(rows {sorted(per_line_rows)})")
  if first <= total_row <= last:
    return f"the total row {total_row} sits inside the per-line block {first}-{last}"
  checked = 0
  periods = _period_columns(ws)
  if not periods:
    return f"no period columns found on the sheet, cannot check the {label!r} total"
  for col in sorted(periods.values()):
    raw = ws.cell(row=total_row, column=col).value
    if not isinstance(raw, str) or not raw.strip().startswith("="):
      continue
    match = _SUM_FORMULA_RE.match(raw.strip())
    letter = ws.cell(row=total_row, column=col).column_letter
    if not match:
      return (f"total row {total_row} column {letter} is {raw.strip()[:60]!r}, "
              "not a SUM over the per-line rows")
    lcol, lrow, rcol, rrow = match.group(1).upper(), int(match.group(2)), \
        match.group(3).upper(), int(match.group(4))
    if lcol != letter or rcol != letter or lrow != first or rrow != last:
      return (f"total row {total_row} column {letter} sums {raw.strip()} but the "
              f"per-line rows are {letter}{first}:{letter}{last}")
    checked += 1
  if not checked:
    return f"total row {total_row} carries no SUM formula in any quarter column"
  return None


def _reconcile_workbook_cogs(wb, label):
  """Law bullet 3: Sigma(line revenue x line pct) == blend == finmo COGS, per quarter.

  All three routes are readable as LITERALS: Revenue Drivers stores capacity,
  unit price, utilization and COGS % per line; Model Inputs stores the blended
  COGS rate; Audit Source stores the engine's own persisted COGS. (The formula
  cells themselves cannot be used — openpyxl writes the workbook and nothing
  recalculates it in place, so every formula cell's cached value is None.)

  Returns (verdict, detail).
  """
  for needed in ("Revenue Drivers", "Model Inputs", "Audit Source"):
    if needed not in wb.sheetnames:
      return ("not_applicable", f"sheet {needed!r} absent, cannot reconcile")
  rd, mi, aud = wb["Revenue Drivers"], wb["Model Inputs"], wb["Audit Source"]

  lines: Dict[str, Dict[str, int]] = {}
  for row, text in _sheet_rows_by_label(rd):
    if " - " not in text:
      continue
    name, _, field = text.rpartition(" - ")
    key = field.strip().lower()
    if key in ("capacity", "unit price", "utilization", "cogs %"):
      lines.setdefault(name.strip(), {})[key] = row
  usable = {n: f for n, f in lines.items()
            if {"capacity", "unit price", "utilization", "cogs %"} <= set(f)}
  if len(usable) < 2:
    return ("not_applicable",
            f"Revenue Drivers carries {len(usable)} line(s) with a per-line COGS rate")

  blend_row = next((r for r, t in _sheet_rows_by_label(mi) if t == label), None)
  audit_row = next((r for r, t in _sheet_rows_by_label(aud) if t == label), None)
  if blend_row is None or audit_row is None:
    return ("not_applicable",
            f"{label!r} absent on Model Inputs or Audit Source, cannot reconcile")

  # Columns are matched BY PERIOD LABEL across the three sheets, not by index:
  # each sheet lays its own rows out independently and only the header label
  # says which quarter a column is.
  rd_periods, mi_periods, aud_periods = (
    _period_columns(rd), _period_columns(mi), _period_columns(aud))
  shared = [p for p in rd_periods if p in mi_periods and p in aud_periods]
  if not shared:
    return ("not_applicable", "the three sheets share no period column")

  compared, failures = 0, []
  for period in shared:
    col = rd_periods[period]
    sigma, total_rev, complete = 0.0, 0.0, True
    for fields in usable.values():
      cap = _numeric(rd.cell(row=fields["capacity"], column=col).value)
      price = _numeric(rd.cell(row=fields["unit price"], column=col).value)
      util = _numeric(rd.cell(row=fields["utilization"], column=col).value)
      pct = _numeric(rd.cell(row=fields["cogs %"], column=col).value)
      if None in (cap, price, util, pct):
        complete = False
        break
      line_rev = cap * price * util
      total_rev += line_rev
      sigma += line_rev * pct
    if not complete or total_rev <= 0:
      continue
    blend_pct = _numeric(mi.cell(row=blend_row, column=mi_periods[period]).value)
    engine = _numeric(aud.cell(row=audit_row, column=aud_periods[period]).value)
    if blend_pct is None or engine is None:
      continue
    blend = blend_pct * total_rev
    compared += 1
    for other_name, other in (("blend", blend), ("finmo COGS", engine)):
      tol = max(_RECONCILE_ABS_TOL, abs(sigma) * _RECONCILE_REL_TOL)
      if abs(sigma - other) > tol:
        failures.append(f"{period}: Sigma(line rev x line pct)={sigma:,.2f} vs "
                        f"{other_name}={other:,.2f}")
  if not compared:
    return ("not_applicable", "no quarter column carried all three COGS routes")
  if failures:
    return ("fail", "; ".join(failures[:3]) + f" ({len(failures)} mismatch(es))")
  return ("pass", f"Sigma(line rev x line pct) == blend == finmo COGS on "
                  f"{compared} quarter column(s)")


def _assert_workbook_cogs_rows(cur, draft_id: str, spec: Dict[str, Any]) -> Dict[str, Any]:
  """The DELIVERED workbook must carry one COGS row per line, a total that
  sums over EXACTLY those rows, and three routes to the quarter's COGS that
  agree — Nick's verification law for this batch, bullets 2 and 3.

  Scoped to the P&L sheet on purpose: the same label legitimately appears on
  Model Inputs (the driver row) and Audit Source (persisted values, no
  formulas), and counting those inflates the count into a false pass.

  The workbook is resolved by BINDING, never by newest-mtime: see
  workbook_delivery_record. An unattributable file is not evidence about this
  draft, so it returns not_applicable rather than judging someone else's run.
  """
  import os
  from client_intake_and_finmo.workbook_delivery_record import (  # type: ignore
    resolve_workbook_for_draft,
  )
  sheet_name = str(spec.get("sheet") or "FINMO")
  label = str(spec.get("label_prefix") or "Cost of Goods Sold")
  min_rows = int(spec.get("min_rows") or 2)
  delivery_dir = (os.getenv("FINMO_MODEL_DELIVERY_DIR") or "").strip()
  resolved = resolve_workbook_for_draft(cur, draft_id, delivery_dir=delivery_dir)
  path = resolved.get("path")
  if not path:
    return {"verdict": "not_applicable", "detail": str(resolved.get("detail") or "")}
  try:
    import openpyxl
    wb = openpyxl.load_workbook(path)
  except Exception as exc:
    return {"verdict": "not_applicable", "detail": f"workbook unreadable: {exc}"}
  try:
    if sheet_name not in wb.sheetnames:
      return {"verdict": "not_applicable", "detail": f"sheet {sheet_name!r} absent"}
    ws = wb[sheet_name]
    labelled = [(r, t) for r, t in _sheet_rows_by_label(ws) if t.startswith(label)]
    per_line = [r for r, t in labelled if t != label]
    totals = [r for r, t in labelled if t == label]
    where = f"{os.path.basename(path)} [{sheet_name}]"
    if len(per_line) < min_rows:
      return {"verdict": "fail",
              "detail": (f"{where} carries {len(per_line)} per-line {label!r} row(s) "
                         f"({len(labelled)} labelled total), expected >= {min_rows}")}
    if spec.get("check_total_sum", True):
      if len(totals) != 1:
        return {"verdict": "fail",
                "detail": (f"{where} carries {len(totals)} total {label!r} rows "
                           f"above {len(per_line)} per-line rows, expected exactly 1")}
      problem = _assert_total_sums_over_lines(ws, per_line, totals[0], label)
      if problem:
        return {"verdict": "fail", "detail": f"{where}: {problem}"}
    detail = f"{where} carries {len(per_line)} per-line {label!r} rows"
    if spec.get("check_total_sum", True):
      detail += f" totalled by =SUM over exactly those rows"
    if spec.get("check_reconciliation", True):
      verdict, recon_detail = _reconcile_workbook_cogs(wb, label)
      if verdict != "pass":
        return {"verdict": verdict, "detail": f"{where}: {recon_detail}"}
      detail += f"; {recon_detail}"
    return {"verdict": "pass", "detail": f"{detail} [{resolved.get('basis')}]"}
  finally:
    try:
      wb.close()
    except Exception:
      pass


_ARTIFACT_DISPATCH = {
  "ops_per_line_cogs": _assert_ops_per_line_cogs,
  "ops_field_non_null": _assert_ops_field_non_null,
  "workbook_cogs_rows": _assert_workbook_cogs_rows,
  "ops_cogs_shared_group": _assert_ops_cogs_shared_group,
}


def _assert_artifacts(cur, draft_id: str, probe: Dict[str, Any]) -> Dict[str, Any]:
  """Read the run's PERSISTED artifacts and judge them.

  Returns present/verdict/details. ``present=False`` means the issue carries
  no artifact assertion at all — the caller must then cap the resolution at
  'observational'. A single failing assertion fails the whole set (a defect
  anywhere in the artifact is the defect); a single not_applicable with no
  failure makes the whole set not_applicable (we did not get to look).
  """
  specs = probe.get("artifact") if isinstance(probe, dict) else None
  if isinstance(specs, dict):
    specs = [specs]
  if not isinstance(specs, list) or not specs:
    return {"present": False, "verdict": "absent", "details": []}
  details: List[str] = []
  verdict = "pass"
  for spec in specs:
    if not isinstance(spec, dict):
      raise ValueError(f"artifact assertion must be an object, got {spec!r}")
    kind = str(spec.get("kind") or "").strip()
    if kind not in _ARTIFACT_DISPATCH:
      raise ValueError(f"artifact kind must be one of {ARTIFACT_KINDS}, got {kind!r}")
    outcome = _ARTIFACT_DISPATCH[kind](cur, draft_id, spec)
    details.append(f"{kind}: {outcome['verdict']} - {outcome['detail']}")
    if outcome["verdict"] == "fail":
      verdict = "fail"
    elif outcome["verdict"] == "not_applicable" and verdict != "fail":
      verdict = "not_applicable"
  return {"present": True, "verdict": verdict, "details": details}


def evaluate_run_for_resolution(
  conn,
  *,
  draft_id: str,
  hard_clean_threshold: int = DEFAULT_HARD_CLEAN_THRESHOLD,
  soft_runs_threshold: int = DEFAULT_SOFT_RUNS_THRESHOLD,
) -> Dict[str, Any]:
  """Evaluate every open/recurring issue against one finished run.

  Per issue: if it was reported again on THIS run (or an opted-in hard
  signal auto-recurred), counters were/are reset. Otherwise, if the run
  exercised the issue's path, tick the clean counters and resolve when the
  class-appropriate threshold is met — 'confirmed' only ever for hard
  issues, 'observational' for soft. Not-exercised runs change nothing
  (absence without opportunity is not evidence).
  """
  draft_id = str(draft_id or "").strip()
  if not draft_id:
    raise ValueError("draft_id is required")
  ensure_tables(conn)
  summary = {
    "draft_id": draft_id, "evaluated": 0, "recurred": 0,
    "exercised_clean": 0, "not_exercised": 0,
    "artifact_verified": 0, "artifact_failed": 0,
    "resolved_confirmed": [], "resolved_observational": [],
  }
  cur = conn.cursor()
  try:
    # 'resolved' is included on purpose: an issue whose probe can READ an
    # artifact stays under audit forever, so a verdict reached in error (or a
    # genuine regression after a real fix) is caught on the next run instead
    # of being frozen by its own resolution.
    cur.execute(
      f"SELECT * FROM {ISSUES_TABLE} WHERE status IN ('open', 'recurring', 'resolved')"
    )
    cols = [d[0] for d in cur.description]
    issues = [dict(zip(cols, r)) for r in cur.fetchall()]
    for issue in issues:
      summary["evaluated"] += 1
      issue_id = int(issue["issue_id"])
      signature = str(issue["signature"])

      cur.execute(
        f"""
        SELECT COUNT(*) FROM {OCCURRENCES_TABLE}
        WHERE issue_id = %s AND draft_id = %s
        """,
        (issue_id, draft_id),
      )
      reported_this_run = int(cur.fetchone()[0] or 0) > 0

      probe = _probe_from_issue(cur, issue)
      auto_reason = None if reported_this_run else _auto_recurrence(cur, draft_id, probe)

      if reported_this_run or auto_reason:
        summary["recurred"] += 1
        if auto_reason:
          # Auto-detected recurrence: record the occurrence + reset, exactly
          # as a reported sighting would (source marks it machine-found).
          report_issue(
            conn,
            signature=signature,
            category=str(issue["category"]),
            severity=str(issue["severity"]),
            observed=f"auto-detected recurrence: {auto_reason}",
            expected="no recurrence of this signature on an exercised run",
            draft_id=draft_id,
            source="auto_check",
          )
        continue

      verdict = _run_exercised(cur, draft_id, probe)
      if not verdict["exercised"]:
        summary["not_exercised"] += 1
        continue

      # The artifact is the authority. Reporter silence on an exercised run is
      # only the ABSENCE of a complaint; a read artifact is EVIDENCE.
      artifact = _assert_artifacts(cur, draft_id, probe)

      if str(issue["status"]) == "resolved":
        # A RESOLVED issue is re-audited, never re-resolved. Resolution used
        # to be terminal, so a verdict reached on weak evidence could never
        # be revisited and #138 stayed 'confirmed' on the very run that
        # disproves it. Now: if the artifact can be read and it is wrong, the
        # issue reopens on its own evidence. If there is no artifact to read,
        # nothing happens -- re-auditing on silence is what got us here.
        if artifact["verdict"] == "fail":
          summary["recurred"] += 1
          summary["artifact_failed"] += 1
          report_issue(
            conn,
            signature=signature,
            category=str(issue["category"]),
            severity=str(issue["severity"]),
            observed=("resolution contradicted by the artifact: "
                      + "; ".join(artifact["details"])),
            expected="a resolved issue's artifact still carries the fix",
            draft_id=draft_id,
            source="artifact_check",
          )
        elif artifact["verdict"] == "pass":
          summary["artifact_verified"] += 1
          _insert_resolution_event(
            cur, issue_id=issue_id, signature=signature, draft_id=draft_id,
            event_type="resolution_reaudited",
            detail={"artifact_verified": artifact["details"]},
          )
        continue
      if artifact["verdict"] == "fail":
        # The registry catches its own issue red-handed: the run walked the
        # path and the persisted artifact is still wrong. That is a
        # recurrence, whatever the reporter did or did not say.
        summary["recurred"] += 1
        summary["artifact_failed"] += 1
        report_issue(
          conn,
          signature=signature,
          category=str(issue["category"]),
          severity=str(issue["severity"]),
          observed="artifact assertion failed: " + "; ".join(artifact["details"]),
          expected="the persisted artifact carries the fix, not just a proposal of it",
          draft_id=draft_id,
          source="artifact_check",
        )
        continue
      if artifact["verdict"] == "not_applicable":
        # We never got to look. Absence of opportunity is not evidence.
        summary["not_exercised"] += 1
        continue

      summary["exercised_clean"] += 1
      if artifact["present"]:
        summary["artifact_verified"] += 1
      clean = int(issue["clean_exercise_count"] or 0) + 1
      quiet = int(issue["runs_since_last_seen"] or 0) + 1
      rclass = str(issue["resolution_class"])
      # 'confirmed' is reserved for a READ artifact that HELD. A hard issue
      # with no artifact assertion has only opportunity + silence behind it,
      # so it is treated exactly like a soft one: observational, and only
      # after the soft threshold of quiet exercised runs.
      artifact_backed = rclass == "hard" and artifact["present"]
      resolve_now = (
        clean >= hard_clean_threshold if artifact_backed
        else quiet >= soft_runs_threshold
      )
      if resolve_now:
        if artifact_backed:
          basis, confidence = "artifact_verified", "confirmed"
        else:
          basis = "retested_clean" if rclass == "hard" else "not_seen_n_runs"
          confidence = "observational"
        cur.execute(
          f"""
          UPDATE {ISSUES_TABLE}
          SET clean_exercise_count=%s, runs_since_last_seen=%s,
              status='resolved', resolved_detected_at=NOW(6),
              resolution_basis=%s, resolution_confidence=%s
          WHERE issue_id=%s
          """,
          (clean, quiet, basis, confidence, issue_id),
        )
        # The event type follows the CONFIDENCE, not the class: a hard issue
        # resolved on silence alone must not leave a 'resolved_confirmed'
        # row in the audit trail.
        event = ("resolved_confirmed" if confidence == "confirmed"
                 else "resolved_observational")
        _insert_resolution_event(
          cur, issue_id=issue_id, signature=signature, draft_id=draft_id,
          event_type=event,
          detail={"basis": basis, "confidence": confidence,
                  "clean_exercise_count": clean, "runs_since_last_seen": quiet,
                  "exercised_via": verdict["reason"],
                  "artifact_verified": artifact["details"] or None},
        )
        summary[f"resolved_{confidence}"].append(signature)
      else:
        cur.execute(
          f"""
          UPDATE {ISSUES_TABLE}
          SET clean_exercise_count=%s, runs_since_last_seen=%s
          WHERE issue_id=%s
          """,
          (clean, quiet, issue_id),
        )
        _insert_resolution_event(
          cur, issue_id=issue_id, signature=signature, draft_id=draft_id,
          event_type="exercised_clean",
          detail={"clean_exercise_count": clean, "runs_since_last_seen": quiet,
                  "exercised_via": verdict["reason"],
                  "artifact_verified": artifact["details"] or None},
        )
    conn.commit()
  finally:
    try:
      cur.close()
    except Exception:
      pass
  return summary


def reclassify_unearned_confirmations(conn, *, dry_run: bool = False) -> Dict[str, Any]:
  """Retire the 'confirmed' verdicts that were never earned.

  Before the artifact gate, EVERY hard resolution was stamped 'confirmed' on
  basis='retested_clean', which meant only "a run finished, visited the same
  section, and nobody re-filed the signature". That is opportunity plus
  silence, not verification. This demotes those verdicts to 'observational'
  and leaves an INSERT-only audit row for each.

  Deliberately NOT touched: status stays 'resolved' (the agenda is not
  flooded with issues nobody has evidence against), basis stays
  'retested_clean' (the history is not rewritten), and basis='manual'
  verdicts stay 'confirmed' (a human looked). Any of these re-earns
  'confirmed' the moment it is given an artifact assertion that passes.

  Idempotent: a second call finds nothing to do.
  """
  ensure_tables(conn)
  cur = conn.cursor()
  try:
    cur.execute(
      f"""
      SELECT issue_id, signature, category, severity FROM {ISSUES_TABLE}
      WHERE resolution_confidence = 'confirmed'
        AND resolution_basis = 'retested_clean'
      ORDER BY issue_id
      """
    )
    targets = [
      {"issue_id": int(r[0]), "signature": str(r[1]),
       "category": str(r[2]), "severity": str(r[3])}
      for r in cur.fetchall()
    ]
    if not dry_run:
      for t in targets:
        cur.execute(
          f"""
          UPDATE {ISSUES_TABLE} SET resolution_confidence = 'observational'
          WHERE issue_id = %s
          """,
          (t["issue_id"],),
        )
        _insert_resolution_event(
          cur, issue_id=t["issue_id"], signature=t["signature"], draft_id="",
          event_type="confidence_demoted",
          detail={
            "from": "confirmed", "to": "observational",
            "reason": ("resolved before the artifact gate: evidence was "
                       "opportunity (run completed / section visited) plus "
                       "reporter silence, with no artifact read"),
            "re_earn": ("attach an 'artifact' assertion via set_probe(); the "
                        "next exercised run re-judges it"),
          },
        )
      conn.commit()
  finally:
    cur.close()
  return {
    "dry_run": dry_run,
    "demoted": len(targets),
    "signatures": [t["signature"] for t in targets],
  }


def probe_audit(conn) -> Dict[str, Any]:
  """What does each issue's detector actually verify? (CW-031 tier 1)

  Buckets every issue by the STRONGEST evidence its probe can produce:
    artifact     - reads a persisted artifact; can earn 'confirmed'
    opportunity  - only proves the run walked the path; 'observational' cap
    metadata     - notes/pins only: states no retest condition, ticks nothing
    manual       - explicitly human-retest-only
    derived      - no probe at all; a section guess is derived from the first
                   occurrence, which is opportunity at its weakest
  """
  ensure_tables(conn)
  cur = conn.cursor()
  try:
    cur.execute(
      f"""SELECT issue_id, signature, status, resolution_class,
                 resolution_confidence, probe_json
          FROM {ISSUES_TABLE} ORDER BY issue_id"""
    )
    rows = cur.fetchall()
  finally:
    cur.close()
  buckets: Dict[str, List[str]] = {
    "artifact": [], "opportunity": [], "metadata": [], "manual": [], "derived": [],
  }
  for issue_id, signature, status, rclass, confidence, raw in rows:
    probe: Any = None
    if raw:
      try:
        probe = json.loads(raw) if isinstance(raw, (str, bytes)) else raw
      except Exception:
        probe = None
    if not isinstance(probe, dict) or not probe:
      bucket = "derived"
    elif probe.get("manual_only"):
      bucket = "manual"
    elif probe.get("artifact"):
      bucket = "artifact"
    elif (set(probe) & PROBE_OPPORTUNITY_KEYS) - {"require_completed"}:
      bucket = "opportunity"
    else:
      bucket = "metadata"
    buckets[bucket].append(f"#{issue_id} {signature}")
  return {
    "total": len(rows),
    "counts": {k: len(v) for k, v in buckets.items()},
    "can_earn_confirmed": len(buckets["artifact"]),
    "buckets": buckets,
  }


# ----------------------------------------------------------------------------
# Read helpers (dashboard / Cowork agenda queries).
# ----------------------------------------------------------------------------

def list_issues(
  conn,
  *,
  status: str = "",
  category: str = "",
  limit: int = 200,
) -> List[Dict[str, Any]]:
  ensure_tables(conn)
  where: List[str] = []
  params: List[Any] = []
  if status:
    where.append("status = %s")
    params.append(_require(status, STATUSES, "status"))
  if category:
    where.append("category = %s")
    params.append(_require(category, CATEGORIES, "category"))
  sql = f"SELECT * FROM {ISSUES_TABLE}"
  if where:
    sql += " WHERE " + " AND ".join(where)
  sql += " ORDER BY last_seen_at DESC LIMIT %s"
  params.append(min(500, max(1, int(limit))))
  cur = conn.cursor()
  try:
    cur.execute(sql, tuple(params))
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]
  finally:
    cur.close()


def recent_occurrences(conn, *, limit: int = 100) -> List[Dict[str, Any]]:
  ensure_tables(conn)
  cur = conn.cursor()
  try:
    cur.execute(
      f"""
      SELECT o.id, o.signature, o.draft_id, o.planning_run_id,
             o.business_name, o.persona, o.turn_index, o.section, o.stage,
             o.severity, LEFT(o.observed, 400) AS observed,
             LEFT(o.expected, 400) AS expected, o.source, o.created_at,
             i.category, i.status
      FROM {OCCURRENCES_TABLE} o
      JOIN {ISSUES_TABLE} i ON i.issue_id = o.issue_id
      ORDER BY o.id DESC LIMIT %s
      """,
      (min(500, max(1, int(limit))),),
    )
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in cur.fetchall()]
  finally:
    cur.close()
