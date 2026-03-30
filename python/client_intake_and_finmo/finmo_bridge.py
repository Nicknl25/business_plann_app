from __future__ import annotations

import json
import re
import subprocess
import tempfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.datetime import from_excel

try:
  from financial_model_engine.finmo_model import calculate_finmo_model
  from financial_model_engine.model_inputs import FinancialModelInputs
except Exception:
  import sys

  ROOT = Path(__file__).resolve().parents[1]
  if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
  from financial_model_engine.finmo_model import calculate_finmo_model
  from financial_model_engine.model_inputs import FinancialModelInputs


def _safe_float(value: Any) -> Optional[float]:
  if value is None or value == "":
    return None
  try:
    return float(value)
  except Exception:
    return None


def _safe_int(value: Any) -> int:
  if value is None or value == "":
    return 0
  try:
    return int(float(value))
  except Exception:
    return 0


def _as_iso_date(value: Any) -> Optional[str]:
  if value is None or value == "":
    return None
  if isinstance(value, datetime):
    return value.date().isoformat()
  if isinstance(value, str):
    cleaned = value.strip()
    if not cleaned:
      return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
      try:
        return datetime.strptime(cleaned, fmt).date().isoformat()
      except Exception:
        continue
    return cleaned
  numeric = _safe_float(value)
  if numeric is None:
    return None
  try:
    return from_excel(numeric).date().isoformat()
  except Exception:
    return None


def _ratio(numerator: Any, denominator: Any) -> float:
  num = _safe_float(numerator) or 0.0
  den = _safe_float(denominator) or 0.0
  if abs(den) < 1e-9:
    return 0.0
  return num / den


def _clone(value: Any) -> Any:
  return deepcopy(value)


def _canonical_model_input_text(value: Any) -> str:
  return str(value or "").strip()


def _revenue_lever_id(lob: Any, product: Any, driver: Any) -> str:
  return "::".join(
    [
      "revenue",
      _canonical_model_input_text(lob),
      _canonical_model_input_text(product),
      _canonical_model_input_text(driver),
    ]
  )


def _simple_lever_id(section: str, label: Any) -> str:
  return "::".join([section, _canonical_model_input_text(label)])


def _full_quarter_scope(slots: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
  full_slots = _full_quarter_slots(slots)
  valid_quarter_indices = list(range(1, len(full_slots) + 1))
  valid_period_columns = [str(slot.get("column_letter") or "").strip() for slot in full_slots if str(slot.get("column_letter") or "").strip()]
  return {
    "valid_quarter_indices": valid_quarter_indices,
    "valid_period_columns": valid_period_columns,
    "writable_full_quarters_only": True,
  }


def _revenue_input_semantics(driver: str) -> Dict[str, str]:
  driver_text = _canonical_model_input_text(driver).lower()
  if driver_text == "capacity":
    return {"value_kind": "direct_number", "input_semantics": "quarter_capacity_units"}
  if driver_text == "unit price":
    return {"value_kind": "direct_number", "input_semantics": "currency_per_unit"}
  if driver_text == "utilization":
    return {"value_kind": "ratio", "input_semantics": "utilization_ratio"}
  return {"value_kind": "direct_number", "input_semantics": "direct_input"}


def _simple_input_semantics(section_key: str, label: str) -> Dict[str, str]:
  normalized_section = _canonical_model_input_text(section_key).lower()
  normalized_label = _canonical_model_input_text(label).lower()
  if normalized_section == "expenses":
    if normalized_label in {
      "cost of goods sold",
      "marketing",
      "research & development",
      "general & administrative",
      "interest rate",
      "depreciation",
      "taxes",
    }:
      return {"value_kind": "ratio", "input_semantics": "percent_of_revenue"}
    if normalized_label in {"lease", "payroll"}:
      return {"value_kind": "direct_number", "input_semantics": "quarter_currency"}
  if normalized_section == "balance_sheet":
    if normalized_label in {"accounts receivable days", "inventory days", "accounts payable days"}:
      return {"value_kind": "day_count", "input_semantics": "days"}
    if normalized_label in {"prepaid expenses", "deferred revnue"}:
      return {"value_kind": "ratio", "input_semantics": "percent_of_revenue"}
    if normalized_label == "short term debt (% of ltd)":
      return {"value_kind": "ratio", "input_semantics": "percent_of_long_term_debt"}
    if normalized_label in {
      "ppe $ (excluding capital leases)",
      "accumulated depreciation",
      "owner's capital",
      "other equity",
    }:
      return {"value_kind": "direct_number", "input_semantics": "quarter_currency"}
  if normalized_section == "schedules":
    return {"value_kind": "direct_number", "input_semantics": "quarter_currency"}
  return {"value_kind": "direct_number", "input_semantics": "direct_input"}


def _safe_ratio(value: Any) -> Optional[float]:
  ratio = _safe_float(value)
  if ratio is None:
    return None
  if ratio > 1.0 and ratio <= 100.0:
    ratio = ratio / 100.0
  return ratio


def _placeholder_index(value: Any, prefix: str) -> Optional[int]:
  match = re.search(rf"{re.escape(prefix)}\s*(\d+)", str(value or "").strip(), re.IGNORECASE)
  if not match:
    return None
  try:
    return max(0, int(match.group(1)) - 1)
  except Exception:
    return None


def _revenue_slot_key(lob_index: int, product_index: int) -> str:
  return f"lob_{max(0, int(lob_index)) + 1}_product_{max(0, int(product_index)) + 1}"


def _revenue_slot_identity(
  *,
  row_lob: Any,
  row_product: Any,
  revenue_row_ordinal: Optional[int] = None,
) -> Dict[str, Any]:
  lob_index = _placeholder_index(row_lob, "LOB")
  product_index = _placeholder_index(row_product, "Product")
  if lob_index is None or product_index is None:
    slot_ordinal = max(0, int(revenue_row_ordinal or 0)) // 3
    lob_index = slot_ordinal // 3
    product_index = slot_ordinal % 3
  return {
    "lob_slot_index": lob_index,
    "product_slot_index": product_index,
    "revenue_slot_key": _revenue_slot_key(lob_index, product_index),
  }


def _named_lob(value: Any, fallback: str) -> str:
  text = _canonical_model_input_text(value)
  return text or fallback


def _named_product(value: Any, fallback: str) -> str:
  text = _canonical_model_input_text(value)
  return text or fallback


def _normalize_finmo_path(finmo_path: Any) -> str:
  cleaned = str(finmo_path or "").strip()
  if not cleaned:
    raise ValueError("finmo_path is required")
  return cleaned


def _defined_range_bounds(wb, name: str) -> Tuple[Any, int, int, int, int]:
  if name not in wb.defined_names:
    raise KeyError(f"Named range not found in workbook: {name}")
  destinations = list(wb.defined_names[name].destinations)
  if not destinations:
    raise ValueError(f"Named range has no destinations: {name}")
  sheet_name, ref = destinations[0]
  min_col, min_row, max_col, max_row = range_boundaries(ref)
  return wb[sheet_name], min_row, max_row, min_col, max_col


def _period_slots_from_model_inputs(wb) -> List[Dict[str, Any]]:
  ws, min_row, _max_row, min_col, max_col = _defined_range_bounds(wb, "model_input_periods")
  label_to_row: Dict[str, int] = {}
  for row_idx in range(min_row, min_row + 5):
    label = str(ws.cell(row=row_idx, column=min_col + 1).value or "").strip().lower()
    if label:
      label_to_row[label] = row_idx
  quarter_row = label_to_row.get("quarter")
  year_row = label_to_row.get("year")
  date_row = label_to_row.get("date")
  fraction_row = label_to_row.get("year fraction")
  if quarter_row is None:
    return []
  slots: List[Dict[str, Any]] = []
  slot_index = 0
  for col_idx in range(min_col + 2, max_col + 1):
    quarter_val = _safe_float(ws.cell(row=quarter_row, column=col_idx).value)
    if quarter_val is None:
      continue
    year_val = _safe_float(ws.cell(row=year_row, column=col_idx).value) if year_row is not None else None
    slots.append(
      {
        "slot_index": slot_index,
        "column_index": col_idx,
        "column_letter": get_column_letter(col_idx),
        "year": year_val,
        "quarter": quarter_val,
        "date": _as_iso_date(ws.cell(row=date_row, column=col_idx).value) if date_row is not None else None,
        "year_fraction": _safe_float(ws.cell(row=fraction_row, column=col_idx).value) if fraction_row is not None else None,
        "is_stub": quarter_val == 0.0,
      }
    )
    slot_index += 1
  return slots


def _period_slots_from_finmo(wb) -> List[Dict[str, Any]]:
  ws, min_row, _max_row, min_col, max_col = _defined_range_bounds(wb, "finmo_periods")
  label_to_row: Dict[str, int] = {}
  for row_idx in range(min_row, min_row + 3):
    label = str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower()
    if label:
      label_to_row[label] = row_idx
  quarter_row = label_to_row.get("quarter")
  year_row = label_to_row.get("year")
  date_row = label_to_row.get("date")
  if quarter_row is None or year_row is None or date_row is None:
    return []
  slots: List[Dict[str, Any]] = []
  slot_index = 0
  for col_idx in range(min_col + 1, max_col + 1):
    quarter_val = ws.cell(row=quarter_row, column=col_idx).value
    if quarter_val in (None, ""):
      continue
    slots.append(
      {
        "slot_index": slot_index,
        "column_index": col_idx,
        "column_letter": get_column_letter(col_idx),
        "year": _safe_float(ws.cell(row=year_row, column=col_idx).value),
        "quarter": _safe_float(quarter_val),
        "date": _as_iso_date(ws.cell(row=date_row, column=col_idx).value),
      }
    )
    slot_index += 1
  return slots


def _read_named_cell(wb, name: str) -> Any:
  if name not in wb.defined_names:
    raise KeyError(f"Named range not found in workbook: {name}")
  destinations = list(wb.defined_names[name].destinations)
  if not destinations:
    raise ValueError(f"Named range has no destinations: {name}")
  sheet_name, cell_ref = destinations[0]
  return wb[sheet_name][cell_ref].value


def _write_named_cell(wb, name: str, value: Any) -> None:
  if name not in wb.defined_names:
    raise KeyError(f"Named range not found in workbook: {name}")
  destinations = list(wb.defined_names[name].destinations)
  if not destinations:
    raise ValueError(f"Named range has no destinations: {name}")
  sheet_name, cell_ref = destinations[0]
  wb[sheet_name][cell_ref].value = value


def _model_input_slot_for_full_quarter(slots: Sequence[Dict[str, Any]], quarter_index: int) -> Optional[Dict[str, Any]]:
  quarter_index = max(1, int(quarter_index or 1))
  full_slots = [slot for slot in (slots or []) if _safe_float(slot.get("quarter")) not in (None, 0.0)]
  if quarter_index <= len(full_slots):
    return full_slots[quarter_index - 1]
  return full_slots[-1] if full_slots else None


def _finmo_slot_for_full_quarter(slots: Sequence[Dict[str, Any]], quarter_index: int) -> Optional[Dict[str, Any]]:
  quarter_index = max(1, int(quarter_index or 1))
  full_slots = [slot for slot in (slots or []) if _safe_float(slot.get("quarter")) not in (None, 0.0)]
  if quarter_index <= len(full_slots):
    return full_slots[quarter_index - 1]
  return full_slots[-1] if full_slots else None


def _read_model_input_json(finmo_path: str) -> Dict[str, Any]:
  wb = load_workbook(finmo_path, data_only=False)
  try:
    slots = _period_slots_from_model_inputs(wb)
    slot_columns = [int(slot["column_index"]) for slot in slots]
    quarter_scope = _full_quarter_scope(slots)
    controller_write_levers: List[Dict[str, Any]] = []
    lever_catalog: Dict[str, Dict[str, Any]] = {}

    def _register_lever(metadata: Dict[str, Any]) -> None:
      lever_id = str(metadata.get("lever_id") or "").strip()
      if not lever_id:
        return
      lever_meta = {
        **_clone(metadata),
        **quarter_scope,
      }
      lever_catalog[lever_id] = lever_meta
      controller_write_levers.append(_clone(lever_meta))

    revenue_rows: List[Dict[str, Any]] = []
    ws, min_row, max_row, min_col, max_col = _defined_range_bounds(wb, "model_input_revenue")
    revenue_row_ordinal = 0
    for row_idx in range(min_row, max_row + 1):
      controller_marker = str(ws.cell(row=row_idx, column=min_col).value or "").strip()
      if controller_marker.lower() != "controller write":
        continue
      lob = str(ws.cell(row=row_idx, column=min_col + 2).value or "").strip()
      product = str(ws.cell(row=row_idx, column=min_col + 3).value or "").strip()
      driver = str(ws.cell(row=row_idx, column=min_col + 4).value or "").strip()
      slot_identity = _revenue_slot_identity(
        row_lob=lob,
        row_product=product,
        revenue_row_ordinal=revenue_row_ordinal,
      )
      lever_id = _revenue_lever_id(lob, product, driver)
      semantics = _revenue_input_semantics(driver)
      revenue_rows.append(
        {
          "named_range": "model_input_revenue",
          "controller_write": True,
          "lever_id": lever_id,
          "placeholder_lob": lob,
          "placeholder_product": product,
          **slot_identity,
          "lob": lob,
          "product": product,
          "driver": driver,
          **semantics,
          **quarter_scope,
          "values": [ws.cell(row=row_idx, column=col_idx).value for col_idx in slot_columns if col_idx <= max_col],
        }
      )
      _register_lever(
        {
          "lever_id": lever_id,
          "named_range": "model_input_revenue",
          "section": "revenue",
          **slot_identity,
          "lob": lob,
          "product": product,
          "driver": driver,
          "label_path": f"{lob} > {product} > {driver}",
          **semantics,
        }
      )
      revenue_row_ordinal += 1

    def _read_simple_rows(range_name: str, *, section_key: str) -> List[Dict[str, Any]]:
      local_ws, local_min_row, local_max_row, local_min_col, local_max_col = _defined_range_bounds(wb, range_name)
      rows: List[Dict[str, Any]] = []
      for row_idx in range(local_min_row, local_max_row + 1):
        controller_marker = str(local_ws.cell(row=row_idx, column=local_min_col).value or "").strip()
        if controller_marker.lower() != "controller write":
          continue
        label = str(local_ws.cell(row=row_idx, column=local_min_col + 1).value or "").strip()
        lever_id = _simple_lever_id(section_key, label)
        semantics = _simple_input_semantics(section_key, label)
        rows.append(
          {
            "named_range": range_name,
            "controller_write": True,
            "lever_id": lever_id,
            "label": label,
            **semantics,
            **quarter_scope,
            "values": [local_ws.cell(row=row_idx, column=col_idx).value for col_idx in slot_columns if col_idx <= local_max_col],
          }
        )
        _register_lever(
          {
            "lever_id": lever_id,
            "named_range": range_name,
            "section": section_key,
            "label": label,
            "label_path": label,
            **semantics,
          }
        )
      return rows

    schedule_ws, sched_min_row, sched_max_row, sched_min_col, sched_max_col = _defined_range_bounds(wb, "model_input_schedules")
    seed_col = slot_columns[0] - 1 if slot_columns else sched_min_col + 2
    schedule_rows = _read_simple_rows("model_input_schedules", section_key="schedules")

    return {
      "contract_version": "finmo_model_input_v3",
      "canonical_lever_vocabulary": "model_inputs_controller_write_only",
      "finmo_path": finmo_path,
      "start_date": _as_iso_date(_read_named_cell(wb, "model_input_startdate")),
      "periods": slots,
      "lever_catalog": lever_catalog,
      "controller_write_levers": controller_write_levers,
      "sections": {
        "revenue": revenue_rows,
        "expenses": _read_simple_rows("model_input_expenses", section_key="expenses"),
        "balance_sheet": _read_simple_rows("model_input_balancehseet", section_key="balance_sheet"),
        "schedules": {
          "debt_opening_balance_seed": schedule_ws.cell(row=sched_min_row + 3, column=seed_col).value,
          "lease_opening_balance_seed": schedule_ws.cell(row=sched_min_row + 12, column=seed_col).value,
          "rows": schedule_rows,
        },
      },
    }
  finally:
    wb.close()


def _read_finmo_json(finmo_path: str) -> Dict[str, Any]:
  wb = load_workbook(finmo_path, data_only=True)
  try:
    slots = _period_slots_from_finmo(wb)
    slot_columns = [int(slot["column_index"]) for slot in slots]

    def _read_rows(range_name: str) -> List[Dict[str, Any]]:
      ws, min_row, max_row, min_col, max_col = _defined_range_bounds(wb, range_name)
      rows: List[Dict[str, Any]] = []
      for row_idx in range(min_row, max_row + 1):
        label = str(ws.cell(row=row_idx, column=min_col).value or "").strip()
        if not label:
          continue
        rows.append(
          {
            "label": label,
            "values": [ws.cell(row=row_idx, column=col_idx).value for col_idx in slot_columns if col_idx <= max_col],
          }
        )
      return rows

    accounting_rows = _read_rows("finmo_accountingcheck")
    pl_rows = _read_rows("finmo_pl")
    balance_rows = _read_rows("finmo_balancesheet")
    cfs_rows = _read_rows("finmo_cfs")

    row_maps = {
      "pl": {row["label"]: row["values"] for row in pl_rows},
      "balance": {row["label"]: row["values"] for row in balance_rows},
      "cfs": {row["label"]: row["values"] for row in cfs_rows},
    }
    quarter_rows: List[Dict[str, Any]] = []
    for slot in slots:
      idx = int(slot["slot_index"])
      quarter_rows.append(
        {
          "slot_index": idx,
          "year": slot.get("year"),
          "quarter": slot.get("quarter"),
          "date": slot.get("date"),
          "revenue": (row_maps["pl"].get("Revenue") or [None])[idx] if idx < len(row_maps["pl"].get("Revenue") or []) else None,
          "cogs": (row_maps["pl"].get("Cost of Goods Sold") or [None])[idx] if idx < len(row_maps["pl"].get("Cost of Goods Sold") or []) else None,
          "gross_profit": (row_maps["pl"].get("Gross Profit") or [None])[idx] if idx < len(row_maps["pl"].get("Gross Profit") or []) else None,
          "marketing": (row_maps["pl"].get("Marketing") or [None])[idx] if idx < len(row_maps["pl"].get("Marketing") or []) else None,
          "research_and_development": (row_maps["pl"].get("Research & Development") or [None])[idx] if idx < len(row_maps["pl"].get("Research & Development") or []) else None,
          "lease_rent": (row_maps["pl"].get("Lease/Rent") or [None])[idx] if idx < len(row_maps["pl"].get("Lease/Rent") or []) else None,
          "payroll": (row_maps["pl"].get("Payroll") or [None])[idx] if idx < len(row_maps["pl"].get("Payroll") or []) else None,
          "g_and_a": (row_maps["pl"].get("General & Administrative") or [None])[idx] if idx < len(row_maps["pl"].get("General & Administrative") or []) else None,
          "ebitda": (row_maps["pl"].get("EBITDA") or [None])[idx] if idx < len(row_maps["pl"].get("EBITDA") or []) else None,
          "interest": (row_maps["pl"].get("Interest") or [None])[idx] if idx < len(row_maps["pl"].get("Interest") or []) else None,
          "depreciation": (row_maps["pl"].get("Depreciation") or [None])[idx] if idx < len(row_maps["pl"].get("Depreciation") or []) else None,
          "taxes": (row_maps["pl"].get("Taxes") or [None])[idx] if idx < len(row_maps["pl"].get("Taxes") or []) else None,
          "net_income": (row_maps["pl"].get("Net Income") or [None])[idx] if idx < len(row_maps["pl"].get("Net Income") or []) else None,
          "cash": (row_maps["balance"].get("Cash") or [None])[idx] if idx < len(row_maps["balance"].get("Cash") or []) else None,
          "total_assets": (row_maps["balance"].get("Total Assets") or [None])[idx] if idx < len(row_maps["balance"].get("Total Assets") or []) else None,
          "total_liabilities_and_equity": (row_maps["balance"].get("Total Liabilities & Equity") or [None])[idx] if idx < len(row_maps["balance"].get("Total Liabilities & Equity") or []) else None,
          "ending_cash": (row_maps["cfs"].get("Ending Cash") or [None])[idx] if idx < len(row_maps["cfs"].get("Ending Cash") or []) else None,
        }
      )

    check_status_values = accounting_rows[0]["values"] if accounting_rows else []
    check_numeric_values = accounting_rows[1]["values"] if len(accounting_rows) > 1 else []
    all_ok = all(str(value or "").strip().upper() == "OK" for value in check_status_values if value not in (None, ""))

    return {
      "contract_version": "finmo_output_v1",
      "finmo_path": finmo_path,
      "periods": slots,
      "accounting_check": {
        "rows": accounting_rows,
        "all_ok": all_ok,
        "status_values": check_status_values,
        "numeric_values": check_numeric_values,
      },
      "pl": pl_rows,
      "balance_sheet": balance_rows,
      "cash_flow": cfs_rows,
      "quarter_rows": quarter_rows,
    }
  finally:
    wb.close()


def build_python_finmo_json(
  *,
  model_input_json: Dict[str, Any],
  finmo_path: Optional[str] = None,
) -> Dict[str, Any]:
  book = FinancialModelInputs.from_model_input_json(model_input_json if isinstance(model_input_json, dict) else {})
  result = calculate_finmo_model(book)
  quarter_rows_raw = result.quarter_rows()
  raw_periods = [
    _clone(item)
    for item in (((model_input_json.get("periods") or []) if isinstance(model_input_json, dict) else []) or [])
    if isinstance(item, dict)
  ]
  periods: List[Dict[str, Any]] = []
  start_date_iso = _as_iso_date((model_input_json or {}).get("start_date")) if isinstance(model_input_json, dict) else None
  if raw_periods:
    opening_year = _safe_float(raw_periods[0].get("year"))
    if opening_year is None and start_date_iso:
      try:
        opening_year = float(datetime.fromisoformat(start_date_iso).year)
      except Exception:
        opening_year = None
    periods.append(
      {
        "slot_index": 0,
        "column_index": 4,
        "column_letter": "D",
        "year": opening_year,
        "quarter": 0.0,
        "date": start_date_iso or raw_periods[0].get("date"),
      }
    )
    for idx, item in enumerate(raw_periods, start=1):
      periods.append(
        {
          "slot_index": idx,
          "column_index": 4 + idx,
          "column_letter": get_column_letter(4 + idx),
          "year": item.get("year"),
          "quarter": item.get("quarter"),
          "date": item.get("date"),
        }
      )
  else:
    opening_date = start_date_iso
    opening_year = None
    if opening_date:
      try:
        opening_year = float(datetime.fromisoformat(opening_date).year)
      except Exception:
        opening_year = None
    periods = [
      {
        "slot_index": 0,
        "column_index": 4,
        "column_letter": "D",
        "year": opening_year,
        "quarter": 0.0,
        "date": opening_date,
      }
    ]
    for idx, row in enumerate(quarter_rows_raw, start=1):
      if not isinstance(row, dict):
        continue
      periods.append(
        {
          "slot_index": idx,
          "column_index": 4 + idx,
          "column_letter": get_column_letter(4 + idx),
          "year": row.get("year"),
          "quarter": row.get("quarter"),
          "date": row.get("date"),
        }
      )

  def _series(metric_key: str) -> List[float]:
    values: List[float] = []
    for row in quarter_rows_raw:
      if not isinstance(row, dict):
        continue
      values.append(round(_safe_float(row.get(metric_key)) or 0.0, 6))
    return values

  pl_rows = [
    {"label": "Revenue", "values": _series("revenue")},
    {"label": "Cost of Goods Sold", "values": _series("cost_of_goods_sold")},
    {"label": "Gross Profit", "values": _series("gross_profit")},
    {"label": "Marketing", "values": _series("marketing")},
    {"label": "Research & Development", "values": _series("research_and_development")},
    {"label": "Lease/Rent", "values": _series("lease_rent")},
    {"label": "Payroll", "values": _series("payroll")},
    {"label": "General & Administrative", "values": _series("general_and_administrative")},
    {"label": "EBITDA", "values": _series("ebitda")},
    {"label": "Interest", "values": _series("interest")},
    {"label": "Depreciation", "values": _series("depreciation")},
    {"label": "Taxes", "values": _series("taxes")},
    {"label": "Net Income", "values": _series("net_income")},
  ]
  balance_rows = [
    {"label": "Cash", "values": _series("cash")},
    {"label": "Accounts Receivable", "values": _series("accounts_receivable")},
    {"label": "Inventory", "values": _series("inventory")},
    {"label": "Current Assets", "values": _series("current_assets")},
    {"label": "PPE", "values": _series("ppe")},
    {"label": "Accumulated Depreciation", "values": _series("accumulated_depreciation")},
    {"label": "Total Assets", "values": _series("total_assets")},
    {"label": "Accounts Payable", "values": _series("accounts_payable")},
    {"label": "Prepaid Expenses", "values": _series("prepaid_expenses")},
    {"label": "Short Term Debt", "values": _series("short_term_debt")},
    {"label": "Deferred Revenue", "values": _series("deferred_revenue")},
    {"label": "Current Liabilites", "values": _series("current_liabilities")},
    {"label": "Long Term Debt", "values": _series("long_term_debt")},
    {"label": "Total Liabilities", "values": _series("total_liabilities")},
    {"label": "Owner's Capital", "values": _series("owners_capital")},
    {"label": "Retained Earnings", "values": _series("retained_earnings")},
    {"label": "Other Equity", "values": _series("other_equity")},
    {"label": "Total Equity", "values": _series("total_equity")},
    {"label": "Total Liabilities & Equity", "values": _series("total_liabilities_and_equity")},
  ]
  cfs_rows = [
    {"label": "Beginning Cash", "values": _series("beginning_cash")},
    {"label": "Net Income", "values": _series("net_income")},
    {"label": "Depreciatoin", "values": _series("depreciation")},
    {"label": "Changes in Current Assets", "values": _series("changes_in_current_assets")},
    {"label": "Changes in Current Liabilites", "values": _series("changes_in_current_liabilities")},
    {"label": "Operating Cash Flow", "values": _series("operating_cash_flow")},
    {"label": "Capital Expenditures", "values": _series("capital_expenditures")},
    {"label": "Investing Cash Flow", "values": _series("investing_cash_flow")},
    {"label": "Dept Receive(Repay)", "values": _series("debt_receive_repay")},
    {"label": "Equity", "values": _series("equity")},
    {"label": "Financing Cash Flow", "values": _series("financing_cash_flow")},
    {"label": "Net Cash Flow", "values": _series("net_cash_flow")},
    {"label": "Ending Cash", "values": _series("ending_cash")},
  ]
  numeric_values = _series("accounting_equation_check")
  tolerance = 1.0
  status_values = ["OK" if abs(value) <= tolerance else "FAIL" for value in numeric_values]
  quarter_rows: List[Dict[str, Any]] = [
    {
      "slot_index": 0,
      "year": periods[0].get("year") if periods else None,
      "quarter": 0.0,
      "date": periods[0].get("date") if periods else None,
      "revenue": 0,
      "cogs": 0,
      "gross_profit": 0,
      "marketing": 0,
      "research_and_development": 0,
      "lease_rent": 0,
      "payroll": 0,
      "g_and_a": 0,
      "ebitda": 0,
      "interest": 0,
      "depreciation": 0,
      "taxes": 0,
      "net_income": 0,
      "cash": 0,
      "total_assets": 0,
      "total_liabilities_and_equity": 0,
      "ending_cash": 0,
    }
  ]
  for idx, row in enumerate(quarter_rows_raw):
    if not isinstance(row, dict):
      continue
    quarter_rows.append(
      {
        "slot_index": idx + 1,
        "quarter_index": int(row.get("quarter_index") or idx + 1),
        "year": row.get("year"),
        "quarter": row.get("quarter"),
        "date": row.get("date"),
        "revenue": row.get("revenue"),
        "cogs": row.get("cost_of_goods_sold"),
        "gross_profit": row.get("gross_profit"),
        "marketing": row.get("marketing"),
        "research_and_development": row.get("research_and_development"),
        "lease_rent": row.get("lease_rent"),
        "payroll": row.get("payroll"),
        "g_and_a": row.get("general_and_administrative"),
        "ebitda": row.get("ebitda"),
        "interest": row.get("interest"),
        "depreciation": row.get("depreciation"),
        "taxes": row.get("taxes"),
        "net_income": row.get("net_income"),
        "cash": row.get("cash"),
        "ending_cash": row.get("ending_cash"),
        "total_assets": row.get("total_assets"),
        "total_liabilities_and_equity": row.get("total_liabilities_and_equity"),
      }
    )

  return {
    "contract_version": "finmo_output_v1",
    "finmo_path": str(finmo_path or "").strip(),
    "periods": periods,
    "accounting_check": {
      "rows": [
        {"label": "Check", "values": status_values},
        {"label": "Accounting Equation Check", "values": numeric_values},
      ],
      "all_ok": all(item == "OK" for item in status_values),
      "status_values": status_values,
      "numeric_values": numeric_values,
    },
    "pl": pl_rows,
    "balance_sheet": balance_rows,
    "cash_flow": cfs_rows,
    "quarter_rows": quarter_rows,
  }


def build_consistency_forecast_view_from_finmo(finmo_json: Dict[str, Any]) -> Dict[str, Any]:
  finmo_obj = finmo_json if isinstance(finmo_json, dict) else {}
  raw_rows = [row for row in (finmo_obj.get("quarter_rows") or []) if isinstance(row, dict)]
  full_quarters = [
    row for row in raw_rows
    if _safe_float(row.get("quarter")) not in (None, 0.0)
  ]
  if not full_quarters:
    full_quarters = raw_rows
  full_quarters = full_quarters[:20]

  quarter_view: List[Dict[str, Any]] = []
  for idx, row in enumerate(full_quarters, start=1):
    r_and_d = _safe_float(row.get("research_and_development")) or 0.0
    lease_rent = _safe_float(row.get("lease_rent")) or 0.0
    g_and_a = _safe_float(row.get("g_and_a")) or 0.0
    quarter_view.append(
      {
        "quarter_index": idx,
        "period_label": f"Year {((idx - 1) // 4) + 1} Q{((idx - 1) % 4) + 1}",
        "year": _safe_float(row.get("year")),
        "quarter": _safe_float(row.get("quarter")),
        "date": row.get("date"),
        "revenue": _safe_float(row.get("revenue")) or 0.0,
        "cogs": _safe_float(row.get("cogs")) or 0.0,
        "gross_profit": _safe_float(row.get("gross_profit")) or 0.0,
        "marketing": _safe_float(row.get("marketing")) or 0.0,
        "research_and_development": r_and_d,
        "lease_rent": lease_rent,
        "payroll": _safe_float(row.get("payroll")) or 0.0,
        "g_and_a": g_and_a,
        "opex": round(r_and_d + lease_rent + g_and_a, 6),
        "ebitda": _safe_float(row.get("ebitda")) or 0.0,
        "interest": _safe_float(row.get("interest")) or 0.0,
        "depreciation": _safe_float(row.get("depreciation")) or 0.0,
        "taxes": _safe_float(row.get("taxes")) or 0.0,
        "net_income": _safe_float(row.get("net_income")) or 0.0,
        "cash": _safe_float(row.get("cash")) or 0.0,
        "ending_cash": _safe_float(row.get("ending_cash")) or 0.0,
        "total_assets": _safe_float(row.get("total_assets")) or 0.0,
        "total_liabilities_and_equity": _safe_float(row.get("total_liabilities_and_equity")) or 0.0,
        "source_level": "finmo",
      }
    )

  forecast_years: List[Dict[str, Any]] = []
  for year_index in range(0, len(quarter_view), 4):
    year_quarters = quarter_view[year_index:year_index + 4]
    if not year_quarters:
      continue
    year_number = (year_index // 4) + 1
    forecast_years.append(
      {
        "year_index": year_number,
        "period_label": f"Year {year_number}",
        "revenue": round(sum(_safe_float(item.get("revenue")) or 0.0 for item in year_quarters), 2),
        "cogs": round(sum(_safe_float(item.get("cogs")) or 0.0 for item in year_quarters), 2),
        "gross_profit": round(sum(_safe_float(item.get("gross_profit")) or 0.0 for item in year_quarters), 2),
        "marketing": round(sum(_safe_float(item.get("marketing")) or 0.0 for item in year_quarters), 2),
        "payroll": round(sum(_safe_float(item.get("payroll")) or 0.0 for item in year_quarters), 2),
        "opex": round(sum(_safe_float(item.get("opex")) or 0.0 for item in year_quarters), 2),
        "ebitda": round(sum(_safe_float(item.get("ebitda")) or 0.0 for item in year_quarters), 2),
        "interest": round(sum(_safe_float(item.get("interest")) or 0.0 for item in year_quarters), 2),
        "depreciation": round(sum(_safe_float(item.get("depreciation")) or 0.0 for item in year_quarters), 2),
        "taxes": round(sum(_safe_float(item.get("taxes")) or 0.0 for item in year_quarters), 2),
        "net_income": round(sum(_safe_float(item.get("net_income")) or 0.0 for item in year_quarters), 2),
        "cash": _safe_float((year_quarters[-1] or {}).get("cash")) or 0.0,
        "ending_cash": _safe_float((year_quarters[-1] or {}).get("ending_cash")) or 0.0,
        "source_level": "finmo",
      }
    )

  return {
    "quarter_driver_path": quarter_view,
    "forecast_years": forecast_years,
  }


def _slot_quarters(period_count: int, forecast_quarters: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
  quarters = [item for item in (forecast_quarters or []) if isinstance(item, dict)]
  if period_count <= 0:
    return []
  if not quarters:
    return [{} for _ in range(period_count)]
  slots: List[Dict[str, Any]] = [_clone(quarters[0])]
  for idx in range(1, period_count):
    source_idx = min(idx - 1, len(quarters) - 1)
    slots.append(_clone(quarters[source_idx]))
  return slots[:period_count]


def _full_quarter_slots(slots: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
  full_slots = [
    _clone(slot) for slot in (slots or [])
    if isinstance(slot, dict) and _safe_float(slot.get("quarter")) not in (None, 0.0)
  ]
  return full_slots if full_slots else [_clone(slot) for slot in (slots or []) if isinstance(slot, dict)]


def _planned_quarter_slots(period_count: int, forecast_quarters: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
  quarters = [item for item in (forecast_quarters or []) if isinstance(item, dict)]
  if period_count <= 0:
    return []
  if not quarters:
    return [{} for _ in range(period_count)]
  slots: List[Dict[str, Any]] = []
  for idx in range(period_count):
    source_idx = min(idx, len(quarters) - 1)
    slots.append(_clone(quarters[source_idx]))
  return slots


def _ops_revenue_catalog(ops_json: Dict[str, Any]) -> Dict[Tuple[int, int], Dict[str, Any]]:
  ops = ops_json if isinstance(ops_json, dict) else {}
  catalog: Dict[Tuple[int, int], Dict[str, Any]] = {}
  lob_models = ops.get("lob_models") if isinstance(ops.get("lob_models"), list) else []
  if lob_models:
    for lob_idx, lob in enumerate(lob_models):
      if not isinstance(lob, dict):
        continue
      lob_name = _named_lob(
        lob.get("lob_name") or lob.get("name") or lob.get("line_of_business_name") or lob.get("label"),
        f"LOB {lob_idx + 1}",
      )
      products = lob.get("products") if isinstance(lob.get("products"), list) else []
      for product_idx, product in enumerate(products):
        if not isinstance(product, dict):
          continue
        product_name = _named_product(
          product.get("product_name") or product.get("name") or product.get("unit_name"),
          f"Product {product_idx + 1}",
        )
        catalog[(lob_idx, product_idx)] = {
          "revenue_slot_key": _revenue_slot_key(lob_idx, product_idx),
          "lob_slot_index": lob_idx,
          "product_slot_index": product_idx,
          "lob": lob_name,
          "product": product_name,
          "capacity": _quarter_capacity_from_ops_product(product=product, ops_json=ops),
          "unit_price": _safe_float(product.get("unit_price") if product.get("unit_price") is not None else ops.get("unit_price")),
          "utilization": _safe_ratio(product.get("utilization_rate") if product.get("utilization_rate") is not None else ops.get("utilization_rate")),
        }
  if catalog:
    return catalog
  catalog[(0, 0)] = {
    "revenue_slot_key": _revenue_slot_key(0, 0),
    "lob_slot_index": 0,
    "product_slot_index": 0,
    "lob": _named_lob(ops.get("business_type") or ops.get("lob_name"), "LOB 1"),
    "product": _named_product(ops.get("product_name") or ops.get("unit_name"), "Product 1"),
    "capacity": _quarter_capacity_from_ops_product(product=ops, ops_json=ops),
    "unit_price": _safe_float(ops.get("unit_price")),
    "utilization": _safe_ratio(ops.get("utilization_rate")),
  }
  return catalog


def _infer_revenue_driver_map_from_forecast_quarter(
  *,
  quarter_slot: Dict[str, Any],
  ops_catalog: Dict[Tuple[int, int], Dict[str, Any]],
  revenue_slot_key: str,
  lob: str,
  product: str,
  fallback_detail: Optional[Dict[str, Any]],
) -> Dict[str, float]:
  slot = quarter_slot if isinstance(quarter_slot, dict) else {}
  target_revenue = max(0.0, _safe_float(slot.get("revenue")) or 0.0)
  slot_details = [item for item in (ops_catalog or {}).values() if isinstance(item, dict)]
  if not slot_details and isinstance(fallback_detail, dict):
    slot_details = [fallback_detail]
  if not slot_details:
    return {}

  normalized: List[Tuple[Dict[str, Any], float, float, float, float]] = []
  matched: Optional[Tuple[Dict[str, Any], float, float, float, float]] = None
  total_baseline_revenue = 0.0
  for detail in slot_details:
    capacity = max(0.0, _safe_float(detail.get("capacity")) or 0.0)
    price = max(0.0, _safe_float(detail.get("unit_price")) or 0.0)
    utilization = max(0.0, _safe_ratio(detail.get("utilization")) or 0.0)
    baseline_revenue = max(0.0, capacity * price * utilization)
    row = (detail, baseline_revenue, capacity, price, utilization)
    normalized.append(row)
    total_baseline_revenue += baseline_revenue
    if (
      str(detail.get("revenue_slot_key") or "").strip() == str(revenue_slot_key or "").strip()
      or (
        str(detail.get("lob") or "").strip() == str(lob or "").strip()
        and str(detail.get("product") or "").strip() == str(product or "").strip()
      )
    ):
      matched = row

  if matched is None:
    matched = normalized[0]
  matched_detail, matched_baseline_revenue, _matched_capacity, matched_price, matched_utilization = matched
  if total_baseline_revenue > 0:
    revenue_share = matched_baseline_revenue / total_baseline_revenue
  else:
    revenue_share = 1.0 / float(len(normalized))
  allocated_revenue = target_revenue * revenue_share
  effective_price = max(0.0, matched_price)
  effective_utilization = max(0.0, matched_utilization)
  if effective_price <= 0.0:
    effective_price = max(1.0, _safe_float(matched_detail.get("unit_price")) or 0.0)
  if effective_utilization <= 0.0:
    effective_utilization = max(1.0, _safe_ratio(matched_detail.get("utilization")) or 0.0)
  denominator = effective_price * effective_utilization
  inferred_capacity = (allocated_revenue / denominator) if denominator > 0 else 0.0
  return {
    "Capacity": round(max(0.0, inferred_capacity), 6),
    "Unit Price": round(max(0.0, effective_price), 6),
    "Utilization": round(max(0.0, effective_utilization), 6),
  }


def _quarter_capacity_from_ops_product(*, product: Dict[str, Any], ops_json: Dict[str, Any]) -> float:
  product_obj = product if isinstance(product, dict) else {}
  ops = ops_json if isinstance(ops_json, dict) else {}
  periods_per_year = _safe_float(
    product_obj.get("operating_periods_per_year")
    if product_obj.get("operating_periods_per_year") is not None else ops.get("operating_periods_per_year")
  )
  units_per_period = _safe_float(
    product_obj.get("units_per_period_capacity")
    if product_obj.get("units_per_period_capacity") is not None else ops.get("units_per_period_capacity")
  )
  if units_per_period is not None and periods_per_year not in (None, 0.0):
    return round((units_per_period * periods_per_year) / 4.0, 6)
  units_per_week = _safe_float(
    product_obj.get("units_per_week_capacity")
    if product_obj.get("units_per_week_capacity") is not None else ops.get("units_per_week_capacity")
  )
  if units_per_week is not None:
    return round(units_per_week * 13.0, 6)
  units_per_month = _safe_float(
    product_obj.get("units_per_month_capacity")
    if product_obj.get("units_per_month_capacity") is not None else ops.get("units_per_month_capacity")
  )
  if units_per_month is not None:
    return round(units_per_month * 3.0, 6)
  concurrent_units = _safe_float(
    product_obj.get("concurrent_capacity_units")
    if product_obj.get("concurrent_capacity_units") is not None else ops.get("concurrent_capacity_units")
  )
  return round(concurrent_units or 0.0, 6)


def _resolve_row_identity_from_catalog(
  *,
  row_lob: Any,
  row_product: Any,
  ops_catalog: Dict[Tuple[int, int], Dict[str, Any]],
  revenue_slot_key: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
  if str(revenue_slot_key or "").strip():
    for catalog_item in ops_catalog.values():
      if not isinstance(catalog_item, dict):
        continue
      if str(catalog_item.get("revenue_slot_key") or "").strip() == str(revenue_slot_key or "").strip():
        return _clone(catalog_item)
  slot_identity = _revenue_slot_identity(row_lob=row_lob, row_product=row_product)
  resolved = ops_catalog.get((int(slot_identity.get("lob_slot_index") or 0), int(slot_identity.get("product_slot_index") or 0)))
  return _clone(resolved) if isinstance(resolved, dict) else None


def _child_driver_map_for_quarter(quarter: Dict[str, Any]) -> Dict[Tuple[str, str], Dict[str, float]]:
  child_map: Dict[Tuple[str, str], Dict[str, float]] = {}
  lobs = quarter.get("lobs") if isinstance(quarter.get("lobs"), list) else []
  for lob_idx, lob in enumerate(lobs):
    if not isinstance(lob, dict):
      continue
    products = lob.get("products") if isinstance(lob.get("products"), list) else []
    for product_idx, product in enumerate(products):
      if not isinstance(product, dict):
        continue
      capacity = _safe_float(product.get("capacity_units"))
      utilization = _safe_float(product.get("utilization"))
      units = _safe_float(product.get("units"))
      price = _safe_float(product.get("price")) or _safe_float(product.get("unit_price"))
      if capacity is None and units is not None and utilization not in (None, 0.0):
        capacity = units / utilization
      driver_payload = {
        "Capacity": round(capacity or 0.0, 6),
        "Unit Price": round(price or 0.0, 6),
        "Utilization": round(utilization or 0.0, 6),
      }
      revenue_slot_key = str(product.get("revenue_slot_key") or lob.get("revenue_slot_key") or _revenue_slot_key(lob_idx, product_idx)).strip()
      if revenue_slot_key:
        child_map[("__slot__", revenue_slot_key)] = _clone(driver_payload)
      child_map[(f"LOB {lob_idx + 1}", f"Product {product_idx + 1}")] = driver_payload
      lob_name = _named_lob(lob.get("lob_name") or lob.get("name") or lob.get("label"), f"LOB {lob_idx + 1}")
      product_name = _named_product(product.get("product_name") or product.get("name") or product.get("unit_name"), f"Product {product_idx + 1}")
      child_map[(lob_name, product_name)] = _clone(driver_payload)
  if child_map:
    return child_map
  capacity = _safe_float(quarter.get("capacity_units"))
  utilization = _safe_float(quarter.get("utilization"))
  units = _safe_float(quarter.get("units"))
  price = _safe_float(quarter.get("price"))
  if capacity is None and units is not None and utilization not in (None, 0.0):
    capacity = units / utilization
  return {
    ("LOB 1", "Product 1"): {
      "Capacity": round(capacity or 0.0, 6),
      "Unit Price": round(price or 0.0, 6),
      "Utilization": round(utilization or 0.0, 6),
    }
  }


def _quarter_lease_amount(financials_json: Dict[str, Any]) -> float:
  monthly_rent = _safe_float((financials_json or {}).get("monthly_rent_expense")) or 0.0
  return round(monthly_rent * 3.0, 6)


def _add_months(base: datetime, months: int) -> datetime:
  month_index = (base.month - 1) + int(months or 0)
  year = base.year + (month_index // 12)
  month = (month_index % 12) + 1
  day = min(
    base.day,
    [
      31,
      29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28,
      31,
      30,
      31,
      30,
      31,
      31,
      30,
      31,
      30,
      31,
    ][month - 1],
  )
  return base.replace(year=year, month=month, day=day)


def _python_model_input_periods(*, start_date_iso: Optional[str], period_count: int = 20) -> List[Dict[str, Any]]:
  normalized_start = _as_iso_date(start_date_iso) or datetime.utcnow().date().isoformat()
  try:
    start_dt = datetime.fromisoformat(normalized_start)
  except Exception:
    start_dt = datetime.utcnow()
    normalized_start = start_dt.date().isoformat()
  slots: List[Dict[str, Any]] = []
  for slot_index in range(max(0, int(period_count or 0))):
    period_date = _add_months(start_dt, slot_index * 3)
    column_index = 8 + slot_index
    slots.append(
      {
        "slot_index": slot_index,
        "column_index": column_index,
        "column_letter": get_column_letter(column_index),
        "year": float(period_date.year),
        "quarter": float(slot_index + 1),
        "date": period_date.date().isoformat(),
        "year_fraction": 1.0,
        "is_stub": False,
      }
    )
  return slots


def _empty_controller_write_row(
  *,
  quarter_scope: Dict[str, Any],
  named_range: str,
  section_key: str,
  label: str,
) -> Dict[str, Any]:
  semantics = _simple_input_semantics(section_key, label)
  return {
    "named_range": named_range,
    "controller_write": True,
    "lever_id": _simple_lever_id(section_key, label),
    "label": label,
    **semantics,
    **quarter_scope,
    "values": [0.0 for _ in range(len(quarter_scope.get("valid_quarter_indices") or []))],
  }


def _python_model_input_template(
  *,
  start_date_iso: Optional[str],
  business_name: Optional[str],
  ops_json: Dict[str, Any],
) -> Dict[str, Any]:
  periods = _python_model_input_periods(start_date_iso=start_date_iso, period_count=20)
  quarter_scope = _full_quarter_scope(periods)
  revenue_catalog = _ops_revenue_catalog(ops_json or {})
  revenue_items = sorted(
    [item for item in revenue_catalog.values() if isinstance(item, dict)],
    key=lambda item: (
      _safe_int(item.get("lob_slot_index")),
      _safe_int(item.get("product_slot_index")),
      str(item.get("revenue_slot_key") or "").strip(),
    ),
  )
  controller_write_levers: List[Dict[str, Any]] = []
  lever_catalog: Dict[str, Dict[str, Any]] = {}

  def _register_lever(metadata: Dict[str, Any]) -> None:
    lever_id = str(metadata.get("lever_id") or "").strip()
    if not lever_id:
      return
    lever_meta = {
      **_clone(metadata),
      **quarter_scope,
    }
    lever_catalog[lever_id] = lever_meta
    controller_write_levers.append(_clone(lever_meta))

  revenue_rows: List[Dict[str, Any]] = []
  for item in revenue_items:
    lob_name = str(item.get("lob") or "").strip() or "LOB 1"
    product_name = str(item.get("product") or "").strip() or "Product 1"
    revenue_slot_key = str(item.get("revenue_slot_key") or "").strip()
    base_meta = {
      "named_range": "model_input_revenue",
      "controller_write": True,
      "placeholder_lob": lob_name,
      "placeholder_product": product_name,
      "lob_slot_index": _safe_int(item.get("lob_slot_index")),
      "product_slot_index": _safe_int(item.get("product_slot_index")),
      "revenue_slot_key": revenue_slot_key,
      "lob": lob_name,
      "product": product_name,
      **quarter_scope,
      "values": [0.0 for _ in range(len(quarter_scope.get("valid_quarter_indices") or []))],
    }
    for driver_name in ("Capacity", "Unit Price", "Utilization"):
      semantics = _revenue_input_semantics(driver_name)
      lever_id = _revenue_lever_id(lob_name, product_name, driver_name)
      revenue_rows.append(
        {
          **_clone(base_meta),
          "lever_id": lever_id,
          "driver": driver_name,
          **semantics,
        }
      )
      _register_lever(
        {
          "lever_id": lever_id,
          "named_range": "model_input_revenue",
          "section": "revenue",
          "lob": lob_name,
          "product": product_name,
          "driver": driver_name,
          "placeholder_lob": lob_name,
          "placeholder_product": product_name,
          "lob_slot_index": _safe_int(item.get("lob_slot_index")),
          "product_slot_index": _safe_int(item.get("product_slot_index")),
          "revenue_slot_key": revenue_slot_key,
          "label_path": f"{lob_name} > {product_name} > {driver_name}",
          **semantics,
        }
      )

  expenses = [
    "Cost of Goods Sold",
    "Marketing",
    "Research & Development",
    "Lease",
    "Payroll",
    "General & Administrative",
    "Interest Rate",
    "Depreciation",
    "Taxes",
  ]
  expense_rows = [
    _empty_controller_write_row(
      quarter_scope=quarter_scope,
      named_range="model_input_expenses",
      section_key="expenses",
      label=label,
    )
    for label in expenses
  ]
  for row in expense_rows:
    _register_lever(
      {
        "lever_id": str(row.get("lever_id") or "").strip(),
        "named_range": "model_input_expenses",
        "section": "expenses",
        "label": str(row.get("label") or "").strip(),
        "label_path": str(row.get("label") or "").strip(),
        "value_kind": str(row.get("value_kind") or "").strip(),
        "input_semantics": str(row.get("input_semantics") or "").strip(),
      }
    )

  balance_sheet = [
    "Accounts Receivable Days",
    "Inventory Days",
    "PPE $ (Excluding Capital Leases)",
    "Accumulated Depreciation",
    "Accounts Payable Days",
    "Prepaid Expenses",
    "Deferred Revnue",
    "Short Term Debt (% of LTD)",
    "Owner's Capital",
    "Other Equity",
  ]
  balance_rows = [
    _empty_controller_write_row(
      quarter_scope=quarter_scope,
      named_range="model_input_balancehseet",
      section_key="balance_sheet",
      label=label,
    )
    for label in balance_sheet
  ]
  for row in balance_rows:
    _register_lever(
      {
        "lever_id": str(row.get("lever_id") or "").strip(),
        "named_range": "model_input_balancehseet",
        "section": "balance_sheet",
        "label": str(row.get("label") or "").strip(),
        "label_path": str(row.get("label") or "").strip(),
        "value_kind": str(row.get("value_kind") or "").strip(),
        "input_semantics": str(row.get("input_semantics") or "").strip(),
      }
    )

  schedule_labels = [
    "Plus: Additions (repayments), net",
    "Less: Principal Repayments",
    "Plus: Net Additions",
  ]
  schedule_rows = [
    _empty_controller_write_row(
      quarter_scope=quarter_scope,
      named_range="model_input_schedules",
      section_key="schedules",
      label=label,
    )
    for label in schedule_labels
  ]
  for row in schedule_rows:
    _register_lever(
      {
        "lever_id": str(row.get("lever_id") or "").strip(),
        "named_range": "model_input_schedules",
        "section": "schedules",
        "label": str(row.get("label") or "").strip(),
        "label_path": str(row.get("label") or "").strip(),
        "value_kind": str(row.get("value_kind") or "").strip(),
        "input_semantics": str(row.get("input_semantics") or "").strip(),
      }
    )

  return {
    "contract_version": "finmo_model_input_v3",
    "canonical_lever_vocabulary": "model_inputs_controller_write_only",
    "finmo_path": "",
    "business_name": str(business_name or "").strip(),
    "start_date": _as_iso_date(start_date_iso) or datetime.utcnow().date().isoformat(),
    "periods": periods,
    "lever_catalog": lever_catalog,
    "controller_write_levers": controller_write_levers,
    "sections": {
      "revenue": revenue_rows,
      "expenses": expense_rows,
      "balance_sheet": balance_rows,
      "schedules": {
        "debt_opening_balance_seed": 0.0,
        "lease_opening_balance_seed": 0.0,
        "rows": schedule_rows,
      },
    },
  }


def build_python_model_input_json(
  *,
  business_facts: Optional[Dict[str, Any]],
  ops_json: Optional[Dict[str, Any]],
  people_json: Optional[Dict[str, Any]],
  financials_json: Optional[Dict[str, Any]],
  financials_year1_json: Optional[Dict[str, Any]],
  marketing_model_json: Optional[Dict[str, Any]],
  controller_input_seed: Optional[Sequence[Dict[str, Any]]] = None,
  forecast_quarters: Sequence[Dict[str, Any]] = (),
  business_name: Optional[str] = None,
) -> Dict[str, Any]:
  start_date = _as_iso_date((business_facts or {}).get("start_date")) or _as_iso_date((ops_json or {}).get("start_date"))
  baseline_model_input = _python_model_input_template(
    start_date_iso=start_date,
    business_name=business_name or (business_facts or {}).get("business_name") or (ops_json or {}).get("business_name"),
    ops_json=ops_json or {},
  )
  return _build_model_input_overlay(
    baseline_model_input=baseline_model_input,
    business_facts=business_facts or {},
    ops_json=ops_json or {},
    people_json=people_json or {},
    financials_json=financials_json or {},
    financials_year1_json=financials_year1_json or {},
    marketing_model_json=marketing_model_json or {},
    controller_input_seed=controller_input_seed,
    forecast_quarters=forecast_quarters,
  )


def _build_model_input_overlay(
  *,
  baseline_model_input: Dict[str, Any],
  business_facts: Dict[str, Any],
  ops_json: Dict[str, Any],
  people_json: Dict[str, Any],
  financials_json: Dict[str, Any],
  financials_year1_json: Dict[str, Any],
  marketing_model_json: Dict[str, Any],
  controller_input_seed: Optional[Sequence[Dict[str, Any]]] = None,
  forecast_quarters: Sequence[Dict[str, Any]],
) -> Dict[str, Any]:
  next_payload = _clone(baseline_model_input if isinstance(baseline_model_input, dict) else {})
  periods = [item for item in (next_payload.get("periods") or []) if isinstance(item, dict)]
  full_periods = _full_quarter_slots(periods)
  seed_slots = [item for item in (controller_input_seed or []) if isinstance(item, dict)]
  target_period_count = len(full_periods) or len(periods)
  if seed_slots:
    slots = [_clone(item) for item in seed_slots[:target_period_count]]
    if len(slots) < target_period_count:
      fallback_slots = _planned_quarter_slots(target_period_count, forecast_quarters)
      slots.extend(_clone(fallback_slots[idx]) for idx in range(len(slots), target_period_count))
  else:
    slots = _planned_quarter_slots(target_period_count, forecast_quarters)
  projection_mode = bool(seed_slots) or bool([item for item in (forecast_quarters or []) if isinstance(item, dict)])
  start_date = _as_iso_date((business_facts or {}).get("start_date"))
  if start_date:
    next_payload["start_date"] = start_date

  sections = next_payload.setdefault("sections", {})
  revenue_rows = [row for row in (sections.get("revenue") or []) if isinstance(row, dict)]
  ops_catalog = _ops_revenue_catalog(ops_json or {})
  quarter_child_maps = [
    _child_driver_map_for_quarter({"lobs": (slot.get("revenue_products") or []) if isinstance(slot, dict) else []})
    if seed_slots else
    _child_driver_map_for_quarter(slot if isinstance(slot, dict) else {})
    for slot in slots
  ]
  for row in revenue_rows:
    revenue_slot_key = str(row.get("revenue_slot_key") or "").strip()
    placeholder_lob = str(row.get("placeholder_lob") or row.get("lob") or "").strip()
    placeholder_product = str(row.get("placeholder_product") or row.get("product") or "").strip()
    resolved_identity = _resolve_row_identity_from_catalog(
      row_lob=placeholder_lob,
      row_product=placeholder_product,
      ops_catalog=ops_catalog,
      revenue_slot_key=revenue_slot_key,
    )
    if isinstance(resolved_identity, dict):
      row["lob"] = resolved_identity.get("lob") or row.get("lob")
      row["product"] = resolved_identity.get("product") or row.get("product")
      row["revenue_slot_key"] = resolved_identity.get("revenue_slot_key") or row.get("revenue_slot_key")
      row["lob_slot_index"] = resolved_identity.get("lob_slot_index") if resolved_identity.get("lob_slot_index") is not None else row.get("lob_slot_index")
      row["product_slot_index"] = resolved_identity.get("product_slot_index") if resolved_identity.get("product_slot_index") is not None else row.get("product_slot_index")
    lob = str(row.get("lob") or placeholder_lob).strip()
    product = str(row.get("product") or placeholder_product).strip()
    driver = str(row.get("driver") or "").strip()
    values: List[float] = []
    if projection_mode:
      for slot_idx, child_map in enumerate(quarter_child_maps):
        driver_map = (
          child_map.get(("__slot__", str(row.get("revenue_slot_key") or "").strip()))
          or
          child_map.get((lob, product))
          or child_map.get((placeholder_lob, placeholder_product))
          or {}
        )
        if not driver_map:
          driver_map = _infer_revenue_driver_map_from_forecast_quarter(
            quarter_slot=(slots[slot_idx] if slot_idx < len(slots) and isinstance(slots[slot_idx], dict) else {}),
            ops_catalog=ops_catalog,
            revenue_slot_key=str(row.get("revenue_slot_key") or "").strip(),
            lob=lob,
            product=product,
            fallback_detail=resolved_identity if isinstance(resolved_identity, dict) else None,
          )
        values.append(round(_safe_float(driver_map.get(driver)) or 0.0, 6))
    else:
      baseline_driver_map = (
        (resolved_identity or {})
        if isinstance(resolved_identity, dict) else {}
      )
      baseline_value = 0.0
      if driver == "Capacity":
        baseline_value = round(_safe_float(baseline_driver_map.get("capacity")) or 0.0, 6)
      elif driver == "Unit Price":
        baseline_value = round(_safe_float(baseline_driver_map.get("unit_price")) or 0.0, 6)
      elif driver == "Utilization":
        baseline_value = round(_safe_ratio(baseline_driver_map.get("utilization")) or 0.0, 6)
      values = [baseline_value for _ in slots]
    row["values"] = values

  lease_amount = _quarter_lease_amount(financials_json or {})
  revenue_total_year1 = max(
    0.0,
    _safe_float((financials_year1_json or {}).get("company_revenue_total_year1"))
    or _safe_float((financials_year1_json or {}).get("revenue_total_year1"))
    or _safe_float((financials_json or {}).get("current_revenue"))
    or 0.0,
  )
  cogs_ratio_baseline = _ratio((financials_json or {}).get("cogs_total_year1"), revenue_total_year1)
  marketing_ratio_baseline = (
    _safe_ratio((marketing_model_json or {}).get("marketing_percent_of_revenue"))
    if isinstance(marketing_model_json, dict) else None
  )
  if marketing_ratio_baseline is None:
    marketing_ratio_baseline = _safe_ratio((financials_json or {}).get("marketing_percent_of_revenue"))
  payroll_total_year1 = (
    _safe_float((financials_json or {}).get("payroll_total_year1"))
    or _safe_float((financials_json or {}).get("current_payroll"))
    or 0.0
  )
  if payroll_total_year1 <= 0:
    payroll_total_year1 = sum(
      max(0.0, _safe_float(item.get("annual_wage")) or 0.0)
      for item in ((people_json or {}).get("people") or [])
      if isinstance(item, dict)
    )
  quarterly_payroll = round(max(0.0, payroll_total_year1) / 4.0, 6) if payroll_total_year1 else 0.0
  non_rent_opex_year1 = max(
    0.0,
    (
      _safe_float((financials_json or {}).get("other_opex_absolute"))
      or _safe_float((financials_json or {}).get("other_operating_expense"))
      or 0.0
    ) - (lease_amount * 4.0)
  )
  g_and_a_ratio_baseline = _ratio(non_rent_opex_year1, revenue_total_year1)
  interest_rate_baseline = _ratio((financials_json or {}).get("annual_interest_payment"), (financials_json or {}).get("total_debt_outstanding"))
  depreciation_ratio_baseline = _ratio((financials_json or {}).get("accumulated_depreciation"), revenue_total_year1)
  expense_rows = [row for row in (sections.get("expenses") or []) if isinstance(row, dict)]
  for row in expense_rows:
    label = str(row.get("label") or "").strip()
    values: List[float] = []
    for slot in slots:
      revenue = _safe_float(slot.get("revenue")) or 0.0
      if seed_slots and label == "Cost of Goods Sold":
        values.append(round(_safe_float(slot.get("cogs_percent")) or 0.0, 6))
      elif seed_slots and label == "Marketing":
        values.append(round(_safe_float(slot.get("marketing_percent")) or 0.0, 6))
      elif seed_slots and label == "Research & Development":
        values.append(round(_safe_float(slot.get("r_and_d_percent")) or 0.0, 6))
      elif seed_slots and label == "Lease":
        values.append(round(_safe_float(slot.get("lease_amount")) or 0.0, 6))
      elif seed_slots and label == "Payroll":
        values.append(round(_safe_float(slot.get("payroll_amount")) or 0.0, 6))
      elif seed_slots and label == "General & Administrative":
        values.append(round(_safe_float(slot.get("g_and_a_percent")) or 0.0, 6))
      elif seed_slots and label == "Interest Rate":
        values.append(round(_safe_float(slot.get("interest_rate")) or 0.0, 6))
      elif seed_slots and label == "Depreciation":
        values.append(round(_safe_float(slot.get("depreciation_percent")) or 0.0, 6))
      elif seed_slots and label == "Taxes":
        values.append(round(_safe_float(slot.get("tax_percent")) or 0.0, 6))
      elif label == "Cost of Goods Sold":
        values.append(round(cogs_ratio_baseline if not projection_mode else _ratio(slot.get("cogs"), revenue), 6))
      elif label == "Marketing":
        values.append(round((marketing_ratio_baseline or 0.0) if not projection_mode else _ratio(slot.get("marketing"), revenue), 6))
      elif label == "Research & Development":
        values.append(0.0)
      elif label == "Lease":
        values.append(round(lease_amount, 6))
      elif label == "Payroll":
        values.append(round(quarterly_payroll if not projection_mode else (_safe_float(slot.get("payroll")) or 0.0), 6))
      elif label == "General & Administrative":
        values.append(round(max(0.0, g_and_a_ratio_baseline if not projection_mode else _ratio((_safe_float(slot.get("opex")) or 0.0) - lease_amount, revenue)), 6))
      elif label == "Interest Rate":
        values.append(round(interest_rate_baseline, 6))
      elif label == "Depreciation":
        values.append(round(depreciation_ratio_baseline if not projection_mode else _ratio(slot.get("depreciation"), revenue), 6))
      elif label == "Taxes":
        values.append(round(_safe_ratio((financials_json or {}).get("taxes_percent")) or (_ratio(slot.get("taxes"), revenue) if projection_mode else 0.0), 6))
      else:
        values.append(round(_safe_float((row.get("values") or [0.0])[0]) or 0.0, 6))
    row["values"] = values

  balance_rows = [row for row in (sections.get("balance_sheet") or []) if isinstance(row, dict)]
  opening_ppe = _safe_float((financials_json or {}).get("current_capex"))
  if opening_ppe is None:
    opening_ppe = _safe_float((financials_json or {}).get("initial_assets")) or 0.0
  opening_accum_dep = _safe_float((financials_json or {}).get("accumulated_depreciation")) or 0.0
  cumulative_capex = 0.0
  cumulative_dep = 0.0
  for row in balance_rows:
    label = str(row.get("label") or "").strip()
    values: List[float] = []
    cumulative_capex = 0.0
    cumulative_dep = 0.0
    base_values = list(row.get("values") or [])
    for slot_idx, slot in enumerate(slots):
      working_capital = slot.get("working_capital") if isinstance(slot.get("working_capital"), dict) else {}
      cumulative_capex += _safe_float(slot.get("capex")) or 0.0
      cumulative_dep += _safe_float(slot.get("depreciation")) or 0.0
      if label == "Accounts Receivable Days":
        values.append(round(_safe_float(working_capital.get("dso")) or 0.0, 6))
      elif label == "Inventory Days":
        values.append(round(_safe_float(working_capital.get("inventory_days")) or 0.0, 6))
      elif label == "PPE $ (Excluding Capital Leases)":
        values.append(round(opening_ppe + cumulative_capex, 6))
      elif label == "Accumulated Depreciation":
        values.append(round(opening_accum_dep - cumulative_dep, 6))
      elif label == "Accounts Payable Days":
        values.append(round(_safe_float(working_capital.get("dpo")) or 0.0, 6))
      elif label == "Prepaid Expenses":
        values.append(round(_safe_float(base_values[min(slot_idx, len(base_values) - 1)]) or 0.0, 6) if base_values else 0.0)
      elif label == "Deferred Revnue":
        values.append(round(_safe_float(base_values[min(slot_idx, len(base_values) - 1)]) or 0.0, 6) if base_values else 0.0)
      elif label == "Short Term Debt (% of LTD)":
        short_term_ratio = _ratio((financials_json or {}).get("short_term_debt"), (financials_json or {}).get("total_debt_outstanding"))
        values.append(round(short_term_ratio, 6))
      elif label == "Owner's Capital":
        values.append(round(_safe_float((financials_json or {}).get("initial_equity")) or (_safe_float(base_values[min(slot_idx, len(base_values) - 1)]) or 0.0), 6) if base_values else round(_safe_float((financials_json or {}).get("initial_equity")) or 0.0, 6))
      elif label == "Other Equity":
        values.append(round(_safe_float(base_values[min(slot_idx, len(base_values) - 1)]) or 0.0, 6) if base_values else 0.0)
      else:
        values.append(round(_safe_float(base_values[min(slot_idx, len(base_values) - 1)]) or 0.0, 6) if base_values else 0.0)
    row["values"] = values

  schedules = sections.get("schedules") if isinstance(sections.get("schedules"), dict) else {}
  debt_seed = _safe_float((financials_json or {}).get("total_debt_outstanding"))
  if debt_seed is not None:
    schedules["debt_opening_balance_seed"] = round(debt_seed, 6)
  lease_seed = _safe_float((financials_json or {}).get("initial_lease"))
  if lease_seed is not None:
    schedules["lease_opening_balance_seed"] = round(lease_seed, 6)
  for row in [item for item in (schedules.get("rows") or []) if isinstance(item, dict)]:
    label = str(row.get("label") or "").strip()
    base_values = list(row.get("values") or [])
    if label == "Plus: Additions (repayments), net":
      row["values"] = [0.0 for _ in slots]
    elif label == "Less: Principal Repayments":
      annual_principal = _safe_float((financials_json or {}).get("annual_principal_payment")) or 0.0
      quarterly = round(-(annual_principal / 4.0), 6) if annual_principal else 0.0
      row["values"] = [quarterly for _ in slots]
    elif label == "Plus: Net Additions":
      row["values"] = [0.0 for _ in slots]
    else:
      row["values"] = [round(_safe_float(base_values[min(idx, len(base_values) - 1)]) or 0.0, 6) if base_values else 0.0 for idx, _slot in enumerate(slots)]
  sections["schedules"] = schedules
  return next_payload


def _write_model_input_json_to_workbook(finmo_path: str, model_input_json: Dict[str, Any]) -> None:
  wb = load_workbook(finmo_path, data_only=False)
  try:
    period_slots = _period_slots_from_model_inputs(wb)
    period_columns = [int(slot["column_index"]) for slot in period_slots]
    full_period_columns = [int(slot["column_index"]) for slot in _full_quarter_slots(period_slots)]

    def _target_period_columns(values: Sequence[Any]) -> List[int]:
      if len(values) == len(full_period_columns) and full_period_columns:
        return list(full_period_columns)
      return list(period_columns)

    if isinstance(model_input_json.get("start_date"), str) and str(model_input_json.get("start_date")).strip():
      _write_named_cell(wb, "model_input_startdate", str(model_input_json.get("start_date")).strip())

    revenue_section = [row for row in (((model_input_json.get("sections") or {}) if isinstance(model_input_json.get("sections"), dict) else {}).get("revenue") or []) if isinstance(row, dict)]
    ws, min_row, max_row, min_col, max_col = _defined_range_bounds(wb, "model_input_revenue")
    revenue_row_lookup: Dict[Tuple[str, str, str], int] = {}
    revenue_slot_lookup: Dict[Tuple[str, str], int] = {}
    revenue_row_ordinal = 0
    for row_idx in range(min_row, max_row + 1):
      marker = str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower()
      if marker != "controller write":
        continue
      row_lob = str(ws.cell(row=row_idx, column=min_col + 2).value or "").strip()
      row_product = str(ws.cell(row=row_idx, column=min_col + 3).value or "").strip()
      row_driver = str(ws.cell(row=row_idx, column=min_col + 4).value or "").strip()
      slot_identity = _revenue_slot_identity(
        row_lob=row_lob,
        row_product=row_product,
        revenue_row_ordinal=revenue_row_ordinal,
      )
      revenue_row_lookup[
        (
          row_lob,
          row_product,
          row_driver,
        )
      ] = row_idx
      revenue_slot_lookup[(str(slot_identity.get("revenue_slot_key") or "").strip(), row_driver)] = row_idx
      revenue_row_ordinal += 1
    for row in revenue_section:
      driver = str(row.get("driver") or "").strip()
      row_idx = revenue_slot_lookup.get((str(row.get("revenue_slot_key") or "").strip(), driver))
      if row_idx is None:
        key = (str(row.get("lob") or "").strip(), str(row.get("product") or "").strip(), driver)
        row_idx = revenue_row_lookup.get(key)
      if row_idx is None:
        fallback_key = (
          str(row.get("placeholder_lob") or "").strip(),
          str(row.get("placeholder_product") or "").strip(),
          driver,
        )
        row_idx = revenue_row_lookup.get(fallback_key)
      if row_idx is None:
        continue
      if str(row.get("lob") or "").strip():
        ws.cell(row=row_idx, column=min_col + 2).value = str(row.get("lob") or "").strip()
      if str(row.get("product") or "").strip():
        ws.cell(row=row_idx, column=min_col + 3).value = str(row.get("product") or "").strip()
      values = list(row.get("values") or [])
      for idx, col_idx in enumerate(_target_period_columns(values)):
        if col_idx > max_col:
          continue
        ws.cell(row=row_idx, column=col_idx).value = values[idx] if idx < len(values) else 0

    def _write_simple_rows(range_name: str, rows_payload: Sequence[Dict[str, Any]]) -> None:
      local_ws, local_min_row, local_max_row, local_min_col, local_max_col = _defined_range_bounds(wb, range_name)
      label_lookup: Dict[str, int] = {}
      for row_idx in range(local_min_row, local_max_row + 1):
        marker = str(local_ws.cell(row=row_idx, column=local_min_col).value or "").strip().lower()
        if marker != "controller write":
          continue
        label_lookup[str(local_ws.cell(row=row_idx, column=local_min_col + 1).value or "").strip()] = row_idx
      for row in rows_payload:
        label = str(row.get("label") or "").strip()
        row_idx = label_lookup.get(label)
        if row_idx is None:
          continue
        values = list(row.get("values") or [])
        for idx, col_idx in enumerate(_target_period_columns(values)):
          if col_idx > local_max_col:
            continue
          local_ws.cell(row=row_idx, column=col_idx).value = values[idx] if idx < len(values) else 0

    sections = model_input_json.get("sections") if isinstance(model_input_json.get("sections"), dict) else {}
    _write_simple_rows("model_input_expenses", [row for row in (sections.get("expenses") or []) if isinstance(row, dict)])
    _write_simple_rows("model_input_balancehseet", [row for row in (sections.get("balance_sheet") or []) if isinstance(row, dict)])

    schedule_section = sections.get("schedules") if isinstance(sections.get("schedules"), dict) else {}
    schedule_ws, sched_min_row, sched_max_row, sched_min_col, sched_max_col = _defined_range_bounds(wb, "model_input_schedules")
    label_lookup: Dict[str, int] = {}
    for row_idx in range(sched_min_row, sched_max_row + 1):
      label = str(schedule_ws.cell(row=row_idx, column=sched_min_col + 1).value or "").strip()
      if label:
        label_lookup[label] = row_idx
    seed_col = period_columns[0] - 1 if period_columns else sched_min_col + 2
    debt_seed_row = label_lookup.get("Closing Balance")
    if debt_seed_row is not None and seed_col <= sched_max_col:
      debt_seed = schedule_section.get("debt_opening_balance_seed")
      if debt_seed is not None:
        schedule_ws.cell(row=debt_seed_row, column=seed_col).value = debt_seed
    lease_seed_row = label_lookup.get("Closing Balance (Total)")
    if lease_seed_row is not None and seed_col <= sched_max_col:
      lease_seed = schedule_section.get("lease_opening_balance_seed")
      if lease_seed is not None:
        schedule_ws.cell(row=lease_seed_row, column=seed_col).value = lease_seed
    for row in [item for item in (schedule_section.get("rows") or []) if isinstance(item, dict)]:
      label = str(row.get("label") or "").strip()
      row_idx = label_lookup.get(label)
      if row_idx is None:
        continue
      values = list(row.get("values") or [])
      for idx, col_idx in enumerate(period_columns):
        if col_idx > sched_max_col:
          continue
        schedule_ws.cell(row=row_idx, column=col_idx).value = values[idx] if idx < len(values) else 0

    wb.save(finmo_path)
  finally:
    wb.close()


def _resolve_finmo_calibration_shell(
  *,
  finmo_path: str,
  model_input_json: Dict[str, Any],
  finmo_json: Dict[str, Any],
  calibration_spec: Dict[str, Any],
) -> Dict[str, Any]:
  wb = load_workbook(finmo_path, data_only=False)
  try:
    model_slots = [item for item in (model_input_json.get("periods") or []) if isinstance(item, dict)]
    finmo_slots = [item for item in (finmo_json.get("periods") or []) if isinstance(item, dict)]

    def _resolve_model_input_cell(spec: Dict[str, Any]) -> Dict[str, Any]:
      section = str(spec.get("section") or "").strip().lower()
      lever_id = str(spec.get("lever_id") or "").strip()
      quarter_index = _safe_int(spec.get("quarter_index")) or 1
      slot = _model_input_slot_for_full_quarter(model_slots, quarter_index)
      if slot is None:
        return {}
      if section == "revenue":
        ws, min_row, max_row, min_col, _max_col = _defined_range_bounds(wb, "model_input_revenue")
        revenue_row_ordinal = 0
        for row_idx in range(min_row, max_row + 1):
          if str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower() != "controller write":
            continue
          row_lob = str(ws.cell(row=row_idx, column=min_col + 2).value or "").strip()
          row_product = str(ws.cell(row=row_idx, column=min_col + 3).value or "").strip()
          row_driver = str(ws.cell(row=row_idx, column=min_col + 4).value or "").strip()
          row_lever_id = _revenue_lever_id(row_lob, row_product, row_driver)
          row_slot_identity = _revenue_slot_identity(
            row_lob=row_lob,
            row_product=row_product,
            revenue_row_ordinal=revenue_row_ordinal,
          )
          revenue_row_ordinal += 1
          if (
            (str(spec.get("revenue_slot_key") or "").strip() and str(row_slot_identity.get("revenue_slot_key") or "").strip() == str(spec.get("revenue_slot_key") or "").strip() and row_driver == str(spec.get("driver") or "").strip())
            or
            (lever_id and row_lever_id == lever_id)
            or (
              not lever_id
              and row_lob == str(spec.get("lob") or "").strip()
              and row_product == str(spec.get("product") or "").strip()
              and row_driver == str(spec.get("driver") or "").strip()
            )
          ):
            col_idx = int(slot["column_index"])
            return {"sheet": ws.title, "cell": f"{get_column_letter(col_idx)}{row_idx}"}
      if section == "expenses":
        ws, min_row, max_row, min_col, _max_col = _defined_range_bounds(wb, "model_input_expenses")
        for row_idx in range(min_row, max_row + 1):
          if str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower() != "controller write":
            continue
          label = str(ws.cell(row=row_idx, column=min_col + 1).value or "").strip()
          row_lever_id = _simple_lever_id("expenses", label)
          if (lever_id and row_lever_id == lever_id) or (not lever_id and label == str(spec.get("label") or "").strip()):
            col_idx = int(slot["column_index"])
            return {"sheet": ws.title, "cell": f"{get_column_letter(col_idx)}{row_idx}"}
      if section == "balance_sheet":
        ws, min_row, max_row, min_col, _max_col = _defined_range_bounds(wb, "model_input_balancehseet")
        for row_idx in range(min_row, max_row + 1):
          if str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower() != "controller write":
            continue
          label = str(ws.cell(row=row_idx, column=min_col + 1).value or "").strip()
          row_lever_id = _simple_lever_id("balance_sheet", label)
          if (lever_id and row_lever_id == lever_id) or (not lever_id and label == str(spec.get("label") or "").strip()):
            col_idx = int(slot["column_index"])
            return {"sheet": ws.title, "cell": f"{get_column_letter(col_idx)}{row_idx}"}
      if section == "schedules":
        ws, min_row, max_row, min_col, _max_col = _defined_range_bounds(wb, "model_input_schedules")
        for row_idx in range(min_row, max_row + 1):
          if str(ws.cell(row=row_idx, column=min_col).value or "").strip().lower() != "controller write":
            continue
          label = str(ws.cell(row=row_idx, column=min_col + 1).value or "").strip()
          row_lever_id = _simple_lever_id("schedules", label)
          if (lever_id and row_lever_id == lever_id) or (not lever_id and label == str(spec.get("label") or "").strip()):
            col_idx = int(slot["column_index"])
            return {"sheet": ws.title, "cell": f"{get_column_letter(col_idx)}{row_idx}"}
      return {}

    def _resolve_finmo_output_cell(objective: Dict[str, Any]) -> Dict[str, Any]:
      range_name = str(objective.get("sheet_range") or "").strip()
      if not range_name:
        return {}
      ws, min_row, max_row, min_col, _max_col = _defined_range_bounds(wb, range_name)
      slot = _finmo_slot_for_full_quarter(finmo_slots, _safe_int(objective.get("quarter_index")) or 1)
      if slot is None:
        return {}
      for row_idx in range(min_row, max_row + 1):
        if str(ws.cell(row=row_idx, column=min_col).value or "").strip() == str(objective.get("line_item") or "").strip():
          col_idx = int(slot["column_index"])
          return {"sheet": ws.title, "cell": f"{get_column_letter(col_idx)}{row_idx}"}
      return {}

    next_shell = _clone(calibration_spec if isinstance(calibration_spec, dict) else {})
    for request in [item for item in (next_shell.get("goal_seek_requests") or []) if isinstance(item, dict)]:
      request["objective_cell"] = _resolve_finmo_output_cell((request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {})
      request["changing_input_cells"] = [
        _resolve_model_input_cell(item)
        for item in (request.get("changing_inputs") or [])
        if isinstance(item, dict)
      ]
    for request in [item for item in (next_shell.get("solver_requests") or []) if isinstance(item, dict)]:
      request["objective_cell"] = _resolve_finmo_output_cell((request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {})
      resolved_inputs = []
      resolved_constraints = [item for item in (request.get("constraints") or []) if isinstance(item, dict)]
      grouped_anchor_cells: Dict[str, str] = {}
      for item in (request.get("changing_inputs") or []):
        if not isinstance(item, dict):
          continue
        resolved = _resolve_model_input_cell(item)
        if not resolved:
          continue
        resolved_inputs.append(resolved)
        band = item.get("band") if isinstance(item.get("band"), dict) else {}
        min_value = _safe_float(band.get("min"))
        max_value = _safe_float(band.get("max"))
        if min_value is not None:
          resolved_constraints.append({"cell": f"{resolved.get('sheet')}!{resolved.get('cell')}", "relation": 3, "value": min_value})
        if max_value is not None:
          resolved_constraints.append({"cell": f"{resolved.get('sheet')}!{resolved.get('cell')}", "relation": 1, "value": max_value})
        if str(item.get("grouping_mode") or "").strip().lower() == "grouped":
          group_key = str(item.get("group_key") or "").strip()
          if group_key:
            resolved_ref = f"{resolved.get('sheet')}!{resolved.get('cell')}"
            anchor_ref = grouped_anchor_cells.get(group_key)
            if anchor_ref:
              resolved_constraints.append({"cell": resolved_ref, "relation": 2, "value_ref": anchor_ref})
            else:
              grouped_anchor_cells[group_key] = resolved_ref
      for band_constraint in [item for item in (request.get("band_constraints") or []) if isinstance(item, dict)]:
        resolved_target = _resolve_finmo_output_cell((band_constraint.get("target") or {}) if isinstance(band_constraint.get("target"), dict) else {})
        if not resolved_target:
          continue
        goal_band = band_constraint.get("goal_band") if isinstance(band_constraint.get("goal_band"), dict) else {}
        min_value = _safe_float(goal_band.get("min"))
        max_value = _safe_float(goal_band.get("max"))
        if min_value is not None:
          resolved_constraints.append({"cell": f"{resolved_target.get('sheet')}!{resolved_target.get('cell')}", "relation": 3, "value": min_value})
        if max_value is not None:
          resolved_constraints.append({"cell": f"{resolved_target.get('sheet')}!{resolved_target.get('cell')}", "relation": 1, "value": max_value})
      request["changing_input_cells"] = resolved_inputs
      request["constraints"] = resolved_constraints
      request["execution_mode"] = "excel_solver_shell"
    return next_shell
  finally:
    wb.close()


def _execute_finmo_calibration_shell(
  *,
  finmo_path: str,
  calibration_shell: Dict[str, Any],
) -> Dict[str, Any]:
  shell = calibration_shell if isinstance(calibration_shell, dict) else {}
  goal_seek_requests: List[Dict[str, Any]] = []
  solver_requests: List[Dict[str, Any]] = []
  for request in [item for item in (shell.get("solver_requests") or []) if isinstance(item, dict)]:
    objective_cell = (request.get("objective_cell") or {}) if isinstance(request.get("objective_cell"), dict) else {}
    changing_cells = [item for item in (request.get("changing_input_cells") or []) if isinstance(item, dict)]
    if not objective_cell or not changing_cells:
      continue
    goal_band = (((request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {}).get("goal_band") or {})
    solver_requests.append(
      {
        "request_id": str(request.get("request_id") or "").strip(),
        "objective_cell": objective_cell,
        "changing_input_cells": changing_cells,
        "objective_mode": str((((request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {}).get("objective_mode") or "maximize")).strip().lower() or "maximize",
        "goal_band": _clone(goal_band if isinstance(goal_band, dict) else {}),
        "constraints": [item for item in (request.get("constraints") or []) if isinstance(item, dict)],
      }
    )
  if not goal_seek_requests and not solver_requests:
    return {"goal_seek_results": [], "solver_results": []}

  workbook_path = _normalize_finmo_path(finmo_path)
  payload = {
    "goal_seek_requests": goal_seek_requests,
    "solver_requests": solver_requests,
  }
  payload_path = ""
  with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8") as handle:
    json.dump(payload, handle)
    payload_path = handle.name
  script = (
    "$path = " + json.dumps(workbook_path) + ";"
    "$payloadPath = " + json.dumps(payload_path) + ";"
    "$payload = Get-Content -Raw $payloadPath | ConvertFrom-Json -Depth 100;"
    "$excel = New-Object -ComObject Excel.Application;"
    "$excel.Visible = $false;"
    "$excel.DisplayAlerts = $false;"
    "$goalResults = New-Object System.Collections.ArrayList;"
    "$solverResults = New-Object System.Collections.ArrayList;"
    "$wb = $null;"
    "try {"
    "$solver = $excel.AddIns.Item('Solver Add-in');"
    "if ($solver -and -not $solver.Installed) { $solver.Installed = $true }"
    "$wb = $excel.Workbooks.Open($path, $false, $false);"
    "foreach ($request in $payload.solver_requests) {"
    "  $excel.Run('Solver.xlam!SolverReset');"
    "  $objectiveSpec = \"'\" + $request.objective_cell.sheet + \"'!\" + $request.objective_cell.cell;"
    "  $changeSpec = (($request.changing_input_cells | ForEach-Object { \"'\" + $_.sheet + \"'!\" + $_.cell }) -join ',');"
    "  $objectiveMode = ($request.objective_mode | Out-String).Trim().ToLower();"
    "  if ($objectiveMode -eq 'minimize') {"
    "    $excel.Run('Solver.xlam!SolverOK', $objectiveSpec, 2, $null, $changeSpec);"
    "  } else {"
    "    $excel.Run('Solver.xlam!SolverOK', $objectiveSpec, 1, $null, $changeSpec);"
    "  }"
    "  if ($request.goal_band.min -ne $null) { $excel.Run('Solver.xlam!SolverAdd', $objectiveSpec, 3, [double]$request.goal_band.min) }"
    "  if ($request.goal_band.max -ne $null) { $excel.Run('Solver.xlam!SolverAdd', $objectiveSpec, 1, [double]$request.goal_band.max) }"
    "  foreach ($constraint in $request.constraints) {"
    "    if (-not $constraint.cell) { continue }"
    "    $parts = $constraint.cell.ToString().Split('!');"
    "    $constraintSpec = \"'\" + $parts[0] + \"'!\" + $parts[1];"
    "    if ($constraint.PSObject.Properties.Match('value_ref').Count -gt 0 -and $constraint.value_ref) {"
    "      $valueParts = $constraint.value_ref.ToString().Split('!');"
    "      $valueSpec = \"'\" + $valueParts[0] + \"'!\" + $valueParts[1];"
    "      $excel.Run('Solver.xlam!SolverAdd', $constraintSpec, [int]$constraint.relation, $valueSpec);"
    "    } else {"
    "      $excel.Run('Solver.xlam!SolverAdd', $constraintSpec, [int]$constraint.relation, $constraint.value);"
    "    }"
    "  }"
    "  $solveResult = $excel.Run('Solver.xlam!SolverSolve', $true);"
    "  $excel.Run('Solver.xlam!SolverFinish', 1);"
    "  $excel.CalculateFullRebuild();"
    "  $solveCode = [int]$solveResult;"
    "  $solveSuccess = ($solveCode -eq 0 -or $solveCode -eq 1);"
    "  [void]$solverResults.Add(@{ request_id = $request.request_id; success = $solveSuccess; solver_result = $solveResult.ToString(); solver_code = $solveCode; objective_mode = $objectiveMode });"
    "}"
    "$wb.Save();"
    "$wb.Close($true);"
    "Write-Output (@{ goal_seek_results = $goalResults; solver_results = $solverResults } | ConvertTo-Json -Compress -Depth 100);"
    "} finally {"
    "if ($wb) { try { $wb.Close($false) } catch {} }"
    "$excel.Quit();"
    "}"
  )
  try:
    try:
      completed = subprocess.run(["powershell", "-NoProfile", "-Command", script], check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as exc:
      return {
        "goal_seek_results": [],
        "solver_results": [],
        "success": False,
        "error": str(exc.stderr or exc.stdout or exc),
      }
    except Exception as exc:
      return {
        "goal_seek_results": [],
        "solver_results": [],
        "success": False,
        "error": str(exc),
      }
    try:
      parsed = json.loads(str(completed.stdout or "").strip() or "{}")
    except Exception:
      parsed = {}
    if isinstance(parsed, dict):
      parsed.setdefault("success", True)
      return parsed
    return {"goal_seek_results": [], "solver_results": [], "success": False, "error": "invalid_calibration_response"}
  finally:
    if payload_path:
      try:
        Path(payload_path).unlink(missing_ok=True)
      except Exception:
        pass


def _recalculate_excel_workbook(finmo_path: str) -> None:
  workbook_path = _normalize_finmo_path(finmo_path)
  script = (
    "$path = " + json.dumps(workbook_path) + ";"
    "$excel = New-Object -ComObject Excel.Application;"
    "$excel.Visible = $false;"
    "$excel.DisplayAlerts = $false;"
    "try {"
    "$wb = $excel.Workbooks.Open($path, $false, $false);"
    "$excel.CalculateFullRebuild();"
    "$wb.Save();"
    "$wb.Close($true);"
    "} finally {"
    "if ($wb) { try { $wb.Close($false) } catch {} }"
    "$excel.Quit();"
    "}"
  )
  subprocess.run(
    ["powershell", "-NoProfile", "-Command", script],
    check=True,
    capture_output=True,
    text=True,
  )


def run_finmo_goal_seek(
  *,
  finmo_path: str,
  objective_cell: Dict[str, Any],
  changing_input_cell: Dict[str, Any],
  goal_value: float,
) -> Dict[str, Any]:
  workbook_path = _normalize_finmo_path(finmo_path)
  objective_ref = f"{objective_cell.get('sheet')}!{objective_cell.get('cell')}"
  changing_ref = f"{changing_input_cell.get('sheet')}!{changing_input_cell.get('cell')}"
  script = (
    "$path = " + json.dumps(workbook_path) + ";"
    "$objective = " + json.dumps(objective_ref) + ";"
    "$changing = " + json.dumps(changing_ref) + ";"
    "$goal = " + json.dumps(goal_value) + ";"
    "$excel = New-Object -ComObject Excel.Application;"
    "$excel.Visible = $false;"
    "$excel.DisplayAlerts = $false;"
    "try {"
    "$wb = $excel.Workbooks.Open($path, $false, $false);"
    "$parts = $objective.Split('!');"
    "$objectiveRange = $wb.Worksheets.Item($parts[0]).Range($parts[1]);"
    "$parts2 = $changing.Split('!');"
    "$changingRange = $wb.Worksheets.Item($parts2[0]).Range($parts2[1]);"
    "$result = $objectiveRange.GoalSeek($goal, $changingRange);"
    "$excel.CalculateFullRebuild();"
    "$wb.Save();"
    "$wb.Close($true);"
    "Write-Output ($result.ToString());"
    "} finally {"
    "if ($wb) { try { $wb.Close($false) } catch {} }"
    "$excel.Quit();"
    "}"
  )
  completed = subprocess.run(["powershell", "-NoProfile", "-Command", script], check=True, capture_output=True, text=True)
  return {
    "objective_cell": objective_cell,
    "changing_input_cell": changing_input_cell,
    "goal_value": goal_value,
    "success": "True" in str(completed.stdout or ""),
  }


def _goal_value_from_band(goal_band: Optional[Dict[str, Any]]) -> Optional[float]:
  band = goal_band if isinstance(goal_band, dict) else {}
  min_value = _safe_float(band.get("min"))
  max_value = _safe_float(band.get("max"))
  if min_value is not None and max_value is not None:
    return (min_value + max_value) / 2.0
  if min_value is not None:
    return min_value
  if max_value is not None:
    return max_value
  return None


def run_finmo_goal_seek_request(
  *,
  finmo_path: str,
  request: Dict[str, Any],
) -> Dict[str, Any]:
  objective = (request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {}
  objective_cell = (request.get("objective_cell") or {}) if isinstance(request.get("objective_cell"), dict) else {}
  changing_inputs = [item for item in (request.get("changing_input_cells") or []) if isinstance(item, dict)]
  goal_value = _goal_value_from_band(objective.get("goal_band") if isinstance(objective, dict) else {})
  if not objective_cell or not changing_inputs or goal_value is None:
    return {
      "request_id": str(request.get("request_id") or "").strip(),
      "success": False,
      "error": "goal_seek_request_incomplete",
    }
  result = run_finmo_goal_seek(
    finmo_path=finmo_path,
    objective_cell=objective_cell,
    changing_input_cell=changing_inputs[0],
    goal_value=goal_value,
  )
  result["request_id"] = str(request.get("request_id") or "").strip()
  result["goal_band"] = _clone(objective.get("goal_band") or {})
  return result


def run_finmo_excel_solver(
  *,
  finmo_path: str,
  objective_cell: Dict[str, Any],
  changing_input_cells: Sequence[Dict[str, Any]],
  goal_band: Optional[Dict[str, Any]] = None,
  constraints: Optional[Sequence[Dict[str, Any]]] = None,
  objective_mode: str = "maximize",
) -> Dict[str, Any]:
  workbook_path = _normalize_finmo_path(finmo_path)
  objective_ref = f"{objective_cell.get('sheet')}!{objective_cell.get('cell')}"
  changing_refs = [
    f"{item.get('sheet')}!{item.get('cell')}"
    for item in (changing_input_cells or [])
    if isinstance(item, dict) and item.get("sheet") and item.get("cell")
  ]
  if not objective_ref or not changing_refs:
    return {
      "objective_cell": objective_cell,
      "changing_input_cells": list(changing_input_cells or []),
      "goal_band": _clone(goal_band or {}),
      "success": False,
      "error": "excel_solver_request_incomplete",
    }
  min_value = _safe_float((goal_band or {}).get("min"))
  max_value = _safe_float((goal_band or {}).get("max"))
  constraints_payload = [item for item in (constraints or []) if isinstance(item, dict)]
  normalized_objective_mode = str(objective_mode or "maximize").strip().lower() or "maximize"
  payload_path = ""
  solver_payload = {
    "objective": objective_ref,
    "changingRefs": changing_refs,
    "objectiveMode": normalized_objective_mode,
    "minGoal": min_value,
    "maxGoal": max_value,
    "constraints": constraints_payload,
  }
  with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8") as handle:
    json.dump(solver_payload, handle)
    payload_path = handle.name
  script = (
    "$path = " + json.dumps(workbook_path) + ";"
    "$payloadPath = " + json.dumps(payload_path) + ";"
    "$payload = Get-Content -Raw $payloadPath | ConvertFrom-Json -Depth 100;"
    "$objective = $payload.objective;"
    "$changingRefs = @($payload.changingRefs);"
    "$objectiveMode = ($payload.objectiveMode | Out-String).Trim().ToLower();"
    "$minGoal = $payload.minGoal;"
    "$maxGoal = $payload.maxGoal;"
    "$constraints = @($payload.constraints);"
    "$excel = New-Object -ComObject Excel.Application;"
    "$excel.Visible = $false;"
    "$excel.DisplayAlerts = $false;"
    "$wb = $null;"
    "try {"
    "$solver = $excel.AddIns.Item('Solver Add-in');"
    "if (-not $solver.Installed) { $solver.Installed = $true }"
    "$wb = $excel.Workbooks.Open($path, $false, $false);"
    "$excel.Run('Solver.xlam!SolverReset');"
    "$changeSpec = ($changingRefs | ForEach-Object { $parts = $_.Split('!'); $sheet = $parts[0]; $cell = $parts[1]; \"'\" + $sheet + \"'!\" + $cell }) -join ',';"
    "$parts = $objective.Split('!');"
    "$objectiveSpec = \"'\" + $parts[0] + \"'!\" + $parts[1];"
    "if ($objectiveMode -eq 'minimize') {"
    "$excel.Run('Solver.xlam!SolverOK', $objectiveSpec, 2, $null, $changeSpec);"
    "} else {"
    "$excel.Run('Solver.xlam!SolverOK', $objectiveSpec, 1, $null, $changeSpec);"
    "}"
    "if ($minGoal -ne $null) { $excel.Run('Solver.xlam!SolverAdd', $objectiveSpec, 3, $minGoal) }"
    "if ($maxGoal -ne $null) { $excel.Run('Solver.xlam!SolverAdd', $objectiveSpec, 1, $maxGoal) }"
    "foreach ($constraint in $constraints) {"
    "if (-not $constraint.cell) { continue }"
    "$relation = [int]($constraint.relation);"
    "if ($relation -lt 1) { continue }"
    "$constraintParts = $constraint.cell.Split('!');"
    "$constraintSpec = \"'\" + $constraintParts[0] + \"'!\" + $constraintParts[1];"
    "if ($constraint.PSObject.Properties.Match('value_ref').Count -gt 0 -and $constraint.value_ref) {"
    "$valueParts = $constraint.value_ref.Split('!');"
    "$valueSpec = \"'\" + $valueParts[0] + \"'!\" + $valueParts[1];"
    "$excel.Run('Solver.xlam!SolverAdd', $constraintSpec, $relation, $valueSpec);"
    "} else {"
    "$formulaText = $constraint.value;"
    "$excel.Run('Solver.xlam!SolverAdd', $constraintSpec, $relation, $formulaText);"
    "}"
    "}"
    "$solveResult = $excel.Run('Solver.xlam!SolverSolve', $true);"
    "$excel.Run('Solver.xlam!SolverFinish', 1);"
    "$excel.CalculateFullRebuild();"
    "$wb.Save();"
    "$wb.Close($true);"
    "$solveCode = [int]$solveResult;"
    "$solveSuccess = ($solveCode -eq 0 -or $solveCode -eq 1);"
    "Write-Output ((@{ solver_result = $solveResult.ToString(); solver_code = $solveCode; success = $solveSuccess; objective_mode = $objectiveMode } | ConvertTo-Json -Compress -Depth 10));"
    "} finally {"
    "if ($wb) { try { $wb.Close($false) } catch {} }"
    "$excel.Quit();"
    "}"
  )
  try:
    completed = subprocess.run(["powershell", "-NoProfile", "-Command", script], check=True, capture_output=True, text=True)
    try:
      parsed = json.loads(str(completed.stdout or "").strip() or "{}")
    except Exception:
      parsed = {}
    return {
      "objective_cell": objective_cell,
      "changing_input_cells": list(changing_input_cells or []),
      "goal_band": _clone(goal_band or {}),
      "constraints": _clone(list(constraints or [])),
      "objective_mode": normalized_objective_mode,
      "success": bool(parsed.get("success")) if isinstance(parsed, dict) else False,
      "solver_result": str((parsed.get("solver_result") if isinstance(parsed, dict) else None) or str(completed.stdout or "").strip()),
      "solver_code": _safe_int(parsed.get("solver_code")) if isinstance(parsed, dict) else 0,
    }
  finally:
    if payload_path:
      try:
        Path(payload_path).unlink(missing_ok=True)
      except Exception:
        pass


def run_finmo_excel_solver_request(
  *,
  finmo_path: str,
  request: Dict[str, Any],
) -> Dict[str, Any]:
  objective = (request.get("objective") or {}) if isinstance(request.get("objective"), dict) else {}
  objective_cell = (request.get("objective_cell") or {}) if isinstance(request.get("objective_cell"), dict) else {}
  changing_inputs = [item for item in (request.get("changing_input_cells") or []) if isinstance(item, dict)]
  if not objective_cell or not changing_inputs:
    return {
      "request_id": str(request.get("request_id") or "").strip(),
      "success": False,
      "error": "excel_solver_request_incomplete",
    }
  result = run_finmo_excel_solver(
    finmo_path=finmo_path,
    objective_cell=objective_cell,
    changing_input_cells=changing_inputs,
    goal_band=(objective.get("goal_band") or {}) if isinstance(objective, dict) else {},
    constraints=request.get("constraints") if isinstance(request.get("constraints"), list) else [],
    objective_mode=str((objective.get("objective_mode") if isinstance(objective, dict) else None) or "maximize"),
  )
  result["request_id"] = str(request.get("request_id") or "").strip()
  return result


def sync_consistency_state_to_finmo(
  *,
  finmo_path: Any,
  business_facts: Optional[Dict[str, Any]],
  ops_json: Optional[Dict[str, Any]],
  people_json: Optional[Dict[str, Any]],
  financials_json: Optional[Dict[str, Any]],
  financials_year1_json: Optional[Dict[str, Any]],
  marketing_model_json: Optional[Dict[str, Any]],
  controller_input_seed: Optional[Sequence[Dict[str, Any]]] = None,
  forecast_quarters: Sequence[Dict[str, Any]],
  calibration_spec: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
  path = str(finmo_path or "").strip()
  try:
    try:
      from consistency_trace import trace_lazy  # type: ignore
    except Exception:
      from client_intake_and_finmo.consistency_trace import trace_lazy  # type: ignore
    trace_lazy(
      "FINMO_SYNC_REQUEST",
      "Finmo workbook sync request",
      lambda: {
        "finmo_path": path,
        "business_facts": _clone(business_facts or {}),
        "ops_json": _clone(ops_json or {}),
        "people_json": _clone(people_json or {}),
        "financials_json": _clone(financials_json or {}),
        "financials_year1_json": _clone(financials_year1_json or {}),
        "marketing_model_json": _clone(marketing_model_json or {}),
        "controller_input_seed": _clone(controller_input_seed or []),
        "forecast_quarters": _clone(forecast_quarters or []),
        "calibration_spec": _clone(calibration_spec or {}),
      },
    )
  except Exception:
    pass
  model_input_json = build_python_model_input_json(
    business_facts=business_facts or {},
    ops_json=ops_json or {},
    people_json=people_json or {},
    financials_json=financials_json or {},
    financials_year1_json=financials_year1_json or {},
    marketing_model_json=marketing_model_json or {},
    controller_input_seed=controller_input_seed or [],
    forecast_quarters=forecast_quarters or [],
    business_name=str((business_facts or {}).get("business_name") or "").strip(),
  )
  finmo_json = build_python_finmo_json(
    model_input_json=model_input_json,
    finmo_path=path,
  )
  calibration_results: Dict[str, Any] = {}
  if isinstance(calibration_spec, dict) and calibration_spec:
    calibration_results = {
      "success": False,
      "status": "python_finmo_calibration_not_supported",
      "request_present": True,
    }
    model_input_json["calibration_results"] = _clone(calibration_results)
    finmo_json["calibration_results"] = _clone(calibration_results)
    try:
      try:
        from consistency_trace import trace_lazy  # type: ignore
      except Exception:
        from client_intake_and_finmo.consistency_trace import trace_lazy  # type: ignore
      trace_lazy(
        "FINMO_CALIBRATION",
        "Finmo calibration shell and results",
        lambda: {
          "finmo_path": path,
          "calibration_shell": {},
          "calibration_results": _clone(calibration_results),
        },
      )
    except Exception:
      pass
  try:
    try:
      from consistency_trace import trace_lazy  # type: ignore
    except Exception:
      from client_intake_and_finmo.consistency_trace import trace_lazy  # type: ignore
      trace_lazy(
        "FINMO_SYNC_RESULT",
        "Current persisted workbook state",
        lambda: {
          "finmo_path": path,
          "model_input_json": _clone(model_input_json),
          "finmo_json": _clone(finmo_json),
        },
      )
  except Exception:
    pass
  return {
    "finmo_path": path,
    "model_input_json": model_input_json,
    "finmo_json": finmo_json,
  }
