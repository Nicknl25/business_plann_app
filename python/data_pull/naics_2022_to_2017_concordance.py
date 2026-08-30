"""NAICS 2022 -> 2017 concordance -> naics_2022_to_2017_concordance (2026-08-30).

WHY: NAICS is revised every five years and codes get renumbered. CBP 2022 files
under NAICS2017; a client classified under a 2022 code that was renumbered
finds nothing - Northgate's 513210 (Software Publishers, 2022) is 511210 in
NAICS2017. Same business, different number.

Census publishes the concordance as an Excel file. Same discipline as the
other loaders: its own table, one create, one load, standalone in data_pull/,
refuses to double-load. Not folded into any existing table (Nick's directive).

Source: https://www.census.gov/naics/concordances/2022_to_2017_NAICS.xlsx
"""
import io
import os
import sys

import mysql.connector
import requests
from dotenv import load_dotenv

ROOT_ENV_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '.env'))
load_dotenv(ROOT_ENV_PATH)

MYSQL_HOST = os.getenv("MYSQL_HOST")
MYSQL_USER = os.getenv("MYSQL_USER")
MYSQL_PASSWORD = os.getenv("MYSQL_PASSWORD")
MYSQL_DB = os.getenv("MYSQL_DB")

URL = "https://www.census.gov/naics/concordances/2022_to_2017_NAICS.xlsx"

TABLE = "naics_2022_to_2017_concordance"

CREATE_SQL = f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
  id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT PRIMARY KEY,
  naics_2022 VARCHAR(8) NOT NULL,
  naics_2022_title VARCHAR(255) NULL,
  naics_2017 VARCHAR(8) NOT NULL,
  naics_2017_title VARCHAR(255) NULL,
  KEY ix_2022 (naics_2022),
  KEY ix_2017 (naics_2017)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
"""

INSERT_SQL = f"""
INSERT INTO {TABLE} (naics_2022, naics_2022_title, naics_2017, naics_2017_title)
VALUES (%s, %s, %s, %s)
"""


def main() -> int:
    print("Downloading the 2022-to-2017 NAICS concordance from Census...")
    resp = requests.get(URL, timeout=300,
                        headers={"User-Agent": "Mozilla/5.0 (data loader)"})
    resp.raise_for_status()

    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    print(f"Sheet '{ws.title}': {len(rows)} raw rows")

    # Find the header row by content rather than assuming position: the Census
    # file leads with title rows. Header carries "2022 NAICS Code" and
    # "2017 NAICS Code" (wording varies slightly between revisions).
    hdr_idx, cols = None, {}
    for i, row in enumerate(rows[:10]):
        texts = [str(c or "").strip().lower() for c in row]
        c22 = next((j for j, t in enumerate(texts) if "2022" in t and "code" in t), None)
        # DISTINCT column required: the sheet's leading title row reads
        # "2022 to 2017 NAICS ..." in ONE cell, which matched both patterns at
        # index 0 and loaded every pair as code->itself (caught by the
        # identical-pairs check below, 1150/1150).
        c17 = next((j for j, t in enumerate(texts) if "2017" in t and "code" in t and j != c22), None)
        if c22 is not None and c17 is not None:
            t22 = next((j for j, t in enumerate(texts) if "2022" in t and "title" in t), None)
            t17 = next((j for j, t in enumerate(texts) if "2017" in t and "title" in t), None)
            hdr_idx, cols = i, {"c22": c22, "t22": t22, "c17": c17, "t17": t17}
            break
    if hdr_idx is None:
        raise SystemExit("could not find the header row; first rows: %r" % rows[:3])
    print("header row %d, columns %s" % (hdr_idx, cols))

    conn = mysql.connector.connect(host=MYSQL_HOST, user=MYSQL_USER,
                                   password=MYSQL_PASSWORD, database=MYSQL_DB)
    cur = conn.cursor()
    cur.execute(CREATE_SQL)
    cur.execute(f"SELECT COUNT(*) FROM {TABLE}")
    if cur.fetchone()[0]:
        raise SystemExit(f"{TABLE} already loaded - drop it first to reload.")

    batch, inserted = [], 0
    for row in rows[hdr_idx + 1:]:
        def cell(j):
            return str(row[j]).strip() if j is not None and j < len(row) and row[j] is not None else None
        c22, c17 = cell(cols["c22"]), cell(cols["c17"])
        if not c22 or not c17 or not c22[:1].isdigit() or not c17[:1].isdigit():
            continue
        # codes arrive as ints or floats from Excel occasionally
        c22 = c22.split(".")[0]
        c17 = c17.split(".")[0]
        # Only real codes: the sheet carries footnote rows and merged-cell
        # spillover where a title lands in a code column ("Data too long").
        if not (c22.isdigit() and c17.isdigit() and 2 <= len(c22) <= 6 and 2 <= len(c17) <= 6):
            continue
        batch.append((c22, cell(cols["t22"]), c17, cell(cols["t17"])))
    identical = sum(1 for b in batch if b[0] == b[2])
    if batch and identical == len(batch):
        raise SystemExit("every pair is code->itself (%d rows) - the header "
                         "detection picked one column twice; refusing to load garbage." % identical)
    if batch:
        cur.executemany(INSERT_SQL, batch)
        inserted = len(batch)
    conn.commit()
    cur.close()
    conn.close()
    print(f"Inserted {inserted} concordance rows into {TABLE}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
