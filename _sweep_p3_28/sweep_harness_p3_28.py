"""Phase 9 P3.28 — 28-draft empirical sweep + architectural audit harness.

OBSERVATION ONLY. No code changes during execution. Iterates through
28 source draft IDs sequentially, invoking the canonical runner
(Test Files/run_persisted_system_run.py) one at a time. For each
draft:

  1. Run runner with 15-min hard timeout
  2. Parse stdout for new_draft_id, business_name, workbook_path
  3. Read the produced workbook for V-1..V-4 analysis:
       V-1: Diagnostics!Acceptance Score == 16/16
       V-2: Checks!B2 == "OK"
       V-3: FINMO trajectory matches realism claims
       V-4: Checks!E166-E172 vs F166-F172 deltas < $50 / 0.01%
  4. Move workbook to docs/architecture/p3_28_sweep_workbooks/<draft_id>.xlsx
  5. Append CSV row to docs/architecture/p3_28_sweep_results.csv
  6. git add CSV + workbook → commit → push

Stop conditions:
  - Python framework traceback in runner
  - 3 consecutive RUNNER_ERROR outcomes
  - Disk < 1GB
  - 3 consecutive git push failures
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


REPO_ROOT = Path(__file__).resolve().parents[1]
RUNNER = REPO_ROOT / "Test Files" / "run_persisted_system_run.py"
SWEEP_DIR = Path(__file__).resolve().parent
LOG_DIR = SWEEP_DIR / "per_draft_logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

ARCH_DIR = REPO_ROOT / "docs" / "architecture"
WORKBOOK_DIR = ARCH_DIR / "p3_28_sweep_workbooks"
WORKBOOK_DIR.mkdir(parents=True, exist_ok=True)

CSV_PATH = ARCH_DIR / "p3_28_sweep_results.csv"
RESULTS_PATH = SWEEP_DIR / "results.json"
STOP_SIGNAL_PATH = SWEEP_DIR / "STOP"

CLIENT_PLANS_DIR = Path(r"C:\dev\Cilient Plans")

# 15-min per-draft hard ceiling per directive.
PER_DRAFT_TIMEOUT_SEC = 900


DRAFT_IDS: List[str] = [
  "25f746500d1d456da638ee216669b78e",
  "201d0ad18ae243dba933703d19cda4df",
  "41f014a5567041d99b2572a67fe6b03d",
  "6d37c6b98ace41ee9c91dd5fbf68b83e",
  "4aaa94e10efc4968866844197f96f24c",
  "290547bbe8cf48899d897cc5ce865111",
  "52ba2e2c1f4f439c9868541606f3affd",
  "4207488106054d72afbe16480e1de100",
  "25b8e17eda804fa7a46adf72a3503900",
  "5dcd919aae314bd5af67849172aa52bb",
  "5af71d361b324f62a3598e8da40c98c7",
  "2aeae92f95f44fbd9a38b6e319c1bcc7",
  "d345194155584588a77cfb0c8dbedc8c",
  "28af3a88a2294237a1aa6294bf02e078",
  "0459723bad0e461f95dbfe5fbf14631f",
  "1bef9076e6504af9a9fe223af128110b",
  "0bda87acc8a945cb81c63a8940ef5044",
  "077e83b679eb4be7be5febabf05f6379",
  "194c29fed27c4fa39498ed95f8d45343",
  "0d0fb60aca754e00954f402a4fdec0ab",
  "0fe6f320519f412eb113f413d803d338",
  "56b8623063a34e6d9c1803568a730825",
  "11d6cd0c19c3430e8aaf8916b550ea7f",
  "26fc7c5b8d1349048a538f65b4f85beb",
  "0e1d881125b543aea68c603782e00da9",
  "03516b40b43d4592b8d7681a4f57dc67",
  "4c6ff03a7e474f57a3a1ff412860cc2f",
  "1aa5d27085e44aee98d401a5cd56bad0",
]


CSV_COLUMNS = [
  "draft_id", "business_name", "naics", "planning_mode", "cash_strategy",
  "business_stage", "outcome", "fail_category", "fail_op_code",
  "wall_clock_s", "handler_fired", "handler_scope", "handler_tool_calls",
  "v1_score_16", "v2_model_status", "v3_pass", "v4_pass",
  "v4_max_abs_delta", "v4_max_pct_delta",
  "ebitda_q1", "ebitda_q11", "ebitda_q20",
  "cash_q1", "cash_q11", "cash_q20",
  "total_equity_q20", "notes",
]


def _now_iso() -> str:
  return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _load_results() -> Dict[str, Any]:
  if RESULTS_PATH.exists():
    try:
      return json.loads(RESULTS_PATH.read_text(encoding="utf-8"))
    except Exception:
      pass
  return {"sweep_started_at": _now_iso(), "drafts": {}}


def _save_results(results: Dict[str, Any]) -> None:
  RESULTS_PATH.write_text(
    json.dumps(results, ensure_ascii=False, indent=2),
    encoding="utf-8",
  )


def _ensure_csv_header() -> None:
  if CSV_PATH.exists() and CSV_PATH.stat().st_size > 0:
    return
  with CSV_PATH.open("w", encoding="utf-8", newline="") as f:
    writer = csv.writer(f)
    writer.writerow(CSV_COLUMNS)


def _append_csv_row(row: Dict[str, Any]) -> None:
  with CSV_PATH.open("a", encoding="utf-8", newline="") as f:
    writer = csv.writer(f)
    writer.writerow([row.get(c, "") for c in CSV_COLUMNS])


def _parse_runner_stdout(text: str) -> Dict[str, Any]:
  out: Dict[str, Any] = {}
  m = re.search(r"Cloned source intake state into new draft:\s+(\S+)", text)
  if m:
    out["new_draft_id"] = m.group(1).strip()
  m = re.search(r"Business Name:\s+(.+)", text)
  if m:
    out["business_name"] = m.group(1).strip()
  m = re.search(r"Saved client financial model workbook:\s+(.+)", text)
  if m:
    out["workbook_path"] = m.group(1).strip()
  m = re.search(r"System run duration:\s+(\d+)\s+ms", text)
  if m:
    try:
      out["system_run_ms"] = int(m.group(1))
    except ValueError:
      pass
  # Extract structured ERROR payload if present.
  err_idx = text.find("ERROR:")
  if err_idx >= 0:
    out["runner_error_blob"] = text[err_idx:err_idx + 200000]
  return out


def _find_workbook_for_business(business_name: str, run_start_ts: float) -> Optional[Path]:
  """Look in C:\\dev\\Cilient Plans\\ for a workbook whose name starts
  with the business name (with non-alphanumeric chars accepted) and
  whose mtime is after run_start_ts. Returns most recent match."""
  if not CLIENT_PLANS_DIR.exists():
    return None
  # Workbook name shape: '<business_name with & dropped/punct adjusted> -- MM-DD-YYYY HH-MM-SS.xlsx'
  bn_clean = re.sub(r"[^A-Za-z0-9 ]+", " ", business_name or "").strip()
  prefix_words = [w for w in bn_clean.split() if w]
  candidates: List[Path] = []
  for p in CLIENT_PLANS_DIR.glob("*.xlsx"):
    try:
      mt = p.stat().st_mtime
    except Exception:
      continue
    if mt < run_start_ts - 5:
      continue
    name = p.stem
    # require at least the first 2 words match
    if not prefix_words:
      continue
    name_norm = re.sub(r"[^A-Za-z0-9 ]+", " ", name)
    ok = all(w in name_norm for w in prefix_words[:3])
    if ok:
      candidates.append(p)
  if not candidates:
    return None
  candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
  return candidates[0]


def _analyze_workbook(wb_path: Path) -> Dict[str, Any]:
  """Run V-1..V-4 analysis on the workbook. Returns a dict of fields
  matching CSV_COLUMNS (subset). Raises on truly unreadable workbook."""
  import openpyxl
  out: Dict[str, Any] = {}
  wb = openpyxl.load_workbook(str(wb_path), data_only=True, read_only=True)
  sheets = set(wb.sheetnames)

  # --- Diagnostics ---
  diag_score = ""
  diag_handler_fired = ""
  diag_handler_scope = ""
  diag_handler_tools = ""
  diag_planning_mode = ""
  diag_cash_strategy = ""
  diag_business_stage = ""
  diag_naics = ""
  if "Diagnostics" in sheets:
    diag = wb["Diagnostics"]
    for row in diag.iter_rows(min_row=1, max_row=200, values_only=True):
      if not row or row[0] is None:
        continue
      label = str(row[0] or "").strip()
      val = row[1] if len(row) > 1 else None
      if label == "Score":
        diag_score = str(val or "").strip()
      elif label == "Fired":
        diag_handler_fired = str(val or "").strip()
      elif label == "Scope":
        diag_handler_scope = str(val or "").strip()
      elif label.startswith("Tool Calls"):
        diag_handler_tools = str(val or "").strip()
      elif label == "Planning Mode":
        diag_planning_mode = str(val or "").strip()
      elif label == "Cash Strategy":
        diag_cash_strategy = str(val or "").strip()
      elif label == "Business Stage":
        diag_business_stage = str(val or "").strip()
      elif label == "NAICS-6":
        diag_naics = str(val or "").strip()

  out["v1_score_16"] = diag_score
  out["handler_fired"] = diag_handler_fired
  out["handler_scope"] = diag_handler_scope
  out["handler_tool_calls"] = diag_handler_tools
  out["planning_mode"] = diag_planning_mode
  out["cash_strategy"] = diag_cash_strategy
  out["business_stage"] = diag_business_stage
  out["naics"] = diag_naics

  # --- Checks B2 model status + baseline deltas (V-2 + V-4) ---
  v2_model_status = ""
  v4_max_abs = 0.0
  v4_max_pct = 0.0
  v4_pass = True
  if "Checks" in sheets:
    ch = wb["Checks"]
    v2_model_status = str(ch["B2"].value or "").strip()
    # Persisted Baseline rows 166-172; col E = FINMO val, F = Audit val, G = delta
    for r in range(166, 173):
      finmo_v = ch.cell(row=r, column=5).value
      audit_v = ch.cell(row=r, column=6).value
      delta_v = ch.cell(row=r, column=7).value
      try:
        delta_abs = abs(float(delta_v or 0.0))
      except Exception:
        delta_abs = 0.0
      try:
        finmo_f = float(finmo_v or 0.0)
        if abs(finmo_f) > 1e-9:
          delta_pct = delta_abs / abs(finmo_f)
        else:
          delta_pct = 0.0
      except Exception:
        delta_pct = 0.0
      if delta_abs > v4_max_abs:
        v4_max_abs = delta_abs
      if delta_pct > v4_max_pct:
        v4_max_pct = delta_pct
      # V-4 thresholds: $50 absolute / 0.01% relative
      if delta_abs > 50.0 and delta_pct > 0.0001:
        v4_pass = False
  out["v2_model_status"] = v2_model_status
  out["v4_max_abs_delta"] = round(v4_max_abs, 2)
  out["v4_max_pct_delta"] = round(v4_max_pct * 100.0, 6)
  out["v4_pass"] = "Y" if v4_pass else "N"

  # --- FINMO trajectories (V-3) ---
  # Quarter columns: Q1=D(4), Q5=H(8), Q11=N(14), Q15=R(18), Q20=W(23)
  # Rows: Revenue=8, EBITDA=16, Net Income=20, Cash=23, Payroll=14,
  #       OCF=51, Total Equity=42
  def _val(row: int, col: int) -> Optional[float]:
    if "FINMO" not in sheets:
      return None
    try:
      v = wb["FINMO"].cell(row=row, column=col).value
      return float(v) if v is not None else None
    except Exception:
      return None

  ebitda_q1 = _val(16, 4)
  ebitda_q5 = _val(16, 8)
  ebitda_q11 = _val(16, 14)
  ebitda_q20 = _val(16, 23)
  cash_q1 = _val(23, 4)
  cash_q5 = _val(23, 8)
  cash_q11 = _val(23, 14)
  cash_q20 = _val(23, 23)
  ni_q11 = _val(20, 14)
  ni_q20 = _val(20, 23)
  ocf_q11 = _val(51, 14)
  total_equity_q20 = _val(42, 23)

  def _r(v: Optional[float]) -> str:
    return f"{v:.2f}" if v is not None else ""
  out["ebitda_q1"] = _r(ebitda_q1)
  out["ebitda_q11"] = _r(ebitda_q11)
  out["ebitda_q20"] = _r(ebitda_q20)
  out["cash_q1"] = _r(cash_q1)
  out["cash_q11"] = _r(cash_q11)
  out["cash_q20"] = _r(cash_q20)
  out["total_equity_q20"] = _r(total_equity_q20)

  # V-3: cross-check viability claims (read from diagnostics realism)
  # Simple proxy checks - if V-1 passed 16/16 but FINMO Q11 EBITDA<0
  # OR cash<0 through Q5 → V-3 fail.
  v3_pass = True
  v3_notes = []
  if diag_score.startswith("16/16"):
    if ebitda_q11 is not None and ebitda_q11 <= 0:
      v3_pass = False
      v3_notes.append(f"ebitda_positive_by_q11 claimed but FINMO Q11 EBITDA={ebitda_q11:.0f}")
    # Cash through Q5 — pull Q1..Q5
    for q_idx in range(1, 6):
      cash_q = _val(23, 3 + q_idx)
      if cash_q is not None and cash_q < 0:
        v3_pass = False
        v3_notes.append(f"loss_window_funded claimed but FINMO Q{q_idx} cash={cash_q:.0f}")
        break
    if ni_q11 is not None and ni_q11 < 0:
      # NI margin may be claimed
      v3_notes.append(f"FINMO Q11 NI={ni_q11:.0f} (warn)")
  out["v3_pass"] = "Y" if v3_pass else "N"
  out["notes"] = "; ".join(v3_notes) if v3_notes else ""

  wb.close()
  return out


def _classify_failure(parsed: Dict[str, Any]) -> Tuple[str, str]:
  """Return (fail_category, fail_op_code) from runner stdout blob."""
  blob = parsed.get("runner_error_blob", "") or ""
  if not blob:
    return ("", "")
  # Look for known fail-fast operation codes
  m = re.search(r"'(payroll_revenue_economic_feasibility_failed|payroll_stage_profitability_feasibility_failed|payroll_headcount_contract_timeout|stage_ramp_revenue_path_not_applied|cash_buffer_minimum_violation|acceptance_gate_failed|post_intake_schedule_marker_missing|accounting_equation_violation|model_status_failed|post_intake_finalize_validation_failed)'", blob)
  op_code = m.group(1) if m else ""

  category = ""
  if "acceptance_gate" in blob:
    category = "acceptance_gate"
  elif "payroll" in blob and "feasibility" in blob:
    category = "payroll_feasibility"
  elif "stage_ramp" in blob:
    category = "stage_ramp"
  elif "cash_buffer" in blob:
    category = "cash_buffer"
  elif "accounting_equation" in blob:
    category = "accounting_equation"
  elif "payroll_headcount_contract_timeout" in blob:
    category = "gpt_timeout"
  elif "model_status" in blob:
    category = "model_status"
  elif "finalize_validation" in blob:
    category = "finalize_validation"
  else:
    category = "other"
  return (category, op_code)


def _classify_outcome(returncode: int, parsed: Dict[str, Any], wb_analysis: Optional[Dict[str, Any]]) -> str:
  """V-1∧V-2∧V-3∧V-4 → outcome label."""
  if returncode == 0:
    if not wb_analysis:
      return "WORKBOOK_ERROR"
    v1 = str(wb_analysis.get("v1_score_16") or "").startswith("16/16")
    v2 = str(wb_analysis.get("v2_model_status") or "").upper() == "OK"
    v3 = str(wb_analysis.get("v3_pass") or "") == "Y"
    v4 = str(wb_analysis.get("v4_pass") or "") == "Y"
    if v1 and v2 and v3 and v4:
      return "GENUINE_PASS"
    if v1 and v2:
      return "FALSE_PASS"
    return "FAIL"
  # Non-zero: check whether workbook exists (FAIL with workbook) vs no workbook (FAIL early)
  if wb_analysis:
    return "FAIL"
  blob = parsed.get("runner_error_blob", "") or ""
  if blob:
    return "FAIL"
  return "RUNNER_ERROR"


def _run_single_draft(draft_id: str, idx: int, total: int) -> Dict[str, Any]:
  log_path = LOG_DIR / f"{draft_id}.log"
  start_iso = _now_iso()
  start_perf = time.perf_counter()
  start_ts = time.time()
  timed_out = False
  try:
    proc = subprocess.run(
      [
        sys.executable,
        str(RUNNER),
        "--draft-id", draft_id,
        "--base-url", "http://127.0.0.1:5050",
      ],
      cwd=str(REPO_ROOT),
      capture_output=True,
      text=True,
      timeout=PER_DRAFT_TIMEOUT_SEC,
      check=False,
    )
    stdout = proc.stdout or ""
    stderr = proc.stderr or ""
    returncode = proc.returncode
  except subprocess.TimeoutExpired as exc:
    stdout = (exc.stdout.decode("utf-8", errors="replace") if isinstance(exc.stdout, bytes) else (exc.stdout or "")) if exc.stdout else ""
    stderr = (exc.stderr.decode("utf-8", errors="replace") if isinstance(exc.stderr, bytes) else (exc.stderr or "")) if exc.stderr else ""
    returncode = -1
    timed_out = True

  elapsed = time.perf_counter() - start_perf
  end_iso = _now_iso()
  log_path.write_text(
    "===== STDOUT =====\n" + stdout + "\n===== STDERR =====\n" + stderr,
    encoding="utf-8",
  )
  parsed = _parse_runner_stdout(stdout + "\n" + stderr)
  business_name = parsed.get("business_name", "")
  wb_path: Optional[Path] = None
  wb_dst: Optional[Path] = None
  wb_analysis: Optional[Dict[str, Any]] = None
  workbook_error: Optional[str] = None

  # Workbook may exist even on FAIL (e.g. acceptance_gate_failed produces one).
  # First try the path from stdout, then fall back to mtime search.
  ws_path = parsed.get("workbook_path") or ""
  if ws_path and Path(ws_path).exists():
    wb_path = Path(ws_path)
  if wb_path is None and business_name:
    wb_path = _find_workbook_for_business(business_name, start_ts)

  if wb_path is not None and wb_path.exists():
    wb_dst = WORKBOOK_DIR / f"{draft_id}.xlsx"
    try:
      shutil.copy2(str(wb_path), str(wb_dst))
    except Exception as e:
      workbook_error = f"copy_failed: {e!r}"
    try:
      wb_analysis = _analyze_workbook(wb_dst or wb_path)
    except Exception as e:
      workbook_error = f"analyze_failed: {type(e).__name__}: {str(e)[:200]}"
      wb_analysis = None

  if timed_out:
    outcome = "RUNNER_ERROR"
  else:
    outcome = _classify_outcome(returncode, parsed, wb_analysis)

  fail_cat, fail_op = _classify_failure(parsed) if outcome != "GENUINE_PASS" else ("", "")

  row: Dict[str, Any] = {
    "draft_id": draft_id,
    "business_name": business_name,
    "naics": (wb_analysis or {}).get("naics", ""),
    "planning_mode": (wb_analysis or {}).get("planning_mode", ""),
    "cash_strategy": (wb_analysis or {}).get("cash_strategy", ""),
    "business_stage": (wb_analysis or {}).get("business_stage", ""),
    "outcome": outcome,
    "fail_category": fail_cat,
    "fail_op_code": fail_op,
    "wall_clock_s": round(elapsed, 2),
    "handler_fired": (wb_analysis or {}).get("handler_fired", ""),
    "handler_scope": (wb_analysis or {}).get("handler_scope", ""),
    "handler_tool_calls": (wb_analysis or {}).get("handler_tool_calls", ""),
    "v1_score_16": (wb_analysis or {}).get("v1_score_16", ""),
    "v2_model_status": (wb_analysis or {}).get("v2_model_status", ""),
    "v3_pass": (wb_analysis or {}).get("v3_pass", ""),
    "v4_pass": (wb_analysis or {}).get("v4_pass", ""),
    "v4_max_abs_delta": (wb_analysis or {}).get("v4_max_abs_delta", ""),
    "v4_max_pct_delta": (wb_analysis or {}).get("v4_max_pct_delta", ""),
    "ebitda_q1": (wb_analysis or {}).get("ebitda_q1", ""),
    "ebitda_q11": (wb_analysis or {}).get("ebitda_q11", ""),
    "ebitda_q20": (wb_analysis or {}).get("ebitda_q20", ""),
    "cash_q1": (wb_analysis or {}).get("cash_q1", ""),
    "cash_q11": (wb_analysis or {}).get("cash_q11", ""),
    "cash_q20": (wb_analysis or {}).get("cash_q20", ""),
    "total_equity_q20": (wb_analysis or {}).get("total_equity_q20", ""),
    "notes": ((wb_analysis or {}).get("notes", "") + ("; " + workbook_error if workbook_error else ""))[:500],
  }

  return {
    "row": row,
    "wb_dst": str(wb_dst) if wb_dst else "",
    "parsed": parsed,
    "outcome": outcome,
    "elapsed": elapsed,
    "returncode": returncode,
    "timed_out": timed_out,
    "start_iso": start_iso,
    "end_iso": end_iso,
  }


def _git_commit_push(draft_id: str, idx: int, business_name: str) -> bool:
  """git add CSV + workbook → commit → push. Returns True on success."""
  short_bn = re.sub(r"[^A-Za-z0-9]+", "_", (business_name or "unknown"))[:40].strip("_") or "unknown"
  rel_csv = "docs/architecture/p3_28_sweep_results.csv"
  rel_wb = f"docs/architecture/p3_28_sweep_workbooks/{draft_id}.xlsx"
  msg = f"phase_9_p3_28_sweep_draft_{idx:02d}_{short_bn}"
  try:
    subprocess.run(["git", "add", rel_csv], cwd=str(REPO_ROOT), check=True, capture_output=True, text=True)
    if (REPO_ROOT / rel_wb).exists():
      subprocess.run(["git", "add", rel_wb], cwd=str(REPO_ROOT), check=True, capture_output=True, text=True)
    # Check if any staged changes
    diff = subprocess.run(["git", "diff", "--cached", "--name-only"], cwd=str(REPO_ROOT), capture_output=True, text=True)
    if not (diff.stdout or "").strip():
      print(f"  no staged changes for draft {idx}, skipping commit", flush=True)
      return True
    subprocess.run(["git", "commit", "-m", msg], cwd=str(REPO_ROOT), check=True, capture_output=True, text=True)
    subprocess.run(["git", "push", "origin", "intake-stable"], cwd=str(REPO_ROOT), check=True, capture_output=True, text=True, timeout=120)
    return True
  except subprocess.CalledProcessError as e:
    print(f"  git op failed: {e.stderr or e.stdout}", flush=True)
    return False
  except subprocess.TimeoutExpired:
    print("  git push timeout", flush=True)
    return False


def _check_disk_gb() -> float:
  try:
    total, used, free = shutil.disk_usage(str(REPO_ROOT))
    return free / 1e9
  except Exception:
    return 999.0


def main() -> int:
  parser = argparse.ArgumentParser()
  parser.add_argument("--limit", type=int, default=None)
  parser.add_argument("--start-from", type=int, default=0)
  args = parser.parse_args()

  _ensure_csv_header()
  results = _load_results()
  drafts_done = results.setdefault("drafts", {})

  ids = DRAFT_IDS
  if args.start_from:
    ids = ids[args.start_from:]
  if args.limit is not None:
    ids = ids[: args.limit]

  consecutive_runner_errors = 0
  consecutive_push_failures = 0

  for idx_raw, draft_id in enumerate(ids, start=1):
    idx = args.start_from + idx_raw
    if STOP_SIGNAL_PATH.exists():
      print(f"[{idx}/{len(DRAFT_IDS)}] STOP signal file present; halting.", flush=True)
      break
    if draft_id in drafts_done and drafts_done[draft_id].get("outcome") not in {None, "", "PENDING"}:
      print(f"[{idx}/{len(DRAFT_IDS)}] {draft_id} already complete ({drafts_done[draft_id]['outcome']}); skipping.", flush=True)
      continue

    disk_gb = _check_disk_gb()
    if disk_gb < 1.0:
      print(f"STOP: disk free < 1GB ({disk_gb:.2f}GB)", flush=True)
      break

    print(f"[{idx}/{len(DRAFT_IDS)}] Running {draft_id} ...", flush=True)
    rec = _run_single_draft(draft_id, idx, len(DRAFT_IDS))
    drafts_done[draft_id] = {
      "outcome": rec["outcome"],
      "wall_clock_s": rec["elapsed"],
      "returncode": rec["returncode"],
      "timed_out": rec["timed_out"],
      "start_iso": rec["start_iso"],
      "end_iso": rec["end_iso"],
      "business_name": rec["row"].get("business_name", ""),
      "wb_dst": rec["wb_dst"],
    }
    _save_results(results)

    # Append CSV row (always)
    _append_csv_row(rec["row"])

    # git commit + push
    push_ok = _git_commit_push(draft_id, idx, rec["row"].get("business_name", ""))
    if push_ok:
      consecutive_push_failures = 0
    else:
      consecutive_push_failures += 1
      if consecutive_push_failures >= 3:
        print("STOP: 3 consecutive git push failures", flush=True)
        break

    if rec["outcome"] == "RUNNER_ERROR":
      consecutive_runner_errors += 1
      if consecutive_runner_errors >= 3:
        print("STOP: 3 consecutive RUNNER_ERROR outcomes", flush=True)
        break
    else:
      consecutive_runner_errors = 0

    print(f"[{idx}/{len(DRAFT_IDS)}] {draft_id} {rec['outcome']} ({rec['elapsed']:.1f}s)", flush=True)

  results["sweep_finished_at"] = _now_iso()
  _save_results(results)
  gp = sum(1 for r in drafts_done.values() if r.get("outcome") == "GENUINE_PASS")
  fp = sum(1 for r in drafts_done.values() if r.get("outcome") == "FALSE_PASS")
  fl = sum(1 for r in drafts_done.values() if r.get("outcome") == "FAIL")
  re_ = sum(1 for r in drafts_done.values() if r.get("outcome") in {"RUNNER_ERROR", "WORKBOOK_ERROR"})
  print(f"DONE. GENUINE_PASS={gp} FALSE_PASS={fp} FAIL={fl} RUNNER/WB_ERROR={re_}", flush=True)
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
