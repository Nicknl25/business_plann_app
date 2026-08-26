"""mini (g): would tests/test_cw043_payroll_chain_grid.py catch the regressions the horizontal sheet can suffer, and
does the block/bridge locator survive 13 roles and a role missing a quarter?
usage: <wt_new (a HEAD worktree the tampers are applied to)> <scratch> <bluestem NEW xlsx>
Tampers are applied to the WORKTREE's schedule_sheets.py (never the real tree) and reverted with git checkout."""
import sys, os, subprocess, json, gzip, tempfile, shutil, time, re, collections
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
WT, SC, BLUESTEM = sys.argv[1], sys.argv[2], sys.argv[3]
PY = r"C:\dev\business_plann_app\.venv\Scripts\python.exe"
SS = os.path.join(WT, "client_statements_output_excel", "schedule_sheets.py")
TEST = "tests/test_cw043_payroll_chain_grid.py"


def pytest_in_wt(tag):
    p = subprocess.run([PY, "-m", "pytest", TEST, "-q", "-p", "no:cacheprovider", "-x" if False else "-q"], cwd=WT, capture_output=True, text=True, timeout=900)
    tail = [l for l in p.stdout.splitlines() if re.search(r"passed|failed|error", l)]
    fails = [l.strip() for l in p.stdout.splitlines() if l.startswith("FAILED") or l.startswith("ERROR")]
    print(f"   [{tag}] rc={p.returncode} {tail[-1] if tail else p.stdout[-300:]}")
    for f in fails[:6]:
        print("      ", f[:160])
    return p.returncode, fails


def patch(old, new, count=1):
    src = open(SS, encoding="utf-8").read()
    assert src.count(old) == count, (old, src.count(old))
    open(SS, "w", encoding="utf-8").write(src.replace(old, new))


def restore():
    subprocess.run(["git", "checkout", "--", "client_statements_output_excel/schedule_sheets.py"], cwd=WT, check=True, capture_output=True)


print("== T0 baseline in the worktree")
rc0, _ = pytest_in_wt("baseline")

print("== T1 bridge row keyed on TITLE ONLY (every column points at the first same-title block)")
patch("    R = role_layout[base_key + (ordinal,)]\n",
      "    R = next(v for k_, v in role_layout.items() if k_[:2] == base_key[:2])\n")
rc1, f1 = pytest_in_wt("T1 title-only bridge")
restore()

print("== T2 ONLY the bridge Starting FTE column points at the wrong (same-title) block; Annual Wage column stays right")
patch("    ws.cell(row=row, column=14, value=_text_ref(R[\"source\"], 2))\n",
      "    ws.cell(row=row, column=14, value=_text_ref(R[\"source\"], 2))\n"
      "    _R2 = next(v for k_, v in role_layout.items() if k_[:2] == base_key[:2])\n"
      "    ws.cell(row=row, column=5, value=f\"={local_ref(_R2['Starting FTE'], qcol)}\")\n")
rc2, f2 = pytest_in_wt("T2 wrong-block Starting FTE only")
restore()

print("== T3 a block loses its 'Average FTE' label row (label not written)")
patch("      ws.cell(row=row + 3 + n, column=1, value=label)\n",
      "      if label != 'Average FTE': ws.cell(row=row + 3 + n, column=1, value=label)\n")
rc3, f3 = pytest_in_wt("T3 missing period-row label")
restore()

print("== T4 the bridge is NOT hidden")
src = open(SS, encoding="utf-8").read()
i = src.index("  for r_ in range(bridge_note_row, max(detail_last_row, detail_header_row) + 1):\n    ws.row_dimensions[r_].hidden = True\n")
open(SS, "w", encoding="utf-8").write(src.replace("  for r_ in range(bridge_note_row, max(detail_last_row, detail_header_row) + 1):\n    ws.row_dimensions[r_].hidden = True\n", "  pass\n", 1))
rc4, f4 = pytest_in_wt("T4 visible bridge")
restore()

print("== T5 the Starting FTE chain drops its ROUND (the part-2 class)")
patch("                value=f\"=ROUND({local_ref(R['Ending FTE'], prev_col)},6)\")",
      "                value=f\"={local_ref(R['Ending FTE'], prev_col)}\")")
rc5, f5 = pytest_in_wt("T5 bare FTE chain")
restore()

# ---- locator survival: 13 roles (Bluestem NEW build) with the TEST's own helpers ----
sys.path.insert(0, os.path.join(WT, "tests")); sys.path.insert(0, WT); sys.path.insert(0, os.path.join(WT, "python"))
import openpyxl
import test_cw043_payroll_chain_grid as T
ws13 = openpyxl.load_workbook(BLUESTEM)["Payroll Schedule"]
b13 = T._role_blocks(ws13); f13, l13 = T._bridge_rows(ws13)
print(f"== T6 13-role draft (Bluestem 52cf5792) through the TEST's locator: blocks={len(b13)} bridge rows={l13 - f13 + 1}")

# ---- a role missing a quarter: Halbrook fixture with one engine row removed, built by the REAL tree, Excel-recalculated ----
print("== T7 a role missing a quarter (Halbrook, role #2 loses its q5 row) on the REAL tree")
sys.path = [p for p in sys.path if not p.startswith(WT)]
for p in (r"C:\dev\business_plann_app", r"C:\dev\business_plann_app\python"):
    sys.path.insert(0, p)
for m in list(sys.modules):
    if m.startswith("client_statements_output_excel"):
        del sys.modules[m]
from client_statements_output_excel.export_client_workbook import export_workbook_for_row
import client_statements_output_excel.schedule_sheets as ss_real
assert ss_real.__file__.lower().startswith(r"c:\dev\business_plann_app\client"), ss_real.__file__
with gzip.open(os.path.join(WT, "tests", "fixtures", "cw043_halbrook_export_row.json.gz"), "rt", encoding="utf-8") as fh:
    fx = json.load(fh)
row = dict(fx["row"]); ph = json.loads(row["payroll_headcount"]); rows = [x for x in ph["rows"] if isinstance(x, dict)]
ids = T._role_identity(rows); order = list(dict.fromkeys(k for _, k in ids)); victim = order[1]
drop = next(i for i, (q, k) in enumerate(ids) if k == victim and q == 5)
kept = [x for i, x in enumerate(rows) if i != drop]; ph["rows"] = kept; row["payroll_headcount"] = json.dumps(ph); row["business_name"] = "GAP probe"
out = os.path.join(SC, "gap"); os.makedirs(out, exist_ok=True)
path = str(export_workbook_for_row(row, output_dir=out, run_diagnostics=fx.get("run_diagnostics")))
wsF = openpyxl.load_workbook(path)["Payroll Schedule"]; blocks = T._role_blocks(wsF); first, last = T._bridge_rows(wsF)
b = blocks[1]; col5 = 3 + 5; col6 = 3 + 6
print(f"   built OK: blocks={len(blocks)} (identities {len(order)}), bridge rows={last - first + 1} == {len(kept)}: {last - first + 1 == len(kept)}")
print(f"   victim block q5 cells all blank: {all(wsF.cell(b[l], col5).value is None for l in T.PERIOD_ROWS)}; q6 Starting/Wage/Benefits are LITERALS (re-seed): {[wsF.cell(b[l], col6).value for l in ('Starting FTE', 'Annual Wage', 'Benefits %')]} engine q6: {[next(x.get(k) for x in kept if T._role_identity([x])[0][1][:3] == victim[:3] and int(x.get('quarter_index')) == 6) for k in ('starting_fte', 'annual_wage', 'payroll_taxes_benefits_percent')]}")
print(f"   q7 chains again: {wsF.cell(b['Starting FTE'], 3 + 7).value!r}")
by_id = {}
for (q, k), it in zip(T._role_identity(kept), kept):
    by_id.setdefault(k, {})[q] = it
misses = []
for blk, (ident, by_q) in zip(blocks, by_id.items()):
    vals = T._evaluate_block(wsF, blk, by_q)
    for q, item in by_q.items():
        for label, key in (("Starting FTE", "starting_fte"), ("Annual Wage", "annual_wage"), ("Benefits %", "payroll_taxes_benefits_percent")):
            if vals[label][q] != float(item.get(key) or 0.0):
                misses.append((ident[1][:16], q, key))
print(f"   the TEST's _evaluate_block reproduces the engine on every block incl. the gapped one: misses={misses[:4]} (n={len(misses)})")
import win32com.client as win32
x = win32.gencache.EnsureDispatch("Excel.Application"); x.Visible = False; x.DisplayAlerts = False
w = x.Workbooks.Open(path)
for _ in range(20):
    try:
        w.Sheets(1).Name; break
    except Exception:
        time.sleep(1.0)
x.CalculateFullRebuild(); w.Save(); w.Close(False); x.Quit()
wv = openpyxl.load_workbook(path, data_only=True); pv = wv["Payroll Schedule"]; ck = wv["Checks"]
xm = []
for i, item in enumerate(kept):
    r = first + i
    for c_, key in ((5, "starting_fte"), (6, "hires"), (9, "annual_wage"), (10, "payroll_taxes_benefits_percent")):
        if float(pv.cell(r, c_).value or 0) != float(item.get(key) or 0):
            xm.append((r, key))
pay = [(ck.cell(r, 2).value[:40], ck.cell(r, 9).value) for r in range(7, ck.max_row + 1) if isinstance(ck.cell(r, 2).value, str) and "payroll" in ck.cell(r, 2).value.lower() and ck.cell(r, 3).value == "Payroll Schedule"]
print(f"   Excel-recalculated bridge inputs == engine on the gapped draft: misses={xm[:3]} (n={len(xm)}); Checks!B2={ck['B2'].value}; payroll checks non-OK={[p for p in pay if p[1] != 'OK']}")
print("\nT-SUMMARY", {"T0": rc0, "T1": rc1, "T2": rc2, "T3": rc3, "T4": rc4, "T5": rc5})
