"""mini's (a)(b)(c)(e) audit over OLD (2e1572b) and NEW (cf87d2f) recalculated builds. usage: <scratch> <draft_prefix>..."""
import sys, os, glob, json, re, collections, openpyxl
from openpyxl.utils import get_column_letter as L
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; drafts = sys.argv[2:]
from dotenv import load_dotenv; load_dotenv(r"C:\dev\business_plann_app\.env")
import mysql.connector
c = mysql.connector.connect(host=os.getenv('MYSQL_HOST'), user=os.getenv('MYSQL_USER'), password=os.getenv('MYSQL_PASSWORD'), database=os.getenv('MYSQL_DB'), autocommit=True)
FIRST = 27; AMBER = "FDF3DF"


def by_addr(ws):
    return {(x.row, x.column): x.value for r in ws.iter_rows() for x in r if x.value is not None}


def amber(ws):
    out = set()
    for r in ws.iter_rows():
        for x in r:
            try:
                if str(x.fill.fgColor.rgb or "").upper().endswith(AMBER):
                    out.add((x.row, x.column))
            except Exception:
                pass
    return out


def labels(ws):
    return {str(ws.cell(r, 1).value).strip(): r for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, str)}


RX_E = re.compile(r"^=ROUND\(G(\d+),6\)$")
RX_FLAT = re.compile(r"^=([IJ])(\d+)$")
RX_STEP = re.compile(r"^=ROUND\(([IJ])(\d+)([+-])([0-9.eE+-]+),6\)$")
GRAND = collections.Counter()
for d in drafts:
    fo = glob.glob(f"{S}/old/OLD {d}*.xlsx")[0]; fn = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]
    vo, vn = openpyxl.load_workbook(fo, data_only=True), openpyxl.load_workbook(fn, data_only=True)
    Fo, Fn = openpyxl.load_workbook(fo), openpyxl.load_workbook(fn)
    print(f"\n######## {d}  old={os.path.basename(fo)}")
    print("sheets old==new:", vo.sheetnames == vn.sheetnames, "| Checks!B2 old/new:", vo["Checks"]["B2"].value, "/", vn["Checks"]["B2"].value)
    tv = tf = n = 0; ps_fdiff = 0
    for name in vo.sheetnames:
        a, b = by_addr(vo[name]), by_addr(vn[name]); dv = [k for k in set(a) | set(b) if a.get(k) != b.get(k)]
        fa, fb = by_addr(Fo[name]), by_addr(Fn[name]); df = [k for k in set(fa) | set(fb) if fa.get(k) != fb.get(k)]
        n += len(a); tv += len(dv)
        if name == "Payroll Schedule":
            ps_fdiff = len(df)
        else:
            tf += len(df)
        if dv or (df and name != "Payroll Schedule"):
            print(f"   DIFF {name}: value diffs={len(dv)} {[(k, a.get(k), b.get(k)) for k in sorted(dv)[:3]]} formula diffs={len(df)} {[(k, fa.get(k), fb.get(k)) for k in sorted(df)[:2]]}")
    fin = vn["FINMO"]; fl = labels(fin); prow = [k for k in fl if "payroll" in k.lower()]
    finmo_same = all([vo["FINMO"].cell(fl[k], cc).value for cc in range(1, 30)] == [fin.cell(fl[k], cc).value for cc in range(1, 30)] for k in prow)
    print(f"   (a) all sheets cells={n} value diffs={tv} formula diffs outside Payroll Schedule={tf} (Payroll Schedule formula diffs={ps_fdiff}); FINMO rows {prow} old==new: {finmo_same}")
    cur = c.cursor(dictionary=True); cur.execute("SELECT payroll_headcount FROM intake_consult_drafts WHERE draft_id LIKE %s", (d + '%',))
    rows = [x for x in json.loads(cur.fetchone()["payroll_headcount"]).get("rows") or [] if isinstance(x, dict)]; cur.close()
    ws, wso, wv = Fn["Payroll Schedule"], Fo["Payroll Schedule"], vn["Payroll Schedule"]
    last = FIRST + len(rows) - 1
    exp_prior = {}; seen = {}; cq = None; keyrow = {}
    dups = collections.defaultdict(list)
    for i, x in enumerate(rows):
        r = FIRST + i; q = int(x.get("quarter_index") or 0)
        if q != cq:
            cq, seen = q, {}
        bk = (str(x.get("staffing_class") or "").strip(), str(x.get("position_title") or "").strip(), str(x.get("person_name") or "").strip())
        o = seen.get(bk, 0); seen[bk] = o + 1; k = bk + (o,)
        p = keyrow.get(k); exp_prior[r] = p[0] if p and p[1] == q - 1 else None
        keyrow[k] = (r, q, x); dups[(q, bk[0], bk[1])].append(r)
    shapes = collections.Counter(); bad = []; fallback = 0; eng_miss = []; old_lit_bad = 0; static_same = 0
    for i, x in enumerate(rows):
        r = FIRST + i; q = int(x.get("quarter_index") or 0); p = exp_prior[r]
        E, F, I, J = (ws.cell(r, cc).value for cc in (5, 6, 9, 10))
        s, h, w, b = (float(x.get(f) or 0) for f in ("starting_fte", "hires", "annual_wage", "payroll_taxes_benefits_percent"))
        if not isinstance(F, str) and float(F or 0) == h:
            shapes["F literal"] += 1
        else:
            bad.append((r, "F", F))
        for cc in (1, 2, 3, 4, 7, 8, 11, 12, 13, 14):
            if ws.cell(r, cc).value == wso.cell(r, cc).value:
                static_same += 1
            else:
                bad.append((r, cc, wso.cell(r, cc).value, ws.cell(r, cc).value))
        if p is None:
            if isinstance(E, str) or isinstance(I, str) or isinstance(J, str):
                bad.append((r, "literal expected", E, I, J))
            else:
                shapes["E/I/J literal q1" if q == 1 else "E/I/J literal FALLBACK"] += 1
                if q > 1:
                    fallback += 1
                if float(E) != s or float(I) != w or float(J) != b:
                    bad.append((r, "literal!=engine", E, I, J))
        else:
            m = RX_E.match(str(E))
            if m and int(m.group(1)) == p:
                shapes["E =ROUND(G prior,6)"] += 1
            else:
                bad.append((r, "E", E, "expected prior", p))
            px = rows[p - FIRST]
            for cc, key, cur_v in ((9, "annual_wage", w), (10, "payroll_taxes_benefits_percent", b)):
                v = ws.cell(r, cc).value; pv = float(px.get(key) or 0); delta = cur_v - pv
                mf, ms = RX_FLAT.match(str(v)), RX_STEP.match(str(v))
                if abs(delta) <= 1e-9:
                    if mf and int(mf.group(2)) == p and mf.group(1) == L(cc):
                        shapes[f"{L(cc)} =prev"] += 1
                    else:
                        bad.append((r, cc, v, "expected flat to", p))
                else:
                    if ms and int(ms.group(2)) == p and ms.group(1) == L(cc) and (ms.group(3) == "-") == (delta < 0) and float(ms.group(4)) == abs(delta):
                        shapes[f"{L(cc)} =ROUND(prev+/-delta,6)"] += 1
                    else:
                        bad.append((r, cc, v, "expected step from", p, delta))
        for cc, key in ((5, "starting_fte"), (9, "annual_wage"), (10, "payroll_taxes_benefits_percent")):
            gv = float(wv.cell(r, cc).value or 0); ev = float(x.get(key) or 0); ov = float(vo["Payroll Schedule"].cell(r, cc).value or 0)
            if gv != ev:
                eng_miss.append((r, key, ev, gv))
            if ov != ev:
                old_lit_bad += 1
    dd = {k: v for k, v in dups.items() if len(v) > 1}
    print(f"   (b) rows={len(rows)} shapes={dict(shapes)} fallback rows={fallback} bad={bad[:4]} (n={len(bad)}); static cols identical old/new={static_same}/{len(rows)*10}")
    print(f"   (a) recalculated E/I/J == engine float-exact: misses={eng_miss[:3]} (n={len(eng_miss)}); OLD literals==engine misses={old_lit_bad}")
    if dd:
        for (q, cls, title), rs in list(dd.items())[:2]:
            refC = []
            for r in rs:
                m = RX_E.match(str(ws.cell(r, 5).value))
                refC.append(ws.cell(int(m.group(1)), 3).value if m else None)
            print(f"   (b) dup-title q{q} {cls}/{title}: rows {rs} C={[ws.cell(r,3).value for r in rs]} person={[rows[r-FIRST].get('person_name') for r in rs]} E={[ws.cell(r,5).value for r in rs]} I={[ws.cell(r,9).value for r in rs]} -> referenced rows' C={refC}")
        own = all(all(exp_prior[r] is None or rows[exp_prior[r] - FIRST].get('person_name') == rows[r - FIRST].get('person_name') for r in rs) for rs in dd.values())
        print(f"   (b) dup-title (quarter,class,title) groups={len(dd)}; every dup row chains to a row of the SAME person (and the grid matched my expected rows: bad=0 above): {own}")
    lost = {}; gained = {}
    for name in vo.sheetnames:
        ao, an = amber(Fo[name]), amber(Fn[name])
        if ao - an:
            lost[name] = sorted(ao - an)[:5]
        if an - ao:
            gained[name] = (len(an - ao), sorted(an - ao)[:3])
    po, pn = amber(Fo["Payroll Schedule"]), amber(Fn["Payroll Schedule"])
    exp_gain = {(r, cc) for r in range(FIRST, last + 1) for cc in (2, 3, 14)}
    print(f"   (c) amber lost anywhere={lost or 'none'}; gained={ {k: v[0] for k, v in gained.items()} }; Payroll amber old={len(po)} new={len(pn)}; new-old == B/C/N x detail rows: {(pn - po) == exp_gain}; old detail amber cols={sorted({cc for (r, cc) in po if FIRST <= r <= last})} new={sorted({cc for (r, cc) in pn if FIRST <= r <= last})}")
    sumifs = [(x.coordinate, x.value) for row in ws.iter_rows(min_row=1, max_row=FIRST - 1) for x in row if isinstance(x.value, str) and "SUMIFS" in x.value.upper()]
    keyA = all(re.search(r"\$?A\$?\d+:\$?A\$?\d+", f) for _, f in sumifs)
    keyBCN = [f for _, f in sumifs if re.search(r"\$?[BCN]\$?\d+:\$?[BCN]\$?\d+", f)]
    chk = Fn["Checks"]; chkv = vn["Checks"]
    notok = [(r, chkv.cell(r, 2).value, chkv.cell(r, 3).value, chkv.cell(r, 9).value) for r in range(7, chkv.max_row + 1) if chkv.cell(r, 9).value not in (None, "OK", "")]
    notok_old = [(r, vo["Checks"].cell(r, 9).value) for r in range(7, vo["Checks"].max_row + 1) if vo["Checks"].cell(r, 9).value not in (None, "OK", "")]
    payroll_chk = [(r, chkv.cell(r, 2).value, chkv.cell(r, 3).value, chkv.cell(r, 9).value) for r in range(7, chkv.max_row + 1) if any("payroll" in str(chkv.cell(r, cc).value or "").lower() for cc in (2, 3, 4))]
    refBCN = [x.value for wsx in Fn.worksheets for row in wsx.iter_rows() for x in row if isinstance(x.value, str) and x.value.startswith("=") and re.search(r"Payroll Schedule'!\$?[BCN]\$?\d", x.value)]
    localBCN = [x.value for row in ws.iter_rows() for x in row if isinstance(x.value, str) and x.value.startswith("=") and re.search(r"(?<![A-Z$'])\$?[BCN]\$?\d+", x.value)]
    print(f"   (e) summary SUMIFS n={len(sumifs)} all keyed on col A={keyA} keyed on B/C/N={len(keyBCN)}; example {sumifs[0] if sumifs else None}")
    print(f"   (e) Checks payroll rows={payroll_chk[:6]}; non-OK check rows new={notok} old={[s for _, s in notok_old]}; formulas referencing Payroll B/C/N cells: cross-sheet={len(refBCN)} local={len(localBCN)} {localBCN[:2]}")
    GRAND.update({"drafts": 1, "value diffs": tv, "formula diffs outside payroll": tf, "eng misses": len(eng_miss), "fallback": fallback, "bad": len(bad), "amber lost": sum(len(v) for v in lost.values()), "old-lit!=engine": old_lit_bad})
print("\nGRAND", dict(GRAND))
