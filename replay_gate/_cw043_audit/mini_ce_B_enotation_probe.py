import openpyxl, time, sys, win32com.client as win32
p = sys.argv[1]
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "T"
ws["C7"] = 12308.0; ws["D7"] = "=ROUND(C7+1.00000761449337e-06,6)"; ws["E7"] = "=ROUND(D7-4.8e-05,6)"; ws["F7"] = "=ROUND(E7+6.999999999999953e-06,6)"; ws["G7"] = "=ROUND(F7+1e-05,6)"
wb.save(p)
x = win32.gencache.EnsureDispatch("Excel.Application"); x.Visible=False; x.DisplayAlerts=False
w = x.Workbooks.Open(p)
for _ in range(20):
    try: w.Sheets(1).Name; break
    except Exception: time.sleep(1)
x.CalculateFullRebuild(); w.Save(); w.Close(False); x.Quit()
v = openpyxl.load_workbook(p, data_only=True)["T"]
print("e-notation formulas recalculated:", [v.cell(7,c).value for c in range(3,8)], "| expect", [12308.0, 12308.000001, 12307.999953, 12307.99996, 12307.99997])
