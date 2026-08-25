"""mini's Cash & Equity audit over OLD (b4949f3) and NEW (HEAD) recalculated builds. usage: <scratch> <draft_prefix>..."""
import sys, os, glob, json, re, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; drafts = sys.argv[2:]
sys.path.insert(0, r"C:\dev\business_plann_app"); sys.path.insert(0, r"C:\dev\business_plann_app\python")
from dotenv import load_dotenv; load_dotenv(r"C:\dev\business_plann_app\.env")
import mysql.connector
from client_statements_output_excel.data import values_21
c = mysql.connector.connect(host=os.getenv('MYSQL_HOST'), user=os.getenv('MYSQL_USER'), password=os.getenv('MYSQL_PASSWORD'), database=os.getenv('MYSQL_DB'), autocommit=True)


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def label_map(ws):
    return {ws.cell(r, 1).value.strip(): [ws.cell(r, cc).value for cc in range(2, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip()}


def by_addr(ws):
    return {(cell.row, cell.column): cell.value for r in ws.iter_rows() for cell in r if cell.value is not None}


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def checks_rows(ws):
    out = {}
    for r in range(7, ws.max_row + 1):
        key = (ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value)
        out.setdefault(key, []).append(tuple(ws.cell(r, cc).value for cc in range(5, 10)))
    return out


def refs(F):
    out = {}
    for ws in F.worksheets:
        for rr in ws.iter_rows():
            for cell in rr:
                if isinstance(cell.value, str) and "Cash Equity Schedule" in cell.value:
                    for m in re.finditer(r"'Cash Equity Schedule'!\$?[A-Z]+\$?(\d+)(?::\$?[A-Z]+\$?(\d+))?", cell.value):
                        key = (ws.title, ws.cell(cell.row, 1).value, m.group(1) + ("-" + m.group(2) if m.group(2) else ""))
                        out[key] = out.get(key, 0) + 1
    return out


for d in drafts:
    fo = glob.glob(f"{S}/old/OLD {d}*.xlsx")[0]; fn = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]
    vo, vn = openpyxl.load_workbook(fo, data_only=True), openpyxl.load_workbook(fn, data_only=True)
    Fo, Fn = openpyxl.load_workbook(fo), openpyxl.load_workbook(fn)
    print(f"\n######## {d}  old={os.path.basename(fo)}  new={os.path.basename(fn)}")
    print("sheets old==new:", vo.sheetnames == vn.sheetnames, "| Checks!B2 old/new:", vo["Checks"]["B2"].value, "/", vn["Checks"]["B2"].value)
    # (a) EXACT value compare by address on every sheet; Cash Equity + Checks by label (rows removed / shifted)
    tot = 0; ncell = 0
    for name in vo.sheetnames:
        if name in ("Cash Equity Schedule", "Checks"):
            continue
        a, b = by_addr(vo[name]), by_addr(vn[name])
        diffs = [k for k in set(a) | set(b) if a.get(k) != b.get(k)]
        ncell += len(a); tot += len(diffs)
        if diffs:
            print(f"   {name}: {len(diffs)} diffs e.g. {[(k, a.get(k), b.get(k)) for k in sorted(diffs)[:4]]}")
    lo, ln = label_map(vo["Cash Equity Schedule"]), label_map(vn["Cash Equity Schedule"])
    ce_d = [(k, lo[k][1:22], ln[k][1:22]) for k in ("Owner's Capital", "Other Equity", "Distributions") if lo[k][1:22] != ln[k][1:22]]
    print(f"   (a) by-address cells={ncell} EXACT diffs={tot} | Cash sheet label rows old={sorted(lo)} new={sorted(ln)} equity-row value diffs={ce_d}")
    co, cn = checks_rows(vo["Checks"]), checks_rows(vn["Checks"])
    gone = [k for k in co if k not in cn]; added = [k for k in cn if k not in co]
    moved = [k for k in co if k in cn and co[k] != cn[k]]
    print(f"   (a) Checks by label: rows old={sum(len(v) for v in co.values())} new={sum(len(v) for v in cn.values())} gone={gone} added={added} value-changed={len(moved)} {[(k, co[k], cn[k]) for k in moved[:3]]}")
    # (b)/(d) chain vs literal on the NEW build, vs the engine's own steps
    cur = c.cursor(dictionary=True)
    cur.execute("SELECT model_input_json, finmo_json FROM intake_consult_drafts WHERE draft_id LIKE %s", (d + '%',))
    row = cur.fetchone(); cur.close()
    mi = json.loads(row["model_input_json"]); fj = json.loads(row["finmo_json"])
    bs = {x["label"]: values_21(x.get("values")) for x in mi["sections"]["balance_sheet"] if isinstance(x, dict)}
    ceF = Fn["Cash Equity Schedule"]; ceO = Fo["Cash Equity Schedule"]
    rows = {ceF.cell(r, 1).value: r for r in range(6, 12) if ceF.cell(r, 1).value}
    for lab in ("Owner's Capital", "Other Equity", "Distributions"):
        r = rows[lab]; eng = bs[lab]
        cells = [ceF.cell(r, 3 + i).value for i in range(21)]
        lits = [i for i, v in enumerate(cells) if not is_formula(v)]
        chain_ok = all(cells[i] == f"={get_column_letter(3 + i - 1)}{r}" for i in range(21) if i not in lits)
        expect_lits = [0] + [i for i in range(1, 21) if abs(eng[i] - eng[i - 1]) > 1e-9]
        old_cells = [ceO.cell(r, 3 + i).value for i in range(21)]
        old_all_lit = all(not is_formula(v) for v in old_cells)
        lit_vals_match = all(float(cells[i]) == eng[i] for i in lits)
        recalc = [vn["Cash Equity Schedule"].cell(r, 3 + i).value for i in range(21)]
        recalc_mm = [(i, recalc[i], eng[i]) for i in range(21) if (recalc[i] or 0) != eng[i]]
        verdict = "MATCH" if (lab == "Distributions" and lits == list(range(21))) or (lab != "Distributions" and lits == expect_lits) else "MISMATCH"
        print(f"   (b/d) {lab} r{r}: literals at {lits}" + (f" expected {expect_lits}" if lab != "Distributions" else " expected ALL 21") +
              f" {verdict}; chain refs =prev ok={chain_ok}; literal values==engine exact={lit_vals_match}; old build all-literal={old_all_lit}; recalculated row==engine series exact mismatches={recalc_mm[:3]}")
    # FINMO equity rows, both builds, vs the engine series
    fin = vn["FINMO"]; finF = Fn["FINMO"]; fino = vo["FINMO"]
    frow = {fin.cell(r, 1).value: r for r in range(1, fin.max_row + 1) if fin.cell(r, 1).value}
    qr = fj.get("quarter_rows") or []
    for lab, key in (("Owner's Capital", "owners_capital"), ("Other Equity", "other_equity"), ("Distributions", "distributions")):
        r = frow[lab]
        newv = [fin.cell(r, 3 + i).value for i in range(21)]; oldv = [fino.cell(r, 3 + i).value for i in range(21)]
        eng_q = [q.get(key) for q in qr[:20]]
        mm_bs = [(i, newv[i], bs[lab][i]) for i in range(21) if (newv[i] or 0) != bs[lab][i]]
        mm_q = [(i + 1, newv[i + 1], eng_q[i]) for i in range(20) if eng_q[i] is not None and (newv[i + 1] or 0) != eng_q[i]]
        print(f"   FINMO r{r} {lab}: new==old exact={newv == oldv}; vs balance_sheet series exact mismatches={mm_bs[:3]}; vs finmo quarter_rows.{key} mismatches={mm_q[:3]} (keys present {sum(1 for v in eng_q if v is not None)}/20)")
    miF = Fn["Model Inputs"]; mrow = {miF.cell(r, 1).value: r for r in range(1, miF.max_row + 1) if miF.cell(r, 1).value}
    print("   FINMO r56 Equity K:", finF.cell(frow["Equity"], 11).value, "| r57 Distributions K:", finF.cell(frow["Distributions"], 11).value,
          "| Model Inputs K: Distributions ->", miF.cell(mrow["Distributions"], 11).value, "| Owner's Capital ->", miF.cell(mrow["Owner's Capital"], 11).value, "| Other Equity ->", miF.cell(mrow["Other Equity"], 11).value)
    # (e) who references the Cash Equity Schedule on the BUILT grids
    print("   (e) OLD refs to Cash Equity rows:", sorted(refs(Fo).items()))
    print("   (e) NEW refs to Cash Equity rows:", sorted(refs(Fn).items()))
    print("   (e) OLD Cash sheet rows 10-12 labels:", [ceO.cell(r, 1).value for r in (10, 11, 12)], "| NEW rows 10-12:", [ceF.cell(r, 1).value for r in (10, 11, 12)])
    mio, min_ = label_map(vo["Model Inputs"]), label_map(vn["Model Inputs"])
    print("   (e) Model Inputs Lease 21 values old==new:", mio["Lease"][1:22] == min_["Lease"][1:22], "annuals old==new:", mio["Lease"][22:27] == min_["Lease"][22:27],
          "| new formula C..E:", [Fn["Model Inputs"].cell(mrow["Lease"], cc).value for cc in (3, 4, 5)], "annual:", Fn["Model Inputs"].cell(mrow["Lease"], 24).value,
          "| old C:", Fo["Model Inputs"].cell(mrow["Lease"], 3).value)
    for k in [k for k in frow if "Lease" in k or "Rent" in k]:
        print(f"   (e) FINMO '{k}' old==new:", [fino.cell(frow[k], 3 + i).value for i in range(21)] == [fin.cell(frow[k], 3 + i).value for i in range(21)], "K formula:", finF.cell(frow[k], 11).value)
