"""mini's audit of the per-quarter WAGE SOURCE ROW (b35b4e8): OLD (785c3bc, per-role header cell) vs NEW (HEAD) recalculated
builds, TEXT-AWARE. usage: <scratch> <draft_prefix>...   (expects <scratch>/old/OLD <d>*.xlsx and <scratch>/new/NEW <d>*.xlsx)
(a) every sheet by address (values AND formulas, strings included) - the only admissible diffs are the OLD/NEW tag, the source row
    (old col B gone / new per-quarter cells) and bridge column N; NEW bridge N == engine per-row label; visible source row == engine
    per quarter; (b) block shapes + recalculated inputs == engine; (c) bridge 13/13 own block incl. N at qcol, empty where engine empty;
(d) SUMIFS/Checks; (e) amber: text 2/role + 1/engine row, numeric 4/engine row, nothing lost elsewhere."""
import sys, os, glob, json, re, collections, itertools, openpyxl
from openpyxl.utils import get_column_letter as L
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; drafts = sys.argv[2:]
sys.path.insert(0, r"C:\dev\business_plann_app")
from client_statements_output_excel.schedule_sheets import _wage_source_plain
from dotenv import load_dotenv; load_dotenv(r"C:\dev\business_plann_app\.env")
import mysql.connector
c = mysql.connector.connect(host=os.getenv('MYSQL_HOST'), user=os.getenv('MYSQL_USER'), password=os.getenv('MYSQL_PASSWORD'), database=os.getenv('MYSQL_DB'), autocommit=True)
AMBER = "FDF3DF"; PSC = 3
PERIOD = ["Wage source", "Starting FTE", "Hires", "Ending FTE", "Average FTE", "Annual Wage", "Benefits %", "Wage Cost", "Taxes & Benefits", "Total Payroll"]
NUMP = PERIOD[1:]
BRIDGE_COLS = {5: "Starting FTE", 6: "Hires", 7: "Ending FTE", 8: "Average FTE", 9: "Annual Wage", 10: "Benefits %", 11: "Wage Cost", 12: "Taxes & Benefits", 13: "Total Payroll"}


def by_addr(ws):
    return {(x.row, x.column): x.value for r in ws.iter_rows() for x in r if x.value is not None}


def amber(ws):
    out = set()
    for r in ws.iter_rows():
        for x in r:
            try:
                if str(x.fill.fgColor.rgb or "").upper().endswith(AMBER):
                    out.add((x.row, x.column))
            except Exception:
                pass
    return out


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def tag(v):
    return isinstance(v, str) and ("OLD " in v or "NEW " in v)


def runs(seq):
    return [(s, len(list(g))) for s, g in itertools.groupby(seq)]


def identities(rows):
    ids, seen, cq = [], {}, None
    for it in rows:
        q = int(it.get("quarter_index") or 0)
        if q != cq:
            cq, seen = q, {}
        b = (str(it.get("staffing_class") or "").strip(), str(it.get("position_title") or "").strip(), str(it.get("person_name") or "").strip())
        o = seen.get(b, 0); seen[b] = o + 1; ids.append((q, b + (o,)))
    return ids


def blocks_of(ws):
    """new geometry: header r-1, OEWS r, then PERIOD rows from r+1 (Wage source first). The OLD sheet matches the same
    locator (its 'Wage source' row is at r+1 too, carrying the per-role label in col B)."""
    out, r = [], 1
    while r <= ws.max_row:
        if ws.cell(r, 1).value == "OEWS title" and ws.cell(r + 1, 1).value == "Wage source":
            b = {"header": r - 1, "oews": r}
            for n, lab in enumerate(PERIOD):
                if ws.cell(r + 1 + n, 1).value != lab:
                    b["BROKEN"] = (r + 1 + n, ws.cell(r + 1 + n, 1).value)
                b[lab] = r + 1 + n
            out.append(b); r += 1 + len(PERIOD)
        else:
            r += 1
    return out


def eng_label(it):
    return _wage_source_plain(it.get("wage_source") or it.get("wage_source_code")) or ""


RX_REF = re.compile(r"^=([A-Z]+)(\d+)$")
RX_ROUND = re.compile(r"^=ROUND\(([A-Z]+)(\d+)(?:([+-])([0-9.eE+-]+))?,6\)$")
RX_TEXT = re.compile(r'^=IF\(([A-Z]+)(\d+)="","",\1\2\)$')
GRAND = collections.Counter()
for d in drafts:
    fo = glob.glob(f"{S}/old/OLD {d}*.xlsx")[0]; fn = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]
    vo, vn = openpyxl.load_workbook(fo, data_only=True), openpyxl.load_workbook(fn, data_only=True)
    Fo, Fn = openpyxl.load_workbook(fo), openpyxl.load_workbook(fn)
    cur = c.cursor(dictionary=True); cur.execute("SELECT business_name, payroll_headcount FROM intake_consult_drafts WHERE draft_id LIKE %s", (d + '%',)); rr = cur.fetchone(); cur.close()
    rows = [x for x in json.loads(rr["payroll_headcount"]).get("rows") or [] if isinstance(x, dict)]
    ids = identities(rows); order = []
    for q, k in ids:
        if k not in order:
            order.append(k)
    by_id = collections.defaultdict(dict)
    for (q, k), it in zip(ids, rows):
        by_id[k][q] = it
    var_roles = [k for k in order if len({eng_label(it) for it in by_id[k].values()}) > 1]
    print(f"\n######## {d} {rr['business_name']!r} engine rows={len(rows)} identities={len(order)} roles whose label VARIES by quarter={[(k[1] or k[2]) for k in var_roles]}")
    print("   sheets old==new:", vo.sheetnames == vn.sheetnames, "| Checks!B2 old/new:", vo["Checks"]["B2"].value, "/", vn["Checks"]["B2"].value)
    # ---------- (a) every OTHER sheet by address, values AND formulas, strings included ----------
    tv = tf = n = 0; ex = []
    for name in vo.sheetnames:
        if name == "Payroll Schedule":
            continue
        for A, B, kind in ((by_addr(vo[name]), by_addr(vn[name]), "value"), (by_addr(Fo[name]), by_addr(Fn[name]), "formula")):
            dd = [k for k in set(A) | set(B) if A.get(k) != B.get(k) and not (tag(A.get(k)) or tag(B.get(k)))]
            if kind == "value":
                n += len(A); tv += len(dd)
            else:
                tf += len(dd)
            ex += [(name, kind, L(k[1]) + str(k[0]), A.get(k), B.get(k)) for k in sorted(dd)[:2]]
    print(f"   (a) OTHER sheets by address, strings included: cells={n} value diffs={tv} formula diffs={tf} {ex[:3]}")
    hdr = [vn["Checks"].cell(6, cc).value for cc in range(1, 12)]
    stat_col = next((i + 1 for i, h in enumerate(hdr) if str(h).strip().lower() == "status"), 9)
    notok_o = [vo["Checks"].cell(r, 2).value for r in range(7, vo["Checks"].max_row + 1) if vo["Checks"].cell(r, stat_col).value not in (None, "OK", "")]
    notok_n = [vn["Checks"].cell(r, 2).value for r in range(7, vn["Checks"].max_row + 1) if vn["Checks"].cell(r, stat_col).value not in (None, "OK", "")]
    print(f"   (a) Checks non-OK rows old={len(notok_o)} new={len(notok_n)} same by label={notok_o==notok_n}: {[str(x)[:38] for x in notok_n]}")
    # ---------- (a) Payroll Schedule by address: classify EVERY diff ----------
    po, pn, Po, Pn = vo["Payroll Schedule"], vn["Payroll Schedule"], Fo["Payroll Schedule"], Fn["Payroll Schedule"]
    bo, bn = blocks_of(Po), blocks_of(Pn)
    src_rows_o = {b["Wage source"] for b in bo}; src_rows_n = {b["Wage source"] for b in bn}
    oh = [r for r in range(1, Po.max_row + 1) if Po.cell(r, 1).value == "Quarter"][-1] + 1
    nh = [r for r in range(1, Pn.max_row + 1) if Pn.cell(r, 1).value == "Quarter"][-1] + 1
    new_n = 0
    while isnum(Pn.cell(nh + new_n, 1).value):
        new_n += 1
    old_n = 0
    while isnum(Po.cell(oh + old_n, 1).value):
        old_n += 1
    cls = collections.Counter(); other = []
    for A, B, kind in ((by_addr(po), by_addr(pn), "value"), (by_addr(Po), by_addr(Pn), "formula")):
        for k in sorted(set(A) | set(B)):
            x, y = A.get(k), B.get(k)
            if x == y or tag(x) or tag(y):
                continue
            r, cc = k
            if r in src_rows_o and cc == 2 and y is None:
                cls[f"{kind}: source row col B gone"] += 1
            elif r in src_rows_n and cc > PSC and x is None:
                cls[f"{kind}: source row per-quarter cell added"] += 1
            elif nh <= r < nh + new_n and cc == 14:
                cls[f"{kind}: bridge col N"] += 1
            else:
                other.append((kind, L(cc) + str(r), x, y))
    print(f"   (a) Payroll Schedule by address (values+formulas): same geometry old==new: blocks {len(bo)}=={len(bn)} bridge from {oh}=={nh} rows {old_n}=={new_n}=={len(rows)}; classified diffs={dict(cls)}; UNCLASSIFIED={other[:3]} (n={len(other)})")
    # bridge N vs engine label, old and new; visible source row vs engine
    blk_of = dict(zip(order, bn))
    old_bad = new_bad = vis_bad = 0; ex_lbl = []; vis_extra = 0
    for i, ((q, k), it) in enumerate(zip(ids, rows)):
        e = eng_label(it)
        if (po.cell(oh + i, 14).value or "") != e:
            old_bad += 1
            if len(ex_lbl) < 2:
                ex_lbl.append((q, k[1] or k[2], "old N", po.cell(oh + i, 14).value, "engine", e))
        if (pn.cell(nh + i, 14).value or "") != e:
            new_bad += 1
        vis = pn.cell(blk_of[k]["Wage source"], PSC + q).value
        if (vis or "") != e:
            vis_bad += 1
    for k, b in blk_of.items():
        for q in range(1, 21):
            if q not in by_id[k] and pn.cell(b["Wage source"], PSC + q).value not in (None, ""):
                vis_extra += 1
    print(f"   (a) bridge col N vs the engine's per-row label: OLD disagreed on {old_bad}, NEW on {new_bad}; visible Wage source row vs engine per quarter: misses={vis_bad}, cells where the engine has no row={vis_extra}; e.g. {ex_lbl}")
    for k in var_roles:
        b = blk_of[k]
        print(f"       visible row for {k[1] or k[2]!r} (header row {b['header']}, source row {b['Wage source']}): {runs([pn.cell(b['Wage source'], PSC + q).value for q in range(1, 21)])}")
        bi = [i for i, (q, kk) in enumerate(ids) if kk == k]
        print(f"       bridge col N for the same rows (recalculated) NEW: {runs([pn.cell(nh + i, 14).value for i in bi])}; OLD: {runs([po.cell(oh + i, 14).value for i in bi])}")
    # numeric bridge value-for-value old vs new + summary values
    vd = sum(1 for i in range(len(rows)) for cc in range(1, 14) if po.cell(oh + i, cc).value != pn.cell(nh + i, cc).value)
    sec_n = [r for r in range(1, Pn.max_row + 1) if str(Pn.cell(r, 1).value).startswith("Payroll Detail")][0]
    sv = sum(1 for r in range(1, sec_n) for cc in range(1, 30) if po.cell(r, cc).value != pn.cell(r, cc).value and not (tag(po.cell(r, cc).value) or tag(pn.cell(r, cc).value)))
    fin = vn["FINMO"]; flab = {str(fin.cell(r, 1).value).strip(): r for r in range(1, fin.max_row + 1) if isinstance(fin.cell(r, 1).value, str)}
    prow = [k for k in flab if "payroll" in k.lower()]
    finmo_same = all([vo["FINMO"].cell(flab[k], cc).value for cc in range(1, 30)] == [fin.cell(flab[k], cc).value for cc in range(1, 30)] for k in prow)
    print(f"   (a) bridge cols A..M value-for-value old==new diffs={vd}; summary rows 1..{sec_n-1} value diffs={sv}; FINMO {prow} identical={finmo_same}")
    # ---------- (b) blocks ----------
    blocks = bn
    broken = [b for b in blocks if "BROKEN" in b]
    print(f"   (b) blocks={len(blocks)} == identities={len(order)}: {len(blocks)==len(order)}; broken={broken[:2]}; last block row {blocks[-1]['Total Payroll']} < bridge {nh}: {blocks[-1]['Total Payroll'] < nh}")
    shapes = collections.Counter(); bad = []; rec_miss = []; hdr_bad = []
    for bi, (k, b) in enumerate(zip(order, blocks)):
        cls_, title, person, ordn = k
        if Pn.cell(b["header"], 1).value != (title or person):
            hdr_bad.append((bi, Pn.cell(b["header"], 1).value, title, person))
        byq = by_id[k]; qs = sorted(byq)
        for q in range(1, 21):
            col = PSC + q
            cells = {lab: Pn.cell(b[lab], col).value for lab in NUMP}
            if q not in byq:
                if any(v is not None for v in cells.values()):
                    bad.append((bi, q, "cell present with no engine row", cells))
                continue
            it = byq[q]; prior = byq.get(q - 1)
            s, h, w, bn_ = (float(it.get(f) or 0) for f in ("starting_fte", "hires", "annual_wage", "payroll_taxes_benefits_percent"))
            if isinstance(cells["Hires"], str) or float(cells["Hires"] or 0) != h:
                bad.append((bi, q, "Hires", cells["Hires"], h))
            else:
                shapes["Hires literal"] += 1
            if prior is None:
                for lab, ev in (("Starting FTE", s), ("Annual Wage", w), ("Benefits %", bn_)):
                    v = cells[lab]
                    if isinstance(v, str) or float(v or 0) != ev:
                        bad.append((bi, q, lab + " literal!=engine", v, ev))
                    else:
                        shapes[f"{lab} literal " + ("first" if q == qs[0] else "RESEED")] += 1
            else:
                m = RX_ROUND.match(str(cells["Starting FTE"]))
                if m and m.group(1) == L(col - 1) and int(m.group(2)) == b["Ending FTE"] and m.group(3) is None:
                    shapes["Starting =ROUND(prev Ending,6)"] += 1
                else:
                    bad.append((bi, q, "Starting FTE", cells["Starting FTE"]))
                for lab, ev, key in (("Annual Wage", w, "annual_wage"), ("Benefits %", bn_, "payroll_taxes_benefits_percent")):
                    v = str(cells[lab]); pv = float(prior.get(key) or 0); delta = ev - pv
                    mf, ms = RX_REF.match(v), RX_ROUND.match(v)
                    if abs(delta) <= 1e-9:
                        if mf and mf.group(1) == L(col - 1) and int(mf.group(2)) == b[lab]:
                            shapes[f"{lab} =prev"] += 1
                        else:
                            bad.append((bi, q, lab, v, "expected flat"))
                    else:
                        if ms and ms.group(1) == L(col - 1) and int(ms.group(2)) == b[lab] and ms.group(3) and (ms.group(3) == "-") == (delta < 0) and float(ms.group(4)) == abs(delta):
                            shapes[f"{lab} =ROUND(prev+/-delta,6)"] += 1
                        else:
                            bad.append((bi, q, lab, v, "expected step", delta))
            C = L(col)
            exp = {"Ending FTE": f"={C}{b['Starting FTE']}+{C}{b['Hires']}", "Average FTE": f"=({C}{b['Starting FTE']}+{C}{b['Ending FTE']})/2",
                   "Wage Cost": f"={C}{b['Average FTE']}*{C}{b['Annual Wage']}/4", "Taxes & Benefits": f"={C}{b['Wage Cost']}*{C}{b['Benefits %']}", "Total Payroll": f"={C}{b['Wage Cost']}+{C}{b['Taxes & Benefits']}"}
            for lab, e in exp.items():
                if cells[lab] == e:
                    shapes["derived formula exact"] += 1
                else:
                    bad.append((bi, q, lab, cells[lab], e))
            for lab, ev in (("Starting FTE", s), ("Hires", h), ("Annual Wage", w), ("Benefits %", bn_)):
                gv = pn.cell(b[lab], col).value
                if float(gv or 0) != ev:
                    rec_miss.append((bi, q, lab, ev, gv))
    print(f"   (b) header mismatches={len(hdr_bad)}; shapes={dict(shapes)}; bad={bad[:3]} (n={len(bad)}); recalculated inputs == engine float-exact misses={len(rec_miss)}")
    # ---------- (c) bridge ----------
    hidden_rows = [r for r in range(nh - 2, nh + new_n) if Pn.row_dimensions[r].hidden]
    cbad = []; nonformula = 0; own_ok = 0; empty_ok = 0; empty_bad = []
    for i, ((q, k), it) in enumerate(zip(ids, rows)):
        r = nh + i; b = blk_of[k]; col = PSC + q
        if Pn.cell(r, 1).value != q:
            cbad.append((r, "A", Pn.cell(r, 1).value, q))
        for cc, (brow, bcol) in {2: (b["header"], 2), 3: (b["header"], 1), 4: (b["oews"], 2), 14: (b["Wage source"], col)}.items():
            v = str(Pn.cell(r, cc).value); m = RX_TEXT.match(v)
            if not (m and m.group(1) == L(bcol) and int(m.group(2)) == brow):
                cbad.append((r, cc, v, "expected", f"{L(bcol)}{brow}"))
        for cc, lab in BRIDGE_COLS.items():
            v = Pn.cell(r, cc).value; m = RX_REF.match(str(v))
            if not (isinstance(v, str) and v.startswith("=")):
                nonformula += 1
            if not (m and m.group(1) == L(col) and int(m.group(2)) == b[lab]):
                cbad.append((r, cc, v, "expected", f"={L(col)}{b[lab]}"))
        if not any(x[0] == r for x in cbad):
            own_ok += 1
        for cc in (4, 14):
            e = (it.get("oews_occ_title") or it.get("oews_matched_title") or "") if cc == 4 else eng_label(it)
            if e == "":
                if pn.cell(r, cc).value in (None, ""):
                    empty_ok += 1
                else:
                    empty_bad.append((r, cc, pn.cell(r, cc).value))
    print(f"   (c) hidden note+header+{new_n} rows={len(hidden_rows)==new_n+2}; bridge rows with all 13 cells on their OWN block (N at its quarter column): {own_ok}/{len(ids)}; non-formula={nonformula}; bad={cbad[:2]} (n={len(cbad)}); engine-empty text cells: empty on the sheet={empty_ok}, not empty={empty_bad[:2]} (n={len(empty_bad)})")
    # ---------- (d) SUMIFS + checks ----------
    sumifs = [(r, cc, Pn.cell(r, cc).value) for r in range(1, sec_n) for cc in range(1, 30) if isinstance(Pn.cell(r, cc).value, str) and "SUMIFS" in Pn.cell(r, cc).value.upper()]
    keyA = [f for _, _, f in sumifs if re.search(rf"\$A\${nh}:\$A\${nh+new_n-1}", f)]
    chkF = Fn["Checks"]; pay = []; fc = []
    for r in range(7, chkF.max_row + 1):
        li = str(chkF.cell(r, 2).value or "")
        if "payroll" in li.lower() and str(chkF.cell(r, 3).value) == "Payroll Schedule":
            pay.append((li[:30], vn["Checks"].cell(r, stat_col).value))
        if li.startswith("Payroll Detail - "):
            m = re.search(r"!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", str(chkF.cell(r, 4).value)); colL, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
            mine = sum(1 for rr_ in range(r1, r2 + 1) if isinstance(Pn[f"{colL}{rr_}"].value, str) and Pn[f"{colL}{rr_}"].value.startswith("="))
            fc.append((li[18:], (r1, r2) == (nh, nh + new_n - 1), chkF.cell(r, 6).value == mine))
    print(f"   (d) SUMIFS keyed on bridge $A${nh}:$A${nh+new_n-1}: {len(keyA)}/{len(sumifs)}; payroll Checks non-OK={[p for p in pay if p[1] != 'OK']} of {len(pay)}; formula-count checks (range==bridge, expected==my count): {fc}")
    # ---------- (e) amber ----------
    lost = {}; gained = {}
    for name in vo.sheetnames:
        if name == "Payroll Schedule":
            continue
        ao, an = amber(Fo[name]), amber(Fn[name])
        if ao - an:
            lost[name] = sorted(ao - an)[:3]
        if an - ao:
            gained[name] = sorted(an - ao)[:3]
    ap_o, ap_n = amber(Po), amber(Pn)
    inp_rows = {b[lab] for b in blocks for lab in ("Starting FTE", "Hires", "Annual Wage", "Benefits %")}
    hdr_rows = {b["header"] for b in blocks}; src_rows = {b["Wage source"] for b in blocks}
    num_n = sum(1 for (r, cc) in ap_n if cc > PSC and r in inp_rows)
    txt_hdr = sum(1 for (r, cc) in ap_n if cc <= 2 and r in hdr_rows)
    txt_src = sum(1 for (r, cc) in ap_n if cc > PSC and r in src_rows)
    src_at_engine = all((b["Wage source"], PSC + q) in ap_n for k, b in blk_of.items() for q in by_id[k])
    other_n = len(ap_n) - num_n - txt_hdr - txt_src
    num_o = sum(1 for (r, cc) in ap_o if cc > PSC and r in {b[lab] for b in bo for lab in ("Starting FTE", "Hires", "Annual Wage", "Benefits %")})
    txt_o = sum(1 for (r, cc) in ap_o if cc <= 2 and r in ({b["header"] for b in bo} | {b["Wage source"] for b in bo}))
    bridge_amber = sum(1 for (r, cc) in ap_n if nh - 2 <= r < nh + new_n)
    above_o = {k for k in ap_o if k[0] < sec_n}; above_n = {k for k in ap_n if k[0] < sec_n}
    print(f"   (e) other sheets amber lost={lost or 'none'} gained={gained or 'none'}; numeric amber old={num_o} new={num_n} == 4 x rows {4*len(rows)}: {num_n==4*len(rows)}; text amber old={txt_o} (3/role) new header={txt_hdr} == 2 x roles {2*len(order)}: {txt_hdr==2*len(order)}, source={txt_src} == engine rows {len(rows)}: {txt_src==len(rows)}, every engine quarter amber={src_at_engine}; bridge amber={bridge_amber}; unclassified amber={other_n}; amber above detail old==new: {above_o==above_n}")
    GRAND.update({"drafts": 1, "other-sheet value diffs": tv, "other-sheet formula diffs": tf, "payroll unclassified diffs": len(other), "bridge N != engine NEW": new_bad, "bridge N != engine OLD": old_bad, "visible source misses": vis_bad, "bridge A..M diffs": vd, "summary diffs": sv, "block bad": len(bad), "recalc misses": len(rec_miss), "bridge bad": len(cbad), "empty-text bad": len(empty_bad), "amber lost": sum(len(v) for v in lost.values()), "unclassified amber": other_n, "rows": len(rows), "roles": len(order)})
print("\nGRAND", dict(GRAND))
