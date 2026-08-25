"""(c) demonstration on chosen quarters: type a new Owner's Capital level into column <col> on OLD and NEW builds, recalc, read FINMO. usage: <scratch> draft:col ..."""
import sys, os, glob, shutil, time, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]


def recalc(path):
    import win32com.client as win32
    x = win32.gencache.EnsureDispatch("Excel.Application"); x.Visible = False; x.DisplayAlerts = False
    w = x.Workbooks.Open(path)
    for _ in range(20):
        try:
            w.Sheets(1).Name; break
        except Exception:
            time.sleep(1.5)
    x.CalculateFullRebuild(); w.Save(); w.Close(False); x.Quit()


os.makedirs(f"{S}/demo", exist_ok=True)
for spec in sys.argv[2:]:
    d, col = spec.split(":"); col = int(col)
    for tag in ("old", "new"):
        src = glob.glob(f"{S}/{tag}/{tag.upper()} {d}*.xlsx")[0]
        dst = f"{S}/demo/{tag}_{d}_{get_column_letter(col)}_demo.xlsx"; shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst); ce = wb["Cash Equity Schedule"]
        r = [rr for rr in range(6, 12) if ce.cell(rr, 1).value == "Owner's Capital"][0]
        before = ce.cell(r, col).value
        formulas = [ce.cell(r, cc).value for cc in range(col - 1, min(col + 4, 24))]
        v0 = openpyxl.load_workbook(src, data_only=True)
        base = v0["Cash Equity Schedule"].cell(r, col).value
        new_level = round(float(base) + 40000.0, 2)
        ce.cell(r, col).value = new_level
        wb.save(dst); recalc(dst)
        v = openpyxl.load_workbook(dst, data_only=True); fin = v["FINMO"]
        frow = {fin.cell(rr, 1).value: rr for rr in range(1, fin.max_row + 1) if fin.cell(rr, 1).value}
        cols = range(col - 2, min(col + 5, 24))
        q = [f"Q{cc-3}" for cc in cols]
        oc = [fin.cell(frow["Owner's Capital"], cc).value for cc in cols]
        oc0 = [v0["FINMO"].cell(frow["Owner's Capital"], cc).value for cc in cols]
        tle = [fin.cell(frow["Total Liabilities & Equity"], cc).value for cc in cols]
        tle0 = [v0["FINMO"].cell(frow["Total Liabilities & Equity"], cc).value for cc in cols]
        ta = [fin.cell(frow["Total Assets"], cc).value for cc in cols]
        eq = [fin.cell(frow["Equity"], cc).value for cc in cols]
        print(f"{d} {tag.upper()} {get_column_letter(col)}{r} (Q{col-3}): cell was {before!r} value {base} -> typed {new_level}; sheet formulas {get_column_letter(col-1)}..: {formulas}")
        print(f"   quarters               {q}")
        print(f"   Owner's Capital before {[round(x or 0) for x in oc0]}")
        print(f"   Owner's Capital after  {[round(x or 0) for x in oc]}")
        print(f"   Equity flow after      {[round(x or 0) for x in eq]}")
        print(f"   Total L&E delta        {[round((a or 0)-(b or 0)) for a, b in zip(tle, tle0)]}   A=L+E diffs {[round((a or 0)-(b or 0), 6) for a, b in zip(ta, tle)]}  Checks!B2 {v['Checks']['B2'].value}")
