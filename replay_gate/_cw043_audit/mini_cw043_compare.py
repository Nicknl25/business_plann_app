"""mini's own three-way compare: delivered vs BEFOREPRIME vs AFTER, formulas AND values."""
import glob, json, sys, math
import openpyxl
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
SC = sys.argv[1]
DELIVERED = r'C:\dev\Cilient Plans\Halbrook Grounds Management -- 08-24-2026 17-57-48.xlsx'
BP = glob.glob(SC + r'/cw043/MINI BEFOREPRIME*.xlsx')[0]; AF = glob.glob(SC + r'/cw043/MINI AFTER*.xlsx')[0]
def cells(ws):
    return {(c.row, c.column): c.value for row in ws.iter_rows() for c in row if c.value is not None}
def isnum(v): return isinstance(v, (int, float)) and not isinstance(v, bool)
def same(a, b, exact=False):
    if isnum(a) and isnum(b):
        return a == b if exact else abs(float(a)-float(b)) <= max(1e-6, abs(float(a))*1e-9)
    return a == b
wd_f, wb_f, wa_f = (openpyxl.load_workbook(p) for p in (DELIVERED, BP, AF))
wd_v, wb_v, wa_v = (openpyxl.load_workbook(p, data_only=True) for p in (DELIVERED, BP, AF))
print("sheets equal:", wd_f.sheetnames == wb_f.sheetnames == wa_f.sheetnames)

print("\n=== LEG 1: delivered vs BEFOREPRIME (sheet rewrite only) - FORMULAS and VALUES, every sheet except Debt Schedule ===")
tot_f = tot_v = 0
for name in wd_f.sheetnames:
    if name == "Debt Schedule": continue
    A, B = cells(wd_f[name]), cells(wb_f[name]); Av, Bv = cells(wd_v[name]), cells(wb_v[name])
    df = [k for k in set(A)|set(B) if not same(A.get(k), B.get(k))]
    df = [k for k in df if not (isinstance(A.get(k), str) and ('Halbrook' in A.get(k) or 'MINI' in str(B.get(k))))]
    dv = [k for k in set(Av)|set(Bv) if not same(Av.get(k), Bv.get(k))]
    dv = [k for k in dv if not (isinstance(Av.get(k), str) and ('Halbrook' in Av.get(k) or 'MINI' in str(Bv.get(k))))]
    tot_f += len(df); tot_v += len(dv)
    print(f"  {name:22s} formulas: {len(A):5d} cells, {len(df)} diffs | values: {len(Av):5d} cells, {len(dv)} diffs" + (f"  e.g. {sorted(dv)[:3]} {[(Av.get(k),Bv.get(k)) for k in sorted(dv)[:2]]}" if dv else ""))
print(f"  TOTAL off-Debt-Schedule diffs: formulas={tot_f} values={tot_v}")
# Debt Schedule: map old->new columns, compare VALUES on the whole table, and list what's new
OLD = dict(d_open=2,d_new=3,d_rate=4,d_pay=5,d_int=6,d_prin=7,d_close=8,l_open=9,l_add=10,l_rate=11,l_term=12,l_pay=13,l_int=14,l_prin=15,l_close=16,t_pay=17,t_int=18,t_close=19)
NEW = dict(d_open=2,d_new=3,d_rate=4,d_pay=10,d_int=9,d_prin=8,d_close=11,l_open=12,l_add=13,l_rate=14,l_term=15,l_pay=20,l_int=19,l_prin=18,l_close=21,t_pay=22,t_int=23,t_close=24)
dd, db = wd_v["Debt Schedule"], wb_v["Debt Schedule"]
bad = 0; exact_bad = 0
for key in OLD:
    for r in range(7, 28):
        a, b = dd.cell(r, OLD[key]).value or 0, db.cell(r, NEW[key]).value or 0
        if not same(a, b): bad += 1; print(f"  DebtSched VALUE diff {key} r{r}: {a} vs {b}")
        if not same(a, b, exact=True): exact_bad += 1
print(f"  Debt Schedule mapped VALUES: {bad} diffs (tolerance), {exact_bad} diffs (bit-exact)")
# bridge rows (below table) values: compare by label
def bridge(ws):
    out = {}
    for r in range(28, ws.max_row+1):
        lab = ws.cell(r,1).value
        if isinstance(lab, str) and lab.strip(): out[lab.strip()] = [ws.cell(r,c).value for c in range(3, 24)]
    return out
bd, bb = bridge(dd), bridge(db)
print(f"  bridge labels delivered={len(bd)} beforeprime={len(bb)} same-set={set(bd)==set(bb)}")
bb_bad = sum(1 for k in bd for i,(x,y) in enumerate(zip(bd[k], bb.get(k, []))) if not same(x or 0, y or 0))
print(f"  bridge VALUE diffs by label: {bb_bad}")

print("\n=== LEG 2: BEFOREPRIME vs AFTER (the lease fix isolated) - VALUES moved per sheet ===")
for name in wb_f.sheetnames:
    Av, Bv = cells(wb_v[name]), cells(wa_v[name])
    dv = [k for k in set(Av)|set(Bv) if not same(Av.get(k), Bv.get(k)) and not isinstance(Av.get(k), str)]
    A, B = cells(wb_f[name]), cells(wa_f[name])
    df = [k for k in set(A)|set(B) if A.get(k) != B.get(k) and isinstance(A.get(k), str) and str(A.get(k)).startswith("=")]
    print(f"  {name:22s} values moved: {len(dv):4d}   formulas changed: {len(df)}")
db_, da = wb_v["Debt Schedule"], wa_v["Debt Schedule"]
moved = [(r,c,db_.cell(r,c).value,da.cell(r,c).value) for c in range(2,12) for r in range(7,28) if not same(db_.cell(r,c).value or 0, da.cell(r,c).value or 0)]
print(f"  DEBT BLOCK (B..K, rows 7-27) moved cells: {len(moved)} {moved[:3]}")

print("\n=== CLAIM 4: recalculated AFTER sheet vs the ENGINE (finmo_after quarter_rows) ===")
fa = json.load(open(SC + '/cw043/finmo_after.json')); rows = fa["quarter_rows"]
print("  finmo keys sample:", [k for k in rows[1].keys() if 'lease' in k or 'debt' in k or 'repay' in k][:14])
def col(ws, c): return [ws.cell(7+i, c).value for i in range(21)]
lease_prin, lease_close, lease_int = col(da, 18), col(da, 21), col(da, 19)
debt_prin, debt_close, debt_sched, debt_extra = col(da, 8), col(da, 11), col(da, 6), col(da, 7)
def cmp(label, sheet_vals, key):
    eng = [rows[i].get(key) for i in range(21)]
    if all(e is None for e in eng): print(f"  {label}: engine key {key!r} absent"); return
    diffs = [abs(float(s or 0)-float(e or 0)) for s,e in zip(sheet_vals, eng)]
    exact = sum(1 for s,e in zip(sheet_vals, eng) if float(s or 0) == float(e or 0))
    print(f"  {label:28s} max|sheet-engine|={max(diffs):.3e}  bit-exact {exact}/21  sheet[Q1,Q10,Q20]={[round(float(sheet_vals[i] or 0),4) for i in (1,10,20)]} engine={[round(float(eng[i] or 0),4) for i in (1,10,20)]}")
for label, vals, key in (("lease principal", lease_prin, "lease_principal_repayments_total"), ("lease closing", lease_close, "lease_closing_balance_total"),
                         ("lease interest", lease_int, "lease_interest_expense_total"), ("debt principal", debt_prin, "actual_debt_repayment"), ("debt closing", debt_close, "closing_debt")):
    cmp(label, vals, key)
# fallback: search keys by fuzzy
print("  debt sched+extra vs principal (bit-exact in the recalculated sheet):", sum(1 for i in range(21) if float(debt_sched[i] or 0)+float(debt_extra[i] or 0) == float(debt_prin[i] or 0)), "/21")
print("  debt extra column:", [round(float(v or 0),4) for v in debt_extra])
print("  lease extra column:", [round(float(v or 0),6) for v in col(da, 17)])
print("  lease rows AFTER Q1/Q10/Q20 [open,sched,extra,prin,int,pay,close]:")
for r in (8, 17, 27): print("   ", [round(float(da.cell(r,c).value or 0),4) for c in (12,16,17,18,19,20,21)])
print("  ROU asset closing Q20 (bridge 'Right-of-Use Asset Closing'):", [ (k, [round(float(x or 0),2) for x in v[-3:]]) for k,v in bridge(da).items() if 'Right-of-Use Asset Closing' in k])

print("\n=== VERDICT SURFACES ===")
for tag, w in (("delivered", wd_v), ("BEFOREPRIME", wb_v), ("AFTER", wa_v)):
    fails = [(n, c.coordinate) for n in w.sheetnames for row in w[n].iter_rows() for c in row if isinstance(c.value, str) and c.value.strip() == "FAIL"]
    dg = w["Diagnostics"]; diag = [(str(dg.cell(r,1).value), dg.cell(r,2).value) for r in range(1, dg.max_row+1) if str(dg.cell(r,1).value or "") in ("Verdict","Score","Acceptance","acceptance_passed") or "capital_lease" in str(dg.cell(r,1).value or "")]
    print(f"  {tag:12s} Checks!B2={w['Checks']['B2'].value!r}  FAIL cells={len(fails)} {fails[:4]}  Diagnostics={diag}")
