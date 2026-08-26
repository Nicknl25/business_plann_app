"""mini's (a)-(e) audit of the HORIZONTAL Payroll Schedule (785c3bc): OLD (730cb98) vs NEW (HEAD) recalculated builds.
usage: <scratch> <draft_prefix>...   (expects <scratch>/old/OLD <d>*.xlsx and <scratch>/new/NEW <d>*.xlsx)"""
import sys, os, glob, json, re, collections, openpyxl
from openpyxl.utils import get_column_letter as L
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; drafts = sys.argv[2:]
from dotenv import load_dotenv; load_dotenv(r"C:\dev\business_plann_app\.env")
import mysql.connector
c = mysql.connector.connect(host=os.getenv('MYSQL_HOST'), user=os.getenv('MYSQL_USER'), password=os.getenv('MYSQL_PASSWORD'), database=os.getenv('MYSQL_DB'), autocommit=True)
AMBER = "FDF3DF"; PSC = 3
PERIOD = ["Starting FTE", "Hires", "Ending FTE", "Average FTE", "Annual Wage", "Benefits %", "Wage Cost", "Taxes & Benefits", "Total Payroll"]
BRIDGE_COLS = {5: "Starting FTE", 6: "Hires", 7: "Ending FTE", 8: "Average FTE", 9: "Annual Wage", 10: "Benefits %", 11: "Wage Cost", 12: "Taxes & Benefits", 13: "Total Payroll"}


def by_addr(ws):
    return {(x.row, x.column): x.value for r in ws.iter_rows() for x in r if x.value is not None}


def amber(ws):
    out = set()
    for r in ws.iter_rows():
        for x in r:
            try:
                if str(x.fill.fgColor.rgb or "").upper().endswith(AMBER):
                    out.add((x.row, x.column))
            except Exception:
                pass
    return out


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def identities(rows):
    ids, seen, cq = [], {}, None
    for it in rows:
        q = int(it.get("quarter_index") or 0)
        if q != cq:
            cq, seen = q, {}
        b = (str(it.get("staffing_class") or "").strip(), str(it.get("position_title") or "").strip(), str(it.get("person_name") or "").strip())
        o = seen.get(b, 0); seen[b] = o + 1; ids.append((q, b + (o,)))
    return ids


def blocks_of(ws):
    out, r = [], 1
    while r <= ws.max_row:
        if ws.cell(r, 1).value == "OEWS title" and ws.cell(r + 1, 1).value == "Wage source":
            b = {"header": r - 1, "oews": r, "source": r + 1}
            for n, lab in enumerate(PERIOD):
                if ws.cell(r + 2 + n, 1).value != lab:
                    b["BROKEN"] = (r + 2 + n, ws.cell(r + 2 + n, 1).value)
                b[lab] = r + 2 + n
            out.append(b); r += 2 + len(PERIOD)
        else:
            r += 1
    return out


RX_REF = re.compile(r"^=([A-Z]+)(\d+)$")
RX_ROUND = re.compile(r"^=ROUND\(([A-Z]+)(\d+)(?:([+-])([0-9.eE+-]+))?,6\)$")
RX_TEXT = re.compile(r'^=IF\(([A-Z]+)(\d+)="","",\1\2\)$')
GRAND = collections.Counter()
for d in drafts:
    fo = glob.glob(f"{S}/old/OLD {d}*.xlsx")[0]; fn = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]
    vo, vn = openpyxl.load_workbook(fo, data_only=True), openpyxl.load_workbook(fn, data_only=True)
    Fo, Fn = openpyxl.load_workbook(fo), openpyxl.load_workbook(fn)
    cur = c.cursor(dictionary=True); cur.execute("SELECT business_name, payroll_headcount FROM intake_consult_drafts WHERE draft_id LIKE %s", (d + '%',)); rr = cur.fetchone(); cur.close()
    rows = [x for x in json.loads(rr["payroll_headcount"]).get("rows") or [] if isinstance(x, dict)]
    ids = identities(rows); order = []
    for q, k in ids:
        if k not in order:
            order.append(k)
    print(f"\n######## {d} {rr['business_name']!r} engine rows={len(rows)} identities={len(order)}")
    print("   sheets old==new:", vo.sheetnames == vn.sheetnames, "| Checks!B2 old/new:", vo["Checks"]["B2"].value, "/", vn["Checks"]["B2"].value)
    # ---------- (a) every other sheet by address (values + formulas) ----------
    tv = tf = n = 0; fdiffs = {}
    for name in vo.sheetnames:
        if name == "Payroll Schedule":
            continue
        a, b = by_addr(vo[name]), by_addr(vn[name]); dv = [k for k in set(a) | set(b) if a.get(k) != b.get(k)]
        fa, fb = by_addr(Fo[name]), by_addr(Fn[name]); df = [k for k in set(fa) | set(fb) if fa.get(k) != fb.get(k)]
        n += len(a); tv += len(dv); tf += len(df)
        if dv:
            print(f"   (a) VALUE DIFF {name}: n={len(dv)} {[(k, a.get(k), b.get(k)) for k in sorted(dv)[:3]]}")
        if df:
            fdiffs[name] = (len(df), [(L(k[1]) + str(k[0]), fa.get(k), fb.get(k)) for k in sorted(df)[:2]])
    print(f"   (a) other sheets: cells={n} value diffs={tv}; formula/static diffs by sheet={ {k: v[0] for k, v in fdiffs.items()} }")
    for name, (cnt, ex) in fdiffs.items():
        for addr, x, y in ex:
            print(f"       {name}!{addr}: {str(x)[:70]!r} -> {str(y)[:70]!r}")

    # Checks layout (row 5 header): A Category, B Line Item, C Sheet, D Range/Cell, E Actual, F Expected, G Difference, H Tolerance, I Status, J Notes
    def bylabel(ws):
        out = collections.OrderedDict()
        for r in range(1, ws.max_row + 1):
            k = (ws.cell(r, 2).value, ws.cell(r, 3).value)
            if isinstance(k[0], str) and k[0].strip():
                out.setdefault(k, []).append([ws.cell(r, cc).value for cc in range(1, ws.max_column + 1) if cc != 4])
        return out
    la, lb = bylabel(vo["Checks"]), bylabel(vn["Checks"]); chk_diff = []
    for k, ra in la.items():
        rb = lb.get(k)
        if rb is None or len(rb) != len(ra):
            chk_diff.append((k, "missing/count")); continue
        for x, y in zip(ra, rb):
            if x != y:
                chk_diff.append((k, x, y))
    print(f"   (a) Checks by (category,line item), all cols but the range string: rows={len(la)} new-only={len(set(lb)-set(la))} diffs={chk_diff[:3]} (n={len(chk_diff)})")
    hdr = [vn["Checks"].cell(6, cc).value for cc in range(1, 12)]
    stat_col = next((i + 1 for i, h in enumerate(hdr) if str(h).strip().lower() == "status"), 9)
    notok_o = [(r, vo["Checks"].cell(r, 2).value) for r in range(7, vo["Checks"].max_row + 1) if vo["Checks"].cell(r, stat_col).value not in (None, "OK", "")]
    notok_n = [(r, vn["Checks"].cell(r, 2).value) for r in range(7, vn["Checks"].max_row + 1) if vn["Checks"].cell(r, stat_col).value not in (None, "OK", "")]
    print(f"   (a) Checks status col={L(stat_col)}; non-OK rows old={[str(x[1])[:40] for x in notok_o]} new={[str(x[1])[:40] for x in notok_n]} same={[x[1] for x in notok_o]==[x[1] for x in notok_n]}")
    fin = vn["FINMO"]; flab = {str(fin.cell(r, 1).value).strip(): r for r in range(1, fin.max_row + 1) if isinstance(fin.cell(r, 1).value, str)}
    prow = [k for k in flab if "payroll" in k.lower()]
    finmo_same = all([vo["FINMO"].cell(flab[k], cc).value for cc in range(1, 30)] == [fin.cell(flab[k], cc).value for cc in range(1, 30)] for k in prow)
    print(f"   (a) FINMO rows {prow} old==new: {finmo_same}")
    # ---------- Payroll Schedule: summary by address, old detail vs bridge ----------
    po, pn, Po, Pn = vo["Payroll Schedule"], vn["Payroll Schedule"], Fo["Payroll Schedule"], Fn["Payroll Schedule"]
    sec_o = [r for r in range(1, Po.max_row + 1) if str(Po.cell(r, 1).value).startswith("Payroll Detail")][0]
    sec_n = [r for r in range(1, Pn.max_row + 1) if str(Pn.cell(r, 1).value).startswith("Payroll Detail")][0]
    sv = sum(1 for r in range(1, sec_o) for cc in range(1, 30) if po.cell(r, cc).value != pn.cell(r, cc).value)
    sf = [(r, cc) for r in range(1, sec_o) for cc in range(1, 30) if Po.cell(r, cc).value != Pn.cell(r, cc).value]
    sumifs = [(r, cc, Pn.cell(r, cc).value) for r in range(1, sec_n) for cc in range(1, 30) if isinstance(Pn.cell(r, cc).value, str) and "SUMIFS" in Pn.cell(r, cc).value.upper()]
    print(f"   (a) summary rows 1..{sec_o-1}: section header same row={sec_o==sec_n}; value diffs={sv}; formula diffs={len(sf)}; SUMIFS cells={len(sumifs)}")
    oh = [r for r in range(1, Po.max_row + 1) if Po.cell(r, 1).value == "Quarter"][-1] + 1
    nh = [r for r in range(1, Pn.max_row + 1) if Pn.cell(r, 1).value == "Quarter"][-1] + 1
    old_n = 0
    while isnum(Po.cell(oh + old_n, 1).value):
        old_n += 1
    new_n = 0
    while isnum(Pn.cell(nh + new_n, 1).value):
        new_n += 1
    vd = []; tx = collections.Counter()
    for k in range(max(old_n, new_n)):
        for cc in range(1, 15):
            x, y = po.cell(oh + k, cc).value, pn.cell(nh + k, cc).value
            if cc in (2, 3, 4, 14):
                if x is None:
                    tx[f"old None -> new {y!r}"] += 1
                if x != y and not (x is None and y == ""):
                    vd.append((k, cc, x, y))
            elif x != y:
                vd.append((k, cc, x, y))
    print(f"   (a) old detail rows={old_n} (from {oh}) vs bridge rows={new_n} (from {nh}) == engine rows {len(rows)}: {old_n==new_n==len(rows)}; value-for-value diffs (exact, 14 cols)={vd[:3]} (n={len(vd)}); text cells where old was empty: {dict(tx)}")
    # ---------- (b) blocks ----------
    blocks = blocks_of(Pn); by_id = collections.defaultdict(dict)
    for (q, k), it in zip(ids, rows):
        by_id[k][q] = it
    broken = [b for b in blocks if "BROKEN" in b]
    print(f"   (b) blocks={len(blocks)} == identities={len(order)}: {len(blocks)==len(order)}; broken blocks={broken[:2]}; blocks within rows {blocks[0]['header'] if blocks else None}..{blocks[-1]['Total Payroll'] if blocks else None} < bridge {nh}: {bool(blocks) and blocks[-1]['Total Payroll'] < nh}")
    shapes = collections.Counter(); bad = []; rec_miss = []; hdr_bad = []
    for bi, (k, b) in enumerate(zip(order, blocks)):
        cls, title, person, ordn = k
        if Pn.cell(b["header"], 1).value != (title or person):
            hdr_bad.append((bi, Pn.cell(b["header"], 1).value, title, person))
        byq = by_id[k]; qs = sorted(byq)
        for q in range(1, 21):
            col = PSC + q
            cells = {lab: Pn.cell(b[lab], col).value for lab in PERIOD}
            if q not in byq:
                if any(v is not None for v in cells.values()):
                    bad.append((bi, q, "cell present with no engine row", cells))
                continue
            it = byq[q]; prior = byq.get(q - 1)
            s, h, w, bn = (float(it.get(f) or 0) for f in ("starting_fte", "hires", "annual_wage", "payroll_taxes_benefits_percent"))
            if isinstance(cells["Hires"], str) or float(cells["Hires"] or 0) != h:
                bad.append((bi, q, "Hires", cells["Hires"], h))
            else:
                shapes["Hires literal"] += 1
            if prior is None:
                for lab, ev in (("Starting FTE", s), ("Annual Wage", w), ("Benefits %", bn)):
                    v = cells[lab]
                    if isinstance(v, str) or float(v or 0) != ev:
                        bad.append((bi, q, lab + " literal!=engine", v, ev))
                    else:
                        shapes[f"{lab} literal " + ("first" if q == qs[0] else "RESEED")] += 1
            else:
                m = RX_ROUND.match(str(cells["Starting FTE"]))
                if m and m.group(1) == L(col - 1) and int(m.group(2)) == b["Ending FTE"] and m.group(3) is None:
                    shapes["Starting =ROUND(prev Ending,6)"] += 1
                else:
                    bad.append((bi, q, "Starting FTE", cells["Starting FTE"]))
                for lab, ev, key in (("Annual Wage", w, "annual_wage"), ("Benefits %", bn, "payroll_taxes_benefits_percent")):
                    v = str(cells[lab]); pv = float(prior.get(key) or 0); delta = ev - pv
                    mf, ms = RX_REF.match(v), RX_ROUND.match(v)
                    if abs(delta) <= 1e-9:
                        if mf and mf.group(1) == L(col - 1) and int(mf.group(2)) == b[lab]:
                            shapes[f"{lab} =prev"] += 1
                        else:
                            bad.append((bi, q, lab, v, "expected flat"))
                    else:
                        if ms and ms.group(1) == L(col - 1) and int(ms.group(2)) == b[lab] and ms.group(3) and (ms.group(3) == "-") == (delta < 0) and float(ms.group(4)) == abs(delta):
                            shapes[f"{lab} =ROUND(prev+/-delta,6)"] += 1
                        else:
                            bad.append((bi, q, lab, v, "expected step", delta))
            C = L(col)
            exp = {"Ending FTE": f"={C}{b['Starting FTE']}+{C}{b['Hires']}", "Average FTE": f"=({C}{b['Starting FTE']}+{C}{b['Ending FTE']})/2",
                   "Wage Cost": f"={C}{b['Average FTE']}*{C}{b['Annual Wage']}/4", "Taxes & Benefits": f"={C}{b['Wage Cost']}*{C}{b['Benefits %']}", "Total Payroll": f"={C}{b['Wage Cost']}+{C}{b['Taxes & Benefits']}"}
            for lab, e in exp.items():
                if cells[lab] == e:
                    shapes["derived formula exact"] += 1
                else:
                    bad.append((bi, q, lab, cells[lab], e))
            for lab, ev in (("Starting FTE", s), ("Hires", h), ("Annual Wage", w), ("Benefits %", bn)):
                gv = pn.cell(b[lab], col).value
                if float(gv or 0) != ev:
                    rec_miss.append((bi, q, lab, ev, gv))
    print(f"   (b) header title mismatches={hdr_bad[:2]} (n={len(hdr_bad)}); shapes={dict(shapes)}; bad={bad[:3]} (n={len(bad)}); recalculated block inputs == engine float-exact: misses={rec_miss[:3]} (n={len(rec_miss)})")
    # ---------- (c) bridge ----------
    hidden_rows = [r for r in range(nh - 2, nh + new_n) if Pn.row_dimensions[r].hidden]
    vis_hidden = [r for r in range(1, nh - 2) if Pn.row_dimensions[r].hidden]
    print(f"   (c) hidden: note+header+{new_n} bridge rows all hidden={len(hidden_rows)==new_n+2} (hidden={len(hidden_rows)}); hidden rows above the bridge note={vis_hidden[:3]}")
    blk_of = dict(zip(order, blocks)); cbad = []; nonformula = 0; own_ok = 0
    for i, (q, k) in enumerate(ids):
        r = nh + i; b = blk_of[k]; col = PSC + q
        if Pn.cell(r, 1).value != q:
            cbad.append((r, "A", Pn.cell(r, 1).value, q))
        for cc, (brow, bcol) in {2: (b["header"], 2), 3: (b["header"], 1), 4: (b["oews"], 2), 14: (b["source"], 2)}.items():
            v = str(Pn.cell(r, cc).value); m = RX_TEXT.match(v)
            if not (m and m.group(1) == L(bcol) and int(m.group(2)) == brow):
                cbad.append((r, cc, v))
        for cc, lab in BRIDGE_COLS.items():
            v = Pn.cell(r, cc).value; m = RX_REF.match(str(v))
            if not (isinstance(v, str) and v.startswith("=")):
                nonformula += 1
            if not (m and m.group(1) == L(col) and int(m.group(2)) == b[lab]):
                cbad.append((r, cc, v, "expected", f"={L(col)}{b[lab]}"))
        if not any(x[0] == r for x in cbad):
            own_ok += 1
    print(f"   (c) bridge rows pointing at their OWN identity's block, right quarter column, all 13 formula cells: {own_ok}/{len(ids)}; non-formula cells={nonformula}; bad={cbad[:3]} (n={len(cbad)})")
    dupq = collections.defaultdict(list)
    for i, (q, k) in enumerate(ids):
        dupq[(q, k[0], k[1])].append(i)
    dd = {kk: v for kk, v in dupq.items() if len(v) > 1}
    if dd:
        (q, cls, title), idx = next(iter(dd.items()))
        refs = [int(RX_REF.match(str(Pn.cell(nh + i, 9).value)).group(2)) for i in idx]
        print(f"   (c) dup-title groups={len(dd)}; e.g. q{q} {title!r}: engine persons={[rows[i].get('person_name') for i in idx]} wages={[rows[i].get('annual_wage') for i in idx]} -> bridge I refs rows {refs} -> block headers {[Pn.cell(r_ - 7, 1).value for r_ in refs]} recalculated I={[pn.cell(nh + i, 9).value for i in idx]}")
    # ---------- (d) SUMIFS + Checks keying ----------
    keyA = [f for _, _, f in sumifs if re.search(rf"\$A\${nh}:\$A\${nh+new_n-1}", f)]
    keyed_elsewhere = [f for _, _, f in sumifs if not re.search(rf"\$A\${nh}:\$A\${nh+new_n-1}", f)]
    chkF = Fn["Checks"]; pay_rows = []
    for r in range(7, chkF.max_row + 1):
        li = str(chkF.cell(r, 2).value or "")
        if "payroll" in li.lower() and str(chkF.cell(r, 3).value) == "Payroll Schedule" and not li.startswith("Payroll Detail - "):
            pay_rows.append((li[:45], chkF.cell(r, 4).value, vn["Checks"].cell(r, stat_col).value, str(chkF.cell(r, 5).value)))
    detail_rows = [x for x in pay_rows if "detail" in x[0].lower() or "summary totals" in x[0].lower()]
    rng_ok = all(re.search(rf"[A-N]{nh}:[A-N]{nh+new_n-1}", str(x[1])) for x in detail_rows)
    sumifs_keyA = all(re.search(rf"\$A\${nh}:\$A\${nh+new_n-1}", x[3]) for x in pay_rows if "SUMIFS" in x[3])
    fc = []
    for r in range(7, chkF.max_row + 1):
        li = str(chkF.cell(r, 2).value or "")
        if li.startswith("Payroll Detail - "):
            m = re.search(r"!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)", str(chkF.cell(r, 4).value)); colL, r1, r2 = m.group(1), int(m.group(2)), int(m.group(4))
            mine = sum(1 for rr_ in range(r1, r2 + 1) if isinstance(Pn[f"{colL}{rr_}"].value, str) and Pn[f"{colL}{rr_}"].value.startswith("="))
            fc.append((li[18:], (r1, r2) == (nh, nh + new_n - 1), chkF.cell(r, 6).value, mine, vn["Checks"].cell(r, stat_col).value))
    print(f"   (d) summary SUMIFS keyed on bridge $A${nh}:$A${nh+new_n-1}: {len(keyA)}/{len(sumifs)} (others: {keyed_elsewhere[:1]}); Checks payroll rows={[(x[0][:32], x[1], x[2]) for x in pay_rows]}; detail ranges cover the bridge={rng_ok}; Checks SUMIFS keyed on bridge col A={sumifs_keyA}")
    print(f"   (d) formula-count checks (label, range==bridge, expected, MY count, status): {fc}")
    # ---------- (e) amber ----------
    lost = {}; gained = {}
    for name in vo.sheetnames:
        if name == "Payroll Schedule":
            continue
        ao, an = amber(Fo[name]), amber(Fn[name])
        if ao - an:
            lost[name] = sorted(ao - an)[:3]
        if an - ao:
            gained[name] = sorted(an - ao)[:3]
    ap_o, ap_n = amber(Po), amber(Pn)
    inp_rows = {b[lab] for b in blocks for lab in ("Starting FTE", "Hires", "Annual Wage", "Benefits %")}
    txt_rows = {b[lab] for b in blocks for lab in ("header", "source")}
    num_n = sum(1 for (r, cc) in ap_n if cc >= 4 and r in inp_rows)
    txt_n = sum(1 for (r, cc) in ap_n if cc <= 2 and r in txt_rows)
    other_n = len(ap_n) - num_n - txt_n
    num_o = sum(1 for (r, cc) in ap_o if oh <= r < oh + old_n and cc in (5, 6, 9, 10)); txt_o = sum(1 for (r, cc) in ap_o if oh <= r < oh + old_n and cc in (2, 3, 14))
    bridge_amber = sum(1 for (r, cc) in ap_n if nh - 2 <= r < nh + new_n)
    above_o = {k for k in ap_o if k[0] < sec_o}; above_n = {k for k in ap_n if k[0] < sec_n}
    print(f"   (e) other sheets amber lost={lost or 'none'} gained={gained or 'none'}; Payroll numeric amber old={num_o} new={num_n} == 4 x rows {4*len(rows)}: {num_n==4*len(rows)}; text amber old={txt_o} new={txt_n} == 3 x roles {3*len(order)}: {txt_n==3*len(order)}; bridge amber={bridge_amber}; unclassified new amber={other_n}; amber above the detail section old==new: {above_o==above_n}")
    GRAND.update({"drafts": 1, "other-sheet value diffs": tv, "checks label diffs": len(chk_diff), "summary value diffs": sv, "detail-vs-bridge diffs": len(vd), "block bad": len(bad), "block recalc misses": len(rec_miss), "bridge bad": len(cbad), "amber lost": sum(len(v) for v in lost.values()), "rows": len(rows), "roles": len(order)})
print("\nGRAND", dict(GRAND))
