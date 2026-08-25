"""(c) demo to Q20 + (f) tamper. usage: <scratch> demo draft:label:col:amount ... | tamper draft ..."""
import sys, os, glob, shutil, time, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; MODE = sys.argv[2]
import win32com.client as win32
X = win32.gencache.EnsureDispatch("Excel.Application"); X.Visible = False; X.DisplayAlerts = False


def recalc(path):
    w = X.Workbooks.Open(path)
    for _ in range(20):
        try:
            w.Sheets(1).Name; break
        except Exception:
            time.sleep(1.5)
    X.CalculateFullRebuild(); w.Save(); w.Close(False)


def label_map(ws):
    return {ws.cell(r, 1).value.strip(): r for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip()}


os.makedirs(f"{S}/demo", exist_ok=True)
if MODE == "demo":
    for spec in sys.argv[3:]:
        d, lab, col, amt = spec.split(":"); col = int(col); amt = float(amt)
        for tag in ("old", "new"):
            src = glob.glob(f"{S}/{tag}/{tag.upper()} {d}*.xlsx")[0]
            dst = f"{S}/demo/{tag}_{d}_{lab[:3]}_{get_column_letter(col)}.xlsx"; shutil.copy(src, dst)
            wb = openpyxl.load_workbook(dst); ce = wb["Cash Equity Schedule"]; r = label_map(ce)[lab]
            before = ce.cell(r, col).value
            v0 = openpyxl.load_workbook(src, data_only=True); base = float(v0["Cash Equity Schedule"].cell(r, col).value or 0)
            ce.cell(r, col).value = base + amt; wb.save(dst); recalc(dst)
            v = openpyxl.load_workbook(dst, data_only=True); fin = v["FINMO"]; fr = label_map(fin)
            cols = range(max(3, col - 2), 24)
            g = lambda ws, row: [float(ws.cell(row, cc).value or 0) for cc in cols]
            row_after = g(fin, fr[lab]); row_before = g(v0["FINMO"], fr[lab])
            tle = g(fin, fr["Total Liabilities & Equity"]); tle0 = g(v0["FINMO"], fr["Total Liabilities & Equity"]); ta = g(fin, fr["Total Assets"])
            print(f"{d} {tag.upper()} {lab} {get_column_letter(col)}{r} (Q{col-3}): cell was {before!r} value {base} -> typed {base+amt}")
            print(f"   quarters      {[f'Q{cc-3}' for cc in cols]}")
            print(f"   {lab[:14]:14} {[round(x) for x in row_before]} -> {[round(x) for x in row_after]}")
            print(f"   row delta     {[round(a-b) for a, b in zip(row_after, row_before)]}")
            print(f"   Total L&E d   {[round(a-b) for a, b in zip(tle, tle0)]}  A=L+E diffs max {max(abs(a-b) for a, b in zip(ta, tle)):.6f}  Checks!B2 {v['Checks']['B2'].value}")
            hold = all(abs((a - b) - amt) < 1e-6 for a, b in zip(row_after, row_before) if True) and all(abs((a - b) - amt) < 1e-6 for a, b in zip(tle, tle0))
            print(f"   HOLDS typed amount on every quarter from Q{col-3} to Q20 (row AND Total L&E): {hold}")
elif MODE == "tamper":
    for d in sys.argv[3:]:
        src = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]; dst = f"{S}/demo/tamper_{d}.xlsx"; shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst); mi = wb["Model Inputs"]; r = label_map(mi)["Lease"]
        v0 = openpyxl.load_workbook(src, data_only=True)
        ck0 = v0["Checks"]; tie0 = [rr for rr in range(1, ck0.max_row + 1) if "Lease/rent reaches" in str(ck0.cell(rr, 2).value or "")]
        was = mi.cell(r, 11).value; mi.cell(r, 11).value = float(was or 0) + 1234.56; wb.save(dst); recalc(dst)
        v = openpyxl.load_workbook(dst, data_only=True); ck = v["Checks"]
        print(f"{d} TAMPER Model Inputs!K{r} (Lease Q8) {was!r} -> {float(was or 0)+1234.56}")
        for rr in tie0:
            print(f"   tie-out r{rr} before: {[ck0.cell(rr, cc).value for cc in range(2, 10)]}")
            print(f"   tie-out r{rr} after : {[ck.cell(rr, cc).value for cc in range(2, 10)]}")
        print(f"   Checks!B2 before/after: {ck0['B2'].value} / {ck['B2'].value} | FINMO Lease/Rent Q8 before/after: {v0['FINMO'].cell(label_map(v0['FINMO'])['Lease/Rent'], 11).value} / {v['FINMO'].cell(label_map(v['FINMO'])['Lease/Rent'], 11).value}")
X.Quit()
