"""mini's option-(B) audit over OLD (7deaefa) and NEW (6cd654e+) recalculated builds. usage: <scratch> <draft_prefix>..."""
import sys, os, glob, json, re, openpyxl
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
S = sys.argv[1]; drafts = sys.argv[2:]
sys.path.insert(0, r"C:\dev\business_plann_app"); sys.path.insert(0, r"C:\dev\business_plann_app\python")
from dotenv import load_dotenv; load_dotenv(r"C:\dev\business_plann_app\.env")
import mysql.connector
from client_statements_output_excel.data import values_21
c = mysql.connector.connect(host=os.getenv('MYSQL_HOST'), user=os.getenv('MYSQL_USER'), password=os.getenv('MYSQL_PASSWORD'), database=os.getenv('MYSQL_DB'), autocommit=True)


def by_addr(ws):
    return {(cell.row, cell.column): cell.value for r in ws.iter_rows() for cell in r if cell.value is not None}


def is_formula(v):
    return isinstance(v, str) and v.startswith("=")


def label_map(ws):
    return {ws.cell(r, 1).value.strip(): r for r in range(1, ws.max_row + 1) if isinstance(ws.cell(r, 1).value, str) and ws.cell(r, 1).value.strip()}


def checks_rows(ws):
    out = {}
    for r in range(7, ws.max_row + 1):
        key = (ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value)
        out.setdefault(key, []).append(tuple(ws.cell(r, cc).value for cc in range(5, 10)))
    return out


RX = re.compile(r"^=ROUND\(([A-Z]+)(\d+)([+-])([0-9.eE+-]+),6\)$")
for d in drafts:
    fo = glob.glob(f"{S}/old/OLD {d}*.xlsx")[0]; fn = glob.glob(f"{S}/new/NEW {d}*.xlsx")[0]
    vo, vn = openpyxl.load_workbook(fo, data_only=True), openpyxl.load_workbook(fn, data_only=True)
    Fo, Fn = openpyxl.load_workbook(fo), openpyxl.load_workbook(fn)
    print(f"\n######## {d}  old={os.path.basename(fo)}  new={os.path.basename(fn)}")
    print("sheets old==new:", vo.sheetnames == vn.sheetnames, "| Checks!B2 old/new:", vo["Checks"]["B2"].value, "/", vn["Checks"]["B2"].value)
    tv = tf = ncell = 0
    for name in vo.sheetnames:
        if name in ("Cash Equity Schedule", "Checks"):
            continue
        a, b = by_addr(vo[name]), by_addr(vn[name]); dv = [k for k in set(a) | set(b) if a.get(k) != b.get(k)]
        fa, fb = by_addr(Fo[name]), by_addr(Fn[name]); df = [k for k in set(fa) | set(fb) if fa.get(k) != fb.get(k)]
        ncell += len(a); tv += len(dv); tf += len(df)
        if dv or df:
            print(f"   {name}: value diffs={len(dv)} {[(k, a.get(k), b.get(k)) for k in sorted(dv)[:3]]} formula diffs={len(df)} {[(k, fa.get(k), fb.get(k)) for k in sorted(df)[:3]]}")
    print(f"   (a) other sheets: cells={ncell} value diffs={tv} formula diffs={tf}")
    lo, ln = label_map(vo["Cash Equity Schedule"]), label_map(vn["Cash Equity Schedule"])
    for lab in ("Owner's Capital", "Other Equity", "Distributions"):
        o = [vo["Cash Equity Schedule"].cell(lo[lab], 3 + i).value for i in range(21)]; n = [vn["Cash Equity Schedule"].cell(ln[lab], 3 + i).value for i in range(21)]
        print(f"   (a) Cash sheet {lab} recalculated old==new exact: {o == n}" + ("" if o == n else f" DIFF {[(i, o[i], n[i]) for i in range(21) if o[i] != n[i]][:3]}"))
    co, cn = checks_rows(vo["Checks"]), checks_rows(vn["Checks"])
    gone = [k for k in co if k not in cn]; added = [k for k in cn if k not in co]; moved = [k for k in co if k in cn and co[k] != cn[k]]
    print(f"   (a) Checks by key: rows old={sum(len(v) for v in co.values())} new={sum(len(v) for v in cn.values())} gone={gone} added={added} value-changed={len(moved)} {[(k, co[k], cn[k]) for k in moved[:3]]}")
    cur = c.cursor(dictionary=True); cur.execute("SELECT model_input_json, finmo_json FROM intake_consult_drafts WHERE draft_id LIKE %s", (d + '%',)); row = cur.fetchone(); cur.close()
    mi = json.loads(row["model_input_json"]); bs = {x["label"]: values_21(x.get("values")) for x in mi["sections"]["balance_sheet"] if isinstance(x, dict)}
    ceF = Fn["Cash Equity Schedule"]; ceO = Fo["Cash Equity Schedule"]
    for lab in ("Owner's Capital", "Other Equity"):
        r = ln[lab]; eng = [float(v or 0) for v in bs[lab]]; cells = [ceF.cell(r, 3 + i).value for i in range(21)]
        stub_lit = not is_formula(cells[0]) and float(cells[0]) == eng[0]
        flat_ok = steps_ok = True; steps = []; probs = []
        for i in range(1, 21):
            prev = f"{get_column_letter(2 + i)}{r}"; delta = eng[i] - eng[i - 1]
            if abs(delta) <= 1e-9:
                if cells[i] != f"={prev}":
                    flat_ok = False; probs.append((i, cells[i]))
            else:
                m = RX.match(str(cells[i]))
                if not m or f"{m.group(1)}{m.group(2)}" != prev or (m.group(3) == "-") != (delta < 0) or float(m.group(4)) != abs(delta):
                    steps_ok = False; probs.append((i, cells[i], delta))
                else:
                    steps.append((i, m.group(3) + m.group(4)))
        recalc = [vn["Cash Equity Schedule"].cell(r, 3 + i).value for i in range(21)]
        mm = [(i, recalc[i], eng[i]) for i in range(21) if float(recalc[i] or 0) != eng[i]]
        old_lits = [i for i in range(21) if not is_formula(ceO.cell(r, 3 + i).value)]
        print(f"   (b) {lab} r{r}: stub literal==engine {stub_lit}; flat cells =prev {flat_ok}; step cells ROUND(prev+/-engine delta,6) {steps_ok} steps={steps}; probs={probs[:3]}; recalculated==engine exact mismatches={mm[:3]}; OLD tree literal positions={old_lits}")
    fin, finF, fino = vn["FINMO"], Fn["FINMO"], vo["FINMO"]; frow = label_map(fin)
    for lab in ("Owner's Capital", "Other Equity", "Distributions"):
        r = frow[lab]; newv = [fin.cell(r, 3 + i).value for i in range(21)]; oldv = [fino.cell(r, 3 + i).value for i in range(21)]
        mm = [(i, newv[i], bs[lab][i]) for i in range(21) if float(newv[i] or 0) != float(bs[lab][i] or 0)]
        print(f"   (a) FINMO r{r} {lab}: new==old exact={newv == oldv}; vs engine series exact mismatches={mm[:3]}")
    r = ln["Distributions"]; dl = all(not is_formula(ceF.cell(r, 3 + i).value) for i in range(21)); dlo = all(not is_formula(ceO.cell(lo["Distributions"], 3 + i).value) for i in range(21))
    dvals_same = [ceF.cell(r, 3 + i).value for i in range(21)] == [ceO.cell(lo["Distributions"], 3 + i).value for i in range(21)]
    miF = Fn["Model Inputs"]; mrow = label_map(miF)
    print(f"   (e) Distributions: all 21 literal new={dl} old={dlo} literals identical={dvals_same}; FINMO r{frow['Distributions']} K: {finF.cell(frow['Distributions'], 11).value} | Model Inputs K: {miF.cell(mrow['Distributions'], 11).value} | FINMO Equity r{frow['Equity']} K: {finF.cell(frow['Equity'], 11).value}")
    print("   (d) A2:", repr(ceF['A2'].value)); print("   (d) B7/B8/B9:", repr(ceF['B7'].value), "|", repr(ceF['B8'].value), "|", repr(ceF['B9'].value))
    ck = Fn["Checks"]; ckv = vn["Checks"]
    for rr in range(1, ck.max_row + 1):
        if any("Lease/rent" in str(ck.cell(rr, cc).value or "") for cc in range(1, 5)):
            print(f"   (f) Checks r{rr}:", [ck.cell(rr, cc).value for cc in range(1, 10)], "| recalculated:", [ckv.cell(rr, cc).value for cc in range(5, 10)])
    src = [s for s in Fn.sheetnames if "Source" in s or "Audit" in s]
    sf = Fn[src[0]]; srow = label_map(sf)
    print(f"   (f) source sheet={src} Lease/Rent row={srow.get('Lease/Rent')} | FINMO Lease/Rent r{frow.get('Lease/Rent')} K formula: {finF.cell(frow['Lease/Rent'], 11).value} | Model Inputs Lease r{mrow.get('Lease')} K: {miF.cell(mrow['Lease'], 11).value!r}")
    fl = [fin.cell(frow['Lease/Rent'], 3 + i).value for i in range(21)]; sl = [vn[src[0]].cell(srow['Lease/Rent'], 3 + i).value for i in range(21)]
    print(f"   (f) FINMO Lease/Rent == Audit Source Lease/Rent (recalculated, 21 cols): {[(a or 0) == (b or 0) for a, b in zip(fl, sl)].count(True)}/21; Q8 {fl[8]} vs {sl[8]}")
