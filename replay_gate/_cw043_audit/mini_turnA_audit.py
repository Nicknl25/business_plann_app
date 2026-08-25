"""mini's CW-043 TURN A audit instruments (2026-08-25). Reproduces the evidence in
turnA_audit_evidence.txt. Run from the repo root with .venv python; needs Excel.

  python replay_gate/_cw043_audit/mini_turnA_audit.py detect  <a.xlsx> [<b.xlsx>]
      residue detector (any numeric 0<|v|<1e-4 on Debt Schedule + CapEx Depreciation,
      any |v|>1e9 anywhere, any #-error), payoff exactness (closing == 0 as int, nothing
      after), Checks!B2 - and with two files an EXACT by-address value compare of every sheet.
  python replay_gate/_cw043_audit/mini_turnA_audit.py tamper  <head_built_harrow.xlsx>
      the four-family tamper: unguard/unround the built FORMULAS, recalc in Excel, show
      the crumb come back. Debt uses Harrow's natural Q7 crumb; lease/ROU/PPE inject a
      one-crumb-short principal/depreciation because no stored draft crumbs there.
  python replay_gate/_cw043_audit/mini_turnA_audit.py sweep
      engine-side sweep of every draft's finmo_json quarter_rows for sub-cent values,
      histogram by key, and the closing-key subset.
Drafts are exported with mini_draft_export.py (same dir).
"""
import collections
import json
import os
import re
import shutil
import sys
import time


def isnum(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def recalc(path):
    import win32com.client as win32
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Visible = False
    x.DisplayAlerts = False
    w = x.Workbooks.Open(path)
    for _ in range(20):
        try:
            w.Sheets(1).Name
            break
        except Exception:
            time.sleep(1.5)
    x.CalculateFullRebuild()
    w.Save()
    w.Close(False)
    x.Quit()


def detect(path):
    import openpyxl
    v = openpyxl.load_workbook(path, data_only=True)
    crumbs = [(ws.title, c.coordinate, c.value) for ws in (v["Debt Schedule"], v["CapEx Depreciation"])
              for row in ws.iter_rows() for c in row
              if isinstance(c.value, float) and c.value != 0 and abs(c.value) < 1e-4]
    big = [(ws.title, c.coordinate, ws.cell(c.row, 1).value, c.value) for ws in v.worksheets
           for row in ws.iter_rows() for c in row if isnum(c.value) and abs(c.value) > 1e9]
    errs = [(ws.title, c.coordinate, c.value) for ws in v.worksheets for row in ws.iter_rows()
            for c in row if isinstance(c.value, str) and c.value.startswith("#")]
    ds = v["Debt Schedule"]

    def payoff(o, p, i, cl):
        rows = [(ds.cell(r, 1).value, ds.cell(r, o).value, ds.cell(r, p).value,
                 ds.cell(r, i).value, ds.cell(r, cl).value) for r in range(7, 28)]
        for k, (lab, op, _pr, _it, close) in enumerate(rows):
            if (op or 0) > 0 and close == 0:
                bad = [x for x in rows[k + 1:]
                       if not ((x[1] or 0) == 0 and (x[3] or 0) == 0 and (x[4] or 0) == 0)]
                return (lab, close, type(close).__name__, "nothing-after" if not bad else bad[:3])
        return None

    return v, dict(crumbs=crumbs, big=big[:6], errs=errs[:6], n_errs=len(errs),
                   debt_payoff=payoff(2, 8, 9, 11), lease_payoff=payoff(12, 18, 19, 21),
                   checks_b2=v["Checks"]["B2"].value)


def compare(va, vb):
    diffs = []
    for name in va.sheetnames:
        for row in va[name].iter_rows():
            for c in row:
                x, y = c.value, vb[name][c.coordinate].value
                if x != y:
                    diffs.append((name, c.coordinate, va[name].cell(c.row, 1).value, x, y))
    return diffs


def unguard(f):
    m = re.match(r"=IF\((.+)<0\.005,0,(.+)\)$", f)
    assert m and m.group(1) == m.group(2), f
    return "=MAX(0,%s)" % m.group(1)


def unround(f):
    return re.sub(r"ROUND\((.+),6\)", r"\1", f)


def old_ratio(f):
    return f.replace("<0.005,", "<=0,")


def tamper(src, out_dir):
    import openpyxl
    os.makedirs(out_dir, exist_ok=True)

    def run(tag, edits, probe):
        dst = os.path.join(out_dir, tag + ".xlsx")
        shutil.copy(src, dst)
        wb = openpyxl.load_workbook(dst)
        for sheet, addr, fn in edits:
            c = wb[sheet][addr]
            c.value = fn(c.value) if callable(fn) else fn
        wb.save(dst)
        recalc(dst)
        v = openpyxl.load_workbook(dst, data_only=True)
        out = {k: v[s][a].value for k, (s, a) in probe.items()}
        out["crumbs"] = [(ws.title, c.coordinate, c.value)
                         for ws in (v["Debt Schedule"], v["CapEx Depreciation"])
                         for row in ws.iter_rows() for c in row
                         if isinstance(c.value, float) and c.value != 0 and abs(c.value) < 1e-4][:5]
        print(tag, out)

    f = openpyxl.load_workbook(src)
    fin, ds, cx = f["FINMO"], f["Debt Schedule"], f["CapEx Depreciation"]
    ic = [r for r in range(1, fin.max_row + 1) if fin.cell(r, 1).value == "Interest Coverage"][0]
    dscr = [r for r in range(1, fin.max_row + 1)
            if fin.cell(r, 1).value == "Debt Service Coverage Ratio (DSCR)"][0]
    rou = [r for r in range(28, ds.max_row + 1) if ds.cell(r, 1).value == "Right-of-Use Asset Closing"][0]
    ppe = [r for r in range(1, cx.max_row + 1) if cx.cell(r, 1).value == "Closing PPE"][0]
    rows = range(8, 28)
    probe = {"K14": ("Debt Schedule", "K14"), "K15": ("Debt Schedule", "K15"),
             "I15": ("Debt Schedule", "I15"), "IC Q8": ("FINMO", "L%d" % ic),
             "DSCR Q8": ("FINMO", "L%d" % dscr), "U27": ("Debt Schedule", "U27"),
             "ROU W": ("Debt Schedule", "W%d" % rou), "PPE W": ("CapEx Depreciation", "W%d" % ppe)}
    run("intact", [], probe)
    run("debt_unguard_only", [("Debt Schedule", "K%d" % r, unguard) for r in rows], probe)
    run("debt_unround_only", [("Debt Schedule", "H%d" % r, unround) for r in rows], probe)
    both = ([("Debt Schedule", "K%d" % r, unguard) for r in rows]
            + [("Debt Schedule", "H%d" % r, unround) for r in rows])
    run("debt_both_0b26ce8_form", both, probe)
    run("debt_both_plus_old_ratio_guard",
        both + [("FINMO", "L%d" % ic, old_ratio), ("FINMO", "L%d" % dscr, old_ratio)], probe)
    run("lease_short_intact", [("Debt Schedule", "R27", "=L27+M27-0.000000000005")], probe)
    run("lease_short_unguard", [("Debt Schedule", "R27", "=L27+M27-0.000000000005"),
                                ("Debt Schedule", "U27", unguard)], probe)
    run("lease_round_absorbs", [("Debt Schedule", "R27", "=MIN(L27+M27,MAX(0,ROUND(P27+Q27-0.000000000005,6)))"),
                                ("Debt Schedule", "U27", unguard)], probe)
    dep = "W%d" % (rou - 1)
    inj = "=W%d+W%d-0.000000000005" % (rou - 2, rou - 3)
    run("rou_inject_intact", [("Debt Schedule", dep, inj)], probe)
    run("rou_inject_unguard", [("Debt Schedule", dep, inj), ("Debt Schedule", "W%d" % rou, unguard)], probe)
    run("ppe_inject_intact", [("CapEx Depreciation", "W12", "=W10+W7-0.000000001")], probe)
    run("ppe_inject_unguard", [("CapEx Depreciation", "W12", "=W10+W7-0.000000001"),
                               ("CapEx Depreciation", "W%d" % ppe, unguard)], probe)


def sweep():
    from dotenv import load_dotenv
    load_dotenv(r"C:\dev\business_plann_app\.env")
    import mysql.connector
    c = mysql.connector.connect(host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
                                password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
                                autocommit=True)
    cur = c.cursor(dictionary=True)
    cur.execute("SELECT draft_id, finmo_json FROM intake_consult_drafts WHERE finmo_json IS NOT NULL")
    closing_keys = ("debt_closing_balance", "debt_opening_balance", "lease_closing_balance_total",
                    "lease_opening_balance_total", "ppe", "right_of_use_asset", "right_of_use_asset_opening",
                    "long_term_debt", "short_term_debt", "capital_lease_obligation", "accumulated_depreciation")
    hist = collections.Counter()
    closing = []
    n = nq = 0
    for r in cur:
        try:
            qr = json.loads(r["finmo_json"]).get("quarter_rows")
        except Exception:
            continue
        if not qr:
            continue
        n += 1
        for q in qr:
            nq += 1
            for k, v in q.items():
                if isnum(v) and v != 0 and abs(v) < 0.005:
                    hist[k] += 1
                    if k in closing_keys:
                        closing.append((r["draft_id"][:8], k, q.get("quarter_index"), v))
    print("drafts", n, "quarter rows", nq)
    print("sub-cent by key:", hist.most_common())
    print("closing-key residues:", closing)


if __name__ == "__main__":
    cmd = sys.argv[1]
    if cmd == "detect":
        va, ra = detect(sys.argv[2])
        print(sys.argv[2])
        print(json.dumps(ra, default=str, indent=1))
        if len(sys.argv) > 3:
            vb, rb = detect(sys.argv[3])
            print(sys.argv[3])
            print(json.dumps(rb, default=str, indent=1))
            d = compare(va, vb)
            print("exact by-address diffs:", len(d),
                  collections.Counter((s, str(l)[:30]) for s, _, l, _, _ in d).most_common(15))
    elif cmd == "tamper":
        tamper(sys.argv[2], os.path.join(os.path.dirname(sys.argv[2]), "tamper"))
    elif cmd == "sweep":
        sweep()
