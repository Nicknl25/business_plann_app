"""X1 numbers-byte-identical instrument (2026-08-18).

X1 is a DESIGN pass: colors, fonts, number formats, chart styling, layout, a
text-only cover sheet and sheet order. It must not change one value or one
formula. This instrument proves that.

  dump    <out.json> <business>   build the workbook through the PRODUCTION
                                  builder from the real draft row, recalculate
                                  it in Excel via COM, and record every cell's
                                  formula string AND its recalculated value.
  compare <base.json> <now.json>  every sheet present in BASE must carry the
                                  identical formula in every cell and the
                                  identical recalculated value; sheets that are
                                  NEW in NOW are reported and must contain zero
                                  formulas (the cover-sheet rule).

Number formats, fonts, fills, column widths, chart styling and sheet ORDER are
deliberately NOT compared - those are exactly what X1 is allowed to change.
"""
from __future__ import annotations

import json
import os
import sys
from typing import Optional

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for p in (REPO, os.path.join(REPO, "python")):
  if p not in sys.path:
    sys.path.insert(0, p)

# NOTHING is excluded by sheet. Excluding "Diagnostics" wholesale is what let a
# reported (and, as it turned out, imagined) content loss go unchecked on
# 2026-08-19 — a proof with a blind spot is not a proof. Only individual cells
# whose value is generation-time metadata are skipped, and they are named.
VOLATILE_CELLS = {("Diagnostics", "B2")}


def _draft_row(business: str) -> dict:
  from dotenv import load_dotenv  # type: ignore

  load_dotenv(os.path.join(REPO, ".env"))
  import mysql.connector  # type: ignore

  conn = mysql.connector.connect(
    host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
  )
  cur = conn.cursor(dictionary=True)
  cur.execute(
    "SELECT * FROM intake_consult_drafts WHERE business_name LIKE %s "
    "AND finmo_json IS NOT NULL ORDER BY planning_run_completed_at DESC LIMIT 1",
    (f"%{business}%",),
  )
  row = cur.fetchone()
  cur.close()
  conn.close()
  if not row:
    raise SystemExit(f"no completed draft matching {business!r}")
  return row


def _run_diagnostics(row: dict):
  from dotenv import load_dotenv  # type: ignore

  load_dotenv(os.path.join(REPO, ".env"))
  import mysql.connector  # type: ignore

  from client_statements_output_excel.export_client_workbook import _load_run_diagnostics  # type: ignore

  conn = mysql.connector.connect(
    host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
  )
  try:
    return _load_run_diagnostics(conn, draft_id=str(row.get("draft_id") or ""))
  finally:
    conn.close()


def dump(out_path: str, business: str) -> None:
  import openpyxl  # type: ignore

  from client_intake_and_finmo.post_intake_runtime_validation.workbook_model_status import (  # type: ignore
    _recalc_workbook_via_excel_com,
  )
  from client_statements_output_excel.data import draft_data_from_row  # type: ignore
  from client_statements_output_excel.workbook_builder import (  # type: ignore
    build_client_financial_model_workbook,
  )

  row = _draft_row(business)
  # Pass run_diagnostics exactly as the production exporter does, so the
  # Diagnostics sheet is compared with real content rather than labels only.
  wb = build_client_financial_model_workbook(draft_data_from_row(row, run_diagnostics=_run_diagnostics(row)))
  xlsx = os.path.splitext(out_path)[0] + ".xlsx"
  wb.save(xlsx)
  err = _recalc_workbook_via_excel_com(xlsx)
  if err:
    raise SystemExit(f"EXCEL REFUSED / recalc failed: {err}")

  formulas = openpyxl.load_workbook(xlsx)
  values = openpyxl.load_workbook(xlsx, data_only=True)
  payload: dict = {"business": row["business_name"], "draft_id": row["draft_id"],
                   "sheets": {}, "sheet_order": list(formulas.sheetnames)}
  for ws in formulas.worksheets:
    vs = values[ws.title]
    cells = {}
    for line in ws.iter_rows():
      for cell in line:
        f = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
        v = vs[cell.coordinate].value
        if f is None and v is None:
          continue
        cells[cell.coordinate] = {"f": f, "v": v if not isinstance(v, float) else round(v, 6)}
    payload["sheets"][ws.title] = cells
  with open(out_path, "w", encoding="utf-8") as fh:
    json.dump(payload, fh, sort_keys=True, default=str)
  total = sum(len(c) for c in payload["sheets"].values())
  fcount = sum(1 for c in payload["sheets"].values() for x in c.values() if x["f"])
  print(f"dumped {len(payload['sheets'])} sheets, {total} populated cells, {fcount} formulas -> {out_path}")
  print(f"order: {payload['sheet_order']}")


def compare(base_path: str, now_path: str) -> int:
  base = json.load(open(base_path, encoding="utf-8"))
  now = json.load(open(now_path, encoding="utf-8"))
  problems: list = []
  compared_cells = compared_formulas = 0

  new_sheets = [s for s in now["sheets"] if s not in base["sheets"]]
  gone_sheets = [s for s in base["sheets"] if s not in now["sheets"]]
  for s in gone_sheets:
    problems.append(f"SHEET REMOVED: {s}")
  for s in new_sheets:
    with_formulas = [c for c, x in now["sheets"][s].items() if x["f"]]
    if with_formulas:
      problems.append(f"NEW SHEET {s} CARRIES {len(with_formulas)} FORMULAS "
                      f"(a new formula-bearing sheet moves the R32 grid): {with_formulas[:5]}")

  for sheet, base_cells in base["sheets"].items():
    if sheet not in now["sheets"]:
      continue
    now_cells = now["sheets"][sheet]
    pass
    for coord, b in base_cells.items():
      n = now_cells.get(coord)
      if n is None:
        if b["f"]:
          problems.append(f"FORMULA LOST {sheet}!{coord}: {b['f']!r}")
        else:
          problems.append(f"VALUE LOST {sheet}!{coord}: {b['v']!r}")
        continue
      if b["f"] != n["f"]:
        problems.append(f"FORMULA CHANGED {sheet}!{coord}\n     base={b['f']!r}\n     now ={n['f']!r}")
      elif b["f"]:
        compared_formulas += 1
      if (sheet, coord) in VOLATILE_CELLS:
        continue
      bv, nv = b["v"], n["v"]
      compared_cells += 1
      if bv != nv:
        if isinstance(bv, (int, float)) and isinstance(nv, (int, float)):
          if abs(float(bv) - float(nv)) <= max(1e-9, abs(float(bv)) * 1e-12):
            continue
        problems.append(f"VALUE CHANGED {sheet}!{coord}\n     base={bv!r}\n     now ={nv!r}")
    for coord, n in now_cells.items():
      if coord not in base_cells and n["f"]:
        problems.append(f"FORMULA ADDED to existing sheet {sheet}!{coord}: {n['f']!r}")

  print(f"compared {compared_cells} values and {compared_formulas} formulas across "
        f"{len(base['sheets'])} pre-existing sheets")
  print(f"new sheets (allowed, must be formula-free): {new_sheets or 'none'}")
  print(f"sheet order base: {base['sheet_order']}")
  print(f"sheet order now : {now['sheet_order']}")
  if problems:
    print(f"\nPROBLEMS ({len(problems)}):")
    for p in problems[:40]:
      print("  -", p)
    if len(problems) > 40:
      print(f"  ... {len(problems) - 40} more")
    return 1
  print("\nIDENTICAL: every pre-existing formula string and every recalculated value "
        "is unchanged. X1 changed appearance only.")
  return 0



# ---------------------------------------------------------------------------
# LABEL-KEYED mode (2026-08-19) — for a STRUCTURAL turn.
#
# When sections move (break-even from under the P&L to under the CFS, a new
# ratios block, a hidden Calc sheet) every row NUMBER shifts, so comparing by
# cell coordinate reports thousands of false changes. What must still hold is
# the thing Nick actually cares about: every number that existed before still
# exists, under the same row label, in the same period column, with the same
# value. That is what this mode proves. Same rule as the R32 grid: a row is
# identified by its first non-formula text cell.
# ---------------------------------------------------------------------------

def _row_label(cells) -> Optional[str]:
  for cell in cells:
    v = cell.value
    if isinstance(v, str) and v.strip() and not v.startswith("="):
      return v.strip()
  return None


def dump_labeled(out_path: str, business: str) -> None:
  import openpyxl  # type: ignore
  from openpyxl.utils import get_column_letter  # type: ignore

  from client_intake_and_finmo.post_intake_runtime_validation.workbook_model_status import (  # type: ignore
    _recalc_workbook_via_excel_com,
  )
  from client_statements_output_excel.data import draft_data_from_row  # type: ignore
  from client_statements_output_excel.workbook_builder import (  # type: ignore
    build_client_financial_model_workbook,
  )

  row = _draft_row(business)
  wb = build_client_financial_model_workbook(
    draft_data_from_row(row, run_diagnostics=_run_diagnostics(row)))
  xlsx = os.path.splitext(out_path)[0] + ".xlsx"
  wb.save(xlsx)
  err = _recalc_workbook_via_excel_com(xlsx)
  if err:
    raise SystemExit(f"EXCEL REFUSED / recalc failed: {err}")

  values = openpyxl.load_workbook(xlsx, data_only=True)
  formulas = openpyxl.load_workbook(xlsx)
  payload: dict = {"business": row["business_name"], "sheets": {}}
  for ws in formulas.worksheets:
    vs = values[ws.title]
    rows: dict = {}
    for line in ws.iter_rows():
      label = _row_label(line)
      if not label:
        continue
      cols = {}
      for cell in line:
        v = vs[cell.coordinate].value
        if isinstance(v, (int, float)):
          cols[get_column_letter(cell.column)] = round(float(v), 6)
      if cols:
        rows.setdefault(label, {}).update(cols)
    payload["sheets"][ws.title] = rows
  with open(out_path, "w", encoding="utf-8") as fh:
    json.dump(payload, fh, sort_keys=True, default=str)
  n = sum(len(c) for r in payload["sheets"].values() for c in r.values())
  print(f"labeled dump: {len(payload['sheets'])} sheets, "
        f"{sum(len(r) for r in payload['sheets'].values())} labelled rows, {n} numeric cells -> {out_path}")


def compare_labeled(base_path: str, now_path: str) -> int:
  base = json.load(open(base_path, encoding="utf-8"))
  now = json.load(open(now_path, encoding="utf-8"))
  problems: list = []
  compared = 0
  moved_sheets = []

  for sheet, base_rows in base["sheets"].items():
    now_rows = now["sheets"].get(sheet)
    if now_rows is None:
      problems.append(f"SHEET REMOVED: {sheet}")
      continue
    for label, base_cols in base_rows.items():
      now_cols = now_rows.get(label)
      if now_cols is None:
        problems.append(f"ROW LOST {sheet} :: {label!r}")
        continue
      for col, bv in base_cols.items():
        nv = now_cols.get(col)
        if nv is None:
          problems.append(f"CELL LOST {sheet} :: {label!r} col {col} (was {bv})")
          continue
        compared += 1
        if abs(bv - nv) > max(1e-6, abs(bv) * 1e-9):
          problems.append(f"VALUE CHANGED {sheet} :: {label!r} col {col}: {bv} -> {nv}")
  for sheet in now["sheets"]:
    if sheet not in base["sheets"]:
      moved_sheets.append(sheet)

  print(f"compared {compared} numeric cells keyed by (sheet, row label, column)")
  print(f"new sheets: {moved_sheets or 'none'}")
  if problems:
    print(f"\nPROBLEMS ({len(problems)}):")
    for p in problems[:40]:
      print("  -", p)
    if len(problems) > 40:
      print(f"  ... {len(problems) - 40} more")
    return 1
  print("\nIDENTICAL: every pre-existing row still exists under the same label, "
        "in the same period column, with the same recalculated value.")
  return 0

if __name__ == "__main__":
  if len(sys.argv) < 2:
    raise SystemExit(__doc__)
  cmd = sys.argv[1]
  if cmd == "dump":
    dump(sys.argv[2], sys.argv[3])
  elif cmd == "compare":
    raise SystemExit(compare(sys.argv[2], sys.argv[3]))
  elif cmd == "dump-labeled":
    dump_labeled(sys.argv[2], sys.argv[3])
  elif cmd == "compare-labeled":
    raise SystemExit(compare_labeled(sys.argv[2], sys.argv[3]))
  else:
    raise SystemExit(f"unknown command {cmd!r}")
