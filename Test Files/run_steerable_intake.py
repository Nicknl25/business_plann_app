"""Steerable dual-GPT intake runner.

Reads an Excel control file (see "Steerable Intake Template.xlsx"), compiles the
scenario + controlled positions into a client-GPT briefing, and launches the
real dual-agent intake conversation (run_dual_agent_intake.py) with it.

The two GPTs have a live, natural conversation; the client-GPT is instructed to
deliver each controlled position from the file, in substance, at the natural
moment. Everything not controlled is improvised in character.

Usage:
  python "Test Files\\run_steerable_intake.py" "path\\to\\my_test.xlsx"
  python "Test Files\\run_steerable_intake.py" "path\\to\\my_tests.xlsx" --sheet "Law Firm 62"

--sheet picks a worksheet by name (default: the first sheet), so one workbook
can hold many test sheets. Any other arguments are passed through to
run_dual_agent_intake.py.
"""
import subprocess
import sys
from pathlib import Path

try:
  from openpyxl import load_workbook
except ImportError:  # pragma: no cover
  print("openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
  raise SystemExit(2)


THIS_DIR = Path(__file__).resolve().parent
DUAL_RUNNER = THIS_DIR / "run_dual_agent_intake.py"

SCENARIO_LABEL = "scenario"
NAME_LABEL = "business name"
ADDRESS_LABEL = "business address"
START_DATE_LABEL = "business start date"
CONTROLS_HEADER = "controlled answers"


def _cell_text(value) -> str:
  return " ".join(str(value if value is not None else "").split())


def parse_control_file(path: Path, sheet_name: str = ""):
  wb = load_workbook(str(path), data_only=True)
  if sheet_name:
    if sheet_name not in wb.sheetnames:
      raise SystemExit(
        f"Sheet {sheet_name!r} not found in {path.name}. Available sheets: {', '.join(wb.sheetnames)}"
      )
    ws = wb[sheet_name]
  else:
    ws = wb.worksheets[0]
  scenario = ""
  business_name = ""
  business_address = ""
  start_date = ""
  controls = []
  in_controls = False
  for row in ws.iter_rows(min_row=1, max_col=2):
    label = _cell_text(row[0].value if len(row) > 0 else "").lower()
    content = _cell_text(row[1].value if len(row) > 1 else "")
    if not label and not content:
      continue
    if CONTROLS_HEADER in label:
      in_controls = True
      continue
    if not in_controls:
      if label.startswith(SCENARIO_LABEL):
        scenario = content
      elif label.startswith(NAME_LABEL):
        business_name = content
      elif label.startswith(ADDRESS_LABEL):
        business_address = content
      elif label.startswith(START_DATE_LABEL):
        start_date = content
      continue
    # Controls section: column B is the position; column A is an optional topic label.
    if content:
      topic = _cell_text(row[0].value)
      controls.append(f"[{topic}] {content}" if topic else content)
  if not scenario:
    raise SystemExit("The control file must have a 'Scenario' row with the business story.")
  return scenario, business_name, business_address, start_date, controls


def compile_seed(scenario, business_name, business_address, controls) -> str:
  parts = [scenario.strip()]
  identity_bits = []
  if business_name:
    identity_bits.append(f'the business is named "{business_name}"')
  if business_address:
    identity_bits.append(f"located at {business_address}")
  if identity_bits:
    parts.append(
      "Business identity (use these exact details when giving business information): "
      + "; ".join(identity_bits) + "."
    )
  if controls:
    lines = "\n".join(f"- {item}" for item in controls)
    parts.append(
      "CONTROLLED POSITIONS. You must work each of the following into the conversation "
      "naturally at the appropriate moment. Your wording can be conversational, but the "
      "substance - every number, choice, and behavior below - must be delivered exactly "
      "as specified and confirmed if the consultant restates it:\n" + lines +
      "\nFor anything not covered above, answer naturally and stay consistent with these positions."
    )
  return "\n\n".join(parts)


def main(argv) -> int:
  if not argv or not _cell_text(argv[0]):
    print(__doc__)
    return 2
  control_path = Path(argv[0]).expanduser()
  if not control_path.exists():
    print(f"Control file not found: {control_path}", file=sys.stderr)
    return 2
  rest = list(argv[1:])
  sheet_name = ""
  if "--sheet" in rest:
    idx = rest.index("--sheet")
    if idx + 1 >= len(rest):
      print("--sheet requires a worksheet name", file=sys.stderr)
      return 2
    sheet_name = rest[idx + 1]
    del rest[idx:idx + 2]
  scenario, name, address, start_date, controls = parse_control_file(control_path, sheet_name)
  seed = compile_seed(scenario, name, address, controls)
  print(f"Loaded control file: {control_path}" + (f" (sheet: {sheet_name})" if sheet_name else ""))
  print(f"Controlled positions: {len(controls)}")
  print("--- compiled client briefing ---")
  print(seed)
  print("--- launching live dual-GPT intake ---", flush=True)
  cmd = [sys.executable, str(DUAL_RUNNER), seed]
  if start_date:
    cmd += ["--business-start-date", start_date]
  cmd += rest
  return subprocess.call(cmd)


if __name__ == "__main__":
  raise SystemExit(main(sys.argv[1:]))
