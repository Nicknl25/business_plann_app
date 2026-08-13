"""CW-031 mini-round-2 red-proof: the artifact detector must not score one
draft's verdict on ANOTHER draft's workbook.

THE BUG mini found inside the tier-1 fix. ``workbook_cogs_rows`` globbed
``<business name>*.xlsx`` and took the newest by mtime. Five drafts share the
name 'Thistledown Cycle and Service', so draft be84629ada44 -- the REAL client
run, whose two product rows carry no per-line COGS at all -- scored PASS on a
workbook produced by draft plcogsd6e3ed0b. The class tier 1 exists to close was
still open inside the fix.

WHAT THIS PROVES, against the production functions on real MySQL rows and real
workbooks on disk (no fixtures, no stubs):

  1. BINDING. be84629ada44 resolves to its OWN workbook (08-12 15-41-24), not
     to the newest file sharing its name, and the verdict flips PASS -> FAIL.
  2. LAW BULLET 2. The total row must be =SUM over exactly the per-line rows.
     A workbook whose total sums the wrong range fails.
  3. LAW BULLET 3. Sigma(line revenue x line pct) == blend == finmo COGS per
     quarter, read from the workbook's own literals.
  4. LAW BULLET 4 / the blend wearing per-line clothing. N rows all carrying
     ONE rate now fails BY DEFAULT, and the opt-out has to be stated.
  5. The positive control: a genuine multi-line workbook + written ops rows
     (plcogs433a848c) still PASSES the finished assertion. A gate that only
     ever says no proves nothing.

  .venv\\Scripts\\python.exe "Test Files\\_redproof_cw031_workbook_binding.py"
"""

from __future__ import annotations

import copy
import os
import sys
from pathlib import Path

from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parents[1]

# The real CW-030 client run: ops product rows 0/2 written. Its own workbook
# (08-12 15-41-24) carries ONE blended COGS row.
THISTLEDOWN_REAL = "be84629ada444df1bb043b7ffdfc0592"
# VS's hand-seeded draft: ops rows written 0.5405 / 0.2287, workbook 08-13
# 10-48-37 carries two per-line rows totalled by =SUM(D9:D10).
THISTLEDOWN_SEEDED = "plcogs433a848cf96145e69f55e063e8"
# The draft that actually produced the 08-13 10-48-37 workbook.
THISTLEDOWN_OTHER = "plcogsd6e3ed0b4f32459c881e4c22f7"
RAVENWOOD = "1070c6a560a04f3d971019a3787180bf"

FAILURES: list = []


def check(label: str, ok: bool, detail: str) -> None:
  print(f"  [{'PASS' if ok else 'FAIL'}] {label}: {detail}")
  if not ok:
    FAILURES.append(label)


def main() -> int:
  load_dotenv(REPO_ROOT / ".env", override=False)
  sys.path.insert(0, str(REPO_ROOT / "python"))
  sys.path.insert(0, str(REPO_ROOT / "python" / "client_intake_and_finmo"))
  from intake_submission import get_mysql_connection  # type: ignore
  from client_intake_and_finmo import issue_registry  # type: ignore
  from client_intake_and_finmo import workbook_delivery_record as wdr  # type: ignore

  delivery_dir = (os.getenv("FINMO_MODEL_DELIVERY_DIR") or "").strip()
  spec = {"kind": "workbook_cogs_rows", "sheet": "FINMO", "min_rows": 2}
  conn = get_mysql_connection()
  cur = conn.cursor()
  try:
    print("STEP 1 - the binding: each draft resolves to ITS OWN workbook")
    resolutions = {}
    for name, draft in (("real client run", THISTLEDOWN_REAL),
                        ("seeded per-line", THISTLEDOWN_SEEDED),
                        ("other draft", THISTLEDOWN_OTHER),
                        ("Ravenwood", RAVENWOOD)):
      res = wdr.resolve_workbook_for_draft(cur, draft, delivery_dir=delivery_dir)
      resolutions[draft] = res
      print(f"  {name:18} {draft[:12]} -> "
            f"{os.path.basename(res['path'] or '(none)')} [{res['basis']}]")
      print(f"      {res['detail']}")
    real_wb = os.path.basename(resolutions[THISTLEDOWN_REAL]["path"] or "")
    other_wb = os.path.basename(resolutions[THISTLEDOWN_OTHER]["path"] or "")
    check("the real client draft gets its own workbook",
          real_wb.startswith("Thistledown Cycle and Service -- 08-12-2026 15-41"),
          real_wb or "(nothing resolved)")
    check("it is NOT the newest file sharing the name",
          real_wb != "Thistledown Cycle and Service -- 08-13-2026 10-48-37.xlsx",
          "the newest-mtime file belongs to another draft")
    check("that newest file is bound to the draft that produced it",
          other_wb == "Thistledown Cycle and Service -- 08-13-2026 10-48-37.xlsx",
          other_wb or "(nothing resolved)")
    check("two drafts sharing a name never share a workbook",
          real_wb != other_wb, f"{real_wb} vs {other_wb}")

    print("\nSTEP 2 - the false PASS is gone (this is the bug, red on the real path)")
    real = issue_registry._assert_workbook_cogs_rows(cur, THISTLEDOWN_REAL, spec)
    print(f"  {real['verdict']} - {real['detail']}")
    check("the real client run FAILS the workbook assertion",
          real["verdict"] == "fail",
          "its own workbook carries one blended COGS row")
    check("the failure names the count, not a missing file",
          "per-line" in real["detail"],
          real["detail"][:90])

    print("\nSTEP 3 - the positive control: a genuine multi-line workbook PASSES")
    seeded = issue_registry._assert_workbook_cogs_rows(cur, THISTLEDOWN_SEEDED, spec)
    print(f"  {seeded['verdict']} - {seeded['detail']}")
    check("the multi-line workbook passes the finished assertion",
          seeded["verdict"] == "pass", seeded["detail"][:120])
    check("the total row was checked, not just the count",
          "SUM over exactly those rows" in seeded["detail"],
          "law bullet 2 asserted")
    check("the three COGS routes were reconciled",
          "== blend == finmo COGS" in seeded["detail"],
          "law bullet 3 asserted")

    print("\nSTEP 4 - law bullet 2 red: a total that sums the WRONG range fails")
    # Mutate the real artifact rather than describe the failure: same workbook,
    # total row re-pointed one row short of the per-line block.
    import openpyxl
    good_path = resolutions[THISTLEDOWN_SEEDED]["path"]
    scratch = REPO_ROOT / "Test Files" / "_redproof_wrongsum.xlsx"
    wb = openpyxl.load_workbook(good_path)
    ws = wb["FINMO"]
    total_row = next(r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=1).value).strip() == "Cost of Goods Sold")
    for col in range(2, ws.max_column + 1):
      val = ws.cell(row=total_row, column=col).value
      if isinstance(val, str) and val.upper().startswith("=SUM("):
        letter = ws.cell(row=total_row, column=col).column_letter
        ws.cell(row=total_row, column=col).value = f"=SUM({letter}{total_row - 1}:{letter}{total_row - 1})"
    wb.save(scratch)
    wb.close()

    original = issue_registry._assert_workbook_cogs_rows

    def _forced(_cur, _draft, _spec, _path=str(scratch)):
      import openpyxl as _x
      _wb = _x.load_workbook(_path)
      try:
        _ws = _wb["FINMO"]
        rows = [(r, t) for r, t in issue_registry._sheet_rows_by_label(_ws)
                if t.startswith("Cost of Goods Sold")]
        per_line = [r for r, t in rows if t != "Cost of Goods Sold"]
        totals = [r for r, t in rows if t == "Cost of Goods Sold"]
        return issue_registry._assert_total_sums_over_lines(
          _ws, per_line, totals[0], "Cost of Goods Sold")
      finally:
        _wb.close()

    problem = _forced(cur, THISTLEDOWN_SEEDED, spec)
    print(f"  wrong-range total -> {problem}")
    check("a total summing the wrong range is caught",
          problem is not None and "sums" in (problem or ""),
          (problem or "NOT CAUGHT")[:100])
    try:
      os.remove(scratch)
    except Exception:
      pass

    print("\nSTEP 5 - law bullet 3 red: a per-line rate that breaks the reconciliation")
    wb = openpyxl.load_workbook(good_path)
    rd = wb["Revenue Drivers"]
    moved = 0
    for r, t in issue_registry._sheet_rows_by_label(rd):
      if t.endswith(" - COGS %"):
        for col in range(2, rd.max_column + 1):
          v = rd.cell(row=r, column=col).value
          if isinstance(v, (int, float)):
            rd.cell(row=r, column=col).value = float(v) * 1.5
            moved += 1
        break
    verdict, detail = issue_registry._reconcile_workbook_cogs(wb, "Cost of Goods Sold")
    wb.close()
    print(f"  one line's rate x1.5 ({moved} cells) -> {verdict}: {detail[:110]}")
    check("a per-line rate that no longer reconciles is caught",
          verdict == "fail", detail[:100])

    print("\nSTEP 6 - the blend wearing per-line clothing fails BY DEFAULT")
    ops = issue_registry._load_ops_model(cur, THISTLEDOWN_SEEDED)
    same_rate = copy.deepcopy(ops)
    for lob in same_rate.get("lob_models") or []:
      for product in lob.get("products") or []:
        product["cogs_percent_of_line_revenue"] = 0.42
    real_loader = issue_registry._load_ops_model
    try:
      issue_registry._load_ops_model = lambda _c, _d: same_rate  # type: ignore
      default_spec = {"kind": "ops_per_line_cogs", "min_lines": 2}
      out = issue_registry._assert_ops_per_line_cogs(cur, THISTLEDOWN_SEEDED, default_spec)
      print(f"  default          -> {out['verdict']}: {out['detail'][:90]}")
      check("N rows at one rate fails with NO flag set",
            out["verdict"] == "fail" and "wearing per-line clothing" in out["detail"],
            "distinct rates are the default")
      opted = issue_registry._assert_ops_per_line_cogs(
        cur, THISTLEDOWN_SEEDED,
        {"kind": "ops_per_line_cogs", "min_lines": 2, "allow_shared_rates": True})
      print(f"  allow_shared_rates -> {opted['verdict']}: {opted['detail'][:90]}")
      check("a STATED opt-out still passes",
            opted["verdict"] == "pass", "the collapse case has a door")
      try:
        issue_registry._assert_ops_per_line_cogs(
          cur, THISTLEDOWN_SEEDED,
          {"kind": "ops_per_line_cogs", "require_distinct_rates": False})
        check("the retired flag fails loud", False, "no exception raised")
      except ValueError as exc:
        check("the retired flag fails loud", "retired" in str(exc), str(exc)[:80])
    finally:
      issue_registry._load_ops_model = real_loader  # type: ignore
    # the real rows are untouched
    after = issue_registry._assert_ops_per_line_cogs(
      cur, THISTLEDOWN_SEEDED, {"kind": "ops_per_line_cogs", "min_lines": 2})
    check("the seeded draft's real rows are unchanged",
          after["verdict"] == "pass" and "distinct" in after["detail"],
          after["detail"][:90])

    print("\nSTEP 7 - an unattributable workbook is not_applicable, never a guess")
    ghost = issue_registry._assert_workbook_cogs_rows(
      cur, "draft_that_never_ran_000000000000", spec)
    print(f"  {ghost['verdict']} - {ghost['detail']}")
    check("a draft with no workbook returns not_applicable",
          ghost["verdict"] == "not_applicable", ghost["detail"][:90])

    print("\nSTEP 8 - Ravenwood, the run that started all of this, still fails")
    rav = issue_registry._assert_workbook_cogs_rows(cur, RAVENWOOD, spec)
    print(f"  {rav['verdict']} - {rav['detail']}")
    check("the four-line business with one blended row fails",
          rav["verdict"] == "fail", rav["detail"][:100])
  finally:
    try:
      cur.close()
      conn.close()
    except Exception:
      pass

  print("\n" + "=" * 72)
  if FAILURES:
    print(f"RED - {len(FAILURES)} check(s) failed: {FAILURES}")
    return 1
  print("GREEN - the workbook assertion is bound to its draft and asserts the whole law.")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
