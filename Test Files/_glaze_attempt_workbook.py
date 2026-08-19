"""Materialize the Glaze restructure ATTEMPT as a workbook: the 3-line
designs the solver evaluated (its chosen corner + the absolute full
corner), per-line revenue by quarter, cost stack, EBITDA/NI — plus the
solve's lever list. The evidence the persisted (reverted) FINMO cannot
show, made visible."""
import json
import os
import sys

import openpyxl
from dotenv import load_dotenv
import mysql.connector

load_dotenv()
sys.path.insert(0, "C:/dev/business_plann_app/python")

conn = mysql.connector.connect(
    host=os.getenv("MYSQL_HOST"),
    user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"),
    database=os.getenv("MYSQL_DB"),
    port=int(os.getenv("MYSQL_PORT") or 3306),
)
cur = conn.cursor(dictionary=True)
cur.execute(
    "SELECT model_input_json, repair_guidance_json, operating_model_json, financials_json, planning_runtime_json "
    "FROM intake_consult_drafts WHERE draft_id=%s",
    ("cc8b7081adec47b4b79f33a6231beb26",),
)
row = cur.fetchone()
mi = json.loads(row["model_input_json"])
g = json.loads(row["repair_guidance_json"])
ops = json.loads(row["operating_model_json"])
fin = json.loads(row["financials_json"])
runtime = json.loads(row["planning_runtime_json"])
cur.close()
conn.close()

bounds = next(it["bounds"] for it in g["restructure"]["history"] if it.get("stage") == "bounds")
solver_candidate = next(
    it["candidate"] for it in g["restructure"]["history"] if it.get("stage") == "search_1"
)

from client_intake_and_finmo.post_intake_restructure.joint_solver import (
    _base_levels, _lever_plan, _prepare_restructure_model,
)
from client_intake_and_finmo.post_intake_restructure.searcher import (
    apply_candidate, line_margins_from_bounds,
)
from client_intake_and_finmo.post_intake_restructure.fast_evaluator import build_fast_finmo

line_margins = line_margins_from_bounds(bounds)
prepared = _prepare_restructure_model(mi, bounds)
base_lv = _base_levels(mi)
stated_wages = float(fin.get("payroll_total_year1") or 0.0)
bf = (base_lv.get("annual_payroll") or 0.0) / stated_wages if stated_wages else 1.0
plan = _lever_plan(prepared, bounds, base_lv, bf)

full_corner = {
    "lines": {"primary line of business/donut": {
        "price_m11": 1.75, "price_m20": 1.75, "volume_m11": 4.0, "volume_m20": 4.0}},
    "new_lines": [
        {"lob": "Beverages", "product": "coffee & drinks", "unit_price": 3.0,
         "gross_margin_pct": 0.75, "q11_quarterly_revenue": 15000.0},
        {"lob": "Pre-orders & Catering", "product": "bulk donut orders", "unit_price": 1.5,
         "gross_margin_pct": 0.6, "q11_quarterly_revenue": 12000.0},
    ],
    "annual_payroll": 115000.0, "quarterly_rent": 4500.0,
    "marketing_pct": 0.04, "g_and_a_pct": 0.12,
}

wb = openpyxl.Workbook()
wb.remove(wb.active)


def line_revenue_series(mi_applied):
    groups = {}
    for r in ((mi_applied.get("sections") or {}).get("revenue") or []):
        if not isinstance(r, dict):
            continue
        key = f"{r.get('lob')} / {r.get('product')}"
        groups.setdefault(key, {})[str(r.get("driver") or "").strip()] = r.get("values") or []
    out = {}
    for key, drv in groups.items():
        series = []
        for q in range(1, 21):
            try:
                series.append(
                    float(drv["Capacity"][q]) * float(drv["Utilization"][q]) * float(drv["Unit Price"][q])
                )
            except (KeyError, TypeError, ValueError, IndexError):
                series.append(0.0)
        out[key] = series
    return out


def add_design_sheet(title, candidate):
    mi_applied = apply_candidate(mi, candidate, line_margins=line_margins or None)
    fm = build_fast_finmo(mi_applied)
    rows = {int(float(r.get("quarter_index"))): r for r in fm.get("quarter_rows") or [] if isinstance(r, dict)}
    per_line = line_revenue_series(mi_applied)
    ws = wb.create_sheet(title)
    header = ["Line / Metric"] + [f"Q{q}" for q in range(1, 21)]
    ws.append(header)
    for key in sorted(per_line.keys()):
        ws.append([f"REVENUE: {key}"] + [round(v) for v in per_line[key]])
    ws.append(["REVENUE: TOTAL"] + [round(float((rows.get(q) or {}).get("revenue") or 0)) for q in range(1, 21)])
    for label, field in (
        ("Payroll", "payroll"), ("Rent", "lease_rent"), ("COGS", "cogs"),
        ("Marketing", "marketing"), ("G&A", "general_and_administrative"),
        ("EBITDA", "ebitda"), ("Net Income", "net_income"),
    ):
        ws.append([label] + [round(float((rows.get(q) or {}).get(field) or 0)) for q in range(1, 21)])
    ws.append(["EBITDA margin"] + [
        round(float((rows.get(q) or {}).get("ebitda") or 0) / max(1.0, float((rows.get(q) or {}).get("revenue") or 0)), 3)
        for q in range(1, 21)
    ])
    ws.column_dimensions["A"].width = 48


add_design_sheet("Solver chosen corner", solver_candidate)
add_design_sheet("Absolute full corner", full_corner)

ws = wb.create_sheet("Solve lever list")
ws.append(["Adjustable cell (lever id)", "Q11 lower bound", "Q11 upper bound", "Kind"])
for lever_id in sorted(plan["bounds"].keys()):
    b11 = plan["bounds"][lever_id].get(11) or ("", "")
    kind = "NEW LINE volume (0..market cap)" if lever_id in plan["new_line_of_lever"] else (
        "existing line" if lever_id.startswith("revenue::") else "cost structure")
    ws.append([lever_id, round(float(b11[0]), 4), round(float(b11[1]), 4), kind])
ws.column_dimensions["A"].width = 70

out_path = r"C:\dev\Cilient Plans\Sunny Glaze Donuts -- RESTRUCTURE ATTEMPT (not shipped, found no viable config).xlsx"
wb.save(out_path)
print("saved:", out_path)

# Console summary of the per-line Q11 split for both designs
for title, candidate in (("SOLVER CHOSEN", solver_candidate), ("FULL CORNER", full_corner)):
    mi_applied = apply_candidate(mi, candidate, line_margins=line_margins or None)
    fm = build_fast_finmo(mi_applied)
    rows = {int(float(r.get("quarter_index"))): r for r in fm.get("quarter_rows") or [] if isinstance(r, dict)}
    per_line = line_revenue_series(mi_applied)
    q11_total = float((rows.get(11) or {}).get("revenue") or 0)
    print(f"\n{title} — Q11 by line (total {q11_total:,.0f}):")
    for key, series in sorted(per_line.items()):
        print(f"  {key}: {series[10]:,.0f}")
    print(f"  payroll {float((rows.get(11) or {}).get('payroll') or 0):,.0f}  "
          f"EBITDA {float((rows.get(11) or {}).get('ebitda') or 0):,.0f}")
