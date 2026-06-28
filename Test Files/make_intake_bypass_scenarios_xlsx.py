"""Generate the EXHAUSTIVE, pre-filled intake-bypass scenario workbook.

Flattens a captured baseline snapshot into one row per leaf, pre-filled with the
baseline value, grouped by section. The user edits any cell to override that
field for a scenario; unedited cells reproduce the baseline exactly.

  python "Test Files/make_intake_bypass_scenarios_xlsx.py"
  python "Test Files/make_intake_bypass_scenarios_xlsx.py" --baseline sunny_glaze_donuts --sheet Sunny_Glaze_Donuts
"""

from __future__ import annotations

import argparse
import importlib.util
from pathlib import Path
from typing import Any, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

THIS_DIR = Path(__file__).resolve().parent


def _load_common():
  spec = importlib.util.spec_from_file_location("intake_bypass_common", str(THIS_DIR / "intake_bypass_common.py"))
  module = importlib.util.module_from_spec(spec)
  spec.loader.exec_module(module)
  return module


C = _load_common()

README_LINES: List[Tuple[str, str]] = [
  ("intake-bypass scenarios — exhaustive override sheet", ""),
  ("", ""),
  ("Each non-underscore sheet is ONE scenario. Sheets starting with '_' are ignored.", ""),
  ("Column A = field path, Column B = value.", ""),
  ("", ""),
  ("HOW IT WORKS", ""),
  ("Every leaf of the baseline is listed and PRE-FILLED with the baseline value.", ""),
  ("To build a scenario, edit only the cells you want to change.", ""),
  ("A row is applied only when its value DIFFERS from the baseline at that path,", ""),
  ("so an unedited sheet reproduces the baseline exactly.", ""),
  ("", ""),
  ("CONVENTIONS", ""),
  ("baseline", "REQUIRED. Name of the baseline snapshot in intake_bypass_baselines/."),
  ("blank cell", "inherit the baseline value (no change)."),
  ("(null)", "explicitly set the field to null."),
  ("rows starting with #", "comments / section headers — ignored."),
  ("", ""),
  ("PATH SYNTAX", ""),
  ("draft.<col>", "a flat draft column, e.g. draft.business_name, draft.address_city."),
  ("<payload>.<path>", "a leaf inside a structured JSON column."),
  ("list items use [i]", "e.g. operating_model_json.lob_models[0].products[0].unit_price"),
  ("payloads exposed", ", ".join(C.STRUCTURED_PAYLOADS)),
  ("", ""),
  ("NUMBERS", "plain (20000), separated ($20,000), or percent (29%) all parse."),
  ("", ""),
  ("DENORMALIZATION — edit consistently", ""),
  ("unit_price / units_per_week_capacity / utilization_rate appear in BOTH", ""),
  ("operating_model_json (top-level AND each product) and financials_year1_json", ""),
  ("products. To change price/capacity/util coherently, edit every occurrence.", ""),
  ("", ""),
  ("ADDING NEW ARRAY ELEMENTS", ""),
  ("You CAN add a brand-new array element by adding a row whose path indexes", ""),
  ("past the baseline (the applier builds the structure). E.g. add a 2nd LOB:", ""),
  ("operating_model_json.lob_models[1].lob_name  +  ...[1].products[0].unit_price etc.", ""),
  ("Fill every leaf of the new element you care about; keep it coherent with", ""),
  ("financials_year1_json.lobs (denormalized). (For headcount scale,", ""),
  ("financials_json.current_num_employees / payroll_total_year1 are the knobs —", ""),
  (" post-intake authors the full headcount grid.)", ""),
]

SECTION_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")


def _build_rows(baseline: dict) -> List[Tuple[str, Any, bool]]:
  """Return (path, value, is_section_header) rows for the scenario sheet."""
  rows: List[Tuple[str, Any, bool]] = []
  flat = baseline.get("flat") or {}
  structured = baseline.get("structured") or {}

  rows.append(("# ===== draft (business identity / address) =====", "", True))
  for _sql_col, key in C.BASELINE_FLAT_COLUMNS:
    rows.append((f"draft.{key}", flat.get(key), False))

  for payload in C.STRUCTURED_PAYLOADS:
    obj = structured.get(payload)
    if obj is None:
      continue
    leaves: List[Tuple[str, Any]] = []
    C.flatten_obj(payload, obj, leaves)
    rows.append((f"# ===== {payload}  ({len(leaves)} fields) =====", "", True))
    rows.extend((p, v, False) for p, v in leaves)
  return rows


def _write_readme(ws) -> None:
  bold = Font(bold=True)
  for i, (a, b) in enumerate(README_LINES, start=1):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
  for r in (1, 6, 12, 18, 27, 32):
    ws.cell(row=r, column=1).font = bold
  ws.column_dimensions["A"].width = 42
  ws.column_dimensions["B"].width = 78


def _write_scenario(ws, baseline_name: str, rows: List[Tuple[str, Any, bool]]) -> None:
  bold = Font(bold=True)
  white_bold = Font(bold=True, color="FFFFFFFF")
  ws.cell(row=1, column=1, value="field").font = white_bold
  ws.cell(row=1, column=2, value="value").font = white_bold
  ws.cell(row=1, column=1).fill = HEADER_FILL
  ws.cell(row=1, column=2).fill = HEADER_FILL
  ws.cell(row=2, column=1, value="baseline").font = bold
  ws.cell(row=2, column=2, value=baseline_name).font = bold

  r = 3
  for path, value, is_header in rows:
    ca = ws.cell(row=r, column=1, value=path)
    if is_header:
      ca.font = bold
      ca.fill = SECTION_FILL
      ws.cell(row=r, column=2).fill = SECTION_FILL
    else:
      cb = ws.cell(row=r, column=2)
      if value is not None:
        cb.value = value  # native int/float/bool/str preserved
    r += 1

  ws.freeze_panes = "A2"
  ws.column_dimensions["A"].width = 62
  ws.column_dimensions["B"].width = 40
  for row in ws.iter_rows(min_row=1, max_col=2):
    row[1].alignment = Alignment(wrap_text=False, vertical="top")


def main(argv=None) -> int:
  parser = argparse.ArgumentParser(description="Generate the exhaustive pre-filled intake-bypass workbook.")
  parser.add_argument("--baseline", default="sunny_glaze_donuts")
  parser.add_argument("--sheet", default="Sunny_Glaze_Donuts")
  parser.add_argument("--baselines-dir", default=str(C.DEFAULT_BASELINES_DIR))
  parser.add_argument("--out", default=str(C.DEFAULT_SCENARIOS_XLSX))
  parser.add_argument(
    "--append", action="store_true",
    help="Add/replace this sheet in the EXISTING workbook (keeps other scenario "
         "sheets) instead of overwriting the file with a single sheet.",
  )
  args = parser.parse_args(argv)

  baseline = C.load_baseline(Path(args.baselines_dir), args.baseline)
  rows = _build_rows(baseline)
  leaf_count = sum(1 for _, _, h in rows if not h)

  out_path = Path(args.out)
  if args.append and out_path.exists():
    wb = openpyxl.load_workbook(str(out_path))
    # Refresh the _README (kept first) so the conventions stay current.
    if "_README" in wb.sheetnames:
      del wb["_README"]
    _write_readme(wb.create_sheet("_README", 0))
    # Replace an existing sheet of the same name; otherwise append a new one.
    if args.sheet in wb.sheetnames:
      del wb[args.sheet]
    _write_scenario(wb.create_sheet(args.sheet), args.baseline, rows)
    mode = "appended to"
  else:
    wb = openpyxl.Workbook()
    _write_readme(wb.active)
    wb.active.title = "_README"
    _write_scenario(wb.create_sheet(args.sheet), args.baseline, rows)
    mode = "wrote"
  wb.save(str(out_path))
  sheets = [s for s in wb.sheetnames if not s.startswith("_")]
  print(f"{mode} {out_path}")
  print(f"  scenario sheet {args.sheet!r}: {leaf_count} editable fields (baseline={args.baseline})")
  print(f"  scenario sheets now in workbook: {sheets}")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
