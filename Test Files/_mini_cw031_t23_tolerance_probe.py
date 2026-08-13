"""CW-031 mini audit, tiers 2/3 -- ITEM 4: is 0.5% too loose to catch a real defect?

VS's reconciliation (law bullet 3) compares three routes to a quarter's COGS and
allows max($1, 0.5% of Sigma). VS checked it on a workbook whose per-line rates
are round-ish; the question is what happens when they are NOT, when the stub
column rounds differently from the quarters, and -- the part that matters --
how big a defect has to be before the check actually fires.

So: take the genuine multi-line workbook (the one VS did NOT hand-build for this
check being unavailable, this is the seeded Thistledown model), rewrite its
per-line rate literals to deliberately ugly numbers with the stub rounded to 4dp
and the quarters to 6dp, rebuild the blend and the engine row so the workbook is
INTERNALLY TRUE, and confirm it still reconciles. Then walk a defect in from
large to small on ONE line and report the smallest rate error the check catches.

  .venv\\Scripts\\python.exe "Test Files\\_mini_cw031_t23_tolerance_probe.py"
"""

from __future__ import annotations

import os
import shutil
import sys
import tempfile
from pathlib import Path

from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parents[1]
COGS_LABEL = "Cost of Goods Sold"

# Deliberately ugly, non-round, and close together (the hard case: two lines
# whose rates nearly agree hide a defect inside the blend better than two that
# differ wildly).
UGLY = [0.536173, 0.210937]


def load(path):
  import openpyxl
  return openpyxl.load_workbook(path)


def describe(wb, reg):
  rd = wb["Revenue Drivers"]
  lines = {}
  for row, text in reg._sheet_rows_by_label(rd):
    if " - " not in text:
      continue
    name, _, field = text.rpartition(" - ")
    key = field.strip().lower()
    if key in ("capacity", "unit price", "utilization", "cogs %"):
      lines.setdefault(name.strip(), {})[key] = row
  return {n: f for n, f in lines.items()
          if {"capacity", "unit price", "utilization", "cogs %"} <= set(f)}


def rewrite(path, reg, rates, *, defect=None):
  """Write `rates` onto the per-line COGS rows and make the blend and the engine
  row TRUE for them, per period column. `defect` = (line_index, delta) perturbs
  ONE line's stored rate AFTER the blend was computed -- i.e. the model's blend
  no longer reflects the line rates, which is exactly the defect bullet 3 exists
  to catch."""
  import openpyxl
  wb = openpyxl.load_workbook(path)
  rd, mi, aud = wb["Revenue Drivers"], wb["Model Inputs"], wb["Audit Source"]
  usable = describe(wb, reg)
  names = list(usable)
  rd_p, mi_p, aud_p = (reg._period_columns(rd), reg._period_columns(mi),
                       reg._period_columns(aud))
  blend_row = next((r for r, t in reg._sheet_rows_by_label(mi) if t == COGS_LABEL), None)
  audit_row = next((r for r, t in reg._sheet_rows_by_label(aud) if t == COGS_LABEL), None)
  shared = [p for p in rd_p if p in mi_p and p in aud_p]
  for period in shared:
    col = rd_p[period]
    sigma = total_rev = 0.0
    complete = True
    for index, name in enumerate(names):
      fields = usable[name]
      cap = reg._numeric(rd.cell(row=fields["capacity"], column=col).value)
      price = reg._numeric(rd.cell(row=fields["unit price"], column=col).value)
      util = reg._numeric(rd.cell(row=fields["utilization"], column=col).value)
      if None in (cap, price, util):
        complete = False
        break
      # The stub column carries a 4dp rate; the quarters carry 6dp. That is the
      # differing-rounding case VS's 0.5% tolerance is said to be absorbing.
      rate = round(rates[index], 4 if period.lower() == "stub" else 6)
      rd.cell(row=fields["cogs %"], column=col).value = rate
      line_rev = cap * price * util
      total_rev += line_rev
      sigma += line_rev * rate
    if not complete or total_rev <= 0:
      continue
    mi.cell(row=blend_row, column=mi_p[period]).value = sigma / total_rev
    aud.cell(row=audit_row, column=aud_p[period]).value = sigma
    if defect is not None:
      index, delta = defect
      fields = usable[names[index]]
      current = reg._numeric(rd.cell(row=fields["cogs %"], column=col).value)
      rd.cell(row=fields["cogs %"], column=col).value = round(current + delta, 6)
  out = path.replace(".xlsx", ".probe.xlsx")
  wb.save(out)
  wb.close()
  return out, names


def main() -> int:
  load_dotenv(REPO_ROOT / ".env", override=False)
  sys.path.insert(0, str(REPO_ROOT / "python"))
  sys.path.insert(0, str(REPO_ROOT / "python" / "client_intake_and_finmo"))
  from client_intake_and_finmo import issue_registry as reg  # type: ignore

  delivery_dir = (os.getenv("FINMO_MODEL_DELIVERY_DIR") or "").strip()
  source = os.path.join(delivery_dir,
                        "Thistledown Cycle and Service -- 08-12-2026 17-16-30.xlsx")
  if not os.path.isfile(source):
    print(f"multi-line workbook missing: {source}")
    return 1

  with tempfile.TemporaryDirectory() as tmp:
    work = os.path.join(tmp, os.path.basename(source))
    shutil.copyfile(source, work)

    print("=" * 78)
    print("0. THE WORKBOOK AS DELIVERED (VS's positive control)")
    print("=" * 78)
    wb = load(work)
    usable = describe(wb, reg)
    print(f"  Revenue Drivers lines with a full driver set: {list(usable)}")
    verdict, detail = reg._reconcile_workbook_cogs(wb, COGS_LABEL)
    wb.close()
    print(f"  reconcile: {verdict} - {detail}")

    print()
    print("=" * 78)
    print("1. UGLY RATES, STUB AT 4dp AND QUARTERS AT 6dp, INTERNALLY TRUE")
    print("=" * 78)
    print(f"  rates: {UGLY}")
    path, names = rewrite(work, reg, UGLY)
    wb = load(path)
    verdict, detail = reg._reconcile_workbook_cogs(wb, COGS_LABEL)
    wb.close()
    print(f"  reconcile: {verdict} - {detail}")
    if verdict != "pass":
      print("  ^ a TRUE workbook fails: the tolerance is too TIGHT for real rounding")

    print()
    print("=" * 78)
    print("2. HOW BIG MUST A DEFECT BE? one line's rate drifts off the blend")
    print("=" * 78)
    print(f"  perturbing {names[0]!r} only (the blend and the engine row keep the")
    print("  TRUE value, so this is 'the split stopped reaching the model')")
    print(f"  {'rate error':>12}  {'stored rate':>12}  verdict")
    caught_at = None
    for delta in (0.20, 0.10, 0.05, 0.02, 0.01, 0.005, 0.002, 0.001, 0.0005):
      path, _ = rewrite(work, reg, UGLY, defect=(0, delta))
      wb = load(path)
      verdict, detail = reg._reconcile_workbook_cogs(wb, COGS_LABEL)
      wb.close()
      print(f"  {delta * 100:>11.2f}%  {UGLY[0] + delta:>12.6f}  {verdict}"
            + ("" if verdict == "fail" else f"   <- MISSED  ({detail[:60]})"))
      if verdict == "fail":
        caught_at = delta
    print(f"  -> smallest rate error caught on this line: "
          f"{(caught_at or 0) * 100:.2f} points")

    print()
    print("=" * 78)
    print("3. THE DEFECT THAT MATTERS: the model reverts to a BLEND")
    print("=" * 78)
    # Every line at the blended rate, blend and engine row unchanged: the exact
    # shape of the Ravenwood failure (one blended COGS row instead of N).
    blended = sum(UGLY) / len(UGLY)
    path, _ = rewrite(work, reg, UGLY)
    wb = load(path)
    rd = wb["Revenue Drivers"]
    usable = describe(wb, reg)
    rd_p = reg._period_columns(rd)
    for name, fields in usable.items():
      for period, col in rd_p.items():
        rd.cell(row=fields["cogs %"], column=col).value = blended
    out = path.replace(".probe.xlsx", ".blend.xlsx")
    wb.save(out)
    wb.close()
    wb = load(out)
    verdict, detail = reg._reconcile_workbook_cogs(wb, COGS_LABEL)
    wb.close()
    print(f"  every line at the blended {blended:.6f}: {verdict} - {detail[:110]}")

    print()
    print("=" * 78)
    print("4. THE SWAP: two lines exchange rates, blend and engine unchanged")
    print("=" * 78)
    path, _ = rewrite(work, reg, UGLY)
    wb = load(path)
    rd = wb["Revenue Drivers"]
    usable = describe(wb, reg)
    rd_p = reg._period_columns(rd)
    names2 = list(usable)
    for period, col in rd_p.items():
      a = usable[names2[0]]["cogs %"]
      b = usable[names2[1]]["cogs %"]
      va = rd.cell(row=a, column=col).value
      vb = rd.cell(row=b, column=col).value
      rd.cell(row=a, column=col).value = vb
      rd.cell(row=b, column=col).value = va
    out = path.replace(".probe.xlsx", ".swap.xlsx")
    wb.save(out)
    wb.close()
    wb = load(out)
    verdict, detail = reg._reconcile_workbook_cogs(wb, COGS_LABEL)
    wb.close()
    print(f"  rates swapped between the two lines: {verdict} - {detail[:110]}")

  return 0


if __name__ == "__main__":
  raise SystemExit(main())
