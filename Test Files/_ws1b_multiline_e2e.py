"""WS1(b) FULL-DEPTH E2E: clone Thistledown (the real two-line draft),
stamp the judge-shaped per-line percents on its ops product rows (the
one home), system-run on the NEW code, then verify the production
artifacts end to end:
  (1) model_input_json carries one 'COGS %' row per slot (N-line),
  (2) the blend expenses row equals the revenue-weighted line sum
      (SIGMA invariant) on every quarter of the final checkpoint,
  (3) finmo cogs == SIGMA(line revenue x line pct) per quarter,
  (4) the exported workbook carries N per-line COGS formula rows and
      the total as a sum over them.
"""
import io
import json
import os
import sys
import uuid

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, "C:/dev/business_plann_app/python")
import requests
from dotenv import load_dotenv

load_dotenv("C:/dev/business_plann_app/.env")
import mysql.connector

BASE = "http://127.0.0.1:5050"
SRC = "be84629a"
new_draft = "plcogs" + uuid.uuid4().hex[:26]
new_client = "plc" + uuid.uuid4().hex[:10]

conn = mysql.connector.connect(
    host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
    autocommit=True,
)
cur = conn.cursor(dictionary=True)
cur.execute("SHOW COLUMNS FROM intake_consult_drafts")
cols = [r["Field"] for r in cur.fetchall()]
col_list = ", ".join(f"`{c}`" for c in cols)
sel_list = ", ".join(
    "%s" if c in ("draft_id", "client_id") else f"`{c}`" for c in cols
)
cur.execute(
    f"INSERT INTO intake_consult_drafts ({col_list}) "
    f"SELECT {sel_list} FROM intake_consult_drafts WHERE draft_id LIKE %s",
    (new_draft, new_client, SRC + "%"),
)
print("cloned ->", new_draft)

# Stamp the per-line percents on the ops rows (bikes 52%, repairs 22% -
# the judge's live proposal from the fixture run).
cur.execute("SELECT operating_model_json FROM intake_consult_drafts WHERE draft_id=%s", (new_draft,))
ops = json.loads(cur.fetchone()["operating_model_json"] or "{}")
stamped = []
for lm in ops.get("lob_models") or []:
    for pr in lm.get("products") or []:
        name = str(pr.get("product_name") or "").lower()
        pct = 0.52 if "sale" in name or "bike" in name else 0.22
        pr["cogs_percent_of_line_revenue"] = pct
        stamped.append((pr.get("product_name"), pct))
cur.execute(
    "UPDATE intake_consult_drafts SET operating_model_json=%s WHERE draft_id=%s",
    (json.dumps(ops), new_draft),
)
print("stamped:", stamped)

print("system run...")
resp = requests.post(f"{BASE}/api/intake-consult/system-run", json={
    "draft_id": new_draft, "client_id": new_client,
}, timeout=1800)
print("HTTP", resp.status_code)

cur.execute(
    "SELECT planning_run_id FROM planning_runs WHERE draft_id=%s ORDER BY created_at DESC LIMIT 1",
    (new_draft,))
run_id = (cur.fetchone() or {}).get("planning_run_id")
cur.execute(
    "SELECT stage, finmo_json, model_input_json FROM planning_run_checkpoints "
    "WHERE planning_run_id=%s AND finmo_json IS NOT NULL "
    "ORDER BY created_at DESC LIMIT 1", (run_id,))
ck = cur.fetchone() or {}
mi = json.loads(ck.get("model_input_json") or "{}")
fj = json.loads(ck.get("finmo_json") or "{}")

rev_rows = (mi.get("sections") or {}).get("revenue") or []
cogs_rows = [r for r in rev_rows if r.get("driver") == "COGS %"]
blend_row = next((r for r in (mi.get("sections") or {}).get("expenses") or []
                  if r.get("label") == "Cost of Goods Sold"), {})
print()
print("(1) COGS % rows:", len(cogs_rows), "slots:", [r.get("revenue_slot_key") for r in cogs_rows])

by = lambda d: {r["revenue_slot_key"]: r for r in rev_rows if r.get("driver") == d}
caps, prices, utils = by("Capacity"), by("Unit Price"), by("Utilization")
blend_vals = blend_row.get("values") or []
sigma_ok = True
worst = 0.0
for i in range(1, min(21, len(blend_vals))):
    tot_rev = 0.0
    line_cogs = 0.0
    for r in cogs_rows:
        k = r["revenue_slot_key"]
        rv = caps[k]["values"][i] * prices[k]["values"][i] * utils[k]["values"][i]
        tot_rev += rv
        line_cogs += rv * r["values"][i]
    blend_cogs = tot_rev * blend_vals[i]
    gap = abs(line_cogs - blend_cogs)
    worst = max(worst, gap / max(1.0, blend_cogs))
    if gap > max(1.0, 0.005 * max(blend_cogs, 1.0)):
        sigma_ok = False
        print(f"  q{i}: SIGMA {line_cogs:.0f} vs blend {blend_cogs:.0f} GAP")
print("(2) SIGMA==blend on all 20 quarters:", sigma_ok, f"(worst rel gap {worst:.5f})")

quarters = fj.get("quarters") or fj.get("quarter_rows") or []
finmo_ok = True
checked = 0
for q in quarters:
    qi = int(q.get("quarter_index") or 0)
    if qi < 1 or qi > 20 or qi >= len(blend_vals):
        continue
    line_cogs = 0.0
    for r in cogs_rows:
        k = r["revenue_slot_key"]
        rv = caps[k]["values"][qi] * prices[k]["values"][qi] * utils[k]["values"][qi]
        line_cogs += rv * r["values"][qi]
    fc = float(q.get("cost_of_goods_sold") or q.get("cogs") or 0.0)
    checked += 1
    if abs(fc - line_cogs) > max(1.0, 0.005 * max(fc, 1.0)):
        finmo_ok = False
        print(f"  finmo q{qi}: {fc:.0f} vs SIGMA {line_cogs:.0f} GAP")
print(f"(3) finmo cogs == SIGMA on {checked} quarters:", finmo_ok)

# (4) the workbook - exported to CLIENT_FINANCIAL_MODELS_DIR (there is
# no path column on the draft row); take the newest matching xlsx.
import glob
_wb_dir = os.getenv("CLIENT_FINANCIAL_MODELS_DIR") or r"C:\dev\Cilient Plans"
_candidates = sorted(
    glob.glob(os.path.join(_wb_dir, "*.xlsx")),
    key=os.path.getmtime, reverse=True,
)
wb_path = _candidates[0] if _candidates else None
print("(4) workbook:", wb_path)
if wb_path and os.path.exists(str(wb_path)):
    import openpyxl
    wb = openpyxl.load_workbook(str(wb_path))
    ws = wb["FINMO"] if "FINMO" in wb.sheetnames else None
    if ws is not None:
        labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 80)]
        pl_lines = [(r, l) for r, l in enumerate(labels, start=1) if l.startswith("Cost of Goods Sold - ")]
        total_row = next((r for r, l in enumerate(labels, start=1) if l == "Cost of Goods Sold"), None)
        print("    per-line P&L rows:", [l for _r, l in pl_lines])
        if total_row and pl_lines:
            f_total = ws.cell(row=total_row, column=4).value
            f_line = ws.cell(row=pl_lines[0][0], column=4).value
            print("    total formula:", f_total)
            print("    first line formula:", f_line)
            print("    total is SUM over lines:", str(f_total or "").startswith("=SUM("))
else:
    print("    workbook file missing")
cur.execute("SELECT planning_run_status FROM intake_consult_drafts WHERE draft_id=%s", (new_draft,))
print("status:", cur.fetchone())
