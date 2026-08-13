"""CW-031 tier-1 mini audit: is each gate mechanism LIVE, and can 'confirmed'
still be earned at all?

The ablation harness showed two mechanisms survive their own removal against
VS's red-proof. That is a coverage statement, not a verdict: it says the proof
does not exercise them, and leaves open whether the code is load-bearing or
dead. This decides it, by seeding synthetic issues whose ONLY difference is the
thing under test and running the PRODUCTION evaluator over the real Ravenwood
draft.

  M1  opportunity-only hard issue      -> must NOT reach 'confirmed'   (gate i)
  M2  hard issue + PASSING artifact    -> must reach 'confirmed'       (gate i, positive)
  M3  metadata-only probe, quiet=4     -> must not tick at all         (gate iii)
  A1  N rows sharing one rate          -> require_distinct_rates       (check 3)
  P1  Thistledown multi-line workbook  -> must PASS workbook_cogs_rows (check 4)

Every synthetic issue is deleted afterwards; the registry is snapshotted and
restored. --worker runs one pass and prints JSON (the driver ablates around it).

  .venv\\Scripts\\python.exe "Test Files\\_mini_cw031_mechanism_probe.py"
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

from dotenv import load_dotenv

REPO_ROOT = Path(__file__).resolve().parents[1]
TARGET = REPO_ROOT / "python" / "client_intake_and_finmo" / "issue_registry.py"
PY = REPO_ROOT / ".venv" / "Scripts" / "python.exe"
DRAFT_ID = "1070c6a560a04f3d971019a3787180bf"

SYNTH = {
  "M1": "mini_audit:financials:opportunity_only_hard_issue_must_not_confirm",
  "M2": "mini_audit:financials:artifact_backed_hard_issue_must_confirm",
  "M3": "mini_audit:financials:metadata_only_probe_must_not_tick",
}
MUTABLE = ["status", "occurrence_count", "reopened_count", "clean_exercise_count",
           "runs_since_last_seen", "resolved_detected_at", "resolution_basis",
           "resolution_confidence", "probe_json", "last_seen_at"]

ABLATIONS = {
  "i": ('artifact_backed = rclass == "hard" and artifact["present"]',
        'artifact_backed = rclass == "hard"  # ABLATED'),
  "iii": ('  if not conditions:\n    return {"exercised": False,\n'
          '            "reason": "probe states no retest condition (metadata/notes only)"}',
          '  if False:  # ABLATED\n    pass'),
}


def _connect():
  load_dotenv(REPO_ROOT / ".env", override=False)
  sys.path.insert(0, str(REPO_ROOT / "python"))
  sys.path.insert(0, str(REPO_ROOT / "python" / "client_intake_and_finmo"))
  from intake_submission import get_mysql_connection  # type: ignore
  return get_mysql_connection()


def _purge(conn):
  cur = conn.cursor()
  for sig in SYNTH.values():
    cur.execute("SELECT issue_id FROM issues WHERE signature=%s", (sig,))
    row = cur.fetchone()
    if row:
      iid = int(row[0])
      cur.execute("DELETE FROM issue_occurrences WHERE issue_id=%s", (iid,))
      cur.execute("DELETE FROM issue_resolution_events WHERE issue_id=%s", (iid,))
      cur.execute("DELETE FROM issues WHERE issue_id=%s", (iid,))
  conn.commit()
  cur.close()


def worker() -> int:
  conn = _connect()
  from client_intake_and_finmo import issue_registry as reg  # type: ignore
  out = {}
  try:
    _purge(conn)
    probes = {
      # opportunity only: exactly the evidence the 51 demoted verdicts had.
      "M1": {"section": "financials"},
      # a real artifact that HOLDS on this draft (product_name is written on
      # all four rows) -- the positive control for 'confirmed' being reachable.
      "M2": {"section": "financials",
             "artifact": [{"kind": "ops_field_non_null",
                           "path": "products[].product_name"}]},
      # metadata only: no retest condition at all.
      "M3": {"note": "mini audit synthetic", "regression_pin": True},
    }
    for key, sig in SYNTH.items():
      reg.report_issue(
        conn, signature=sig, category="flow", severity="major",
        observed="synthetic issue seeded by the mini audit",
        expected="deleted before this script exits",
        draft_id="mini-audit-seed", probe=probes[key], source="probe",
      )
    cur = conn.cursor()
    # Age them so the SOFT path is one run from resolving too: that way a
    # verdict of 'nothing happened' cannot be blamed on a threshold.
    cur.execute(
      "UPDATE issues SET runs_since_last_seen=4, clean_exercise_count=0, "
      "occurrence_count=1 WHERE signature IN (%s, %s, %s)", tuple(SYNTH.values()))
    conn.commit()
    cur.close()

    summary = reg.evaluate_run_for_resolution(conn, draft_id=DRAFT_ID)
    out["summary"] = {k: v for k, v in summary.items()
                      if k not in ("resolved_confirmed", "resolved_observational")}
    for key, sig in SYNTH.items():
      issue = reg.get_issue(conn, signature=sig)
      out[key] = {"status": str(issue["status"]),
                  "basis": issue["resolution_basis"],
                  "confidence": issue["resolution_confidence"],
                  "clean": int(issue["clean_exercise_count"] or 0),
                  "quiet": int(issue["runs_since_last_seen"] or 0)}
  finally:
    try:
      _purge(conn)
      conn.close()
    except Exception:
      pass
  print("WORKER_JSON " + json.dumps(out))
  return 0


def _snapshot(conn):
  cur = conn.cursor()
  cur.execute(f"SELECT issue_id, {', '.join(MUTABLE)} FROM issues")
  issues = cur.fetchall()
  cur.execute("SELECT COALESCE(MAX(id),0) FROM issue_occurrences")
  mo = int(cur.fetchone()[0] or 0)
  cur.execute("SELECT COALESCE(MAX(id),0) FROM issue_resolution_events")
  me = int(cur.fetchone()[0] or 0)
  cur.close()
  return issues, mo, me


def _restore(conn, issues, mo, me):
  cur = conn.cursor()
  cur.execute("DELETE FROM issue_occurrences WHERE id > %s", (mo,))
  cur.execute("DELETE FROM issue_resolution_events WHERE id > %s", (me,))
  sets = ", ".join(f"{c}=%s" for c in MUTABLE)
  for row in issues:
    cur.execute(f"UPDATE issues SET {sets} WHERE issue_id=%s", (*row[1:], row[0]))
  conn.commit()
  cur.close()


def _run_worker():
  proc = subprocess.run([str(PY), str(Path(__file__).resolve()), "--worker"],
                        capture_output=True, text=True)
  for line in proc.stdout.splitlines():
    if line.startswith("WORKER_JSON "):
      return json.loads(line[len("WORKER_JSON "):])
  return {"error": (proc.stdout + proc.stderr)[-800:]}


def adversarial_and_positive(conn):
  """A1 (all rows share one rate) and P1 (Thistledown positive control)."""
  from client_intake_and_finmo import issue_registry as reg  # type: ignore
  print("\n== A1 - the blend wearing per-line clothing ==")
  fake_ops = {"lob_models": [{"products": [
    {"product_name": n, "cogs_percent_of_line_revenue": 0.42}
    for n in ("Plant sale", "Hard goods sale", "Install project", "Design consult")]}]}
  real_loader = reg._load_ops_model
  reg._load_ops_model = lambda cur, draft_id: fake_ops  # type: ignore
  try:
    lax = reg._assert_ops_per_line_cogs(None, DRAFT_ID, {"min_lines": 2})
    strict = reg._assert_ops_per_line_cogs(
      None, DRAFT_ID, {"min_lines": 2, "require_distinct_rates": True})
  finally:
    reg._load_ops_model = real_loader  # type: ignore
  print(f"  without require_distinct_rates: {lax['verdict']} - {lax['detail']}")
  print(f"  with    require_distinct_rates: {strict['verdict']} - {strict['detail']}")

  print("\n== P1 - workbook_cogs_rows scoping + the Thistledown positive control ==")
  cur = conn.cursor()
  cur.execute("SELECT draft_id, business_name FROM intake_consult_drafts "
              "WHERE business_name LIKE 'Thistledown%' ORDER BY created_at DESC LIMIT 5")
  cands = cur.fetchall()
  for did, name in cands:
    print(f"  candidate draft {did[:12]} {name!r}")
  for did, name in cands[:2]:
    v = reg._assert_workbook_cogs_rows(cur, did, {"sheet": "FINMO", "min_rows": 2})
    print(f"    -> {name!r} [{did[:12]}]: {v['verdict']} - {v['detail']}")
    o = reg._assert_ops_per_line_cogs(cur, did, {"min_lines": 2})
    print(f"       ops_per_line_cogs: {o['verdict']} - {o['detail']}")
    o2 = reg._assert_ops_per_line_cogs(
      cur, did, {"min_lines": 2, "require_distinct_rates": True})
    print(f"       + require_distinct_rates: {o2['verdict']} - {o2['detail']}")

  print("\n  per-sheet 'Cost of Goods Sold' label census (is FINMO the right scope?)")
  import glob
  import os
  import openpyxl
  ddir = os.getenv("FINMO_MODEL_DELIVERY_DIR") or ""
  for biz in ("Ravenwood Garden Company", "Thistledown Cycle and Service"):
    paths = sorted(glob.glob(os.path.join(ddir, f"{biz}*.xlsx")), key=os.path.getmtime)
    if not paths:
      print(f"    {biz}: no workbook")
      continue
    wb = openpyxl.load_workbook(paths[-1])
    print(f"    {os.path.basename(paths[-1])}")
    for sh in wb.sheetnames:
      ws = wb[sh]
      hits = [(c.row, str(c.value)) for (c,) in ws.iter_rows(min_col=1, max_col=1)
              if c.value is not None and str(c.value).strip().startswith("Cost of Goods Sold")]
      if hits:
        for r, v in hits:
          formula = ws.cell(row=r, column=4).value
          print(f"      [{sh}] row {r}: {v!r} D={formula!r}")
    wb.close()
  cur.close()


def main() -> int:
  if "--worker" in sys.argv:
    return worker()
  conn = _connect()
  # BYTES, not text. read_text/write_text round-trips LF -> CRLF on Windows, so
  # a "restored" file comes back content-identical and byte-different. Ablating
  # app code that is not mine has to leave it bit-for-bit as found.
  original_b = TARGET.read_bytes()
  original = original_b.decode("utf-8")
  try:
    snap = _snapshot(conn)
    results = {}
    print("== M1/M2/M3 - baseline (gate intact) ==")
    results["baseline"] = _run_worker()
    _restore(conn, *snap)
    print("  " + json.dumps(results["baseline"], indent=2)[:1200])

    for tag, (old, new) in ABLATIONS.items():
      assert original.count(old) == 1, f"anchor {tag} not unique"
      TARGET.write_bytes(original.replace(old, new).encode("utf-8"))
      try:
        results[f"ablate-{tag}"] = _run_worker()
      finally:
        TARGET.write_bytes(original_b)
      _restore(conn, *snap)
      print(f"\n== ablated ({tag}) ==")
      print("  " + json.dumps(results[f"ablate-{tag}"], indent=2)[:1200])

    print("\n== VERDICTS ==")
    b, ai, aiii = (results["baseline"], results["ablate-i"], results["ablate-iii"])
    for key, label in (("M1", "gate (i): opportunity-only must not confirm"),
                       ("M2", "gate (i) positive: artifact-backed CAN confirm"),
                       ("M3", "gate (iii): metadata-only must not tick")):
      other = aiii if key == "M3" else ai
      same = b.get(key) == other.get(key)
      print(f"  {key} {label}")
      print(f"      intact  : {b.get(key)}")
      print(f"      ablated : {other.get(key)}")
      print(f"      -> {'DECORATIVE (no behavioural difference)' if same else 'LIVE (removal changes the verdict)'}")

    adversarial_and_positive(conn)
    _restore(conn, *snap)
  finally:
    TARGET.write_bytes(original_b)
    try:
      conn.close()
    except Exception:
      pass
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
