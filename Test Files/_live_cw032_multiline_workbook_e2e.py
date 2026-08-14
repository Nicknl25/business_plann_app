"""CW-032 multi-line workbook E2E, the CLIENT PATH end to end.

Chain: clone of the REAL completed Alderfen draft (158f6816) -> the client's
own four-rate sentence through the LIVE wall router (no hand-stamping - the
rates enter the model the way a client's words do) -> a real system-run
(rerun POST names its own new run) -> the DELIVERED workbook.

Asserts the ruled CW-032 layout on the real product:
  (1) the final checkpoint's model_input carries FOUR 'COGS %' rows,
  (2) the Model Inputs sheet renders the four driver rows,
  (3) FINMO carries EXACTLY ONE 'Cost of Goods Sold' row whose formula is
      the four-term Sigma(line revenue x line driver) roll-up - no
      per-line P&L rows,
  (4) _assert_workbook_cogs_rows PASSES via the delivery-record binding,
  (5) the opening-PPE schedule reached the workbook (Q1 depreciation rate
      covers the straight-line share, ~5%/quarter at 5y).

  .venv\\Scripts\\python.exe "Test Files\\_live_cw032_multiline_workbook_e2e.py"
"""
import io
import json
import os
import re
import sys
import uuid

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.path.insert(0, "C:/dev/business_plann_app/python")
import requests
from dotenv import load_dotenv

load_dotenv("C:/dev/business_plann_app/.env")
import mysql.connector

BASE = "http://127.0.0.1:5050"
SRC = "158f6816"
new_draft = "cw32wb" + uuid.uuid4().hex[:26]
new_client = "CW32" + uuid.uuid4().hex[:10].upper()

RATES_MESSAGE = (
  "Close, but let me give you my actual numbers. Plants are 46%. Hardgoods "
  "are 73% - that's the pallet-of-pavers problem. Install is 17% in "
  "materials because the labour is all on my payroll. And design is 3%, "
  "just printing and the odd soil test."
)

FAILURES = []


def check(label, ok, detail):
  print(f"  [{'PASS' if ok else 'FAIL'}] {label}: {detail}")
  if not ok:
    FAILURES.append(label)


conn = mysql.connector.connect(
    host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
    password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
    autocommit=True,
)
cur = conn.cursor(dictionary=True)
cur.execute("SHOW COLUMNS FROM intake_consult_drafts")
cols = [r["Field"] for r in cur.fetchall()]
col_list = ", ".join(f"`{c}`" for c in cols)
_reset_null = {
    "planning_run_id", "planning_run_status", "planning_stage",
    "planning_status", "planning_latest_checkpoint_id",
    "planning_resume_from_checkpoint_id", "planning_source_run_id",
    "planning_superseded_by_run_id",
}
sel_list = ", ".join(
    "%s" if c in ("draft_id", "client_id") else ("NULL" if c in _reset_null else f"`{c}`")
    for c in cols
)
cur.execute(
    f"INSERT INTO intake_consult_drafts ({col_list}) "
    f"SELECT {sel_list} FROM intake_consult_drafts WHERE draft_id LIKE %s",
    (new_draft, new_client, SRC + "%"),
)
print("cloned completed Alderfen ->", new_draft)

print("\nSTEP 1 - the client's four rates through the LIVE wall router")
resp = requests.post(
    f"{BASE}/api/intake-consult",
    json={"draft_id": new_draft, "client_id": new_client, "message": RATES_MESSAGE},
    timeout=300,
)
body = resp.json() if resp.status_code == 200 else {}
print(f"  < [{resp.status_code}] {str(body.get('assistant_message') or '')[:250]}")
cur.execute("SELECT operating_model_json FROM intake_consult_drafts WHERE draft_id=%s", (new_draft,))
ops = json.loads(cur.fetchone()["operating_model_json"] or "{}")
rates = {str(p.get("product_name")): p.get("cogs_percent_of_line_revenue")
         for lm in ops.get("lob_models") or [] for p in lm.get("products") or []}
print("  ops rates:", rates)
check("wall turn landed all four client rates",
      rates == {"Plant sale": 0.46, "Hardgoods sale": 0.73,
                "Install job": 0.17, "Design project": 0.03}, str(rates))

print("\nSTEP 2 - real system-run (foreground; rerun names its own run)")
resp = requests.post(f"{BASE}/api/intake-consult/system-run", json={
    "draft_id": new_draft, "client_id": new_client,
}, timeout=1800)
print("  HTTP", resp.status_code)
cur.execute(
    "SELECT planning_run_status FROM intake_consult_drafts WHERE draft_id=%s",
    (new_draft,))
status = (cur.fetchone() or {}).get("planning_run_status")
check("system run completed", str(status) == "completed", str(status))

cur.execute(
    "SELECT planning_run_id FROM planning_runs WHERE draft_id=%s ORDER BY created_at DESC LIMIT 1",
    (new_draft,))
run_id = (cur.fetchone() or {}).get("planning_run_id")
cur.execute(
    "SELECT stage, model_input_json FROM planning_run_checkpoints "
    "WHERE planning_run_id=%s AND finmo_json IS NOT NULL "
    "ORDER BY created_at DESC LIMIT 1", (run_id,))
ck = cur.fetchone() or {}
mi = json.loads(ck.get("model_input_json") or "{}")
cogs_rows = [r for r in (mi.get("sections") or {}).get("revenue") or []
             if r.get("driver") == "COGS %"]
check("(1) model_input carries FOUR COGS % rows", len(cogs_rows) == 4,
      f"{len(cogs_rows)} rows: {[r.get('revenue_slot_key') for r in cogs_rows]}")

cur.execute(
    "SELECT source_path FROM workbook_deliveries WHERE draft_id=%s "
    "ORDER BY delivered_at DESC LIMIT 1", (new_draft,))
wb_path = (cur.fetchone() or {}).get("source_path")
print("\nSTEP 3 - the delivered workbook:", wb_path)
if not wb_path or not os.path.exists(str(wb_path)):
  check("delivered workbook exists", False, str(wb_path))
else:
  import openpyxl
  wb = openpyxl.load_workbook(wb_path)
  mi_ws = wb["Model Inputs"]
  drivers = [(r, str(mi_ws.cell(row=r, column=1).value or ""))
             for r in range(1, 200)
             if str(mi_ws.cell(row=r, column=1).value or "").endswith(" - COGS %")]
  check("(2) Model Inputs renders FOUR COGS driver rows", len(drivers) == 4,
        str([d[1] for d in drivers]))
  ws = wb["FINMO"]
  labelled = [(r, str(ws.cell(row=r, column=1).value or ""))
              for r in range(1, 90)
              if str(ws.cell(row=r, column=1).value or "").startswith("Cost of Goods Sold")]
  per_line = [t for _r, t in labelled if t != "Cost of Goods Sold"]
  totals = [r for r, t in labelled if t == "Cost of Goods Sold"]
  check("(3a) FINMO carries EXACTLY ONE COGS row, zero per-line rows",
        len(totals) == 1 and not per_line,
        f"totals={len(totals)} per_line={per_line}")
  formula = str(ws.cell(row=totals[0], column=4).value or "") if totals else ""
  terms = formula.lstrip("=").split("+")
  term_re = re.compile(r"^'?Model Inputs'?![A-Z]{1,3}\d+\*'?Model Inputs'?![A-Z]{1,3}\d+$")
  check("(3b) the one cell IS the four-term roll-up",
        len(terms) == 4 and all(term_re.match(t.strip()) for t in terms),
        formula[:120])
  # (5) opening-PPE depreciation reached the schedule
  cap = wb[[n for n in wb.sheetnames if "CapEx" in n][0]]
  dep_q1 = None
  for r in range(1, 40):
    if str(cap.cell(row=r, column=1).value or "") == "Depreciation Rate":
      dep_q1 = cap.cell(row=r, column=5).value  # Q1
  check("(5) Q1 depreciation rate carries the straight-line share",
        isinstance(dep_q1, (int, float)) and dep_q1 >= 0.049,
        f"q1 rate={dep_q1} (>= 0.05 - 5y straight line + capex vintages)")
  wb.close()

  # (4) the artifact assertion, via the real delivery binding. Plain
  # (tuple) cursor - the production evaluator's shape; a dictionary
  # cursor breaks the resolver's row[0] reads.
  import client_intake_and_finmo.issue_registry as reg
  r4 = reg._assert_workbook_cogs_rows(
      conn.cursor(), new_draft,
      {"min_rows": 4})
  check("(4) workbook_cogs_rows assertion PASSES on the delivery binding",
        r4.get("verdict") == "pass", str(r4.get("detail"))[:200])

print("\n" + "=" * 72)
if FAILURES:
  print(f"RED - {len(FAILURES)} check(s) failed: {FAILURES}")
  sys.exit(1)
print("GREEN - client words -> live router -> ops rows -> system run -> "
      "four drivers on Model Inputs + ONE roll-up P&L line, delivered and "
      "assertion-bound.")
