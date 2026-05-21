"""Phase 9 P3.32 — V-4 workbook reconciliation verifier (L-4 tooling).

Opens a generated workbook via Excel COM, forces a full recalc, saves,
and reads back the Checks sheet's "Persisted Baseline Reconciliation"
rows. Each row compares one FINMO line item (formula-rebuilt from
schedule sheets) against the corresponding Audit Source value
(hardcoded from `finmo_json`).

V-4 verdict:
  PASS if every baseline row's |delta| <= ABS_TOLERANCE
       OR |delta| / max(|actual|, |expected|, 1.0) <= REL_TOLERANCE
  FAIL otherwise.

Doctrine: ABS_TOLERANCE = $50, REL_TOLERANCE = 0.01% per P3.32
directive. These match the directive's V-4 thresholds.

Usage:
  python "Test Files/v4_workbook_verifier.py" --workbook <path.xlsx>
  python "Test Files/v4_workbook_verifier.py" --dir <path/to/dir/>
  python "Test Files/v4_workbook_verifier.py" --csv <path.csv> [--dir <dir>]
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Optional


ABS_TOLERANCE = 50.0
REL_TOLERANCE = 0.0001  # 0.01%

CHECKS_SHEET = "Checks"
BASELINE_CATEGORY = "Persisted Baseline"

ACTUAL_COL = 5       # column E
EXPECTED_COL = 6     # column F
DIFFERENCE_COL = 7   # column G
STATUS_COL = 9       # column I
LINE_ITEM_COL = 2    # column B
CATEGORY_COL = 1     # column A


@dataclass
class BaselineRow:
  line_item: str
  actual: Optional[float]
  expected: Optional[float]
  difference: Optional[float]
  status: Optional[str]
  abs_delta: Optional[float]
  rel_delta: Optional[float]
  within_tolerance: bool


@dataclass
class V4Result:
  workbook_path: str
  recalc_status: str
  recalc_error: Optional[str]
  baseline_rows: list[BaselineRow]
  max_abs_delta: float
  max_rel_delta: float
  v4_pass: bool
  reason: str

  def to_dict(self) -> dict[str, Any]:
    return {
      "workbook_path": self.workbook_path,
      "recalc_status": self.recalc_status,
      "recalc_error": self.recalc_error,
      "baseline_rows": [asdict(r) for r in self.baseline_rows],
      "max_abs_delta": self.max_abs_delta,
      "max_rel_delta": self.max_rel_delta,
      "v4_pass": self.v4_pass,
      "reason": self.reason,
    }


def _recalc_via_excel_com(workbook_path: str) -> Optional[str]:
  """Open in Excel, CalculateFull(), Save. Returns None on success
  or a short error description on failure."""
  try:
    import win32com.client as _w32  # type: ignore
  except Exception as exc:
    return f"pywin32_unavailable: {type(exc).__name__}: {str(exc)[:200]}"
  excel = None
  wb = None
  try:
    excel = _w32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb = excel.Workbooks.Open(str(workbook_path))
    if wb is None:
      return "excel_workbook_open_returned_none"
    excel.CalculateFull()
    wb.Save()
    return None
  except Exception as exc:
    return f"excel_com_failure: {type(exc).__name__}: {str(exc)[:200]}"
  finally:
    try:
      if wb is not None:
        wb.Close(SaveChanges=False)
    except Exception:
      pass
    try:
      if excel is not None:
        excel.Quit()
    except Exception:
      pass


def _coerce_float(value: Any) -> Optional[float]:
  if value is None:
    return None
  try:
    if isinstance(value, bool):
      return float(value)
    if isinstance(value, (int, float)):
      return float(value)
    text = str(value).strip()
    if not text or text.startswith("="):
      return None
    return float(text.replace(",", "").replace("$", ""))
  except Exception:
    return None


def _within_tolerance(actual: Optional[float], expected: Optional[float], delta: Optional[float]) -> tuple[bool, float, float]:
  """Returns (within_tolerance, abs_delta, rel_delta)."""
  if delta is None:
    if actual is None or expected is None:
      return False, float("inf"), float("inf")
    delta = actual - expected
  abs_d = abs(delta)
  if actual is None and expected is None:
    return False, abs_d, float("inf")
  scale = max(abs(actual or 0.0), abs(expected or 0.0), 1.0)
  rel_d = abs_d / scale
  ok = (abs_d <= ABS_TOLERANCE) or (rel_d <= REL_TOLERANCE)
  return ok, abs_d, rel_d


def _read_baseline_rows(workbook_path: str) -> tuple[list[BaselineRow], Optional[str]]:
  try:
    import openpyxl  # type: ignore
  except Exception as exc:
    return [], f"openpyxl_unavailable: {exc}"
  try:
    wb = openpyxl.load_workbook(str(workbook_path), data_only=True)
  except Exception as exc:
    return [], f"openpyxl_load_failed: {type(exc).__name__}: {str(exc)[:200]}"
  try:
    if CHECKS_SHEET not in wb.sheetnames:
      return [], f"checks_sheet_missing"
    ws = wb[CHECKS_SHEET]
    rows: list[BaselineRow] = []
    for r in range(1, ws.max_row + 1):
      category = ws.cell(row=r, column=CATEGORY_COL).value
      if category != BASELINE_CATEGORY:
        continue
      line_item = str(ws.cell(row=r, column=LINE_ITEM_COL).value or "").strip()
      actual = _coerce_float(ws.cell(row=r, column=ACTUAL_COL).value)
      expected = _coerce_float(ws.cell(row=r, column=EXPECTED_COL).value)
      difference = _coerce_float(ws.cell(row=r, column=DIFFERENCE_COL).value)
      status = ws.cell(row=r, column=STATUS_COL).value
      ok, abs_d, rel_d = _within_tolerance(actual, expected, difference)
      rows.append(BaselineRow(
        line_item=line_item,
        actual=actual,
        expected=expected,
        difference=difference,
        status=str(status) if status is not None else None,
        abs_delta=abs_d if abs_d != float("inf") else None,
        rel_delta=rel_d if rel_d != float("inf") else None,
        within_tolerance=ok,
      ))
    return rows, None
  finally:
    try:
      wb.close()
    except Exception:
      pass


def verify_workbook(workbook_path: str) -> V4Result:
  path = Path(workbook_path).resolve()
  if not path.exists():
    return V4Result(
      workbook_path=str(path),
      recalc_status="missing",
      recalc_error="workbook_file_missing",
      baseline_rows=[],
      max_abs_delta=float("inf"),
      max_rel_delta=float("inf"),
      v4_pass=False,
      reason="workbook_not_found",
    )
  recalc_error = _recalc_via_excel_com(str(path))
  recalc_status = "ok" if recalc_error is None else "failed"
  if recalc_error is not None:
    return V4Result(
      workbook_path=str(path),
      recalc_status=recalc_status,
      recalc_error=recalc_error,
      baseline_rows=[],
      max_abs_delta=float("inf"),
      max_rel_delta=float("inf"),
      v4_pass=False,
      reason=f"recalc_failed: {recalc_error}",
    )
  rows, read_error = _read_baseline_rows(str(path))
  if read_error is not None:
    return V4Result(
      workbook_path=str(path),
      recalc_status=recalc_status,
      recalc_error=read_error,
      baseline_rows=[],
      max_abs_delta=float("inf"),
      max_rel_delta=float("inf"),
      v4_pass=False,
      reason=f"baseline_read_failed: {read_error}",
    )
  if not rows:
    return V4Result(
      workbook_path=str(path),
      recalc_status=recalc_status,
      recalc_error=None,
      baseline_rows=[],
      max_abs_delta=float("inf"),
      max_rel_delta=float("inf"),
      v4_pass=False,
      reason="no_baseline_rows_found",
    )
  max_abs = max((r.abs_delta or 0.0) for r in rows)
  max_rel = max((r.rel_delta or 0.0) for r in rows)
  all_ok = all(r.within_tolerance for r in rows)
  reason = "all_baseline_rows_within_tolerance" if all_ok else (
    "baseline_rows_exceed_tolerance: " +
    ", ".join(f"{r.line_item}=delta_abs={r.abs_delta:.2f}/rel={r.rel_delta:.6f}" for r in rows if not r.within_tolerance)
  )
  return V4Result(
    workbook_path=str(path),
    recalc_status=recalc_status,
    recalc_error=None,
    baseline_rows=rows,
    max_abs_delta=max_abs,
    max_rel_delta=max_rel,
    v4_pass=all_ok,
    reason=reason,
  )


def verify_directory(dir_path: str) -> list[V4Result]:
  d = Path(dir_path)
  results: list[V4Result] = []
  for f in sorted(d.glob("*.xlsx")):
    if f.name.startswith("~$"):
      continue
    results.append(verify_workbook(str(f)))
  return results


def _format_result_short(r: V4Result) -> str:
  name = Path(r.workbook_path).name
  if not r.v4_pass:
    return f"FAIL  {name}  max_abs=${r.max_abs_delta:,.2f}  max_rel={r.max_rel_delta:.6f}  reason={r.reason}"
  return f"PASS  {name}  max_abs=${r.max_abs_delta:,.2f}  max_rel={r.max_rel_delta:.6f}"


def main(argv: Optional[list[str]] = None) -> int:
  parser = argparse.ArgumentParser(description="Phase 9 P3.32 V-4 workbook reconciliation verifier.")
  parser.add_argument("--workbook", default="", help="Path to a single .xlsx to verify.")
  parser.add_argument("--dir", default="", help="Directory of .xlsx workbooks to verify.")
  parser.add_argument("--json-out", default="", help="Optional path to write structured results JSON.")
  parser.add_argument("--csv-out", default="", help="Optional path to write a summary CSV.")
  args = parser.parse_args(argv)

  if not args.workbook and not args.dir:
    print("ERROR: --workbook or --dir is required.", file=sys.stderr)
    return 2

  results: list[V4Result] = []
  if args.workbook:
    results.append(verify_workbook(args.workbook))
  if args.dir:
    results.extend(verify_directory(args.dir))

  for r in results:
    print(_format_result_short(r))

  total = len(results)
  passes = sum(1 for r in results if r.v4_pass)
  print(f"\nV-4 summary: {passes}/{total} pass")

  if args.json_out:
    Path(args.json_out).parent.mkdir(parents=True, exist_ok=True)
    Path(args.json_out).write_text(json.dumps([r.to_dict() for r in results], indent=2), encoding="utf-8")
    print(f"JSON written: {args.json_out}")

  if args.csv_out:
    Path(args.csv_out).parent.mkdir(parents=True, exist_ok=True)
    with open(args.csv_out, "w", newline="", encoding="utf-8") as fh:
      writer = csv.writer(fh)
      writer.writerow([
        "workbook_name", "v4_pass", "recalc_status", "max_abs_delta", "max_rel_delta",
        "row_count", "first_failing_row", "first_failing_abs_delta", "reason",
      ])
      for r in results:
        first_fail = next((row for row in r.baseline_rows if not row.within_tolerance), None)
        writer.writerow([
          Path(r.workbook_path).name,
          "PASS" if r.v4_pass else "FAIL",
          r.recalc_status,
          f"{r.max_abs_delta:.4f}",
          f"{r.max_rel_delta:.8f}",
          len(r.baseline_rows),
          first_fail.line_item if first_fail else "",
          f"{first_fail.abs_delta:.4f}" if first_fail and first_fail.abs_delta is not None else "",
          r.reason,
        ])
    print(f"CSV written: {args.csv_out}")

  return 0 if all(r.v4_pass for r in results) else 1


if __name__ == "__main__":
  raise SystemExit(main())
