"""Intake-bypass runner: exercise the post-intake pipeline straight from Excel.

Reads scenarios from an Excel workbook (one sheet per scenario; column A is a
field name, column B is its value), builds an intake-complete draft by overlaying
the sheet's overrides onto a captured baseline snapshot, writes it directly to
the intake SQL tables, then triggers the post-intake pipeline -- the same
/api/intake-consult/system-run path real intake uses on completion.

This skips the GPT intake conversation entirely. It does NOT replace intake for
real users; it is a test harness for the post-intake architecture.

Prerequisites:
  - The Flask API must be running (e.g. `python context/run_api_5050_single.py`).
  - A baseline snapshot must exist (see capture_intake_baseline.py).
  - .env must have MYSQL_* configured.

Usage:
  python "Test Files/run_intake_bypass.py"                       # run every scenario sheet
  python "Test Files/run_intake_bypass.py" --scenario Sunny_Glaze_Donuts
  python "Test Files/run_intake_bypass.py" --list                # list scenario sheets
  python "Test Files/run_intake_bypass.py" --dry-run             # build + write SQL, skip system-run
"""

from __future__ import annotations

import argparse
import copy
import importlib.util
import json
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


THIS_DIR = Path(__file__).resolve().parent


def _load_module(filename: str, mod_name: str):
  spec = importlib.util.spec_from_file_location(mod_name, str(THIS_DIR / filename))
  if spec is None or spec.loader is None:
    raise RuntimeError(f"Unable to load {filename}")
  module = importlib.util.module_from_spec(spec)
  spec.loader.exec_module(module)
  return module


C = _load_module("intake_bypass_common.py", "intake_bypass_common")
_DUAL = _load_module("run_dual_agent_intake.py", "run_dual_agent_intake_bypass")


def _string(value: Any) -> str:
  return str(value if value is not None else "").strip()


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------
def _read_scenarios(excel_path: Path) -> "Dict[str, List[Tuple[str, Any]]]":
  import openpyxl  # type: ignore

  if not excel_path.exists():
    raise RuntimeError(f"Scenario workbook not found: {excel_path}")
  wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
  scenarios: "Dict[str, List[Tuple[str, Any]]]" = {}
  try:
    for sheet_name in wb.sheetnames:
      if sheet_name.startswith("_"):
        continue  # documentation / non-scenario sheets
      ws = wb[sheet_name]
      rows: List[Tuple[str, Any]] = []
      for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        field = _string(row[0]) if row and len(row) >= 1 else ""
        value = row[1] if row and len(row) >= 2 else None
        if not field or field.startswith("#"):
          continue
        if field.lower() in {"field", "question", "key"}:
          continue  # header row
        rows.append((field, value))
      if rows:
        scenarios[sheet_name] = rows
  finally:
    wb.close()
  return scenarios


def _scenario_to_dict(rows: List[Tuple[str, Any]]) -> Dict[str, Any]:
  out: Dict[str, Any] = {}
  for field, value in rows:
    out[field] = value
  return out


# ---------------------------------------------------------------------------
# Draft assembly + SQL write
# ---------------------------------------------------------------------------
def _intake_complete_planning_run_payload() -> Dict[str, Any]:
  return {
    "contract_version": "planning_run_v1",
    "stage": "intake_complete",
    "status": "pending",
    "gpt_narrative": "Intake complete (intake-bypass runner). Ready for backend planning.",
  }


def _bypass_messages_json(scenario_name: str, baseline_name: str) -> List[Dict[str, Any]]:
  return [
    {
      "role": "system",
      "content": (
        f"Draft synthesized by run_intake_bypass.py for scenario {scenario_name!r} "
        f"from baseline {baseline_name!r}. No GPT intake conversation occurred."
      ),
    }
  ]


def _write_draft_state(
  conn,
  *,
  target_draft_id: str,
  flat: Dict[str, Any],
  structured: Dict[str, Any],
  messages_json: List[Dict[str, Any]],
) -> None:
  now = _DUAL._eastern_now().strftime("%Y-%m-%d %H:%M:%S.%f")
  cur = conn.cursor()
  try:
    cur.execute(
      """
      UPDATE intake_consult_drafts
      SET
        status = %s,
        active_focus = %s,
        ops_confirmed = 1,
        market_confirmed = 1,
        people_confirmed = 1,
        financials_confirmed = 1,
        business_name = %s,
        business_address = %s,
        address_street = %s,
        address_city = %s,
        address_state = %s,
        address_zip = %s,
        address_country = %s,
        business_start_date = %s,
        messages_json = %s,
        operating_model_json = %s,
        target_market_json = %s,
        people_json = %s,
        financials_json = %s,
        marketing_model_json = %s,
        financials_year1_json = %s,
        realism_memo_json = %s,
        model_input_json = NULL,
        finmo_json = NULL,
        planning_run_json = %s,
        pending_ops_milestone_json = %s,
        fulfillment_json = %s,
        ops_finalize_proposed = 1,
        market_finalize_proposed = 1,
        people_finalize_proposed = 1,
        financials_finalize_proposed = 1,
        completed_at = %s,
        submitted_at = NULL,
        intake_submission_id = NULL,
        updated_at = %s
      WHERE draft_id = %s
      """,
      (
        "completed",
        "done",
        _string(flat.get("business_name")) or None,
        _string(flat.get("business_address")) or None,
        _string(flat.get("address_street")) or None,
        _string(flat.get("address_city")) or None,
        _string(flat.get("address_state")) or None,
        _string(flat.get("address_zip")) or None,
        _string(flat.get("address_country")) or None,
        _string(flat.get("business_start_date")) or None,
        C.sql_json_value(messages_json),
        C.sql_json_value(structured.get("operating_model_json")),
        C.sql_json_value(structured.get("target_market_json")),
        C.sql_json_value(structured.get("people_json")),
        C.sql_json_value(structured.get("financials_json")),
        C.sql_json_value(structured.get("marketing_model_json")),
        C.sql_json_value(structured.get("financials_year1_json")),
        C.sql_json_value(structured.get("realism_memo_json")),
        C.sql_json_value(_intake_complete_planning_run_payload()),
        C.sql_json_value(structured.get("pending_ops_milestone_json")),
        C.sql_json_value(structured.get("fulfillment_json")),
        now,
        now,
        _string(target_draft_id),
      ),
    )
    conn.commit()
  finally:
    try:
      cur.close()
    except Exception:
      pass


def _mint_draft_offline(conn) -> "Tuple[str, str]":
  """Insert a fresh in_progress draft directly (mirrors intake create_draft).

  Used by --dry-run so the runner can populate the intake tables without the
  API/GPT being up. client_id is CHAR(20); generate a unique 20-char id.
  """
  draft_id = uuid.uuid4().hex
  client_id = ("bp" + uuid.uuid4().hex)[:20]
  now = _DUAL._eastern_now().strftime("%Y-%m-%d %H:%M:%S.%f")
  cur = conn.cursor()
  try:
    cur.execute(
      """
      INSERT INTO intake_consult_drafts
        (draft_id, client_id, status, messages_json, created_at, updated_at)
      VALUES (%s, %s, 'in_progress', %s, %s, %s)
      """,
      (draft_id, client_id, "[]", now, now),
    )
    conn.commit()
  finally:
    try:
      cur.close()
    except Exception:
      pass
  return draft_id, client_id


def _verify_draft(conn, draft_id: str) -> Dict[str, Any]:
  cur = conn.cursor(dictionary=True)
  try:
    cur.execute(
      """
      SELECT draft_id, business_name, status, active_focus,
             (operating_model_json IS NOT NULL) AS has_ops,
             (target_market_json IS NOT NULL) AS has_mkt,
             (people_json IS NOT NULL) AS has_ppl,
             (financials_json IS NOT NULL) AS has_fin,
             (financials_year1_json IS NOT NULL) AS has_fy1
      FROM intake_consult_drafts WHERE draft_id=%s
      """,
      (draft_id,),
    )
    return cur.fetchone() or {}
  finally:
    cur.close()


# ---------------------------------------------------------------------------
# Per-scenario execution
# ---------------------------------------------------------------------------
def _run_scenario(
  *,
  conn,
  base_url: str,
  baselines_dir: Path,
  scenario_name: str,
  overrides: Dict[str, Any],
  dry_run: bool,
) -> Dict[str, Any]:
  baseline_name = _string(overrides.get("baseline"))
  baseline = C.load_baseline(baselines_dir, baseline_name)

  flat = copy.deepcopy(baseline.get("flat") or {})
  structured = copy.deepcopy(baseline.get("structured") or {})

  audit = C.apply_overrides(flat=flat, structured=structured, overrides=overrides)

  print(f"\n=== Scenario: {scenario_name} (baseline={baseline_name}) ===")
  if audit:
    print(f"  Applied {len(audit)} override(s) vs baseline:")
    for entry in audit:
      print(f"    - {entry['field']}: {entry['old']!r} -> {entry['new']!r}")
  else:
    print("  No edits vs baseline (faithful reproduction).")

  # 1. Mint a fresh draft. Dry-run inserts directly (fully offline); a real run
  #    uses the session endpoint so the row is initialized exactly as intake does.
  if dry_run:
    draft_id, client_id = _mint_draft_offline(conn)
  else:
    session = _DUAL._post_json(f"{base_url}/api/intake-consult/session", {})
    draft_id = _string(session.get("draft_id"))
    client_id = _string(session.get("client_id"))
    if not draft_id or not client_id:
      raise RuntimeError(f"Failed to create fresh intake session: {session}")

  # 2. Write the intake-complete state directly to SQL.
  messages_json = _bypass_messages_json(scenario_name, baseline_name)
  _write_draft_state(
    conn,
    target_draft_id=draft_id,
    flat=flat,
    structured=structured,
    messages_json=messages_json,
  )
  verify = _verify_draft(conn, draft_id)
  ok = (
    _string(verify.get("active_focus")).lower() == "done"
    and bool(verify.get("has_ops")) and bool(verify.get("has_fin"))
    and bool(verify.get("has_ppl")) and bool(verify.get("has_mkt"))
  )
  print(f"  Draft {draft_id} written. active_focus={verify.get('active_focus')} "
        f"ops/mkt/ppl/fin/fy1={verify.get('has_ops')}{verify.get('has_mkt')}{verify.get('has_ppl')}{verify.get('has_fin')}{verify.get('has_fy1')}")
  if not ok:
    raise RuntimeError(f"Draft {draft_id} did not populate the intake tables as expected: {verify}")

  result: Dict[str, Any] = {
    "scenario": scenario_name,
    "baseline": baseline_name,
    "draft_id": draft_id,
    "client_id": client_id,
    "overrides_applied": audit,
    "sql_populated": True,
  }

  if dry_run:
    print("  --dry-run: skipping system-run trigger.")
    result["status"] = "dry_run_sql_ok"
    return result

  # 3. Trigger the post-intake pipeline (identical to real intake completion).
  trace_name = f"intake_bypass -- {scenario_name}"
  trace_headers = {"X-Planning-Trace-Run-Name": trace_name, "X-Planning-Trace-Reset": "1"}
  print("  Triggering post-intake system-run...")
  started = time.perf_counter()
  response = _DUAL._post_json(
    f"{base_url}/api/intake-consult/system-run",
    {"draft_id": draft_id, "client_id": client_id},
    timeout=None,
    headers=trace_headers,
  )
  elapsed_ms = int(round((time.perf_counter() - started) * 1000.0))
  workbook_path = _string(response.get("client_workbook_path"))
  message = _string(response.get("assistant_message")) or "System run complete."
  print(f"  {message}")
  print(f"  Duration: {elapsed_ms} ms")
  if workbook_path:
    print(f"  Workbook: {workbook_path}")

  draft = _DUAL._get_json(f"{base_url}/api/intake-consult/draft", {"draft_id": draft_id})
  result.update({
    "status": "system_run_complete",
    "duration_ms": elapsed_ms,
    "client_workbook_path": workbook_path,
    "assistant_message": message,
    "final_flags": {
      "ops_confirmed": draft.get("ops_confirmed"),
      "market_confirmed": draft.get("market_confirmed"),
      "people_confirmed": draft.get("people_confirmed"),
      "financials_confirmed": draft.get("financials_confirmed"),
    },
  })
  return result


def _check_api(base_url: str) -> None:
  try:
    _DUAL._post_json(f"{base_url}/api/intake-consult/session", {}, timeout=15)
  except Exception as exc:
    raise RuntimeError(
      f"Cannot reach the API at {base_url} ({exc}). Start it first, e.g.: "
      f"python context/run_api_5050_single.py"
    ) from exc


def main(argv: Optional[list] = None) -> int:
  parser = argparse.ArgumentParser(description="Run post-intake scenarios from an Excel workbook (intake bypass).")
  parser.add_argument("--excel", default=str(C.DEFAULT_SCENARIOS_XLSX))
  parser.add_argument("--baselines-dir", default=str(C.DEFAULT_BASELINES_DIR))
  parser.add_argument("--scenario", default="", help="Run only this sheet (default: all scenario sheets).")
  parser.add_argument("--base-url", default="http://127.0.0.1:5050")
  parser.add_argument("--list", action="store_true", help="List scenario sheets and exit.")
  parser.add_argument("--dry-run", action="store_true", help="Build + write SQL, but do not trigger system-run.")
  args = parser.parse_args(argv)

  C.load_env()
  excel_path = Path(args.excel)
  baselines_dir = Path(args.baselines_dir)
  base_url = _string(args.base_url).rstrip("/")

  scenarios = _read_scenarios(excel_path)
  if not scenarios:
    print(f"No scenario sheets found in {excel_path}.")
    return 1

  if args.list:
    print(f"Scenario sheets in {excel_path}:")
    for name, rows in scenarios.items():
      ov = _scenario_to_dict(rows)
      print(f"  {name}  (baseline={_string(ov.get('baseline')) or '?'}, {len(rows)} rows)")
    return 0

  selected = scenarios
  if _string(args.scenario):
    if args.scenario not in scenarios:
      print(f"Scenario {args.scenario!r} not found. Available: {list(scenarios)}")
      return 1
    selected = {args.scenario: scenarios[args.scenario]}

  if not args.dry_run:
    _check_api(base_url)

  conn = C.mysql_connect()
  results: List[Dict[str, Any]] = []
  failures = 0
  try:
    for name, rows in selected.items():
      overrides = _scenario_to_dict(rows)
      try:
        results.append(_run_scenario(
          conn=conn,
          base_url=base_url,
          baselines_dir=baselines_dir,
          scenario_name=name,
          overrides=overrides,
          dry_run=bool(args.dry_run),
        ))
      except Exception as exc:
        failures += 1
        print(f"  ERROR in scenario {name!r}: {exc}")
        results.append({"scenario": name, "status": "error", "error": str(exc)})
  finally:
    try:
      conn.close()
    except Exception:
      pass

  print("\n=== Summary ===")
  for r in results:
    extra = ""
    if r.get("draft_id"):
      extra = f" draft={r['draft_id']}"
    if r.get("client_workbook_path"):
      extra += f" workbook={r['client_workbook_path']}"
    print(f"  {r['scenario']:28} {r.get('status', '?')}{extra}")
  return 1 if failures else 0


if __name__ == "__main__":
  raise SystemExit(main())
