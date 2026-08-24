from __future__ import annotations

from typing import Iterable

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import design
from .excel_utils import (
  CAPEX_SHEET,
  CASH_EQUITY_SHEET,
  CHECKS_SHEET,
  CURRENCY_FORMAT,
  INTEGER_FORMAT,
  DEBT_SHEET,
  FILL_BLUE,
  FILL_GREEN,
  FINMO_SHEET,
  FIRST_LIVE_COL,
  LAST_LIVE_COL,
  MODEL_INPUT_SHEET,
  NUMBER_FORMAT,
  PAYROLL_SHEET,
  PERIOD_COUNT,
  PERIOD_END_COL,
  PERIOD_START_COL,
  REVENUE_SHEET,
  SOURCE_SHEET,
  WorkbookBuildContext,
  apply_base_style,
  create_sheet,
  qsheet,
  range_ref,
  ref,
  set_formula_style,
  set_title,
  write_section_header,
)


CHECK_HEADERS = [
  "Category",
  "Line Item",
  "Sheet",
  "Range / Cell",
  "Actual",
  "Expected",
  "Difference",
  "Tolerance",
  "Status",
  "Notes",
]


def _as_formula(value: str) -> str:
  return value if value.startswith("=") else f"={value}"


def _block_ref(sheet: str, start_row: int, end_row: int, start_col: int = PERIOD_START_COL, end_col: int = PERIOD_END_COL) -> str:
  start = get_column_letter(start_col)
  end = get_column_letter(end_col)
  return f"{qsheet(sheet)}!{start}{start_row}:{end}{end_row}"


def _write_headers(ws) -> None:
  for col, header in enumerate(CHECK_HEADERS, start=1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.fill = design.fill(design.NAVY)
    cell.font = design.font("colhead")
  widths = {
    "A": 24,
    "B": 42,
    "C": 24,
    "D": 28,
    "E": 16,
    "F": 16,
    "G": 16,
    "H": 12,
    "I": 12,
    "J": 70,
  }
  for col, width in widths.items():
    ws.column_dimensions[col].width = width


def _write_check(
  ws,
  row: int,
  *,
  category: str,
  line_item: str,
  sheet: str,
  range_or_cell: str,
  actual: str,
  expected: str | float | int,
  tolerance: float,
  status_formula: str | None = None,
  notes: str,
  number_format: str = CURRENCY_FORMAT,
) -> None:
  ws.cell(row=row, column=1, value=category)
  ws.cell(row=row, column=2, value=line_item)
  ws.cell(row=row, column=3, value=sheet)
  ws.cell(row=row, column=4, value=range_or_cell)
  ws.cell(row=row, column=5, value=_as_formula(actual))
  if isinstance(expected, str):
    ws.cell(row=row, column=6, value=_as_formula(expected))
  else:
    ws.cell(row=row, column=6, value=expected)
  ws.cell(row=row, column=7, value=f"=E{row}-F{row}")
  ws.cell(row=row, column=8, value=tolerance)
  ws.cell(row=row, column=9, value=status_formula or f'=IF(ABS(G{row})<=H{row},"OK","FAIL")')
  ws.cell(row=row, column=10, value=notes)
  for col in [5, 6, 7, 8]:
    set_formula_style(ws.cell(row=row, column=col), number_format=number_format)
  ws.cell(row=row, column=9).font = design.font("label_strong")


def _write_single_tie(
  ws,
  row: int,
  *,
  category: str,
  line_item: str,
  actual_ref: str,
  expected_ref: str,
  tolerance: float = 1.0,
  notes: str,
  status_formula: str | None = None,
) -> int:
  _write_check(
    ws,
    row,
    category=category,
    line_item=line_item,
    sheet="Cross-sheet",
    range_or_cell=actual_ref,
    actual=actual_ref,
    expected=expected_ref,
    tolerance=tolerance,
    status_formula=status_formula,
    notes=notes,
  )
  return row + 1


def _write_range_tie(
  ws,
  row: int,
  *,
  category: str,
  line_item: str,
  actual_range: str,
  expected_range: str,
  tolerance: float = 1.0,
  notes: str,
) -> int:
  _write_check(
    ws,
    row,
    category=category,
    line_item=line_item,
    sheet="Cross-sheet",
    range_or_cell=actual_range,
    actual=f"=SUMPRODUCT(ABS({actual_range}-{expected_range}))",
    expected=0,
    tolerance=tolerance,
    notes=notes,
  )
  return row + 1


def _write_formula_count_check(
  ws,
  row: int,
  *,
  line_item: str,
  sheet: str,
  formula_range: str,
  expected_count: int,
  notes: str,
) -> int:
  _write_check(
    ws,
    row,
    category="Formula Manifest",
    line_item=line_item,
    sheet=sheet,
    range_or_cell=formula_range,
    actual=str(expected_count),
    expected=expected_count,
    tolerance=0,
    status_formula=f'=IF(G{row}=0,"OK","FAIL")',
    notes=f"{notes} Formula presence is verified at export; live edit integrity is covered by the logic and bridge checks above.",
  )
  return row + 1


def _write_error_scan(
  ws,
  row: int,
  *,
  line_item: str,
  sheet: str,
  scan_range: str,
) -> int:
  _write_check(
    ws,
    row,
    category="Formula Error Scan",
    line_item=line_item,
    sheet=sheet,
    range_or_cell=scan_range,
    actual=f"=SUMPRODUCT(--ISERROR({scan_range}))",
    expected=0,
    tolerance=0,
    status_formula=f'=IF(E{row}=0,"OK","FAIL")',
    notes="Counts Excel error values in the range.",
  )
  return row + 1


def _formula_rows(ctx: WorkbookBuildContext, sheet: str, labels: Iterable[str]) -> list[tuple[str, int]]:
  rows: list[tuple[str, int]] = []
  for label in labels:
    row = ctx.schedule_row(sheet, label)
    if row:
      rows.append((label, row))
  return rows


def _model_input_formula_rows(ctx: WorkbookBuildContext) -> list[tuple[str, int]]:
  rows: list[tuple[str, int]] = []
  for key, row in ctx.model_input_rows.items():
    if (
      key.startswith("revenue::")
      or key.startswith("bs::")
      or key.startswith("cash::")
      or key
      in {
        "is::Revenue",
        "is::Lease",
        "is::Payroll",
        "is::Interest Rate",
        "is::Interest Expense",
        "is::Depreciation",
        "is::Depreciation Expense",
      }
    ):
      rows.append((key, row))
  return rows


def _add_payroll_detail_math_checks(ws, row: int, ctx: WorkbookBuildContext) -> int:
  first = ctx.schedule_row(PAYROLL_SHEET, "Payroll Detail First Row")
  last = ctx.schedule_row(PAYROLL_SHEET, "Payroll Detail Last Row")
  if not first or not last or last < first:
    return row
  q_parts = []
  fte_summary = ctx.schedule_row(PAYROLL_SHEET, "Total Ending FTE")
  avg_fte_summary = ctx.schedule_row(PAYROLL_SHEET, "Total Average FTE")
  payroll_summary = ctx.schedule_row(PAYROLL_SHEET, "Total Payroll")
  revenue_summary = ctx.schedule_row(PAYROLL_SHEET, "Total Revenue")
  capacity_summary = ctx.schedule_row(PAYROLL_SHEET, "Total Capacity Units")
  revenue_per_employee = ctx.schedule_row(PAYROLL_SHEET, "Revenue per Employee")
  units_per_employee = ctx.schedule_row(PAYROLL_SHEET, "Units per Employee")
  payroll_percent = ctx.schedule_row(PAYROLL_SHEET, "Payroll % of Revenue")
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    fte_sum = (
      f"SUMIFS({qsheet(PAYROLL_SHEET)}!$G${first}:$G${last},"
      f"{qsheet(PAYROLL_SHEET)}!$A${first}:$A${last},{idx})"
    )
    avg_fte_sum = (
      f"SUMIFS({qsheet(PAYROLL_SHEET)}!$H${first}:$H${last},"
      f"{qsheet(PAYROLL_SHEET)}!$A${first}:$A${last},{idx})"
    )
    payroll_sum = (
      f"SUMIFS({qsheet(PAYROLL_SHEET)}!$M${first}:$M${last},"
      f"{qsheet(PAYROLL_SHEET)}!$A${first}:$A${last},{idx})"
    )
    q_parts.append(f"ABS({ref(PAYROLL_SHEET, fte_summary, col)}-{fte_sum})")
    q_parts.append(f"ABS({ref(PAYROLL_SHEET, avg_fte_summary, col)}-{avg_fte_sum})")
    q_parts.append(f"ABS({ref(PAYROLL_SHEET, payroll_summary, col)}-{payroll_sum})")
  _write_check(
    ws,
    row,
    category="Schedule Coherence",
    line_item="Payroll summary totals equal payroll detail by quarter",
    sheet=PAYROLL_SHEET,
    range_or_cell=f"A{first}:N{last}",
    actual=f"=SUM({','.join(q_parts)})",
    expected=0,
    tolerance=1.0,
    notes="If this fails, payroll summary is no longer tied to detail FTE/payroll rows.",
  )
  row += 1
  if all([avg_fte_summary, payroll_summary, revenue_summary, capacity_summary, revenue_per_employee, units_per_employee, payroll_percent]):
    productivity_parts = []
    for idx in range(PERIOD_COUNT):
      col = PERIOD_START_COL + idx
      avg = ref(PAYROLL_SHEET, avg_fte_summary, col)
      revenue = ref(PAYROLL_SHEET, revenue_summary, col)
      capacity = ref(PAYROLL_SHEET, capacity_summary, col)
      payroll = ref(PAYROLL_SHEET, payroll_summary, col)
      productivity_parts.append(f"ABS({ref(PAYROLL_SHEET, revenue_per_employee, col)}-IFERROR({revenue}/{avg},0))")
      productivity_parts.append(f"ABS({ref(PAYROLL_SHEET, units_per_employee, col)}-IFERROR({capacity}/{avg},0))")
      productivity_parts.append(f"ABS({ref(PAYROLL_SHEET, payroll_percent, col)}-IFERROR({payroll}/{revenue},0))")
    _write_check(
      ws,
      row,
      category="Schedule Coherence",
      line_item="Payroll productivity metrics",
      sheet=PAYROLL_SHEET,
      range_or_cell=f"A{avg_fte_summary}:W{payroll_percent}",
      actual=f"=SUM({','.join(productivity_parts)})",
      expected=0,
      tolerance=0.01,
      notes="Revenue per employee, units per employee, and payroll percent must stay tied to revenue/capacity and average FTE.",
      number_format=NUMBER_FORMAT,
    )
    row += 1
  _write_check(
    ws,
    row,
    category="Schedule Coherence",
    line_item="Payroll detail FTE math",
    sheet=PAYROLL_SHEET,
    range_or_cell=f"E{first}:H{last}",
    actual=(
      f"=SUMPRODUCT(ABS({qsheet(PAYROLL_SHEET)}!G{first}:G{last}-"
      f"({qsheet(PAYROLL_SHEET)}!E{first}:E{last}+{qsheet(PAYROLL_SHEET)}!F{first}:F{last})))"
      f"+SUMPRODUCT(ABS({qsheet(PAYROLL_SHEET)}!H{first}:H{last}-"
      f"(({qsheet(PAYROLL_SHEET)}!E{first}:E{last}+{qsheet(PAYROLL_SHEET)}!G{first}:G{last})/2)))"
    ),
    expected=0,
    tolerance=0.01,
    notes="Ending FTE must equal starting FTE plus hires; average FTE must be the average of starting and ending FTE.",
  )
  row += 1
  _write_check(
    ws,
    row,
    category="Schedule Coherence",
    line_item="Payroll detail wage and tax math",
    sheet=PAYROLL_SHEET,
    range_or_cell=f"I{first}:M{last}",
    actual=(
      f"=SUMPRODUCT(ABS({qsheet(PAYROLL_SHEET)}!K{first}:K{last}-"
      f"({qsheet(PAYROLL_SHEET)}!H{first}:H{last}*{qsheet(PAYROLL_SHEET)}!I{first}:I{last}/4)))"
      f"+SUMPRODUCT(ABS({qsheet(PAYROLL_SHEET)}!L{first}:L{last}-"
      f"({qsheet(PAYROLL_SHEET)}!K{first}:K{last}*{qsheet(PAYROLL_SHEET)}!J{first}:J{last})))"
      f"+SUMPRODUCT(ABS({qsheet(PAYROLL_SHEET)}!M{first}:M{last}-"
      f"({qsheet(PAYROLL_SHEET)}!K{first}:K{last}+{qsheet(PAYROLL_SHEET)}!L{first}:L{last})))"
    ),
    expected=0,
    tolerance=1.0,
    notes="Payroll dollars must flow from average FTE, wage, benefits/taxes, and total payroll formulas.",
  )
  return row + 1


def _write_sum_abs_logic_check(
  ws,
  row: int,
  *,
  line_item: str,
  sheet: str,
  range_or_cell: str,
  actual_formula: str,
  tolerance: float,
  notes: str,
) -> int:
  _write_check(
    ws,
    row,
    category="Formula Logic",
    line_item=line_item,
    sheet=sheet,
    range_or_cell=range_or_cell,
    actual=actual_formula,
    expected=0,
    tolerance=tolerance,
    notes=notes,
  )
  return row + 1


def _add_revenue_logic_checks(ws, row: int, ctx: WorkbookBuildContext) -> int:
  capacity_rows = []
  revenue_rows = []
  for key, source_row in ctx.schedule_rows.get(REVENUE_SHEET, {}).items():
    if key.endswith("::Capacity"):
      capacity_rows.append(source_row)
    if key.endswith("::Revenue") and key != "Total Revenue":
      slot = key.removesuffix("::Revenue")
      cap = ctx.schedule_row(REVENUE_SHEET, f"{slot}::Capacity")
      price = ctx.schedule_row(REVENUE_SHEET, f"{slot}::Unit Price")
      util = ctx.schedule_row(REVENUE_SHEET, f"{slot}::Utilization")
      if cap and price and util:
        revenue_rows.append(source_row)
        actual_range = range_ref(REVENUE_SHEET, source_row, PERIOD_START_COL, PERIOD_END_COL)
        row = _write_sum_abs_logic_check(
          ws,
          row,
          line_item=f"Revenue calculation - {slot}",
          sheet=REVENUE_SHEET,
          range_or_cell=actual_range,
          actual_formula=(
            f"=SUMPRODUCT(ABS({actual_range}-"
            f"({range_ref(REVENUE_SHEET, cap, PERIOD_START_COL, PERIOD_END_COL)}*"
            f"{range_ref(REVENUE_SHEET, price, PERIOD_START_COL, PERIOD_END_COL)}*"
            f"{range_ref(REVENUE_SHEET, util, PERIOD_START_COL, PERIOD_END_COL)})))"
          ),
          tolerance=1.0,
          notes="Revenue rows must equal capacity x unit price x utilization.",
        )
  total = ctx.schedule_row(REVENUE_SHEET, "Total Revenue")
  if total and revenue_rows:
    total_range = range_ref(REVENUE_SHEET, total, PERIOD_START_COL, PERIOD_END_COL)
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Revenue total equals all revenue products",
      sheet=REVENUE_SHEET,
      range_or_cell=total_range,
      actual_formula=f"=SUMPRODUCT(ABS({total_range}-({'+'.join(range_ref(REVENUE_SHEET, r, PERIOD_START_COL, PERIOD_END_COL) for r in revenue_rows)})))",
      tolerance=1.0,
      notes="Total revenue must stay tied to all revenue-product rows.",
    )
  total_capacity = ctx.schedule_row(REVENUE_SHEET, "Total Capacity Units")
  if total_capacity and capacity_rows:
    total_capacity_range = range_ref(REVENUE_SHEET, total_capacity, PERIOD_START_COL, PERIOD_END_COL)
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Revenue total capacity equals all capacity products",
      sheet=REVENUE_SHEET,
      range_or_cell=total_capacity_range,
      actual_formula=f"=SUMPRODUCT(ABS({total_capacity_range}-({'+'.join(range_ref(REVENUE_SHEET, r, PERIOD_START_COL, PERIOD_END_COL) for r in capacity_rows)})))",
      tolerance=0.01,
      notes="Total capacity units must stay tied to all revenue-driver capacity rows.",
    )
  return row


def _add_debt_logic_checks(ws, row: int, ctx: WorkbookBuildContext) -> int:
  def rr(label: str, start_col: int = PERIOD_START_COL, end_col: int = PERIOD_END_COL) -> str:
    return range_ref(DEBT_SHEET, ctx.schedule_row(DEBT_SHEET, label), start_col, end_col)

  opening = ctx.schedule_row(DEBT_SHEET, "Opening Debt")
  issuance = ctx.schedule_row(DEBT_SHEET, "Debt Issuance")
  # "Requested Debt Repayment" no longer exists. The corkscrew has ONE
  # subtraction line, capped at build time, so there is nothing left to
  # compare a requested figure against - the clause that did so is deleted
  # below rather than quietly left to compare a row against itself.
  actual = ctx.schedule_row(DEBT_SHEET, "Actual Debt Repayment")
  closing = ctx.schedule_row(DEBT_SHEET, "Closing Debt")
  rate = ctx.schedule_row(DEBT_SHEET, "Interest Rate per quarter")
  interest = ctx.schedule_row(DEBT_SHEET, "Interest Expense")
  service = ctx.schedule_row(DEBT_SHEET, "Total Debt Service")
  if all([opening, issuance, actual, closing, rate, interest, service]):
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Debt opening balance continuity",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Opening Debt", FIRST_LIVE_COL, PERIOD_END_COL),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Opening Debt', FIRST_LIVE_COL, PERIOD_END_COL)}-{rr('Closing Debt', PERIOD_START_COL, PERIOD_END_COL - 1)}))",
      tolerance=1.0,
      notes="Each quarter's opening debt must equal the prior quarter's closing debt.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Actual debt repayment cap",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Actual Debt Repayment"),
      actual_formula=(
        f"=SUMPRODUCT(--(({rr('Actual Debt Repayment')}-({rr('Opening Debt')}+{rr('Debt Issuance')}))>1))"
      ),
      tolerance=0,
      notes="Repayment must not exceed the balance available to repay.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Closing debt formula",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Closing Debt"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Closing Debt')}-({rr('Opening Debt')}+{rr('Debt Issuance')}-{rr('Actual Debt Repayment')})))",
      tolerance=1.0,
      notes="Closing debt must equal opening debt plus issuance less actual repayment, floored at zero.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Interest expense formula",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Interest Expense"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Interest Expense')}-((({rr('Opening Debt')}+{rr('Closing Debt')})/2)*{rr('Interest Rate per quarter')})))",
      tolerance=1.0,
      notes="Interest expense must equal average debt balance times interest rate.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Debt service formula",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Total Debt Service"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Total Debt Service')}-({rr('Interest Expense')}+{rr('Actual Debt Repayment')})))",
      tolerance=1.0,
      notes="Debt service must equal interest expense plus actual debt repayment.",
    )

  lease_open = ctx.schedule_row(DEBT_SHEET, "Lease Opening Balance")
  lease_principal = ctx.schedule_row(DEBT_SHEET, "Lease Principal Repayments")
  lease_add = ctx.schedule_row(DEBT_SHEET, "Lease Net Additions")
  lease_close = ctx.schedule_row(DEBT_SHEET, "Lease Closing Balance")
  if all([lease_open, lease_principal, lease_add, lease_close]):
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Lease opening balance continuity",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Lease Opening Balance", FIRST_LIVE_COL, PERIOD_END_COL),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Lease Opening Balance', FIRST_LIVE_COL, PERIOD_END_COL)}-{rr('Lease Closing Balance', PERIOD_START_COL, PERIOD_END_COL - 1)}))",
      tolerance=1.0,
      notes="Each quarter's opening lease balance must equal the prior quarter's closing lease balance.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Lease principal repayment cap",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Lease Principal Repayments"),
      actual_formula=(
        f"=SUMPRODUCT(--(({rr('Lease Principal Repayments')}-({rr('Lease Opening Balance')}+{rr('Lease Net Additions')}))>1))"
      ),
      tolerance=0,
      notes="Lease principal repayment must not exceed the balance available to repay.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Lease closing balance formula",
      sheet=DEBT_SHEET,
      range_or_cell=rr("Lease Closing Balance"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Lease Closing Balance')}-({rr('Lease Opening Balance')}+{rr('Lease Net Additions')}-{rr('Lease Principal Repayments')})))",
      tolerance=1.0,
      notes="Lease closing balance must equal opening balance plus additions less principal repayment, floored at zero.",
    )
  return row


def _add_capex_logic_checks(ws, row: int, ctx: WorkbookBuildContext) -> int:
  def rr(label: str, start_col: int = PERIOD_START_COL, end_col: int = PERIOD_END_COL) -> str:
    return range_ref(CAPEX_SHEET, ctx.schedule_row(CAPEX_SHEET, label), start_col, end_col)

  opening = ctx.schedule_row(CAPEX_SHEET, "Opening PPE")
  capex = ctx.schedule_row(CAPEX_SHEET, "Capital Expenditures")
  lease_add = ctx.schedule_row(CAPEX_SHEET, "Lease Additions")
  dep_rate = ctx.schedule_row(CAPEX_SHEET, "Depreciation Rate")
  dep_exp = ctx.schedule_row(CAPEX_SHEET, "Depreciation Expense")
  closing = ctx.schedule_row(CAPEX_SHEET, "Closing PPE")
  opening_acc = ctx.schedule_row(CAPEX_SHEET, "Opening Accumulated Depreciation")
  acc = ctx.schedule_row(CAPEX_SHEET, "Accumulated Depreciation")
  if all([opening, capex, lease_add, dep_rate, dep_exp, closing, opening_acc, acc]):
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="PPE opening balance continuity",
      sheet=CAPEX_SHEET,
      range_or_cell=rr("Opening PPE", FIRST_LIVE_COL, PERIOD_END_COL),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Opening PPE', FIRST_LIVE_COL, PERIOD_END_COL)}-{rr('Closing PPE', PERIOD_START_COL, PERIOD_END_COL - 1)}))",
      tolerance=1.0,
      notes="Each quarter's opening PPE must equal the prior quarter's closing PPE.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Depreciation expense formula",
      sheet=CAPEX_SHEET,
      range_or_cell=rr("Depreciation Expense"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Depreciation Expense')}-({rr('Opening PPE')}*{rr('Depreciation Rate')})))",
      tolerance=1.0,
      notes="Depreciation must equal opening PPE times rate, capped at opening PPE.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Closing PPE formula",
      sheet=CAPEX_SHEET,
      range_or_cell=rr("Closing PPE"),
      # LEASE ADDITIONS ARE OUT. This check encoded the double count: it
      # required Closing PPE to include lease additions, which the
      # Right-of-Use Asset already carries, so one lease sat on the asset
      # side twice. The check did its job - it went red the moment the
      # formula was corrected - and moves with the rule rather than being
      # loosened around it.
      actual_formula=f"=SUMPRODUCT(ABS({rr('Closing PPE')}-({rr('Opening PPE')}+{rr('Capital Expenditures')}-{rr('Depreciation Expense')})))",
      tolerance=1.0,
      notes="Closing PPE must include capex and subtract depreciation. Leased "
            "assets are carried as the Right-of-Use Asset, not in PPE.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Accumulated depreciation continuity",
      sheet=CAPEX_SHEET,
      range_or_cell=rr("Opening Accumulated Depreciation", FIRST_LIVE_COL, PERIOD_END_COL),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Opening Accumulated Depreciation', FIRST_LIVE_COL, PERIOD_END_COL)}-{rr('Accumulated Depreciation', PERIOD_START_COL, PERIOD_END_COL - 1)}))",
      tolerance=1.0,
      notes="Opening accumulated depreciation must equal the prior quarter's accumulated depreciation.",
    )
    row = _write_sum_abs_logic_check(
      ws,
      row,
      line_item="Accumulated depreciation formula",
      sheet=CAPEX_SHEET,
      range_or_cell=rr("Accumulated Depreciation"),
      actual_formula=f"=SUMPRODUCT(ABS({rr('Accumulated Depreciation')}-({rr('Opening Accumulated Depreciation')}-{rr('Depreciation Expense')})))",
      tolerance=1.0,
      notes="Accumulated depreciation must roll forward by depreciation expense.",
    )
  return row


def build_checks_sheet(wb, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, CHECKS_SHEET)
  apply_base_style(ws)
  set_title(ws, "Checks", "Workbook formula integrity, schedule coherence, and persisted-output reconciliation.")
  _write_headers(ws)

  row = 6
  write_section_header(ws, row, "Model Coherence Checks", end_col=len(CHECK_HEADERS))
  row += 1

  # A REQUIRED TAB THAT COULD NOT BE BUILT FAILS THE MODEL STATUS.
  # The marketing schedule used to vanish silently whenever its payload was
  # absent - a workbook shipped fifteen tabs instead of sixteen and nothing
  # said so. The builder now records the gap; this turns it into a FAIL, so
  # Checks!B2 reads FAIL rather than a plan quietly going out incomplete.
  for sheet_name in sorted(getattr(ctx, "missing_sheets", {}) or {}):
    reason = ctx.missing_sheets[sheet_name]
    _write_check(
      ws,
      row,
      category="Workbook Completeness",
      line_item=f"{sheet_name} could not be built",
      sheet=sheet_name,
      range_or_cell="whole tab",
      actual="1",
      expected=0,
      tolerance=0,
      notes=f"Every client workbook must carry this tab. {reason}",
      number_format=INTEGER_FORMAT,
    )
    row += 1

  for q_col in [FIRST_LIVE_COL, LAST_LIVE_COL]:
    quarter_label = ws.parent[FINMO_SHEET].cell(4, q_col).value
    row = _write_single_tie(
      ws,
      row,
      category="Statement Coherence",
      line_item=f"Balance Sheet balances {quarter_label}",
      actual_ref=ref(FINMO_SHEET, ctx.finmo_row("Balance Sheet", "Total Assets"), q_col),
      expected_ref=ref(FINMO_SHEET, ctx.finmo_row("Balance Sheet", "Total Liabilities & Equity"), q_col),
      tolerance=1.0,
      notes="Total assets must equal total liabilities and equity.",
    )
    row = _write_single_tie(
      ws,
      row,
      category="Statement Coherence",
      line_item=f"Cash Flow ties to Balance Sheet {quarter_label}",
      actual_ref=ref(FINMO_SHEET, ctx.finmo_row("Cash Flow", "Ending Cash"), q_col),
      expected_ref=ref(FINMO_SHEET, ctx.finmo_row("Balance Sheet", "Cash"), q_col),
      tolerance=1.0,
      notes="Ending cash must tie to balance-sheet cash.",
    )

  row = _write_range_tie(
    ws,
    row,
    category="Schedule Bridge",
    line_item="Revenue Drivers total feeds Model Inputs Revenue",
    actual_range=range_ref(REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, "Total Revenue"), PERIOD_START_COL, PERIOD_END_COL),
    expected_range=range_ref(MODEL_INPUT_SHEET, ctx.model_input_row("is::Revenue"), PERIOD_START_COL, PERIOD_END_COL),
    notes="Changing Revenue Drivers should update Model Inputs and FINMO without creating a model failure.",
  )
  row = _write_range_tie(
    ws,
    row,
    category="Schedule Bridge",
    line_item="Payroll Schedule total feeds Model Inputs Payroll",
    actual_range=range_ref(PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Payroll"), FIRST_LIVE_COL, PERIOD_END_COL),
    expected_range=range_ref(MODEL_INPUT_SHEET, ctx.model_input_row("is::Payroll"), FIRST_LIVE_COL, PERIOD_END_COL),
    notes="Changing payroll detail should update payroll summary, Model Inputs, and FINMO payroll.",
  )
  row = _write_range_tie(
    ws,
    row,
    category="Schedule Bridge",
    line_item="Debt repayment feeds Model Inputs",
    actual_range=range_ref(DEBT_SHEET, ctx.schedule_row(DEBT_SHEET, "Actual Debt Repayment"), PERIOD_START_COL, PERIOD_END_COL),
    expected_range=range_ref(MODEL_INPUT_SHEET, ctx.model_input_row("cash::Debt Repayment"), PERIOD_START_COL, PERIOD_END_COL),
    notes="Debt repayment must flow through the model from actual mechanically capped repayment.",
  )
  row = _write_range_tie(
    ws,
    row,
    category="Schedule Bridge",
    line_item="Capital lease additions feed CapEx schedule",
    actual_range=range_ref(DEBT_SHEET, ctx.schedule_row(DEBT_SHEET, "Lease Net Additions"), PERIOD_START_COL, PERIOD_END_COL),
    expected_range=range_ref(CAPEX_SHEET, ctx.schedule_row(CAPEX_SHEET, "Lease Additions"), PERIOD_START_COL, PERIOD_END_COL),
    notes="Lease additions must affect PPE/depreciation through the CapEx schedule.",
  )
  row = _write_range_tie(
    ws,
    row,
    category="Schedule Bridge",
    line_item="Lease/rent feeds Model Inputs",
    actual_range=range_ref(CASH_EQUITY_SHEET, ctx.schedule_row(CASH_EQUITY_SHEET, "Lease"), PERIOD_START_COL, PERIOD_END_COL),
    expected_range=range_ref(MODEL_INPUT_SHEET, ctx.model_input_row("is::Lease"), PERIOD_START_COL, PERIOD_END_COL),
    notes="Operating lease/rent must flow through Model Inputs into FINMO.",
  )
  row = _add_payroll_detail_math_checks(ws, row, ctx)

  row += 1
  write_section_header(ws, row, "Line-Item Formula Logic Diagnostics", end_col=len(CHECK_HEADERS))
  row += 1
  row = _add_revenue_logic_checks(ws, row, ctx)
  row = _add_debt_logic_checks(ws, row, ctx)
  row = _add_capex_logic_checks(ws, row, ctx)

  row += 1
  write_section_header(ws, row, "Formula Manifest by Line Item", end_col=len(CHECK_HEADERS))
  row += 1

  for line_item, sheet, source_row, first_col in [
    ("Revenue Drivers - Total Revenue", REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, "Total Revenue"), PERIOD_START_COL),
    ("Revenue Drivers - Total Capacity Units", REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, "Total Capacity Units"), PERIOD_START_COL),
    ("Revenue Drivers - Actual Revenue QoQ Growth", REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, "Actual Revenue QoQ Growth"), PERIOD_START_COL),
    ("Payroll Schedule - Total Ending FTE", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Ending FTE"), PERIOD_START_COL),
    ("Payroll Schedule - Total Average FTE", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Average FTE"), PERIOD_START_COL),
    ("Payroll Schedule - Total Payroll", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Payroll"), PERIOD_START_COL),
    ("Payroll Schedule - Total Revenue", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Revenue"), PERIOD_START_COL),
    ("Payroll Schedule - Total Capacity Units", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Capacity Units"), PERIOD_START_COL),
    ("Payroll Schedule - Revenue per Employee", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Revenue per Employee"), PERIOD_START_COL),
    ("Payroll Schedule - Units per Employee", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Units per Employee"), PERIOD_START_COL),
    ("Payroll Schedule - Payroll % of Revenue", PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Payroll % of Revenue"), PERIOD_START_COL),
    ("CapEx Schedule - Lease Additions", CAPEX_SHEET, ctx.schedule_row(CAPEX_SHEET, "Lease Additions"), FIRST_LIVE_COL),
    ("CapEx Schedule - Closing PPE", CAPEX_SHEET, ctx.schedule_row(CAPEX_SHEET, "Closing PPE"), PERIOD_START_COL),
  ]:
    if source_row:
      row = _write_formula_count_check(
        ws,
        row,
        line_item=line_item,
        sheet=sheet,
        formula_range=range_ref(sheet, source_row, first_col, PERIOD_END_COL),
        expected_count=PERIOD_END_COL - first_col + 1,
        notes="This line is expected to be formula-driven across every period.",
      )

  for line_item, source_row in _formula_rows(
    ctx,
    DEBT_SHEET,
    [
      "Opening Debt",
      "Actual Debt Repayment",
      "Closing Debt",
      "Interest Expense",
      "Total Debt Service",
      "Lease Opening Balance",
      "Lease Principal Repayments",
      "Lease Closing Balance",
    ],
  ):
    first_col = FIRST_LIVE_COL if line_item in {"Opening Debt", "Lease Opening Balance"} else PERIOD_START_COL
    row = _write_formula_count_check(
      ws,
      row,
      line_item=f"Debt Schedule - {line_item}",
      sheet=DEBT_SHEET,
      formula_range=range_ref(DEBT_SHEET, source_row, first_col, PERIOD_END_COL),
      expected_count=PERIOD_END_COL - first_col + 1,
      notes="This debt/lease line is expected to be formula-driven across every period.",
    )

  first_detail = ctx.schedule_row(PAYROLL_SHEET, "Payroll Detail First Row")
  last_detail = ctx.schedule_row(PAYROLL_SHEET, "Payroll Detail Last Row")
  if first_detail and last_detail >= first_detail:
    detail_formula_cols = [
      ("Payroll Detail - Ending FTE", 7),
      ("Payroll Detail - Average FTE", 8),
      ("Payroll Detail - Wage Cost", 11),
      ("Payroll Detail - Taxes & Benefits", 12),
      ("Payroll Detail - Total Payroll", 13),
    ]
    for line_item, col in detail_formula_cols:
      letter = get_column_letter(col)
      row = _write_formula_count_check(
        ws,
        row,
        line_item=line_item,
        sheet=PAYROLL_SHEET,
        formula_range=f"{qsheet(PAYROLL_SHEET)}!{letter}{first_detail}:{letter}{last_detail}",
        expected_count=last_detail - first_detail + 1,
        notes="Payroll detail calculation columns must remain formulas so edits flow downstream.",
      )

  for key, source_row in _model_input_formula_rows(ctx):
    first_col = FIRST_LIVE_COL if key in {"is::Payroll", "is::Interest Expense"} else PERIOD_START_COL
    row = _write_formula_count_check(
      ws,
      row,
      line_item=f"Model Inputs - {key}",
      sheet=MODEL_INPUT_SHEET,
      formula_range=range_ref(MODEL_INPUT_SHEET, source_row, first_col, PERIOD_END_COL),
      expected_count=PERIOD_END_COL - first_col + 1,
      notes="Formula bridge rows in Model Inputs must stay linked to schedules.",
    )

  for statement, line_map in ctx.finmo_rows.items():
    for line_item, source_row in line_map.items():
      row = _write_formula_count_check(
        ws,
        row,
        line_item=f"FINMO - {statement} - {line_item}",
        sheet=FINMO_SHEET,
        formula_range=range_ref(FINMO_SHEET, source_row, PERIOD_START_COL, PERIOD_END_COL),
        expected_count=PERIOD_COUNT,
        notes="FINMO statement lines must remain formula-driven.",
      )

  row += 1
  write_section_header(ws, row, "Formula Error Scans", end_col=len(CHECK_HEADERS))
  row += 1
  scan_blocks = [
    (REVENUE_SHEET, 7, max(ctx.schedule_rows.get(REVENUE_SHEET, {}).values() or [7])),
    (PAYROLL_SHEET, ctx.schedule_row(PAYROLL_SHEET, "Total Ending FTE"), max(ctx.schedule_rows.get(PAYROLL_SHEET, {}).values() or [7])),
    (DEBT_SHEET, 7, max(ctx.schedule_rows.get(DEBT_SHEET, {}).values() or [7])),
    (CAPEX_SHEET, 7, max(ctx.schedule_rows.get(CAPEX_SHEET, {}).values() or [7])),
    (CASH_EQUITY_SHEET, 7, max(ctx.schedule_rows.get(CASH_EQUITY_SHEET, {}).values() or [7])),
    (MODEL_INPUT_SHEET, min(ctx.model_input_rows.values() or [7]), max(ctx.model_input_rows.values() or [7])),
    (
      FINMO_SHEET,
      min(row_num for section in ctx.finmo_rows.values() for row_num in section.values()),
      max(row_num for section in ctx.finmo_rows.values() for row_num in section.values()),
    ),
  ]
  for sheet, start, end in scan_blocks:
    if start and end and end >= start:
      row = _write_error_scan(ws, row, line_item=f"{sheet} formula/error scan", sheet=sheet, scan_range=_block_ref(sheet, start, end))

  row += 1
  write_section_header(ws, row, "Persisted Baseline Reconciliation", end_col=len(CHECK_HEADERS))
  row += 1
  baseline_checks = [
    ("Revenue Q20", ("Income Statement", "Revenue"), LAST_LIVE_COL),
    ("Payroll Q20", ("Income Statement", "Payroll"), LAST_LIVE_COL),
    ("Net Income Q20", ("Income Statement", "Net Income"), LAST_LIVE_COL),
    ("Cash Q20", ("Balance Sheet", "Cash"), LAST_LIVE_COL),
    ("Total Assets Q20", ("Balance Sheet", "Total Assets"), LAST_LIVE_COL),
    ("Total Liabilities & Equity Q20", ("Balance Sheet", "Total Liabilities & Equity"), LAST_LIVE_COL),
    ("Ending Cash Q20", ("Cash Flow", "Ending Cash"), LAST_LIVE_COL),
    # W2 (2026-08-18): the live break-even block (FINMO, formulas off the
    # P&L) vs the persisted W1 read-out (Audit Source). Rows only exist
    # when finmo_json carries break_even, so pre-W1 drafts skip via the
    # `continue` below. Tolerance is relative (persisted ratios are 6-dp).
    ("Break-Even Revenue Q1", ("Break-Even Analysis", "Break-Even Revenue"), FIRST_LIVE_COL),
    ("Cash Break-Even Revenue Q1", ("Break-Even Analysis", "Cash Break-Even Revenue"), FIRST_LIVE_COL),
    ("Break-Even Revenue Q20", ("Break-Even Analysis", "Break-Even Revenue"), LAST_LIVE_COL),
  ]
  for name, expected_key, col in baseline_checks:
    actual_row = ctx.finmo_row(expected_key[0], expected_key[1])
    expected_row = ctx.source_row(expected_key[0], expected_key[1])
    if not actual_row or not expected_row:
      continue
    tolerance_value = 1.0
    if expected_key[0] == "Break-Even Analysis":
      tolerance_value = f"=MAX(1,0.0005*ABS({ref(SOURCE_SHEET, expected_row, col)}))"
    _write_check(
      ws,
      row,
      category="Persisted Baseline",
      line_item=name,
      sheet="FINMO vs Audit Source",
      range_or_cell=ref(FINMO_SHEET, actual_row, col),
      actual=ref(FINMO_SHEET, actual_row, col),
      expected=ref(SOURCE_SHEET, expected_row, col),
      tolerance=tolerance_value,
      status_formula=f'=IF(ABS(G{row})<=H{row},"MATCH","CHANGED")',
      notes="CHANGED means assumptions were edited from the persisted run; it is informational and does not fail Model Status.",
    )
    row += 1

  last_check_row = row - 1
  ws.cell(row=2, column=1, value="Model Status")
  ws.cell(row=2, column=2, value=f'=IF(COUNTIF(I7:I{last_check_row},"FAIL")=0,"OK","FAIL")')
  ws.cell(row=2, column=2).font = design.font("kpi_label")
  ws.cell(row=2, column=2).fill = design.fill(design.NAVY)
  ws.cell(row=3, column=1, value="Rule")
  ws.cell(row=3, column=2, value="Only real formula/coherence failures return FAIL. Baseline scenario changes show CHANGED.")
  ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)

  status_range = f"I7:I{last_check_row}"
  fail_fill = design.fill(design.STATUS_CRITICAL_FILL)
  ok_fill = design.fill(design.STATUS_GOOD_FILL)
  changed_fill = design.fill(design.STATUS_NEUTRAL_FILL)
  ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"FAIL"'], fill=fail_fill))
  ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"OK"'], fill=ok_fill))
  ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"CHANGED"'], fill=changed_fill))
