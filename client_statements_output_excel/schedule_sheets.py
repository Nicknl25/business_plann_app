from __future__ import annotations

from typing import Dict, List

from openpyxl.styles import Alignment, Font, PatternFill

from . import design
from .data import DraftWorkbookData, live_values, number, row_by_label, text, values_21
from openpyxl.utils import get_column_letter

from .excel_utils import (
  ANNUAL_ANNUALIZE,
  ANNUAL_AVERAGE,
  ANNUAL_SUM,
  ANNUAL_YEAR_END,
  ANNUAL_YEAR_START,
  ANNUAL_START_COL,
  CAPEX_SHEET,
  CASH_EQUITY_SHEET,
  CURRENCY_FORMAT,
  DATE_FORMAT,
  DEBT_SHEET,
  FILL_BLUE,
  FILL_GRAY,
  FILL_GREEN,
  FILL_LIGHT,
  FIRST_LIVE_COL,
  INTEGER_FORMAT,
  LAST_LIVE_COL,
  MODEL_INPUT_SHEET,
  NUMBER_FORMAT,
  PAYROLL_SHEET,
  PERCENT_FORMAT,
  PERIOD_COUNT,
  PERIOD_END_COL,
  PERIOD_START_COL,
  REVENUE_SHEET,
  WORKING_CAPITAL_SHEET,
  WorkbookBuildContext,
  add_annual_formulas,
  apply_base_style,
  create_sheet,
  hide_stub_column,
  local_ref,
  qsheet,
  ref,
  set_formula_style,
  set_input_style,
  set_title,
  style_row,
  write_period_headers,
  write_section_header,
  write_values_row,
)


def _fmt_for_row(row: Dict[str, object]) -> str:
  kind = text(row.get("value_kind"))
  semantics = text(row.get("input_semantics")).lower()
  label = text(row.get("label")).lower()
  if kind == "ratio" or "percent" in semantics or "rate" in label or "percent" in label:
    return PERCENT_FORMAT
  if kind == "count" or "days" in semantics or "fte" in semantics or "day" in label:
    return NUMBER_FORMAT
  if kind == "currency":
    return CURRENCY_FORMAT
  return CURRENCY_FORMAT


def _annual_quarter_bounds(year_index: int) -> tuple[int, int]:
  start_col = FIRST_LIVE_COL + ((year_index - 1) * 4)
  return start_col, start_col + 3


def _add_annual_average_formulas(ws, row: int, *, number_format: str) -> None:
  for year in range(1, 6):
    start_col, end_col = _annual_quarter_bounds(year)
    cell = ws.cell(
      row=row,
      column=ANNUAL_START_COL + year - 1,
      value=f"=AVERAGE({local_ref(row, start_col)}:{local_ref(row, end_col)})",
    )
    set_formula_style(cell, number_format=number_format)


def _add_annual_ratio_formulas(
  ws,
  row: int,
  *,
  numerator_row: int,
  denominator_row: int,
  number_format: str,
) -> None:
  for year in range(1, 6):
    col = ANNUAL_START_COL + year - 1
    cell = ws.cell(
      row=row,
      column=col,
      value=f"=IFERROR({local_ref(numerator_row, col)}/{local_ref(denominator_row, col)},0)",
    )
    set_formula_style(cell, number_format=number_format)


#: Half a cent. Any balance below it is zero - the threshold under which a
#: rounding crumb (an ulp left by Excel reproducing python) is treated as the
#: payoff it is, rather than a balance that keeps accruing interest or
#: depreciating forever. Shared by every closing balance on these sheets.
SUB_CENT = "0.005"


def build_revenue_drivers_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, REVENUE_SHEET)
  apply_base_style(ws)
  set_title(ws, "Revenue Drivers", "Source operating drivers. Model Inputs links to these rows.")
  write_period_headers(ws, data.periods)
  write_section_header(ws, 6, "Revenue Driver Source Rows")
  row = 7
  revenue_groups: Dict[str, Dict[str, Dict[str, object]]] = {}
  ordered: List[str] = []
  for source_row in data.revenue_rows:
    slot = text(source_row.get("revenue_slot_key")) or f"{text(source_row.get('lob'))}::{text(source_row.get('product'))}"
    if slot not in revenue_groups:
      ordered.append(slot)
      revenue_groups[slot] = {}
    revenue_groups[slot][text(source_row.get("driver"))] = source_row

  total_capacity_rows: List[int] = []
  total_revenue_rows: List[int] = []
  for slot in ordered:
    group = revenue_groups[slot]
    display = " / ".join(
      [
        text(next(iter(group.values())).get("lob")) or "LOB",
        text(next(iter(group.values())).get("product")) or "Product",
      ]
    )
    write_section_header(ws, row, display)
    row += 1
    # WS1(b): the per-line "COGS %" source row exists only on multi-line
    # drafts whose lines carry their own COGS percents — single-line
    # workbooks are unchanged.
    slot_drivers = ["Capacity", "Unit Price", "Utilization"] + (["COGS %"] if "COGS %" in group else [])
    for driver in slot_drivers:
      source = group.get(driver, {"values": []})
      fmt = NUMBER_FORMAT if driver == "Capacity" else CURRENCY_FORMAT if driver == "Unit Price" else PERCENT_FORMAT
      write_values_row(
        ws,
        row,
        f"{display} - {driver}",
        values_21(source.get("values")),
        detail="Source driver",
        number_format=fmt,
      )
      ctx.add_schedule_row(REVENUE_SHEET, f"{slot}::{driver}", row)
      if driver == "Capacity":
        total_capacity_rows.append(row)
      style_row(ws, row, number_format=fmt)
      row += 1
    revenue_row = row
    ws.cell(row=row, column=1, value=f"{display} - Revenue")
    ws.cell(row=row, column=2, value="Capacity x Unit Price x Utilization")
    for idx in range(PERIOD_COUNT):
      col = PERIOD_START_COL + idx
      cap = local_ref(ctx.schedule_row(REVENUE_SHEET, f"{slot}::Capacity"), col)
      price = local_ref(ctx.schedule_row(REVENUE_SHEET, f"{slot}::Unit Price"), col)
      util = local_ref(ctx.schedule_row(REVENUE_SHEET, f"{slot}::Utilization"), col)
      cell = ws.cell(row=row, column=col, value=f"={cap}*{price}*{util}")
      set_formula_style(cell, number_format=CURRENCY_FORMAT)
    add_annual_formulas(ws, row, label=f"{display} - Revenue")
    style_row(ws, row, fill=FILL_GREEN, bold=True, number_format=CURRENCY_FORMAT, border_top=True)
    ctx.add_schedule_row(REVENUE_SHEET, f"{slot}::Revenue", row)
    total_revenue_rows.append(revenue_row)
    row += 2

  ws.cell(row=row, column=1, value="Total Capacity Units")
  ws.cell(row=row, column=2, value="All revenue driver capacity units")
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    refs = [local_ref(r, col) for r in total_capacity_rows]
    ws.cell(row=row, column=col, value=f"=SUM({','.join(refs)})" if refs else "=0")
    set_formula_style(ws.cell(row=row, column=col), number_format=NUMBER_FORMAT)
  add_annual_formulas(ws, row, label="Total Capacity Units", number_format=NUMBER_FORMAT)
  style_row(ws, row, fill=FILL_LIGHT, bold=True, number_format=NUMBER_FORMAT, border_top=True)
  ctx.add_schedule_row(REVENUE_SHEET, "Total Capacity Units", row)
  row += 1

  ws.cell(row=row, column=1, value="Total Revenue")
  ws.cell(row=row, column=2, value="All revenue products")
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    refs = [local_ref(r, col) for r in total_revenue_rows]
    ws.cell(row=row, column=col, value=f"=SUM({','.join(refs)})" if refs else "=0")
    set_formula_style(ws.cell(row=row, column=col), number_format=CURRENCY_FORMAT)
  add_annual_formulas(ws, row, label="Total Revenue")
  style_row(ws, row, fill=FILL_BLUE, bold=True, number_format=CURRENCY_FORMAT, border_top=True)
  ctx.add_schedule_row(REVENUE_SHEET, "Total Revenue", row)
  total_revenue_row = row
  row += 2

  # ACTUAL REVENUE QoQ GROWTH lives HERE now, in the live driver section
  # directly under revenue, because it is a RESULT of the drivers above and it
  # recomputes when a client edits them. It used to sit inside a "Stage Ramp
  # Contract" block that has been omitted (R_RAMP_01, 7f9be65) - leaving the one
  # live row under a header for a block that no longer exists would have
  # mislabelled it. Addressed through the ctx key, never a literal row number,
  # because its row moves with the number of products.
  ws.cell(row=row, column=1, value="Actual Revenue QoQ Growth")
  ws.cell(row=row, column=2, value="Total revenue growth from modeled revenue drivers")
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    formula = "=0" if idx == 0 else f"=IFERROR({local_ref(total_revenue_row, col)}/{local_ref(total_revenue_row, col - 1)}-1,0)"
    cell = ws.cell(row=row, column=col, value=formula)
    set_formula_style(cell, number_format=PERCENT_FORMAT, internal_link=True)
  _add_annual_average_formulas(ws, row, number_format=PERCENT_FORMAT)
  style_row(ws, row, fill=FILL_LIGHT, number_format=PERCENT_FORMAT)
  ctx.add_schedule_row(REVENUE_SHEET, "Actual Revenue QoQ Growth", row)
  row += 1

  hide_stub_column(ws)

  # THE STAGE RAMP CONTRACT BLOCK IS OMITTED (Nick's ruling on R_RAMP_01 A1).
  # Eleven rows x 21 columns of engine constants that NOTHING in the workbook
  # consumed - proven across formulas, defined names, data validation,
  # conditional formatting, chart series and hyperlinks, on both fixtures - and
  # all 231 cells carried the amber input styling, so a client scanning for what
  # they could change found eleven rows of solver constraints dressed as levers.
  #
  # The record is NOT lost: stage_ramp_contract stays in planning_run_json with
  # its rationale, decision_source and business_stage. This omits a RENDERING.
  #
  # Marketing % Max, which the Marketing Schedule shows as context, never came
  # from here either - it reads stage_ramp_contract.quarter_ramp_grid directly.


# ---------------------------------------------------------------------------
# PAYROLL SCHEDULE - CLIENT LANGUAGE (Nick, 2026-08-25). The payroll payload
# carries code values (key_person, oews_title_catalog:oews_median,
# labor_driven ...). They were printed verbatim into client-facing cells.
# Nothing computed reads these strings - the summary SUMIFS and the Checks
# tie-out key on the numeric Quarter in column A and on fixed column letters -
# so the mapping is presentation only. Unknown values fall through unchanged
# rather than failing a build over a label.
# ---------------------------------------------------------------------------
_STAFFING_CLASS_TEXT = {
  "key_person": "Named person",
  "supporting_staff": "Staffed role",
}
_WAGE_SOURCE_TEXT = {
  "client_override": "Your stated wage",
  "intake_oews_key_person": "Your stated wage",
  "oews_title_catalog:oews_median": "Market median for this title (BLS)",
  "oews_median": "Market median for this title (BLS)",
  "oews_pct10": "Market 10th percentile for this title (BLS)",
  "oews_pct25": "Market 25th percentile for this title (BLS)",
  "oews_pct75": "Market 75th percentile for this title (BLS)",
  "oews_pct90": "Market 90th percentile for this title (BLS)",
}
_WAGE_SOURCE_SUFFIX_TEXT = {
  "floor_adapted": "raised to the wage floor",
  "part_time_hours_adapted": "adjusted for part-time hours",
  "owner_draw_deferred": "owner draw deferred",
}
_CAPACITY_MODEL_TEXT = {
  "labor_driven": "Capacity grows with headcount",
  "hybrid": "Capacity grows with headcount and systems",
  "system_driven": "Capacity grows with systems, not headcount",
  "expert_driven": "Capacity grows with expertise",
}
_LABOR_INTENSITY_TEXT = {
  "low": "Low - few staff per unit of revenue",
  "medium": "Medium",
  "high": "High - many staff per unit of revenue",
  "expert": "Expert - a few specialists",
}
_WAGE_TIER_TEXT = {
  "floor": "Floor - 1.00 to 1.10x the market reference",
  "market": "Market - 1.10 to 1.40x the market reference",
  "premium": "Premium - 1.35 to 1.80x the market reference",
  "specialized": "Specialized - 1.70 to 2.50x the market reference",
}


def _plain(value, table):
  key = str(value or "").strip()
  return table.get(key.lower(), key) if key else key


def _wage_source_plain(value) -> str:
  raw = str(value or "").strip()
  if not raw:
    return ""
  base, _, rest = raw.partition("|")
  words = _WAGE_SOURCE_TEXT.get(base.lower(), base)
  notes = [_WAGE_SOURCE_SUFFIX_TEXT.get(x.strip().lower(), x.strip())
           for x in rest.split("|") if x.strip()]
  return words + (" (" + ", ".join(notes) + ")" if notes else "")


#: Column A of the payroll detail keeps its NUMERIC quarter index - the
#: summary SUMIFS and the Checks tie-out key on it - and only DISPLAYS as
#: Stub / Q1 ... Q20 like every other sheet's period headers.
QUARTER_INDEX_FORMAT = design.FMT_QUARTER_INDEX


def build_payroll_schedule_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, PAYROLL_SHEET)
  apply_base_style(ws)
  set_title(ws, "Payroll Schedule",
            "Who is on payroll, quarter by quarter: each role's headcount, wage and "
            "benefits, and what they add up to. Model Inputs takes Total Payroll from "
            "this sheet.")
  write_period_headers(ws, data.periods)
  root = data.payroll_headcount
  summary_row = 6
  write_section_header(ws, summary_row, "Payroll Assumptions")
  assumptions = [
    ("Capacity Labor Model", _plain(root.get("capacity_labor_model"), _CAPACITY_MODEL_TEXT)),
    ("Labor Intensity Class", _plain(root.get("labor_intensity_class"), _LABOR_INTENSITY_TEXT)),
    ("Wage Positioning Tier", _plain(root.get("wage_positioning_tier"), _WAGE_TIER_TEXT)),
    ("Wage Positioning Multiplier", root.get("wage_positioning_multiplier")),
    ("Capacity Units per Supporting FTE", root.get("capacity_units_per_supporting_fte")),
    ("Target Payroll % of Revenue", root.get("target_payroll_percent_of_revenue")),
  ]
  row = summary_row + 1
  for label, value in assumptions:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    row += 1

  row += 1
  write_section_header(ws, row, "Quarter Summary")
  row += 1
  summary_rows: Dict[str, tuple[int, str, str]] = {}
  for label, key, fmt in [
    ("Total Ending FTE", "ending_fte", NUMBER_FORMAT),
    ("Total Average FTE", "average_fte", NUMBER_FORMAT),
    ("Total Payroll", "payroll", CURRENCY_FORMAT),
    ("Total Revenue", "revenue", CURRENCY_FORMAT),
    ("Total Capacity Units", "capacity_units", NUMBER_FORMAT),
    ("Revenue per Employee", "revenue_per_employee", CURRENCY_FORMAT),
    ("Units per Employee", "units_per_employee", NUMBER_FORMAT),
    ("Payroll % of Revenue", "payroll_percent_of_revenue", PERCENT_FORMAT),
  ]:
    ws.cell(row=row, column=1, value=label)
    detail = "Formula output from payroll detail"
    if key in {"revenue", "capacity_units"}:
      detail = "Linked from Revenue Drivers"
    elif key == "revenue_per_employee":
      detail = "Revenue divided by average FTE"
    elif key == "units_per_employee":
      detail = "Capacity units divided by average FTE"
    elif key == "payroll_percent_of_revenue":
      detail = "Total payroll divided by revenue"
    ws.cell(row=row, column=2, value=detail)
    ctx.add_schedule_row(PAYROLL_SHEET, label, row)
    summary_rows[label] = (row, key, fmt)
    row += 1

  row += 2
  write_section_header(ws, row, "Payroll Detail")
  row += 1
  detail_header_row = row
  headers = [
    "Quarter",
    "Staffing Class",
    "Title / Person",
    "OEWS Title",
    "Starting FTE",
    "Hires",
    "Ending FTE",
    "Average FTE",
    "Annual Wage",
    "Benefits %",
    "Wage Cost",
    "Taxes & Benefits",
    "Total Payroll",
    "Wage Source",
  ]
  for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col, value=header)
    cell.fill = design.fill(design.NAVY_DEEP)
    cell.font = design.font("colhead")
  row += 1
  detail_start_row = row
  # THE DETAIL BLOCK IS CHAINED (Nick, 2026-08-25, Payroll Schedule part 2).
  # It was 80 rows of independent literals - a wage raise meant editing 20
  # cells, and a hire in Q4 never reached Q5's starting headcount. Now, for a
  # role with a row in the PRIOR quarter:
  #   Starting FTE  = ROUND(prior Ending FTE, 6)
  #   Annual Wage   = prior wage, or ROUND(prior +/- the engine's delta, 6)
  #                   at the engine's own bump quarters (3%/yr as authored)
  #   Benefits %    = prior, same rule
  #   Hires         = literal - a per-quarter flow, like Distributions
  # The ROUND is the same guard the Debt Schedule payoff and the equity chain
  # use: the engine authors FTE and hires on a 2-dp grid (6.41), and
  # 6.06 + 0.35 in IEEE is 6.409999999999999 - without the grid the crumb
  # compounds through Ending, Average, Wage Cost and the P&L (28 cells on
  # Halbrook in the prototype). A role with NO prior-quarter row keeps its
  # literals (0 such rows in the population today; the fallback ships anyway).
  # Every chained cell stays amber: the client types over it and the chain
  # picks their number up from that point down. Nothing editable was lost;
  # Staffing Class, Title/Person and Wage Source become editable too
  # (nothing computed reads them - the SUMIFS key on the numeric Quarter).
  # ROLE IDENTITY across quarters is (staffing class, title, person, ordinal):
  # 73 of 390 drafts carry two named people with the SAME title (two "Dental
  # Hygienist" key persons), and the ordinal covers two identical supporting
  # rows in one quarter. Keyed on title alone they collided and fell back to
  # literals - right numbers, no chain.
  prior_row_by_role: Dict[tuple, tuple] = {}
  seen_this_quarter: Dict[tuple, int] = {}
  current_quarter = None
  for item in root.get("rows") or []:
    if not isinstance(item, dict):
      continue
    quarter = number(item.get("quarter_index"))
    if quarter != current_quarter:
      current_quarter, seen_this_quarter = quarter, {}
    base_key = (
      str(item.get("staffing_class") or "").strip(),
      str(item.get("position_title") or "").strip(),
      str(item.get("person_name") or "").strip(),
    )
    ordinal = seen_this_quarter.get(base_key, 0)
    seen_this_quarter[base_key] = ordinal + 1
    role_key = base_key + (ordinal,)
    prior = prior_row_by_role.get(role_key)
    prior_row = prior[0] if prior and quarter is not None and prior[1] == quarter - 1 else None
    starting = number(item.get("starting_fte"))
    wage = number(item.get("annual_wage"))
    benefits = number(item.get("payroll_taxes_benefits_percent"))

    def _chained(col: int, value, prev_value):
      """prev, or ROUND(prev +/- delta, 6) where the engine's series steps."""
      prev_ref = local_ref(prior_row, col)
      delta = (value or 0.0) - (prev_value or 0.0)
      if abs(delta) <= 1e-9:
        return f"={prev_ref}"
      sign = "-" if delta < 0 else "+"
      return f"=ROUND({prev_ref}{sign}{abs(delta)!r},6)"

    ws.cell(row=row, column=1, value=quarter)
    ws.cell(row=row, column=1).number_format = QUARTER_INDEX_FORMAT
    ws.cell(row=row, column=2, value=_plain(item.get("staffing_class"), _STAFFING_CLASS_TEXT))
    ws.cell(row=row, column=3, value=text(item.get("position_title") or item.get("person_name")))
    ws.cell(row=row, column=4, value=text(item.get("oews_occ_title") or item.get("oews_matched_title")))
    if prior_row is not None:
      ws.cell(row=row, column=5, value=f"=ROUND({local_ref(prior_row, 7)},6)")
      ws.cell(row=row, column=9, value=_chained(9, wage, prior[2]))
      ws.cell(row=row, column=10, value=_chained(10, benefits, prior[3]))
    else:
      ws.cell(row=row, column=5, value=starting)
      ws.cell(row=row, column=9, value=wage)
      ws.cell(row=row, column=10, value=benefits)
    ws.cell(row=row, column=6, value=number(item.get("hires")))
    ws.cell(row=row, column=7, value=f"={local_ref(row, 5)}+{local_ref(row, 6)}")
    ws.cell(row=row, column=8, value=f"=({local_ref(row, 5)}+{local_ref(row, 7)})/2")
    ws.cell(row=row, column=11, value=f"={local_ref(row, 8)}*{local_ref(row, 9)}/4")
    ws.cell(row=row, column=12, value=f"={local_ref(row, 11)}*{local_ref(row, 10)}")
    ws.cell(row=row, column=13, value=f"={local_ref(row, 11)}+{local_ref(row, 12)}")
    ws.cell(row=row, column=14, value=_wage_source_plain(item.get("wage_source") or item.get("wage_source_code")))
    for col in [2, 3, 14]:
      set_input_style(ws.cell(row=row, column=col), number_format=design.FMT_TEXT)
    for col in [5, 6, 9, 10]:
      set_input_style(ws.cell(row=row, column=col), number_format=PERCENT_FORMAT if col == 10 else CURRENCY_FORMAT if col == 9 else NUMBER_FORMAT)
    for col in [7, 8]:
      set_formula_style(ws.cell(row=row, column=col), number_format=NUMBER_FORMAT, internal_link=True)
    for col in [11, 12, 13]:
      set_formula_style(ws.cell(row=row, column=col), number_format=CURRENCY_FORMAT, internal_link=True)
    prior_row_by_role[role_key] = (row, quarter, wage, benefits)
    row += 1
  detail_last_row = row - 1
  ctx.add_schedule_row(PAYROLL_SHEET, "Payroll Detail First Row", detail_start_row)
  ctx.add_schedule_row(PAYROLL_SHEET, "Payroll Detail Last Row", detail_last_row)

  for label, (summary_output_row, key, fmt) in summary_rows.items():
    for q in range(PERIOD_COUNT):
      col = PERIOD_START_COL + q
      if key in {"ending_fte", "average_fte", "payroll"} and detail_last_row >= detail_start_row:
        source_col_letter = {"ending_fte": "G", "average_fte": "H", "payroll": "M"}[key]
        formula = (
          f"=SUMIFS(${source_col_letter}${detail_start_row}:${source_col_letter}${detail_last_row},"
          f"$A${detail_start_row}:$A${detail_last_row},{q})"
        )
      elif key == "revenue":
        source_row = ctx.schedule_row(REVENUE_SHEET, "Total Revenue")
        formula = f"={ref(REVENUE_SHEET, source_row, col)}" if source_row else "=0"
      elif key == "capacity_units":
        source_row = ctx.schedule_row(REVENUE_SHEET, "Total Capacity Units")
        formula = f"={ref(REVENUE_SHEET, source_row, col)}" if source_row else "=0"
      elif key == "revenue_per_employee":
        formula = f"=IFERROR({local_ref(summary_rows['Total Revenue'][0], col)}/{local_ref(summary_rows['Total Average FTE'][0], col)},0)"
      elif key == "units_per_employee":
        formula = f"=IFERROR({local_ref(summary_rows['Total Capacity Units'][0], col)}/{local_ref(summary_rows['Total Average FTE'][0], col)},0)"
      elif key == "payroll_percent_of_revenue":
        formula = f"=IFERROR({local_ref(summary_rows['Total Payroll'][0], col)}/{local_ref(summary_rows['Total Revenue'][0], col)},0)"
      else:
        formula = "=0"
      cell = ws.cell(row=summary_output_row, column=col, value=formula)
      set_formula_style(cell, number_format=fmt, internal_link=True)
    if key == "average_fte":
      _add_annual_average_formulas(ws, summary_output_row, number_format=fmt)
    elif key == "revenue_per_employee":
      _add_annual_ratio_formulas(
        ws,
        summary_output_row,
        numerator_row=summary_rows["Total Revenue"][0],
        denominator_row=summary_rows["Total Average FTE"][0],
        number_format=fmt,
      )
    elif key == "units_per_employee":
      _add_annual_ratio_formulas(
        ws,
        summary_output_row,
        numerator_row=summary_rows["Total Capacity Units"][0],
        denominator_row=summary_rows["Total Average FTE"][0],
        number_format=fmt,
      )
    elif key == "payroll_percent_of_revenue":
      _add_annual_ratio_formulas(
        ws,
        summary_output_row,
        numerator_row=summary_rows["Total Payroll"][0],
        denominator_row=summary_rows["Total Revenue"][0],
        number_format=fmt,
      )
    else:
      add_annual_formulas(ws, summary_output_row,
                          mode=ANNUAL_YEAR_END if key == "ending_fte" else None,
                          label=label, number_format=fmt)
    style_row(
      ws,
      summary_output_row,
      fill=FILL_GREEN if key in {"payroll", "revenue_per_employee", "units_per_employee"} else FILL_LIGHT,
      bold=(key in {"payroll", "revenue_per_employee", "units_per_employee"}),
      number_format=fmt,
    )

  for col in range(1, len(headers) + 1):
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = max(ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width or 12, 16)
  ws.auto_filter.ref = f"A{detail_header_row}:N{max(detail_header_row, detail_last_row)}"


def build_debt_schedule_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  """A vertical amortization schedule: periods DOWN, one row per quarter.

  Debt columns and capital-lease columns sit side by side against the same
  period rows, with combined totals on the right. A client scans down and sees
  each balance walk to zero and exactly which quarter it gets there.

  WHY VERTICAL AT ALL. The rest of the workbook is a 21-column horizontal
  grid, and this sheet used to match it - which made it a grid of drivers
  rather than a schedule. An amortization table reads down.

  HOW IT STILL FEEDS THE STATEMENTS (the part that had to be solved). A HIDDEN
  HORIZONTAL BRIDGE sits below the table: one row per figure the rest of the
  workbook consumes, 21 columns wide, in the standard geometry, each cell a
  direct reference back into the vertical table. Every bridge row is
  registered under the SAME ctx key the sheet has always published, so
  Model Inputs, Checks, FINMO and CapEx need no change at all - they ask for
  "Closing Debt" and get a horizontal row exactly as before.

  Chosen over INDEX lookups down the schedule because the references are
  plain and non-volatile, and over re-pointing the consumers because that
  would have meant editing four other sheets to rearrange one.

  WHAT IS NOT HERE. The right-of-use asset. It depreciates rather than
  amortizing, so it does not belong in an amortization table; it is computed
  in the bridge from the lease columns and published under its existing keys.

  PRINCIPAL COMPUTES; NOTHING IS PASTED (CW-043). The first vertical build
  pasted principal as a literal - the one pasted cell in each row - so a
  zero authored upstream froze the lease at its opening balance for five
  years while the ROU asset depreciated to nothing, and every downstream
  tie-out reconciled the frozen number with itself. Now, per block:

    Scheduled principal = (Opening + New) / remaining term   (formula)
    Extra principal     = engine repayment - scheduled       (INPUT column)
    Principal           = MIN(Opening + New, MAX(0, ROUND(Scheduled + Extra, 6)))
    Closing             = Opening + New - Principal, zero below half a cent
    Payment             = Interest + Principal

  Opening-over-remaining-term reproduces the engine's straight-line base
  exactly (180,000/32 = 5,625; 174,375/31 = 5,625; ...) and re-amortizes on
  its own after any client edit. The engine's cash sweeps - real, and never
  to be smoothed away - land in Extra principal, explicit and editable, on
  top of the scheduled amount. Scheduled + Extra equals the engine's
  repayment by construction, so delivered numbers do not move.

  THE DEBT TERM IS THE CASH PASS'S OWN JUDGMENT, read from
  model_input_json.solver_input.cash_judgment.debt_term_quarters (the
  earlier note here claiming no term existed in the payload was wrong -
  it lives under solver_input, not schedules). When a draft predates that
  judgment the term is implied from the engine's own base repayment
  (seed / first repayment), and failing that the horizon; either way the
  split is presentational - Extra absorbs any difference, so totals match
  the engine regardless of the term shown.
  """
  ws = create_sheet(wb, DEBT_SHEET)
  apply_base_style(ws)
  set_title(
    ws,
    "Debt & Lease Amortization Schedule",
    "One row per quarter. Each balance opens where the quarter above it "
    "closed. Amber cells are yours: new borrowing, rates, terms, extra "
    "principal and lease additions. Every other cell computes - scheduled "
    "principal comes from the balance over the remaining term, and payment "
    "is interest plus principal.",
  )
  schedule_by_label = row_by_label(data.schedule_rows)
  expense_by_label = row_by_label(data.expense_rows)
  schedules = data.schedules

  issuance_values = values_21((schedule_by_label.get("Debt Issuance (New Borrowing)") or {}).get("values"))
  repay_values = values_21((schedule_by_label.get("Debt Repayment (Scheduled)") or {}).get("values"))
  rate_values = values_21((expense_by_label.get("Interest Rate") or {}).get("values"))
  lease_principal_values = values_21((schedule_by_label.get("Less: Principal Repayments") or {}).get("values"))
  lease_add_values = values_21((schedule_by_label.get("Plus: Net Additions") or {}).get("values"))
  lease_life_quarters = int(number(schedules.get("lease_life_quarters")) or 20)
  debt_seed = number(schedules.get("debt_opening_balance_seed"))
  lease_seed = number(schedules.get("lease_opening_balance_seed"))

  def _walk(seed, adds, takes):
    """Balances walked forward with each draw capped at what is owed.

    The repayment a client sees is already capped here, in Python, exactly as
    the MIN chain used to do it in Excel - so the schedule has ONE principal
    line rather than a requested one and an actual one, and the figures are
    unchanged.
    """
    capped = []
    balance = number(seed) or 0.0
    for idx in range(PERIOD_COUNT):
      add = 0.0 if idx == 0 else (adds[idx] or 0.0)
      want = 0.0 if idx == 0 else (takes[idx] or 0.0)
      take = min(want, balance + add)
      capped.append(take)
      balance = max(0.0, balance + add - take)
    return capped

  debt_principal = _walk(debt_seed, issuance_values, repay_values)
  lease_principal = _walk(lease_seed, lease_add_values, lease_principal_values)

  def _judged_debt_term() -> int:
    """The cash pass's own judged term, from the payload - never invented.

    Fallbacks for drafts that predate the judgment: the term implied by the
    engine's own base repayment, then the horizon. The choice only moves the
    scheduled/extra SPLIT, never the totals - Extra absorbs the difference.
    """
    solver = data.model_input_json.get("solver_input")
    judged = ((solver if isinstance(solver, dict) else {}).get("cash_judgment")
              or {}).get("debt_term_quarters")
    term = int(number(judged) or 0)
    if term >= 1:
      return term
    seed = number(debt_seed) or 0.0
    base = next((v for v in debt_principal[1:] if v and v > 0), 0.0)
    if seed > 0 and base > 0:
      return max(1, int(round(seed / base)))
    return PERIOD_COUNT - 1

  debt_term_quarters = _judged_debt_term()

  def _extra_split(seed, adds, principal, term):
    """Per-row EXTRA principal: engine repayment minus the scheduled amount.

    Mirrors the sheet's own recurrence operation for operation - opening from
    prior closing, scheduled = (opening + additions) / remaining term - so the
    authored Extra plus the Excel-computed Scheduled reproduces the engine's
    repayment in every row. Zero everywhere the engine amortizes on schedule;
    the cash sweeps (and nothing else) surface here.
    """
    extra = []
    balance = number(seed) or 0.0
    for idx in range(PERIOD_COUNT):
      add = 0.0 if idx == 0 else (adds[idx] or 0.0)
      sched = 0.0 if idx == 0 else (balance + add) / max(1, term - (idx - 1))
      take = principal[idx]
      extra.append(take - sched)
      balance = max(0.0, balance + add - take)
    return extra

  debt_extra = _extra_split(debt_seed, issuance_values, debt_principal, debt_term_quarters)
  lease_extra = _extra_split(lease_seed, lease_add_values, lease_principal, lease_life_quarters)

  # ---- geometry -----------------------------------------------------------
  C_PERIOD = 1
  (C_D_OPEN, C_D_NEW, C_D_RATE, C_D_TERM, C_D_SCHED, C_D_EXTRA,
   C_D_PRIN, C_D_INT, C_D_PAY, C_D_CLOSE) = range(2, 12)
  (C_L_OPEN, C_L_ADD, C_L_RATE, C_L_TERM, C_L_SCHED, C_L_EXTRA,
   C_L_PRIN, C_L_INT, C_L_PAY, C_L_CLOSE) = range(12, 22)
  C_T_PAY, C_T_INT, C_T_CLOSE = 22, 23, 24
  GROUP_ROW, HEAD_ROW = 5, 6
  FIRST_ROW = 7

  groups = [("Debt", C_D_OPEN, C_D_CLOSE),
            ("Capital lease", C_L_OPEN, C_L_CLOSE),
            ("Combined", C_T_PAY, C_T_CLOSE)]
  for label, first, last in groups:
    cell = ws.cell(GROUP_ROW, first, value=label)
    cell.font = design.font("colhead")
    cell.fill = design.fill(design.NAVY_DEEP)
    cell.alignment = Alignment(horizontal="center")
    for c in range(first, last + 1):
      ws.cell(GROUP_ROW, c).fill = design.fill(design.NAVY_DEEP)
    ws.merge_cells(start_row=GROUP_ROW, start_column=first,
                   end_row=GROUP_ROW, end_column=last)

  headers = [
    (C_PERIOD, "Period", None), (C_D_OPEN, "Opening", CURRENCY_FORMAT),
    (C_D_NEW, "New borrowing", CURRENCY_FORMAT), (C_D_RATE, "Rate", PERCENT_FORMAT),
    (C_D_TERM, "Term (q)", INTEGER_FORMAT),
    (C_D_SCHED, "Scheduled principal", CURRENCY_FORMAT),
    (C_D_EXTRA, "Extra principal", CURRENCY_FORMAT),
    (C_D_PRIN, "Principal", CURRENCY_FORMAT), (C_D_INT, "Interest", CURRENCY_FORMAT),
    (C_D_PAY, "Payment", CURRENCY_FORMAT), (C_D_CLOSE, "Closing", CURRENCY_FORMAT),
    (C_L_OPEN, "Opening", CURRENCY_FORMAT), (C_L_ADD, "Additions", CURRENCY_FORMAT),
    (C_L_RATE, "Rate", PERCENT_FORMAT), (C_L_TERM, "Term (q)", INTEGER_FORMAT),
    (C_L_SCHED, "Scheduled principal", CURRENCY_FORMAT),
    (C_L_EXTRA, "Extra principal", CURRENCY_FORMAT),
    (C_L_PRIN, "Principal", CURRENCY_FORMAT), (C_L_INT, "Interest", CURRENCY_FORMAT),
    (C_L_PAY, "Payment", CURRENCY_FORMAT), (C_L_CLOSE, "Closing", CURRENCY_FORMAT),
    (C_T_PAY, "Total payment", CURRENCY_FORMAT),
    (C_T_INT, "Total interest", CURRENCY_FORMAT),
    (C_T_CLOSE, "Total closing", CURRENCY_FORMAT),
  ]
  for col, text, _fmt in headers:
    cell = ws.cell(HEAD_ROW, col, value=text)
    cell.font = design.font("colhead")
    cell.fill = design.fill(design.NAVY)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
  ws.column_dimensions[get_column_letter(C_PERIOD)].width = 16
  for col, _t, _f in headers[1:]:
    ws.column_dimensions[get_column_letter(col)].width = 13
  ws.freeze_panes = ws.cell(FIRST_ROW, C_D_OPEN)

  INPUT_COLS = {C_D_NEW, C_D_RATE, C_D_TERM, C_D_EXTRA,
                C_L_ADD, C_L_RATE, C_L_TERM, C_L_EXTRA}
  periods = data.periods or []

  for idx in range(PERIOD_COUNT):
    r = FIRST_ROW + idx
    stub = idx == 0
    prev = r - 1
    period = periods[idx] if idx < len(periods) else {}
    ws.cell(r, C_PERIOD, value="Stub" if stub else f"Q{idx}")
    ws.cell(r, C_PERIOD).font = design.font("label_strong")

    period_ref = local_ref(r, C_PERIOD)
    elapsed = max(0, idx - 1)

    def _block(c_open, c_new, c_rate, c_term, c_sched, c_extra,
               c_prin, c_int, c_pay, c_close, *,
               seed, new_value, rate_value, term_value, extra_value,
               interest_kind):
      """One amortization block, principal COMPUTED - the only values are the
      client's inputs (new/rate/term/extra) and the stub's opening seed."""
      ws.cell(r, c_open, value=seed if stub else f"={local_ref(prev, c_close)}")
      ws.cell(r, c_new, value=new_value)
      ws.cell(r, c_rate, value=rate_value)
      ws.cell(r, c_term, value=term_value)
      # Straight-line on the CURRENT balance over the REMAINING term - which
      # reproduces the engine's level base exactly on an untouched schedule
      # and re-amortizes by itself after any edit or sweep. The stub is the
      # historical anchor period: nothing is ever scheduled in it.
      ws.cell(r, c_sched,
              value=f'=IF({period_ref}="Stub",0,({local_ref(r, c_open)}'
                    f'+{local_ref(r, c_new)})'
                    f'/MAX(1,{local_ref(r, c_term)}-{elapsed}))')
      ws.cell(r, c_extra, value=extra_value)
      # A ROUNDING CRUMB MUST NEVER BREAK THE MODEL (CW-043 TURN A). Excel's
      # sched + extra reproduces the engine's repayment only to within an
      # ulp; on Harrow 85f5825d the Q7 payoff landed at 11756.999999999998,
      # MAX(0, open - prin) kept the 5e-12 of debt left over, interest accrued
      # on it for thirteen quarters and FINMO's coverage ratios printed 1e17.
      # So principal lands on the engine's own 6-dp grid, and a closing
      # balance under half a cent IS zero - a payoff is a payoff.
      ws.cell(r, c_prin,
              value=f"=MIN({local_ref(r, c_open)}+{local_ref(r, c_new)},"
                    f"MAX(0,ROUND({local_ref(r, c_sched)}+{local_ref(r, c_extra)},6)))")
      _left = f"{local_ref(r, c_open)}+{local_ref(r, c_new)}-{local_ref(r, c_prin)}"
      ws.cell(r, c_close, value=f"=IF({_left}<{SUB_CENT},0,{_left})")
      if interest_kind == "average_balance":
        live_interest = (f"(({local_ref(r, c_open)}+{local_ref(r, c_close)})/2)"
                         f"*{local_ref(r, c_rate)}")
      else:
        live_interest = f"{local_ref(r, c_open)}*{local_ref(r, c_rate)}"
      # The stub is the historical anchor period: the engine schedules no
      # interest in it (debt_interest_expense = 0 at the stub, lease likewise),
      # and the stub's P&L interest is the client's STATED figure carried on
      # Model Inputs - not this cell. Live rows carry the bare formula.
      ws.cell(r, c_int, value=(f'=IF({period_ref}="Stub",0,{live_interest})' if stub
                               else f"={live_interest}"))
      ws.cell(r, c_pay, value=f"={local_ref(r, c_int)}+{local_ref(r, c_prin)}")

    # ---- debt
    _block(C_D_OPEN, C_D_NEW, C_D_RATE, C_D_TERM, C_D_SCHED, C_D_EXTRA,
           C_D_PRIN, C_D_INT, C_D_PAY, C_D_CLOSE,
           seed=debt_seed, new_value=0 if stub else issuance_values[idx],
           rate_value=rate_values[idx], term_value=debt_term_quarters,
           extra_value=debt_extra[idx], interest_kind="average_balance")

    # ---- capital lease
    _block(C_L_OPEN, C_L_ADD, C_L_RATE, C_L_TERM, C_L_SCHED, C_L_EXTRA,
           C_L_PRIN, C_L_INT, C_L_PAY, C_L_CLOSE,
           seed=lease_seed, new_value=0 if stub else lease_add_values[idx],
           rate_value=rate_values[idx], term_value=lease_life_quarters,
           extra_value=lease_extra[idx], interest_kind="opening_balance")

    # ---- combined
    ws.cell(r, C_T_PAY, value=f"={local_ref(r, C_D_PAY)}+{local_ref(r, C_L_PAY)}")
    ws.cell(r, C_T_INT, value=f"={local_ref(r, C_D_INT)}+{local_ref(r, C_L_INT)}")
    ws.cell(r, C_T_CLOSE, value=f"={local_ref(r, C_D_CLOSE)}+{local_ref(r, C_L_CLOSE)}")

    for col, _text, fmt in headers[1:]:
      cell = ws.cell(r, col)
      if col in INPUT_COLS:
        set_input_style(cell, number_format=fmt)
      else:
        set_formula_style(cell, number_format=fmt)
    if stub:
      for col, _t, _f in headers:
        ws.cell(r, col).font = design.font("note")

  last_row = FIRST_ROW + PERIOD_COUNT - 1

  # PRINCIPAL COMPUTED, NOT PASTED - enforced at build time (CW-043). The
  # first vertical build pasted principal literals; a zero authored upstream
  # then froze a $62,000 lease for five years and every tie-out reconciled
  # the frozen number with itself. The input whitelist is exactly the amber
  # columns plus the stub row's opening seeds (balance-sheet anchors, stated
  # facts); a bare value anywhere else in the table kills the build.
  for r_ in range(FIRST_ROW, last_row + 1):
    for col, _text, _fmt in headers[1:]:
      if col in INPUT_COLS:
        continue
      if r_ == FIRST_ROW and col in (C_D_OPEN, C_L_OPEN):
        continue
      v = ws.cell(r_, col).value
      if not (isinstance(v, str) and v.startswith("=")):
        raise AssertionError(
          f"Debt Schedule literal where a formula belongs: "
          f"{get_column_letter(col)}{r_} = {v!r}")

  # ---- THE HIDDEN BRIDGE --------------------------------------------------
  # One row per figure the rest of the workbook consumes, 21 columns wide, in
  # the standard horizontal geometry, each cell pointing straight back into
  # the table above. Registered under the keys this sheet has always
  # published, so nothing downstream changes.
  bridge_top = last_row + 3
  ws.cell(bridge_top - 1, 1,
          value="Feed to Model Inputs, FINMO, CapEx and Checks - hidden by "
                "design; edit the schedule above, not these rows"
          ).font = design.font("note")

  #: (ctx key, column in the table above)
  DIRECT = [
    ("Opening Debt", C_D_OPEN),
    ("Debt Issuance", C_D_NEW),
    ("Interest Rate per quarter", C_D_RATE),
    ("Total Debt Service", C_D_PAY),
    ("Interest Expense", C_D_INT),
    ("Actual Debt Repayment", C_D_PRIN),
    ("Closing Debt", C_D_CLOSE),
    ("Lease Opening Balance", C_L_OPEN),
    ("Lease Net Additions", C_L_ADD),
    ("Lease Life in quarters", C_L_TERM),
    ("Lease Interest Expense", C_L_INT),
    ("Lease Principal Repayments", C_L_PRIN),
    ("Lease Closing Balance", C_L_CLOSE),
    ("Lease Additions (asset)", C_L_ADD),
  ]
  row = bridge_top
  lease_add_bridge_row = 0
  for key, col in DIRECT:
    if key == "Lease Net Additions":
      lease_add_bridge_row = row
    ws.cell(row, 1, value=key).font = design.font("note")
    for idx in range(PERIOD_COUNT):
      ws.cell(row, PERIOD_START_COL + idx,
              value=f"={local_ref(FIRST_ROW + idx, col)}")
    ctx.add_schedule_row(DEBT_SHEET, key, row)
    if key == "Lease Additions (asset)":
      rou_add_row = row
    if key == "Lease Life in quarters":
      lease_term_bridge_row = row
    row += 1

  # ONE INPUT, ONE HOP. The asset side reads the LIABILITY's additions row
  # rather than reaching into the table a second time, so there is exactly one
  # cell a client types a lease into and both sides are downstream of it.
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    ws.cell(rou_add_row, col, value=f"={local_ref(lease_add_bridge_row, col)}")

  # The right-of-use asset: off the schedule, computed here from the lease
  # columns so the lease TERM a client edits still drives it. Straight line on
  # ORIGINAL COST over the term, each tranche from the quarter it is signed.
  rou_open_row, rou_dep_row, rou_close_row = row, row + 1, row + 2
  for key, r_ in (("Right-of-Use Asset Opening", rou_open_row),
                  ("Lease Asset Depreciation", rou_dep_row),
                  ("Right-of-Use Asset Closing", rou_close_row)):
    ws.cell(r_, 1, value=key).font = design.font("note")
    ctx.add_schedule_row(DEBT_SHEET, key, r_)
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    prev_col = col - 1
    add = local_ref(rou_add_row, col)
    term = local_ref(lease_term_bridge_row, col)
    ws.cell(rou_open_row, col,
            value=lease_seed if idx == 0 else f"={local_ref(rou_close_row, prev_col)}")
    if idx == 0:
      ws.cell(rou_dep_row, col, value=0)
    else:
      # FIRST_LIVE_COL, not the stub. The stub column is hidden and R50 bans a
      # scalar that pins it across every live quarter - and the two cells hold
      # the same original cost, since nothing moves in the stub period.
      seed_ref = f"${get_column_letter(FIRST_LIVE_COL)}${rou_open_row}"
      first_add = f"${get_column_letter(FIRST_LIVE_COL)}${rou_add_row}"
      window = f"{first_add}:{local_ref(rou_add_row, col)}"
      cost = (f"IF({idx}<={term},{seed_ref},0)"
              f"+SUMPRODUCT((COLUMN({window})>COLUMN({local_ref(rou_add_row, col)})-{term})*{window})")
      cap = f"{local_ref(rou_open_row, col)}+{add}"
      ws.cell(rou_dep_row, col, value=f"=MIN(({cost})/{term},{cap})")
    _rou_left = f"{local_ref(rou_open_row, col)}+{add}-{local_ref(rou_dep_row, col)}"
    ws.cell(rou_close_row, col, value=f"=IF({_rou_left}<{SUB_CENT},0,{_rou_left})")

  for r_ in range(bridge_top - 1, rou_close_row + 1):
    ws.row_dimensions[r_].hidden = True


def build_capex_depreciation_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, CAPEX_SHEET)
  apply_base_style(ws)
  # CW-032 #266: the depreciation assumption is STATED where the client
  # reads the schedule - no intake question asks equipment age, so the
  # policy must be visible rather than implied.
  _life_years = number(data.schedules.get("useful_life_years")) or 5
  set_title(
    ws,
    "CapEx & Depreciation Schedule",
    "Capital expenditure and depreciation mechanics. Model Inputs links "
    "to calculated outputs. Existing equipment and new capital spending "
    f"are written down over about {_life_years:g} years, at a rate fitted "
    "each quarter to keep the expense level - see the note below the "
    "schedule.",
  )
  write_period_headers(ws, data.periods)
  schedule_by_label = row_by_label(data.schedule_rows)
  expense_by_label = row_by_label(data.expense_rows)
  schedules = data.schedules
  # ADJUSTABLE ROWS FIRST, RESULTS BELOW - the same shape as the Debt
  # Schedule. Capital Expenditures sat at row 8 and Depreciation Rate at row
  # 10, with calculated rows above, between and below them, so nothing on the
  # sheet said which two numbers were the client's.
  capex_inputs = [
    ("Capital Expenditures", "What you spend on equipment this quarter",
     CURRENCY_FORMAT, ANNUAL_SUM),
    # A CLIENT INPUT, amber, under "What you can change" (Nick, 2026-08-25,
    # having been through the sheet: the mechanics are correct and this is a
    # lever like any other). It was de-ambered on 08-21 as a fitted solver
    # artifact; that ruling is reversed here. The note below the schedule
    # still says what the rate is and what flattening it does.
    ("Depreciation Rate", "Derived - fitted each quarter, see the note below",
     PERCENT_FORMAT, ANNUAL_AVERAGE),
  ]
  capex_outputs = [
    ("Opening PPE", "What the equipment was worth at the start",
     CURRENCY_FORMAT, ANNUAL_YEAR_START),
    ("Depreciation Expense", "Opening balance x the rate above",
     CURRENCY_FORMAT, ANNUAL_SUM),
    ("Closing PPE", "What it is worth at the end, after depreciation",
     CURRENCY_FORMAT, ANNUAL_YEAR_END),
    ("Opening Accumulated Depreciation", "Written off before this quarter",
     CURRENCY_FORMAT, ANNUAL_YEAR_START),
    ("Accumulated Depreciation", "Written off in total, to date",
     CURRENCY_FORMAT, ANNUAL_YEAR_END),
    ("Lease Additions", "Memo - leased assets are carried as the Right-of-Use "
                        "Asset on the Debt Schedule, not in PPE",
     CURRENCY_FORMAT, ANNUAL_SUM),
  ]
  labels = [(l, f, m) for l, _d, f, m in capex_inputs + capex_outputs]
  row = 6
  write_section_header(ws, row, "What you can change")
  row += 1
  for label, detail, _fmt, _mode in capex_inputs:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=detail)
    ctx.add_schedule_row(CAPEX_SHEET, label, row)
    row += 1
  row += 1
  write_section_header(ws, row, "What that produces")
  row += 1
  for label, detail, _fmt, _mode in capex_outputs:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=detail)
    ctx.add_schedule_row(CAPEX_SHEET, label, row)
    row += 1
  capex_values = values_21((schedule_by_label.get("Capital Expenditures") or {}).get("values"))
  dep_rate_values = values_21((expense_by_label.get("Depreciation") or {}).get("values"))
  opening = ctx.schedule_row(CAPEX_SHEET, "Opening PPE")
  capex = ctx.schedule_row(CAPEX_SHEET, "Capital Expenditures")
  lease_add = ctx.schedule_row(CAPEX_SHEET, "Lease Additions")
  dep_rate = ctx.schedule_row(CAPEX_SHEET, "Depreciation Rate")
  dep_exp = ctx.schedule_row(CAPEX_SHEET, "Depreciation Expense")
  closing = ctx.schedule_row(CAPEX_SHEET, "Closing PPE")
  opening_acc = ctx.schedule_row(CAPEX_SHEET, "Opening Accumulated Depreciation")
  acc = ctx.schedule_row(CAPEX_SHEET, "Accumulated Depreciation")
  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    ws.cell(opening, col, value=number(schedules.get("forecast_ppe_opening_balance_seed") or schedules.get("ppe_opening_balance_seed")) if idx == 0 else f"={local_ref(closing, col - 1)}")
    ws.cell(capex, col, value=0 if idx == 0 else capex_values[idx])
    ws.cell(
      lease_add,
      col,
      value=0 if idx == 0 else f"={ref(DEBT_SHEET, ctx.schedule_row(DEBT_SHEET, 'Lease Net Additions'), col)}",
    )
    ws.cell(dep_rate, col, value=dep_rate_values[idx])
    ws.cell(dep_exp, col, value=f"=MIN({local_ref(opening, col)}*{local_ref(dep_rate, col)},{local_ref(opening, col)})")
    # LEASED ASSETS ARE NOT IN PPE. This used to add lease additions here AND
    # carry them again as the Right-of-Use Asset, so one lease landed on the
    # asset side twice while the liability counted it once - the balance sheet
    # came out over by the full amount of the lease. MEASURED: a single 40,000
    # lease moved PPE +37,688 and the ROU asset +26,000 against a liability of
    # +40,000, and Checks "Balance Sheet balances Q20" read FAIL by exactly
    # 40,000. It predates the corkscrew rebuild - the same gap reproduces at
    # aa014c7 - and it needs a lease ADDITION to appear, which no sampled
    # draft has (0 of 150), so nothing delivered carried it.
    #
    # The row stays as a memo and keeps its Checks tie-out to the Debt
    # Schedule; it just no longer double-counts into owned PPE.
    _ppe_left = f"{local_ref(opening, col)}+{local_ref(capex, col)}-{local_ref(dep_exp, col)}"
    ws.cell(closing, col, value=f"=IF({_ppe_left}<{SUB_CENT},0,{_ppe_left})")
    ws.cell(opening_acc, col, value=number(schedules.get("accumulated_depreciation_opening_seed")) if idx == 0 else f"={local_ref(acc, col - 1)}")
    ws.cell(acc, col, value=f"={local_ref(opening_acc, col)}-{local_ref(dep_exp, col)}")
  for label, fmt, annual_mode in labels:
    r = ctx.schedule_row(CAPEX_SHEET, label)
    for col in range(PERIOD_START_COL, PERIOD_END_COL + 1):
      if label in {"Capital Expenditures", "Depreciation Rate"}:
        set_input_style(ws.cell(r, col), number_format=fmt)
      elif label == "Lease Additions":
        set_formula_style(ws.cell(r, col), number_format=fmt, internal_link=True)
      else:
        set_formula_style(ws.cell(r, col), number_format=fmt)
    add_annual_formulas(ws, r, mode=annual_mode, number_format=fmt)
    # Capital Expenditures is OUT of the emphasis set. style_row repaints the
    # row, so the FILL_GREEN band was overwriting the amber that
    # set_input_style had just applied - which left this sheet with the fitted
    # Depreciation Rate as its ONLY amber row, and the one thing a client can
    # actually change painted like a subtotal. Exactly backwards.
    style_row(ws, r, fill=FILL_GREEN if label in {"Depreciation Expense", "Closing PPE"} else None, bold=label in {"Depreciation Expense", "Closing PPE"}, number_format=fmt)

  # WHAT THE RATE ACTUALLY IS. The row above is editable and amber, and its
  # values climb - measured on three drafts, 5.1% to 54.4% across the twenty
  # quarters. That is not a rate anyone can reason about, and it is not an
  # assumption: the engine FITS it quarter by quarter so that the expense
  # comes out level (759, 768, 777 ... 859 on the fixture) while the balance
  # it applies to keeps falling. The straight line is in the RESULT, not in
  # the rate. A client who "corrects" this row to a flat 5% would change the
  # whole schedule and be reasonable to think they were fixing a typo.
  #
  # Said plainly on the sheet rather than fixed here, because replacing the
  # mechanism with a useful-life row - the shape the lease block now uses -
  # changes delivered numbers, and that is Nick's call, not a tidy-up.
  row += 1
  ws.cell(row=row, column=1, value="About the depreciation rate")
  ws.cell(row=row, column=2, value="Read this before changing it").font = design.font("note")
  note = ws.cell(
    row=row, column=PERIOD_START_COL,
    value=("This rate is fitted quarter by quarter so the depreciation expense "
           "stays level as the equipment balance falls - that is why it rises "
           "over time. It is not an annual rate, and flattening it will change "
           "every figure below."))
  note.font = design.font("note")
  ctx.add_schedule_row(CAPEX_SHEET, "CapEx note", row)

  hide_stub_column(ws)


def build_working_capital_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, WORKING_CAPITAL_SHEET)
  apply_base_style(ws)
  set_title(ws, "Working Capital Schedule", "Working-capital drivers. Model Inputs links to these source rows.")
  write_period_headers(ws, data.periods)
  row = 6
  write_section_header(ws, row, "Working Capital Drivers")
  row += 1
  balance_by_label = row_by_label(data.balance_sheet_rows)
  labels = [
    "Accounts Receivable Days",
    "Inventory Days",
    "Accounts Payable Days",
    "Prepaid Expenses (% of Revenue)",
    "Deferred Revenue (% of Revenue)",
    "Short Term Debt (% of LTD)",
  ]
  for label in labels:
    source = balance_by_label.get(label, {"values": []})
    fmt = _fmt_for_row(source)
    write_values_row(ws, row, label, values_21(source.get("values")), detail="Working capital driver", number_format=fmt)
    ctx.add_schedule_row(WORKING_CAPITAL_SHEET, label, row)
    style_row(ws, row, number_format=fmt)
    row += 1
  row += 1
  write_section_header(ws, row, "Opening Working Capital Seeds")
  row += 1
  seed_rows = [
    ("Cash Opening Balance", "cash_opening_balance_seed"),
    ("Accounts Receivable Opening Balance", "accounts_receivable_opening_balance_seed"),
    ("Inventory Opening Balance", "inventory_opening_balance_seed"),
    ("Accounts Payable Opening Balance", "accounts_payable_opening_balance_seed"),
    ("Short Term Debt Opening Balance", "short_term_debt_opening_balance_seed"),
  ]
  for label, key in seed_rows:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value="Opening balance seed")
    ws.cell(row=row, column=PERIOD_START_COL, value=number(data.schedules.get(key)))
    set_input_style(ws.cell(row=row, column=PERIOD_START_COL), number_format=CURRENCY_FORMAT)
    for col in range(FIRST_LIVE_COL, LAST_LIVE_COL + 1):
      ws.cell(row=row, column=col, value=0)
      set_formula_style(ws.cell(row=row, column=col), number_format=CURRENCY_FORMAT)
    ctx.add_schedule_row(WORKING_CAPITAL_SHEET, label, row)
    style_row(ws, row, fill=FILL_LIGHT, number_format=CURRENCY_FORMAT)
    row += 1


def build_cash_equity_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, CASH_EQUITY_SHEET)
  apply_base_style(ws)
  set_title(ws, "Cash & Equity Schedule",
            "Equity balances and distributions. Model Inputs links to these rows. "
            "Owner's Capital and Other Equity are BALANCES: type a new level into a "
            "quarter and it carries to every quarter after it - any contributions "
            "your plan already makes in later quarters are added on top of the level "
            "you set. Distributions are paid in the quarter you enter them.")
  write_period_headers(ws, data.periods)
  balance_by_label = row_by_label(data.balance_sheet_rows)
  row = 6
  write_section_header(ws, row, "Equity and Distributions")
  row += 1
  # OWNER'S CAPITAL AND OTHER EQUITY ARE CHAINED (Nick, 2026-08-25). They
  # were 21 independent literals that happened to hold the same number, so a
  # client typing 250,000 into Q8 changed one cell: Q9 fell back to its own
  # literal and the balance sheet read a jump and a reversal. These rows are
  # BALANCES (FINMO reads the level, r39, and derives the equity FLOW as the
  # change in it, r56), so each quarter references the quarter before it and
  # one edit flows all the way right. The stub seeds the chain.
  #
  # WHERE THE ENGINE'S OWN SERIES STEPS (the cash pass injects equity - 652
  # of 2,006 stored drafts step at Q2 or later), the step quarter is written
  # as ROUND(previous + the authored contribution, 6): the contribution rides
  # on top of whatever level is in force, so a client's edit before a step
  # still carries all the way to Q20 and the plan's own injections still land.
  # A step written as a LEVEL literal (the first version) snapped the client's
  # edit back to the engine's number at the step - the very reversal this
  # chain exists to remove (mini, 2026-08-25, Third Coast Q10 -> Q11). The
  # ROUND to the engine's 6-dp grid is the same guard as the Debt Schedule's
  # payoff fix: previous + delta reproduces the authored level only to an ulp
  # in Excel, and the grid lands it exactly. The stub is the one literal.
  #
  # DISTRIBUTIONS ARE NOT CHAINED. They are a per-quarter FLOW - FINMO r57
  # pays them out in the quarter they appear and retained earnings subtracts
  # them there - so a Q8 distribution is a Q8 distribution, not one that
  # repeats forever.
  for label in ["Owner's Capital", "Other Equity", "Distributions"]:
    source = balance_by_label.get(label, {"values": []})
    values = values_21(source.get("values"))
    if label == "Distributions":
      write_values_row(ws, row, label, values, detail="Paid out in the quarter you enter it",
                       number_format=CURRENCY_FORMAT)
    else:
      ws.cell(row=row, column=1, value=label)
      ws.cell(row=row, column=2,
              value="Balance - a new level carries forward; planned contributions add on top")
      for idx in range(PERIOD_COUNT):
        col = PERIOD_START_COL + idx
        value = number(values[idx]) or 0.0
        if idx == 0:
          formula = value  # the stub seeds the chain
        else:
          prev_ref = local_ref(row, col - 1)
          delta = value - (number(values[idx - 1]) or 0.0)
          if abs(delta) <= 1e-9:
            formula = f"={prev_ref}"
          else:
            sign = "-" if delta < 0 else "+"
            formula = f"=ROUND({prev_ref}{sign}{abs(delta)!r},6)"
        cell = ws.cell(row=row, column=col, value=formula)
        set_input_style(cell, number_format=CURRENCY_FORMAT)
    ctx.add_schedule_row(CASH_EQUITY_SHEET, label, row)
    style_row(ws, row, number_format=CURRENCY_FORMAT)
    row += 1
