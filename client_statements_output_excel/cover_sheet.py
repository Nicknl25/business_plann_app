"""X1 — the Cover sheet (2026-08-18).

The first thing that opens. Deliberately **text only**: no formulas anywhere, so
the sheet never enters the R32 formula grid (a formula-bearing cover would move
the digest — docs/WORKBOOK_ANALYTICS_RESEARCH.md §1.6) and never carries a
number that could disagree with the model.

Hyperlinks are cell attributes, not formulas, so the contents list is free.
The logo placeholder stays a styled empty box until X6 (Pillow).
"""
from __future__ import annotations

from datetime import date, datetime
from typing import Any, Optional

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from . import design
from .data import DraftWorkbookData, text

COVER_SHEET = "Cover"

_CONTENTS = [
  ("Dashboard", "Headline numbers and the charts that carry them"),
  ("FINMO", "The three statements, quarterly and annual, with break-even"),
  ("Revenue Drivers", "Capacity, price and utilisation per line of business"),
  ("Payroll Schedule", "Roles, headcount and payroll by quarter"),
  ("Debt Schedule", "Debt and capital-lease amortisation"),
  ("CapEx Depreciation", "Capital expenditure and the depreciation schedule"),
  ("Working Capital", "Receivable, inventory and payable assumptions"),
  ("Cash Equity Schedule", "Owner capital, other equity and distributions"),
  ("Model Inputs", "Every driver the statements are built from"),
  ("Checks", "The model's own tie-outs and status"),
  ("Diagnostics", "How this model run was produced"),
]

_LEGEND = [
  (design.INPUT_FILL, "Amber cells are inputs — change them and the whole model recalculates"),
  (design.WHITE, "Plain cells are calculated from the inputs"),
  (design.TINT_1, "Shaded rows are subtotals and headline results"),
]


def _meta(row: Any, key: str) -> str:
  try:
    return text(row.get(key))
  except Exception:
    return ""


def _pretty_date(value: Any) -> str:
  if isinstance(value, (datetime, date)):
    return value.strftime("%d %B %Y")
  cleaned = text(value)
  return cleaned[:10] if cleaned else date.today().strftime("%d %B %Y")


def build_cover_sheet(wb, data: DraftWorkbookData) -> None:
  ws = wb.create_sheet(COVER_SHEET)
  ws.sheet_view.showGridLines = False
  ws.sheet_view.zoomScale = 100
  ws.column_dimensions["A"].width = 4
  ws.column_dimensions["B"].width = 46
  for col in range(3, 11):
    ws.column_dimensions[get_column_letter(col)].width = 15
  for row in range(1, 9):
    ws.row_dimensions[row].height = 26

  row_meta = data.draft_row or {}
  business = data.business_name or _meta(row_meta, "business_name") or "Business"

  # --- the band -------------------------------------------------------------
  for row in range(1, 9):
    for col in range(1, 11):
      ws.cell(row=row, column=col).fill = design.fill(design.NAVY_DEEP)
  ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=9)
  title = ws.cell(row=3, column=2, value=business)
  title.font = design.font("cover_title")
  title.alignment = Alignment(horizontal="left", vertical="center")
  ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=9)
  sub = ws.cell(row=5, column=2, value="Financial Model & Analysis")
  sub.font = design.font("cover_sub")
  sub.alignment = Alignment(horizontal="left", vertical="center")

  # --- meta block -----------------------------------------------------------
  address = " ".join(x for x in [
    _meta(row_meta, "address_city"), _meta(row_meta, "address_state"),
  ] if x) or _meta(row_meta, "business_address")
  meta_rows = [
    ("Prepared for", business),
    ("Location", address),
    ("Prepared on", _pretty_date(row_meta.get("planning_run_completed_at"))),
    ("Model version", "Five-year quarterly model — 20 quarters"),
    ("Model run", (_meta(row_meta, "planning_run_id") or "")[:12]),
  ]
  row = 10
  for label, value in meta_rows:
    if not value:
      continue
    lab = ws.cell(row=row, column=2, value=label.upper())
    lab.font = design.font("cover_meta_label")
    val = ws.cell(row=row, column=3, value=value)
    val.font = design.font("cover_meta_value")
    val.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    row += 1

  # --- logo placeholder (X6 replaces this with the client's mark) -----------
  ws.merge_cells(start_row=10, start_column=8, end_row=13, end_column=10)
  logo = ws.cell(row=10, column=8, value="[ logo ]")
  logo.font = design.font("footnote")
  logo.alignment = Alignment(horizontal="center", vertical="center")
  logo.fill = design.fill(design.TINT_2)

  # --- contents -------------------------------------------------------------
  row += 2
  design.section_band(ws, row, "What's in this workbook", end_col=10)
  row += 1
  for name, blurb in _CONTENTS:
    if name not in wb.sheetnames:
      continue
    cell = ws.cell(row=row, column=2, value=name)
    cell.font = design.font("label_strong")
    cell.hyperlink = f"#'{name}'!A1"
    cell.alignment = Alignment(horizontal="left")
    note = ws.cell(row=row, column=3, value=blurb)
    note.font = design.font("note")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    row += 1

  # --- how to read it -------------------------------------------------------
  row += 1
  design.section_band(ws, row, "How to read this model", end_col=10)
  row += 1
  for color, explanation in _LEGEND:
    swatch = ws.cell(row=row, column=2, value="")
    swatch.fill = design.fill(color)
    swatch.border = design.BORDER_HAIRLINE
    note = ws.cell(row=row, column=3, value=explanation)
    note.font = design.font("label")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    row += 1

  row += 1
  design.footnote(ws, row, "Confidential. Prepared from the figures supplied by the business "
                           "owner and the industry data cited in the model.", col=2)
  ws.freeze_panes = "A9"
  design.page_setup(ws, landscape=False, fit_width=True, footer=f"{business} — Financial Model")
