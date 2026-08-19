from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import design
from .data import PERIOD_COUNT


MODEL_INPUT_SHEET = "Model Inputs"
FINMO_SHEET = "FINMO"
CHECKS_SHEET = "Checks"
SOURCE_SHEET = "Audit Source"
REVENUE_SHEET = "Revenue Drivers"
PAYROLL_SHEET = "Payroll Schedule"
DEBT_SHEET = "Debt Schedule"
CAPEX_SHEET = "CapEx Depreciation"
WORKING_CAPITAL_SHEET = "Working Capital"
CASH_EQUITY_SHEET = "Cash Equity Schedule"
DASHBOARD_SHEET = "Dashboard"  # W2 (2026-08-18)

PERIOD_START_COL = 3
PERIOD_END_COL = PERIOD_START_COL + PERIOD_COUNT - 1
FIRST_LIVE_COL = PERIOD_START_COL + 1
LAST_LIVE_COL = PERIOD_END_COL
ANNUAL_START_COL = PERIOD_END_COL + 1

# X1 DESIGN SYSTEM (2026-08-18): these names are kept so sheet modules keep
# compiling, but every value now comes from design.py. Add no new color or
# format here — add a ROLE in design.py.
FILL_NAVY = design.NAVY
FILL_BLUE = design.TINT_1          # subtotal / emphasis band
FILL_LIGHT = design.TINT_2         # supporting band
FILL_GREEN = design.TINT_1         # schedule "answer" row -> emphasis band
FILL_YELLOW = design.INPUT_FILL    # client input
FILL_GRAY = design.TINT_2
FONT_BLUE = design.INPUT_INK
FONT_GREEN = design.INK            # link-vs-local color split retired (see design.py)
FONT_BLACK = design.INK
FONT_WHITE = design.WHITE

THIN_GRAY = Side(style="thin", color=design.HAIRLINE)
MEDIUM_BLUE = Side(style="thin", color=design.RULE)

CURRENCY_FORMAT = design.FMT_MONEY
NUMBER_FORMAT = design.FMT_UNITS
INTEGER_FORMAT = design.FMT_INTEGER
PERCENT_FORMAT = design.FMT_PERCENT
DATE_FORMAT = design.FMT_DATE


@dataclass
class WorkbookBuildContext:
  period_cols: Dict[int, int] = field(default_factory=dict)
  annual_cols: Dict[int, int] = field(default_factory=dict)
  schedule_rows: Dict[str, Dict[str, int]] = field(default_factory=dict)
  model_input_rows: Dict[str, int] = field(default_factory=dict)
  finmo_rows: Dict[str, Dict[str, int]] = field(default_factory=dict)
  source_rows: Dict[str, Dict[str, int]] = field(default_factory=dict)

  def add_schedule_row(self, sheet: str, key: str, row: int) -> None:
    self.schedule_rows.setdefault(sheet, {})[key] = row

  def schedule_row(self, sheet: str, key: str) -> int:
    return self.schedule_rows.get(sheet, {}).get(key, 0)

  def add_model_input_row(self, key: str, row: int) -> None:
    self.model_input_rows[key] = row

  def model_input_row(self, key: str) -> int:
    return self.model_input_rows.get(key, 0)

  def add_finmo_row(self, statement: str, key: str, row: int) -> None:
    self.finmo_rows.setdefault(statement, {})[key] = row

  def finmo_row(self, statement: str, key: str) -> int:
    return self.finmo_rows.get(statement, {}).get(key, 0)

  def add_source_row(self, statement: str, key: str, row: int) -> None:
    self.source_rows.setdefault(statement, {})[key] = row

  def source_row(self, statement: str, key: str) -> int:
    return self.source_rows.get(statement, {}).get(key, 0)


def qsheet(name: str) -> str:
  return "'" + name.replace("'", "''") + "'"


def ref(sheet: str, row: int, col: int, *, abs_ref: bool = False) -> str:
  prefix = "$" if abs_ref else ""
  return f"{qsheet(sheet)}!{prefix}{get_column_letter(col)}{prefix}{row}"


def local_ref(row: int, col: int, *, abs_ref: bool = False) -> str:
  prefix = "$" if abs_ref else ""
  return f"{prefix}{get_column_letter(col)}{prefix}{row}"


def range_ref(sheet: str, row: int, start_col: int, end_col: int) -> str:
  return f"{qsheet(sheet)}!{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}"


def create_workbook() -> Workbook:
  wb = Workbook()
  default = wb.active
  wb.remove(default)
  try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
  except Exception:
    pass
  return wb


def create_sheet(wb: Workbook, title: str) -> Worksheet:
  ws = wb.create_sheet(title)
  ws.sheet_view.showGridLines = False
  return ws


def apply_base_style(ws: Worksheet) -> None:
  ws.freeze_panes = "C6"
  ws.sheet_view.showGridLines = False
  ws.sheet_view.zoomScale = 100
  design.base_columns(
    ws,
    period_start=PERIOD_START_COL,
    period_end=PERIOD_END_COL,
    annual_end=ANNUAL_START_COL + 4,
  )
  design.page_setup(ws, title_rows="4:5")


def set_title(ws: Worksheet, title: str, subtitle: str = "") -> None:
  design.title_block(ws, title, subtitle)


def write_period_headers(ws: Worksheet, periods: List[dict], *, row: int = 4, include_annual: bool = True) -> None:
  labels = ["Stub"] + [f"Q{i}" for i in range(1, PERIOD_COUNT)]
  for index, label in enumerate(labels):
    col = PERIOD_START_COL + index
    cell = ws.cell(row=row, column=col, value=label)
    cell.fill = design.fill(design.NAVY)
    cell.font = design.font("colhead")
    cell.alignment = Alignment(horizontal="center")
    period = periods[index] if index < len(periods) else {}
    date_value = period.get("date") or ""
    parsed_date = None
    if isinstance(date_value, str) and date_value:
      try:
        parsed_date = datetime.strptime(date_value[:10], "%Y-%m-%d").date()
      except Exception:
        parsed_date = None
    date_cell = ws.cell(row=row + 1, column=col, value=parsed_date or str(date_value or ""))
    if parsed_date:
      date_cell.number_format = DATE_FORMAT
    date_cell.font = design.font("colhead_sub")
    date_cell.alignment = Alignment(horizontal="center")
  if include_annual:
    for year in range(1, 6):
      col = ANNUAL_START_COL + year - 1
      cell = ws.cell(row=row, column=col, value=f"Y{year}")
      cell.fill = design.fill(design.NAVY_DEEP)
      cell.font = design.font("colhead")
      cell.alignment = Alignment(horizontal="center")
      ws.cell(row=row + 1, column=col, value="Annual").font = design.font("colhead_sub")


def write_section_header(ws: Worksheet, row: int, title: str, *, end_col: int = ANNUAL_START_COL + 4) -> None:
  design.section_band(ws, row, title, end_col=end_col)


def style_row(
  ws: Worksheet,
  row: int,
  *,
  start_col: int = 1,
  end_col: int = ANNUAL_START_COL + 4,
  fill: Optional[str] = None,
  bold: bool = False,
  font_color: str = FONT_BLACK,
  number_format: Optional[str] = None,
  border_top: bool = False,
) -> None:
  """X1: delegates to the design system. The legacy `fill` argument is read as
  a ROLE — the emphasis tints mean subtotal, the light tint means a supporting
  band — so callers keep working while the palette lives in one place."""
  emphasis = bool(bold) or fill in {FILL_BLUE, FILL_GREEN}
  band = fill == FILL_LIGHT
  design.data_row(
    ws,
    row,
    start_col=start_col,
    end_col=end_col,
    number_format=number_format or design.FMT_MONEY,
    emphasis=emphasis,
    band=band,
    top_rule=border_top,
  )


def set_input_style(cell: Cell, *, number_format: Optional[str] = None) -> None:
  design.input_cell(cell, number_format=number_format or design.FMT_MONEY)


def set_formula_style(cell: Cell, *, number_format: Optional[str] = None, internal_link: bool = False) -> None:
  # X1: the green "cross-sheet link" convention is retired — the reader needs
  # editable-vs-calculated, not link provenance (design.py header).
  design.calculated_cell(cell, number_format=number_format or design.FMT_MONEY)


def write_values_row(
  ws: Worksheet,
  row: int,
  label: str,
  values: Iterable[float],
  *,
  detail: str = "",
  number_format: str = CURRENCY_FORMAT,
  input_style: bool = True,
  include_stub: bool = True,
) -> None:
  ws.cell(row=row, column=1, value=label)
  ws.cell(row=row, column=2, value=detail)
  value_list = list(values)
  if include_stub and len(value_list) == PERIOD_COUNT:
    start_index = 0
  elif include_stub:
    value_list = [0.0] + value_list
    start_index = 0
  else:
    start_index = 1
  for idx, value in enumerate(value_list[start_index:], start=start_index):
    col = PERIOD_START_COL + idx
    cell = ws.cell(row=row, column=col, value=value)
    if input_style:
      set_input_style(cell, number_format=number_format)
    else:
      set_formula_style(cell, number_format=number_format)


# ---------------------------------------------------------------------------
# ANNUAL AGGREGATION — routed by what the row MEANS (2026-08-19).
#
# The annual columns used to have two modes, SUM or last-quarter, and every
# caller chose with a boolean. A RATE row therefore got SUMMED, and a client
# opening the workbook saw a unit price of $2,599 for a $640 service, a 247%
# utilisation and a 107% tax rate. The mode is now a SEMANTIC property of the
# row, resolved from its label and number format, so a rate CANNOT be summed by
# construction - `mode` is required and there is no boolean to get backwards.
# ---------------------------------------------------------------------------

ANNUAL_SUM = "sum"                # a FLOW: revenue, costs, cash movements
ANNUAL_AVERAGE = "average"        # a RATE or LEVEL: %, price, utilisation, days
ANNUAL_ANNUALIZE = "annualize"    # a PER-QUARTER rate quoted annually (cost of debt)
ANNUAL_YEAR_END = "year_end"      # a BALANCE at the close of the year
ANNUAL_YEAR_START = "year_start"  # a BALANCE at the open of the year

_ANNUAL_MODES = {ANNUAL_SUM, ANNUAL_AVERAGE, ANNUAL_ANNUALIZE, ANNUAL_YEAR_END, ANNUAL_YEAR_START}

#: Rows whose annual figure is the rate quoted the way a reader expects it -
#: a cost of debt is an ANNUAL rate, even though the model carries it quarterly.
_ANNUALIZED_LABELS = {"Interest Rate"}

#: Labels that are a LEVEL or a RATE whatever their number format - a unit
#: price is money-formatted but four quarters of it do not add up to a year's
#: price. This is what put $2,599 in front of a client for a $640 service.
_LEVEL_HINTS = ("unit price", "price per", "utilization", "utilisation", "rate", "days", "% of")

#: Money-formatted rows that are BALANCES, not flows.
_BALANCE_HINTS = ("opening", "closing", "balance", "ppe", "accumulated depreciation",
                  "right-of-use asset", "owner's capital", "other equity")


def annual_mode_for(label: str, number_format: str) -> str:
  """The annual mode a row EARNS from its own meaning.

  Number format is the honest signal for rate-ness (a percent/ratio/days row is
  never a flow); the label decides balance-vs-flow among the money rows, and
  whether a balance is a year-START (an "Opening ..." row) or a year-END one.
  """
  text_label = (label or "").strip().lower()
  fmt = number_format or ""
  if label in _ANNUALIZED_LABELS:
    return ANNUAL_ANNUALIZE
  if fmt in {design.FMT_PERCENT, design.FMT_RATIO, design.FMT_DAYS}:
    return ANNUAL_AVERAGE
  if any(hint in text_label for hint in _LEVEL_HINTS):
    return ANNUAL_AVERAGE
  if fmt == design.FMT_UNITS and any(
      h in text_label for h in ("rate", "utilization", "utilisation", "price", "fte", "days")):
    return ANNUAL_AVERAGE
  # "Opening" anywhere, not just at the front: the row is "Debt Opening Balance"
  # on Model Inputs and "Opening Debt" on the schedule, and both are the balance
  # the year STARTS with.
  if "opening" in text_label:
    return ANNUAL_YEAR_START
  if any(hint in text_label for hint in _BALANCE_HINTS):
    return ANNUAL_YEAR_END
  return ANNUAL_SUM


def annual_formula_for_row(row: int, year_index: int, *, mode: str) -> str:
  if mode not in _ANNUAL_MODES:
    raise ValueError(f"unknown annual mode {mode!r}")
  start_col = FIRST_LIVE_COL + ((year_index - 1) * 4)
  end_col = start_col + 3
  span = f"{local_ref(row, start_col)}:{local_ref(row, end_col)}"
  if mode == ANNUAL_YEAR_END:
    return f"={local_ref(row, end_col)}"
  if mode == ANNUAL_YEAR_START:
    return f"={local_ref(row, start_col)}"
  if mode == ANNUAL_AVERAGE:
    return f"=IFERROR(AVERAGE({span}),0)"
  if mode == ANNUAL_ANNUALIZE:
    return f"=IFERROR(AVERAGE({span})*4,0)"
  return f"=SUM({span})"


def add_annual_formulas(
  ws: Worksheet,
  row: int,
  *,
  mode: Optional[str] = None,
  label: str = "",
  number_format: str = CURRENCY_FORMAT,
) -> None:
  """Write the Y1..Y5 columns for one row. Pass an explicit `mode`, or a `label`
  and let `annual_mode_for` resolve it - never a bare boolean."""
  resolved = mode or annual_mode_for(label, number_format)
  for year in range(1, 6):
    col = ANNUAL_START_COL + year - 1
    cell = ws.cell(row=row, column=col, value=annual_formula_for_row(row, year, mode=resolved))
    set_formula_style(cell, number_format=number_format)


def set_tab_colors(wb: Workbook) -> None:
  design.apply_tab_colors(wb)
