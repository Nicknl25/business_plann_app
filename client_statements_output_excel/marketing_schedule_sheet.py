"""The Marketing Schedule tab — where the marketing percentage is now produced.

THE CHAIN RUNS FORWARD (Nick's A1). It used to read the settled percentage from
Model Inputs and divide its way back to CAC, which meant the two levers moved
nothing: spend was revenue x percentage, fixed before retention entered, so
raising retention only moved CAC. Spend, new customers and CAC are a closed
loop — any two fix the third — so the loop is now closed the other way:

    customers       = units sold / purchases per customer   (units from Revenue Drivers)
    returning       = last quarter's customers x retention
    new customers   = customers - returning
    marketing spend = new customers x CAC                   <- CAC is HELD, not divided
    marketing %     = spend / revenue                       -> Model Inputs -> FINMO

CAC is seeded per quarter from the same back-derivation that used to compute it,
so the delivered file reproduces the agreed percentage to the cent. After that
the economics run the right way round: better retention means fewer customers to
buy, which means less spend, which means a lower percentage.

WHAT THAT COSTS, AND IT IS WORTH SAYING. The marketing percentage is no longer
structurally exact — it is the OUTPUT of three assumptions (retention, purchases
per customer, CAC) that happen to reproduce the agreed number on delivery. The
labels say so: the percentage and spend are no longer marked Exact, because
their exactness holds only at the moment of delivery.

NO CYCLE. This sheet reads Revenue Drivers, which holds literals; Model Inputs
reads this sheet; FINMO reads Model Inputs. The old ='Model Inputs'!D19 link is
CUT in the same change that adds the reverse one, or the workbook would carry a
circular reference.

THE FOUR CLASS RULES, and one changed shape under the reversal:
  R1  a quarter with no net acquisition CANNOT derive spend from new x CAC —
      that would zero out real marketing spend and change the plan. Those
      quarters hold the planned figure, and the note row says so.
  R2  zero marketing: CAC seeds at zero, so spend stays zero on its own.
  R3  pre-revenue: no revenue means no units sold, so the stub's customers are
      zero and Q1's new customers equal its customers. Still no special case.
  R4  no usable audience: no customer rows at all, and the percentage row keeps
      the settled literals so Model Inputs still has something to point at.
"""
from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from . import design
from .data import DraftWorkbookData, text
from .excel_utils import (
  MARKETING_SCHEDULE_SHEET,
  PERIOD_COUNT,
  PERIOD_START_COL,
  REVENUE_SHEET,
  WorkbookBuildContext,
  apply_base_style,
  create_sheet,
  set_formula_style,
  set_input_style,
  set_title,
  style_row,
  write_period_headers,
  write_section_header,
)

#: The key Model Inputs looks up to repoint its marketing driver row at this
#: sheet. If the sheet is not built the lookup misses and Model Inputs keeps its
#: literals — which is exactly the absent-tolerant behaviour we want.
MARKETING_PERCENT_ROW_KEY = "Marketing percent of revenue"

_NOTE_TEXT = {
  "no_marketing_spend": "No marketing spend in this plan",
  "no_net_acquisition": "No net new customers — spend held at the planned figure",
  "thin_acquisition_count": "Few new customers — read with care",
  "not_modelled": "Not modelled for this business",
}


def _ref(column_index: int, row: int) -> str:
  return f"{get_column_letter(column_index)}{row}"


def _driver_rows(ctx: WorkbookBuildContext) -> List[Tuple[int, int]]:
  """(capacity_row, utilisation_row) per product on Revenue Drivers.

  Units sold = capacity x utilisation, summed. The payload computes the SAME
  thing from the same drivers, which is what lets CAC be seeded from one and
  multiplied back by the other without losing a cent.
  """
  rows = ctx.schedule_rows.get(REVENUE_SHEET, {})
  pairs: List[Tuple[int, int]] = []
  for key in sorted(rows):
    if key.endswith("::Capacity"):
      stem = key[: -len("::Capacity")]
      utilisation = rows.get(f"{stem}::Utilization")
      if utilisation:
        pairs.append((rows[key], utilisation))
  return pairs


def build_marketing_schedule_sheet(
  wb, data: DraftWorkbookData, ctx: WorkbookBuildContext
) -> None:
  payload: Dict[str, Any] = data.marketing_schedule or {}
  if not payload or payload.get("status") != "ok":
    return
  periods = payload.get("periods") or []
  if not periods:
    return

  schedule_class = str(payload.get("schedule_class") or "")
  assumptions = payload.get("assumptions") or {}
  retention_meta = assumptions.get("retention") or {}
  repeat_meta = assumptions.get("repeat_units_per_customer") or {}
  context = payload.get("context") or {}
  noun = str(context.get("entity_noun") or "customers")
  singular = "customer" if " and " in noun else (noun[:-1] if noun.endswith("s") else noun)

  ws = create_sheet(wb, MARKETING_SCHEDULE_SHEET)
  apply_base_style(ws)
  set_title(
    ws, "Marketing Schedule",
    f"What drives your marketing spend. Change retention, purchases per "
    f"{singular} or the cost to win one, and your marketing budget follows.")
  write_period_headers(ws, data.periods)

  columns = [PERIOD_START_COL + i for i in range(min(len(periods), PERIOD_COUNT))]
  pairs = _driver_rows(ctx)
  revenue_src = ctx.schedule_row(REVENUE_SHEET, "Total Revenue")
  modelled = schedule_class == "audience_modelled" and bool(pairs) and bool(revenue_src)

  row = 6
  write_section_header(ws, row, "What you can change")
  row += 1

  def editable_row(label: str, key: str, values, number_format: str,
                   detail: str) -> int:
    """A full 21-column editable row, like every other driver in the workbook.

    These were two single cells anchored at $C$7 and $C$8 — one value applying
    to all twenty quarters, and anchored in the STUB column at that. A driver a
    client can only set once is not a schedule.
    """
    nonlocal row
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=detail).font = design.font("note")
    for index, column in enumerate(columns):
      value = values[index] if index < len(values) else None
      cell = ws.cell(row=row, column=column, value=value)
      set_input_style(cell, number_format=number_format)
    ctx.add_schedule_row(MARKETING_SCHEDULE_SHEET, key, row)
    written = row
    row += 1
    return written

  retention_row = editable_row(
    "Customer retention, quarter over quarter", "Retention",
    [retention_meta.get("retention_rate")] * len(columns), design.FMT_PERCENT,
    "ASSUMPTION — expert estimate from your business model, not a sourced figure")

  # PER QUARTER, not per year. The payload carries an annual rate and the old
  # formula divided by four inside itself; a hidden conversion in a quarterly
  # grid is exactly what produced a four-fold customer error the first time.
  # The label and the number now agree and no formula divides.
  repeat_row = editable_row(
    f"Purchases per {singular} per quarter", "Repeat units per customer",
    [repeat_meta.get("per_quarter")] * len(columns), design.FMT_UNITS,
    "ASSUMPTION — implied by your own plan volume, not a measured rate")

  cac_row: Optional[int] = None
  if modelled:
    cac_row = editable_row(
      f"Cost to acquire one {singular}", "Customer acquisition cost",
      [p.get("customer_acquisition_cost") for p in periods], design.FMT_MONEY,
      "ASSUMPTION — the softest number here, and the one your budget is built "
      "on. Derived from the marketing figure agreed in your plan, then held.")
  row += 1

  write_section_header(ws, row, "What that produces")
  row += 1

  def computed_row(label: str, key: str, formula_for, number_format: str,
                   detail: str, emphasis: bool = False) -> int:
    nonlocal row
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=detail).font = design.font("note")
    for index, column in enumerate(columns):
      cell = ws.cell(row=row, column=column, value=formula_for(index, column))
      set_formula_style(cell, number_format=number_format)
    ctx.add_schedule_row(MARKETING_SCHEDULE_SHEET, key, row)
    style_row(ws, row, number_format=number_format, bold=emphasis)
    written = row
    row += 1
    return written

  settled_spend = [p.get("marketing_dollars") or 0.0 for p in periods]
  settled_percent = [p.get("marketing_percent_of_revenue") or 0.0 for p in periods]

  if modelled:
    def units_expr(column: int) -> str:
      return "+".join(
        f"'{REVENUE_SHEET}'!{_ref(column, cap)}*'{REVENUE_SHEET}'!{_ref(column, util)}"
        for cap, util in pairs)

    customers_row = computed_row(
      noun.capitalize(), "Customers",
      lambda i, c: (
        f"=IF('{REVENUE_SHEET}'!{_ref(c, revenue_src)}<=0,0,"
        f"IFERROR(({units_expr(c)})/{_ref(c, repeat_row)},0))"),
      design.FMT_UNITS, f"Units sold divided by purchases per {singular}")

    retained_row = computed_row(
      f"Returning {noun}", "Retained customers",
      lambda i, c: ("0" if i == 0
                    else f"={_ref(c - 1, customers_row)}*{_ref(c, retention_row)}"),
      design.FMT_UNITS, "Last quarter's customers who come back")

    new_row = computed_row(
      f"New {noun}", "New customers",
      lambda i, c: f"={_ref(c, customers_row)}-{_ref(c, retained_row)}",
      design.FMT_UNITS, "The customers your marketing has to win")

    # THE AGREED RATE, ON ITS OWN VISIBLE ROW. The fallback below used to carry
    # hardcoded dollar amounts inside the formula, and that was wrong three ways:
    # those dollars were frozen at build-time revenue, so a client editing a
    # revenue driver left the held branch describing a plan that no longer
    # existed and the percentage silently dropped; the number was invisible and
    # uneditable because it only surfaced when the condition tripped; and it
    # converted a settled RATE into a fixed sum, cutting its tie to revenue.
    # As a rate on a row, it stays a rate and stays tied.
    agreed_row = computed_row(
      "Marketing % agreed in your plan", "Marketing percent agreed",
      lambda i, c: settled_percent[i], design.FMT_PERCENT,
      "The figure your plan was built on — used when a quarter wins no new "
      f"{noun}")

    # R1 under the reversal. A quarter with no net acquisition cannot derive
    # spend from new x CAC — that would zero out real marketing spend and change
    # the plan. It falls back to revenue x the agreed rate above, so it still
    # moves with revenue, and the note row says so.
    spend_row = computed_row(
      "Marketing spend", "Marketing spend",
      lambda i, c: (f"=IF({_ref(c, new_row)}>0,{_ref(c, new_row)}*{_ref(c, cac_row)},"
                    f"'{REVENUE_SHEET}'!{_ref(c, revenue_src)}*{_ref(c, agreed_row)})"),
      design.FMT_MONEY, f"New {noun} x the cost to win one", emphasis=True)

    percent_row = computed_row(
      "Marketing % of revenue", MARKETING_PERCENT_ROW_KEY,
      lambda i, c: (f"=IFERROR({_ref(c, spend_row)}/"
                    f"'{REVENUE_SHEET}'!{_ref(c, revenue_src)},0)"),
      design.FMT_PERCENT,
      "Spend divided by revenue — the figure the rest of your plan uses",
      emphasis=True)

    ws.cell(row=row, column=1, value="Why a spend is held").font = design.font("note")
    for index, column in enumerate(columns):
      note_key = str(periods[index].get("customer_acquisition_cost_note") or "")
      cell = ws.cell(row=row, column=column, value=_NOTE_TEXT.get(note_key, ""))
      cell.font = design.font("note")
    ctx.add_schedule_row(MARKETING_SCHEDULE_SHEET, "Acquisition note", row)
    row += 2
  else:
    # R2 / R4: no acquisition economics to run, so the percentage keeps the
    # settled values and Model Inputs still has a row to point at.
    computed_row(
      "Marketing % of revenue", MARKETING_PERCENT_ROW_KEY,
      lambda i, c: settled_percent[i], design.FMT_PERCENT,
      "The marketing figure agreed in your plan", emphasis=True)
    ws.cell(
      row=row, column=1,
      value=("This plan has no usable audience estimate, so the customer and "
             "acquisition-cost lines are left out rather than invented."
             if schedule_class != "zero_marketing"
             else "This plan carries no marketing spend."),
    ).font = design.font("note")
    row += 2

  write_section_header(ws, row, "Where these figures come from")
  row += 1
  provenance = [
    ("Marketing percentage",
     "Produced by this sheet and read by the rest of the model. On delivery it "
     "equals the figure agreed when your plan was built."),
    ("Customer retention",
     text(retention_meta.get("source")) or "Not available for this plan"),
  ]
  rationale = text(retention_meta.get("rationale"))
  if rationale:
    provenance.append(("Retention reasoning", rationale))
  provenance.append((
    f"Purchases per {singular}",
    text(repeat_meta.get("source")) or "Not available for this plan"))
  if modelled:
    provenance.append((
      f"Cost to acquire one {singular}",
      "Derived from the marketing figure agreed in your plan, then held as an "
      "input so your budget responds when you change the drivers above."))
  reachable = context.get("reachable_market")
  if reachable is not None:
    provenance.append((
      f"Reachable {noun}",
      f"{reachable} — the audience estimated for this business at intake"))
  for label, detail in provenance:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=detail).font = design.font("note")
    row += 1
