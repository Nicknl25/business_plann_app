"""The Dashboard — a live, macro-free, period-selectable view (2026-08-19).

Nothing is computed here. Every card and every chart reads the hidden Calc
sheet (the data engine):

  * period-specific cards and charts read Calc's CURRENT-VIEW region, so they
    reslice the moment the selector changes;
  * full-arc charts read Calc's RAW series, so the whole five-year picture is
    always on screen;
  * the CVP chart reads Calc's CVP block, which is itself driven by the
    selector - it used to be hard-wired to Q1.

The selector is a data-validation dropdown driving MATCH/INDEX. No macros, no
form controls, no ActiveX: it works on Excel for Windows, Mac and the web
(docs/WORKBOOK_ANALYTICS_RESEARCH.md §5/§6).
"""
from __future__ import annotations

from typing import Dict, Optional

from openpyxl.chart import Reference
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from . import design
from .break_even_sheet import build_cvp_chart
from .calc_sheet import CALC_KEY, CALC_SHEET, FIRST_COL, LAST_Q_COL, QUARTERS, YEARS
from .data import DraftWorkbookData
from .excel_utils import WorkbookBuildContext, create_sheet, ref

DASHBOARD_SHEET = "Dashboard"

QUARTER_LIST_NAME = "PeriodQuarters"
YEAR_LIST_NAME = "PeriodYears"

#: (label, calc key, number format, note)
_CARDS = [
  [("Revenue", "revenue", design.FMT_MONEY, ""),
   ("EBITDA", "ebitda", design.FMT_MONEY, ""),
   ("EBITDA Margin", "ebitda_margin", design.FMT_PERCENT, ""),
   ("Net Income", "net_income", design.FMT_MONEY, ""),
   ("Ending Cash", "cash", design.FMT_MONEY, "")],
  [("Break-Even Revenue", "be_revenue", design.FMT_MONEY, "revenue needed to cover every cost"),
   ("Margin of Safety", "be_mos", design.FMT_PERCENT, "how far above break-even"),
   ("Net Cash Flow", "net_cf", design.FMT_MONEY_SIGNED, "negative means cash was used"),
   ("Operating Cash Flow", "operating_cf", design.FMT_MONEY, ""),
   ("Total Debt", "total_debt", design.FMT_MONEY, "")],
  [("Gross Margin", "gross_margin", design.FMT_PERCENT, ""),
   ("Net Margin", "net_margin", design.FMT_PERCENT, ""),
   ("Debt Service Coverage", "dscr", design.FMT_RATIO, "EBITDA over debt service"),
   ("Cash Low Point", "cash_low", design.FMT_MONEY, "lowest cash across all 20 quarters"),
   ("Headcount (FTE)", "headcount", design.FMT_UNITS, "")],
]


def _calc(ctx: WorkbookBuildContext) -> Dict[str, int]:
  return ctx.schedule_rows.get(CALC_KEY) or {}


def _cur(ctx: WorkbookBuildContext, key: str) -> Optional[str]:
  row = _calc(ctx).get(f"cur::{key}")
  return ref(CALC_SHEET, row, FIRST_COL) if row else None


def _anchor(row: int, col: str) -> str:
  return f"{col}{row}"


def build_dashboard_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  calc_rows = _calc(ctx)
  if not calc_rows:
    return
  ws = create_sheet(wb, DASHBOARD_SHEET)
  calc_ws = wb[CALC_SHEET]
  ws.sheet_view.zoomScale = 90
  ws.sheet_view.showGridLines = False
  for col in range(1, 30):
    ws.column_dimensions[get_column_letter(col)].width = 11
  design.page_setup(ws, landscape=True, fit_width=True,
                    footer=f"{data.business_name} - Dashboard")

  title = ws.cell(row=1, column=1, value=f"{data.business_name} - Dashboard")
  title.font = design.font("title")
  sub = ws.cell(row=2, column=1,
                value="Choose a view and a period; every card and the period charts below "
                      "reslice to match. The trend charts always show all twenty quarters.")
  sub.font = design.font("subtitle")

  # ---------------- the selector -------------------------------------------
  q_row = calc_rows.get("__labels_q__")
  y_row = calc_rows.get("__labels_y__")
  wb.defined_names.add(DefinedName(
    QUARTER_LIST_NAME,
    attr_text=f"{CALC_SHEET}!${get_column_letter(FIRST_COL)}${q_row}:"
              f"${get_column_letter(LAST_Q_COL)}${q_row}"))
  wb.defined_names.add(DefinedName(
    YEAR_LIST_NAME,
    attr_text=f"{CALC_SHEET}!${get_column_letter(FIRST_COL)}${y_row}:"
              f"${get_column_letter(FIRST_COL + YEARS - 1)}${y_row}"))

  for row, label, value, source in (
    (3, "View", "Quarterly", '"Quarterly,Annual"'),
    (4, "Quarter", "Q1", f"={QUARTER_LIST_NAME}"),
    (5, "Year", "Y1", f"={YEAR_LIST_NAME}"),
  ):
    lab = ws.cell(row=row, column=1, value=label)
    lab.font = design.font("label_strong")
    cell = ws.cell(row=row, column=3, value=value)
    design.input_cell(cell, number_format=design.FMT_TEXT)
    # Never set showDropDown=True: in OOXML that HIDES the arrow.
    dv = DataValidation(type="list", formula1=source, allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell)

  banner = ws.cell(row=3, column=5, value=f'="Showing: "&{ref(CALC_SHEET, calc_rows["sel::sel_label"], 2)}'
                                          f'&" ("&{ref(CALC_SHEET, calc_rows["sel::sel_view"], 2)}&" view)"')
  banner.font = design.font("kpi_value")
  ws.merge_cells(start_row=3, start_column=5, end_row=4, end_column=11)
  hint = ws.cell(row=5, column=5,
                 value="The Quarter box drives the Quarterly view; the Year box drives the Annual view.")
  hint.font = design.font("footnote")
  ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=13)

  # ---------------- cards ---------------------------------------------------
  row = 7
  for card_row in _CARDS:
    col = 1
    for label, key, fmt, note in card_row:
      source = _cur(ctx, key)
      if source:
        design.kpi_tile(ws, row, col, label, f"={source}", fmt, note, span=3)
      col += 4
    row += 4

  # ---------------- break-even volume, 1..N lines ---------------------------
  be_first, be_last = calc_rows.get("be_units_first"), calc_rows.get("be_units_last")
  if be_first and be_last:
    single = (be_last - be_first) == 0
    heading = ("Break-even volume - selected period" if single
               else "Break-even volume at the planned mix - selected period")
    design.section_band(ws, row, heading, end_col=21)
    row += 1
    for r in range(be_first, be_last + 1):
      name = ws.cell(row=row, column=1, value=calc_ws.cell(row=r, column=1).value)
      name.font = design.font("label")
      # Line names run long on multi-line businesses ("Landscaping &
      # installation services / Landscaping-installation job"), so the name
      # gets six columns before the number rather than four.
      ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
      value = ws.cell(row=row, column=7, value=f"={ref(CALC_SHEET, r, FIRST_COL)}")
      design.calculated_cell(value, number_format=design.FMT_UNITS)
      unit = ws.cell(row=row, column=8, value="units to break even")
      unit.font = design.font("note")
      ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=13)
      row += 1
    row += 1

  # ---------------- charts --------------------------------------------------
  def raw(key: str) -> Reference:
    return Reference(calc_ws, min_col=FIRST_COL, max_col=LAST_Q_COL,
                     min_row=calc_rows[f"q::{key}"])

  q_cats = Reference(calc_ws, min_col=FIRST_COL, max_col=LAST_Q_COL, min_row=q_row)

  # 1 Revenue. The app builds anything from one line of business to N, so the
  #   chart adapts: a multi-line business gets the MIX stacked over time (the
  #   total is the stack height), a single-line business gets revenue against
  #   EBITDA, which is the more useful pair when there is no mix to show.
  line_count = calc_rows.get("line_count") or 0
  if line_count > 1:
    c1 = design.chart("stacked_column", title="Revenue by line of business - all twenty quarters",
                      y_format=design.FMT_AXIS_MONEY, legend="b")
    for i in range(line_count):
      row_i = calc_rows.get(f"line::{i}::revenue")
      if not row_i:
        continue
      label = calc_ws.cell(row=row_i, column=1).value or f"Line {i + 1}"
      design.add_series(c1, Reference(calc_ws, min_col=FIRST_COL, max_col=LAST_Q_COL, min_row=row_i),
                        title=str(label), slot=i, line=False)
  else:
    c1 = design.chart("line", title="Revenue and EBITDA - all twenty quarters",
                      y_format=design.FMT_AXIS_MONEY, legend="b")
    design.add_series(c1, raw("revenue"), title="Revenue", slot=0)
    design.add_series(c1, raw("ebitda"), title="EBITDA", slot=1)
  design.set_categories(c1, q_cats)
  design.place(ws, c1, _anchor(row, "A"))

  # 2 Margins, full arc
  c2 = design.chart("line", title="Margins - all twenty quarters",
                    y_format=design.FMT_AXIS_PERCENT, legend="b")
  design.add_series(c2, raw("gross_margin"), title="Gross margin", slot=0)
  design.add_series(c2, raw("ebitda_margin"), title="EBITDA margin", slot=1)
  design.add_series(c2, raw("net_margin"), title="Net margin", slot=6)
  design.set_categories(c2, q_cats)
  design.place(ws, c2, _anchor(row, "K"))

  # 3 Cash and debt, full arc
  c3 = design.chart("line", title="Ending cash and total debt - all twenty quarters",
                    y_format=design.FMT_AXIS_MONEY, legend="b")
  design.add_series(c3, raw("cash"), title="Ending cash", slot=0)
  design.add_series(c3, raw("total_debt"), title="Total debt", color=design.SERIES_COST)
  design.set_categories(c3, q_cats)
  design.place(ws, c3, _anchor(row + 18, "A"))

  # 4 Revenue against break-even, full arc
  c4 = design.chart("line", title="Revenue against break-even - all twenty quarters",
                    y_format=design.FMT_AXIS_MONEY, legend="b")
  design.add_series(c4, raw("revenue"), title="Revenue", slot=0)
  design.add_series(c4, raw("be_revenue"), title="Break-even revenue",
                    color=design.SERIES_ATTENTION)
  design.set_categories(c4, q_cats)
  design.place(ws, c4, _anchor(row + 18, "K"))

  # 5 Cash built (+) and cash used (-), full arc. A single "cash burn" series
  #   is flat at zero for a business that never burns cash (Bellweather is one),
  #   which reads as a broken chart; splitting the sign shows the build AND the
  #   burn, and a burning quarter appears in red the moment it exists.
  c5 = design.chart("column", title="Cash built (+) and cash used (-) by quarter",
                    y_format=design.FMT_AXIS_MONEY, legend="b")
  design.add_series(c5, raw("cash_built"), title="Cash built", slot=0, line=False)
  design.add_series(c5, raw("cash_used"), title="Cash used", color=design.SERIES_COST, line=False)
  design.set_categories(c5, q_cats)
  design.place(ws, c5, _anchor(row + 36, "A"))

  # 6 Headcount, full arc
  c6 = design.chart("line", title="Headcount by quarter (ending FTE)",
                    y_format=design.FMT_AXIS_UNITS, legend=None)
  design.add_series(c6, raw("headcount"), title="Total ending FTE", slot=6)
  design.set_categories(c6, q_cats)
  design.place(ws, c6, _anchor(row + 36, "K"))

  # 7 Cost structure - the SELECTED period
  c7 = design.chart("bar", title="Cost structure - selected period",
                    y_format=design.FMT_AXIS_MONEY, legend=None, height=9.5)
  design.add_series(c7, Reference(calc_ws, min_col=FIRST_COL,
                                  min_row=calc_rows["cost_first"], max_row=calc_rows["cost_last"]),
                    slot=0, line=False, labels=True, label_position="outEnd")
  design.set_categories(c7, Reference(calc_ws, min_col=1, min_row=calc_rows["cost_first"],
                                      max_row=calc_rows["cost_last"]))
  design.place(ws, c7, _anchor(row + 54, "A"))

  # 8 Sources and uses - the SELECTED period
  c8 = design.chart("bar", title="Sources (+) and uses (-) of cash - selected period",
                    y_format=design.FMT_AXIS_MONEY, legend=None, height=9.5)
  design.add_series(c8, Reference(calc_ws, min_col=FIRST_COL,
                                  min_row=calc_rows["su_first"], max_row=calc_rows["su_last"]),
                    slot=0, line=False, labels=True, label_position="outEnd")
  design.set_categories(c8, Reference(calc_ws, min_col=1, min_row=calc_rows["su_first"],
                                      max_row=calc_rows["su_last"]))
  design.place(ws, c8, _anchor(row + 54, "K"))

  # 9 CVP - the SELECTED period (reads Calc's selector-driven CVP block)
  c9 = build_cvp_chart(calc_ws, first=calc_rows["cvp::first"], last=calc_rows["cvp::last"],
                       title="Break-even (cost-volume-profit) - selected period")
  c9.height, c9.width = 9.5, 16.5
  design.place(ws, c9, _anchor(row + 72, "A"))

  ws.sheet_properties.tabColor = design.NAVY
