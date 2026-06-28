"""Phase 9 P3.9 — Workbook 'Diagnostics' sheet builder.

Renders the per-run diagnostic payload produced by
`python.client_intake_and_finmo.post_intake_run_diagnostics`. The
workbook is a pure reflection of the SQL diagnostic record -- never a
source of truth, never a separate compute.

Sheet structure:
  Section 1 -- Run metadata (planning mode, cash strategy, business
              stage / name / start_date, draft id, planning_run_id)
  Section 2 -- Acceptance gate (X/X score + verdict)
  Section 3 -- Realism check detail (every metric_key + pass/fail
              marker, ordered by metric_key)

Additive: appended to the workbook as the LAST sheet. If the
diagnostic payload is missing or empty, a placeholder sheet is still
rendered so the workbook layout is consistent across runs.
"""

from __future__ import annotations

from typing import Any, Dict, Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .data import DraftWorkbookData
from .excel_utils import (
  FILL_LIGHT,
  FILL_NAVY,
  FILL_GRAY,
  FONT_WHITE,
  FONT_BLACK,
  WorkbookBuildContext,
)


DIAGNOSTICS_SHEET = "Diagnostics"

_PASS_MARK = "✓"   # ✓
_FAIL_MARK = "X"


def _bold(size: int = 11) -> Font:
  return Font(name="Calibri", size=size, bold=True, color=FONT_BLACK)


def _section_header_fill() -> PatternFill:
  return PatternFill(fill_type="solid", fgColor=FILL_NAVY)


def _section_header_font() -> Font:
  return Font(name="Calibri", size=12, bold=True, color=FONT_WHITE)


def _row_alt_fill() -> PatternFill:
  return PatternFill(fill_type="solid", fgColor=FILL_LIGHT)


def _section_header(ws: Worksheet, row: int, text: str) -> int:
  cell = ws.cell(row=row, column=1, value=text)
  cell.font = _section_header_font()
  cell.fill = _section_header_fill()
  cell.alignment = Alignment(horizontal="left", vertical="center")
  for col in range(2, 4):
    ws.cell(row=row, column=col).fill = _section_header_fill()
  ws.row_dimensions[row].height = 22
  return row + 1


def _kv_row(ws: Worksheet, row: int, key: str, value: Any) -> int:
  label_cell = ws.cell(row=row, column=1, value=str(key))
  label_cell.font = _bold(11)
  label_cell.alignment = Alignment(horizontal="left", vertical="center")
  val_cell = ws.cell(row=row, column=2, value="" if value is None else str(value))
  val_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
  return row + 1


def _check_row(
  ws: Worksheet, row: int, metric_key: str, passed: bool, *, alt: bool,
) -> int:
  k = ws.cell(row=row, column=1, value=str(metric_key))
  k.alignment = Alignment(horizontal="left", vertical="center")
  status_text = _PASS_MARK if passed else _FAIL_MARK
  v = ws.cell(row=row, column=2, value=status_text)
  v.alignment = Alignment(horizontal="center", vertical="center")
  v.font = Font(
    name="Calibri", size=11, bold=True,
    color="006100" if passed else "9C0006",
  )
  if alt:
    fill = _row_alt_fill()
    k.fill = fill
    v.fill = fill
  return row + 1


def build_diagnostics_sheet(
  wb: Workbook,
  data: DraftWorkbookData,
  ctx: WorkbookBuildContext,
  diagnostics_payload: Optional[Dict[str, Any]] = None,
) -> Worksheet:
  """Append the Diagnostics sheet to the workbook as the LAST sheet.

  ``diagnostics_payload`` is the dict returned by
  ``post_intake_run_diagnostics.build_run_diagnostics_payload`` (or
  loaded via ``load_latest_diagnostics_for_draft``). When None, the
  sheet still renders with placeholder values so the workbook layout
  stays consistent across runs.
  """
  ws = wb.create_sheet(title=DIAGNOSTICS_SHEET)
  ws.sheet_view.showGridLines = False
  ws.column_dimensions["A"].width = 48
  ws.column_dimensions["B"].width = 36
  ws.column_dimensions["C"].width = 14

  payload = diagnostics_payload or {}

  row = 1
  title = ws.cell(row=row, column=1, value="Run Diagnostics")
  title.font = Font(name="Calibri", size=16, bold=True, color=FONT_BLACK)
  ws.row_dimensions[row].height = 24
  row += 2

  # Section 1 -- Run metadata.
  row = _section_header(ws, row, "Run Metadata")
  row = _kv_row(ws, row, "Planning Mode", payload.get("planning_mode"))
  row = _kv_row(ws, row, "Cash Strategy", payload.get("cash_strategy_name"))
  row = _kv_row(ws, row, "Business Stage", payload.get("business_stage"))
  row = _kv_row(ws, row, "Business Name", payload.get("business_name"))
  row = _kv_row(ws, row, "Business Start Date", payload.get("business_start_date"))
  row = _kv_row(ws, row, "NAICS-6", payload.get("business_naics_6"))
  row = _kv_row(ws, row, "Draft ID", payload.get("draft_id") or data.draft_id)
  row = _kv_row(ws, row, "Planning Run ID", payload.get("planning_run_id"))
  row += 1

  # Section 2 -- Acceptance gate.
  row = _section_header(ws, row, "Acceptance Gate")
  # Prefer the human "ok/total" label ("15/16"); fall back to the numeric.
  score = payload.get("acceptance_score_label") or payload.get("acceptance_score")
  passed = payload.get("acceptance_passed")
  verdict_text = (
    "PASSED" if passed is True
    else "FAILED" if passed is False
    else "UNKNOWN"
  )
  row = _kv_row(ws, row, "Score", score)
  row = _kv_row(ws, row, "Verdict", verdict_text)
  row += 1

  # Section 2b -- Handler summary (only when handler fired).
  handler_fired = bool(payload.get("handler_fired"))
  row = _section_header(ws, row, "GPT Exhaustion Handler")
  if handler_fired:
    row = _kv_row(ws, row, "Fired", "Yes")
    row = _kv_row(ws, row, "Status", payload.get("handler_status"))
    row = _kv_row(ws, row, "Scope", payload.get("handler_scope"))
    row = _kv_row(ws, row, "Tool calls used", payload.get("tool_calls_used"))
    row = _kv_row(ws, row, "Budget extension triggered",
                  payload.get("budget_extension_triggered"))
  else:
    row = _kv_row(ws, row, "Fired", "No (restoration loop landed deterministically)")
  row += 1

  # Section 3 -- Realism check detail.
  row = _section_header(ws, row, "Realism Check Detail")
  header_k = ws.cell(row=row, column=1, value="Metric Key")
  header_v = ws.cell(row=row, column=2, value="Status")
  for c in (header_k, header_v):
    c.font = _bold(11)
    c.fill = PatternFill(fill_type="solid", fgColor=FILL_GRAY)
    c.alignment = Alignment(horizontal="left", vertical="center")
  header_v.alignment = Alignment(horizontal="center", vertical="center")
  row += 1

  checks = payload.get("realism_checks") or []
  if not isinstance(checks, list):
    checks = []
  if checks:
    for idx, entry in enumerate(checks):
      if not isinstance(entry, dict):
        continue
      mk = str(entry.get("metric_key") or "")
      passed_metric = bool(entry.get("passed"))
      row = _check_row(ws, row, mk, passed_metric, alt=(idx % 2 == 1))
  else:
    ws.cell(row=row, column=1, value="(no realism gate results recorded for this run)")
    row += 1

  # Freeze the title row so scrolling keeps the heading visible.
  ws.freeze_panes = "A2"
  return ws
