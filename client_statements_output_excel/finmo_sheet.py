from __future__ import annotations

from typing import Dict, List, Tuple

from .finmo_ratios import RATIOS_STATEMENT, fill_ratio_formulas, write_ratio_rows
from .break_even_sheet import (
  BREAK_EVEN_STATEMENT,
  fill_break_even_formulas,
  write_break_even_rows,
)
from openpyxl.utils import get_column_letter

from .data import DraftWorkbookData, text
from .excel_utils import (
  ANNUAL_SUM,
  ANNUAL_YEAR_END,
  annual_mode_for,
  ANNUAL_START_COL,
  CURRENCY_FORMAT,
  DEBT_SHEET,
  FILL_BLUE,
  FILL_GREEN,
  FILL_LIGHT,
  FINMO_SHEET,
  FIRST_LIVE_COL,
  LAST_LIVE_COL,
  MODEL_INPUT_SHEET,
  NUMBER_FORMAT,
  PERIOD_COUNT,
  PERIOD_END_COL,
  PERIOD_START_COL,
  WorkbookBuildContext,
  add_annual_formulas,
  apply_base_style,
  create_sheet,
  local_ref,
  range_ref,
  ref,
  set_formula_style,
  set_title,
  style_row,
  write_period_headers,
  write_section_header,
)


PL_LINES = [
  "Revenue",
  "Cost of Goods Sold",
  "Gross Profit",
  "Marketing",
  "Research & Development",
  "Lease/Rent",
  "Payroll",
  "General & Administrative",
  "EBITDA",
  "Interest",
  "Depreciation",
  "Taxes",
  "Net Income",
]

BS_LINES = [
  "Cash",
  "Accounts Receivable",
  "Inventory",
  "Prepaid Expenses",
  "Current Assets",
  "PPE",
  # Phase 9 P3.16 — Right-of-Use Asset (Capital Lease) is a separate
  # asset line parallel to PPE. Q0 seeded from capital lease balance,
  # depreciates straight-line over CAPITAL_LEASE_DEPRECIATION_QUARTERS.
  "Right-of-Use Asset (Capital Lease)",
  "Accumulated Depreciation",
  "Total Assets",
  "Accounts Payable",
  "Short Term Debt",
  "Deferred Revenue",
  "Current Liabilities",
  "Long Term Debt",
  # Phase 9 P3.16 — Capital Lease Obligation surfaces the lease
  # liability as its own line so the balance sheet shows the offset
  # to the new ROU asset (was previously folded into Total Liabilities
  # with no asset-side counterpart).
  "Capital Lease Obligation",
  "Total Liabilities",
  "Owner's Capital",
  "Retained Earnings",
  "Other Equity",
  "Total Equity",
  "Total Liabilities & Equity",
]

CF_LINES = [
  "Beginning Cash",
  "Net Income",
  "Depreciation",
  "Changes in Current Assets",
  "Changes in Current Liabilities",
  "Operating Cash Flow",
  "Capital Expenditures",
  "Investing Cash Flow",
  "Debt Issuance (New Borrowing)",
  "Debt Repayment",
  "Equity",
  "Distributions",
  # Phase 9 P3.16 — Capital Lease Principal Payments shown as its
  # own financing line (was folded into Financing Cash Flow with no
  # display row).
  "Capital Lease Principal Payments",
  "Financing Cash Flow",
  "Net Cash Flow",
  "Ending Cash",
]


def _mi(ctx: WorkbookBuildContext, key: str, col: int) -> str:
  return ref(MODEL_INPUT_SHEET, ctx.model_input_row(key), col)


def _fr(ctx: WorkbookBuildContext, statement: str, label: str, col: int) -> str:
  return local_ref(ctx.finmo_row(statement, label), col)


def _prior(ctx: WorkbookBuildContext, statement: str, label: str, col: int) -> str:
  return local_ref(ctx.finmo_row(statement, label), col - 1)


def _days_in_quarter_formula(col: int) -> str:
  return local_ref(6, col)


def _short_term_debt_formula(ctx: WorkbookBuildContext, col: int) -> str:
  """Phase 9 P3.10 STD canonical-source layer 2 — FINMO sheet STD cell
  references the Debt Schedule sheet's `Actual Debt Repayment` row at
  cells col+1..col+4 (exclusive of the current quarter), clipped to
  the live-period range. Q20's window is entirely beyond the horizon
  so the cell is literal "=0".
  """
  start_col = col + 1
  end_col = col + 4
  if start_col > LAST_LIVE_COL:
    return "=0"
  end_col = min(end_col, LAST_LIVE_COL)
  actual_repay_row = ctx.schedule_row(DEBT_SHEET, "Actual Debt Repayment")
  return f"=SUM({range_ref(DEBT_SHEET, actual_repay_row, start_col, end_col)})"


def _write_statement_rows(ws, ctx: WorkbookBuildContext, *, statement: str, lines: List[str], start_row: int) -> int:
  write_section_header(ws, start_row, statement)
  row = start_row + 1
  for line in lines:
    ws.cell(row=row, column=1, value=line)
    ctx.add_finmo_row(statement, line, row)
    row += 1
  return row + 1


def _set_formula(ws, row: int, col: int, formula: str, *, number_format: str = CURRENCY_FORMAT) -> None:
  cell = ws.cell(row=row, column=col, value=formula)
  set_formula_style(cell, number_format=number_format, internal_link=True)


def build_finmo_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, FINMO_SHEET)
  apply_base_style(ws)
  set_title(ws, "FINMO", "Three-statement financial model. All formulas reference Model Inputs and in-sheet statement rows.")
  write_period_headers(ws, data.periods)
  ws.cell(row=6, column=1, value="Days in Quarter")
  for idx in range(PERIOD_COUNT):
    period = data.periods[idx] if idx < len(data.periods) else {}
    ws.cell(row=6, column=PERIOD_START_COL + idx, value=period.get("days_in_quarter") or 0)

  # CW-032 layout (Nick's ruling): the per-line COGS breakout lives on
  # MODEL INPUTS (one "<line> - COGS %" driver row per line, already
  # rendered there); the P&L stays clean with ONE Cost of Goods Sold row
  # whose formula rolls the drivers up - Sigma(line revenue x that line's
  # COGS driver), summed in the one cell. Per-line P&L rows are gone; a
  # client reads their four assumptions on Model Inputs and one honest
  # total on the P&L. Single-line drafts have no COGS % slots and render
  # the exact legacy layout.
  per_line_cogs_slots: List[Tuple[str, str]] = []
  _seen_cogs_slots = set()
  for source in data.revenue_rows:
    if text(source.get("driver")) != "COGS %":
      continue
    slot = text(source.get("revenue_slot_key")) or f"{text(source.get('lob'))}::{text(source.get('product'))}"
    if slot in _seen_cogs_slots:
      continue
    _seen_cogs_slots.add(slot)
    display = " / ".join([text(source.get("lob")) or "LOB", text(source.get("product")) or "Product"])
    per_line_cogs_slots.append((slot, display))
  pl_lines = list(PL_LINES)

  row = 7
  row = _write_statement_rows(ws, ctx, statement="Income Statement", lines=pl_lines, start_row=row)
  row = _write_statement_rows(ws, ctx, statement="Balance Sheet", lines=BS_LINES, start_row=row)
  row = _write_statement_rows(ws, ctx, statement="Cash Flow", lines=CF_LINES, start_row=row)
  # 2026-08-19 (Nick's structure ruling): FINMO reads top to bottom as
  # Income Statement -> Balance Sheet -> Cash Flow -> BREAK-EVEN -> RATIOS.
  # Both analysis blocks reserve their rows here so every reference resolves
  # by label through ctx; their formulas are filled after the statement loop
  # because they read the statement cells above them.
  row = write_break_even_rows(ws, data, ctx, start_row=row)
  row = write_ratio_rows(ws, ctx, start_row=row)

  for idx in range(PERIOD_COUNT):
    col = PERIOD_START_COL + idx
    q0 = idx == 0
    owner_capital_ref = _fr(ctx, "Balance Sheet", "Owner's Capital", col)
    prior_owner_capital_ref = (
      _prior(ctx, "Balance Sheet", "Owner's Capital", col)
      if not q0
      else owner_capital_ref
    )
    owner_capital_input_ref = _mi(ctx, "bs::Owner's Capital", col)

    # Income Statement
    _set_formula(ws, ctx.finmo_row("Income Statement", "Revenue"), col, f"={_mi(ctx, 'is::Revenue', col)}")
    if per_line_cogs_slots:
      # ONE consolidated COGS cell: each line's revenue times that line's
      # Model Inputs COGS driver, summed here. The breakout is readable on
      # Model Inputs; the P&L carries the single provable roll-up.
      _rollup = "+".join(
        f"{_mi(ctx, f'revenue::{slot}::Revenue', col)}*{_mi(ctx, f'revenue::{slot}::COGS %', col)}"
        for slot, _display in per_line_cogs_slots
      )
      _set_formula(
        ws,
        ctx.finmo_row("Income Statement", "Cost of Goods Sold"),
        col,
        f"={_rollup}",
      )
    else:
      _set_formula(ws, ctx.finmo_row("Income Statement", "Cost of Goods Sold"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}*{_mi(ctx, 'is::Cost of Goods Sold', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Gross Profit"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}-{_fr(ctx, 'Income Statement', 'Cost of Goods Sold', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Marketing"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}*{_mi(ctx, 'is::Marketing', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Research & Development"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}*{_mi(ctx, 'is::Research & Development', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Lease/Rent"), col, f"={_mi(ctx, 'is::Lease', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Payroll"), col, f"={_mi(ctx, 'is::Payroll', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "General & Administrative"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}*{_mi(ctx, 'is::General & Administrative', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "EBITDA"), col, f"={_fr(ctx, 'Income Statement', 'Gross Profit', col)}-SUM({_fr(ctx, 'Income Statement', 'Marketing', col)}:{_fr(ctx, 'Income Statement', 'General & Administrative', col)})")
    # Phase 9 P3.17 Phase 3c — P&L Interest and Depreciation are
    # COMBINED totals per FORMULA_REGISTRY (debt+lease for Interest,
    # PPE+lease for Depreciation). Pre-iter the workbook P&L cells
    # read only the legacy `is::Interest Expense` / `is::Depreciation
    # Expense` Model Inputs cells, which point at the debt-only /
    # PPE-only source rows on the Debt Schedule and CapEx sheets.
    # The lease components (cash::Lease Interest Expense, cash::Lease
    # Asset Depreciation) were exposed on Model Inputs but never read
    # by the P&L, so the workbook P&L silently dropped them while the
    # persisted FINMO `interest` and `depreciation` fields correctly
    # included them. Doctrine Pattern 1 / Mirror Flavor 1 — the
    # workbook P&L formulas now reference both components directly.
    _set_formula(ws, ctx.finmo_row("Income Statement", "Interest"), col, f"={_mi(ctx, 'is::Interest Expense', col)}+{_mi(ctx, 'cash::Lease Interest Expense', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Depreciation"), col, f"={_mi(ctx, 'is::Depreciation Expense', col)}+{_mi(ctx, 'cash::Lease Asset Depreciation', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Taxes"), col, f"=MAX(0,{_fr(ctx, 'Income Statement', 'EBITDA', col)}-{_fr(ctx, 'Income Statement', 'Interest', col)}-{_fr(ctx, 'Income Statement', 'Depreciation', col)})*{_mi(ctx, 'is::Taxes', col)}")
    _set_formula(ws, ctx.finmo_row("Income Statement", "Net Income"), col, f"={_fr(ctx, 'Income Statement', 'EBITDA', col)}-SUM({_fr(ctx, 'Income Statement', 'Interest', col)}:{_fr(ctx, 'Income Statement', 'Taxes', col)})")

    # Balance Sheet
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Cash"), col, f"={_fr(ctx, 'Cash Flow', 'Ending Cash', col)}")
    if q0:
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Accounts Receivable"), col, f"={_mi(ctx, 'cash::Accounts Receivable Opening Balance', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Inventory"), col, f"={_mi(ctx, 'cash::Inventory Opening Balance', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Prepaid Expenses"), col, "=0")
    else:
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Accounts Receivable"), col, f"=({_mi(ctx, 'bs::Accounts Receivable Days', col)}/{_days_in_quarter_formula(col)})*{_fr(ctx, 'Income Statement', 'Revenue', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Inventory"), col, f"=({_mi(ctx, 'bs::Inventory Days', col)}/{_days_in_quarter_formula(col)})*{_fr(ctx, 'Income Statement', 'Cost of Goods Sold', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Prepaid Expenses"), col, f"={_fr(ctx, 'Income Statement', 'Revenue', col)}*{_mi(ctx, 'bs::Prepaid Expenses (% of Revenue)', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Current Assets"), col, f"=SUM({_fr(ctx, 'Balance Sheet', 'Cash', col)}:{_fr(ctx, 'Balance Sheet', 'Prepaid Expenses', col)})")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "PPE"), col, f"={_mi(ctx, 'cash::PPE Closing Balance', col)}")
    # Phase 9 P3.16 — Right-of-Use Asset references the Debt Schedule
    # sheet's ROU Asset Closing row (via cash::Right-of-Use Asset).
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Right-of-Use Asset (Capital Lease)"), col, f"={_mi(ctx, 'cash::Right-of-Use Asset', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Accumulated Depreciation"), col, f"={_mi(ctx, 'cash::Accumulated Depreciation', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Total Assets"), col, f"={_fr(ctx, 'Balance Sheet', 'Current Assets', col)}+{_fr(ctx, 'Balance Sheet', 'PPE', col)}+{_fr(ctx, 'Balance Sheet', 'Right-of-Use Asset (Capital Lease)', col)}")
    operating_expense_sum = f"SUM({_fr(ctx, 'Income Statement', 'Marketing', col)}:{_fr(ctx, 'Income Statement', 'General & Administrative', col)})"
    if q0:
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Accounts Payable"), col, f"={_mi(ctx, 'cash::Accounts Payable Opening Balance', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Short Term Debt"), col, f"={_mi(ctx, 'cash::Short Term Debt Opening Balance', col)}")
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Deferred Revenue"), col, "=0")
    else:
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Accounts Payable"), col, f"=({_mi(ctx, 'bs::Accounts Payable Days', col)}/{_days_in_quarter_formula(col)})*{operating_expense_sum}")
      # Phase 9 P3.10 STD canonical-source layer 2 — workbook FINMO STD
      # references the Debt Schedule sheet's Actual Debt Repayment row
      # for the next 4 quarters (q+1..q+4), exclusive of the current
      # quarter. Out-of-horizon columns naturally drop out by clipping
      # the SUM range to PERIOD_END_COL; the last live column (Q20) has
      # no quarters after it, so its formula is the literal "=0".
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Short Term Debt"), col, _short_term_debt_formula(ctx, col))
      _set_formula(ws, ctx.finmo_row("Balance Sheet", "Deferred Revenue"), col, f"={_mi(ctx, 'bs::Deferred Revenue (% of Revenue)', col)}*{_fr(ctx, 'Income Statement', 'Revenue', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Current Liabilities"), col, f"=SUM({_fr(ctx, 'Balance Sheet', 'Accounts Payable', col)}:{_fr(ctx, 'Balance Sheet', 'Deferred Revenue', col)})")
    # Phase 9 P3.10 iter 15 — workbook LTD is the NON-CURRENT portion
    # of the debt closing balance: closing - STD. STD + LTD = closing
    # by construction; pre-iter-15 the workbook displayed full closing
    # in LTD AND added STD separately to Current Liabilities, so Total
    # Liabilities double-counted current-portion debt every quarter
    # where STD > 0. Mirrors the Layer 1 fix in finmo_model.py.
    _set_formula(
      ws,
      ctx.finmo_row("Balance Sheet", "Long Term Debt"),
      col,
      f"=MAX(0,{_mi(ctx, 'cash::Debt Opening Balance', col) if q0 else _mi(ctx, 'cash::Debt Closing Balance', col)}-{_fr(ctx, 'Balance Sheet', 'Short Term Debt', col)})",
    )
    # Phase 9 P3.16 — Capital Lease Obligation displays the lease
    # closing balance as its own line so the BS shows it explicitly
    # (rather than folding into the Total Liabilities formula).
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Capital Lease Obligation"), col, f"={_mi(ctx, 'cash::Lease Closing Balance', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Total Liabilities"), col, f"={_fr(ctx, 'Balance Sheet', 'Current Liabilities', col)}+{_fr(ctx, 'Balance Sheet', 'Long Term Debt', col)}+{_fr(ctx, 'Balance Sheet', 'Capital Lease Obligation', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Owner's Capital"), col, f"={owner_capital_input_ref}")
    if q0:
      retained_formula = f"={_fr(ctx, 'Balance Sheet', 'Total Assets', col)}-{_fr(ctx, 'Balance Sheet', 'Total Liabilities', col)}-{owner_capital_ref}-{_fr(ctx, 'Balance Sheet', 'Other Equity', col)}"
    else:
      retained_formula = f"={_prior(ctx, 'Balance Sheet', 'Retained Earnings', col)}+{_fr(ctx, 'Income Statement', 'Net Income', col)}-{_mi(ctx, 'bs::Distributions', col)}"
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Retained Earnings"), col, retained_formula)
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Other Equity"), col, f"={_mi(ctx, 'bs::Other Equity', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Total Equity"), col, f"={owner_capital_ref}+{_fr(ctx, 'Balance Sheet', 'Retained Earnings', col)}+{_fr(ctx, 'Balance Sheet', 'Other Equity', col)}")
    _set_formula(ws, ctx.finmo_row("Balance Sheet", "Total Liabilities & Equity"), col, f"={_fr(ctx, 'Balance Sheet', 'Total Liabilities', col)}+{_fr(ctx, 'Balance Sheet', 'Total Equity', col)}")

    # Cash Flow
    if q0:
      begin_cash_formula = f"={_mi(ctx, 'cash::Cash Opening Balance', col)}"
      equity_formula = "=0"
      current_asset_change = "=0"
      current_liability_change = "=0"
    else:
      begin_cash_formula = f"={_prior(ctx, 'Cash Flow', 'Ending Cash', col)}"
      equity_formula = f"=({owner_capital_ref}-{prior_owner_capital_ref})+({_fr(ctx, 'Balance Sheet', 'Other Equity', col)}-{_prior(ctx, 'Balance Sheet', 'Other Equity', col)})"
      current_asset_change = f"=-(({_fr(ctx, 'Balance Sheet', 'Accounts Receivable', col)}+{_fr(ctx, 'Balance Sheet', 'Inventory', col)}+{_fr(ctx, 'Balance Sheet', 'Prepaid Expenses', col)})-({_prior(ctx, 'Balance Sheet', 'Accounts Receivable', col)}+{_prior(ctx, 'Balance Sheet', 'Inventory', col)}+{_prior(ctx, 'Balance Sheet', 'Prepaid Expenses', col)}))"
      # Phase 9 P3.10 iter 16 — operational subset (AP + DR) only.
      # STD reclassification is a balance-sheet presentation change,
      # not an operating cash event. Pre-iter-16 this referenced
      # the full Current Liabilities cell (AP + STD + DR), so ΔSTD
      # was inflating OCF by accumulated ΔSTD and the balance sheet
      # did not reconcile once iter 15 corrected LTD double-counting.
      # The displayed Current Liabilities row remains AP + STD + DR.
      current_liability_change = (
        f"=({_fr(ctx, 'Balance Sheet', 'Accounts Payable', col)}+{_fr(ctx, 'Balance Sheet', 'Deferred Revenue', col)})"
        f"-({_prior(ctx, 'Balance Sheet', 'Accounts Payable', col)}+{_prior(ctx, 'Balance Sheet', 'Deferred Revenue', col)})"
      )
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Beginning Cash"), col, begin_cash_formula)
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Net Income"), col, "=0" if q0 else f"={_fr(ctx, 'Income Statement', 'Net Income', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Depreciation"), col, "=0" if q0 else f"={_fr(ctx, 'Income Statement', 'Depreciation', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Changes in Current Assets"), col, current_asset_change)
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Changes in Current Liabilities"), col, current_liability_change)
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Operating Cash Flow"), col, "=0" if q0 else f"=SUM({_fr(ctx, 'Cash Flow', 'Net Income', col)}:{_fr(ctx, 'Cash Flow', 'Changes in Current Liabilities', col)})")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Capital Expenditures"), col, f"={_mi(ctx, 'cash::Capital Expenditures', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Investing Cash Flow"), col, "=0" if q0 else f"=-{_fr(ctx, 'Cash Flow', 'Capital Expenditures', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Debt Issuance (New Borrowing)"), col, f"={_mi(ctx, 'cash::Debt Issuance', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Debt Repayment"), col, f"={_mi(ctx, 'cash::Debt Repayment', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Equity"), col, equity_formula)
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Distributions"), col, f"={_mi(ctx, 'bs::Distributions', col)}")
    # Phase 9 P3.16 — Capital Lease Principal Payments shown as
    # its own financing-section line. Financing Cash Flow now
    # references this cell rather than folding the model input.
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Capital Lease Principal Payments"), col, f"={_mi(ctx, 'cash::Lease Principal Repayments', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Financing Cash Flow"), col, "=0" if q0 else f"={_fr(ctx, 'Cash Flow', 'Debt Issuance (New Borrowing)', col)}-{_fr(ctx, 'Cash Flow', 'Debt Repayment', col)}+{_fr(ctx, 'Cash Flow', 'Equity', col)}-{_fr(ctx, 'Cash Flow', 'Distributions', col)}-{_fr(ctx, 'Cash Flow', 'Capital Lease Principal Payments', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Net Cash Flow"), col, "=0" if q0 else f"={_fr(ctx, 'Cash Flow', 'Operating Cash Flow', col)}+{_fr(ctx, 'Cash Flow', 'Investing Cash Flow', col)}+{_fr(ctx, 'Cash Flow', 'Financing Cash Flow', col)}")
    _set_formula(ws, ctx.finmo_row("Cash Flow", "Ending Cash"), col, f"={_fr(ctx, 'Cash Flow', 'Beginning Cash', col)}+{_fr(ctx, 'Cash Flow', 'Net Cash Flow', col)}")

  subtotal_rows = {
    "Revenue",
    "Gross Profit",
    "EBITDA",
    "Net Income",
    "Current Assets",
    "Total Assets",
    "Current Liabilities",
    "Total Liabilities",
    "Total Equity",
    "Total Liabilities & Equity",
    "Operating Cash Flow",
    "Investing Cash Flow",
    "Financing Cash Flow",
    "Net Cash Flow",
    "Ending Cash",
  }
  for statement, line_map in ctx.finmo_rows.items():
    if statement in (BREAK_EVEN_STATEMENT, RATIOS_STATEMENT):
      # Both analysis blocks write EVERY column themselves, annual ones
      # included and correctly re-derived per ratio. Letting the statement-style
      # aggregation also run over them is what printed $0 under the six ratio
      # section headers.
      continue
    for label, r in line_map.items():
      # Route by the ROW, not by the statement. A blanket "sum unless it is the
      # balance sheet" is what let Beginning/Ending Cash - balances living on the
      # cash-flow statement - be summed across the year.
      mode = ANNUAL_YEAR_END if statement == "Balance Sheet" else annual_mode_for(label, CURRENCY_FORMAT)
      add_annual_formulas(ws, r, mode=mode)
      style_row(
        ws,
        r,
        fill=FILL_BLUE if label in subtotal_rows else None,
        bold=label in subtotal_rows,
        number_format=CURRENCY_FORMAT,
        border_top=label in subtotal_rows,
      )

  # Fill the two analysis blocks (live formulas off the statement cells above).
  # The CVP helper data no longer lives here - it moved to the hidden Calc
  # sheet (Nick: "it's helper data - invisible"), and the CVP chart is placed
  # from there so it follows the dashboard's period selector.
  fill_break_even_formulas(ws, data, ctx)
  fill_ratio_formulas(ws, ctx)

  # Client-facing cleanup (Nick, 2026-08-19): the stub column is a modelling
  # period, not a plan period, and "Days in Quarter" is a formula input - both
  # stay in the model and out of sight. Hiding is a display attribute: every
  # formula that references them is untouched.
  ws.column_dimensions[get_column_letter(PERIOD_START_COL)].hidden = True
  ws.row_dimensions[6].hidden = True
