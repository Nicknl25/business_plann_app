"""The hidden CALC sheet — the dashboard's single data engine (2026-08-19).

Nick's architecture call:

  * it holds ALL series for all 20 quarters AND all 5 years — the raw material,
    always present;
  * it has a SELECTOR-DRIVEN "current view" region that slices the selected
    period out of that raw material;
  * every card and chart on the Dashboard reads from here — period-specific
    ones from the current-view region (so they reslice with the toggle),
    full-arc ones from the raw series (so they always show the whole picture);
  * wire the selector in ONCE and everything follows: no per-chart wiring, no
    macros, and a new chart later just plugs in.

Mechanism (empirically verified in docs/WORKBOOK_ANALYTICS_RESEARCH.md §5): a
data-validation dropdown on the Dashboard drives MATCH/INDEX here. No form
controls, no ActiveX, no VBA — so it works on Excel for Windows, Mac and the
web alike.

Layout, all driven off the block registry so the Dashboard never hardcodes a
row number:

  RAW QUARTERLY   one row per series, columns C..V  = Q1..Q20
  RAW ANNUAL      one row per series, columns C..G  = Y1..Y5
  CURRENT         one cell per series, column C     = the selected period
  ARC             one row per series, columns C..V  = what a view-following
                  chart plots (20 quarterly points, or 5 annual points padded
                  with NA() so the line simply stops instead of dropping to 0)
  CVP             the cost-volume-profit helper for the SELECTED period
"""
from __future__ import annotations

from typing import Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

from . import design
from .break_even_sheet import BREAK_EVEN_STATEMENT
from .data import DraftWorkbookData, text
from .excel_utils import (
  ANNUAL_START_COL,
  FINMO_SHEET,
  FIRST_LIVE_COL,
  WorkbookBuildContext,
  create_sheet,
  local_ref,
  ref,
  write_section_header,
)
from .finmo_ratios import RATIOS_STATEMENT

CALC_SHEET = "Calc"
CALC_KEY = "__calc__"

QUARTERS = 20
YEARS = 5
FIRST_COL = 3                      # column C
LAST_Q_COL = FIRST_COL + QUARTERS - 1
LAST_Y_COL = FIRST_COL + YEARS - 1

#: Selector cells (they live on the Dashboard so a person can change them).
SEL_VIEW = "$C$3"
SEL_QUARTER = "$C$4"
SEL_YEAR = "$C$5"

#: (key, display, source) — source is ("finmo", statement, label) or
#: ("schedule", sheet, key) or ("derived", expression-builder-name).
_SERIES: List[Tuple[str, str, tuple]] = [
  ("revenue", "Revenue", ("finmo", "Income Statement", "Revenue")),
  ("cogs", "Cost of Goods Sold", ("finmo", "Income Statement", "Cost of Goods Sold")),
  ("gross_profit", "Gross Profit", ("finmo", "Income Statement", "Gross Profit")),
  ("marketing", "Marketing", ("finmo", "Income Statement", "Marketing")),
  ("r_and_d", "Research & Development", ("finmo", "Income Statement", "Research & Development")),
  ("lease", "Lease/Rent", ("finmo", "Income Statement", "Lease/Rent")),
  ("payroll", "Payroll", ("finmo", "Income Statement", "Payroll")),
  ("g_and_a", "General & Administrative", ("finmo", "Income Statement", "General & Administrative")),
  ("ebitda", "EBITDA", ("finmo", "Income Statement", "EBITDA")),
  ("depreciation", "Depreciation", ("finmo", "Income Statement", "Depreciation")),
  ("interest", "Interest", ("finmo", "Income Statement", "Interest")),
  ("taxes", "Taxes", ("finmo", "Income Statement", "Taxes")),
  ("net_income", "Net Income", ("finmo", "Income Statement", "Net Income")),
  ("cash", "Ending Cash", ("finmo", "Balance Sheet", "Cash")),
  ("total_assets", "Total Assets", ("finmo", "Balance Sheet", "Total Assets")),
  ("total_equity", "Total Equity", ("finmo", "Balance Sheet", "Total Equity")),
  ("operating_cf", "Operating Cash Flow", ("finmo", "Cash Flow", "Operating Cash Flow")),
  ("net_cf", "Net Cash Flow", ("finmo", "Cash Flow", "Net Cash Flow")),
  ("capex", "Capital Expenditures", ("finmo", "Cash Flow", "Capital Expenditures")),
  ("debt_issuance", "Debt Issuance", ("finmo", "Cash Flow", "Debt Issuance (New Borrowing)")),
  ("debt_repayment", "Debt Repayment", ("finmo", "Cash Flow", "Debt Repayment")),
  ("equity_in", "Owner Equity In", ("finmo", "Cash Flow", "Equity")),
  ("distributions", "Distributions", ("finmo", "Cash Flow", "Distributions")),
  ("lease_principal", "Lease Principal", ("finmo", "Cash Flow", "Capital Lease Principal Payments")),
  # analysis blocks
  ("be_revenue", "Break-Even Revenue", ("finmo", BREAK_EVEN_STATEMENT, "Break-Even Revenue")),
  ("be_cash", "Cash Break-Even Revenue", ("finmo", BREAK_EVEN_STATEMENT, "Cash Break-Even Revenue")),
  ("be_fixed", "Fixed Costs", ("finmo", BREAK_EVEN_STATEMENT, "Fixed Costs")),
  ("be_cm", "Contribution Margin Ratio", ("finmo", BREAK_EVEN_STATEMENT, "Contribution Margin Ratio")),
  ("be_mos", "Margin of Safety", ("finmo", BREAK_EVEN_STATEMENT, "Margin of Safety")),
  ("gross_margin", "Gross Margin", ("finmo", RATIOS_STATEMENT, "Gross Margin")),
  ("ebitda_margin", "EBITDA Margin", ("finmo", RATIOS_STATEMENT, "EBITDA Margin")),
  ("net_margin", "Net Margin", ("finmo", RATIOS_STATEMENT, "Net Margin")),
  ("total_debt", "Total Debt", ("finmo", RATIOS_STATEMENT, "Total Debt")),
  ("net_debt", "Net Debt", ("finmo", RATIOS_STATEMENT, "Net Debt")),
  ("current_ratio", "Current Ratio", ("finmo", RATIOS_STATEMENT, "Current Ratio")),
  ("dscr", "DSCR", ("finmo", RATIOS_STATEMENT, "Debt Service Coverage Ratio (DSCR)")),
  ("debt_to_equity", "Debt to Equity", ("finmo", RATIOS_STATEMENT, "Debt to Equity")),
  ("roe", "Return on Equity", ("finmo", RATIOS_STATEMENT, "Return on Equity")),
  # schedules
  ("headcount", "Headcount (ending FTE)", ("schedule", "Payroll Schedule", "Total Ending FTE")),
  # derived
  ("cash_burn", "Cash Burn (net cash used)", ("derived", "cash_burn")),
  ("cash_built", "Cash Built (net cash added)", ("derived", "cash_built")),
  ("cash_used", "Cash Used (shown negative)", ("derived", "cash_used")),
  ("cash_low", "Cash Low Point (20 quarters)", ("derived", "cash_low")),
]

_PERCENT_KEYS = {"gross_margin", "ebitda_margin", "net_margin", "be_mos", "be_cm"}
_RATIO_KEYS = {"current_ratio", "dscr", "debt_to_equity"}
_UNIT_KEYS = {"headcount"}


def _fmt_for(key: str) -> str:
  if key in _PERCENT_KEYS:
    return design.FMT_PERCENT
  if key in _RATIO_KEYS:
    return design.FMT_RATIO
  if key in _UNIT_KEYS:
    return design.FMT_UNITS
  return design.FMT_MONEY


def _derived(kind: str, ctx: WorkbookBuildContext, col: int) -> str:
  """Series the statements do not carry directly."""
  cf = ctx.finmo_row("Cash Flow", "Net Cash Flow")
  cash = ctx.finmo_row("Balance Sheet", "Cash")
  if not cf or not cash:
    return "=0"
  ncf = ref(FINMO_SHEET, cf, col)
  if kind == "cash_burn":
    return f"=-MIN(0,{ncf})"
  if kind == "cash_built":
    return f"=MAX(0,{ncf})"
  if kind == "cash_used":
    return f"=MIN(0,{ncf})"
  if kind == "cash_low":
    # the same figure in every column: the trough is a property of the arc
    first = ref(FINMO_SHEET, cash, FIRST_LIVE_COL)
    last = ref(FINMO_SHEET, cash, FIRST_LIVE_COL + QUARTERS - 1)
    return f"=MIN({first}:{last})"
  return "=0"


def _source_ref(source: tuple, ctx: WorkbookBuildContext, col: int) -> Optional[str]:
  kind = source[0]
  if kind == "finmo":
    row = ctx.finmo_row(source[1], source[2])
    return ref(FINMO_SHEET, row, col) if row else None
  if kind == "schedule":
    row = ctx.schedule_row(source[1], source[2])
    return ref(source[1], row, col) if row else None
  return None


def _revenue_lines(data: DraftWorkbookData, ctx: WorkbookBuildContext):
  """(display, revenue-row, break-even-units-row) per line of business.

  The app builds anything from one line to N, so the dashboard cannot assume
  either shape: this is what lets it show a single revenue line for a one-line
  business and a stacked mix for a multi-line one.
  """
  out = []
  seen = set()
  for source in data.revenue_rows or []:
    if text(source.get("driver")) != "Unit Price":
      continue
    slot = text(source.get("revenue_slot_key")) or ""
    if not slot or slot in seen:
      continue
    seen.add(slot)
    display = " / ".join(x for x in [text(source.get("lob")), text(source.get("product"))] if x)
    rev_row = ctx.schedule_row("Revenue Drivers", f"{slot}::Revenue")
    be_row = 0
    for label, row in (ctx.finmo_rows.get(BREAK_EVEN_STATEMENT) or {}).items():
      if label.startswith("Break-Even Units") and label.endswith(display):
        be_row = row
        break
    if rev_row:
      out.append((display or slot, rev_row, be_row))
  return out


def build_calc_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  ws = create_sheet(wb, CALC_SHEET)
  ws.sheet_state = "hidden"
  ws.column_dimensions["A"].width = 34
  ws.column_dimensions["B"].width = 12
  for col in range(FIRST_COL, LAST_Q_COL + 1):
    ws.column_dimensions[get_column_letter(col)].width = 13

  title = ws.cell(row=1, column=1, value="Calc - dashboard data engine")
  title.font = design.font("title")
  sub = ws.cell(row=2, column=1,
                value="Hidden by design. Every dashboard card and chart reads from this sheet; "
                      "the selector region below slices the chosen period out of the raw series.")
  sub.font = design.font("subtitle")

  rows: Dict[str, int] = {}

  # ---------------- selector region ----------------------------------------
  write_section_header(ws, 4, "Selector (driven from the Dashboard)", end_col=LAST_Q_COL)
  def _kv(row: int, label: str, formula: str, key: str, fmt: str = design.FMT_GENERAL) -> None:
    ws.cell(row=row, column=1, value=label).font = design.font("label")
    cell = ws.cell(row=row, column=2, value=formula)
    design.calculated_cell(cell, number_format=fmt)
    rows[key] = row

  dash = "Dashboard"
  _kv(5, "View", f"={dash}!{SEL_VIEW}", "sel_view")
  _kv(6, "Selected quarter", f"={dash}!{SEL_QUARTER}", "sel_quarter")
  _kv(7, "Selected year", f"={dash}!{SEL_YEAR}", "sel_year")
  _kv(8, "Is annual view", f'=IF({local_ref(5, 2)}="Annual",1,0)', "is_annual")
  # The period LABEL rows are also the chart categories, so they are written
  # before the index that MATCHes against them.
  ws.cell(row=10, column=1, value="Quarter labels").font = design.font("note")
  for i in range(QUARTERS):
    ws.cell(row=10, column=FIRST_COL + i, value=f"Q{i + 1}").font = design.font("colhead_sub")
  rows["q_labels"] = 10
  ws.cell(row=11, column=1, value="Year labels").font = design.font("note")
  for i in range(YEARS):
    ws.cell(row=11, column=FIRST_COL + i, value=f"Y{i + 1}").font = design.font("colhead_sub")
  rows["y_labels"] = 11

  q_label_range = f"{local_ref(10, FIRST_COL, abs_ref=True)}:{local_ref(10, LAST_Q_COL, abs_ref=True)}"
  y_label_range = f"{local_ref(11, FIRST_COL, abs_ref=True)}:{local_ref(11, LAST_Y_COL, abs_ref=True)}"
  _kv(12, "Selected index (within view)",
      f'=IFERROR(IF({local_ref(8, 2)}=1,MATCH({local_ref(7, 2)},{y_label_range},0),'
      f'MATCH({local_ref(6, 2)},{q_label_range},0)),1)', "sel_index")
  _kv(13, "Selected period label",
      f'=IF({local_ref(8, 2)}=1,{local_ref(7, 2)},{local_ref(6, 2)})', "sel_label")
  _kv(14, "Points in view", f"=IF({local_ref(8, 2)}=1,{YEARS},{QUARTERS})", "points")

  idx_ref = local_ref(12, 2, abs_ref=True)
  idx = idx_ref
  is_annual = local_ref(8, 2, abs_ref=True)

  # ---------------- raw quarterly ------------------------------------------
  row = 17
  write_section_header(ws, row - 1, "Raw series - quarterly (Q1..Q20)", end_col=LAST_Q_COL)
  q_rows: Dict[str, int] = {}
  for key, display, source in _SERIES:
    ws.cell(row=row, column=1, value=display).font = design.font("label")
    for i in range(QUARTERS):
      col = FIRST_COL + i
      finmo_col = FIRST_LIVE_COL + i
      if source[0] == "derived":
        formula = _derived(source[1], ctx, finmo_col)
      else:
        src = _source_ref(source, ctx, finmo_col)
        formula = f"=IFERROR(IF(ISTEXT({src}),0,{src}),0)" if src else "=0"
      cell = ws.cell(row=row, column=col, value=formula)
      design.calculated_cell(cell, number_format=_fmt_for(key))
    q_rows[key] = row
    row += 1

  # ---------------- raw annual ---------------------------------------------
  row += 1
  write_section_header(ws, row - 1, "Raw series - annual (Y1..Y5)", end_col=LAST_Q_COL)
  y_rows: Dict[str, int] = {}
  for key, display, source in _SERIES:
    ws.cell(row=row, column=1, value=display).font = design.font("label")
    for i in range(YEARS):
      col = FIRST_COL + i
      finmo_col = ANNUAL_START_COL + i
      if source[0] == "derived":
        formula = _derived(source[1], ctx, finmo_col)
      else:
        src = _source_ref(source, ctx, finmo_col)
        formula = f"=IFERROR(IF(ISTEXT({src}),0,{src}),0)" if src else "=0"
      cell = ws.cell(row=row, column=col, value=formula)
      design.calculated_cell(cell, number_format=_fmt_for(key))
    y_rows[key] = row
    row += 1

  # ---------------- current view (one cell per series) ---------------------
  row += 1
  write_section_header(ws, row - 1, "CURRENT VIEW - the selected period", end_col=LAST_Q_COL)
  cur_rows: Dict[str, int] = {}
  for key, display, _ in _SERIES:
    ws.cell(row=row, column=1, value=display).font = design.font("label")
    q_range = f"{local_ref(q_rows[key], FIRST_COL, abs_ref=True)}:{local_ref(q_rows[key], LAST_Q_COL, abs_ref=True)}"
    y_range = f"{local_ref(y_rows[key], FIRST_COL, abs_ref=True)}:{local_ref(y_rows[key], LAST_Y_COL, abs_ref=True)}"
    cell = ws.cell(row=row, column=FIRST_COL,
                   value=f"=IF({is_annual}=1,INDEX({y_range},{idx}),INDEX({q_range},{idx}))")
    design.calculated_cell(cell, number_format=_fmt_for(key))
    cur_rows[key] = row
    row += 1

  # ---------------- arc (what a view-following chart plots) ----------------
  row += 1
  write_section_header(ws, row - 1, "ARC - the series as the selected view", end_col=LAST_Q_COL)
  ws.cell(row=row, column=1, value="Period label").font = design.font("label")
  for i in range(QUARTERS):
    col = FIRST_COL + i
    q_lbl = local_ref(10, col, abs_ref=True)
    y_lbl = local_ref(11, FIRST_COL + i, abs_ref=True) if i < YEARS else '""'
    ws.cell(row=row, column=col,
            value=f'=IF({is_annual}=1,{y_lbl if i < YEARS else chr(34) + chr(34)},{q_lbl})')
  arc_labels_row = row
  row += 1
  arc_rows: Dict[str, int] = {}
  for key, display, _ in _SERIES:
    ws.cell(row=row, column=1, value=display).font = design.font("label")
    for i in range(QUARTERS):
      col = FIRST_COL + i
      q_cell = local_ref(q_rows[key], col, abs_ref=True)
      if i < YEARS:
        y_cell = local_ref(y_rows[key], FIRST_COL + i, abs_ref=True)
        formula = f"=IF({is_annual}=1,{y_cell},{q_cell})"
      else:
        formula = f"=IF({is_annual}=1,NA(),{q_cell})"
      cell = ws.cell(row=row, column=col, value=formula)
      design.calculated_cell(cell, number_format=_fmt_for(key))
    arc_rows[key] = row
    row += 1

  # ---------------- per line of business -----------------------------------
  lines = _revenue_lines(data, ctx)
  line_rows: List[Dict[str, int]] = []
  if lines:
    row += 1
    write_section_header(ws, row - 1, "Revenue by line of business - quarterly", end_col=LAST_Q_COL)
    for display, rev_row, be_row in lines:
      ws.cell(row=row, column=1, value=display).font = design.font("label")
      for i in range(QUARTERS):
        src = ref("Revenue Drivers", rev_row, FIRST_LIVE_COL + i)
        cell = ws.cell(row=row, column=FIRST_COL + i, value=f"=IFERROR({src},0)")
        design.calculated_cell(cell, number_format=design.FMT_MONEY)
      line_rows.append({"revenue": row, "display_row": row, "be_units": 0})
      row += 1

    row += 1
    write_section_header(ws, row - 1,
                         "Break-even volume by line at the planned mix - selected period",
                         end_col=LAST_Q_COL)
    be_first = row
    for idx, (display, _rev, be_row) in enumerate(lines):
      ws.cell(row=row, column=1, value=display).font = design.font("label")
      if be_row:
        q_src = ",".join([])  # placeholder, built below
        parts_q = f"{ref(FINMO_SHEET, be_row, FIRST_LIVE_COL)}:{ref(FINMO_SHEET, be_row, FIRST_LIVE_COL + QUARTERS - 1)}"
        parts_y = f"{ref(FINMO_SHEET, be_row, ANNUAL_START_COL)}:{ref(FINMO_SHEET, be_row, ANNUAL_START_COL + YEARS - 1)}"
        formula = (f"=IFERROR(IF({is_annual}=1,INDEX({parts_y},{idx_ref}),"
                   f"INDEX({parts_q},{idx_ref})),0)")
      else:
        formula = "=0"
      cell = ws.cell(row=row, column=FIRST_COL, value=formula)
      design.calculated_cell(cell, number_format=design.FMT_UNITS)
      line_rows[idx]["be_units"] = row
      row += 1
    be_last = row - 1
  else:
    be_first = be_last = 0

  # ---------------- category blocks for the SELECTED period ----------------
  # A chart series needs ONE contiguous range, so the period-specific
  # breakdowns get their own little ordered blocks that read the current-view
  # cells. They reslice with the selector like everything else here.
  row += 1
  write_section_header(ws, row - 1, "Cost structure - the selected period", end_col=LAST_Q_COL)
  cost_first = row
  for key, label in (("cogs", "Cost of Goods Sold"), ("payroll", "Payroll"),
                     ("lease", "Lease/Rent"), ("g_and_a", "General & Administrative"),
                     ("marketing", "Marketing"), ("r_and_d", "Research & Development"),
                     ("depreciation", "Depreciation"), ("interest", "Interest")):
    ws.cell(row=row, column=1, value=label).font = design.font("label")
    cell = ws.cell(row=row, column=FIRST_COL,
                   value=f"={local_ref(cur_rows[key], FIRST_COL, abs_ref=True)}")
    design.calculated_cell(cell, number_format=design.FMT_MONEY)
    row += 1
  cost_last = row - 1

  row += 1
  write_section_header(ws, row - 1, "Sources (+) and uses (-) of cash - the selected period", end_col=LAST_Q_COL)
  su_first = row
  for key, label, sign in (("operating_cf", "Operating cash flow", 1),
                           ("debt_issuance", "New borrowing", 1),
                           ("equity_in", "Owner equity in", 1),
                           ("capex", "Capital expenditures", -1),
                           ("debt_repayment", "Debt repayment", -1),
                           ("lease_principal", "Lease principal", -1),
                           ("distributions", "Distributions", -1)):
    ws.cell(row=row, column=1, value=label).font = design.font("label")
    src = local_ref(cur_rows[key], FIRST_COL, abs_ref=True)
    formula = f"={src}" if sign > 0 else f"=-ABS({src})"
    cell = ws.cell(row=row, column=FIRST_COL, value=formula)
    design.calculated_cell(cell, number_format=design.FMT_MONEY)
    row += 1
  su_last = row - 1

  # ---------------- CVP helper for the SELECTED period ---------------------
  row += 1
  write_section_header(ws, row - 1, "Cost-volume-profit - the selected period", end_col=LAST_Q_COL)
  cvp = _write_cvp(ws, row, cur_rows)
  row = cvp["last"] + 2

  registry = {"__labels_q__": 10, "__labels_y__": 11, "__arc_labels__": arc_labels_row}
  registry.update({f"q::{k}": v for k, v in q_rows.items()})
  registry.update({f"y::{k}": v for k, v in y_rows.items()})
  registry.update({f"cur::{k}": v for k, v in cur_rows.items()})
  registry.update({f"arc::{k}": v for k, v in arc_rows.items()})
  registry.update({f"cvp::{k}": v for k, v in cvp.items()})
  registry.update({f"sel::{k}": v for k, v in rows.items()})
  registry.update({"cost_first": cost_first, "cost_last": cost_last,
                   "su_first": su_first, "su_last": su_last,
                   "line_count": len(lines),
                   "be_units_first": be_first, "be_units_last": be_last})
  for i, entry in enumerate(line_rows):
    registry[f"line::{i}::revenue"] = entry["revenue"]
    registry[f"line::{i}::be_units"] = entry["be_units"]
  ctx.schedule_rows[CALC_KEY] = registry


def _write_cvp(ws, start_row: int, cur_rows: Dict[str, int]) -> Dict[str, int]:
  """The CVP series for the SELECTED period.

  The revenue grid is ANCHORED on the two points that matter - break-even and
  planned revenue are exact grid points - so the shaded profit and loss regions
  meet exactly at the crossing instead of a step or two away from it. The grid
  stays monotonic whether the plan sits above break-even or below it, because
  the anchors are taken as MIN/MAX rather than assumed in order.

  Columns: A revenue (x) | B revenue line | C total cost | D band base
           E loss band | F profit band | G break-even marker | H planned marker
  D/E/F are a stacked area under the lines: the base is invisible, the loss
  band is a red wash where cost exceeds revenue, the profit band a blue wash
  where revenue exceeds cost. Only one of the two is non-zero at any x.
  """
  fixed = local_ref(cur_rows["be_fixed"], FIRST_COL, abs_ref=True)
  cm = local_ref(cur_rows["be_cm"], FIRST_COL, abs_ref=True)
  be = local_ref(cur_rows["be_revenue"], FIRST_COL, abs_ref=True)
  planned = local_ref(cur_rows["revenue"], FIRST_COL, abs_ref=True)

  header = start_row
  for i, name in enumerate(("Revenue (x)", "Total revenue", "Total cost", "Band base",
                            "Loss band", "Profit band", "Break-even point", "Planned point")):
    ws.cell(row=header, column=1 + i, value=name).font = design.font("note")

  anchors = header + 1
  ws.cell(row=anchors, column=1, value="Anchors").font = design.font("note")
  ws.cell(row=anchors, column=2, value=f"=MIN({be},{planned})")     # lo
  ws.cell(row=anchors, column=3, value=f"=MAX({be},{planned})")     # hi
  ws.cell(row=anchors, column=4, value=f"=MAX({be},{planned})*1.45")  # x max
  lo = local_ref(anchors, 2, abs_ref=True)
  hi = local_ref(anchors, 3, abs_ref=True)
  xmax = local_ref(anchors, 4, abs_ref=True)

  # 0 .. lo (3 steps) | lo .. hi (3 steps) | hi .. xmax (3 steps) = 10 points,
  # with `lo` and `hi` landing exactly on points 4 and 7.
  fractions = [("=0", None)]
  for k in (1, 2, 3):
    fractions.append((f"={lo}*{k}/3", None))
  for k in (1, 2, 3):
    fractions.append((f"={lo}+({hi}-{lo})*{k}/3", None))
  for k in (1, 2, 3):
    fractions.append((f"={hi}+({xmax}-{hi})*{k}/3", None))

  first = anchors + 1
  for i, (formula, _) in enumerate(fractions):
    r = first + i
    ws.cell(row=r, column=1, value=formula)
    x = local_ref(r, 1)
    ws.cell(row=r, column=2, value=f"={x}")
    ws.cell(row=r, column=3, value=f"={fixed}+(1-{cm})*{x}")
    rev, cost = local_ref(r, 2), local_ref(r, 3)
    ws.cell(row=r, column=4, value=f"=MIN({rev},{cost})")
    ws.cell(row=r, column=5, value=f"=MAX(0,{cost}-{rev})")
    ws.cell(row=r, column=6, value=f"=MAX(0,{rev}-{cost})")
    # The markers sit on the grid points that ARE break-even and planned
    # revenue; every other point is NA() so the series shows a single dot.
    tol = f"MAX(1,{be})*0.0001"
    ws.cell(row=r, column=7, value=f"=IF(ABS({x}-{be})<={tol},{cost},NA())")
    ws.cell(row=r, column=8, value=f"=IF(ABS({x}-{planned})<=MAX(1,{planned})*0.0001,{rev},NA())")
  last = first + len(fractions) - 1

  for r in range(header, last + 1):
    for c in range(1, 9):
      cell = ws.cell(row=r, column=c)
      if isinstance(cell.value, str) and cell.value.startswith("="):
        cell.number_format = design.FMT_MONEY
  return {"hdr": header, "first": first, "last": last, "anchors": anchors}
