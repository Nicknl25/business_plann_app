"""X5 — the Valuation sheet: a discounted-cash-flow the reader can audit.

A DCF is mostly assumptions, so this sheet is built to make the difference
visible rather than to look authoritative. Every input is one of three things,
and the sheet says which on its face:

  GROUNDED (model)   the cash flows, the tax rate, the cost of debt and the
                     capital weights come from the model's own cells
  GROUNDED (market)  the risk-free rate is the 10-year Treasury from FRED
  ASSUMPTION         the equity risk premium, the size and company-specific
                     premiums, the terminal growth rate and the exit multiple
                     come from `valuation_reference_constants`, each carrying
                     its published source and the date that figure was published

Cost of equity is a BUILD-UP, not a derived beta: `rf + ERP + size + specific`.
That is one method for every client. A beta derived from public comparables
would exist for about a quarter of our clients and would be two tickers wide for
most of them, so it would make the method depend on the industry a client
happens to be in.

Universality is the other design rule. Every reference input resolves through a
NAICS-specific row where one exists and the ALL default where it does not, so a
landscaper with no comparable data still gets a complete valuation - never a
gap, never a crash, never "no valuation".
"""
from __future__ import annotations

import json
import os
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.chart import Reference

from . import design
from .data import DraftWorkbookData, text
from .excel_utils import (
  ANNUAL_START_COL,
  FINMO_SHEET,
  FIRST_LIVE_COL,
  PERIOD_COUNT,
  WorkbookBuildContext,
  create_sheet,
  local_ref,
  ref,
  write_section_header,
)
from .finmo_ratios import RATIOS_STATEMENT

VALUATION_SHEET = "Valuation"
QUARTERS = 20
LAST_Q_COL = FIRST_LIVE_COL + QUARTERS - 1

#: Fallback figures used only when the reference table is unreachable at build
#: time. They are the values the loader seeded, and the sheet labels them as a
#: fallback so a reader is never told a stale constant came from a live source.
_FALLBACK: Dict[str, Dict[str, Any]] = {
  "risk_free_rate": {"value": 0.0472, "citation": "FRED DGS10 (build-time fallback)", "as_of": "2026-08-17"},
  "equity_risk_premium": {"value": 0.0428, "citation": "Damodaran implied ERP (build-time fallback)", "as_of": "2026-08-01"},
  "size_premium_micro_cap": {"value": 0.112, "citation": "Kroll CRSP deciles (build-time fallback)", "as_of": "2023-12-31"},
  "company_specific_risk_premium": {"value": 0.03, "citation": "Build-up judgment component", "as_of": ""},
  "terminal_growth_rate": {"value": 0.023, "citation": "FRED-derived (build-time fallback)", "as_of": "2026-08-18", "max": 0.0428},
  "exit_multiple_sde": {"value": 2.7, "citation": "BizBuySell Insight Report (build-time fallback)", "as_of": "2026-06-30"},
  "wacc_minus_growth_floor": {
    "value": 0.03,
    "citation": "Structural guard: as WACC approaches the growth rate the perpetuity "
                "value diverges, so the method is withheld below this spread", "as_of": ""},
  "maintenance_capex_percent_of_revenue": {
    "value": 0.028,
    "citation": "Maintenance capital expenditure floor. This model's depreciation exceeds "
                "its capital spending, so an unfloored free cash flow would add back "
                "wear the business never actually replaces", "as_of": ""},
}


def _load_constants(naics: str) -> Dict[str, Dict[str, Any]]:
  """Resolve every constant for THIS client: the NAICS-specific row when one
  exists, otherwise the ALL default. Longest matching prefix wins, so an
  8111-scoped multiple beats the ALL one for NAICS 811111 and a landscaper with
  no scoped row still gets a complete set."""
  resolved: Dict[str, Dict[str, Any]] = {}
  try:
    from dotenv import load_dotenv  # type: ignore
    import mysql.connector  # type: ignore

    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"))
    conn = mysql.connector.connect(
      host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
      password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"),
    )
    cur = conn.cursor(dictionary=True)
    cur.execute(
      "SELECT constant_key, applies_to, value_min, value_default, value_max, "
      "source_citation, source_as_of, data_source, confidence_tier "
      "FROM valuation_reference_constants WHERE active=1"
    )
    digits = "".join(ch for ch in (naics or "") if ch.isdigit())
    for row in cur.fetchall():
      scope = str(row["applies_to"] or "ALL")
      if scope != "ALL" and not digits.startswith(scope):
        continue
      key = str(row["constant_key"])
      current = resolved.get(key)
      # A longer NAICS prefix is more specific and wins; ALL is the floor.
      if current and len(str(current.get("scope") or "")) >= len(scope if scope != "ALL" else ""):
        if not (scope != "ALL" and current.get("scope") == "ALL"):
          continue
      resolved[key] = {
        "value": float(row["value_default"]),
        "min": float(row["value_min"]) if row["value_min"] is not None else None,
        "max": float(row["value_max"]) if row["value_max"] is not None else None,
        "citation": text(row["source_citation"]),
        "as_of": text(row["source_as_of"]),
        "source": text(row["data_source"]),
        "scope": "" if scope == "ALL" else scope,
        "tier": text(row["confidence_tier"]),
      }
    cur.close()
    conn.close()
  except Exception:
    pass
  for key, fallback in _FALLBACK.items():
    if key not in resolved:
      resolved[key] = {
        "value": fallback["value"], "min": None, "max": fallback.get("max"),
        "citation": fallback["citation"], "as_of": fallback.get("as_of", ""),
        "source": "build_time_fallback", "scope": "", "tier": "low",
      }
  return resolved


def _owner_compensation(data: DraftWorkbookData) -> float:
  """Owner compensation PER QUARTER.

  Owner compensation sits inside payroll (prior ruling), so seller's
  discretionary earnings adds it back - that is the base main-street buyers
  actually price, and the exit multiples we hold are SDE multiples.

  The stored field is MONTHLY (`capture_receipt.py:44` records the unit as
  "month"; `intake_consult.py:16217` divides an annual figure by 12 to write
  it), so a quarter is three of them. Reading it as quarterly understates SDE
  three-fold and would drag every multiple-based valuation down with it.
  """
  raw = (data.draft_row or {}).get("financials_json")
  try:
    parsed = json.loads(raw) if isinstance(raw, str) else (raw or {})
    value = parsed.get("owner_compensation")
    return float(value) * 3.0 if value not in (None, "") else 0.0
  except Exception:
    return 0.0


def _naics(data: DraftWorkbookData) -> str:
  raw = (data.draft_row or {}).get("operating_model_json")
  try:
    parsed = json.loads(raw) if isinstance(raw, str) else (raw or {})
    return text(parsed.get("business_naics_6"))
  except Exception:
    return ""


def build_valuation_sheet(wb, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  if not ctx.finmo_rows.get("Income Statement"):
    return
  naics = _naics(data)
  const = _load_constants(naics)
  owner_comp_q = _owner_compensation(data)

  ws = create_sheet(wb, VALUATION_SHEET)
  ws.sheet_view.showGridLines = False
  ws.sheet_view.zoomScale = 90
  ws.column_dimensions["A"].width = 38
  ws.column_dimensions["B"].width = 14
  for col in range(3, 30):
    ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = 12
  design.title_block(
    ws, "Valuation", "Discounted cash flow built from this model's own figures. "
    "Every input below is labelled GROUNDED or ASSUMPTION with its source.")
  design.page_setup(ws, landscape=True, fit_width=True, footer=f"{data.business_name} - Valuation")

  def fin(statement: str, label: str, col: int) -> str:
    row = ctx.finmo_row(statement, label)
    return ref(FINMO_SHEET, row, col) if row else "0"

  rows: Dict[str, int] = {}

  def put(row: int, label: str, value, fmt: str, *, note: str = "", input_cell: bool = False,
          bold: bool = False) -> None:
    cell_label = ws.cell(row=row, column=1, value=label)
    cell_label.font = design.font("label_strong" if bold else "label")
    cell = ws.cell(row=row, column=2, value=value)
    if input_cell:
      design.input_cell(cell, number_format=fmt)
    else:
      design.calculated_cell(cell, number_format=fmt)
    if note:
      cell_note = ws.cell(row=row, column=3, value=note)
      cell_note.font = design.font("note")
      ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)

  # ---------------- assumptions and their provenance -----------------------
  row = 4
  write_section_header(ws, row, "Inputs — what is grounded and what is assumed", end_col=12)
  row += 1
  header = ("Input", "Value", "Basis", "Source", "As of")
  for idx, name in enumerate(header):
    cell = ws.cell(row=row, column=1 + (idx if idx < 2 else idx + 0), value=name)
    cell.font = design.font("colhead")
    cell.fill = design.fill(design.NAVY)
  row += 1

  def assumption(label: str, key: str, fmt: str, basis: str) -> int:
    nonlocal row
    entry = const[key]
    put(row, label, entry["value"], fmt, input_cell=(basis == "ASSUMPTION"))
    tag = ws.cell(row=row, column=3, value=basis)
    tag.font = design.font("status_good" if basis.startswith("GROUNDED") else "label_strong")
    scope = f" [NAICS {entry['scope']}]" if entry.get("scope") else " [all industries]"
    cite = ws.cell(row=row, column=4, value=(entry["citation"] or "")[:150] + (scope if key.startswith("exit_multiple") else ""))
    cite.font = design.font("note")
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=11)
    as_of = ws.cell(row=row, column=12, value=entry["as_of"] or "—")
    as_of.font = design.font("note")
    rows[key] = row
    row += 1
    return rows[key]

  assumption("Risk-free rate (10-year Treasury)", "risk_free_rate", design.FMT_PERCENT, "GROUNDED (market)")
  assumption("Equity risk premium", "equity_risk_premium", design.FMT_PERCENT, "ASSUMPTION")
  assumption("Size premium (micro-cap)", "size_premium_micro_cap", design.FMT_PERCENT, "ASSUMPTION")
  assumption("Company-specific risk premium", "company_specific_risk_premium", design.FMT_PERCENT, "ASSUMPTION")
  assumption("Terminal growth rate", "terminal_growth_rate", design.FMT_PERCENT, "ASSUMPTION")
  assumption("Exit multiple (x SDE)", "exit_multiple_sde", design.FMT_RATIO, "ASSUMPTION")
  assumption("Maintenance capital expenditure (% of revenue)",
             "maintenance_capex_percent_of_revenue", design.FMT_PERCENT, "ASSUMPTION")
  assumption("Minimum WACC-minus-growth spread", "wacc_minus_growth_floor",
             design.FMT_PERCENT, "ASSUMPTION")

  # cost of debt and tax come from the model, not from the table
  tax_row = ctx.model_input_row("is::Taxes")
  put(row, "Effective tax rate", f"={ref('Model Inputs', tax_row, FIRST_LIVE_COL)}" if tax_row else 0.0,
      design.FMT_PERCENT)
  ws.cell(row=row, column=3, value="GROUNDED (model)").font = design.font("status_good")
  ws.cell(row=row, column=4, value="The model's own effective tax rate").font = design.font("note")
  rows["tax"] = row
  row += 1

  rate_row = ctx.schedule_row("Debt Schedule", "Interest Rate")
  put(row, "Cost of debt (annual)",
      f"={ref('Debt Schedule', rate_row, FIRST_LIVE_COL)}*4" if rate_row else 0.0, design.FMT_PERCENT)
  ws.cell(row=row, column=3, value="GROUNDED (model)").font = design.font("status_good")
  ws.cell(row=row, column=4, value="SBA 7(a) median rate for this industry and state, "
                                   "carried by the model").font = design.font("note")
  rows["kd"] = row
  row += 2

  # ---------------- cost of capital ---------------------------------------
  write_section_header(ws, row, "Cost of capital — a build-up, applied the same way for every business", end_col=12)
  row += 1
  # One accessor for every assumption cell, so ordering inside this function can
  # never break a reference: short aliases resolve to the constant they name.
  _ALIAS = {"maint_capex": "maintenance_capex_percent_of_revenue",
            "floor": "wacc_minus_growth_floor"}

  def b(key: str) -> str:
    return local_ref(rows[_ALIAS.get(key, key)], 2, abs_ref=True)
  put(row, "Cost of equity (build-up)",
      f"={b('risk_free_rate')}+{b('equity_risk_premium')}+{b('size_premium_micro_cap')}+{b('company_specific_risk_premium')}",
      design.FMT_PERCENT, note="risk-free + equity risk premium + size premium + company-specific premium",
      bold=True)
  rows["ke"] = row
  row += 1
  equity_col = ANNUAL_START_COL   # Y1 weights
  eq = fin("Balance Sheet", "Total Equity", equity_col)
  debt = fin(RATIOS_STATEMENT, "Total Debt", equity_col)
  put(row, "Equity weight", f"=IFERROR({eq}/({eq}+{debt}),1)", design.FMT_PERCENT,
      note="book weights from the Year 1 balance sheet")
  rows["we"] = row
  row += 1
  put(row, "Debt weight", f"=1-{local_ref(rows['we'], 2, abs_ref=True)}", design.FMT_PERCENT)
  rows["wd"] = row
  row += 1
  put(row, "WACC",
      f"={b('ke')}*{b('we')}+{b('kd')}*(1-{b('tax')})*{b('wd')}",
      design.FMT_PERCENT, note="cost of equity and after-tax cost of debt, weighted", bold=True)
  rows["wacc"] = row
  row += 1
  put(row, "WACC less terminal growth",
      f"={b('wacc')}-{b('terminal_growth_rate')}", design.FMT_PERCENT,
      note="the perpetuity method is only shown when this clears the 3-point floor")
  rows["spread"] = row
  row += 2

  # ---------------- free cash flow ----------------------------------------
  write_section_header(ws, row, "Unlevered free cash flow", end_col=12)
  fcf_header = row
  row += 1
  for i in range(QUARTERS):
    cell = ws.cell(row=row, column=3 + i, value=f"Q{i + 1}")
    cell.font = design.font("colhead")
    cell.fill = design.fill(design.NAVY)
  ws.cell(row=row, column=3 + QUARTERS, value="Total").font = design.font("colhead")
  ws.cell(row=row, column=3 + QUARTERS).fill = design.fill(design.NAVY)
  label_row = row
  row += 1

  def series(label: str, builder, fmt: str = design.FMT_MONEY, *, total: bool = True,
             bold: bool = False) -> int:
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = design.font("label_strong" if bold else "label")
    for i in range(QUARTERS):
      ws.cell(row=row, column=3 + i, value=builder(FIRST_LIVE_COL + i))
      design.calculated_cell(ws.cell(row=row, column=3 + i), number_format=fmt)
    if total:
      first = local_ref(row, 3)
      last = local_ref(row, 3 + QUARTERS - 1)
      ws.cell(row=row, column=3 + QUARTERS, value=f"=SUM({first}:{last})")
      design.calculated_cell(ws.cell(row=row, column=3 + QUARTERS), number_format=fmt)
    return row

  rows["ebitda"] = series("EBITDA", lambda c: f"={fin('Income Statement', 'EBITDA', c)}")
  row += 1
  rows["sde"] = series(
    "Seller's discretionary earnings (SDE)",
    lambda c: f"={fin('Income Statement', 'EBITDA', c)}+{owner_comp_q}")
  ws.cell(row=rows["sde"], column=2,
          value=f"EBITDA + ${owner_comp_q:,.0f}/qtr owner pay").font = design.font("note")
  row += 1
  rows["ebit"] = series("EBIT", lambda c: f"={fin(RATIOS_STATEMENT, 'EBIT', c)}")
  row += 1
  rows["nopat"] = series(
    "NOPAT (EBIT after tax)",
    lambda c: f"={fin(RATIOS_STATEMENT, 'EBIT', c)}*(1-{b('tax')})")
  row += 1
  rows["dep"] = series("Add: depreciation", lambda c: f"={fin('Income Statement', 'Depreciation', c)}")
  row += 1
  rows["capex"] = series(
    "Less: capital expenditure",
    lambda c: f"=-MAX({fin('Cash Flow', 'Capital Expenditures', c)},"
              f"{fin('Income Statement', 'Revenue', c)}*{b('maint_capex')})")
  row += 1
  rows["nwc"] = series(
    "Less: change in working capital",
    lambda c: f"={fin('Cash Flow', 'Changes in Current Assets', c)}+"
              f"{fin('Cash Flow', 'Changes in Current Liabilities', c)}")
  row += 1
  rows["ufcf"] = series(
    "Unlevered free cash flow",
    lambda c: "=" + "+".join(local_ref(rows[k], 3 + (c - FIRST_LIVE_COL)) for k in ("nopat", "dep", "capex", "nwc")),
    bold=True)
  row += 1
  rows["df"] = series(
    "Discount factor",
    lambda c: f"=1/(1+{b('wacc')}/4)^{c - FIRST_LIVE_COL + 1}",
    fmt=design.FMT_RATIO, total=False)
  row += 1
  rows["pv"] = series(
    "Present value",
    lambda c: f"={local_ref(rows['ufcf'], 3 + (c - FIRST_LIVE_COL))}*"
              f"{local_ref(rows['df'], 3 + (c - FIRST_LIVE_COL))}",
    bold=True)
  row += 2
  return _finish(ws, data, ctx, const, rows, row, b, fin, owner_comp_q, naics)


def _finish(ws, data, ctx, const, rows, row, b, fin, owner_comp_q, naics):
  """Terminal value, the equity bridge, the sensitivity grid and the note."""
  pv_total = f"SUM({local_ref(rows['pv'], 3)}:{local_ref(rows['pv'], 3 + QUARTERS - 1)})"
  ufcf_y5 = f"SUM({local_ref(rows['ufcf'], 3 + 16)}:{local_ref(rows['ufcf'], 3 + 19)})"
  sde_y5 = f"SUM({local_ref(rows['sde'], 3 + 16)}:{local_ref(rows['sde'], 3 + 19)})"
  last_df = local_ref(rows["df"], 3 + QUARTERS - 1)

  def put(label, value, fmt=design.FMT_MONEY, *, note="", bold=False):
    nonlocal row
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = design.font("label_strong" if bold else "label")
    target = ws.cell(row=row, column=2, value=value)
    design.calculated_cell(target, number_format=fmt)
    if note:
      n = ws.cell(row=row, column=3, value=note)
      n.font = design.font("note")
      ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
    key = row
    row += 1
    return key

  write_section_header(ws, row, "Terminal value — both methods, side by side", end_col=12)
  row += 1
  rows["tv_perp"] = put(
    "Terminal value — perpetuity growth",
    f'=IF({b("spread")}<{b("wacc_minus_growth_floor")},"—",({ufcf_y5})*(1+{b("terminal_growth_rate")})/{b("spread")})',
    note="last year's free cash flow grown forever, discounted at WACC less growth; "
         "shown only when the spread clears the floor")
  rows["tv_mult"] = put(
    "Terminal value — exit multiple",
    f"=({sde_y5})*{b('exit_multiple_sde')}",
    note="year-5 seller's discretionary earnings at the multiple businesses of this "
         "kind actually transact at")
  rows["tv_used"] = put(
    "Terminal value used (average of the two)",
    f'=IF(ISNUMBER({local_ref(rows["tv_perp"], 2)}),AVERAGE({local_ref(rows["tv_perp"], 2)},'
    f'{local_ref(rows["tv_mult"], 2)}),{local_ref(rows["tv_mult"], 2)})', bold=True,
    note="neither method is authoritative; the sensitivity grid below is the honest answer")
  rows["tv_pv"] = put("Present value of terminal value",
                      f"={local_ref(rows['tv_used'], 2)}*{last_df}")
  row += 1

  write_section_header(ws, row, "Value", end_col=12)
  row += 1
  rows["pv_ops"] = put("Present value of forecast cash flows", f"={pv_total}")
  rows["ev"] = put("Enterprise value",
                   f"={local_ref(rows['pv_ops'], 2)}+{local_ref(rows['tv_pv'], 2)}", bold=True)
  net_debt = fin(RATIOS_STATEMENT, "Net Debt", FIRST_LIVE_COL + QUARTERS - 1)
  rows["net_debt"] = put("Less: net debt at the end of year 5", f"={net_debt}",
                         note="negative means the business holds more cash than debt, which adds to value")
  rows["equity"] = put("Equity value",
                       f"={local_ref(rows['ev'], 2)}-{local_ref(rows['net_debt'], 2)}", bold=True)
  rows["implied_sde"] = put(
    "Implied multiple of year-5 SDE",
    f'=IFERROR({local_ref(rows["ev"], 2)}/({sde_y5}),"—")', design.FMT_RATIO,
    note="the cross-check: if this is far from the exit multiple above, the two methods disagree")
  row += 1

  # ---------------- sensitivity -------------------------------------------
  write_section_header(ws, row, "Sensitivity — the range is the answer, not the single number", end_col=12)
  row += 1
  ws.cell(row=row, column=1, value="Equity value at WACC (down) x terminal growth (across)").font = design.font("note")
  row += 1
  grid_top = row
  growth_steps = (-0.01, -0.005, 0.0, 0.005, 0.01)
  wacc_steps = (-0.03, -0.015, 0.0, 0.015, 0.03)
  ws.cell(row=grid_top, column=2, value="WACC \\ growth").font = design.font("colhead")
  ws.cell(row=grid_top, column=2).fill = design.fill(design.NAVY)
  for j, gstep in enumerate(growth_steps):
    cell = ws.cell(row=grid_top, column=3 + j, value=f"={b('terminal_growth_rate')}+{gstep}")
    design.calculated_cell(cell, number_format=design.FMT_PERCENT)
    cell.fill = design.fill(design.TINT_1)
  for i, wstep in enumerate(wacc_steps):
    r = grid_top + 1 + i
    head = ws.cell(row=r, column=2, value=f"={b('wacc')}+{wstep}")
    design.calculated_cell(head, number_format=design.FMT_PERCENT)
    head.fill = design.fill(design.TINT_1)
    for j, gstep in enumerate(growth_steps):
      w = local_ref(r, 2, abs_ref=False)
      g = local_ref(grid_top, 3 + j)
      cell = ws.cell(
        row=r, column=3 + j,
        value=(f'=IFERROR(IF(${w[0]}${r}-{g}<{b("wacc_minus_growth_floor")},"—",'
               f'{pv_total}+(({ufcf_y5})*(1+{g})/(${w[0]}${r}-{g}))*{last_df}-{net_debt}),"—")'),
      )
      design.calculated_cell(cell, number_format=design.FMT_MONEY)
  row = grid_top + len(wacc_steps) + 2

  design.footnote(
    ws, row,
    "How to read this: the cash flows, tax rate, cost of debt and capital weights come from this "
    "model. The risk-free rate is the 10-year US Treasury. The equity risk premium, size and "
    "company-specific premiums, terminal growth rate and exit multiple are assumptions - each is "
    "shown above with the source it came from and the date that figure was published, and each is "
    "an amber cell you can change. A valuation of a private business is a range, not a number; the "
    "grid above shows how the answer moves when the two assumptions it is most sensitive to move.")
  row += 2
  design.footnote(
    ws, row,
    "Seller's discretionary earnings (SDE) adds the owner's pay back to EBITDA, because a buyer of "
    "an owner-operated business acquires that role too - and the transaction multiples above are "
    "quoted against SDE.")
  ws.sheet_properties.tabColor = design.NAVY
