"""X1 — THE DESIGN SYSTEM (2026-08-18).

One home for the workbook's visual standard. Ruled in
docs/WORKBOOK_ANALYTICS_RESEARCH.md §1 with Nick's Q1/Q2 rulings baked in:
the dataviz-validated palette, and Calibri pinned (universal on older installs).

Everything visual comes from here — colors, type, number formats, layout
geometry, and the SINGLE DOOR every chart is built through. Sheet modules import
roles, never raw hex; `tests/test_x1_design_system.py` walks the built workbook
and fails on any cell or chart that bypasses this module, so a later sheet
cannot regress the standard without going red.

WHAT THIS MODULE MUST NEVER DO: change a value or a formula. It styles cells;
it does not author them.

Palette provenance — run through dataviz/scripts/validate_palette.js against
the workbook's real surface (white):
  categorical order (blue→orange→aqua→amber→magenta→green→violet→red):
    lightness band PASS · chroma PASS · worst adjacent CVD dE 9.1 PASS ·
    worst adjacent normal-vision dE 19.6 PASS · contrast WARN on aqua/amber/
    magenta -> those slots always ship with a legend or a direct label.
  semantic trio (revenue blue / cost red / attention amber), ALL-PAIRS:
    CVD dE 15.3 PASS · normal-vision dE 20.8 PASS.
  (blue+orange+amber FAILED the normal-vision floor at dE 13.7 — which is why
  the CVP chart uses red for cost, not orange.)
"""
from __future__ import annotations

from typing import Iterable, List, Optional, Sequence

from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.chart.legend import LegendEntry
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
  CharacterProperties,
  Paragraph,
  ParagraphProperties,
  RichTextProperties,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. PALETTE  (no default Excel/Office colors anywhere)
# ---------------------------------------------------------------------------

NAVY = "12314B"           # brand primary — 13.4:1 with white text
NAVY_DEEP = "0C2338"      # cover band
TINT_1 = "E8EEF4"         # KPI tiles, subtotal bands
TINT_2 = "F4F7FA"         # section wash, helper blocks
INK = "1A2733"            # body text — 15.2:1 on white
INK_SECONDARY = "5B6B7B"  # row notes — 5.5:1
INK_MUTED = "71808F"      # axis labels, footnotes — 3.6:1
HAIRLINE = "DDE3E9"       # cell borders
GRID = "EDF1F5"           # chart gridlines - one whisper off the surface
RULE = "C3CCD5"           # axis lines, block dividers
WHITE = "FFFFFF"

#: Editable-input convention (kept — a buyer edits this workbook), restyled.
INPUT_FILL = "FDF3DF"
INPUT_INK = "1B4F8A"

#: Categorical series — the validated ORDER is the colorblind-safety mechanism.
SERIES: Sequence[str] = (
  "2A78D6",  # 1 blue
  "EB6834",  # 2 orange
  "1BAF7A",  # 3 aqua
  "EDA100",  # 4 amber
  "E87BA4",  # 5 magenta
  "008300",  # 6 green
  "4A3AA7",  # 7 violet
  "E34948",  # 8 red
)
SERIES_REVENUE = SERIES[0]
SERIES_COST = SERIES[7]
SERIES_ATTENTION = SERIES[3]
SERIES_REFERENCE = INK_MUTED   # thresholds/bands are chrome, not identity
#: Region fills for the CVP chart. Pale tints of the profit/loss series hues -
#: openpyxl cannot express fill transparency, so the wash is a lighter STEP of
#: the same hue rather than an alpha, which keeps the lines on top readable.
BAND_PROFIT = "D6E4F5"
BAND_LOSS = "F8DCDB"

#: Status — reserved, never reused as a series color, always paired with a word.
STATUS_GOOD = "0CA30C"
STATUS_WARNING = "FAB219"
STATUS_CRITICAL = "D03B3B"
STATUS_GOOD_FILL = "E6F4E6"
STATUS_CRITICAL_FILL = "FBE3E3"
STATUS_NEUTRAL_FILL = TINT_1

#: Every fill and font color the workbook is allowed to use (the guard test's
#: allow-list). Anything else is a bypass.
ALLOWED_FILLS = frozenset({
  NAVY, NAVY_DEEP, TINT_1, TINT_2, WHITE, INPUT_FILL, GRID,
  STATUS_GOOD_FILL, STATUS_CRITICAL_FILL, BAND_PROFIT, BAND_LOSS,
})
ALLOWED_FONT_COLORS = frozenset({
  INK, INK_SECONDARY, INK_MUTED, NAVY, WHITE, INPUT_INK,
  STATUS_GOOD, STATUS_WARNING, STATUS_CRITICAL,
})
ALLOWED_SERIES_COLORS = frozenset(SERIES) | {SERIES_REFERENCE, BAND_PROFIT, BAND_LOSS}

# ---------------------------------------------------------------------------
# 2. TYPOGRAPHY  (Calibri pinned — Nick's Q2)
# ---------------------------------------------------------------------------

FONT_FAMILY = "Calibri"

_ROLES = {
  "cover_title": dict(size=30, bold=True, color=WHITE),
  "cover_sub": dict(size=13, bold=False, color=WHITE),
  "cover_meta_label": dict(size=9, bold=True, color=INK_MUTED),
  "cover_meta_value": dict(size=12, bold=False, color=INK),
  "title": dict(size=18, bold=True, color=NAVY),
  "subtitle": dict(size=10, italic=True, color=INK_SECONDARY),
  "section": dict(size=11, bold=True, color=WHITE),
  "colhead": dict(size=10, bold=True, color=WHITE),
  "colhead_sub": dict(size=9, bold=False, color=INK_MUTED),
  "label": dict(size=11, bold=False, color=INK),
  "label_strong": dict(size=11, bold=True, color=INK),
  "note": dict(size=9, italic=True, color=INK_SECONDARY),
  "value": dict(size=11, bold=False, color=INK),
  "value_strong": dict(size=11, bold=True, color=INK),
  "input": dict(size=11, bold=False, color=INPUT_INK),
  "kpi_label": dict(size=9, bold=True, color=WHITE),
  "kpi_value": dict(size=18, bold=True, color=NAVY),
  "footnote": dict(size=8, italic=True, color=INK_MUTED),
  "status_good": dict(size=11, bold=True, color=STATUS_GOOD),
  "status_bad": dict(size=11, bold=True, color=STATUS_CRITICAL),
}

#: Every (name, size) pair the workbook may use — the guard test's allow-list.
ALLOWED_FONT_SIZES = frozenset(spec["size"] for spec in _ROLES.values())


def font(role: str) -> Font:
  """The Font for a type role. Never construct Font() in a sheet module."""
  spec = _ROLES[role]
  return Font(
    name=FONT_FAMILY,
    size=spec["size"],
    bold=bool(spec.get("bold")),
    italic=bool(spec.get("italic")),
    color=spec["color"],
  )


def fill(color: str) -> PatternFill:
  return PatternFill("solid", fgColor=color)


# ---------------------------------------------------------------------------
# 3. NUMBER FORMATS  (one set, applied by role — replaces 3 competing sniffers)
# ---------------------------------------------------------------------------

FMT_MONEY = '$#,##0;($#,##0);"-"'
FMT_MONEY_SIGNED = '+$#,##0;-$#,##0;"-"'
FMT_PERCENT = '0.0%;(0.0%);"-"'
FMT_RATIO = '0.00"x"'
FMT_UNITS = '#,##0.0;(#,##0.0);"-"'
FMT_INTEGER = '#,##0;(#,##0);"-"'
FMT_DAYS = '#,##0" days"'
FMT_DATE = "mmm yyyy"
FMT_TEXT = "@"
FMT_GENERAL = "General"
#: A numeric quarter index that DISPLAYS as Stub / Q1 ... Q20 (Payroll Schedule
#: column A): the value stays numeric so SUMIFS and the Checks tie-out that key
#: on it are untouched, and a client reads the same period labels as every
#: other sheet.
FMT_QUARTER_INDEX = '[=0]"Stub";"Q"0'

#: Axis formats (no negative/zero sections — Excel axes take the simple form).
FMT_AXIS_MONEY = "$#,##0"
FMT_AXIS_PERCENT = "0%"
FMT_AXIS_UNITS = "#,##0.0"

ALLOWED_NUMBER_FORMATS = frozenset({
  FMT_MONEY, FMT_MONEY_SIGNED, FMT_PERCENT, FMT_RATIO, FMT_UNITS, FMT_INTEGER,
  FMT_DAYS, FMT_DATE, FMT_TEXT, FMT_GENERAL,
  FMT_QUARTER_INDEX,
})

# ---------------------------------------------------------------------------
# 4. LAYOUT
# ---------------------------------------------------------------------------

COL_LABEL_WIDTH = 34
COL_NOTE_WIDTH = 30
COL_PERIOD_WIDTH = 13
COL_ANNUAL_WIDTH = 14
BLOCK_SPACER_ROWS = 1
TITLE_ROW = 1
SUBTITLE_ROW = 2
HEADER_ROW = 4
HEADER_SUB_ROW = 5
FIRST_BLOCK_ROW = 6

_THIN = Side(style="thin", color=HAIRLINE)
_RULE = Side(style="thin", color=RULE)
BORDER_HAIRLINE = Border(top=_THIN, bottom=_THIN)
BORDER_SWATCH = Border(top=_RULE, bottom=_RULE, left=_RULE, right=_RULE)
BORDER_BLOCK_TOP = Border(top=_RULE, bottom=_THIN)

#: One tab color scheme, ordered by the reader's journey rather than Accent1-6.
TAB_COLORS = {
  "Cover": NAVY_DEEP,
  "Dashboard": NAVY,
  "FINMO": NAVY,
  "Break-Even": NAVY,
  "Ratios": NAVY,
  "Valuation": NAVY,
  "Revenue Drivers": INK_SECONDARY,
  "Payroll Schedule": INK_SECONDARY,
  "Debt Schedule": INK_SECONDARY,
  "CapEx Depreciation": INK_SECONDARY,
  "Working Capital": INK_SECONDARY,
  "Cash Equity Schedule": INK_SECONDARY,
  "Model Inputs": INK_SECONDARY,
  "Calc": INK_MUTED,
  "Audit Source": INK_MUTED,
  "Checks": INK_MUTED,
  "Diagnostics": INK_MUTED,
}

#: Reading order (Nick's Q6). Sheets absent from a given build are skipped;
#: sheets not listed keep their build position at the end.
SHEET_ORDER = (
  "Cover", "Dashboard", "FINMO", "Break-Even", "Ratios", "Valuation",
  "Revenue Drivers", "Payroll Schedule", "Debt Schedule", "CapEx Depreciation",
  "Working Capital", "Cash Equity Schedule", "Model Inputs",
  "Audit Source", "Calc", "Checks", "Diagnostics",
)


def apply_sheet_order(wb) -> None:
  """Re-order sheets for reading (build order is unchanged — the builder still
  builds FINMO before the Dashboard so the row registry is populated)."""
  ordered = [name for name in SHEET_ORDER if name in wb.sheetnames]
  ordered += [name for name in wb.sheetnames if name not in ordered]
  wb._sheets = [wb[name] for name in ordered]


def apply_tab_colors(wb) -> None:
  for name, color in TAB_COLORS.items():
    if name in wb.sheetnames:
      wb[name].sheet_properties.tabColor = color


def page_setup(ws, *, landscape: bool = True, fit_width: bool = True,
               title_rows: Optional[str] = None, footer: str = "") -> None:
  """Print/PDF setup — so 'Save as PDF' yields a report, not a grid dump."""
  ws.page_setup.orientation = "landscape" if landscape else "portrait"
  if fit_width:
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
  ws.page_margins.left = ws.page_margins.right = 0.4
  ws.page_margins.top = ws.page_margins.bottom = 0.5
  if title_rows:
    ws.print_title_rows = title_rows
  ws.oddFooter.left.text = footer
  ws.oddFooter.left.size = 8
  ws.oddFooter.left.color = INK_MUTED.lower()
  ws.oddFooter.right.text = "Page &P of &N"
  ws.oddFooter.right.size = 8
  ws.oddFooter.right.color = INK_MUTED.lower()


def base_columns(ws, *, period_start: int, period_end: int, annual_end: int) -> None:
  ws.column_dimensions["A"].width = COL_LABEL_WIDTH
  ws.column_dimensions["B"].width = COL_NOTE_WIDTH
  for col in range(period_start, period_end + 1):
    ws.column_dimensions[get_column_letter(col)].width = COL_PERIOD_WIDTH
  for col in range(period_end + 1, annual_end + 1):
    ws.column_dimensions[get_column_letter(col)].width = COL_ANNUAL_WIDTH


def title_block(ws, title: str, subtitle: str = "") -> None:
  cell = ws.cell(row=TITLE_ROW, column=1, value=title)
  cell.font = font("title")
  if subtitle:
    sub = ws.cell(row=SUBTITLE_ROW, column=1, value=subtitle)
    sub.font = font("subtitle")


def section_band(ws, row: int, title: str, *, end_col: int) -> None:
  ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
  cell = ws.cell(row=row, column=1, value=title)
  cell.fill = fill(NAVY)
  cell.font = font("section")
  cell.alignment = Alignment(horizontal="left", vertical="center")


def input_cell(cell, *, number_format: str = FMT_MONEY) -> None:
  cell.font = font("input")
  cell.fill = fill(INPUT_FILL)
  cell.alignment = Alignment(horizontal="right")
  cell.number_format = number_format


def calculated_cell(cell, *, number_format: str = FMT_MONEY) -> None:
  cell.font = font("value")
  cell.alignment = Alignment(horizontal="right")
  cell.number_format = number_format


def data_row(ws, row: int, *, start_col: int = 1, end_col: int,
             number_format: str = FMT_MONEY, emphasis: bool = False,
             band: bool = False, top_rule: bool = False) -> None:
  """Style one data row. `emphasis` = subtotal (bold + tint + rule)."""
  for col in range(start_col, end_col + 1):
    cell = ws.cell(row=row, column=col)
    if band or emphasis:
      cell.fill = fill(TINT_1 if emphasis else TINT_2)
    if col == 1:
      cell.font = font("label_strong" if emphasis else "label")
      cell.alignment = Alignment(horizontal="left")
    elif col == 2:
      cell.font = font("note")
      cell.alignment = Alignment(horizontal="left")
    else:
      if cell.font is None or cell.font.color is None or cell.font.color.rgb != f"00{INPUT_INK}":
        cell.font = font("value_strong" if emphasis else "value")
      cell.alignment = Alignment(horizontal="right")
      cell.number_format = number_format
    cell.border = BORDER_BLOCK_TOP if (top_rule or emphasis) else BORDER_HAIRLINE


def kpi_tile(ws, row: int, col: int, label: str, formula: str,
             number_format: str, note: str = "", *, span: int = 3) -> None:
  ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
  ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + span - 1)
  lab = ws.cell(row=row, column=col, value=label)
  lab.font = font("kpi_label")
  lab.fill = fill(NAVY)
  lab.alignment = Alignment(horizontal="center", vertical="center")
  val = ws.cell(row=row + 1, column=col, value=formula)
  val.font = font("kpi_value")
  val.fill = fill(TINT_1)
  val.alignment = Alignment(horizontal="center", vertical="center")
  val.number_format = number_format
  if note:
    cell = ws.cell(row=row + 2, column=col, value=note)
    cell.font = font("footnote")
    cell.alignment = Alignment(horizontal="center")


def footnote(ws, row: int, text: str, *, col: int = 1) -> None:
  cell = ws.cell(row=row, column=col, value=text)
  cell.font = font("footnote")
  cell.alignment = Alignment(horizontal="left")


# ---------------------------------------------------------------------------
# 5. THE CHART HELPER — the single door
# ---------------------------------------------------------------------------

MARKER_ATTR = "_design_system"
_LINE_W = 28575          # 2.25pt
_LINE_W_THIN = 12700     # 1pt
_HAIRLINE_W = 9525       # 0.75pt

_KINDS = {"line", "bar", "column", "stacked_column", "scatter", "pie", "area"}


def _char_props(*, size: int, color: str, bold: bool = False) -> CharacterProperties:
  return CharacterProperties(latin=None, sz=size * 100, b=bold, solidFill=color)


def _text_props(*, size: int, color: str, bold: bool = False) -> RichText:
  return RichText(
    bodyPr=RichTextProperties(),
    p=[Paragraph(pPr=ParagraphProperties(defRPr=_char_props(size=size, color=color, bold=bold)), endParaRPr=_char_props(size=size, color=color, bold=bold))],
  )


def _style_axis(axis, *, number_format: Optional[str], gridlines: bool, title: Optional[str],
                axis_line: bool = True) -> None:
  # openpyxl HIDES axes unless delete is explicitly False — the #1 gotcha
  # (docs/WORKBOOK_ANALYTICS_RESEARCH.md §1.5).
  axis.delete = False
  axis.majorTickMark = "none"
  axis.minorTickMark = "none"
  # A modern chart draws no box: the value axis has no rule at all (its
  # gridlines carry the scale) and only the category baseline is drawn.
  axis.spPr = (GraphicalProperties(ln=LineProperties(solidFill=RULE, w=_HAIRLINE_W))
               if axis_line else GraphicalProperties(ln=LineProperties(noFill=True)))
  axis.txPr = _text_props(size=8, color=INK_MUTED)
  if number_format:
    axis.number_format = number_format
  if gridlines:
    axis.majorGridlines = ChartLines(
      spPr=GraphicalProperties(ln=LineProperties(solidFill=GRID, w=_HAIRLINE_W))
    )
  else:
    axis.majorGridlines = None
  if title:
    axis.title = title
    try:
      axis.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=_char_props(size=9, color=INK_MUTED))
    except Exception:
      pass
  else:
    axis.title = None


def chart(kind: str, *, title: str, y_format: Optional[str] = FMT_AXIS_MONEY,
          x_format: Optional[str] = None, y_title: Optional[str] = None,
          x_title: Optional[str] = None, legend: Optional[str] = "b",
          width: float = 16.5, height: float = 8.5):
  """Create a chart with the house style already applied. This is the ONLY way
  a chart may be created — `tests/test_x1_design_system.py` fails the build on
  any chart without the design marker."""
  if kind not in _KINDS:
    raise ValueError(f"unsupported chart kind {kind!r}")
  if kind == "line":
    obj = LineChart()
  elif kind == "area":
    obj = AreaChart()
  elif kind in {"bar", "column", "stacked_column"}:
    obj = BarChart()
    obj.type = "bar" if kind == "bar" else "col"
    if kind == "stacked_column":
      obj.grouping = "stacked"
      obj.overlap = 100
    obj.gapWidth = 60
  elif kind == "scatter":
    obj = ScatterChart()
    obj.scatterStyle = "lineMarker"
  else:
    obj = PieChart()

  obj.title = title
  try:  # title in house type, not Excel's default
    obj.title.tx.rich.p[0].pPr = ParagraphProperties(defRPr=_char_props(size=11, color=INK, bold=True))
  except Exception:
    pass
  obj.width = width
  obj.height = height
  obj.roundedCorners = False
  # No chart-area border, no plot-area fill: the data is the only ink.
  obj.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
  obj.graphicalProperties = GraphicalProperties(ln=LineProperties(noFill=True))
  obj.plot_area_no_fill = True
  # Excel legends every POINT when varyColors is on — the rainbow-legend bug.
  obj.varyColors = False

  if kind != "pie":
    # openpyxl ALWAYS writes x_axis as the category axis and y_axis as the
    # value axis, including for a horizontal bar (type="bar") where Excel
    # then RENDERS them swapped. Mapping by rendered orientation instead of
    # by role sends the number format and the tick-label position to the
    # wrong axis - which is what left the sources-and-uses labels colliding.
    value_axis = obj.y_axis
    cat_axis = obj.x_axis
    _style_axis(value_axis, number_format=y_format, gridlines=True, title=y_title,
                axis_line=False)
    _style_axis(cat_axis, number_format=x_format, gridlines=False, title=x_title)
    if kind in {"bar", "column", "stacked_column"}:
      # Category labels go to the OUTSIDE edge, not against the zero line —
      # otherwise a negative bar's value label lands on top of them (the
      # sources-and-uses collision).
      cat_axis.tickLblPos = "low"
  if legend is None:
    obj.legend = None
  else:
    obj.legend.position = legend
    obj.legend.overlay = False
    obj.legend.txPr = _text_props(size=9, color=INK_SECONDARY)
  setattr(obj, MARKER_ATTR, True)
  return obj


def add_series(obj, values: Reference, *, title: Optional[str] = None,
               end_index: int = 19,
               x_values: Optional[Reference] = None, slot: Optional[int] = None,
               color: Optional[str] = None, dashed: bool = False,
               line: bool = True, marker: Optional[str] = None,
               marker_size: int = 8, labels=False,
               label_position: Optional[str] = None, thin: bool = False,
               no_fill: bool = False, end_label: bool = False):
  """Add one series in the house style. Color comes from the validated slot
  order unless an explicit palette color is passed (semantic charts)."""
  hex_color = color or (SERIES[slot % len(SERIES)] if slot is not None else SERIES[0])
  if no_fill:
    hex_color = SERIES[0]
  if hex_color not in ALLOWED_SERIES_COLORS:
    raise ValueError(f"series color {hex_color!r} is outside the design palette")
  series = Series(values, x_values, title=title) if x_values is not None else Series(values, title=title)
  props = GraphicalProperties()
  if line:
    props.line = LineProperties(solidFill=hex_color, w=_LINE_W_THIN if thin else _LINE_W)
    if dashed:
      props.line.dashStyle = "dash"
  else:
    props.solidFill = hex_color
    props.line = LineProperties(noFill=True)
  if no_fill:
    props = GraphicalProperties(noFill=True, ln=LineProperties(noFill=True))
  series.graphicalProperties = props
  # Excel hollows out a negative bar unless told otherwise, which reads as an
  # unfinished chart on a sources-and-uses view.
  series.invertIfNegative = False
  if marker:
    series.marker = Marker(symbol=marker, size=marker_size,
                           spPr=GraphicalProperties(solidFill=hex_color,
                                                    ln=LineProperties(solidFill=WHITE, w=_HAIRLINE_W)))
  elif isinstance(obj, (LineChart, ScatterChart)):
    series.marker = Marker(symbol="none")
  series.smooth = False
  if labels:
    # openpyxl IGNORES dLbls.numFmt — the format is inherited from the SOURCE
    # CELLS, so callers must format those (research §1.5).
    # labels=True -> the value; labels="name" -> the series name (region
    # annotations), never both, never one per point on a dense series.
    by_name = labels == "name"
    series.dLbls = DataLabelList()
    series.dLbls.showVal = not by_name
    series.dLbls.showSerName = by_name
    series.dLbls.showCatName = False
    series.dLbls.showLegendKey = False
    series.dLbls.showBubbleSize = False
    if label_position:
      series.dLbls.dLblPos = label_position
    series.dLbls.txPr = _text_props(size=9, color=INK_SECONDARY)
  if end_label:
    # Direct labelling: ONE label riding the end of the line instead of a
    # legend row. The modern idiom, and it keeps the eye on the data.
    series.dLbls = DataLabelList(dLbl=[DataLabel(
      idx=end_index, showVal=False, showSerName=True, showCatName=False,
      showLegendKey=False, showBubbleSize=False, showPercent=False,
      txPr=_text_props(size=9, color=INK_SECONDARY))])
    series.dLbls.showVal = False
    series.dLbls.showSerName = False
    series.dLbls.showCatName = False
    series.dLbls.showLegendKey = False
    series.dLbls.showBubbleSize = False
  obj.series.append(series)
  return series


def set_categories(obj, categories: Reference, *, skip: Optional[int] = None) -> None:
  obj.set_categories(categories)
  if skip and skip > 1:
    obj.x_axis.tickLblSkip = skip
    obj.x_axis.tickMarkSkip = skip


def combine(base, overlay):
  """Layer `overlay` on top of `base` (e.g. the CVP lines over the shaded
  profit/loss bands). Both must be category-axis charts so they share axes."""
  if not getattr(base, MARKER_ATTR, False) or not getattr(overlay, MARKER_ATTR, False):
    raise ValueError("both charts must come from design.chart()")
  overlay.y_axis.majorGridlines = None
  # ChartBase defines __iadd__ (not __add__): `+=` appends the overlay to the
  # base chart's plot area. `base + overlay` raises "cannot combine instances
  # of different types" because Serialisable.__add__ requires one type.
  base += overlay
  setattr(base, MARKER_ATTR, True)
  return base


def hide_legend_entry(obj, index: int) -> None:
  """Drop ONE series from the legend while keeping it plotted - used for the
  invisible base of a stacked band."""
  if obj.legend is None:
    return
  obj.legend.legendEntry = list(obj.legend.legendEntry or []) + [LegendEntry(idx=index, delete=True)]


def place(ws, obj, anchor: str) -> None:
  """Finalize and anchor. Re-applies the polish that must survive series adds."""
  if not getattr(obj, MARKER_ATTR, False):
    raise ValueError("chart was not created through design.chart()")
  obj.varyColors = False
  for axis in (getattr(obj, "x_axis", None), getattr(obj, "y_axis", None)):
    if axis is not None:
      axis.delete = False
  if obj.legend is not None and len(obj.series) < 2:
    obj.legend = None          # a single series needs no legend box
  ws.add_chart(obj, anchor)


# ---------------------------------------------------------------------------
# 6. THE GUARD — what makes the standard permanent
# ---------------------------------------------------------------------------

def _rgb(color) -> Optional[str]:
  """Normalise an openpyxl color to 'RRGGBB', or None when it carries no
  explicit rgb (theme/indexed/unset all mean 'inherited', which is fine)."""
  if color is None:
    return None
  rgb = getattr(color, "rgb", None)
  if not isinstance(rgb, str):
    return None
  rgb = rgb.upper()
  if len(rgb) == 8:            # AARRGGBB
    if rgb[:2] == "00" and rgb[2:] == "000000":
      return None              # openpyxl's "unset" sentinel
    rgb = rgb[2:]
  return rgb if len(rgb) == 6 else None


def audit_workbook(wb, *, max_report: int = 40) -> List[str]:
  """Return every place the workbook bypasses this module.

  Used by tests/test_x1_design_system.py. Keep it cheap and total: it walks
  every populated cell and every chart of every sheet, so a new sheet is
  covered the moment it exists — nobody has to remember to add it.
  """
  problems: List[str] = []

  def note(msg: str) -> None:
    if len(problems) < max_report:
      problems.append(msg)

  for ws in wb.worksheets:
    for line in ws.iter_rows():
      for cell in line:
        if cell.value is None and not cell.has_style:
          continue
        where = f"{ws.title}!{cell.coordinate}"
        f = cell.font
        if f is not None:
          if f.name and f.name != FONT_FAMILY:
            note(f"{where}: font family {f.name!r} (only {FONT_FAMILY} is allowed)")
          if f.size and int(f.size) not in ALLOWED_FONT_SIZES:
            note(f"{where}: font size {f.size} is outside the type scale")
          color = _rgb(f.color)
          if color and color not in ALLOWED_FONT_COLORS:
            note(f"{where}: font color #{color} is outside the palette")
        if cell.fill is not None and cell.fill.fill_type == "solid":
          color = _rgb(cell.fill.fgColor)
          if color and color not in ALLOWED_FILLS:
            note(f"{where}: fill #{color} is outside the palette")
        fmt = cell.number_format
        if fmt and fmt not in ALLOWED_NUMBER_FORMATS:
          note(f"{where}: number format {fmt!r} is outside the design set")

    for idx, obj in enumerate(getattr(ws, "_charts", []) or []):
      title = f"{ws.title} chart #{idx + 1}"
      if not getattr(obj, MARKER_ATTR, False):
        note(f"{title}: built outside design.chart() — the single door")
        continue
      if getattr(obj, "varyColors", None):
        note(f"{title}: varyColors is on (Excel will legend every point)")
      for axis_name in ("x_axis", "y_axis"):
        axis = getattr(obj, axis_name, None)
        if axis is not None and getattr(axis, "delete", False):
          note(f"{title}: {axis_name} is deleted (openpyxl hides axes by default)")
      allowed = {c.upper() for c in ALLOWED_SERIES_COLORS}
      for series in getattr(obj, "series", []) or []:
        props = getattr(series, "graphicalProperties", None)
        marker_props = getattr(getattr(series, "marker", None), "graphicalProperties", None)
        for source in (getattr(props, "solidFill", None),
                       getattr(getattr(props, "line", None), "solidFill", None),
                       getattr(marker_props, "solidFill", None)):
          # openpyxl wraps a fill in a ColorChoice(srgbClr=...) once it round-
          # trips through the model, so a bare isinstance(str) check misses it.
          color = source if isinstance(source, str) else getattr(source, "srgbClr", None)
          if isinstance(color, str) and color.upper() not in allowed | {WHITE}:
            note(f"{title}: series color #{color.upper()} is outside the palette")
  return problems


__all__ = [n for n in dir() if not n.startswith("_")]
