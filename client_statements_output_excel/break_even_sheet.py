"""W2 (2026-08-18) — Break-Even Analysis rendered on the FINMO sheet.

Ruled in docs/WRITING_PHASE_RESEARCH_2.md (R5): the block sits DIRECTLY
BELOW the Income Statement (below Net Income), with LIVE FORMULAS that
reference the sheet's own P&L cells (and the Revenue Drivers sheet for
prices / line revenue), so it recomputes if the model changes. Nothing
here recomputes W1 - ``finmo_json["break_even"]`` (W1, 5c9a8b9) is the
persisted read-out; the Audit Source sheet mirrors it and the Checks sheet
ties Q1 out against this block. The block therefore renders even on
pre-W1 drafts (no persisted block -> no tie-out row, no crash).

Rows (statement key "Break-Even Analysis", registered through
``ctx.add_finmo_row`` so Checks / Audit Source resolve by label):
  Fixed Costs                       = Payroll + Lease/Rent + Depreciation + Interest
  Variable Costs                    = COGS + Marketing + R&D + G&A   ($, helper for annual ratios)
  Variable Cost Ratio               = Variable Costs / Revenue
  Contribution Margin Ratio         = 1 - Variable Cost Ratio
  Break-Even Revenue                = Fixed / CM            (HEADLINE, pre-tax accounting)
  Cash Break-Even Revenue           = (Payroll + Lease + Interest + Debt Repayment + Lease Principal) / CM
  EBITDA-Basis Break-Even Revenue   = (Payroll + Lease) / CM  (comparator)
  Planned Revenue                   = Revenue
  Margin of Safety                  = (Revenue - Break-Even Revenue) / Revenue
  Break-Even Revenue (G&A as fixed) = (Fixed + G&A) / (CM + G&A / Revenue)   (sensitivity)
  Break-Even Units [at planned mix] - <line>  = Break-Even Revenue x (line revenue / total revenue) / unit price
The CVP chart (Part B) is a native openpyxl ScatterChart over a helper
range written below the statements: revenue on the X axis, the total-
revenue line, the total-cost line (fixed intercept + variable slope), the
break-even point, the planned-revenue marker, and LOSS / PROFIT labels.
"""
from __future__ import annotations

from typing import Dict, List, Optional, Tuple

from openpyxl.chart import Reference, ScatterChart

from . import design
from .data import DraftWorkbookData, text
from .excel_utils import (
  ANNUAL_START_COL,
  CURRENCY_FORMAT,
  FILL_BLUE,
  FIRST_LIVE_COL,
  LAST_LIVE_COL,
  NUMBER_FORMAT,
  PERCENT_FORMAT,
  PERIOD_START_COL,
  REVENUE_SHEET,
  WorkbookBuildContext,
  local_ref,
  ref,
  set_formula_style,
  style_row,
  write_section_header,
)

BREAK_EVEN_STATEMENT = "Break-Even Analysis"
_UNIT_SLOT_KEY = "__break_even_slot__"

_ROW_NOTES: Dict[str, str] = {
  "Fixed Costs": "Payroll + Lease/Rent + Depreciation + Interest (owner compensation is inside payroll)",
  "Variable Costs": "COGS + Marketing + R&D + G&A (the model applies these as % of revenue)",
  "Variable Cost Ratio": "Variable Costs / Revenue",
  "Contribution Margin Ratio": "1 - Variable Cost Ratio",
  "Break-Even Revenue": "HEADLINE - pre-tax accounting: Fixed Costs / Contribution Margin (0 = unreachable)",
  "Cash Break-Even Revenue": "(Payroll + Lease + Interest + scheduled principal) / Contribution Margin",
  "EBITDA-Basis Break-Even Revenue": "Comparator: (Payroll + Lease) / Contribution Margin",
  "Planned Revenue": "Revenue from the P&L above",
  "Margin of Safety": "(Planned Revenue - Break-Even Revenue) / Planned Revenue",
  "Break-Even Revenue (G&A as fixed)": "Sensitivity: G&A treated as fixed instead of % of revenue",
}


def _fr(ctx: WorkbookBuildContext, statement: str, label: str, col: int) -> str:
  return local_ref(ctx.finmo_row(statement, label), col)


def _be(ctx: WorkbookBuildContext, label: str, col: int) -> str:
  return local_ref(ctx.finmo_row(BREAK_EVEN_STATEMENT, label), col)


def _revenue_slots(data: DraftWorkbookData) -> List[Tuple[str, str]]:
  """[(slot_key, display)] in Revenue Drivers sheet order."""
  ordered: List[str] = []
  display: Dict[str, str] = {}
  for source_row in data.revenue_rows:
    slot = text(source_row.get("revenue_slot_key")) or f"{text(source_row.get('lob'))}::{text(source_row.get('product'))}"
    if slot not in display:
      ordered.append(slot)
      display[slot] = " / ".join([text(source_row.get("lob")) or "LOB", text(source_row.get("product")) or "Product"])
  return [(slot, display[slot]) for slot in ordered]


def _set(ws, row: int, col: int, formula: str, *, number_format: str = CURRENCY_FORMAT) -> None:
  cell = ws.cell(row=row, column=col, value=formula)
  set_formula_style(cell, number_format=number_format, internal_link=True)


def write_break_even_rows(ws, data: DraftWorkbookData, ctx: WorkbookBuildContext, *, start_row: int) -> int:
  """Reserve + label the block rows directly below the Income Statement.
  Formulas are written by ``fill_break_even_formulas`` once every P&L row
  is registered (the P&L formulas themselves are filled after all
  statements are laid out, in build_finmo_sheet's column loop)."""
  slots = _revenue_slots(data)
  multi_line = len(slots) > 1
  labels = list(_ROW_NOTES.keys())
  unit_labels: List[Tuple[str, str, str]] = []
  for slot, display in slots:
    label = (f"Break-Even Units at planned mix - {display}" if multi_line else f"Break-Even Units - {display}")
    unit_labels.append((label, slot, display))
  write_section_header(ws, start_row, BREAK_EVEN_STATEMENT)
  row = start_row + 1
  for label in labels:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=_ROW_NOTES[label])
    ctx.add_finmo_row(BREAK_EVEN_STATEMENT, label, row)
    row += 1
  for label, slot, display in unit_labels:
    ws.cell(row=row, column=1, value=label)
    ws.cell(
      row=row,
      column=2,
      value=("Break-Even Revenue x this line's revenue share / unit price" if multi_line else "Break-Even Revenue / unit price"),
    )
    ctx.add_finmo_row(BREAK_EVEN_STATEMENT, label, row)
    row += 1
  # Unit-row label -> revenue slot key, for the formula pass. Kept under a
  # private key of the schedule-row registry (values are slot strings, not
  # rows); no sheet consumer enumerates this key.
  ctx.schedule_rows.setdefault(_UNIT_SLOT_KEY, {})
  for label, slot, display in unit_labels:
    ctx.schedule_rows[_UNIT_SLOT_KEY][label] = slot  # type: ignore[assignment]
  return row + 1


def fill_break_even_formulas(ws, data: DraftWorkbookData, ctx: WorkbookBuildContext) -> None:
  """Live formulas per period column (the stub column too - it reads the
  client's stated 'today' P&L, so it shows today's break-even; every FINMO
  row must be formula-driven across all 21 periods for the Checks sheet),
  plus annual columns aggregated the way W1 aggregates (sum of $ rows;
  ratios re-derived from the annual sums)."""
  IS = "Income Statement"
  CF = "Cash Flow"
  slots = _revenue_slots(data)
  for col in range(PERIOD_START_COL, LAST_LIVE_COL + 1):
    rev = _fr(ctx, IS, "Revenue", col)
    fixed = f"{_fr(ctx, IS, 'Payroll', col)}+{_fr(ctx, IS, 'Lease/Rent', col)}+{_fr(ctx, IS, 'Depreciation', col)}+{_fr(ctx, IS, 'Interest', col)}"
    variable = f"{_fr(ctx, IS, 'Cost of Goods Sold', col)}+{_fr(ctx, IS, 'Marketing', col)}+{_fr(ctx, IS, 'Research & Development', col)}+{_fr(ctx, IS, 'General & Administrative', col)}"
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Fixed Costs"), col, f"={fixed}")
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Variable Costs"), col, f"={variable}")
    _write_ratio_and_be_formulas(ws, ctx, col, rev=rev)
    principal = f"ABS({_fr(ctx, CF, 'Debt Repayment', col)})+ABS({_fr(ctx, CF, 'Capital Lease Principal Payments', col)})"
    cm = _be(ctx, "Contribution Margin Ratio", col)
    _set(
      ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Cash Break-Even Revenue"), col,
      f"=IF({cm}>0,({_fr(ctx, IS, 'Payroll', col)}+{_fr(ctx, IS, 'Lease/Rent', col)}+{_fr(ctx, IS, 'Interest', col)}+{principal})/{cm},0)",
    )
    _set(
      ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "EBITDA-Basis Break-Even Revenue"), col,
      f"=IF({cm}>0,({_fr(ctx, IS, 'Payroll', col)}+{_fr(ctx, IS, 'Lease/Rent', col)})/{cm},0)",
    )
    gna = _fr(ctx, IS, "General & Administrative", col)
    _set(
      ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Break-Even Revenue (G&A as fixed)"), col,
      f"=IF({rev}>0,IF(({cm}+{gna}/{rev})>0,({_be(ctx, 'Fixed Costs', col)}+{gna})/({cm}+{gna}/{rev}),0),0)",
    )
    # Volume rows
    total_rev_row = ctx.schedule_row(REVENUE_SHEET, "Total Revenue")
    for label, slot in (ctx.schedule_rows.get(_UNIT_SLOT_KEY) or {}).items():
      unit_row = ctx.finmo_row(BREAK_EVEN_STATEMENT, label)
      price = ref(REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, f"{slot}::Unit Price"), col)
      line_rev = ref(REVENUE_SHEET, ctx.schedule_row(REVENUE_SHEET, f"{slot}::Revenue"), col)
      total_rev = ref(REVENUE_SHEET, total_rev_row, col)
      be_rev = _be(ctx, "Break-Even Revenue", col)
      if len(slots) > 1:
        formula = f"=IF(AND({price}>0,{total_rev}>0),{be_rev}*({line_rev}/{total_rev})/{price},0)"
      else:
        formula = f"=IF({price}>0,{be_rev}/{price},0)"
      _set(ws, unit_row, col, formula, number_format=NUMBER_FORMAT)

  # Annual columns: $ rows summed; ratio / BE rows re-derived from the sums.
  for year in range(1, 6):
    col = ANNUAL_START_COL + year - 1
    start = FIRST_LIVE_COL + (year - 1) * 4
    end = start + 3
    def _sum(label: str, statement: str = BREAK_EVEN_STATEMENT) -> str:
      r = ctx.finmo_row(statement, label)
      return f"SUM({local_ref(r, start)}:{local_ref(r, end)})"
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Fixed Costs"), col, f"={_sum('Fixed Costs')}")
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Variable Costs"), col, f"={_sum('Variable Costs')}")
    rev = f"{_sum('Revenue', IS)}"
    _write_ratio_and_be_formulas(ws, ctx, col, rev=rev)
    cm = _be(ctx, "Contribution Margin Ratio", col)
    cash_fixed = f"{_sum('Payroll', IS)}+{_sum('Lease/Rent', IS)}+{_sum('Interest', IS)}+ABS({_sum('Debt Repayment', CF)})+ABS({_sum('Capital Lease Principal Payments', CF)})"
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Cash Break-Even Revenue"), col, f"=IF({cm}>0,({cash_fixed})/{cm},0)")
    _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "EBITDA-Basis Break-Even Revenue"), col, f"=IF({cm}>0,({_sum('Payroll', IS)}+{_sum('Lease/Rent', IS)})/{cm},0)")
    gna = _sum("General & Administrative", IS)
    _set(
      ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Break-Even Revenue (G&A as fixed)"), col,
      f"=IF(({rev})>0,IF(({cm}+({gna})/({rev}))>0,({_be(ctx, 'Fixed Costs', col)}+{gna})/({cm}+({gna})/({rev})),0),0)",
    )
    total_rev_row = ctx.schedule_row(REVENUE_SHEET, "Total Revenue")
    for label, slot in (ctx.schedule_rows.get(_UNIT_SLOT_KEY) or {}).items():
      unit_row = ctx.finmo_row(BREAK_EVEN_STATEMENT, label)
      # Annual units = sum of the quarterly break-even units.
      _set(ws, unit_row, col, f"=SUM({local_ref(unit_row, start)}:{local_ref(unit_row, end)})", number_format=NUMBER_FORMAT)

  # Styling
  for label, r in ctx.finmo_rows.get(BREAK_EVEN_STATEMENT, {}).items():
    if label.startswith("__"):
      continue
    is_ratio = label in {"Variable Cost Ratio", "Contribution Margin Ratio", "Margin of Safety"}
    is_units = label.startswith("Break-Even Units")
    headline = label == "Break-Even Revenue"
    style_row(
      ws, r,
      fill=FILL_BLUE if headline else None,
      bold=headline,
      number_format=PERCENT_FORMAT if is_ratio else NUMBER_FORMAT if is_units else CURRENCY_FORMAT,
      border_top=headline,
    )
    ws.cell(row=r, column=2).font = design.font("note")


def _write_ratio_and_be_formulas(ws, ctx: WorkbookBuildContext, col: int, *, rev: str) -> None:
  fixed = _be(ctx, "Fixed Costs", col)
  variable = _be(ctx, "Variable Costs", col)
  ratio_row = ctx.finmo_row(BREAK_EVEN_STATEMENT, "Variable Cost Ratio")
  cm_row = ctx.finmo_row(BREAK_EVEN_STATEMENT, "Contribution Margin Ratio")
  be_row = ctx.finmo_row(BREAK_EVEN_STATEMENT, "Break-Even Revenue")
  _set(ws, ratio_row, col, f"=IF(({rev})>0,{variable}/({rev}),0)", number_format=PERCENT_FORMAT)
  _set(ws, cm_row, col, f"=1-{local_ref(ratio_row, col)}", number_format=PERCENT_FORMAT)
  _set(ws, be_row, col, f"=IF({local_ref(cm_row, col)}>0,{fixed}/{local_ref(cm_row, col)},0)")
  _set(ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Planned Revenue"), col, f"={rev}")
  _set(
    ws, ctx.finmo_row(BREAK_EVEN_STATEMENT, "Margin of Safety"), col,
    f"=IF(({rev})>0,(({rev})-{local_ref(be_row, col)})/({rev}),0)", number_format=PERCENT_FORMAT,
  )


# ---------------------------------------------------------------------------
# Part B - the CVP chart (helper range + native scatter chart)
# ---------------------------------------------------------------------------

CVP_POINTS = 11  # 0%..150% of max(planned, break-even) in 15% steps
CVP_HELPER_KEY = "__cvp_helper__"


def write_cvp_helper_and_chart(ws, ctx: WorkbookBuildContext, *, start_row: int, anchor: Optional[str] = None, quarter_col: int = FIRST_LIVE_COL, title_suffix: str = "Q1") -> int:
  """Write the CVP helper range (all live formulas off the block's
  quarter column) below ``start_row`` and add a native ScatterChart
  anchored at ``anchor`` (default: beside the break-even block). Returns
  the next free row. The chart renders on Excel open (fullCalcOnLoad)."""
  col_x, col_rev, col_cost, col_be, col_plan, col_lbl = 1, 2, 3, 4, 5, 6
  write_section_header(ws, start_row, f"Cost-Volume-Profit Chart Data ({title_suffix}) - helper range for the break-even chart", end_col=8)
  hdr = start_row + 1
  for c, name in ((col_x, "Revenue (x)"), (col_rev, "Total Revenue"), (col_cost, "Total Cost"), (col_be, "Break-Even Point"), (col_plan, "Planned Revenue"), (col_lbl, "Region")):
    ws.cell(row=hdr, column=c, value=name).font = design.font("label_strong")
  fixed = local_ref(ctx.finmo_row(BREAK_EVEN_STATEMENT, "Fixed Costs"), quarter_col, abs_ref=True)
  ratio = local_ref(ctx.finmo_row(BREAK_EVEN_STATEMENT, "Variable Cost Ratio"), quarter_col, abs_ref=True)
  be = local_ref(ctx.finmo_row(BREAK_EVEN_STATEMENT, "Break-Even Revenue"), quarter_col, abs_ref=True)
  planned = local_ref(ctx.finmo_row(BREAK_EVEN_STATEMENT, "Planned Revenue"), quarter_col, abs_ref=True)
  # X max = 1.5 x max(planned, break-even) so both markers sit inside the plot.
  xmax_row = hdr + 1
  ws.cell(row=xmax_row, column=col_x, value="X max").font = design.font("note")
  _set(ws, xmax_row, col_rev, f"=1.5*MAX({planned},{be},1)")
  xmax = local_ref(xmax_row, col_rev, abs_ref=True)
  first = xmax_row + 1
  for i in range(CVP_POINTS):
    r = first + i
    _set(ws, r, col_x, f"={xmax}*{i}/{CVP_POINTS - 1}")
    _set(ws, r, col_rev, f"={local_ref(r, col_x)}")
    _set(ws, r, col_cost, f"={fixed}+{ratio}*{local_ref(r, col_x)}")
  last = first + CVP_POINTS - 1
  # Break-even point (one marker), planned-revenue vertical (two points), region labels.
  be_r = last + 1
  _set(ws, be_r, col_x, f"={be}")
  _set(ws, be_r, col_be, f"={be}")
  ws.cell(row=be_r, column=col_lbl, value="Break-even").font = design.font("note")
  plan_r1, plan_r2 = be_r + 1, be_r + 2
  _set(ws, plan_r1, col_x, f"={planned}")
  _set(ws, plan_r1, col_plan, "=0")
  _set(ws, plan_r2, col_x, f"={planned}")
  _set(ws, plan_r2, col_plan, f"={xmax}")
  ws.cell(row=plan_r1, column=col_lbl, value="Planned revenue").font = design.font("note")
  # LOSS / PROFIT label anchors: midway left of BE on the cost line, and
  # midway right of BE on the revenue line.
  loss_r, profit_r = plan_r2 + 1, plan_r2 + 2
  _set(ws, loss_r, col_x, f"={be}/2")
  _set(ws, loss_r, col_be, f"={fixed}+{ratio}*{be}/2")
  ws.cell(row=loss_r, column=col_lbl, value="LOSS").font = design.font("note")
  _set(ws, profit_r, col_x, f"=({be}+{xmax})/2")
  _set(ws, profit_r, col_be, f"=({be}+{xmax})/2")
  ws.cell(row=profit_r, column=col_lbl, value="PROFIT").font = design.font("note")
  ctx.schedule_rows.setdefault(CVP_HELPER_KEY, {})
  ctx.schedule_rows[CVP_HELPER_KEY].update({"first": first, "last": last, "be": be_r, "plan1": plan_r1, "plan2": plan_r2, "loss": loss_r, "profit": profit_r, "hdr": hdr})

  chart = build_cvp_chart(ws, first=first, last=last, be_r=be_r, plan_r1=plan_r1, plan_r2=plan_r2, loss_r=loss_r, profit_r=profit_r, title=f"Break-Even (Cost-Volume-Profit) - {title_suffix}")
  design.place(ws, chart, anchor or f"AD{ctx.finmo_row(BREAK_EVEN_STATEMENT, 'Fixed Costs')}")
  return profit_r + 2


def build_cvp_chart(ws, *, first: int, last: int, be_r: int, plan_r1: int, plan_r2: int, loss_r: int, profit_r: int, title: str) -> ScatterChart:
  """X1: built through the design system's single door. Semantic colors
  (revenue blue / cost red / break-even amber) were validated ALL-PAIRS;
  the planned-revenue line is chrome (muted dashed), not a series identity."""
  chart = design.chart(
    "scatter", title=title, y_format=design.FMT_AXIS_MONEY,
    x_format=design.FMT_AXIS_MONEY, legend="b",  # no axis title: an axis title
    # sits inside the plot band in Excel and collides with the tick labels;
    # both axes are dollars and the legend names every line.
    width=18, height=10,
  )
  x = Reference(ws, min_col=1, min_row=first, max_row=last)
  design.add_series(chart, Reference(ws, min_col=2, min_row=first, max_row=last),
                    x_values=x, title="Total revenue", color=design.SERIES_REVENUE)
  design.add_series(chart, Reference(ws, min_col=3, min_row=first, max_row=last),
                    x_values=x, title="Total cost (fixed + variable)", color=design.SERIES_COST)
  design.add_series(chart, Reference(ws, min_col=4, min_row=be_r, max_row=be_r),
                    x_values=Reference(ws, min_col=1, min_row=be_r, max_row=be_r),
                    title="Break-even", color=design.SERIES_ATTENTION, line=False,
                    marker="diamond", marker_size=11, labels=True, label_position="r")
  design.add_series(chart, Reference(ws, min_col=5, min_row=plan_r1, max_row=plan_r2),
                    x_values=Reference(ws, min_col=1, min_row=plan_r1, max_row=plan_r2),
                    title="Planned revenue", color=design.SERIES_REFERENCE, dashed=True, thin=True)
  design.add_series(chart, Reference(ws, min_col=4, min_row=loss_r, max_row=loss_r),
                    x_values=Reference(ws, min_col=1, min_row=loss_r, max_row=loss_r),
                    title="Loss", color=design.SERIES_COST, line=False,
                    marker="triangle", marker_size=8, labels="name", label_position="t")
  design.add_series(chart, Reference(ws, min_col=4, min_row=profit_r, max_row=profit_r),
                    x_values=Reference(ws, min_col=1, min_row=profit_r, max_row=profit_r),
                    title="Profit", color=design.SERIES[5], line=False,
                    marker="triangle", marker_size=8, labels="name", label_position="b")
  return chart
