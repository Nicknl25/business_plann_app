"""THE VALUATION DIVERGENCE GUARD (Nick's ruling 2, 2026-08-31).

The equity-value FACT and the workbook's Valuation sheet must agree - same
run, same numbers. This script builds the draft's workbook, recalculates it
through Excel, reads the sheet's equity value back, computes the fact's value
through writing_phase.facts.valuation, and FAILS on divergence beyond 0.5%.

Per the CoInitialize law, an environment where the sheet cannot be read is a
FAILURE of the guard, never a pass.

usage: python scripts/writing_phase_valuation_guard.py [--draft c9095a31]
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import tempfile
import time

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
for p in (ROOT, os.path.join(ROOT, "python")):
  if p not in sys.path:
    sys.path.insert(0, p)
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import mysql.connector  # noqa: E402
from dotenv import load_dotenv  # noqa: E402

from writing_phase.facts import valuation as V  # noqa: E402

load_dotenv(os.path.join(ROOT, ".env"))

TOLERANCE = 0.005   # same numbers means same numbers; 0.5% covers float noise


def main() -> int:
  ap = argparse.ArgumentParser()
  ap.add_argument("--draft", default="c9095a31")
  a = ap.parse_args()
  conn = mysql.connector.connect(host=os.getenv("MYSQL_HOST"), user=os.getenv("MYSQL_USER"),
                                 password=os.getenv("MYSQL_PASSWORD"), database=os.getenv("MYSQL_DB"))
  cur = conn.cursor(dictionary=True)
  cur.execute("SELECT * FROM intake_consult_drafts WHERE draft_id LIKE %s", (a.draft + "%",))
  d = cur.fetchone()
  if not d:
    print("no such draft"); return 2

  fact = V.compute_valuation(conn.cursor(), d)
  if fact:
    print("--- fact components ---")
    print("   component_totals        %s" % fact.get("component_totals"))
    for k in ("cost_of_equity", "wacc", "spread", "ufcf_total", "perp_tv", "exit_tv",
              "tv_used", "pv_explicit", "tv_pv", "net_debt",
              "enterprise_value", "equity_value", "sde_y5", "value_at_exit_multiple"):
      v = fact.get(k)
      print("   %-24s %s" % (k, round(v, 2) if isinstance(v, (int, float)) else v))
  if not fact or fact.get("equity_value") is None:
    print("fact side: ABSENT (%s) - nothing to guard, and nothing may be written"
          % ("spread below floor" if fact and not fact.get("spread_ok") else "insufficient model"))
    return 0

  # build the workbook fresh from the same stored run
  from client_statements_output_excel.export_client_workbook import export_workbook_for_row
  tmp = tempfile.mkdtemp(prefix="val_guard_")
  row = dict(d); row["business_name"] = "Valuation Guard Probe"
  path = str(export_workbook_for_row(row, output_dir=tmp))

  try:
    import pythoncom, win32com.client as win32
    pythoncom.CoInitialize()
    x = win32.gencache.EnsureDispatch("Excel.Application")
    x.Visible = False; x.DisplayAlerts = False
    wb = x.Workbooks.Open(path)
    for _ in range(20):
      try:
        wb.Sheets(1).Name; break
      except Exception:
        time.sleep(1.5)
    x.CalculateFullRebuild()
    wb.Save(); wb.Close(False); x.Quit()
    pythoncom.CoUninitialize()
  except Exception as exc:
    print("GUARD FAILURE: workbook could not be recalculated (%s) - a guard that "
          "cannot run does not pass" % exc)
    return 1

  import openpyxl
  w = openpyxl.load_workbook(path, data_only=True)
  if "Valuation" not in w.sheetnames:
    print("GUARD FAILURE: no Valuation sheet in the built workbook"); return 1
  ws = w["Valuation"]
  sheet_val = None
  # component print: when the headline diverges, the term that moved names
  # itself here instead of being guessed at.
  WATCH = ("cost of equity", "wacc", "effective tax", "terminal value",
           "present value", "enterprise value", "less: net debt",
           "unlevered free cash flow", "nopat", "add: depreciation",
           "less: capital expenditure", "less: change in working capital",
           "ebitda", "ebit", "seller's discretionary")
  print("--- sheet components ---")
  for r in range(1, ws.max_row + 1):
    label = str(ws.cell(r, 1).value or "").strip()
    ll = label.lower()
    if any(wd in ll for wd in WATCH):
      v = ws.cell(r, 2).value
      t = ws.cell(r, 3 + 20).value   # the Total column for series rows
      print("   %-42s B=%-16s TOT=%s" % (label[:42],
            round(v, 2) if isinstance(v, (int, float)) else v,
            round(t, 2) if isinstance(t, (int, float)) else ""))
    if "equity value" in ll and "implied" not in ll:
      v = ws.cell(r, 2).value
      if isinstance(v, (int, float)):
        sheet_val = float(v)
  shutil.rmtree(tmp, ignore_errors=True)
  if sheet_val is None:
    print("GUARD FAILURE: equity value cell not found or not numeric"); return 1

  fact_val = float(fact["equity_value"])
  diff = abs(fact_val - sheet_val) / max(1.0, abs(sheet_val))
  print("fact equity value:  %s" % "{:,.0f}".format(fact_val))
  print("sheet equity value: %s" % "{:,.0f}".format(sheet_val))
  print("divergence: %.3f%%  (tolerance %.1f%%)" % (diff * 100, TOLERANCE * 100))
  if diff > TOLERANCE:
    print("GUARD FAILURE: the document and the workbook disagree - the fact may not ship")
    return 1
  print("GUARD PASS: same run, same number")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
